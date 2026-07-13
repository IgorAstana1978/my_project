"""Extract a preliminary switchboard composition from PDF/workbook sources."""

from __future__ import annotations

import hashlib
import json
import re
from collections import defaultdict
from collections.abc import Iterable, Mapping, Sequence
from dataclasses import dataclass, field
from datetime import UTC, datetime
from functools import partial
from pathlib import Path
from typing import Any

from openpyxl import load_workbook  # type: ignore[import-untyped]
from pypdf import PdfReader
from pypdf.errors import FileNotDecryptedError, PdfReadError

PROJECT_ROOT = Path(__file__).resolve().parents[1]
PDF_PAGE_STATUSES = {
    "text_available",
    "low_text_confidence",
    "image_only",
    "unreadable",
    "encrypted_or_protected",
    "corrupt",
}
MANUAL_PAGE_STATUSES = PDF_PAGE_STATUSES - {"text_available"}
SUPPORTED_WORKBOOK_SUFFIXES = {".xlsx", ".xlsm", ".xls"}
SPACE_RE = re.compile(r"\s+")
DASH_RE = re.compile(r"[‐‑‒–—−]")
NON_WORD_RE = re.compile(r"[^0-9a-zа-яё]+", re.IGNORECASE)
BOARD_RE = re.compile(
    r"(?<![0-9A-ZА-ЯЁ])"
    r"(?P<prefix>ЩР\s*ИТП|НЩР|ВРУ|АВР|ШРС|ЩЭ|УКРМ|ЩОВ|ЩАО|ЩР|ЩО|РП|ВП|"
    r"NSCHR|VRU|AVR|SHRS|SHCHOV|SHCHAO|SHCHR|SHCHO|RP|VP)"
    r"(?:\s*[-‐‑‒–—−]?\s*"
    r"(?P<suffix>\d+(?:[./-]\d+)*(?:[A-ZА-ЯЁ])?"
    r"(?:\s*[-‐‑‒–—−]\s*\d+\s*[AА])?))?",
    re.IGNORECASE,
)
BOARD_CONTEXT_RE = re.compile(
    r"\b(щит|панел|шкаф|спецификац|board|panel|switchboard|schedule)\w*",
    re.IGNORECASE,
)
SCHEMATIC_TITLE_RE = re.compile(
    r"принципиальн\w*\s+схем\w*\s+группов\w*\s+щит\w*",
    re.IGNORECASE,
)
SCHEMATIC_BOARD_RE = re.compile(
    r"(?P<prefix>ЩР\s*ИТП|НЩР|ВРУ|АВР|ШРС|ЩЭ|УКРМ|ЩОВ|ЩАО|ЩР|ЩО|РП|ВП|"
    r"NSCHR|VRU|AVR|SHRS|SHCHOV|SHCHAO|SHCHR|SHCHO|RP|VP)"
    r"\s*[-‐‑‒–—−.]?\s*"
    r"(?P<suffix>\d+(?:[./-]\d+)*(?:[A-ZА-ЯЁ])?)",
    re.IGNORECASE,
)
SOURCE_REFERENCE_RE = re.compile(
    r"(?:^|\b)(?:от|питани[ея]\s+от|существующ\w*\s+групп\w*\s+от)\s+ВРУ\b",
    re.IGNORECASE,
)
QF_ANCHOR_RE = re.compile(r"(?<![0-9A-ZА-ЯЁ])QF\s*(?P<number>\d+)", re.IGNORECASE)
QF_APPARATUS_RE = re.compile(
    r"^(?P<qf>QF\s*\d+)\s*"
    r"(?P<model>ВА88-?32|ВН-?32|АВДТ32)\s*"
    r"(?P<poles>\d\s*P)?\s*"
    r"(?:(?P<trip>[A-ZА-Я])\s*)?"
    r"(?P<current>\d+(?:[.,]\d+)?)\s*(?P<amp>A|А)?"
    r"(?:\s*/\s*(?P<residual>\d+(?:[.,]\d+)?\s*мА))?",
    re.IGNORECASE,
)
BOARD_QUANTITY_RE = re.compile(
    r"(?:кол(?:ичество|-во)?|qty|quantity)\s*[:=]?\s*(\d+)", re.IGNORECASE
)
COMPONENT_QUANTITY_RE = re.compile(
    r"^(?P<label>.+?)\s+(?P<qty>\d+(?:[.,]\d+)?)\s*"
    r"(?P<unit>шт\.?|pcs?|компл\.?|комплект(?:а|ов)?|sets?|м\.?)$",
    re.IGNORECASE,
)
RATING_RE = re.compile(
    r"\b\d+(?:[.,]\d+)?\s*(?:A|А|V|В|кВт|kW|мА|mA|IP\d{2})\b",
    re.IGNORECASE,
)
PROJECT_NOTE_RE = re.compile(
    r"\b(примечан|границ|замен|корпус|степен[ьи] защиты|схем|"
    r"note|boundary|replace|enclosure|scheme|IP\s*\d{2})\w*",
    re.IGNORECASE,
)
MODEL_RE = re.compile(
    r"\b(?=[A-ZА-ЯЁ0-9./-]{4,}\b)(?=\S*[A-ZА-ЯЁ])(?=\S*\d)\S+", re.IGNORECASE
)
BRANDS = (
    "Schneider Electric",
    "Schneider",
    "DEKraft",
    "Legrand",
    "Siemens",
    "CHINT",
    "KEAZ",
    "КЭАЗ",
    "EKF",
    "IEK",
    "ABB",
)

HEADER_ALIASES: Mapping[str, frozenset[str]] = {
    "board": frozenset(
        {
            "щит",
            "обозначение щита",
            "марка щита",
            "панель",
            "board",
            "switchboard",
        }
    ),
    "board_name": frozenset({"наименование щита", "название щита", "board name"}),
    "board_quantity": frozenset(
        {"количество щитов", "кол-во щитов", "кол во щитов", "board quantity"}
    ),
    "name": frozenset(
        {
            "наименование",
            "наименование позиции",
            "оборудование",
            "позиция",
            "item",
            "description",
        }
    ),
    "model": frozenset(
        {"модель", "артикул", "обозначение", "тип", "model", "catalog number"}
    ),
    "brand": frozenset({"бренд", "производитель", "марка", "brand", "manufacturer"}),
    "rating": frozenset(
        {"номинал", "характеристика", "параметры", "rating", "characteristic"}
    ),
    "unit": frozenset({"ед", "ед.", "единица", "единица измерения", "unit"}),
    "quantity": frozenset(
        {"количество", "кол-во", "кол во", "кол.", "qty", "quantity"}
    ),
    "note": frozenset({"примечание", "комментарий", "note", "remarks"}),
}


class ExtractionError(Exception):
    """Expected input, extraction, or output-policy failure."""


@dataclass(frozen=True)
class Provenance:
    source_file: str
    source_type: str
    locator: str
    raw_text: str
    confidence: float
    reason: str
    page: int | None = None
    block_coordinates: str | None = None
    sheet: str | None = None
    row: int | None = None
    cell_range: str | None = None

    def as_dict(self) -> dict[str, Any]:
        data: dict[str, Any] = {
            "source_file": self.source_file,
            "source_type": self.source_type,
            "locator": self.locator,
            "raw_text": self.raw_text,
            "confidence": self.confidence,
            "reason": self.reason,
        }
        for key in ("page", "block_coordinates", "sheet", "row", "cell_range"):
            value = getattr(self, key)
            if value is not None:
                data[key] = value
        return data


@dataclass(frozen=True)
class PdfBlock:
    text: str
    x: float
    y: float
    number: int


@dataclass(frozen=True)
class SchematicBoardTitle:
    designation: str
    normalized: str
    title: str
    block: PdfBlock


@dataclass(frozen=True)
class QfSegment:
    designation: str
    raw_text: str
    block: PdfBlock


@dataclass
class ComponentCandidate:
    label: str
    quantity: int | float | None
    unit: str | None
    model: str | None
    brand: str | None
    rating: str | None
    note: str | None
    provenance: list[Provenance]
    confidence: float
    red_flags: list[str] = field(default_factory=list)
    conflicts: list[dict[str, Any]] = field(default_factory=list)


@dataclass
class BoardCandidate:
    designation: str
    normalized: str
    title: str
    quantity: int | None
    provenance: list[Provenance]
    confidence: float
    components: list[ComponentCandidate] = field(default_factory=list)
    red_flags: list[str] = field(default_factory=list)
    conflicts: list[dict[str, Any]] = field(default_factory=list)
    source_types: set[str] = field(default_factory=set)


@dataclass
class SourceExtraction:
    file_name: str
    source_type: str
    sha256: str
    status: str
    boards: list[BoardCandidate] = field(default_factory=list)
    pages: list[dict[str, Any]] = field(default_factory=list)
    sheets: list[dict[str, Any]] = field(default_factory=list)
    red_flags: list[str] = field(default_factory=list)

    def source_metadata(self) -> dict[str, Any]:
        data: dict[str, Any] = {
            "file_name": self.file_name,
            "source_type": self.source_type,
            "sha256": self.sha256,
            "status": self.status,
        }
        if self.pages:
            data["pages"] = self.pages
        if self.sheets:
            data["sheets"] = self.sheets
        return data


@dataclass(frozen=True)
class ExtractionArtifacts:
    manifest_text: str
    draft: dict[str, Any]
    summary: dict[str, Any]


def compact_text(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        value = int(value)
    return SPACE_RE.sub(" ", str(value).replace("\r", " ").replace("\n", " ")).strip()


def normalize_header(value: object) -> str:
    text = DASH_RE.sub("-", compact_text(value).casefold().replace("ё", "е"))
    return SPACE_RE.sub(" ", NON_WORD_RE.sub(" ", text)).strip()


def normalize_designation(value: str) -> str:
    text = DASH_RE.sub("-", compact_text(value).upper().replace("Ё", "Е"))
    text = re.sub(r"\s*([./-])\s*", r"\1", text)
    match = BOARD_RE.search(text)
    if match is None:
        return text
    prefix = SPACE_RE.sub(" ", match.group("prefix").upper()).strip()
    suffix = compact_text(match.group("suffix") or "")
    return f"{prefix}-{suffix}" if suffix else prefix


def find_board_designations(text: str) -> list[str]:
    found: list[str] = []
    for match in BOARD_RE.finditer(text):
        suffix = match.group("suffix")
        if not suffix and BOARD_CONTEXT_RE.search(text) is None:
            continue
        designation = compact_text(match.group(0))
        if designation and designation not in found:
            found.append(designation)
    return found


def is_source_reference(text: str) -> bool:
    return SOURCE_REFERENCE_RE.search(compact_text(text)) is not None


def pdf_blocks(raw_blocks: Sequence[tuple[str, float, float]]) -> list[PdfBlock]:
    return [
        PdfBlock(compact_text(text), x, y, index)
        for index, (text, x, y) in enumerate(raw_blocks, start=1)
        if compact_text(text)
    ]


def schematic_board_designations(text: str) -> list[str]:
    compact = compact_text(text)
    if SCHEMATIC_TITLE_RE.search(compact) is None:
        return []
    found: list[str] = []
    for match in SCHEMATIC_BOARD_RE.finditer(compact):
        designation = compact_text(match.group(0))
        if designation and designation not in found:
            found.append(designation)
    return found


def find_schematic_board_titles(
    blocks: Sequence[PdfBlock],
    page_text: str,
) -> list[SchematicBoardTitle]:
    titles: list[SchematicBoardTitle] = []
    seen: set[str] = set()
    for index, block in enumerate(blocks):
        if SCHEMATIC_TITLE_RE.search(block.text) is None:
            continue
        context = " ".join(value.text for value in blocks[index : index + 8])
        for designation in schematic_board_designations(context):
            key = normalize_designation(designation)
            if key not in seen:
                titles.append(SchematicBoardTitle(designation, key, context, block))
                seen.add(key)

    for raw_line in page_text.splitlines():
        line = compact_text(raw_line)
        for designation in schematic_board_designations(line):
            key = normalize_designation(designation)
            if key not in seen:
                titles.append(
                    SchematicBoardTitle(
                        designation,
                        key,
                        line,
                        PdfBlock(line, 0.0, 0.0, 0),
                    )
                )
                seen.add(key)
    return titles


def qf_designation(value: str) -> str:
    match = QF_ANCHOR_RE.search(value)
    return f"QF{match.group('number')}" if match else "QF"


def split_qf_segments_from_text(text: str) -> list[str]:
    compact = compact_text(text)
    matches = list(QF_ANCHOR_RE.finditer(compact))
    if not matches:
        return []
    segments: list[str] = []
    for index, match in enumerate(matches):
        end = matches[index + 1].start() if index + 1 < len(matches) else len(compact)
        segment = compact_text(compact[match.start() : end])
        if segment:
            segments.append(segment)
    return segments


def qf_segments_from_blocks(
    blocks: Sequence[PdfBlock],
    page_text: str,
) -> list[QfSegment]:
    segments: list[QfSegment] = []
    for index, block in enumerate(blocks):
        if QF_ANCHOR_RE.search(block.text) is None:
            continue
        inline_segments = split_qf_segments_from_text(block.text)
        if len(inline_segments) > 1 or block.text != qf_designation(block.text):
            segments.extend(
                QfSegment(qf_designation(segment), segment, block)
                for segment in inline_segments
            )
            continue

        parts = [block.text]
        for next_block in blocks[index + 1 : index + 12]:
            if QF_ANCHOR_RE.search(next_block.text):
                break
            parts.append(next_block.text)
        raw_segment = compact_text(" ".join(parts))
        segments.append(QfSegment(qf_designation(raw_segment), raw_segment, block))

    if segments:
        return segments

    synthetic_block = PdfBlock("page text", 0.0, 0.0, 0)
    for raw_line in page_text.splitlines():
        for segment in split_qf_segments_from_text(raw_line):
            segments.append(
                QfSegment(qf_designation(segment), segment, synthetic_block)
            )
    return segments


def parse_qf_apparatus(segment: str) -> dict[str, str | None]:
    collapsed = re.sub(r"\s+", " ", compact_text(segment))
    collapsed = re.sub(r"\b(ВА88-32|ВН-32)(?=\dP)", r"\1 ", collapsed)
    collapsed = re.sub(r"\b(АВДТ32)\s*(\dP)([A-ZА-Я])", r"\1 \2 \3", collapsed)
    match = QF_APPARATUS_RE.match(collapsed)
    if match is None:
        return {
            "model": None,
            "poles": None,
            "trip": None,
            "current": None,
            "residual": None,
            "rating": None,
        }
    model = compact_text(match.group("model"))
    poles = compact_text(match.group("poles")).replace(" ", "") or None
    trip = compact_text(match.group("trip")).upper() or None
    current = compact_text(match.group("current")) or None
    amp = compact_text(match.group("amp")).upper()
    residual = compact_text(match.group("residual")) or None
    if current and amp:
        current = f"{current}{amp}"
    rating_parts = []
    if trip:
        rating_parts.append(trip)
    if current:
        rating_parts.append(current)
    rating = "".join(rating_parts) if rating_parts else None
    if residual:
        rating = f"{rating}/{residual}" if rating else residual
    return {
        "model": model,
        "poles": poles,
        "trip": trip,
        "current": current,
        "residual": residual,
        "rating": rating,
    }


def qf_component_from_segment(
    segment: QfSegment,
    file_name: str,
    page_number: int,
    confidence: float,
) -> ComponentCandidate:
    parsed = parse_qf_apparatus(segment.raw_text)
    qf = qf_designation(segment.raw_text)
    label_parts = [qf]
    for field_name in ("model", "poles", "rating"):
        value = parsed[field_name]
        if value:
            label_parts.append(value)
    missing = [
        field_name
        for field_name in ("model", "poles", "current")
        if parsed[field_name] in (None, "")
    ]
    note_values = [f"qf_designation={qf}"]
    for field_name in ("poles", "trip", "current", "residual"):
        value = parsed[field_name]
        if value:
            note_values.append(f"{field_name}={value}")
    coordinates = f"x={segment.block.x:.1f}, y={segment.block.y:.1f}"
    provenance = Provenance(
        source_file=file_name,
        source_type="pdf",
        page=page_number,
        locator=(
            f"page={page_number}; block={segment.block.number or 'line'}; "
            f"qf={qf}; {coordinates}"
        ),
        block_coordinates=coordinates,
        raw_text=segment.raw_text,
        confidence=confidence,
        reason="schematic QF anchor matched a confirmed schematic board title",
    )
    red_flags = ["schematic QF requires Igor review"]
    if missing:
        red_flags.append("schematic QF has missing apparatus fields")
    return ComponentCandidate(
        label=compact_text(" ".join(label_parts)),
        quantity=1,
        unit="шт.",
        model=parsed["model"],
        brand=None,
        rating=parsed["rating"],
        note="; ".join(note_values),
        provenance=[provenance],
        confidence=confidence,
        red_flags=red_flags,
    )


def nearest_schematic_board(
    segment: QfSegment,
    titles: Sequence[SchematicBoardTitle],
) -> SchematicBoardTitle | None:
    if not titles:
        return None
    if len(titles) == 1:
        return titles[0]
    scored = sorted(
        (
            (
                abs(segment.block.y - title.block.y)
                + abs(segment.block.x - title.block.x) / 10,
                title,
            )
            for title in titles
        ),
        key=lambda value: value[0],
    )
    if len(scored) > 1 and abs(scored[0][0] - scored[1][0]) < 50:
        return None
    return scored[0][1]


def parse_number(value: object) -> int | float | None:
    text = compact_text(value).replace(",", ".")
    if not text:
        return None
    try:
        number = float(text)
    except ValueError:
        return None
    if number <= 0:
        return None
    return int(number) if number.is_integer() else number


def file_sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as source:
        for chunk in iter(lambda: source.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def source_locator_for_pdf(
    file_name: str,
    page_number: int,
    line: str,
    blocks: Sequence[tuple[str, float, float]],
    confidence: float,
    reason: str,
) -> Provenance:
    block_number = 0
    x = 0.0
    y = 0.0
    needle = compact_text(line)[:24]
    for index, (block_text, block_x, block_y) in enumerate(blocks, start=1):
        if needle and needle in compact_text(block_text):
            block_number, x, y = index, block_x, block_y
            break
    coordinates = f"x={x:.1f}, y={y:.1f}" if block_number else "unavailable"
    locator = f"page={page_number}; block={block_number or 'line'}; {coordinates}"
    return Provenance(
        source_file=file_name,
        source_type="pdf",
        page=page_number,
        locator=locator,
        block_coordinates=coordinates,
        raw_text=compact_text(line),
        confidence=confidence,
        reason=reason,
    )


def explicit_brand(text: str) -> str | None:
    lowered = text.casefold()
    for brand in BRANDS:
        if brand.casefold() in lowered:
            return brand
    match = re.search(r"(?:бренд|производитель|brand)\s*[:=]\s*([^,;|]+)", text, re.I)
    return compact_text(match.group(1)) if match else None


def explicit_model(text: str) -> str | None:
    match = re.search(r"(?:модель|артикул|model)\s*[:=]\s*([^,;|]+)", text, re.I)
    if match:
        return compact_text(match.group(1))
    for candidate in MODEL_RE.findall(text):
        if RATING_RE.fullmatch(candidate) is None:
            return compact_text(candidate.strip(".,;"))
    return None


def component_from_text(
    line: str,
    provenance: Provenance,
    confidence: float,
) -> ComponentCandidate | None:
    match = COMPONENT_QUANTITY_RE.match(compact_text(line))
    if match is None:
        return None
    label = compact_text(match.group("label").replace("|", " "))
    if len(label) < 3 or normalize_header(label).startswith(("итого", "total")):
        return None
    quantity = parse_number(match.group("qty"))
    rating_match = RATING_RE.search(label)
    return ComponentCandidate(
        label=label,
        quantity=quantity,
        unit=compact_text(match.group("unit")),
        model=explicit_model(label),
        brand=explicit_brand(label),
        rating=compact_text(rating_match.group(0)) if rating_match else None,
        note=None,
        provenance=[provenance],
        confidence=confidence,
    )


def suspicious_block_order(blocks: Sequence[tuple[str, float, float]]) -> bool:
    coordinates = [(x, y) for text, x, y in blocks if compact_text(text)]
    if len(coordinates) < 4:
        return False
    upward_moves = sum(
        1
        for previous, current in zip(coordinates, coordinates[1:], strict=False)
        if current[1] > previous[1] + 18
    )
    return upward_moves >= 2


def collect_pdf_block(
    blocks: list[tuple[str, float, float]],
    text: str,
    _cm: Sequence[float],
    tm: Sequence[float],
    _font: Mapping[str, Any] | None,
    _font_size: float,
) -> None:
    if compact_text(text):
        blocks.append((text, float(tm[4]), float(tm[5])))


def page_has_image_xobject(page: Any) -> bool:
    try:
        resources = page.get("/Resources")
        if resources is None:
            return False
        resources = resources.get_object()
        xobjects = resources.get("/XObject")
        if xobjects is None:
            return False
        xobjects = xobjects.get_object()
        return any(
            image.get_object().get("/Subtype") == "/Image"
            for image in xobjects.values()
        )
    except AttributeError, KeyError, TypeError, ValueError:
        return False


def extract_pdf(path: Path) -> SourceExtraction:
    path = path.expanduser().resolve(strict=False)
    if path.suffix.casefold() != ".pdf":
        raise ExtractionError(f"project PDF must have .pdf suffix: {path}")
    if not path.is_file():
        raise ExtractionError(f"project PDF does not exist: {path}")
    result = SourceExtraction(path.name, "pdf", file_sha256(path), "processed")
    try:
        reader = PdfReader(str(path), strict=False)
        if reader.is_encrypted and reader.decrypt("") == 0:
            result.status = "encrypted_or_protected"
            result.red_flags.append(
                "PDF is encrypted or protected and requires manual review"
            )
            result.pages.append({"page": 0, "status": "encrypted_or_protected"})
            return result
    except FileNotDecryptedError, PermissionError:
        result.status = "encrypted_or_protected"
        result.red_flags.append(
            "PDF is encrypted or protected and requires manual review"
        )
        result.pages.append({"page": 0, "status": "encrypted_or_protected"})
        return result
    except PdfReadError, OSError, ValueError:
        result.status = "corrupt"
        result.red_flags.append("PDF is corrupt and could not be processed")
        result.pages.append({"page": 0, "status": "corrupt"})
        return result

    boards_by_key: dict[str, BoardCandidate] = {}
    for page_number, page in enumerate(reader.pages, start=1):
        blocks: list[tuple[str, float, float]] = []

        try:
            page_text = (
                page.extract_text(visitor_text=partial(collect_pdf_block, blocks)) or ""
            )
        except FileNotDecryptedError, PdfReadError, KeyError, TypeError, ValueError:
            result.pages.append({"page": page_number, "status": "unreadable"})
            result.red_flags.append(f"PDF page {page_number} is unreadable")
            continue

        compact_page = compact_text(page_text)
        structured_blocks = pdf_blocks(blocks)
        schematic_titles = find_schematic_board_titles(structured_blocks, page_text)
        qf_segments = qf_segments_from_blocks(structured_blocks, page_text)
        order_suspect = suspicious_block_order(blocks)
        if not compact_page:
            if page_has_image_xobject(page):
                status = "image_only"
                result.red_flags.append(
                    f"PDF page {page_number} has no usable text layer and requires "
                    "OCR/manual review"
                )
            else:
                status = "unreadable"
                result.red_flags.append(
                    f"PDF page {page_number} is blank or unreadable and requires "
                    "manual review"
                )
        elif len(compact_page) < 30 or len(compact_page.split()) < 4 or order_suspect:
            status = "low_text_confidence"
            result.red_flags.append(
                f"PDF page {page_number} has low text confidence or uncertain "
                "block order"
            )
        else:
            status = "text_available"
        page_qf_detected = len(qf_segments)
        page_qf_extracted = 0
        page_qf_unresolved = (
            page_qf_detected if qf_segments and not schematic_titles else 0
        )
        page_metadata = {
            "page": page_number,
            "status": status,
            "text_characters": len(compact_page),
            "block_count": len(blocks),
            "block_order_suspect": order_suspect,
            "qf_tokens_detected": page_qf_detected,
            "qf_components_extracted": page_qf_extracted,
            "qf_unresolved_count": page_qf_unresolved,
        }
        result.pages.append(page_metadata)
        if status == "image_only":
            continue

        confidence = 0.82 if status == "text_available" else 0.45
        schematic_page = bool(schematic_titles)
        if schematic_page:
            for title in schematic_titles:
                current_board = boards_by_key.get(title.normalized)
                provenance = source_locator_for_pdf(
                    path.name,
                    page_number,
                    title.title,
                    blocks,
                    confidence,
                    "schematic board title matched a controlled title pattern",
                )
                if current_board is None:
                    current_board = BoardCandidate(
                        designation=title.designation,
                        normalized=title.normalized,
                        title=title.title,
                        quantity=None,
                        provenance=[provenance],
                        confidence=confidence,
                        source_types={"pdf"},
                    )
                    boards_by_key[title.normalized] = current_board
                else:
                    current_board.provenance.append(provenance)

            unresolved_qf = 0
            for segment in qf_segments:
                matched_title = nearest_schematic_board(segment, schematic_titles)
                component = qf_component_from_segment(
                    segment, path.name, page_number, confidence
                )
                if matched_title is None:
                    unresolved_qf += 1
                    key = f"UNASSIGNED-QF-P{page_number}"
                    current_board = boards_by_key.get(key)
                    if current_board is None:
                        provenance = component.provenance[0]
                        current_board = BoardCandidate(
                            designation=key,
                            normalized=key,
                            title=f"Unassigned schematic QF on page {page_number}",
                            quantity=None,
                            provenance=[provenance],
                            confidence=min(confidence, 0.35),
                            source_types={"pdf"},
                            red_flags=[
                                "schematic QF board assignment is ambiguous",
                            ],
                        )
                        boards_by_key[key] = current_board
                    component.red_flags.append(
                        "schematic QF board assignment is ambiguous"
                    )
                    current_board.components.append(component)
                    continue
                boards_by_key[matched_title.normalized].components.append(component)
                page_qf_extracted += 1

            page_qf_unresolved = unresolved_qf
            page_metadata["qf_components_extracted"] = page_qf_extracted
            page_metadata["qf_unresolved_count"] = page_qf_unresolved
            accounted = page_qf_extracted + page_qf_unresolved
            if page_qf_detected > accounted:
                result.red_flags.append(
                    f"PDF page {page_number} has incomplete schematic QF extraction"
                )
            continue

        if qf_segments:
            result.red_flags.append(
                f"PDF page {page_number} has QF tokens without schematic board title"
            )

        line_current_board: BoardCandidate | None = None
        for raw_line in page_text.splitlines():
            line = compact_text(raw_line)
            if not line:
                continue
            designations = find_board_designations(line)
            if designations:
                if is_source_reference(line):
                    continue
                designation = designations[0]
                key = normalize_designation(designation)
                provenance = source_locator_for_pdf(
                    path.name,
                    page_number,
                    line,
                    blocks,
                    confidence,
                    "switchboard designation matched a controlled designation pattern",
                )
                quantity_match = BOARD_QUANTITY_RE.search(line)
                quantity = int(quantity_match.group(1)) if quantity_match else None
                line_current_board = boards_by_key.get(key)
                if line_current_board is None:
                    line_current_board = BoardCandidate(
                        designation=designation,
                        normalized=key,
                        title=line,
                        quantity=quantity,
                        provenance=[provenance],
                        confidence=confidence,
                        source_types={"pdf"},
                    )
                    boards_by_key[key] = line_current_board
                else:
                    line_current_board.provenance.append(provenance)
                    if quantity is not None and line_current_board.quantity not in (
                        None,
                        quantity,
                    ):
                        line_current_board.red_flags.append(
                            "different switchboard quantities inside PDF"
                        )
                        line_current_board.quantity = None
                continue
            if line_current_board is None:
                continue
            provenance = source_locator_for_pdf(
                path.name,
                page_number,
                line,
                blocks,
                confidence,
                "line ended with an explicit positive quantity and unit",
            )
            line_component = component_from_text(line, provenance, confidence)
            if line_component is not None:
                line_current_board.components.append(line_component)
            elif PROJECT_NOTE_RE.search(line):
                line_current_board.provenance.append(provenance)
                line_current_board.red_flags.append(
                    f"project note requires Igor review: {line[:120]}"
                )

    result.boards = list(boards_by_key.values())
    return result


def header_field(value: object) -> str | None:
    normalized = normalize_header(value)
    for field_name, aliases in HEADER_ALIASES.items():
        if normalized in aliases:
            return field_name
    return None


def header_mapping(row: Sequence[str]) -> dict[str, int]:
    mapping: dict[str, int] = {}
    for index, value in enumerate(row):
        field_name = header_field(value)
        if field_name is not None and field_name not in mapping:
            mapping[field_name] = index
    return mapping if "name" in mapping and "quantity" in mapping else {}


def cell_value(row: Sequence[str], mapping: Mapping[str, int], field_name: str) -> str:
    index = mapping.get(field_name)
    return row[index] if index is not None and index < len(row) else ""


def range_for_indexes(indexes: Iterable[int], row_number: int) -> str:
    from openpyxl.utils import get_column_letter  # type: ignore[import-untyped]

    values = sorted(set(indexes))
    if not values:
        return f"row {row_number}"
    start = f"{get_column_letter(values[0] + 1)}{row_number}"
    end = f"{get_column_letter(values[-1] + 1)}{row_number}"
    return start if start == end else f"{start}:{end}"


def workbook_provenance(
    file_name: str,
    sheet_name: str,
    row_number: int,
    cell_range: str,
    row: Sequence[str],
    confidence: float,
    reason: str,
) -> Provenance:
    return Provenance(
        source_file=file_name,
        source_type="workbook",
        sheet=sheet_name,
        row=row_number,
        cell_range=cell_range,
        locator=f"sheet={sheet_name}; row={row_number}; cells={cell_range}",
        raw_text=" | ".join(value for value in row if value),
        confidence=confidence,
        reason=reason,
    )


def read_openxml_sheets(path: Path) -> list[tuple[str, list[list[str]]]]:
    workbook = load_workbook(
        path,
        read_only=False,
        data_only=True,
        keep_vba=path.suffix.casefold() == ".xlsm",
        keep_links=False,
    )
    sheets: list[tuple[str, list[list[str]]]] = []
    try:
        for worksheet in workbook.worksheets:
            merged_values: dict[tuple[int, int], str] = {}
            for merged_range in worksheet.merged_cells.ranges:
                anchor = compact_text(
                    worksheet.cell(merged_range.min_row, merged_range.min_col).value
                )
                for row_number in range(merged_range.min_row, merged_range.max_row + 1):
                    for column_number in range(
                        merged_range.min_col, merged_range.max_col + 1
                    ):
                        merged_values[(row_number, column_number)] = anchor
            rows: list[list[str]] = []
            for row_number in range(1, worksheet.max_row + 1):
                rows.append(
                    [
                        merged_values.get(
                            (row_number, column_number),
                            compact_text(
                                worksheet.cell(row_number, column_number).value
                            ),
                        )
                        for column_number in range(1, worksheet.max_column + 1)
                    ]
                )
            sheets.append((worksheet.title, rows))
    finally:
        workbook.close()
    return sheets


def read_legacy_sheets(path: Path) -> list[tuple[str, list[list[str]]]]:
    import xlrd  # type: ignore[import-untyped]

    try:
        workbook = xlrd.open_workbook(str(path), on_demand=True, formatting_info=False)
    except Exception as error:  # pragma: no cover - error types vary by xlrd version.
        raise ExtractionError(
            f"failed to read legacy .xls workbook: {error}"
        ) from error
    sheets: list[tuple[str, list[list[str]]]] = []
    try:
        for sheet in workbook.sheets():
            merged_values: dict[tuple[int, int], str] = {}
            for row_low, row_high, col_low, col_high in sheet.merged_cells:
                anchor = compact_text(sheet.cell_value(row_low, col_low))
                for row_index in range(row_low, row_high):
                    for column_index in range(col_low, col_high):
                        merged_values[(row_index, column_index)] = anchor
            rows = [
                [
                    merged_values.get(
                        (row_index, column_index),
                        compact_text(sheet.cell_value(row_index, column_index)),
                    )
                    for column_index in range(sheet.ncols)
                ]
                for row_index in range(sheet.nrows)
            ]
            sheets.append((sheet.name, rows))
    finally:
        workbook.release_resources()
    return sheets


def component_from_workbook_row(
    path: Path,
    sheet_name: str,
    row_number: int,
    row: Sequence[str],
    mapping: Mapping[str, int],
) -> ComponentCandidate | None:
    name = cell_value(row, mapping, "name")
    if not name or normalize_header(name).startswith(("итого", "total")):
        return None
    quantity = parse_number(cell_value(row, mapping, "quantity"))
    used_indexes = [
        index for index in mapping.values() if index < len(row) and row[index]
    ]
    cell_range = range_for_indexes(used_indexes, row_number)
    provenance = workbook_provenance(
        path.name,
        sheet_name,
        row_number,
        cell_range,
        row,
        0.9,
        "structured workbook row matched item and quantity headers",
    )
    component = ComponentCandidate(
        label=name,
        quantity=quantity,
        unit=cell_value(row, mapping, "unit") or None,
        model=cell_value(row, mapping, "model") or None,
        brand=cell_value(row, mapping, "brand") or None,
        rating=cell_value(row, mapping, "rating") or None,
        note=cell_value(row, mapping, "note") or None,
        provenance=[provenance],
        confidence=0.9 if quantity is not None else 0.58,
    )
    if quantity is None:
        component.red_flags.append("missing or doubtful component quantity")
    return component


def extract_workbook(path: Path) -> SourceExtraction:
    path = path.expanduser().resolve(strict=False)
    suffix = path.suffix.casefold()
    if suffix not in SUPPORTED_WORKBOOK_SUFFIXES:
        raise ExtractionError(f"spec workbook must be .xlsx, .xlsm, or .xls: {path}")
    if not path.is_file():
        raise ExtractionError(f"spec workbook does not exist: {path}")
    result = SourceExtraction(path.name, "workbook", file_sha256(path), "processed")
    try:
        sheets = (
            read_legacy_sheets(path) if suffix == ".xls" else read_openxml_sheets(path)
        )
    except (OSError, ValueError, KeyError) as error:
        raise ExtractionError(f"failed to read workbook: {error}") from error

    boards_by_key: dict[str, BoardCandidate] = {}
    for sheet_name, rows in sheets:
        result.sheets.append({"sheet": sheet_name, "rows_checked": len(rows)})
        current_board: BoardCandidate | None = None
        mapping: dict[str, int] = {}
        seen_components: set[tuple[str, str, str, str]] = set()
        for row_number, row in enumerate(rows, start=1):
            if not any(row):
                continue
            candidate_mapping = header_mapping(row)
            if candidate_mapping:
                mapping = candidate_mapping
                continue

            board_text = cell_value(row, mapping, "board") if mapping else ""
            designation_source = board_text or " | ".join(row)
            designations = find_board_designations(designation_source)
            name_value = cell_value(row, mapping, "name") if mapping else ""
            if designations and (board_text or not name_value):
                designation = designations[0]
                key = normalize_designation(designation)
                used = [index for index, value in enumerate(row) if value]
                provenance = workbook_provenance(
                    path.name,
                    sheet_name,
                    row_number,
                    range_for_indexes(used, row_number),
                    row,
                    0.92,
                    "switchboard designation found in workbook row",
                )
                board_quantity_value = (
                    parse_number(cell_value(row, mapping, "board_quantity"))
                    if mapping
                    else None
                )
                board_quantity = (
                    int(board_quantity_value)
                    if isinstance(board_quantity_value, int)
                    else None
                )
                current_board = boards_by_key.get(key)
                if current_board is None:
                    current_board = BoardCandidate(
                        designation=designation,
                        normalized=key,
                        title=(
                            cell_value(row, mapping, "board_name")
                            if mapping
                            else compact_text(" | ".join(row))
                        )
                        or designation,
                        quantity=board_quantity,
                        provenance=[provenance],
                        confidence=0.92,
                        source_types={"workbook"},
                    )
                    boards_by_key[key] = current_board
                else:
                    current_board.provenance.append(provenance)
                    if board_quantity is not None and current_board.quantity not in (
                        None,
                        board_quantity,
                    ):
                        current_board.red_flags.append(
                            "different switchboard quantities inside workbook"
                        )
                        current_board.quantity = None
                    elif current_board.quantity is None:
                        current_board.quantity = board_quantity

            if not mapping or not name_value:
                continue
            if current_board is None:
                key = "UNASSIGNED"
                current_board = boards_by_key.get(key)
                if current_board is None:
                    provenance = workbook_provenance(
                        path.name,
                        sheet_name,
                        row_number,
                        range_for_indexes(range(len(row)), row_number),
                        row,
                        0.35,
                        "composition row had no reliable switchboard association",
                    )
                    current_board = BoardCandidate(
                        designation="UNASSIGNED",
                        normalized=key,
                        title="Composition without a determined switchboard",
                        quantity=None,
                        provenance=[provenance],
                        confidence=0.35,
                        red_flags=["composition without a determined switchboard"],
                        source_types={"workbook"},
                    )
                    boards_by_key[key] = current_board
            component = component_from_workbook_row(
                path, sheet_name, row_number, row, mapping
            )
            if component is None:
                continue
            signature = (
                normalize_header(component.label),
                normalize_header(component.model or ""),
                normalize_header(component.rating or ""),
                compact_text(component.quantity),
            )
            if signature in seen_components:
                component.red_flags.append("possible duplicate workbook row")
            seen_components.add(signature)
            current_board.components.append(component)

    result.boards = list(boards_by_key.values())
    return result


def conflict(
    conflict_id: str,
    conflict_type: str,
    field_name: str,
    message: str,
    provenances: Sequence[Provenance],
) -> dict[str, Any]:
    return {
        "conflict_id": conflict_id,
        "type": conflict_type,
        "field": field_name,
        "message": message,
        "sources": [value.as_dict() for value in provenances],
    }


def component_key(component: ComponentCandidate) -> str:
    if component.note:
        qf_match = re.search(r"\bqf_designation=(QF\d+)\b", component.note)
        if qf_match:
            return f"qf:{qf_match.group(1)}"
    if component.model:
        return f"model:{normalize_header(component.model)}"
    label_without_rating = RATING_RE.sub("", component.label)
    return f"label:{normalize_header(label_without_rating)}"


def merge_components(
    board: BoardCandidate,
    components: Sequence[ComponentCandidate],
    both_sources: bool,
) -> tuple[list[ComponentCandidate], int]:
    groups: dict[str, list[ComponentCandidate]] = defaultdict(list)
    for component in components:
        groups[component_key(component)].append(component)
    merged: list[ComponentCandidate] = []
    merged_without_conflict = 0
    for index, values in enumerate(groups.values(), start=1):
        sources = {value.provenance[0].source_type for value in values}
        representative = values[0]
        all_provenance = [item for value in values for item in value.provenance]
        quantities = {value.quantity for value in values if value.quantity is not None}
        attributes = {
            field_name: {
                compact_text(getattr(value, field_name))
                for value in values
                if getattr(value, field_name) not in (None, "")
            }
            for field_name in ("model", "brand", "rating", "unit", "note")
        }
        red_flags = list(
            dict.fromkeys(flag for value in values for flag in value.red_flags)
        )
        conflicts = [item for value in values for item in value.conflicts]
        if len(quantities) > 1:
            conflicts.append(
                conflict(
                    f"{board.normalized}-COMP-{index}-QTY",
                    "component_quantity_mismatch",
                    "quantity_guess",
                    "different component quantities between sources",
                    all_provenance,
                )
            )
        for field_name, field_values in attributes.items():
            if len(field_values) > 1:
                conflicts.append(
                    conflict(
                        f"{board.normalized}-COMP-{index}-{field_name.upper()}",
                        f"component_{field_name}_mismatch",
                        f"{field_name}_guess",
                        f"different component {field_name} values between sources",
                        all_provenance,
                    )
                )
        if both_sources and sources != {"pdf", "workbook"}:
            only_source = next(iter(sources))
            conflicts.append(
                conflict(
                    f"{board.normalized}-COMP-{index}-ONLY",
                    "component_present_in_one_source",
                    "source_presence",
                    f"component is present only in {only_source}",
                    all_provenance,
                )
            )
        if not conflicts and sources == {"pdf", "workbook"}:
            merged_without_conflict += 1
        merged.append(
            ComponentCandidate(
                label=representative.label,
                quantity=next(iter(quantities)) if len(quantities) == 1 else None,
                unit=(
                    next(iter(attributes["unit"]))
                    if len(attributes["unit"]) == 1
                    else None
                ),
                model=(
                    next(iter(attributes["model"]))
                    if len(attributes["model"]) == 1
                    else None
                ),
                brand=(
                    next(iter(attributes["brand"]))
                    if len(attributes["brand"]) == 1
                    else None
                ),
                rating=(
                    next(iter(attributes["rating"]))
                    if len(attributes["rating"]) == 1
                    else None
                ),
                note=(
                    next(iter(attributes["note"]))
                    if len(attributes["note"]) == 1
                    else None
                ),
                provenance=all_provenance,
                confidence=min(value.confidence for value in values),
                red_flags=red_flags,
                conflicts=conflicts,
            )
        )
    return merged, merged_without_conflict


def merge_boards(
    source_results: Sequence[SourceExtraction],
) -> tuple[list[BoardCandidate], dict[str, int]]:
    pdf_boards = [
        board
        for source in source_results
        if source.source_type == "pdf"
        for board in source.boards
    ]
    workbook_boards = [
        board
        for source in source_results
        if source.source_type == "workbook"
        for board in source.boards
    ]
    for pdf_board in pdf_boards:
        for workbook_board in workbook_boards:
            if pdf_board.normalized == workbook_board.normalized:
                continue
            pdf_prefix = re.split(r"\d", pdf_board.normalized, maxsplit=1)[0]
            workbook_prefix = re.split(r"\d", workbook_board.normalized, maxsplit=1)[0]
            pdf_numbers = re.findall(r"\d+", pdf_board.normalized)
            workbook_numbers = re.findall(r"\d+", workbook_board.normalized)
            if pdf_prefix == workbook_prefix and pdf_numbers == workbook_numbers:
                ambiguity = conflict(
                    f"AMBIGUOUS-{pdf_board.normalized}-{workbook_board.normalized}",
                    "ambiguous_switchboard_match",
                    "normalized_designation",
                    "similar designations were not merged automatically",
                    pdf_board.provenance + workbook_board.provenance,
                )
                pdf_board.conflicts.append(ambiguity)
                workbook_board.conflicts.append(ambiguity)
    groups: dict[str, list[BoardCandidate]] = defaultdict(list)
    for source in source_results:
        for board in source.boards:
            groups[board.normalized].append(board)
    both_sources = {source.source_type for source in source_results} == {
        "pdf",
        "workbook",
    }
    merged_boards: list[BoardCandidate] = []
    matched = 0
    rows_merged = 0
    for index, (key, values) in enumerate(groups.items(), start=1):
        sources = {source for value in values for source in value.source_types}
        all_provenance = [item for value in values for item in value.provenance]
        all_components = [item for value in values for item in value.components]
        quantities = {value.quantity for value in values if value.quantity is not None}
        red_flags = list(
            dict.fromkeys(flag for value in values for flag in value.red_flags)
        )
        conflicts = [item for value in values for item in value.conflicts]
        if len(quantities) > 1:
            conflicts.append(
                conflict(
                    f"BOARD-{index}-QTY",
                    "switchboard_quantity_mismatch",
                    "quantity_guess",
                    "different switchboard quantities between sources",
                    all_provenance,
                )
            )
        if both_sources and sources != {"pdf", "workbook"}:
            only_source = next(iter(sources))
            conflicts.append(
                conflict(
                    f"BOARD-{index}-ONLY",
                    "switchboard_present_in_one_source",
                    "source_presence",
                    f"switchboard is present only in {only_source}",
                    all_provenance,
                )
            )
        if sources == {"pdf", "workbook"}:
            matched += 1
        merged_components, component_merged_count = merge_components(
            values[0], all_components, both_sources
        )
        rows_merged += component_merged_count
        merged_boards.append(
            BoardCandidate(
                designation=values[0].designation,
                normalized=key,
                title=next((value.title for value in values if value.title), key),
                quantity=next(iter(quantities)) if len(quantities) == 1 else None,
                provenance=all_provenance,
                confidence=min(value.confidence for value in values),
                components=merged_components,
                red_flags=red_flags,
                conflicts=conflicts,
                source_types=sources,
            )
        )
    return merged_boards, {
        "switchboards_matched": matched,
        "switchboards_unmatched": len(merged_boards) - matched if both_sources else 0,
        "rows_merged_without_conflict": rows_merged,
    }


def component_to_draft(
    component: ComponentCandidate, component_id: str
) -> dict[str, Any]:
    missing = [
        field_name
        for field_name, value in (
            ("quantity_guess", component.quantity),
            ("model_guess", component.model),
            ("brand_guess", component.brand),
            ("rating_guess", component.rating),
            ("unit_guess", component.unit),
        )
        if value in (None, "")
    ]
    return {
        "component_id": component_id,
        "component_code_guess": component.model,
        "component_label_guess": component.label,
        "quantity_guess": component.quantity,
        "install_type_guess": "manual_review_required",
        "confidence": component.confidence,
        "evidence": [
            f"{value.source_type}: {value.locator}: "
            f"{compact_text(value.raw_text)[:160]}"
            for value in component.provenance
        ],
        "red_flags": list(
            dict.fromkeys(
                component.red_flags
                + [item["message"] for item in component.conflicts]
                + (["component has missing values"] if missing else [])
            )
        ),
        "assumptions": [],
        "requires_igor_confirmation": True,
        "model_guess": component.model,
        "brand_guess": component.brand,
        "rating_guess": component.rating,
        "unit_guess": component.unit,
        "note_guess": component.note,
        "provenance": [value.as_dict() for value in component.provenance],
        "conflicts": component.conflicts,
        "missing_fields": missing,
        "review_status": "requires_igor_review",
    }


def board_to_draft(board: BoardCandidate, item_id: str) -> dict[str, Any]:
    missing = []
    if board.quantity is None:
        missing.append("quantity_guess")
    if not board.components:
        missing.append("components")
    evidence = [
        f"{value.source_type}: {value.locator}: {compact_text(value.raw_text)[:160]}"
        for value in board.provenance
    ] or ["No reliable source locator was available; Igor review is required."]
    return {
        "item_id": item_id,
        "product_name_guess": board.title or board.designation,
        "product_type_guess": "switchboard",
        "quantity_guess": board.quantity,
        "cabinet_guess": {
            "code_guess": None,
            "label_guess": None,
            "confidence": 0.0,
            "evidence": ["Cabinet data was not reliably extracted."],
            "red_flags": ["cabinet or enclosure requires Igor review"],
        },
        "components": [
            component_to_draft(component, f"{item_id}-COMP-{index:03d}")
            for index, component in enumerate(board.components, start=1)
        ],
        "confidence": board.confidence,
        "evidence": evidence,
        "red_flags": list(
            dict.fromkeys(
                board.red_flags
                + [item["message"] for item in board.conflicts]
                + (["switchboard has missing values"] if missing else [])
            )
        ),
        "assumptions": [],
        "requires_igor_confirmation": True,
        "normalized_designation": board.normalized,
        "provenance": [value.as_dict() for value in board.provenance],
        "conflicts": board.conflicts,
        "missing_fields": missing,
        "questions_for_igor": [
            "Confirm or correct all conflicts and missing values for this switchboard.",
            "Confirm cabinet, enclosure, and ingress-protection requirements.",
            "Confirm scheme and supply-boundary requirements from the project.",
        ],
        "review_status": "requires_igor_review",
    }


def build_artifacts(
    project_pdf: Path | None,
    spec_workbook: Path | None,
) -> ExtractionArtifacts:
    if project_pdf is None and spec_workbook is None:
        raise ExtractionError("at least one source must be provided")
    sources: list[SourceExtraction] = []
    if project_pdf is not None:
        sources.append(extract_pdf(project_pdf))
    if spec_workbook is not None:
        sources.append(extract_workbook(spec_workbook))
    boards, merge_counts = merge_boards(sources)

    manual_pages = sum(
        1
        for source in sources
        for page in source.pages
        if page.get("status") in MANUAL_PAGE_STATUSES
    )
    pdf_pages = sum(
        len(source.pages) for source in sources if source.source_type == "pdf"
    )
    pdf_extractable = sum(
        1
        for source in sources
        for page in source.pages
        if page.get("status") in {"text_available", "low_text_confidence"}
    )
    extracted_rows = sum(
        len(board.components) for source in sources for board in source.boards
    )
    qf_tokens_detected = sum(
        int(page.get("qf_tokens_detected", 0))
        for source in sources
        for page in source.pages
    )
    qf_components_extracted = sum(
        int(page.get("qf_components_extracted", 0))
        for source in sources
        for page in source.pages
    )
    qf_unresolved_count = sum(
        int(page.get("qf_unresolved_count", 0))
        for source in sources
        for page in source.pages
    )
    conflicts_found = sum(
        len(board.conflicts)
        + sum(len(component.conflicts) for component in board.components)
        for board in boards
    )
    root_red_flags = list(
        dict.fromkeys(flag for source in sources for flag in source.red_flags)
    )
    if qf_tokens_detected > qf_components_extracted + qf_unresolved_count:
        root_red_flags.append("schematic QF extraction is incomplete")
    review_rows = (
        manual_pages
        + conflicts_found
        + len(boards)
        + sum(
            len(board.red_flags)
            + sum(len(component.red_flags) for component in board.components)
            for board in boards
        )
    )
    summary: dict[str, Any] = {
        "files_processed": len(sources),
        "file_types": [source.source_type for source in sources],
        "pdf_pages_checked": pdf_pages,
        "pdf_pages_extractable": pdf_extractable,
        "pdf_pages_manual_review": manual_pages,
        "workbook_sheets_processed": sum(len(source.sheets) for source in sources),
        "switchboards_pdf": sum(
            len(source.boards) for source in sources if source.source_type == "pdf"
        ),
        "switchboards_workbook": sum(
            len(source.boards) for source in sources if source.source_type == "workbook"
        ),
        **merge_counts,
        "composition_rows_extracted": extracted_rows,
        "qf_tokens_detected": qf_tokens_detected,
        "qf_components_extracted": qf_components_extracted,
        "qf_unresolved_count": qf_unresolved_count,
        "rows_merged_without_conflict": merge_counts["rows_merged_without_conflict"],
        "conflicts_found": conflicts_found,
        "review_rows": review_rows,
        "ready_for_preliminary_workflow": bool(boards),
    }
    manifest_data = {
        "manifest_version": "mixed_source_bundle.v0.1",
        "sources": [source.source_metadata() for source in sources],
    }
    manifest_text = (
        json.dumps(manifest_data, ensure_ascii=False, indent=2, sort_keys=True) + "\n"
    )
    manifest_hash = hashlib.sha256(manifest_text.encode("utf-8")).hexdigest()
    draft = {
        "schema_version": "preliminary_composition_draft.v0.1",
        "draft_id": f"PRELIM-MIXED-{manifest_hash[:12].upper()}",
        "created_by": "project_spec_extraction",
        "created_at": datetime.now(UTC).isoformat(),
        "source": {
            "source_type": "other",
            "source_summary": "Mixed-source PDF/workbook extraction bundle.",
            "raw_input_sha256": manifest_hash,
            "source_files": [source.source_metadata() for source in sources],
        },
        "safety": {
            "status": "preliminary_only",
            "confirmed_by_igor": False,
            "price_execution_authorized": False,
            "commercial_csv_authorized": False,
            "client_style_export_authorized": False,
            "sending_authorized": False,
            "production_authorized": False,
        },
        "items": [
            board_to_draft(board, f"ITEM-{index:03d}")
            for index, board in enumerate(boards, start=1)
        ],
        "overall_confidence": min((board.confidence for board in boards), default=0.0),
        "red_flags": root_red_flags,
        "assumptions": [
            "Extraction is heuristic and preliminary; no engineering substitutions "
            "were made."
        ],
        "next_required_human_actions": [
            "Igor reviews conflicts, missing values, and source-only rows before "
            "confirming composition."
        ],
        "extraction_summary": summary,
    }
    return ExtractionArtifacts(
        manifest_text=manifest_text, draft=draft, summary=summary
    )
