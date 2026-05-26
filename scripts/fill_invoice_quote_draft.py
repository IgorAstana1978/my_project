"""Fill a draft invoice-quote workbook from a constrained JSON contract."""

from __future__ import annotations

import argparse
import json
import posixpath
import shutil
import subprocess
import sys
import tempfile
from collections.abc import Mapping, Sequence
from dataclasses import dataclass
from pathlib import Path
from typing import Any, cast
from xml.etree import ElementTree
from zipfile import ZIP_DEFLATED, ZipFile

from openpyxl import load_workbook  # type: ignore[import-untyped]
from openpyxl.workbook.workbook import Workbook  # type: ignore[import-untyped]
from openpyxl.worksheet.worksheet import Worksheet  # type: ignore[import-untyped]

PROJECT_ROOT = Path(__file__).resolve().parents[1]
SHEET_NAME = "Счёт-КП шаблон"
NEEDS_CLARIFICATION = "нужно уточнить"
MAX_ITEMS = 5
ITEM_START_ROW = 17
ITEM_END_ROW = 21
FORMULA_CELLS = ("I17", "I18", "I19", "I20", "I21", "I22")
SIGNATURE_RANGE = ("B32", "I34")
HEADER_RANGES = (("C2", "I6"), ("B4", "B6"))
BANNED_ITEM_WORDS = ("монтаж", "пнр", "пусконалад")

SPREADSHEET_NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
OFFICE_REL_NS = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
PACKAGE_REL_NS = "http://schemas.openxmlformats.org/package/2006/relationships"
CONTENT_TYPES_NS = "http://schemas.openxmlformats.org/package/2006/content-types"
DRAWINGML_NS = "http://schemas.openxmlformats.org/drawingml/2006/main"
DRAWING_REL_TYPE = (
    "http://schemas.openxmlformats.org/officeDocument/2006/relationships/drawing"
)
IMAGE_REL_TYPE = (
    "http://schemas.openxmlformats.org/officeDocument/2006/relationships/image"
)
DRAWING_CONTENT_TYPE = "application/vnd.openxmlformats-officedocument.drawing+xml"
NS = {
    "main": SPREADSHEET_NS,
    "r": OFFICE_REL_NS,
    "rel": PACKAGE_REL_NS,
    "a": DRAWINGML_NS,
}

ElementTree.register_namespace("", SPREADSHEET_NS)
ElementTree.register_namespace("r", OFFICE_REL_NS)
ElementTree.register_namespace("rel", PACKAGE_REL_NS)
ElementTree.register_namespace("ct", CONTENT_TYPES_NS)


class DraftFillError(Exception):
    """Expected validation or verification error."""


@dataclass(frozen=True)
class CheckResult:
    name: str
    passed: bool
    detail: str = ""


@dataclass(frozen=True)
class DrawingChain:
    sheet_part: str
    sheet_rels_part: str
    drawing_reference_id: str
    drawing_part: str
    drawing_rels_part: str
    image_reference_id: str
    image_part: str
    image_bytes: bytes


@dataclass(frozen=True)
class TemplateSnapshot:
    template_bytes: bytes
    formulas: dict[str, Any]
    merged_ranges: tuple[str, ...]
    signatures: tuple[tuple[Any, ...], ...]
    headers: dict[str, Any]
    drawing_chain: DrawingChain | None


def parse_args(argv: Sequence[str] | None = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Fill a draft invoice-quote workbook from JSON."
    )
    parser.add_argument("--template", required=True, type=Path, help="Path to .xlsx")
    parser.add_argument("--input", required=True, type=Path, help="Path to JSON input")
    parser.add_argument("--output", required=True, type=Path, help="Output .xlsx path")
    return parser.parse_args(argv)


def fail(message: str) -> None:
    raise DraftFillError(message)


def resolved(path: Path) -> Path:
    return path.expanduser().resolve(strict=False)


def ensure_safe_paths(template: Path, input_path: Path, output: Path) -> None:
    template_path = resolved(template)
    input_json_path = resolved(input_path)
    output_path = resolved(output)

    if not template_path.is_file():
        fail(f"template не существует: {template_path}")
    if not input_json_path.is_file():
        fail(f"input JSON не существует: {input_json_path}")
    if output_path == template_path:
        fail("output совпадает с template")
    if output_path.exists():
        fail(f"output уже существует: {output_path}")
    if not output_path.parent.is_dir():
        fail(f"parent-папка output не существует: {output_path.parent}")
    if output_path.is_relative_to(PROJECT_ROOT):
        fail(f"output находится внутри Git-проекта: {output_path}")


def load_json(path: Path) -> Mapping[str, Any]:
    try:
        raw_data = json.loads(path.read_text(encoding="utf-8"))
    except json.JSONDecodeError as error:
        fail(f"input JSON невалидный: {error.msg}")

    if not isinstance(raw_data, Mapping):
        fail("input JSON должен быть объектом")
    return cast(Mapping[str, Any], raw_data)


def require_mapping(data: Mapping[str, Any], key: str) -> Mapping[str, Any]:
    value = data.get(key)
    if not isinstance(value, Mapping):
        fail(f"обязательное поле отсутствует или не объект: {key}")
    return cast(Mapping[str, Any], value)


def require_field(data: Mapping[str, Any], key: str, prefix: str) -> Any:
    if key not in data:
        fail(f"обязательное поле отсутствует: {prefix}.{key}")
    value = data[key]
    if value is None or value == "":
        fail(f"обязательное поле пустое: {prefix}.{key}")
    return value


def optional_text(value: Any) -> str:
    if value is None or value == "":
        return NEEDS_CLARIFICATION
    return str(value)


def validate_contract(data: Mapping[str, Any]) -> None:
    document = require_mapping(data, "document")
    customer = require_mapping(data, "customer")
    project = require_mapping(data, "project")

    status = require_field(document, "status", "document")
    if status != "черновик":
        fail('document.status должен быть "черновик"')

    require_field(customer, "payer_name", "customer")
    require_field(project, "object_name", "project")

    items = data.get("items")
    if not isinstance(items, list):
        fail("обязательное поле отсутствует или не список: items")
    item_list = cast(list[Any], items)
    if not item_list:
        fail("items пустой")
    if len(item_list) > MAX_ITEMS:
        fail(f"items больше {MAX_ITEMS}")

    for index, item in enumerate(item_list, start=1):
        if not isinstance(item, Mapping) or not item:
            fail(f"items[{index}] должен быть непустым объектом")

        name = require_field(item, "name", f"items[{index}]")
        require_field(item, "unit", f"items[{index}]")
        require_field(item, "quantity", f"items[{index}]")
        if "price_kzt" not in item:
            fail(f"обязательное поле отсутствует: items[{index}].price_kzt")
        if "price_confirmed_by_igor" not in item:
            fail(
                "обязательное поле отсутствует: "
                f"items[{index}].price_confirmed_by_igor"
            )
        if item["price_kzt"] is not None:
            fail(f"items[{index}].price_kzt должен быть null")
        if item["price_confirmed_by_igor"] is not False:
            fail(f"items[{index}].price_confirmed_by_igor должен быть false")

        normalized_name = str(name).casefold()
        if any(word in normalized_name for word in BANNED_ITEM_WORDS):
            fail(f"items[{index}].name содержит запрещённую для MVP позицию: {name}")

    commercial_terms = data.get("commercial_terms")
    if isinstance(commercial_terms, Mapping):
        delivery_confirmed = commercial_terms.get("delivery_confirmed_by_igor")
        delivery_lead_time = commercial_terms.get("delivery_lead_time_working_days")
        if delivery_confirmed is False and delivery_lead_time not in (None, ""):
            fail(
                "delivery_lead_time_working_days заполнен "
                "при delivery_confirmed_by_igor: false"
            )

        total_confirmed = commercial_terms.get("total_amount_confirmed_by_igor")
        total_words = commercial_terms.get("total_amount_words")
        if total_confirmed is False and total_words != NEEDS_CLARIFICATION:
            fail(
                'total_amount_words должен быть "нужно уточнить" '
                "при total_amount_confirmed_by_igor: false"
            )


def worksheet_values(
    worksheet: Worksheet, start_cell: str, end_cell: str
) -> tuple[tuple[Any, ...], ...]:
    return tuple(
        tuple(cell.value for cell in row)
        for row in worksheet[start_cell:end_cell]  # type: ignore[misc]
    )


def worksheet_value_map(
    worksheet: Worksheet, ranges: Sequence[tuple[str, str]]
) -> dict[str, Any]:
    values: dict[str, Any] = {}
    for start_cell, end_cell in ranges:
        for row in worksheet[start_cell:end_cell]:  # type: ignore[misc]
            for cell in row:
                values[cell.coordinate] = cell.value
    return values


def zip_part_bytes(path: Path, part_name: str) -> bytes | None:
    with ZipFile(path) as archive:
        try:
            return archive.read(part_name)
        except KeyError:
            return None


def archive_names(path: Path) -> set[str]:
    with ZipFile(path) as archive:
        return set(archive.namelist())


def read_xml(path: Path, part_name: str) -> ElementTree.Element | None:
    content = zip_part_bytes(path, part_name)
    if content is None:
        return None
    return ElementTree.fromstring(content)


def xml_bytes(root: ElementTree.Element) -> bytes:
    return cast(
        bytes,
        ElementTree.tostring(root, encoding="utf-8", xml_declaration=True),
    )


def normalize_workbook_target(target: str) -> str:
    if target.startswith("/"):
        return target.lstrip("/")
    return posixpath.normpath(posixpath.join("xl", target))


def resolve_relationship_target(source_part: str, target: str) -> str:
    if target.startswith("/"):
        return target.lstrip("/")
    return posixpath.normpath(posixpath.join(posixpath.dirname(source_part), target))


def relationship_target(source_part: str, target_part: str) -> str:
    return posixpath.relpath(target_part, posixpath.dirname(source_part))


def rels_part_for(part_name: str) -> str:
    directory = posixpath.dirname(part_name)
    filename = posixpath.basename(part_name)
    return posixpath.join(directory, "_rels", f"{filename}.rels")


def worksheet_part_for_sheet(path: Path, sheet_name: str) -> str | None:
    workbook = read_xml(path, "xl/workbook.xml")
    workbook_rels = read_xml(path, "xl/_rels/workbook.xml.rels")
    if workbook is None or workbook_rels is None:
        return None

    targets = {
        relationship.get("Id"): normalize_workbook_target(
            relationship.get("Target", "")
        )
        for relationship in workbook_rels.findall(f"{{{PACKAGE_REL_NS}}}Relationship")
    }
    for sheet in workbook.findall("main:sheets/main:sheet", NS):
        if sheet.get("name") != sheet_name:
            continue
        relationship_id = sheet.get(f"{{{OFFICE_REL_NS}}}id")
        target = targets.get(relationship_id)
        if target is not None:
            return target
    return None


def worksheet_drawing_reference(path: Path, sheet_part: str) -> str | None:
    worksheet = read_xml(path, sheet_part)
    if worksheet is None:
        return None
    drawing = worksheet.find("main:drawing", NS)
    if drawing is None:
        return None
    return drawing.get(f"{{{OFFICE_REL_NS}}}id")


def relationship_by_id(
    path: Path, rels_part: str, relationship_id: str, relationship_type: str
) -> str | None:
    root = read_xml(path, rels_part)
    if root is None:
        return None
    for relationship in root.findall(f"{{{PACKAGE_REL_NS}}}Relationship"):
        if (
            relationship.get("Id") == relationship_id
            and relationship.get("Type") == relationship_type
        ):
            target = relationship.get("Target")
            if target is not None:
                return target
    return None


def drawing_image_reference(path: Path, drawing_part: str) -> str | None:
    drawing = read_xml(path, drawing_part)
    if drawing is None:
        return None
    blip = drawing.find(".//a:blip", NS)
    if blip is None:
        return None
    return blip.get(f"{{{OFFICE_REL_NS}}}embed")


def extract_drawing_chain(path: Path) -> DrawingChain | None:
    sheet_part = worksheet_part_for_sheet(path, SHEET_NAME)
    if sheet_part is None:
        return None

    drawing_reference_id = worksheet_drawing_reference(path, sheet_part)
    if drawing_reference_id is None:
        return None

    sheet_rels_part = rels_part_for(sheet_part)
    drawing_target = relationship_by_id(
        path, sheet_rels_part, drawing_reference_id, DRAWING_REL_TYPE
    )
    if drawing_target is None:
        return None
    drawing_part = resolve_relationship_target(sheet_part, drawing_target)

    image_reference_id = drawing_image_reference(path, drawing_part)
    if image_reference_id is None:
        return None

    drawing_rels_part = rels_part_for(drawing_part)
    image_target = relationship_by_id(
        path, drawing_rels_part, image_reference_id, IMAGE_REL_TYPE
    )
    if image_target is None:
        return None
    image_part = resolve_relationship_target(drawing_part, image_target)
    image_bytes = zip_part_bytes(path, image_part)
    if image_bytes is None:
        return None

    return DrawingChain(
        sheet_part=sheet_part,
        sheet_rels_part=sheet_rels_part,
        drawing_reference_id=drawing_reference_id,
        drawing_part=drawing_part,
        drawing_rels_part=drawing_rels_part,
        image_reference_id=image_reference_id,
        image_part=image_part,
        image_bytes=image_bytes,
    )


def content_types_ok(path: Path, chain: DrawingChain) -> bool:
    content_types = read_xml(path, "[Content_Types].xml")
    if content_types is None:
        return False

    image_extension = Path(chain.image_part).suffix.lstrip(".")
    has_image_default = any(
        item.get("Extension") == image_extension
        and item.get("ContentType", "").startswith("image/")
        for item in content_types.findall(f"{{{CONTENT_TYPES_NS}}}Default")
    )
    has_drawing_override = any(
        item.get("PartName") == f"/{chain.drawing_part}"
        and item.get("ContentType") == DRAWING_CONTENT_TYPE
        for item in content_types.findall(f"{{{CONTENT_TYPES_NS}}}Override")
    )
    return has_image_default and has_drawing_override


def merge_content_types(
    output_content: bytes, template_content: bytes, chain: DrawingChain
) -> bytes:
    output_root = ElementTree.fromstring(output_content)
    template_root = ElementTree.fromstring(template_content)
    image_extension = Path(chain.image_part).suffix.lstrip(".")

    has_image_default = any(
        item.get("Extension") == image_extension
        for item in output_root.findall(f"{{{CONTENT_TYPES_NS}}}Default")
    )
    if not has_image_default:
        for item in template_root.findall(f"{{{CONTENT_TYPES_NS}}}Default"):
            if item.get("Extension") == image_extension:
                output_root.append(item)
                break

    drawing_part_name = f"/{chain.drawing_part}"
    has_drawing_override = any(
        item.get("PartName") == drawing_part_name
        for item in output_root.findall(f"{{{CONTENT_TYPES_NS}}}Override")
    )
    if not has_drawing_override:
        for item in template_root.findall(f"{{{CONTENT_TYPES_NS}}}Override"):
            if item.get("PartName") == drawing_part_name:
                output_root.append(item)
                break

    return xml_bytes(output_root)


def next_relationship_id(root: ElementTree.Element) -> str:
    used_ids = {
        relationship.get("Id")
        for relationship in root.findall(f"{{{PACKAGE_REL_NS}}}Relationship")
    }
    index = 1
    while f"rId{index}" in used_ids:
        index += 1
    return f"rId{index}"


def patch_worksheet_drawing_reference(
    worksheet_content: bytes, drawing_reference_id: str
) -> bytes:
    worksheet = ElementTree.fromstring(worksheet_content)
    drawing = worksheet.find("main:drawing", NS)
    if drawing is None:
        drawing = ElementTree.Element(f"{{{SPREADSHEET_NS}}}drawing")
        worksheet.append(drawing)
    drawing.set(f"{{{OFFICE_REL_NS}}}id", drawing_reference_id)
    return xml_bytes(worksheet)


def patch_sheet_relationships(
    rels_content: bytes | None, sheet_part: str, drawing_part: str
) -> tuple[bytes, str]:
    if rels_content is None:
        root = ElementTree.Element(f"{{{PACKAGE_REL_NS}}}Relationships")
    else:
        root = ElementTree.fromstring(rels_content)

    for relationship in root.findall(f"{{{PACKAGE_REL_NS}}}Relationship"):
        if relationship.get("Type") == DRAWING_REL_TYPE:
            root.remove(relationship)

    drawing_reference_id = next_relationship_id(root)
    relationship = ElementTree.Element(f"{{{PACKAGE_REL_NS}}}Relationship")
    relationship.set("Id", drawing_reference_id)
    relationship.set("Type", DRAWING_REL_TYPE)
    relationship.set("Target", relationship_target(sheet_part, drawing_part))
    root.append(relationship)
    return xml_bytes(root), drawing_reference_id


def rewrite_zip_entries(path: Path, updates: Mapping[str, bytes]) -> None:
    with tempfile.NamedTemporaryFile(
        delete=False, dir=path.parent, suffix=".xlsx"
    ) as temporary_file:
        temporary_path = Path(temporary_file.name)

    try:
        with ZipFile(path) as source_archive:
            source_entries = {
                name: source_archive.read(name)
                for name in source_archive.namelist()
                if name not in updates
            }

        with ZipFile(temporary_path, "w", ZIP_DEFLATED) as target_archive:
            for name, content in source_entries.items():
                target_archive.writestr(name, content)
            for name, content in updates.items():
                target_archive.writestr(name, content)

        shutil.move(str(temporary_path), path)
    finally:
        if temporary_path.exists():
            temporary_path.unlink()


def restore_drawing_chain(
    template: Path, output: Path, chain: DrawingChain | None
) -> None:
    if chain is None:
        return

    with ZipFile(template) as template_archive, ZipFile(output) as output_archive:
        required_template_parts = (
            chain.drawing_part,
            chain.drawing_rels_part,
            chain.image_part,
            "[Content_Types].xml",
        )
        missing_template_parts = [
            part
            for part in required_template_parts
            if part not in template_archive.namelist()
        ]
        if missing_template_parts:
            fail(f"template lacks drawing parts: {', '.join(missing_template_parts)}")

        sheet_rels_content = (
            output_archive.read(chain.sheet_rels_part)
            if chain.sheet_rels_part in output_archive.namelist()
            else None
        )
        patched_rels, drawing_reference_id = patch_sheet_relationships(
            sheet_rels_content, chain.sheet_part, chain.drawing_part
        )
        updates = {
            chain.sheet_part: patch_worksheet_drawing_reference(
                output_archive.read(chain.sheet_part), drawing_reference_id
            ),
            chain.sheet_rels_part: patched_rels,
            chain.drawing_part: template_archive.read(chain.drawing_part),
            chain.drawing_rels_part: template_archive.read(chain.drawing_rels_part),
            chain.image_part: template_archive.read(chain.image_part),
            "[Content_Types].xml": merge_content_types(
                output_archive.read("[Content_Types].xml"),
                template_archive.read("[Content_Types].xml"),
                chain,
            ),
        }

    rewrite_zip_entries(output, updates)


def snapshot_template(path: Path) -> TemplateSnapshot:
    template_bytes = path.read_bytes()
    workbook = load_workbook(path, data_only=False)
    worksheet = workbook[SHEET_NAME]
    return TemplateSnapshot(
        template_bytes=template_bytes,
        formulas={cell: worksheet[cell].value for cell in FORMULA_CELLS},
        merged_ranges=tuple(str(item) for item in worksheet.merged_cells.ranges),
        signatures=worksheet_values(worksheet, *SIGNATURE_RANGE),
        headers=worksheet_value_map(worksheet, HEADER_RANGES),
        drawing_chain=extract_drawing_chain(path),
    )


def load_template_workbook(template: Path) -> Workbook:
    workbook = load_workbook(template, data_only=False)
    if SHEET_NAME not in workbook.sheetnames:
        fail(f'лист "{SHEET_NAME}" не найден')
    return workbook


def fill_allowed_cells(workbook: Workbook, data: Mapping[str, Any]) -> None:
    worksheet = workbook[SHEET_NAME]
    customer = require_mapping(data, "customer")
    project = require_mapping(data, "project")
    items = data["items"]
    if not isinstance(items, list):
        fail("items должен быть списком")

    worksheet["B10"] = f"Плательщик: {customer['payer_name']}"
    worksheet["B11"] = f"Объект: {project['object_name']}"
    worksheet["B12"] = (
        f"Основание / проект: {optional_text(project.get('basis_or_project'))}"
    )
    worksheet["B13"] = "Статус документа: Черновик"
    if worksheet["C16"].value in (None, ""):
        worksheet["C16"] = "Раздел / объект / позиция проекта: нужно уточнить"

    for offset, item in enumerate(items):
        if not isinstance(item, Mapping):
            fail(f"items[{offset + 1}] должен быть объектом")
        row = ITEM_START_ROW + offset
        worksheet[f"C{row}"] = item["name"]
        worksheet[f"D{row}"] = item["unit"]
        worksheet[f"E{row}"] = item["quantity"]
        worksheet[f"F{row}"] = optional_text(item.get("instruments_and_devices"))
        worksheet[f"G{row}"] = optional_text(
            item.get("cabinet_type_dimensions_material")
        )
        worksheet[f"H{row}"] = NEEDS_CLARIFICATION

    for row in range(ITEM_START_ROW + len(items), ITEM_END_ROW + 1):
        worksheet[f"C{row}"] = NEEDS_CLARIFICATION
        worksheet[f"D{row}"] = "шт"
        worksheet[f"E{row}"] = 1
        worksheet[f"F{row}"] = NEEDS_CLARIFICATION
        worksheet[f"G{row}"] = NEEDS_CLARIFICATION
        worksheet[f"H{row}"] = NEEDS_CLARIFICATION


def save_output(
    workbook: Workbook, template: Path, output: Path, before: TemplateSnapshot
) -> None:
    workbook.save(output)
    restore_drawing_chain(template, output, before.drawing_chain)


def inspect_output(output: Path) -> subprocess.CompletedProcess[str]:
    inspector = PROJECT_ROOT / "scripts" / "inspect_excel_template.py"
    return subprocess.run(
        [sys.executable, str(inspector), str(output)],
        capture_output=True,
        text=True,
        check=False,
    )


def output_snapshot(path: Path) -> TemplateSnapshot:
    workbook = load_workbook(path, data_only=False)
    worksheet = workbook[SHEET_NAME]
    return TemplateSnapshot(
        template_bytes=b"",
        formulas={cell: worksheet[cell].value for cell in FORMULA_CELLS},
        merged_ranges=tuple(str(item) for item in worksheet.merged_cells.ranges),
        signatures=worksheet_values(worksheet, *SIGNATURE_RANGE),
        headers=worksheet_value_map(worksheet, HEADER_RANGES),
        drawing_chain=extract_drawing_chain(path),
    )


def verify_drawing_chain(
    output: Path, expected: DrawingChain | None
) -> list[CheckResult]:
    if expected is None:
        return [
            CheckResult(
                "template содержит валидную цепочку worksheet -> drawing -> image",
                False,
            )
        ]

    names = archive_names(output)
    sheet_rid = worksheet_drawing_reference(output, expected.sheet_part)
    sheet_has_reference = sheet_rid is not None

    sheet_target = (
        relationship_by_id(
            output, expected.sheet_rels_part, sheet_rid, DRAWING_REL_TYPE
        )
        if sheet_rid is not None
        else None
    )
    sheet_target_part = (
        resolve_relationship_target(expected.sheet_part, sheet_target)
        if sheet_target is not None
        else None
    )
    sheet_rels_ok = sheet_target_part == expected.drawing_part

    image_rid = drawing_image_reference(output, expected.drawing_part)
    image_target = (
        relationship_by_id(
            output, expected.drawing_rels_part, image_rid, IMAGE_REL_TYPE
        )
        if image_rid is not None
        else None
    )
    image_target_part = (
        resolve_relationship_target(expected.drawing_part, image_target)
        if image_target is not None
        else None
    )
    drawing_rels_ok = image_target_part == expected.image_part
    image_bytes = zip_part_bytes(output, expected.image_part)

    return [
        CheckResult("sheet1.xml содержит drawing reference", sheet_has_reference),
        CheckResult(
            "xl/worksheets/_rels/sheet1.xml.rels содержит relationship на drawing",
            sheet_rels_ok,
        ),
        CheckResult(
            "xl/drawings/drawing1.xml существует",
            expected.drawing_part in names,
        ),
        CheckResult(
            "xl/drawings/_rels/drawing1.xml.rels содержит relationship на image",
            drawing_rels_ok,
        ),
        CheckResult(
            "xl/media/image1.png существует и совпадает по байтам",
            expected.image_part == "xl/media/image1.png"
            and image_bytes == expected.image_bytes,
        ),
        CheckResult(
            "[Content_Types].xml содержит drawing/image content types",
            content_types_ok(output, expected),
        ),
    ]


def verify_output(
    template: Path, output: Path, before: TemplateSnapshot
) -> list[CheckResult]:
    results: list[CheckResult] = []
    results.append(CheckResult("output создан", output.is_file()))
    results.append(
        CheckResult(
            "template не изменился по байтам",
            template.read_bytes() == before.template_bytes,
        )
    )

    if not output.is_file():
        return results

    after = output_snapshot(output)
    results.extend(
        [
            CheckResult(
                "формулы I17:I22 совпадают с шаблоном",
                after.formulas == before.formulas,
            ),
            CheckResult(
                "merged ranges совпадают",
                after.merged_ranges == before.merged_ranges,
            ),
            CheckResult(
                "подписи B32:I34 совпадают",
                after.signatures == before.signatures,
            ),
            CheckResult(
                "шапка и реквизиты C2:I6, B4:B6 совпадают",
                after.headers == before.headers,
            ),
        ]
    )
    results.extend(verify_drawing_chain(output, before.drawing_chain))

    inspection = inspect_output(output)
    results.append(
        CheckResult(
            "inspect_excel_template.py открывает output без ошибки",
            inspection.returncode == 0,
            inspection.stderr.strip(),
        )
    )
    return results


def print_report(results: Sequence[CheckResult]) -> None:
    for result in results:
        status = "PASS" if result.passed else "FAIL"
        detail = f" — {result.detail}" if result.detail else ""
        print(f"{status}: {result.name}{detail}")


def main(argv: Sequence[str] | None = None) -> int:
    args = parse_args(argv)
    template = resolved(args.template)
    input_path = resolved(args.input)
    output = resolved(args.output)

    try:
        ensure_safe_paths(template, input_path, output)
        data = load_json(input_path)
        validate_contract(data)
        before = snapshot_template(template)
        workbook = load_template_workbook(template)
        fill_allowed_cells(workbook, data)
        save_output(workbook, template, output, before)
        results = verify_output(template, output, before)
    except DraftFillError as error:
        print(f"ERROR: {error}", file=sys.stderr)
        return 1

    print_report(results)
    if any(not result.passed for result in results):
        return 1
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
