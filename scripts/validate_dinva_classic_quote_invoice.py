"""Independently validate a rendered DINVA classic quote/invoice XLSX."""

from __future__ import annotations

import argparse
import base64
import hashlib
import json
import os
from collections.abc import Mapping, Sequence
from pathlib import Path
from typing import Any, NoReturn, cast
from xml.etree import ElementTree
from zipfile import BadZipFile, ZipFile

from openpyxl import load_workbook  # type: ignore[import-untyped]
from openpyxl.cell.cell import Cell  # type: ignore[import-untyped]

PROJECT_ROOT = Path(__file__).resolve().parents[1]
FAMILY = "DINVA_CLASSIC_QUOTE_INVOICE_V0_1"
TEST_MODE_ENV = "DINVA_RENDERER_TEST_MODE"
PROFILE_SCHEMA_VERSION = "dinva_classic_presentation_profile.v0.1"
DOCUMENT_SCHEMA_VERSION = "dinva_quote_invoice_document.v0.1"
PACKAGE_REL_NS = "http://schemas.openxmlformats.org/package/2006/relationships"
OFFICE_REL_NS = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
CONTENT_TYPES_NS = "http://schemas.openxmlformats.org/package/2006/content-types"
CUSTOM_PROPS_NS = (
    "http://schemas.openxmlformats.org/officeDocument/2006/custom-properties"
)
DRAWING_NS = "http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing"
SPREADSHEET_NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
NS = {
    "rel": PACKAGE_REL_NS,
    "ct": CONTENT_TYPES_NS,
    "cp": CUSTOM_PROPS_NS,
    "xdr": DRAWING_NS,
    "main": SPREADSHEET_NS,
}
EXPECTED_PARTS = {
    "[Content_Types].xml",
    "_rels/.rels",
    "docProps/app.xml",
    "docProps/core.xml",
    "docProps/custom.xml",
    "xl/_rels/workbook.xml.rels",
    "xl/drawings/_rels/drawing1.xml.rels",
    "xl/drawings/drawing1.xml",
    "xl/media/image1.png",
    "xl/styles.xml",
    "xl/theme/theme1.xml",
    "xl/workbook.xml",
    "xl/worksheets/_rels/sheet1.xml.rels",
    "xl/worksheets/sheet1.xml",
}


class ValidationError(ValueError):
    """The workbook does not match its governed inputs."""


def fail(message: str) -> NoReturn:
    raise ValidationError(message)


def require(condition: bool, message: str) -> None:
    if not condition:
        fail(message)


def mapping(value: Any, label: str) -> Mapping[str, Any]:
    require(isinstance(value, Mapping), f"{label} must be an object")
    return cast(Mapping[str, Any], value)


def integer(value: Any, label: str) -> int:
    require(type(value) is int, f"{label} must be an integer")
    return cast(int, value)


def canonical_json(value: object) -> bytes:
    return json.dumps(
        value,
        ensure_ascii=False,
        sort_keys=True,
        separators=(",", ":"),
        allow_nan=False,
    ).encode("utf-8")


def sha256_bytes(raw: bytes) -> str:
    return hashlib.sha256(raw).hexdigest()


def test_mode(requested: bool) -> bool:
    return requested and os.environ.get(TEST_MODE_ENV) == "1"


def color_spec(color: Any) -> dict[str, object] | None:
    if color is None or color.type is None:
        return None
    value: str | int
    if color.type == "rgb":
        value = str(color.rgb)
    elif color.type == "indexed":
        value = int(color.indexed)
    elif color.type == "theme":
        value = int(color.theme)
    else:
        fail(f"unsupported workbook color type: {color.type}")
    return {"type": color.type, "value": value, "tint": float(color.tint or 0)}


def style_spec(cell: Cell) -> dict[str, Any]:
    return {
        "font": {
            "name": cell.font.name,
            "size": float(cell.font.sz or 0),
            "bold": bool(cell.font.b),
            "italic": bool(cell.font.i),
            "underline": cell.font.u,
            "color": color_spec(cell.font.color),
        },
        "fill": {
            "type": cell.fill.fill_type,
            "foreground": color_spec(cell.fill.fgColor),
        },
        "border": {
            "left": cell.border.left.style,
            "right": cell.border.right.style,
            "top": cell.border.top.style,
            "bottom": cell.border.bottom.style,
        },
        "alignment": {
            "horizontal": cell.alignment.horizontal,
            "vertical": cell.alignment.vertical,
            "wrap_text": bool(cell.alignment.wrap_text),
            "shrink_to_fit": bool(cell.alignment.shrink_to_fit),
        },
        "number_format": cell.number_format,
    }


def expected_item_height(item: Mapping[str, Any], rule: Mapping[str, Any]) -> float:
    longest = max(
        len(cast(str, item[field]))
        for field in ("name", "detailed_technical_composition", "enclosure")
    )
    extra = max(
        0,
        (longest - 1) // integer(rule["characters_per_increment"], "height threshold"),
    )
    return min(
        float(rule["maximum"]),
        float(rule["base"]) + extra * float(rule["increment"]),
    )


def validate_governance(
    profile: Mapping[str, Any],
    profile_sha256: str,
    document: Mapping[str, Any],
    document_sha256: str,
    *,
    allow_test_profile: bool,
) -> Mapping[str, Any]:
    require(
        len(profile_sha256) == 64 and len(document_sha256) == 64,
        "binding SHA-256 is invalid",
    )
    require(
        profile.get("schema_version") == PROFILE_SCHEMA_VERSION,
        "profile schema mismatch",
    )
    require(profile.get("profile_id") == FAMILY, "profile id mismatch")
    require(profile.get("document_family") == FAMILY, "profile family mismatch")
    require(
        document.get("schema_version") == DOCUMENT_SCHEMA_VERSION,
        "document schema mismatch",
    )
    require(document.get("document_family") == FAMILY, "document family mismatch")
    contract = mapping(profile.get("presentation_contract"), "presentation contract")
    fingerprint = sha256_bytes(canonical_json(contract))
    require(
        profile.get("presentation_contract_fingerprint") == fingerprint,
        "profile fingerprint mismatch",
    )
    approval = mapping(profile.get("approval_provenance"), "profile approval")
    if test_mode(allow_test_profile):
        require(
            approval.get("status") in {"DRAFT_UNAPPROVED", "APPROVED"},
            "test profile status invalid",
        )
    else:
        require(
            profile.get("artifact_status") == "IMMUTABLE_APPROVED_PROFILE",
            "profile is not immutable",
        )
        require(approval.get("status") == "APPROVED", "profile is not approved")
        require(
            approval.get("authority") == "IGOR_DIRECT_HUMAN_APPROVAL",
            "profile authority mismatch",
        )
        require(
            approval.get("approved_contract_fingerprint") == fingerprint,
            "profile approval binding mismatch",
        )
    document_approval = mapping(
        document.get("approval_provenance"), "document approval"
    )
    require(document_approval.get("status") == "APPROVED", "document is not approved")
    require(
        document_approval.get("rendering_authorized") is True,
        "rendering is not authorized",
    )
    require(
        document_approval.get("client_send_authorized") is False,
        "client-send boundary is open",
    )
    return contract


def expected_cells(
    contract: Mapping[str, Any], document: Mapping[str, Any]
) -> tuple[dict[str, Any], dict[str, str], dict[int, float], int]:
    layout = mapping(contract["layout"], "layout")
    fixed = mapping(contract["fixed_blocks"], "fixed blocks")
    cells: dict[str, Any] = dict(mapping(fixed["company"], "company block"))
    terms = mapping(document["terms"], "terms")
    cells.update(
        {
            "B9": (
                f"Счёт-КП № {document['document_number']} "
                f"от {document['document_date']}"
            ),
            "B10": f"Плательщик: {document['payer']}",
            "B11": f"Объект: {document['object_name'] or 'не указан'}",
            "B12": f"Основание / проект: {document['basis'] or 'не указано'}",
            "B13": "Статус документа: DRAFT XLSX",
            "G9": "ВНИМАНИЕ!",
            "G10": f"Срок изготовления: {terms['manufacturing_lead_time']}",
            "G11": f"Оплата: {terms['payment']}",
            "G12": f"Поставка: {terms['delivery']}",
            "G13": "Technical PASS не является разрешением на отправку",
        }
    )
    styles: dict[str, str] = {}
    for coordinate in mapping(fixed["company"], "company block"):
        styles[coordinate] = "company_title" if coordinate == "C2" else "company_info"
    styles.update({key: "company_info" for key in ("B9", "B10", "B11", "B12", "B13")})
    styles.update(
        {
            key: "warning" if key == "G9" else "company_info"
            for key in ("G9", "G10", "G11", "G12", "G13")
        }
    )
    header_row = integer(layout["table_header_row"], "header row")
    headers = dict(mapping(fixed["table_headers"], "table headers"))
    headers["F"] = document["apparatus_heading"]
    for column in "BCDEFGHI":
        cells[f"{column}{header_row}"] = headers[column]
        styles[f"{column}{header_row}"] = "table_header"
    first_row = integer(layout["first_item_row"], "first item row")
    section_row = integer(layout["section_row"], "section row")
    capacity = integer(layout["item_capacity"], "item capacity")
    section_value = (
        document["object_name"] or document["basis"] or document["document_id"]
    )
    cells[f"C{section_row}"] = f"{fixed['section_label_prefix']} {section_value}"
    styles[f"C{section_row}"] = "terms"
    row_heights: dict[int, float] = {}
    style_names = {
        "B": "position",
        "C": "item_name",
        "D": "unit",
        "E": "quantity",
        "F": "technical_composition",
        "G": "enclosure",
        "H": "money",
        "I": "line_total",
    }
    items = cast(list[Mapping[str, Any]], document["items"])
    require(len(items) <= capacity, "document exceeds governed item capacity")
    rule = mapping(layout["item_height_rule"], "height rule")
    for row in range(first_row, first_row + capacity):
        row_heights[row] = float(rule["base"])
        for column, style_name in style_names.items():
            styles[f"{column}{row}"] = style_name
    line_template = cast(
        str, mapping(contract["formulas"], "formulas")["line_total_template"]
    )
    for offset, item in enumerate(items):
        row = first_row + offset
        values = {
            "B": item["position"],
            "C": item["name"],
            "D": item["unit"],
            "E": item["quantity"],
            "F": item["detailed_technical_composition"],
            "G": item["enclosure"],
            "H": item["approved_unit_price_kzt"],
            "I": line_template.format(row=row),
        }
        for column, value in values.items():
            cells[f"{column}{row}"] = value
            styles[f"{column}{row}"] = style_names[column]
        row_heights[row] = expected_item_height(item, rule)
    total_row = integer(layout["total_row"], "total row")
    cells[f"H{total_row}"] = fixed["total_label"]
    grand_template = cast(
        str, mapping(contract["formulas"], "formulas")["grand_total_template"]
    )
    cells[f"I{total_row}"] = grand_template.format(
        start=first_row, end=first_row + capacity - 1
    )
    styles[f"H{total_row}"] = styles[f"I{total_row}"] = "total"
    vat_row = integer(layout["vat_row"], "VAT row")
    vat = mapping(document["vat"], "VAT")
    cells[f"H{vat_row}"] = vat["approved_text"]
    cells[f"I{vat_row}"] = vat["approved_amount_kzt"]
    styles[f"H{vat_row}"] = styles[f"I{vat_row}"] = "money"
    amount_row = integer(layout["amount_words_row"], "amount row")
    cells[f"C{amount_row}"] = mapping(document["amount_words"], "amount words")[
        "approved_text"
    ]
    styles[f"C{amount_row}"] = "amount_words"
    guard_lines = mapping(fixed["guard_lines"], "guard lines")
    term_values = [
        f"Срок действия: {terms['validity'] or 'не указан'}",
        f"Условия оплаты: {terms['payment']}. Условия поставки: {terms['delivery']}.",
        f"Срок изготовления: {terms['manufacturing_lead_time']}",
        guard_lines["specification"],
        guard_lines["no_send"],
    ]
    term_rows = cast(list[int], layout["terms_rows"])
    require(len(term_rows) == len(term_values), "terms region mismatch")
    for row, value in zip(term_rows, term_values, strict=True):
        cells[f"C{row}"] = value
        styles[f"C{row}"] = "terms"
    signatures = mapping(document["signatures"], "signatures")
    signature_rows = cast(list[int], layout["signature_rows"])
    require(len(signature_rows) == 3, "signature region mismatch")
    director_row, executor_row, review_row = signature_rows
    cells.update(
        {
            f"B{director_row}": signatures["director_title"],
            f"F{director_row}": signatures["director_name"],
            f"B{executor_row}": signatures["executor_title"],
            f"F{executor_row}": signatures["executor_name"],
            f"B{review_row}": guard_lines["review_date"],
        }
    )
    for coordinate in (
        f"B{director_row}",
        f"F{director_row}",
        f"B{executor_row}",
        f"F{executor_row}",
        f"B{review_row}",
    ):
        styles[coordinate] = "signature"
    return cells, styles, row_heights, integer(layout["final_row"], "final row")


def validate_workbook(
    path: Path, contract: Mapping[str, Any], document: Mapping[str, Any]
) -> None:
    try:
        workbook = load_workbook(
            path, data_only=False, read_only=False, keep_links=True
        )
    except (OSError, ValueError, BadZipFile) as exc:
        raise ValidationError(f"XLSX cannot be reopened: {exc}") from exc
    try:
        sheet_contracts = cast(
            list[Mapping[str, Any]], mapping(contract["workbook"], "workbook")["sheets"]
        )
        expected_names = [cast(str, item["name"]) for item in sheet_contracts]
        require(workbook.sheetnames == expected_names, "sheet set/order drift")
        worksheet = workbook[expected_names[0]]
        cells, style_names, row_heights, final_row = expected_cells(contract, document)
        for coordinate, expected in cells.items():
            require(
                worksheet[coordinate].value == expected,
                f"business cell drift: {coordinate}",
            )
        allowed = set(cells)
        for row in worksheet.iter_rows(
            min_row=1, max_row=max(worksheet.max_row, final_row), min_col=2, max_col=9
        ):
            for cell in row:
                if cell.value is not None:
                    require(
                        cell.coordinate in allowed,
                        f"unexpected business cell: {cell.coordinate}",
                    )
        profile_styles = mapping(contract["styles"], "styles")
        for coordinate, style_name in style_names.items():
            require(
                style_spec(worksheet[coordinate]) == profile_styles[style_name],
                f"style drift: {coordinate}",
            )
            require(
                worksheet[coordinate].font.name == "Times New Roman",
                f"font drift: {coordinate}",
            )
        layout = mapping(contract["layout"], "layout")
        for column, expected in mapping(
            layout["column_widths"], "column widths"
        ).items():
            actual = worksheet.column_dimensions[column].width
            require(
                actual is not None and abs(float(actual) - float(expected)) < 1e-9,
                f"width drift: {column}",
            )
        for row, expected in row_heights.items():
            actual = worksheet.row_dimensions[row].height
            require(
                actual is not None and abs(float(actual) - expected) < 1e-9,
                f"height drift: {row}",
            )
        expected_merges = set(
            cast(list[str], mapping(layout["merged_cells"], "merges")["ranges"])
        )
        require(
            {str(value) for value in worksheet.merged_cells.ranges} == expected_merges,
            "merged-cell drift",
        )
        print_contract = mapping(contract["print"], "print contract")
        require(
            str(worksheet.page_setup.paperSize) == str(print_contract["paper_size"]),
            "paper-size drift",
        )
        require(
            worksheet.page_setup.orientation == print_contract["orientation"],
            "orientation drift",
        )
        require(
            worksheet.page_setup.scale == print_contract["scale"], "print scale drift"
        )
        require(
            worksheet.page_setup.fitToHeight == print_contract["fit_to_height"],
            "fit-height drift",
        )
        require(
            bool(worksheet.sheet_properties.pageSetUpPr.fitToPage)
            == bool(print_contract["fit_to_page"]),
            "fit-to-page drift",
        )
        for name, expected in mapping(print_contract["margins"], "margins").items():
            require(
                abs(float(getattr(worksheet.page_margins, name)) - float(expected))
                < 1e-9,
                f"margin drift: {name}",
            )
        expected_area = f"'{expected_names[0]}'!$B$1:$I${final_row}"
        require(str(worksheet.print_area) == expected_area, "print-area drift")
    finally:
        workbook.close()


def relationship_nodes(parts: Mapping[str, bytes]) -> list[ElementTree.Element]:
    nodes: list[ElementTree.Element] = []
    for name, raw in parts.items():
        if name.endswith(".rels"):
            try:
                root = ElementTree.fromstring(raw)
            except ElementTree.ParseError as exc:
                raise ValidationError(f"invalid relationship part: {name}") from exc
            nodes.extend(root.findall("rel:Relationship", NS))
    return nodes


def validate_calc_chain(
    parts: Mapping[str, bytes], formula_coordinates: set[str]
) -> None:
    nodes = relationship_nodes(parts)
    calc_rels = [
        node
        for node in nodes
        if "calcChain" in cast(str, node.get("Type", ""))
        or "calcChain" in cast(str, node.get("Target", ""))
    ]
    content_types = ElementTree.fromstring(parts["[Content_Types].xml"])
    calc_types = [
        node
        for node in content_types
        if "calcChain" in cast(str, node.get("PartName", ""))
        or "calcChain" in cast(str, node.get("ContentType", ""))
    ]
    if "xl/calcChain.xml" not in parts:
        require(not calc_rels and not calc_types, "calcChain residue without part")
        return
    require(
        len(calc_rels) == 1 and len(calc_types) == 1,
        "calcChain package binding invalid",
    )
    try:
        root = ElementTree.fromstring(parts["xl/calcChain.xml"])
    except ElementTree.ParseError as exc:
        raise ValidationError("calcChain XML invalid") from exc
    refs = [node.get("r") for node in root.findall("main:c", NS)]
    require(all(isinstance(ref, str) and ref for ref in refs), "calcChain ref invalid")
    require(len(refs) == len(set(refs)), "calcChain duplicate ref")
    require(
        set(cast(list[str], refs)) == formula_coordinates,
        "calcChain stale/orphan/missing ref",
    )


def validate_package(
    path: Path,
    contract: Mapping[str, Any],
    profile: Mapping[str, Any],
    profile_sha256: str,
    document_sha256: str,
) -> None:
    try:
        with ZipFile(path) as archive:
            names = archive.namelist()
            require(len(names) == len(set(names)), "duplicate OOXML part")
            parts = {name: archive.read(name) for name in names}
    except (OSError, BadZipFile) as exc:
        raise ValidationError(f"invalid OOXML package: {exc}") from exc
    actual_parts = set(parts)
    require(
        actual_parts == EXPECTED_PARTS
        or actual_parts == EXPECTED_PARTS | {"xl/calcChain.xml"},
        "unexpected/missing OOXML part",
    )
    package = mapping(contract["package"], "package contract")
    for name in parts:
        require(
            not any(
                name.startswith(prefix)
                for prefix in cast(list[str], package["forbidden_part_prefixes"])
            ),
            f"forbidden part: {name}",
        )
        require(
            not any(
                name.endswith(suffix)
                for suffix in cast(list[str], package["forbidden_part_suffixes"])
            ),
            f"forbidden part: {name}",
        )
    relationships = relationship_nodes(parts)
    require(
        not any(node.get("TargetMode") == "External" for node in relationships),
        "external relationship found",
    )
    expected_custom = {
        "DINVA_PROFILE_ID": FAMILY,
        "DINVA_DOCUMENT_FAMILY": FAMILY,
        "DINVA_PROFILE_SHA256": profile_sha256,
        "DINVA_PRESENTATION_FINGERPRINT": cast(
            str, profile["presentation_contract_fingerprint"]
        ),
        "DINVA_DOCUMENT_SHA256": document_sha256,
    }
    custom_root = ElementTree.fromstring(parts["docProps/custom.xml"])
    custom: dict[str, str] = {}
    for node in custom_root.findall("cp:property", NS):
        require(node.get("name") not in custom, "duplicate custom property")
        value_node = next(iter(node), None)
        require(
            value_node is not None and value_node.text is not None,
            "empty custom property",
        )
        custom[cast(str, node.get("name"))] = cast(
            str, cast(ElementTree.Element, value_node).text
        )
    require(custom == expected_custom, "profile/document package binding drift")
    asset = mapping(cast(list[Any], contract["assets"])[0], "logo asset")
    logo = parts["xl/media/image1.png"]
    require(sha256_bytes(logo) == asset["sha256"], "logo hash drift")
    require(
        logo == base64.b64decode(cast(str, asset["data_base64"]), validate=True),
        "logo bytes drift",
    )
    drawing = ElementTree.fromstring(parts["xl/drawings/drawing1.xml"])
    anchors = drawing.findall("xdr:oneCellAnchor", NS)
    require(len(anchors) == 1, "logo anchor count drift")
    placement = mapping(asset["placement"], "logo placement")
    require(placement.get("anchor_type") == "ONE_CELL", "logo anchor type drift")
    anchor_node = anchors[0].find("xdr:from", NS)
    require(anchor_node is not None, "logo from anchor missing")
    expected = mapping(placement["from"], "logo from")
    for xml_name, key in (
        ("col", "column"),
        ("colOff", "column_offset"),
        ("row", "row"),
        ("rowOff", "row_offset"),
    ):
        child = cast(ElementTree.Element, anchor_node).find(f"xdr:{xml_name}", NS)
        require(
            child is not None and child.text == str(expected[key]),
            f"logo anchor drift: from.{key}",
        )
    extent = anchors[0].find("xdr:ext", NS)
    require(extent is not None, "logo extent missing")
    expected_extent = mapping(placement["extent"], "logo extent")
    require(
        cast(ElementTree.Element, extent).get("cx") == str(expected_extent["cx"])
        and cast(ElementTree.Element, extent).get("cy") == str(expected_extent["cy"]),
        "logo extent drift",
    )
    drawing_rels = ElementTree.fromstring(parts["xl/drawings/_rels/drawing1.xml.rels"])
    image_rels = [
        node
        for node in drawing_rels.findall("rel:Relationship", NS)
        if node.get("Type") == f"{OFFICE_REL_NS}/image"
    ]
    require(
        len(image_rels) == 1 and image_rels[0].get("Target") == "../media/image1.png",
        "logo relationship drift",
    )
    sheet_rels = ElementTree.fromstring(parts["xl/worksheets/_rels/sheet1.xml.rels"])
    drawing_references = [
        node
        for node in sheet_rels.findall("rel:Relationship", NS)
        if node.get("Type") == f"{OFFICE_REL_NS}/drawing"
    ]
    require(
        len(drawing_references) == 1
        and drawing_references[0].get("Target") == "../drawings/drawing1.xml",
        "drawing relationship drift",
    )
    sheet_xml = ElementTree.fromstring(parts["xl/worksheets/sheet1.xml"])
    formulas = {
        cast(str, cell.get("r"))
        for cell in sheet_xml.findall(".//main:c", NS)
        if cell.find("main:f", NS) is not None
    }
    validate_calc_chain(parts, formulas)


def validate_or_raise(
    workbook_path: Path,
    profile: Mapping[str, Any],
    profile_sha256: str,
    document: Mapping[str, Any],
    document_sha256: str,
    *,
    allow_test_profile: bool = False,
) -> None:
    path = workbook_path.resolve(strict=True)
    require(path.suffix.casefold() == ".xlsx", "candidate suffix is not .xlsx")
    contract = validate_governance(
        profile,
        profile_sha256,
        document,
        document_sha256,
        allow_test_profile=allow_test_profile,
    )
    validate_workbook(path, contract, document)
    validate_package(path, contract, profile, profile_sha256, document_sha256)


def reject_duplicate_keys(pairs: list[tuple[str, Any]]) -> dict[str, Any]:
    result: dict[str, Any] = {}
    for key, value in pairs:
        require(key not in result, f"duplicate JSON key: {key}")
        result[key] = value
    return result


def load_json(path: Path, expected_sha256: str) -> dict[str, Any]:
    raw = path.resolve(strict=True).read_bytes()
    require(sha256_bytes(raw) == expected_sha256, f"input SHA mismatch: {path}")
    try:
        value = json.loads(
            raw.decode("utf-8", errors="strict"),
            object_pairs_hook=reject_duplicate_keys,
        )
    except (UnicodeDecodeError, json.JSONDecodeError) as exc:
        raise ValidationError(f"invalid JSON: {path}") from exc
    require(isinstance(value, Mapping), f"JSON root is not an object: {path}")
    return dict(value)


def parse_args(argv: Sequence[str] | None = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--workbook", type=Path, required=True)
    parser.add_argument("--profile", type=Path, required=True)
    parser.add_argument("--profile-sha256", required=True)
    parser.add_argument("--document", type=Path, required=True)
    parser.add_argument("--document-sha256", required=True)
    parser.add_argument(
        "--test-only-allow-unapproved-profile",
        action="store_true",
        help=argparse.SUPPRESS,
    )
    return parser.parse_args(argv)


def main(argv: Sequence[str] | None = None) -> int:
    args = parse_args(argv)
    try:
        profile = load_json(cast(Path, args.profile), cast(str, args.profile_sha256))
        document = load_json(cast(Path, args.document), cast(str, args.document_sha256))
        validate_or_raise(
            cast(Path, args.workbook),
            profile,
            cast(str, args.profile_sha256),
            document,
            cast(str, args.document_sha256),
            allow_test_profile=bool(args.test_only_allow_unapproved_profile),
        )
    except (OSError, ValidationError) as exc:
        print(f"HOLD: {exc}")
        return 1
    print("DINVA_CLASSIC_VALIDATION=PASS_CANDIDATE_ONLY")
    print("CLIENT_SEND=CLOSED")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
