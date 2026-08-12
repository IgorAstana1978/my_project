"""Calculate a read-only preliminary price draft from confirmed composition CSV."""

from __future__ import annotations

import argparse
import csv
import re
from collections.abc import Sequence
from dataclasses import dataclass, field
from decimal import ROUND_HALF_UP, Decimal, InvalidOperation
from pathlib import Path
from typing import Any

from openpyxl import load_workbook  # type: ignore[import-untyped]

CSV_DELIMITER = ";"
KRN_SHEET_NAME = "КРН"
FORBIDDEN_PRICE_SHEET_NAME = "Прайс"
MAX_LOOKUP_ROW = 200
REQUIRED_COLUMNS = (
    "product_name",
    "cabinet_code",
    "consumables_factor",
    "component_code",
    "component_qty",
    "install_type",
)
TECHNICAL_COLUMNS = REQUIRED_COLUMNS + ("component_label", "cabinet_label")
POSITIVE_INTEGER_RE = re.compile(r"[1-9][0-9]*\Z")
POSITIVE_DECIMAL_RE = re.compile(r"(?:0|[1-9][0-9]*)(?:\.[0-9]+)?\Z")
MATERIAL_MULTIPLIER = Decimal("1.25")
FINAL_MULTIPLIER = Decimal("1.15")


@dataclass(frozen=True)
class ComponentDefinition:
    workbook_label: str | None
    install_type: str


COMPONENT_DEFINITIONS = {
    "EKF-VA47-29-1P": ComponentDefinition(
        workbook_label="ВА47 1 полюсный",
        install_type="modular_1p",
    ),
    "EKF-VA47-29-2P": ComponentDefinition(
        workbook_label="ВА47 2 полюсный",
        install_type="modular_2p",
    ),
    "EKF-VA47-29-3P": ComponentDefinition(
        workbook_label="ВА47 3 полюсный до 63А",
        install_type="modular_3p",
    ),
    "EKF-VN-32-1P": ComponentDefinition(
        workbook_label="ВН-32 1Р 16-25-40-63А",
        install_type="load_switch_1p",
    ),
    "EKF-VN-32-2P": ComponentDefinition(
        workbook_label="ВН-32 2Р 16-25-40-63А",
        install_type="load_switch_2p",
    ),
    "EKF-VN-32-3P": ComponentDefinition(
        workbook_label="ВН-32 3Р 16-25-40-63-80-100А",
        install_type="load_switch_3p",
    ),
    "EKF-AD32-1P-N": ComponentDefinition(
        workbook_label="УЗО АД-32 1Р+N до 63А EKF",
        install_type="diff_1p_n",
    ),
    "EKF-AD12-1P-N-C16-30MA-4P5KA": ComponentDefinition(
        workbook_label=None,
        install_type="diff_1p_n",
    ),
    "EKF-RN-47": ComponentDefinition(
        workbook_label="независимый расцепитель для ВА47 РН47",
        install_type="modular_1p",
    ),
}
CABINET_DEFINITIONS = {
    "CAB-KURN-038-24": "Корпус КУРН-0,38-24 540х490х170",
    "CAB-KRN-18": "Корпус КРН-18 265х440х100",
    "CAB-KRN-12": "Корпус КРН-12 265х330х100",
    "CAB-KRN-24": "Корпус КРН-24 395х330х100",
    "CAB-SCHE-BI-900X900X120-M12": ("Встроенный ЩЭ, 900×900×120 мм, металл 1.2 мм"),
}
CABINET_TECHNICAL_LABELS = {
    "CAB-KURN-038-24": "Корпус КУРН-0,38-24 540×490×170 мм, металл",
    "CAB-KRN-18": "Корпус КРН-18 265×440×100 мм, металл",
    "CAB-KRN-12": "Корпус КРН-12 265×330×100 мм, металл",
    "CAB-KRN-24": "Корпус КРН-24 395×330×100 мм, металл",
    "CAB-SCHE-BI-900X900X120-M12": ("Встроенный ЩЭ, 900×900×120 мм, металл 1.2 мм"),
}
CABINET_SOURCE_TEMPLATE_CODES = {
    "ПР": "CAB-KURN-038-24",
    "Щоф": "CAB-KRN-18",
    "ШУ-Т2": "CAB-KRN-12",
    "ЩАО-1Ж": "CAB-KRN-12",
    "ЩАО-2Ж": "CAB-KRN-12",
    "ЩАО-3Ж": "CAB-KRN-12",
    "ЩО-1Ж": "CAB-KRN-12",
    "ЩО-2Ж": "CAB-KRN-12",
    "ЩО-3Ж": "CAB-KRN-12",
    "ЩС": "CAB-KRN-24",
}
UNRESOLVED_COMPONENT_MAPPING_REQUESTS: dict[str, str] = {}
PRICING_DECISION_ARTIFACT_SHA256 = (
    "777faed80c8ef92782378dd2a788160af8ad2252d8cb4f539560f15657a1d96e"
)
AD12_PRICE_MAPPING_DECISION_ARTIFACT_SHA256 = (
    "f67c0d79ec404a739ad5bdc3650a6259b9dc496a6f23ebffb7f29e7a9a24a17a"
)
AD12_PRICE_MAPPING_DECISION_ARTIFACT_SCHEMA = (
    "technical_ad12_price_mapping_human_decisions.v0.1"
)
AD12_PRICE_MAPPING_DECISION_ARTIFACT_STATUS = (
    "IGOR_AD12_SHARED_PRICE_MAPPING_APPROVED_NOT_APPLIED"
)
AD12_PRICE_MAPPING_DECISION_ID = "IGOR-AD12-SHARED-PRICE-MAPPING-2024-086-001"
RESOLVED_COMPONENT_MAPPING_PROVENANCE = {
    "COMPONENT-MAPPING-005": {
        "article": "D63N46ES16C100",
        "component_code": "EKF-AVDT63N-3P-N-C16-100MA-6KA-S",
        "pricing_decision_artifact_sha256": PRICING_DECISION_ARTIFACT_SHA256,
    },
    "COMPONENT-MAPPING-012": {
        "article": "DA32-6-16-30-ac-pro",
        "component_code": "EKF-AD32-1P-N",
        "pricing_decision_artifact_sha256": PRICING_DECISION_ARTIFACT_SHA256,
    },
    "COMPONENT-MAPPING-009": {
        "article": "DA12-16-30-bas",
        "component_code": "EKF-AD12-1P-N-C16-30MA-4P5KA",
        "row_draft_ids": (
            "ROW-DRAFT-0024",
            "ROW-DRAFT-0025",
            "ROW-DRAFT-0026",
            "ROW-DRAFT-0027",
        ),
        "human_decision_artifact_sha256": (
            "5d6e0de7af052c959abff015f41081c8bddc10e834fe4f971e8a7d2e60f19c46"
        ),
        "pricing_decision_artifact_sha256": (
            AD12_PRICE_MAPPING_DECISION_ARTIFACT_SHA256
        ),
        "pricing_decision_artifact_schema": (
            AD12_PRICE_MAPPING_DECISION_ARTIFACT_SCHEMA
        ),
        "pricing_decision_artifact_status": (
            AD12_PRICE_MAPPING_DECISION_ARTIFACT_STATUS
        ),
        "pricing_decision_id": AD12_PRICE_MAPPING_DECISION_ID,
        "direct_human_shared_price_decision": True,
        "ad32_fallback_used_for_ad12": False,
        "scope_expansion": False,
    },
    "COMPONENT-MAPPING-016": {
        "article": "DA12-16-30-bas",
        "component_code": "EKF-AD12-1P-N-C16-30MA-4P5KA",
        "row_draft_ids": (
            "ROW-DRAFT-0074",
            "ROW-DRAFT-0075",
        ),
        "human_decision_artifact_sha256": (
            "5d6e0de7af052c959abff015f41081c8bddc10e834fe4f971e8a7d2e60f19c46"
        ),
        "pricing_decision_artifact_sha256": (
            AD12_PRICE_MAPPING_DECISION_ARTIFACT_SHA256
        ),
        "pricing_decision_artifact_schema": (
            AD12_PRICE_MAPPING_DECISION_ARTIFACT_SCHEMA
        ),
        "pricing_decision_artifact_status": (
            AD12_PRICE_MAPPING_DECISION_ARTIFACT_STATUS
        ),
        "pricing_decision_id": AD12_PRICE_MAPPING_DECISION_ID,
        "direct_human_shared_price_decision": True,
        "ad32_fallback_used_for_ad12": False,
        "scope_expansion": False,
    },
}
APPROVED_MAPPING_005_SOURCE_LABELS = ("АВДТ-34, 4P, C16, 100мА",)


@dataclass(frozen=True)
class TechnicalSignature:
    apparatus_category: str
    poles: int
    rating_a: int
    residual_current_ma: int | None
    trip_curve: str | None
    install_type: str
    breaking_capacity_ka: Decimal | None = None


@dataclass(frozen=True)
class ApprovedComponentPriceMapping:
    signature: TechnicalSignature
    sheet_name: str
    row: int
    expected_label: str
    expected_material_price: int
    expected_work_price: int
    component_code: str | None = None
    strict_raw_label: bool = False


@dataclass(frozen=True)
class CabinetSignature:
    cabinet_code: str
    width_mm: int
    height_mm: int
    depth_mm: int
    material: str


@dataclass(frozen=True)
class ApprovedCabinetPriceMapping:
    signature: CabinetSignature
    sheet_name: str
    row: int
    expected_label: str
    expected_price: int


APPROVED_COMPONENT_PRICE_MAPPINGS = (
    ApprovedComponentPriceMapping(
        TechnicalSignature("mccb", 3, 63, None, None, "mccb_up_to_100a"),
        "ЩР",
        8,
        "ВА55/57/59, АМ1 3 полюсные от 16 до 63А",
        13000,
        1800,
    ),
    ApprovedComponentPriceMapping(
        TechnicalSignature("rcbo", 2, 16, 30, "C", "diff_1p_n"),
        "КРН",
        5,
        "УЗО АД-32 1Р+N до 63А EKF",
        4100,
        432,
    ),
    ApprovedComponentPriceMapping(
        TechnicalSignature("rcbo", 2, 20, 30, "C", "diff_1p_n"),
        "КРН",
        5,
        "УЗО АД-32 1Р+N до 63А EKF",
        4100,
        432,
    ),
    ApprovedComponentPriceMapping(
        TechnicalSignature("load_switch", 3, 32, None, None, "load_switch_3p"),
        "КРН",
        14,
        "ВН-32 3Р 16-25-40-63-80-100А",
        2750,
        540,
    ),
    ApprovedComponentPriceMapping(
        TechnicalSignature(
            "rcbo",
            2,
            16,
            30,
            "C",
            "diff_1p_n",
            Decimal("6"),
        ),
        "КРН",
        5,
        "УЗО АД-32 1Р+N до 63А EKF",
        4100,
        432,
        component_code="EKF-AD32-1P-N",
        strict_raw_label=True,
    ),
    ApprovedComponentPriceMapping(
        TechnicalSignature(
            "rcbo",
            2,
            16,
            30,
            "C",
            "diff_1p_n",
            Decimal("4.5"),
        ),
        "КРН",
        5,
        "УЗО АД-32 1Р+N до 63А EKF",
        4100,
        432,
        component_code="EKF-AD12-1P-N-C16-30MA-4P5KA",
        strict_raw_label=True,
    ),
    ApprovedComponentPriceMapping(
        TechnicalSignature("rcbo", 4, 16, 100, "C", "diff_3p_4p"),
        "КРН",
        28,
        "УЗО АД-32 1Р+N до 63А 100мА-300мА EKF ",
        8000,
        432,
        component_code="EKF-AVDT63N-3P-N-C16-100MA-6KA-S",
        strict_raw_label=True,
    ),
    ApprovedComponentPriceMapping(
        TechnicalSignature(
            "rcbo",
            4,
            16,
            100,
            "C",
            "diff_3p_4p",
            Decimal("6"),
        ),
        "КРН",
        28,
        "УЗО АД-32 1Р+N до 63А 100мА-300мА EKF ",
        8000,
        432,
        component_code="EKF-AVDT63N-3P-N-C16-100MA-6KA-S",
        strict_raw_label=True,
    ),
)

APPROVED_CABINET_PRICE_MAPPINGS = (
    ApprovedCabinetPriceMapping(
        CabinetSignature("ПР", 800, 600, 250, "metal"),
        "ЩР",
        8,
        "800х600х250",
        21336,
    ),
    ApprovedCabinetPriceMapping(
        CabinetSignature("КРН-36", 540, 330, 100, "metal"),
        "КРН",
        9,
        "Корпус КРН-36 540х330х100",
        9405,
    ),
)


@dataclass(frozen=True)
class CompositionRow:
    product_name: str
    cabinet_code: str
    consumables_factor: Decimal
    component_code: str
    component_qty: int
    install_type: str
    component_label: str | None = None
    cabinet_label: str | None = None
    component_mapping: ApprovedComponentPriceMapping | None = None
    cabinet_mapping: ApprovedCabinetPriceMapping | None = None
    technical_mapping_validated: bool = False


@dataclass
class PriceCalculationResult:
    price_workbook: Path
    input_csv: Path
    status: str = "FAIL"
    product_name: str | None = None
    input_rows_count: int = 0
    cabinet_code: str | None = None
    cabinet_label: str | None = None
    cabinet_price: int | None = None
    component_material_total: int | None = None
    work_total: int | None = None
    additional_materials_total: Decimal | None = None
    consumables_factor: Decimal | None = None
    base: Decimal | None = None
    total_preliminary_price: int | None = None
    red_flags: list[str] = field(default_factory=list)


def parse_args(argv: Sequence[str] | None = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "Calculate a read-only preliminary price draft from confirmed "
            "composition CSV using approved worksheet mappings."
        )
    )
    parser.add_argument(
        "--price-workbook",
        required=True,
        type=Path,
        help="Path to the approved .xlsx price workbook",
    )
    parser.add_argument(
        "--input-csv",
        required=True,
        type=Path,
        help="Path to confirmed semicolon-delimited composition CSV",
    )
    parser.add_argument(
        "--custom-cabinet-base-cost",
        type=int,
        help=("Checked positive integer base cost for " "CAB-SCHE-BI-900X900X120-M12"),
    )
    return parser.parse_args(argv)


def resolved(path: Path) -> Path:
    return path.expanduser().resolve(strict=False)


def add_red_flag(result: PriceCalculationResult, message: str) -> None:
    if message not in result.red_flags:
        result.red_flags.append(message)


def parse_positive_decimal(value: str) -> Decimal | None:
    if POSITIVE_DECIMAL_RE.fullmatch(value) is None:
        return None
    try:
        parsed = Decimal(value)
    except InvalidOperation:
        return None
    if parsed <= 0:
        return None
    return parsed


def parse_breaking_capacity_ka(component_label: str) -> Decimal | None:
    normalized = normalize_workbook_label(component_label)
    if normalized is None:
        return None
    match = re.search(
        r"\b(\d+(?:[.,]\d+)?)\s*[kк]\s*[aа]\b",
        normalized.casefold(),
    )
    return Decimal(match.group(1).replace(",", ".")) if match is not None else None


def parse_nominal_current_a(component_label: str) -> int | None:
    folded = component_label.casefold()
    patterns = (
        r"\b(?:i|и)[рpн]\s*=\s*(\d+)\s*[aа]\b",
        r"\b\d+\s*/\s*(\d+)\s*[aа]\b",
        r"\bc\s*(\d+)\b",
    )
    for pattern in patterns:
        match = re.search(pattern, folded)
        if match is not None:
            return int(match.group(1))
    matches = re.findall(r"\b(\d+)\s*[aа]\b", folded)
    return int(matches[-1]) if matches else None


def parse_component_signature(
    component_label: str,
    install_type: str,
) -> TechnicalSignature | None:
    normalized = normalize_workbook_label(component_label)
    if normalized is None:
        return None
    folded = normalized.casefold()
    poles_match = re.search(r"\b([1-4])\s*[pр]\b", folded)
    if poles_match is None:
        return None
    poles = int(poles_match.group(1))
    breaking_capacity = parse_breaking_capacity_ka(normalized)

    if install_type in {"modular_1p", "modular_2p", "modular_3p"} and (
        "автоматический выключатель" in folded or "выключатель автоматический" in folded
    ):
        expected_poles = int(install_type.removeprefix("modular_").removesuffix("p"))
        if poles != expected_poles or breaking_capacity is None:
            return None
        category = "mcb"
        rating = parse_nominal_current_a(normalized)
        rating_match = re.match(r"(\d+)", str(rating)) if rating is not None else None
        curve = "C" if re.search(r"(?:\(|\b)(?:c|с)(?=\d|\)|\b)", folded) else None
        if curve is None:
            return None
        residual_current = None
    elif install_type == "mccb_up_to_100a" and "автоматический выключатель" in folded:
        category = "mccb"
        rating_match = re.search(r"\b(\d+)\s*а\b", folded)
        curve = None
        residual_current = None
    elif install_type == "diff_1p_n" and ("авдт" in folded or "ад12" in folded):
        category = "rcbo"
        rating = parse_nominal_current_a(normalized)
        rating_match = re.match(r"(\d+)", str(rating)) if rating is not None else None
        residual_match = re.search(r"(?:/|,)\s*(\d+)\s*ма\b", folded)
        if (
            residual_match is None
            or re.search(r"(?:\b|\()(?:c|с)(?=\d|\)|\b)", folded) is None
        ):
            return None
        curve = "C"
        residual_current = int(residual_match.group(1))
    elif install_type == "diff_3p_4p" and any(
        normalized.startswith(label) for label in APPROVED_MAPPING_005_SOURCE_LABELS
    ):
        category = "rcbo"
        rating_match = re.search(r"\bc\s*(\d+)\b", folded)
        residual_match = re.search(r",\s*(\d+)\s*ма\b", folded)
        if residual_match is None:
            return None
        curve = "C"
        residual_current = int(residual_match.group(1))
    elif install_type == "load_switch_3p" and "выключатель нагрузки" in folded:
        category = "load_switch"
        rating_match = re.search(r"\b(\d+)\s*а\b", folded)
        curve = None
        residual_current = None
    else:
        return None

    if rating_match is None:
        return None
    return TechnicalSignature(
        apparatus_category=category,
        poles=poles,
        rating_a=int(rating_match.group(1)),
        residual_current_ma=residual_current,
        trip_curve=curve,
        install_type=install_type,
        breaking_capacity_ka=breaking_capacity,
    )


def technical_component_definition_matches(
    signature: TechnicalSignature | None,
    component_code: str,
    install_type: str,
) -> bool:
    definition = COMPONENT_DEFINITIONS.get(component_code)
    if (
        signature is None
        or definition is None
        or definition.install_type != install_type
        or signature.install_type != install_type
    ):
        return False
    if install_type.startswith("modular_"):
        expected_poles = {
            "EKF-VA47-29-1P": 1,
            "EKF-VA47-29-2P": 2,
            "EKF-VA47-29-3P": 3,
        }.get(component_code)
        return (
            signature.apparatus_category == "mcb"
            and signature.poles == expected_poles
            and signature.trip_curve == "C"
            and signature.breaking_capacity_ka is not None
        )
    if component_code == "EKF-AD12-1P-N-C16-30MA-4P5KA":
        return signature == TechnicalSignature(
            "rcbo",
            2,
            16,
            30,
            "C",
            "diff_1p_n",
            Decimal("4.5"),
        )
    return False


def exact_component_price_mapping_required(component_code: str | None) -> bool:
    return any(
        mapping.component_code == component_code
        for mapping in APPROVED_COMPONENT_PRICE_MAPPINGS
        if mapping.component_code is not None
    )


def technical_cabinet_definition_matches(
    cabinet_code: str,
    cabinet_label: str,
) -> bool:
    expected = CABINET_TECHNICAL_LABELS.get(cabinet_code)
    if expected is None:
        return False
    return (
        normalize_workbook_label(cabinet_label) == normalize_workbook_label(expected)
        and parse_cabinet_signature(cabinet_code, cabinet_label) is not None
    )


def parse_cabinet_signature(
    cabinet_code: str,
    cabinet_label: str,
) -> CabinetSignature | None:
    normalized = normalize_workbook_label(cabinet_label)
    if normalized is None:
        return None
    dimensions = re.search(
        r"(\d+)\s*[xх×*]\s*(\d+)\s*[xх×*]\s*(\d+)",
        normalized.casefold(),
    )
    if dimensions is None or "металл" not in normalized.casefold():
        return None
    return CabinetSignature(
        cabinet_code=normalize_workbook_label(cabinet_code) or cabinet_code,
        width_mm=int(dimensions.group(1)),
        height_mm=int(dimensions.group(2)),
        depth_mm=int(dimensions.group(3)),
        material="metal",
    )


def resolve_component_mapping(
    signature: TechnicalSignature,
    component_code: str | None = None,
) -> ApprovedComponentPriceMapping | None:
    exact_mapping_required = exact_component_price_mapping_required(component_code)
    matches = [
        mapping
        for mapping in APPROVED_COMPONENT_PRICE_MAPPINGS
        if mapping.signature == signature
        and (
            mapping.component_code == component_code
            if exact_mapping_required
            else mapping.component_code is None
            or mapping.component_code == component_code
        )
    ]
    return matches[0] if len(matches) == 1 else None


def signature_requires_component_code(signature: TechnicalSignature) -> bool:
    return any(
        mapping.signature == signature and mapping.component_code is not None
        for mapping in APPROVED_COMPONENT_PRICE_MAPPINGS
    )


def resolve_cabinet_mapping(
    signature: CabinetSignature,
) -> ApprovedCabinetPriceMapping | None:
    matches = [
        mapping
        for mapping in APPROVED_CABINET_PRICE_MAPPINGS
        if mapping.signature == signature
    ]
    return matches[0] if len(matches) == 1 else None


def load_composition_rows(result: PriceCalculationResult) -> list[CompositionRow]:
    path = result.input_csv
    if not path.is_file():
        add_red_flag(result, f"input CSV does not exist: {path}")
        return []
    if path.suffix.casefold() != ".csv":
        add_red_flag(result, "input composition suffix must be .csv")
        return []

    try:
        with path.open("r", encoding="utf-8-sig", newline="") as csv_file:
            reader = csv.reader(csv_file, delimiter=CSV_DELIMITER, strict=True)
            try:
                header = next(reader)
            except StopIteration:
                add_red_flag(result, "input composition CSV is empty")
                return []
            raw_rows = list(reader)
    except UnicodeDecodeError:
        add_red_flag(result, "input composition CSV must be valid UTF-8")
        return []
    except csv.Error:
        add_red_flag(result, "input composition CSV is invalid")
        return []
    except OSError:
        add_red_flag(result, "input composition CSV could not be read")
        return []

    result.input_rows_count = len(raw_rows)
    header_tuple = tuple(header)
    if header_tuple not in (REQUIRED_COLUMNS, TECHNICAL_COLUMNS):
        add_red_flag(
            result,
            "input header must exactly match a supported composition contract",
        )
        return []
    if not raw_rows:
        add_red_flag(result, "input composition CSV must contain at least one row")
        return []

    rows: list[CompositionRow] = []
    for row_number, values in enumerate(raw_rows, start=2):
        if len(values) != len(header_tuple):
            add_red_flag(result, f"row {row_number}: field count mismatch")
            continue

        row = dict(zip(header_tuple, values, strict=True))
        empty_columns = [column for column in header_tuple if row[column].strip() == ""]
        if empty_columns:
            add_red_flag(
                result,
                f"row {row_number}: required fields are empty: "
                f"{', '.join(empty_columns)}",
            )
            continue

        factor = parse_positive_decimal(row["consumables_factor"])
        if factor is None:
            add_red_flag(
                result,
                f"row {row_number}: consumables_factor must be a positive "
                "dot-decimal number",
            )
            continue

        quantity_text = row["component_qty"]
        if POSITIVE_INTEGER_RE.fullmatch(quantity_text) is None:
            add_red_flag(
                result,
                f"row {row_number}: component_qty must be a positive integer",
            )
            continue

        cabinet_code = row["cabinet_code"]
        component_code = row["component_code"]
        component_label: str | None = None
        cabinet_label: str | None = None
        component_mapping: ApprovedComponentPriceMapping | None = None
        cabinet_mapping: ApprovedCabinetPriceMapping | None = None
        technical_mapping_validated = False

        if header_tuple == TECHNICAL_COLUMNS:
            component_label = row["component_label"]
            cabinet_label = row["cabinet_label"]
            component_signature = parse_component_signature(
                component_label,
                row["install_type"],
            )
            component_mapping = (
                resolve_component_mapping(component_signature, component_code)
                if component_signature is not None
                else None
            )
            if component_mapping is None:
                legacy_definition = COMPONENT_DEFINITIONS.get(component_code)
                explicitly_validated = technical_component_definition_matches(
                    component_signature,
                    component_code,
                    row["install_type"],
                )
                legacy_compatible = (
                    parse_breaking_capacity_ka(component_label) is None
                    and legacy_definition is not None
                    and legacy_definition.install_type == row["install_type"]
                    and component_code != "EKF-AD32-1P-N"
                    and not exact_component_price_mapping_required(component_code)
                    and not (
                        component_signature is not None
                        and signature_requires_component_code(component_signature)
                    )
                )
                if not explicitly_validated and not legacy_compatible:
                    add_red_flag(
                        result,
                        f"row {row_number}: unknown or ambiguous technical "
                        f"component mapping for {component_label}; ask Igor",
                    )
                    continue
                technical_mapping_validated = explicitly_validated or legacy_compatible
            else:
                technical_mapping_validated = True

            cabinet_signature = parse_cabinet_signature(
                cabinet_code,
                cabinet_label,
            )
            cabinet_mapping = (
                resolve_cabinet_mapping(cabinet_signature)
                if cabinet_signature is not None
                else None
            )
            if cabinet_mapping is None and not technical_cabinet_definition_matches(
                cabinet_code,
                cabinet_label,
            ):
                add_red_flag(
                    result,
                    f"row {row_number}: unknown or ambiguous technical "
                    f"cabinet mapping for {cabinet_code} / {cabinet_label}; "
                    "ask Igor",
                )
                continue
        else:
            component_definition = COMPONENT_DEFINITIONS.get(component_code)
            if component_definition is None:
                add_red_flag(
                    result,
                    f"row {row_number}: component_code is not confirmed: "
                    f"{component_code}; ask Igor",
                )
                continue
            if row["install_type"] != component_definition.install_type:
                add_red_flag(
                    result,
                    f"row {row_number}: install_type does not match confirmed "
                    f"component map for {component_code}; ask Igor",
                )
                continue
            if cabinet_code not in CABINET_DEFINITIONS:
                add_red_flag(
                    result,
                    f"row {row_number}: cabinet_code is not confirmed: "
                    f"{cabinet_code}; ask Igor",
                )
                continue

        rows.append(
            CompositionRow(
                product_name=row["product_name"],
                cabinet_code=cabinet_code,
                consumables_factor=factor,
                component_code=component_code,
                component_qty=int(quantity_text),
                install_type=row["install_type"],
                component_label=component_label,
                cabinet_label=cabinet_label,
                component_mapping=component_mapping,
                cabinet_mapping=cabinet_mapping,
                technical_mapping_validated=technical_mapping_validated,
            )
        )

    if len(rows) != len(raw_rows):
        return []

    product_names = {row.product_name for row in rows}
    cabinet_codes = {row.cabinet_code for row in rows}
    factors = {row.consumables_factor for row in rows}
    if len(product_names) != 1:
        add_red_flag(result, "all rows must have the same product_name")
    if len(cabinet_codes) != 1:
        add_red_flag(result, "all rows must have the same cabinet_code")
    if len(factors) != 1:
        add_red_flag(result, "all rows must have the same consumables_factor")
    if result.red_flags:
        return []

    result.product_name = rows[0].product_name
    result.cabinet_code = rows[0].cabinet_code
    result.cabinet_label = (
        rows[0].cabinet_label
        if rows[0].cabinet_label is not None
        else CABINET_DEFINITIONS[rows[0].cabinet_code]
    )
    result.consumables_factor = rows[0].consumables_factor
    return rows


def normalize_workbook_label(value: Any) -> str | None:
    if not isinstance(value, str):
        return None
    return " ".join(value.replace("\xa0", " ").split())


def positive_integer_price(value: Any) -> int | None:
    if isinstance(value, bool) or value is None:
        return None
    if isinstance(value, int):
        return value if value > 0 else None
    if isinstance(value, float):
        return int(value) if value > 0 and value.is_integer() else None
    if isinstance(value, Decimal):
        integral = value.to_integral_value()
        return int(integral) if value > 0 and value == integral else None
    return None


def mapping_worksheet(
    workbook: Any,
    sheet_name: str,
    result: PriceCalculationResult,
) -> Any | None:
    if sheet_name.casefold() == FORBIDDEN_PRICE_SHEET_NAME.casefold():
        add_red_flag(result, "worksheet Прайс is forbidden for price lookup")
        return None
    try:
        return workbook[sheet_name]
    except KeyError:
        add_red_flag(result, f"required worksheet {sheet_name} was not found; ask Igor")
        return None


def read_approved_component_price(
    workbook: Any,
    mapping: ApprovedComponentPriceMapping,
    result: PriceCalculationResult,
) -> tuple[int, int] | None:
    worksheet = mapping_worksheet(workbook, mapping.sheet_name, result)
    if worksheet is None:
        return None
    raw_label = worksheet.cell(mapping.row, 1).value
    actual_label = (
        raw_label if mapping.strict_raw_label else normalize_workbook_label(raw_label)
    )
    expected_label = (
        mapping.expected_label
        if mapping.strict_raw_label
        else normalize_workbook_label(mapping.expected_label)
    )
    if actual_label != expected_label:
        add_red_flag(
            result,
            f"approved component mapping signature mismatch at "
            f"{mapping.sheet_name}!A{mapping.row}: expected "
            f"{mapping.expected_label}; ask Igor",
        )
        return None
    material_price = positive_integer_price(worksheet.cell(mapping.row, 2).value)
    work_price = positive_integer_price(worksheet.cell(mapping.row, 3).value)
    if (
        material_price != mapping.expected_material_price
        or work_price != mapping.expected_work_price
    ):
        add_red_flag(
            result,
            f"approved component mapping price mismatch at "
            f"{mapping.sheet_name}!B{mapping.row}:C{mapping.row}; ask Igor",
        )
        return None
    return material_price, work_price


def read_approved_cabinet_price(
    workbook: Any,
    mapping: ApprovedCabinetPriceMapping,
    result: PriceCalculationResult,
) -> int | None:
    worksheet = mapping_worksheet(workbook, mapping.sheet_name, result)
    if worksheet is None:
        return None
    actual_label = normalize_workbook_label(worksheet.cell(mapping.row, 12).value)
    expected_label = normalize_workbook_label(mapping.expected_label)
    if actual_label != expected_label:
        add_red_flag(
            result,
            f"approved cabinet mapping signature mismatch at "
            f"{mapping.sheet_name}!L{mapping.row}: expected "
            f"{mapping.expected_label}; ask Igor",
        )
        return None
    price = positive_integer_price(worksheet.cell(mapping.row, 13).value)
    if price != mapping.expected_price:
        add_red_flag(
            result,
            f"approved cabinet mapping price mismatch at "
            f"{mapping.sheet_name}!M{mapping.row}; ask Igor",
        )
        return None
    return price


def read_component_prices(
    worksheet: Any,
    required_codes: set[str],
    result: PriceCalculationResult,
) -> dict[str, tuple[int, int]]:
    label_to_code = {
        normalize_workbook_label(definition.workbook_label): code
        for code, definition in COMPONENT_DEFINITIONS.items()
        if code in required_codes and definition.workbook_label is not None
    }
    found: dict[str, tuple[int, int]] = {}

    for label_value, material_value, work_value in worksheet.iter_rows(
        min_row=1,
        max_row=MAX_LOOKUP_ROW,
        min_col=1,
        max_col=3,
        values_only=True,
    ):
        label = normalize_workbook_label(label_value)
        code = label_to_code.get(label)
        if code is None:
            continue
        if code in found:
            add_red_flag(
                result,
                f"duplicate component price row in КРН for {code}; ask Igor",
            )
            continue

        material_price = positive_integer_price(material_value)
        work_price = positive_integer_price(work_value)
        if material_price is None:
            add_red_flag(
                result,
                f"material price is missing or invalid in КРН for {code}; ask Igor",
            )
        if work_price is None:
            add_red_flag(
                result,
                f"work price is missing or invalid in КРН for {code}; ask Igor",
            )
        if material_price is not None and work_price is not None:
            found[code] = (material_price, work_price)

    for code in sorted(required_codes - found.keys()):
        if not any(code in flag for flag in result.red_flags):
            add_red_flag(
                result,
                f"component price row was not found in КРН for {code}; ask Igor",
            )
    return found


def read_cabinet_price(
    worksheet: Any,
    cabinet_code: str,
    result: PriceCalculationResult,
) -> int | None:
    expected_label = normalize_workbook_label(CABINET_DEFINITIONS[cabinet_code])
    found_prices: list[int] = []

    for label_value, price_value in worksheet.iter_rows(
        min_row=1,
        max_row=MAX_LOOKUP_ROW,
        min_col=12,
        max_col=13,
        values_only=True,
    ):
        if normalize_workbook_label(label_value) != expected_label:
            continue
        price = positive_integer_price(price_value)
        if price is None:
            add_red_flag(
                result,
                f"cabinet price is missing or invalid in КРН for "
                f"{cabinet_code}; ask Igor",
            )
        else:
            found_prices.append(price)

    if not found_prices:
        if not any(cabinet_code in flag for flag in result.red_flags):
            add_red_flag(
                result,
                f"cabinet price row was not found in КРН for "
                f"{cabinet_code}; ask Igor",
            )
        return None
    if len(found_prices) > 1:
        add_red_flag(
            result,
            f"duplicate cabinet price row in КРН for {cabinet_code}; ask Igor",
        )
        return None
    return found_prices[0]


def calculate_price_draft(
    price_workbook: Path,
    input_csv: Path,
    custom_cabinet_base_cost: int | None = None,
) -> PriceCalculationResult:
    result = PriceCalculationResult(
        price_workbook=resolved(price_workbook),
        input_csv=resolved(input_csv),
    )
    rows = load_composition_rows(result)
    if not rows:
        return result

    workbook_path = result.price_workbook
    if not workbook_path.is_file():
        add_red_flag(result, f"price workbook does not exist: {workbook_path}")
        return result
    if workbook_path.suffix.casefold() != ".xlsx":
        add_red_flag(result, "price workbook suffix must be .xlsx")
        return result

    workbook: Any | None = None
    try:
        workbook = load_workbook(
            filename=workbook_path,
            read_only=True,
            data_only=False,
            keep_links=False,
        )
        approved_component_prices: dict[
            ApprovedComponentPriceMapping, tuple[int, int]
        ] = {}
        for row in rows:
            mapping = row.component_mapping
            if mapping is not None and mapping not in approved_component_prices:
                price = read_approved_component_price(workbook, mapping, result)
                if price is not None:
                    approved_component_prices[mapping] = price

        worksheet = mapping_worksheet(workbook, KRN_SHEET_NAME, result)
        if worksheet is None:
            return result
        required_codes = {
            row.component_code for row in rows if row.component_mapping is None
        }
        component_prices: dict[str, tuple[int, int]] = {}
        if required_codes:
            component_prices = read_component_prices(
                worksheet,
                required_codes,
                result,
            )

        cabinet_mapping = rows[0].cabinet_mapping
        if any(row.cabinet_mapping != cabinet_mapping for row in rows):
            add_red_flag(result, "technical cabinet mapping is inconsistent; ask Igor")
            cabinet_price = None
        elif rows[0].cabinet_code == "CAB-SCHE-BI-900X900X120-M12":
            if custom_cabinet_base_cost is None or custom_cabinet_base_cost <= 0:
                add_red_flag(
                    result,
                    "checked custom ЩЭ cabinet base cost is required; ask Igor",
                )
                cabinet_price = None
            else:
                cabinet_price = custom_cabinet_base_cost
        elif cabinet_mapping is not None:
            cabinet_price = read_approved_cabinet_price(
                workbook,
                cabinet_mapping,
                result,
            )
        else:
            cabinet_price = read_cabinet_price(
                worksheet,
                rows[0].cabinet_code,
                result,
            )
    except (OSError, ValueError):  # fmt: skip
        add_red_flag(result, "price workbook could not be opened safely")
        return result
    finally:
        if workbook is not None:
            workbook.close()

    if result.red_flags or cabinet_price is None:
        return result

    material_total = sum(
        (
            approved_component_prices[row.component_mapping][0]
            if row.component_mapping is not None
            else component_prices[row.component_code][0]
        )
        * row.component_qty
        for row in rows
    )
    work_total = sum(
        (
            approved_component_prices[row.component_mapping][1]
            if row.component_mapping is not None
            else component_prices[row.component_code][1]
        )
        * row.component_qty
        for row in rows
    )
    factor = rows[0].consumables_factor
    additional_materials = Decimal(material_total) * (factor - Decimal("1"))
    base = (
        Decimal(cabinet_price)
        + Decimal(material_total)
        + additional_materials
        + Decimal(work_total)
    )
    total = int(
        (base * MATERIAL_MULTIPLIER * FINAL_MULTIPLIER).quantize(
            Decimal("1"),
            rounding=ROUND_HALF_UP,
        )
    )

    result.cabinet_price = cabinet_price
    result.component_material_total = material_total
    result.work_total = work_total
    result.additional_materials_total = additional_materials
    result.base = base
    result.total_preliminary_price = total
    result.status = "PASS"
    return result


def format_amount(value: int | Decimal | None) -> str:
    if value is None:
        return "not calculated"
    decimal_value = Decimal(value)
    if decimal_value == decimal_value.to_integral_value():
        return f"{int(decimal_value):,}".replace(",", " ")
    integer_part, fractional_part = format(decimal_value.normalize(), "f").split(".")
    grouped_integer = f"{int(integer_part):,}".replace(",", " ")
    return f"{grouped_integer}.{fractional_part}"


def format_red_flags(red_flags: Sequence[str]) -> list[str]:
    if not red_flags:
        return ["none"]
    return [f"- {flag}" for flag in red_flags]


def format_report(result: PriceCalculationResult) -> str:
    cabinet = "not resolved"
    if result.cabinet_code is not None and result.cabinet_label is not None:
        cabinet = f"{result.cabinet_code} / {result.cabinet_label}"
    factor = (
        f"{result.consumables_factor:.2f}"
        if result.consumables_factor is not None
        else "not resolved"
    )
    lines = [
        "PRICE_CALCULATION_DRAFT_REPORT_START",
        "",
        "Status:",
        result.status,
        "",
        "Mode:",
        "read-only preliminary price draft",
        "",
        "Product name:",
        result.product_name or "not resolved",
        "",
        "Workbook path:",
        str(result.price_workbook),
        "",
        "Input CSV path:",
        str(result.input_csv),
        "",
        "Input rows count:",
        str(result.input_rows_count),
        "",
        "Cabinet:",
        cabinet,
        "",
        "Cabinet price:",
        format_amount(result.cabinet_price),
        "",
        "Component material total:",
        format_amount(result.component_material_total),
        "",
        "Work total:",
        format_amount(result.work_total),
        "",
        "Additional materials total:",
        format_amount(result.additional_materials_total),
        "",
        "Consumables factor:",
        factor,
        "",
        "Base:",
        format_amount(result.base),
        "",
        "Total preliminary price:",
        format_amount(result.total_preliminary_price),
        "",
        "Red flags:",
    ]
    lines.extend(format_red_flags(result.red_flags))
    lines.extend(
        [
            "",
            "Commercial status:",
            "preliminary only; PASS is not commercial approval",
            "",
            "Before transfer to commercial CSV:",
            "Igor approval required",
            "",
            "Manual Igor check:",
            "required",
            "",
            "Human Approval:",
            "required before using price in commercial КП",
            "",
            "PRICE_CALCULATION_DRAFT_REPORT_END",
        ]
    )
    return "\n".join(lines)


def main(argv: Sequence[str] | None = None) -> int:
    args = parse_args(argv)
    result = calculate_price_draft(
        args.price_workbook,
        args.input_csv,
        custom_cabinet_base_cost=args.custom_cabinet_base_cost,
    )
    print(format_report(result))
    return 0 if result.status == "PASS" else 1


if __name__ == "__main__":
    raise SystemExit(main())
