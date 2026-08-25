"""Build the exact SHU-T2 RT-820 Invoice 519 pricing-profile successor.

The builder performs a controlled replacement inside a full immutable profile
snapshot.  It never calculates, approves, or applies a price.
"""

from __future__ import annotations

import argparse
import copy
import hashlib
import json
import os
import tempfile
from collections.abc import Mapping, Sequence
from dataclasses import dataclass
from pathlib import Path
from typing import Any, cast

from openpyxl import load_workbook  # type: ignore[import-untyped]

PROJECT_ID = "2024/086"
PROFILE_SCHEMA = "technical_invoice519_pricing_profile_human_decisions.v0.1"
PROFILE_STATUS = "IGOR_INVOICE519_PRICING_PROFILE_APPROVED_NOT_APPLIED"
TECHNICAL_SCHEMA = "price_calculator_input_draft.v0.2"
TECHNICAL_STATUS = "V02_TECHNICAL_COMPLETION_APPLIED_NOT_PRICED"
TECHNICAL_CONTRACT = "controlled_shu_t2_rt820_technical_successor.v0.1"
DECISION_SCHEMA = "technical_shu_t2_rt820_scope_human_decision.v0.1"
DECISION_ID = "IGOR-SHU-T2-RT820-SCOPE-2024-086-001"
DECISION_STATUS = "IGOR_SHU_T2_RT820_SCOPE_APPROVED_NOT_APPLIED"
AUTHORITY = "IGOR_DIRECT_HUMAN_APPROVAL"
APPLICATION_STATUS = "NOT_APPLIED"
SUCCESSOR_CONTRACT = "controlled_shu_t2_rt820_pricing_profile_successor.v0.1"
PUBLICATION_AUTHORIZATION = (
    "IGOR_SHU_T2_RT820_PRICING_PROFILE_SUCCESSOR_PUBLICATION_AUTHORIZED"
)
OUTPUT_FILENAME = "invoice519-pricing-profile-shu-t2-rt820-successor.json"
REPO_ROOT = Path(__file__).resolve().parents[1]

PARENT_PROFILE = Path(
    r"C:\Users\IgorN\Documents\production_ai_cases\CASE-QF-PROJECT-2024-086-"
    "SHU-T1-PRICING-PROFILE-SUCCESSOR-20260820-001\\"
    "invoice519-pricing-profile-additive-successor.json"
)
PARENT_PROFILE_SHA256 = (
    "10d4301923b1ae141ae228c319f38e7281810e40c6990f0b2d533e9e20763424"
)
TECHNICAL_SUCCESSOR = Path(
    r"C:\Users\IgorN\Documents\production_ai_cases\CASE-QF-PROJECT-2024-086-"
    "SHU-T2-RT820-TECHNICAL-SUCCESSOR-20260824-001\\"
    "price-calculator-input-v0.2-completed-shu-t2-rt820-successor.json"
)
TECHNICAL_SUCCESSOR_SHA256 = (
    "c27c2c3032699cb07c981aeb4af429b27ec18180225319f45ce65ab77fedee44"
)
HUMAN_DECISION = Path(
    r"C:\Users\IgorN\Documents\production_ai_cases\CASE-QF-PROJECT-2024-086-"
    "SHU-T2-RT820-SCOPE-DECISION-20260820-001\\"
    "technical-shu-t2-rt820-scope-human-decision-v0.1.json"
)
HUMAN_DECISION_SHA256 = (
    "92a79401591fa6202af493848dd979a227ae20da8e66b8dea6e8084fc80c2ac6"
)
PRICING_WORKBOOK = (
    Path(r"C:\Users\IgorN\Documents\invoice_quote_filler_data\prices\current")
    / "Таблица 05.01.2026 верная.xlsx"
)
PRICING_WORKBOOK_SHA256 = (
    "f8bd69da1f61612d3853e608333486dcd3b6ecd572cd98beb2247c6accb31b5f"
)

TARGET_GROUP_ID = "CABINET-GROUP-003"
SHU_T1_GROUP_ID = "CABINET-GROUP-015"
OLD_FINGERPRINT = "99db78a5c3c7688a9e2cebbbe57f41489af797bbc61f2b1fa38492a42329cb79"
NEW_FINGERPRINT = "4b5cf23236653dfd33e27eefa8034ad2a779b5e2b40f0adc972ee49912dbc0ec"
NEW_COMPONENTS = [
    {
        "component_code": "EKF-AD12-1P-N-C16-30MA-4P5KA",
        "component_qty": 1,
        "install_type": "diff_1p_n",
    },
    {
        "component_code": "EKF-RT-820",
        "component_qty": 1,
        "install_type": "temperature_relay_din_2mod",
    },
    {
        "component_code": "EKF-VA47-29-2P",
        "component_qty": 1,
        "install_type": "modular_2p",
    },
]
POSITION_SCOPE = (
    ("PRICE-POSITION-009", "TFE-016", "ROW-DRAFT-0113", 112),
    ("PRICE-POSITION-023", "TFE-041", "ROW-DRAFT-0114", 113),
    ("PRICE-POSITION-035", "TFE-061", "ROW-DRAFT-0115", 114),
    ("PRICE-POSITION-047", "TFE-083", "ROW-DRAFT-0116", 115),
)
SHU_T1_POSITION_IDS = (
    "PRICE-POSITION-052",
    "PRICE-POSITION-053",
    "PRICE-POSITION-054",
    "PRICE-POSITION-055",
)
EXPECTED_COVERAGE = {
    "technical_cabinet_groups": 15,
    "section_aware_pricing_positions": 55,
    "physical_cabinets": 137,
    "composition_fingerprints": 11,
}
EXPECTED_AUTHORITY = {
    "authority": AUTHORITY,
    "decision_source": "DIRECT_IGOR_INSTRUCTION_2026-08-14",
    "no_scope_expansion": True,
}
EXPECTED_SAFETY_FLAGS = {
    "pricing_profile_decision_recorded": True,
    "pricing_profile_applied": False,
    "current_scope_pricing_calculated": False,
    "reserved_formula_rules_applied": False,
    "calculator_run_authorized": False,
    "checked_calculator_run_authorized": False,
    "quote_generation_authorized": False,
    "price_approval_for_client": False,
    "lead_time_approved": False,
    "client_send_authorized": False,
    "procurement_authorized": False,
    "production_authorized": False,
    "scope_expansion": False,
}
EXPECTED_NON_APPROVALS = {
    "project_total_approved": False,
    "remaining_current_position_prices_approved": False,
    "reserved_family_prices_approved": False,
    "lead_time_approved": False,
    "final_invoice_or_quote_approved": False,
    "client_send_authorized": False,
}


class ContractError(ValueError):
    """Raised when any exact contract boundary fails closed."""


class DuplicateJsonKeyError(ValueError):
    """Raised when an input JSON object repeats a key."""


@dataclass(frozen=True)
class InputPaths:
    parent_profile: Path
    technical_successor: Path
    human_decision: Path
    pricing_workbook: Path


@dataclass(frozen=True)
class ExpectedShas:
    parent_profile: str
    technical_successor: str
    human_decision: str
    pricing_workbook: str


@dataclass(frozen=True)
class LoadedInput:
    role: str
    path: Path
    expected_sha256: str
    raw: bytes
    value: Mapping[str, Any] | None = None


@dataclass(frozen=True)
class LoadedInputs:
    parent: LoadedInput
    technical: LoadedInput
    decision: LoadedInput
    workbook: LoadedInput


@dataclass(frozen=True)
class PublicationResult:
    sha256: str
    size: int
    encoded: bytes


def require(condition: bool, message: str) -> None:
    if not condition:
        raise ContractError(message)


def reject_duplicate_keys(pairs: list[tuple[str, Any]]) -> dict[str, Any]:
    result: dict[str, Any] = {}
    for key, value in pairs:
        if key in result:
            raise DuplicateJsonKeyError(key)
        result[key] = value
    return result


def sha256_bytes(payload: bytes) -> str:
    return hashlib.sha256(payload).hexdigest()


def resolved(path: Path) -> Path:
    return path.expanduser().resolve(strict=False)


def load_json(path: Path, description: str) -> tuple[dict[str, Any], bytes]:
    try:
        raw = path.read_bytes()
    except OSError as exc:
        raise ContractError(f"{description} could not be read: {exc}") from exc
    try:
        value = json.loads(raw.decode("utf-8"), object_pairs_hook=reject_duplicate_keys)
    except UnicodeDecodeError as exc:
        raise ContractError(f"{description} must be UTF-8") from exc
    except DuplicateJsonKeyError as exc:
        raise ContractError(f"{description} contains duplicate key: {exc}") from exc
    except json.JSONDecodeError as exc:
        raise ContractError(f"{description} is malformed JSON") from exc
    require(isinstance(value, dict), f"{description} root must be an object")
    return cast(dict[str, Any], value), raw


def serialize(payload: Mapping[str, Any]) -> bytes:
    return (json.dumps(payload, ensure_ascii=False, indent=2) + "\n").encode("utf-8")


def fingerprint(components: Sequence[Mapping[str, Any]]) -> str:
    ordered = sorted(
        (dict(component) for component in components),
        key=lambda item: (
            item["component_code"],
            item["component_qty"],
            item["install_type"],
        ),
    )
    encoded = json.dumps(
        ordered, ensure_ascii=False, sort_keys=True, separators=(",", ":")
    ).encode("utf-8")
    return sha256_bytes(encoded)


def _validate_exact_cli_bindings(paths: InputPaths, shas: ExpectedShas) -> None:
    expected = (
        (
            paths.parent_profile,
            shas.parent_profile,
            PARENT_PROFILE,
            PARENT_PROFILE_SHA256,
        ),
        (
            paths.technical_successor,
            shas.technical_successor,
            TECHNICAL_SUCCESSOR,
            TECHNICAL_SUCCESSOR_SHA256,
        ),
        (
            paths.human_decision,
            shas.human_decision,
            HUMAN_DECISION,
            HUMAN_DECISION_SHA256,
        ),
        (
            paths.pricing_workbook,
            shas.pricing_workbook,
            PRICING_WORKBOOK,
            PRICING_WORKBOOK_SHA256,
        ),
    )
    for actual_path, actual_sha, expected_path, expected_sha in expected:
        require(
            resolved(actual_path) == resolved(expected_path),
            "input path binding mismatch",
        )
        require(actual_sha == expected_sha, "input expected SHA binding mismatch")


def _load_bound_json(role: str, path: Path, expected_sha: str) -> LoadedInput:
    value, raw = load_json(path, role)
    require(sha256_bytes(raw) == expected_sha, f"{role} SHA-256 mismatch")
    return LoadedInput(role, resolved(path), expected_sha, raw, value)


def _load_bound_workbook(path: Path, expected_sha: str) -> LoadedInput:
    try:
        raw = path.read_bytes()
    except OSError as exc:
        raise ContractError(f"pricing workbook could not be read: {exc}") from exc
    require(sha256_bytes(raw) == expected_sha, "pricing workbook SHA-256 mismatch")
    workbook: Any | None = None
    try:
        workbook = load_workbook(
            path, read_only=True, data_only=False, keep_links=False
        )
        require("КРН" in workbook.sheetnames, "pricing workbook КРН sheet missing")
        sheet = workbook["КРН"]
        require(
            [sheet.cell(19, column).value for column in range(1, 4)]
            == ["Терморегулятор RT-820", 15000, 900],
            "pricing workbook exact КРН!A19:C19 contract mismatch",
        )
    except ContractError:
        raise
    except (OSError, ValueError, KeyError) as exc:  # fmt: skip
        raise ContractError("pricing workbook could not be validated safely") from exc
    finally:
        if workbook is not None:
            workbook.close()
    return LoadedInput("pricing_workbook", resolved(path), expected_sha, raw)


def _scope(profile: Mapping[str, Any]) -> Mapping[str, Any]:
    value = profile.get("current_completed_technical_scope")
    require(isinstance(value, Mapping), "pricing profile current scope missing")
    return cast(Mapping[str, Any], value)


def validate_parent_profile(profile: Mapping[str, Any]) -> None:
    require(
        profile.get("schema_version") == PROFILE_SCHEMA,
        "parent profile schema mismatch",
    )
    require(profile.get("project_id") == PROJECT_ID, "parent profile project mismatch")
    require(profile.get("status") == PROFILE_STATUS, "parent profile status mismatch")
    require(profile.get("authority") == EXPECTED_AUTHORITY, "parent authority mismatch")
    require(
        profile.get("application_status") == APPLICATION_STATUS,
        "parent profile application mismatch",
    )
    require(
        profile.get("immutable_state") == {"immutable": True, "no_overwrite": True},
        "parent profile immutability mismatch",
    )
    require(profile.get("scope_expansion") is False, "parent scope expansion mismatch")
    require(
        profile.get("safety_flags") == EXPECTED_SAFETY_FLAGS,
        "parent safety flags mismatch",
    )
    require(
        profile.get("non_approvals") == EXPECTED_NON_APPROVALS,
        "parent non-approvals mismatch",
    )
    additive = profile.get("additive_successor")
    require(
        isinstance(additive, Mapping)
        and additive.get("contract")
        == "controlled_additive_invoice519_pricing_profile_successor.v0.1"
        and additive.get("candidate_project_total_kzt") == 11841516
        and additive.get("price_approval_status") == "REQUIRES_IGOR_PRICE_APPROVAL",
        "parent SHU-T1 successor contract mismatch",
    )
    scope = _scope(profile)
    require(
        scope.get("coverage")
        == {
            "technical_cabinet_groups": 15,
            "section_aware_pricing_positions": 55,
            "physical_cabinets": 137,
            "composition_fingerprints": 12,
        },
        "parent coverage mismatch",
    )
    groups = scope.get("cabinet_groups")
    positions = scope.get("pricing_positions")
    fingerprints = scope.get("composition_fingerprints")
    require(isinstance(groups, list) and len(groups) == 15, "parent groups mismatch")
    require(
        isinstance(positions, list) and len(positions) == 55,
        "parent positions mismatch",
    )
    require(
        isinstance(fingerprints, list) and len(fingerprints) == 12,
        "parent fingerprints mismatch",
    )
    group3 = groups[2]
    require(
        isinstance(group3, Mapping)
        and group3.get("cabinet_group_id") == TARGET_GROUP_ID
        and group3.get("product_name") == "ШУ-Т2"
        and len(group3.get("row_draft_ids", [])) == 8,
        "parent SHU-T2 group mismatch",
    )
    by_position = {
        item.get("pricing_position_id"): item
        for item in positions
        if isinstance(item, Mapping)
    }
    for position_id, source_id, _row_id, _index in POSITION_SCOPE:
        position = by_position.get(position_id)
        require(
            isinstance(position, Mapping)
            and position.get("source_position_id") == source_id
            and position.get("product_name") == "ШУ-Т2"
            and position.get("composition_fingerprint_sha256") == OLD_FINGERPRINT
            and position.get("approved_unit_price_kzt") is None
            and position.get("approved_unit_price_decision_status")
            == "NOT_CALCULATED_NOT_APPROVED",
            f"parent position mismatch: {position_id}",
        )
    by_fingerprint = {
        item.get("fingerprint_sha256"): item
        for item in fingerprints
        if isinstance(item, Mapping)
    }
    require(
        set((OLD_FINGERPRINT, NEW_FINGERPRINT)).issubset(by_fingerprint),
        "parent fingerprint contract mismatch",
    )
    require(
        fingerprint(NEW_COMPONENTS) == NEW_FINGERPRINT,
        "independent RT-820 fingerprint mismatch",
    )


def _expected_positions() -> list[dict[str, Any]]:
    evidence = (
        ("COMP-031", "COMP-034"),
        ("COMP-085", "COMP-088"),
        ("COMP-128", "COMP-131"),
        ("COMP-178", "COMP-181"),
    )
    sections = ("10", "12", "14", "16")
    return [
        {
            "section": section,
            "technical_position_id": source_id,
            "pricing_position_id": position_id,
            "relay_evidence_id": relay,
            "sensor_evidence_id": sensor,
            "physical_multiplicity": 1,
        }
        for (position_id, source_id, _row, _index), section, (relay, sensor) in zip(
            POSITION_SCOPE, sections, evidence, strict=True
        )
    ]


def _expected_appended_rows() -> list[dict[str, Any]]:
    return [
        {
            "row_id": row_id,
            "cabinet_group_id": TARGET_GROUP_ID,
            "calculator_values": {
                "product_name": "ШУ-Т2",
                "cabinet_code": "CAB-KRN-12",
                "consumables_factor": 1.2,
                "component_code": "EKF-RT-820",
                "component_qty": 1,
                "install_type": "temperature_relay_din_2mod",
            },
            "source_quantity": {
                "decision_id": DECISION_ID,
                "decision_kind": "DIRECT_PER_CABINET_COMPLETE_SET",
                "technical_position_id": expected["technical_position_id"],
                "pricing_position_id": expected["pricing_position_id"],
                "section": expected["section"],
                "quantity_per_individual_cabinet": 1,
                "physical_multiplicity": 1,
                "applies_once_per_cabinet": True,
                "multiply_by_member_count": False,
                "scope_expansion": False,
            },
            "source_component_evidence_ids": [
                expected["relay_evidence_id"],
                expected["sensor_evidence_id"],
            ],
            "approved_signature": {
                "manufacturer": "EKF",
                "product": "Реле температуры RT-820 EKF PROxima",
                "manufacturer_article": "RT-820",
                "supply_form": (
                    "ONE_TEMPERATURE_RELAY_WITH_ONE_EXTERNAL_TEMPERATURE_SENSOR"
                ),
                "module_width_din": 2,
                "TST05_evidence_included_as_provenance_only": True,
                "TST05_separate_component_row": False,
            },
            "mapping_status": "APPROVED_HUMAN_DECISIONS_APPLIED",
            "component_label": (
                "Реле температуры RT-820 EKF PROxima с внешним датчиком"
            ),
        }
        for (_position_id, _source_id, row_id, _index), expected in zip(
            POSITION_SCOPE, _expected_positions(), strict=True
        )
    ]


def validate_decision(decision: Mapping[str, Any], parent_sha: str) -> None:
    require(
        decision.get("schema_version") == DECISION_SCHEMA,
        "Human Decision schema mismatch",
    )
    require(
        decision.get("artifact_type") == "IMMUTABLE_HUMAN_DECISION_CAPTURE",
        "Human Decision artifact type mismatch",
    )
    require(decision.get("project_id") == PROJECT_ID, "Human Decision project mismatch")
    require(decision.get("decision_id") == DECISION_ID, "Human Decision ID mismatch")
    require(decision.get("status") == DECISION_STATUS, "Human Decision status mismatch")
    require(decision.get("authority") == AUTHORITY, "Human Decision authority mismatch")
    require(
        decision.get("application_status") == APPLICATION_STATUS,
        "Human Decision application mismatch",
    )
    bindings = decision.get("input_bindings")
    pricing_binding = (
        next(
            (
                item
                for item in bindings
                if isinstance(item, Mapping) and item.get("role") == "pricing_profile"
            ),
            None,
        )
        if isinstance(bindings, list)
        else None
    )
    require(
        isinstance(pricing_binding, Mapping)
        and pricing_binding.get("expected_sha256") == parent_sha
        and pricing_binding.get("actual_sha256") == parent_sha,
        "Human Decision parent pricing binding mismatch",
    )
    scope = decision.get("exact_scope")
    require(
        isinstance(scope, Mapping)
        and scope.get("product") == "ШУ-Т2"
        and scope.get("cabinet_group_id") == TARGET_GROUP_ID
        and scope.get("positions") == _expected_positions()
        and scope.get("source_evidence_row_count") == 8
        and scope.get("future_component_row_count") == 4,
        "Human Decision exact scope mismatch",
    )
    require(
        decision.get("rt820_contract")
        == {
            "component_code": "EKF-RT-820",
            "component_qty_per_physical_cabinet": 1,
            "install_type": "temperature_relay_din_2mod",
            "module_width": 2,
            "source_range": "КРН!A19:C19",
            "source_label": "Терморегулятор RT-820",
            "material_kzt": 15000,
            "work_kzt": 900,
            "generic_work_432_prohibited": True,
            "family_fallback_prohibited": True,
            "fuzzy_fallback_prohibited": True,
            "similar_relay_fallback_prohibited": True,
        },
        "Human Decision RT-820 contract mismatch",
    )
    bundle = decision.get("bundle_semantics")
    require(
        isinstance(bundle, Mapping)
        and bundle.get("tst05_provenance_only") is True
        and bundle.get("separate_tst05_component_row") is False
        and bundle.get("separate_tst05_material_charge") is False
        and bundle.get("separate_tst05_work_charge") is False
        and bundle.get("separate_tst05_pricing_row") is False,
        "Human Decision TST05 semantics mismatch",
    )
    supersession = decision.get("supersession")
    require(
        isinstance(supersession, Mapping)
        and supersession.get("outside_cabinet_exclusion_count_must_be_derived") is True
        and supersession.get("outside_cabinet_exclusion_count_override_prohibited")
        is True
        and supersession.get("shu_t1_unchanged") is True,
        "Human Decision exclusion/SHU-T1 safety mismatch",
    )


def validate_technical(technical: Mapping[str, Any], decision_sha: str) -> None:
    require(
        technical.get("schema_version") == TECHNICAL_SCHEMA, "technical schema mismatch"
    )
    completion = technical.get("completion")
    require(
        isinstance(completion, Mapping)
        and completion.get("status") == TECHNICAL_STATUS
        and completion.get("scope")
        == {
            "component_groups": 35,
            "rows": "116/116",
            "cabinet_groups": "15/15",
            "duplicate_component_membership": 0,
            "duplicate_cabinet_membership": 0,
            "scope_expansion": False,
        },
        "technical completion scope mismatch",
    )
    source = technical.get("source")
    metadata = (
        source.get("shu_t2_rt820_technical_successor")
        if isinstance(source, Mapping)
        else None
    )
    human = metadata.get("human_decision") if isinstance(metadata, Mapping) else None
    projection = (
        metadata.get("technical_projection") if isinstance(metadata, Mapping) else None
    )
    require(
        isinstance(metadata, Mapping)
        and metadata.get("contract") == TECHNICAL_CONTRACT
        and metadata.get("append_only") is True
        and metadata.get("scope_expansion") is False
        and isinstance(human, Mapping)
        and human.get("sha256") == decision_sha
        and human.get("decision_id") == DECISION_ID
        and human.get("application_status") == APPLICATION_STATUS,
        "technical successor metadata mismatch",
    )
    expected_evidence = [
        evidence_id
        for item in _expected_positions()
        for evidence_id in (item["relay_evidence_id"], item["sensor_evidence_id"])
    ]
    require(
        isinstance(projection, Mapping)
        and projection.get("row_ids") == [item[2] for item in POSITION_SCOPE]
        and projection.get("evidence_count") == 8
        and projection.get("evidence_ids") == expected_evidence
        and projection.get("outside_cabinet_membership_asserted") is False
        and projection.get("outside_cabinet_count_transition_asserted") is False,
        "technical projection/evidence mismatch",
    )
    pricing = metadata.get("rt820_pricing_provenance_only")
    require(
        isinstance(pricing, Mapping)
        and pricing.get("source_range") == "КРН!A19:C19"
        and pricing.get("material_kzt") == 15000
        and pricing.get("work_kzt") == 900
        and pricing.get("pricing_calculation_executed") is False
        and all(
            pricing.get(key) is True
            for key in (
                "generic_work_432_prohibited",
                "family_fallback_prohibited",
                "fuzzy_fallback_prohibited",
                "similar_relay_fallback_prohibited",
            )
        ),
        "technical RT-820 pricing provenance mismatch",
    )
    groups = technical.get("cabinet_groups")
    rows_value = technical.get("calculator_input_format")
    rows = rows_value.get("row_drafts") if isinstance(rows_value, Mapping) else None
    require(isinstance(groups, list) and len(groups) == 15, "technical groups mismatch")
    require(isinstance(rows, list) and len(rows) == 116, "technical rows mismatch")
    require(
        isinstance(groups[2], Mapping)
        and groups[2].get("row_draft_ids")[-4:] == [item[2] for item in POSITION_SCOPE]
        and len(groups[2].get("row_draft_ids", [])) == 12,
        "technical SHU-T2 group mismatch",
    )
    require(
        isinstance(groups[14], Mapping)
        and groups[14].get("cabinet_group_id") == SHU_T1_GROUP_ID
        and groups[14].get("row_draft_ids")
        == ["ROW-DRAFT-0110", "ROW-DRAFT-0111", "ROW-DRAFT-0112"],
        "technical SHU-T1 integrity mismatch",
    )
    expected_rows = _expected_appended_rows()
    require(rows[-4:] == expected_rows, "technical appended row envelope mismatch")
    appended_evidence = [
        evidence_id
        for row in rows[-4:]
        for evidence_id in row["source_component_evidence_ids"]
    ]
    require(
        appended_evidence == expected_evidence
        and len(appended_evidence) == len(set(appended_evidence)) == 8,
        "technical appended evidence membership mismatch",
    )
    require(
        sum(
            isinstance(row, Mapping)
            and isinstance(row.get("calculator_values"), Mapping)
            and row["calculator_values"].get("component_code") == "EKF-RT-820"
            for row in rows
        )
        == 5,
        "technical RT-820 row count mismatch",
    )
    require(
        all(
            "TST05"
            not in str(
                row.get("calculator_values", {}).get("component_code", "")
                if isinstance(row, Mapping)
                else ""
            )
            for row in rows
        ),
        "separate TST05 row is forbidden",
    )


def load_and_validate_inputs(paths: InputPaths, shas: ExpectedShas) -> LoadedInputs:
    _validate_exact_cli_bindings(paths, shas)
    parent = _load_bound_json(
        "parent pricing profile", paths.parent_profile, shas.parent_profile
    )
    technical = _load_bound_json(
        "technical successor", paths.technical_successor, shas.technical_successor
    )
    decision = _load_bound_json(
        "Human Decision", paths.human_decision, shas.human_decision
    )
    workbook = _load_bound_workbook(paths.pricing_workbook, shas.pricing_workbook)
    validate_parent_profile(cast(Mapping[str, Any], parent.value))
    validate_decision(cast(Mapping[str, Any], decision.value), parent.expected_sha256)
    validate_technical(
        cast(Mapping[str, Any], technical.value), decision.expected_sha256
    )
    return LoadedInputs(parent, technical, decision, workbook)


def _input_binding(item: LoadedInput, schema_or_type: str) -> dict[str, Any]:
    return {
        "path": str(item.path),
        "sha256": item.expected_sha256,
        "schema_or_type": schema_or_type,
    }


def successor_metadata(loaded: LoadedInputs) -> dict[str, Any]:
    return {
        "contract": SUCCESSOR_CONTRACT,
        "project_id": PROJECT_ID,
        "parent_pricing_profile": _input_binding(loaded.parent, PROFILE_SCHEMA),
        "technical_successor": {
            **_input_binding(loaded.technical, TECHNICAL_SCHEMA),
            "contract": TECHNICAL_CONTRACT,
        },
        "human_decision": {
            **_input_binding(loaded.decision, DECISION_SCHEMA),
            "decision_id": DECISION_ID,
            "status": DECISION_STATUS,
            "authority": AUTHORITY,
            "application_status": APPLICATION_STATUS,
        },
        "pricing_workbook": {
            **_input_binding(
                loaded.workbook,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            ),
            "source_range": "КРН!A19:C19",
            "source_label": "Терморегулятор RT-820",
            "material_kzt": 15000,
            "work_kzt": 900,
        },
        "controlled_replacement": True,
        "append_only": False,
        "scope_expansion": False,
        "coverage_transition": {
            "cabinet_groups": "15->15",
            "pricing_positions": "55->55",
            "physical_cabinets": "137->137",
            "technical_rows": "112->116",
            "composition_fingerprints": "12->11",
        },
        "old_fingerprint_removed": OLD_FINGERPRINT,
        "merged_fingerprint": NEW_FINGERPRINT,
        "tst05_provenance_only": True,
        "generic_work_432_prohibited": True,
        "fallback_prohibited": True,
        "pricing_calculation_executed": False,
        "approved_unit_price_kzt": None,
        "application_status": APPLICATION_STATUS,
        "price_approval_status": "REQUIRES_IGOR_PRICE_APPROVAL",
        "preliminary_not_approved_invariants": {
            "status": "NOT_CALCULATED_NOT_APPROVED",
            "X_cabinet_base_kzt": 6936,
            "G_material_kzt": 20450,
            "H_work_kzt": 1764,
            "formula_base_kzt": 33240,
            "raw_unit_candidate_kzt": "53762.72702586206896551724138",
            "unit_candidate_kzt": 53763,
            "four_position_candidate_kzt": 215052,
            "delta_from_prior_checked_candidate_kzt": 122276,
            "preliminary_project_candidate_kzt": 11963792,
            "approved": False,
            "applied": False,
        },
    }


def build_successor_payload(loaded: LoadedInputs) -> dict[str, Any]:
    payload = copy.deepcopy(cast(Mapping[str, Any], loaded.parent.value))
    scope = cast(dict[str, Any], payload["current_completed_technical_scope"])
    groups = cast(list[dict[str, Any]], scope["cabinet_groups"])
    positions = cast(list[dict[str, Any]], scope["pricing_positions"])
    fingerprints = cast(list[dict[str, Any]], scope["composition_fingerprints"])
    groups[2]["row_draft_ids"] = [
        *groups[2]["row_draft_ids"],
        *(item[2] for item in POSITION_SCOPE),
    ]
    by_position = {item["pricing_position_id"]: item for item in positions}
    for position_id, _source_id, row_id, row_index in POSITION_SCOPE:
        position = by_position[position_id]
        position["row_draft_ids"] = [*position["row_draft_ids"], row_id]
        position["row_draft_json_paths"] = [
            *position["row_draft_json_paths"],
            f"$.calculator_input_format.row_drafts[{row_index}]",
        ]
        position["composition_fingerprint_sha256"] = NEW_FINGERPRINT
    old_record = next(
        item for item in fingerprints if item["fingerprint_sha256"] == OLD_FINGERPRINT
    )
    merged_record = next(
        item for item in fingerprints if item["fingerprint_sha256"] == NEW_FINGERPRINT
    )
    merged_record["source_position_ids"] = [
        *old_record["source_position_ids"],
        *merged_record["source_position_ids"],
    ]
    merged_record["pricing_position_ids"] = [
        *old_record["pricing_position_ids"],
        *merged_record["pricing_position_ids"],
    ]
    scope["composition_fingerprints"] = [
        item for item in fingerprints if item["fingerprint_sha256"] != OLD_FINGERPRINT
    ]
    scope["coverage"] = copy.deepcopy(EXPECTED_COVERAGE)
    scope["shu_t2_rt820_preliminary_candidate"] = {
        "status": "NOT_CALCULATED_NOT_APPROVED",
        "approved_unit_price_kzt": None,
        "application_status": APPLICATION_STATUS,
        "unit_candidate_kzt": 53763,
        "four_position_candidate_kzt": 215052,
        "preliminary_project_candidate_kzt": 11963792,
    }
    payload["authoritative_inputs"] = [
        *payload["authoritative_inputs"],
        {
            "role": "parent_pricing_profile_successor",
            **_input_binding(loaded.parent, PROFILE_SCHEMA),
        },
        {
            "role": "completed_technical_input_shu_t2_rt820_successor",
            **_input_binding(loaded.technical, TECHNICAL_SCHEMA),
            "purpose": "exact 15-group/116-row SHU-T2 RT-820 technical authority",
        },
        {
            "role": "shu_t2_rt820_scope_human_decision",
            **_input_binding(loaded.decision, DECISION_SCHEMA),
            "decision_id": DECISION_ID,
            "status": DECISION_STATUS,
            "authority": AUTHORITY,
            "application_status": APPLICATION_STATUS,
        },
        {
            "role": "main_price_workbook_shu_t2_rt820_revalidated",
            **_input_binding(
                loaded.workbook,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            ),
            "source_range": "КРН!A19:C19",
        },
    ]
    payload["shu_t2_rt820_pricing_profile_successor"] = successor_metadata(loaded)
    validation = cast(dict[str, Any], payload["validation_summary"])
    validation.update(
        {
            "current_coverage": "55/137/11",
            "shu_t2_rt820_controlled_replacement_validation": "PASS",
            "publication_status": "PUBLISHED_IMMUTABLE_NO_OVERWRITE",
        }
    )
    return payload


def validate_successor_payload(
    payload: Mapping[str, Any], loaded: LoadedInputs
) -> None:
    parent = cast(Mapping[str, Any], loaded.parent.value)
    validate_parent_profile(parent)
    expected_payload = build_successor_payload(loaded)
    require(
        set(payload) == {*parent, "shu_t2_rt820_pricing_profile_successor"},
        "successor root key envelope mismatch",
    )
    require(payload.get("project_id") == PROJECT_ID, "successor project mismatch")
    require(
        payload.get("shu_t2_rt820_pricing_profile_successor")
        == successor_metadata(loaded),
        "successor metadata mismatch",
    )
    scope = _scope(payload)
    parent_scope = _scope(parent)
    require(
        set(scope) == {*parent_scope, "shu_t2_rt820_preliminary_candidate"},
        "successor scope key envelope mismatch",
    )
    require(scope.get("coverage") == EXPECTED_COVERAGE, "successor coverage mismatch")
    groups = cast(list[Any], scope.get("cabinet_groups"))
    parent_groups = cast(list[Any], parent_scope.get("cabinet_groups"))
    positions = cast(list[Any], scope.get("pricing_positions"))
    parent_positions = cast(list[Any], parent_scope.get("pricing_positions"))
    fingerprints = cast(list[Any], scope.get("composition_fingerprints"))
    require(
        len(groups) == 15 and len(positions) == 55 and len(fingerprints) == 11,
        "successor inventory mismatch",
    )
    require(
        groups[:2] == parent_groups[:2] and groups[3:] == parent_groups[3:],
        "non-target cabinet group changed",
    )
    require(
        groups[2].get("row_draft_ids")
        == [*parent_groups[2]["row_draft_ids"], *(item[2] for item in POSITION_SCOPE)],
        "target cabinet group rows mismatch",
    )
    target_ids = {item[0] for item in POSITION_SCOPE}
    for current, previous in zip(positions, parent_positions, strict=True):
        if current.get("pricing_position_id") not in target_ids:
            require(current == previous, "non-target pricing position changed")
            continue
        position_id = current["pricing_position_id"]
        expected = next(item for item in POSITION_SCOPE if item[0] == position_id)
        for key in current:
            if key not in {
                "row_draft_ids",
                "row_draft_json_paths",
                "composition_fingerprint_sha256",
            }:
                require(
                    current[key] == previous[key],
                    f"target position identity/state changed: {position_id}",
                )
        require(
            current["row_draft_ids"] == [*previous["row_draft_ids"], expected[2]],
            f"target row binding mismatch: {position_id}",
        )
        require(
            current["row_draft_json_paths"]
            == [
                *previous["row_draft_json_paths"],
                f"$.calculator_input_format.row_drafts[{expected[3]}]",
            ],
            f"target row path mismatch: {position_id}",
        )
        require(
            current["composition_fingerprint_sha256"] == NEW_FINGERPRINT,
            f"target fingerprint mismatch: {position_id}",
        )
        require(
            current["approved_unit_price_kzt"] is None
            and current["approved_unit_price_decision_status"]
            == "NOT_CALCULATED_NOT_APPROVED",
            f"target price state changed: {position_id}",
        )
    fingerprint_ids = [
        item.get("fingerprint_sha256")
        for item in fingerprints
        if isinstance(item, Mapping)
    ]
    require(OLD_FINGERPRINT not in fingerprint_ids, "old fingerprint was not removed")
    require(
        fingerprint_ids.count(NEW_FINGERPRINT) == 1, "merged fingerprint must be unique"
    )
    merged = next(
        item for item in fingerprints if item["fingerprint_sha256"] == NEW_FINGERPRINT
    )
    require(
        merged["components"] == NEW_COMPONENTS
        and fingerprint(merged["components"]) == NEW_FINGERPRINT,
        "merged fingerprint components mismatch",
    )
    require(
        merged["source_position_ids"]
        == [
            "TFE-016",
            "TFE-041",
            "TFE-061",
            "TFE-083",
            "TFE-006",
            "TFE-029",
            "TFE-052",
            "TFE-074",
        ]
        and merged["pricing_position_ids"]
        == [*target_ids_order(), *SHU_T1_POSITION_IDS],
        "merged eight-position ordering mismatch",
    )
    require(
        set(item["pricing_position_id"] for item in positions)
        == {f"PRICE-POSITION-{index:03d}" for index in range(1, 56)},
        "pricing position identity coverage mismatch",
    )
    require(
        sum(item["physical_multiplicity"] for item in positions) == 137,
        "physical cabinet coverage mismatch",
    )
    require(
        all(
            current.get("approved_unit_price_kzt")
            == previous.get("approved_unit_price_kzt")
            and current.get("approved_unit_price_decision_status")
            == previous.get("approved_unit_price_decision_status")
            for current, previous in zip(positions, parent_positions, strict=True)
        ),
        "successor position price states changed from parent",
    )
    require(
        payload.get("status") == PROFILE_STATUS
        and payload.get("authority") == EXPECTED_AUTHORITY
        and payload.get("application_status") == APPLICATION_STATUS
        and payload.get("immutable_state") == {"immutable": True, "no_overwrite": True}
        and payload.get("scope_expansion") is False,
        "successor immutable/application state mismatch",
    )
    require(
        payload.get("safety_flags") == EXPECTED_SAFETY_FLAGS,
        "successor safety flags mismatch",
    )
    require(
        payload.get("non_approvals") == EXPECTED_NON_APPROVALS,
        "successor non-approvals mismatch",
    )
    metadata = cast(
        Mapping[str, Any], payload["shu_t2_rt820_pricing_profile_successor"]
    )
    invariants = metadata.get("preliminary_not_approved_invariants")
    require(
        isinstance(invariants, Mapping)
        and invariants.get("unit_candidate_kzt") == 53763
        and invariants.get("four_position_candidate_kzt") == 215052
        and invariants.get("preliminary_project_candidate_kzt") == 11963792
        and invariants.get("approved") is False
        and invariants.get("applied") is False,
        "preliminary non-approval invariants mismatch",
    )
    forbidden = json.dumps(payload, ensure_ascii=False)
    require('"work_kzt": 432' not in forbidden, "generic RT-820 work 432 is forbidden")
    require(
        '"separate_tst05' not in forbidden.casefold(),
        "separate TST05 charge/row is forbidden",
    )
    require(
        payload == expected_payload,
        "successor closed-envelope mismatch against deterministic expected payload",
    )


def target_ids_order() -> list[str]:
    return [item[0] for item in POSITION_SCOPE]


def validate_real_inputs_read_only(
    paths: InputPaths, shas: ExpectedShas
) -> dict[str, Any]:
    loaded = load_and_validate_inputs(paths, shas)
    payload = build_successor_payload(loaded)
    validate_successor_payload(payload, loaded)
    return {
        "status": "PASS",
        "coverage": EXPECTED_COVERAGE,
        "fingerprint": NEW_FINGERPRINT,
        "publication_called": False,
        "price_approved": False,
    }


def _recheck_inputs(loaded: LoadedInputs, phase: str) -> None:
    for item in (loaded.parent, loaded.technical, loaded.decision, loaded.workbook):
        try:
            current = item.path.read_bytes()
        except OSError as exc:
            raise ContractError(
                f"{phase} TOCTOU reread failed: {item.role}: {exc}"
            ) from exc
        require(current == item.raw, f"{phase} TOCTOU bytes changed: {item.role}")
        require(
            sha256_bytes(current) == item.expected_sha256,
            f"{phase} TOCTOU SHA mismatch: {item.role}",
        )


def _path_identity(path: Path) -> tuple[int, int]:
    stat = os.lstat(path)
    return stat.st_dev, stat.st_ino


def _rollback_publication(
    output: Path,
    staging: Path | None,
    final_link_created: bool,
    staged_identity: tuple[int, int] | None,
) -> list[str]:
    blockers: list[str] = []
    if final_link_created and os.path.lexists(output):
        try:
            if staged_identity is None or _path_identity(output) != staged_identity:
                blockers.append("foreign final replacement preserved")
            else:
                output.unlink()
        except OSError as exc:
            blockers.append(f"owned final cleanup failed: {exc}")
    if staging is not None and os.path.lexists(staging):
        try:
            staging.unlink()
        except OSError as exc:
            blockers.append(f"staging cleanup failed: {exc}")
    if output.parent.exists():
        try:
            output.parent.rmdir()
        except OSError:
            if not os.path.lexists(output):
                blockers.append("output directory cleanup failed")
    return blockers


def publish_successor(
    paths: InputPaths, shas: ExpectedShas, output: Path
) -> PublicationResult:
    require(output.name == OUTPUT_FILENAME, "output filename mismatch")
    require(output.parent != output, "output directory mismatch")
    require(output.parent.parent.is_dir(), "output directory owner must already exist")
    require(not output.parent.exists(), "output directory already exists")
    try:
        resolved(output).relative_to(REPO_ROOT)
    except ValueError:
        pass
    else:
        raise ContractError("successor output must be outside the repository")
    require(
        resolved(output) not in {resolved(path) for path in vars(paths).values()},
        "output must not alias an input",
    )
    loaded = load_and_validate_inputs(paths, shas)
    payload = build_successor_payload(loaded)
    validate_successor_payload(payload, loaded)
    encoded = serialize(payload)
    output.parent.mkdir()
    descriptor = -1
    staging: Path | None = None
    staged_identity: tuple[int, int] | None = None
    final_link_created = False
    try:
        descriptor, staging_name = tempfile.mkstemp(
            prefix=f".{output.name}.", suffix=".staging", dir=output.parent
        )
        staging = Path(staging_name)
        with os.fdopen(descriptor, "wb") as stream:
            descriptor = -1
            stream.write(encoded)
            stream.flush()
            os.fsync(stream.fileno())
        staged, staged_raw = load_json(staging, "staged pricing successor")
        require(staged_raw == encoded, "staged bytes mismatch")
        validate_successor_payload(staged, loaded)
        require(
            set(output.parent.iterdir()) == {staging},
            "output directory contains unexpected pre-publication entries",
        )
        _recheck_inputs(loaded, "pre-publication")
        require(not output.exists(), "output appeared before publication")
        staged_identity = _path_identity(staging)
        try:
            os.link(staging, output)
        except OSError as exc:
            raise ContractError(
                f"atomic no-overwrite publication failed: {exc}"
            ) from exc
        final_link_created = True
        require(
            _path_identity(output) == staged_identity, "published identity mismatch"
        )
        published, published_raw = load_json(output, "published pricing successor")
        require(published_raw == encoded, "published bytes mismatch")
        validate_successor_payload(published, loaded)
        _recheck_inputs(loaded, "final")
        staging.unlink()
        require(_path_identity(output) == staged_identity, "final identity changed")
        require(
            set(output.parent.iterdir()) == {output},
            "output directory final inventory mismatch",
        )
        return PublicationResult(
            sha256_bytes(published_raw), len(published_raw), encoded
        )
    except BaseException as error:
        blockers: list[str] = []
        if descriptor >= 0:
            try:
                os.close(descriptor)
            except OSError as exc:
                blockers.append(f"staging descriptor cleanup failed: {exc}")
        blockers.extend(
            _rollback_publication(output, staging, final_link_created, staged_identity)
        )
        if blockers:
            raise ContractError(
                "publication rollback cleanup blocked: " + "; ".join(blockers)
            ) from error
        raise


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--parent-pricing-profile", required=True, type=Path)
    parser.add_argument("--parent-pricing-profile-sha256", required=True)
    parser.add_argument("--technical-successor", required=True, type=Path)
    parser.add_argument("--technical-successor-sha256", required=True)
    parser.add_argument("--human-decision", required=True, type=Path)
    parser.add_argument("--human-decision-sha256", required=True)
    parser.add_argument("--pricing-workbook", required=True, type=Path)
    parser.add_argument("--pricing-workbook-sha256", required=True)
    parser.add_argument("--output", required=True, type=Path)
    parser.add_argument("--authorization", required=True)
    return parser


def parse_args(argv: Sequence[str] | None = None) -> argparse.Namespace:
    return build_parser().parse_args(argv)


def main(argv: Sequence[str] | None = None) -> int:
    args = parse_args(argv)
    require(
        args.authorization == PUBLICATION_AUTHORIZATION,
        "exact SHU-T2 RT-820 pricing-profile publication authorization is required",
    )
    paths = InputPaths(
        args.parent_pricing_profile,
        args.technical_successor,
        args.human_decision,
        args.pricing_workbook,
    )
    shas = ExpectedShas(
        args.parent_pricing_profile_sha256,
        args.technical_successor_sha256,
        args.human_decision_sha256,
        args.pricing_workbook_sha256,
    )
    result = publish_successor(paths, shas, args.output)
    print(
        f"PUBLISHED_IMMUTABLE_NO_OVERWRITE {args.output} "
        f"SHA256={result.sha256} SIZE={result.size}"
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
