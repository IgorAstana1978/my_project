"""Resolve the approved custom ЩЭ cabinet base cost without Excel recalculation."""

from __future__ import annotations

import argparse
import hashlib
import json
import re
from decimal import ROUND_UP, Decimal, InvalidOperation
from pathlib import Path
from typing import Any

from openpyxl import load_workbook  # type: ignore[import-untyped]

SCHEMA_VERSION = "custom_sche_cabinet_base_cost_resolution.v0.1"
APPROVED_WORKBOOK_SHA256 = (
    "b51d7087e0bd8f92e48985294062ead6826c6b50ce3cfacd0f9d0dc22c05f7f2"
)
APPROVED_INTERNAL_CABINET_CODE = "CAB-SCHE-BI-900X900X120-M12"
APPROVED_METAL_THICKNESS = Decimal("1.2")
APPROVED_SHEET = "Лист1"
APPROVED_ROW = 82
APPROVED_SOURCE_LABEL = "ЩЭ 5кв 900х900х120"
APPROVED_DIMENSIONS_MM = (Decimal("900"), Decimal("900"), Decimal("120"))
APPROVED_SHARED_SOURCE_TEMPLATES = ("ЩЭ-3кв", "ЩЭ-4кв", "ЩЭ-5кв", "ЩЭ-6кв")
SHA256_RE = re.compile(r"[0-9a-f]{64}\Z")

EXPECTED_FORMULAS = {
    "B82": "=ROUNDUP(D82*$B$5,0)",
    "C82": "=ROUNDUP(D82*$C$5,0)",
    "D82": "=ROUNDUP((M82+N82)*R82+O82+P82+S82+Q82,0)",
    "H82": "=(E82/1000)",
    "I82": "=(F82/1000)",
    "J82": "=(G82/1000)",
    "K82": "=(H82+0.066)*(I82+0.066)+0.1*((H82-0.01)*4+I82*2+I82-0.09+J82*8)",
    "L82": "=K82*$B$1*8.42",
    "M82": "=L82*$B$2",
    "N82": "=K82*2*0.25*$B$3*1.33",
}
EXPECTED_MANUAL_INPUTS = {
    "O82": Decimal("3750"),
    "P82": Decimal("900"),
    "R82": Decimal("1.05"),
    "S82": Decimal("1700"),
}
EXPECTED_Q82_FORMULA = "=1250"
EXPECTED_Q82_VALUE = Decimal("1250")


class ResolutionError(ValueError):
    """Raised when an approved resolver invariant is not satisfied."""


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as source:
        for block in iter(lambda: source.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def _decimal_from_numeric(value: object, *, cell: str) -> Decimal:
    if isinstance(value, bool) or not isinstance(value, (int, float, Decimal)):
        raise ResolutionError(f"{cell} must contain a numeric value")
    try:
        number = Decimal(str(value))
    except InvalidOperation as exc:
        raise ResolutionError(f"{cell} must contain a valid numeric value") from exc
    if not number.is_finite():
        raise ResolutionError(f"{cell} must contain a finite numeric value")
    return number


def _decimal_from_input(value: str, *, field: str) -> Decimal:
    try:
        number = Decimal(value)
    except InvalidOperation as exc:
        raise ResolutionError(f"{field} must be numeric") from exc
    if not number.is_finite() or number <= 0:
        raise ResolutionError(f"{field} must be a positive finite number")
    return number


def _decimal_text(value: Decimal) -> str:
    return format(value, "f")


def _failure_result(
    *,
    workbook_path: Path,
    expected_workbook_sha256: str,
    internal_cabinet_code: str,
    metal_thickness: str,
    expected_sheet: str,
    expected_row: int,
    error: str,
) -> dict[str, Any]:
    return {
        "schema_version": SCHEMA_VERSION,
        "status": "CUSTOM_SCHE_BASE_COST_RESOLUTION_FAILED",
        "errors": [error],
        "request": {
            "metal_workbook_path": str(workbook_path),
            "expected_workbook_sha256": expected_workbook_sha256,
            "internal_cabinet_code": internal_cabinet_code,
            "metal_thickness_mm": metal_thickness,
            "expected_sheet": expected_sheet,
            "expected_row": expected_row,
        },
    }


def _validate_request(
    *,
    expected_workbook_sha256: str,
    internal_cabinet_code: str,
    metal_thickness: str,
    expected_sheet: str,
    expected_row: int,
) -> Decimal:
    if not SHA256_RE.fullmatch(expected_workbook_sha256):
        raise ResolutionError("expected_workbook_sha256 must be lowercase SHA-256")
    if expected_workbook_sha256 != APPROVED_WORKBOOK_SHA256:
        raise ResolutionError(
            "expected_workbook_sha256 is not the approved workbook hash"
        )
    if internal_cabinet_code != APPROVED_INTERNAL_CABINET_CODE:
        raise ResolutionError("internal_cabinet_code is not approved")
    thickness = _decimal_from_input(metal_thickness, field="metal_thickness")
    if thickness != APPROVED_METAL_THICKNESS:
        raise ResolutionError("metal_thickness must equal the approved 1.2 mm")
    if expected_sheet != APPROVED_SHEET:
        raise ResolutionError("expected_sheet is not the approved source sheet")
    if expected_row != APPROVED_ROW:
        raise ResolutionError("expected_row is not the approved source row")
    return thickness


def _assert_formula(sheet: Any, cell: str, expected: str) -> None:
    actual = sheet[cell].value
    if actual != expected:
        raise ResolutionError(
            f"{cell} formula drift: expected {expected!r}, got {actual!r}"
        )


def _assert_decimal(sheet: Any, cell: str, expected: Decimal) -> Decimal:
    actual = _decimal_from_numeric(sheet[cell].value, cell=cell)
    if actual != expected:
        raise ResolutionError(
            f"{cell} value drift: expected {_decimal_text(expected)}, "
            f"got {_decimal_text(actual)}"
        )
    return actual


def _calculate_roles(
    *,
    width_mm: Decimal,
    height_mm: Decimal,
    depth_mm: Decimal,
    thickness: Decimal,
    paint_rate: Decimal,
    labor_rate: Decimal,
) -> dict[str, Decimal]:
    width_m = width_mm / Decimal("1000")
    height_m = height_mm / Decimal("1000")
    depth_m = depth_mm / Decimal("1000")
    area = (width_m + Decimal("0.066")) * (height_m + Decimal("0.066")) + (
        Decimal("0.1")
        * (
            (width_m - Decimal("0.01")) * Decimal("4")
            + height_m * Decimal("2")
            + height_m
            - Decimal("0.09")
            + depth_m * Decimal("8")
        )
    )
    metal_mass = area * thickness * Decimal("8.42")
    metal_cost = metal_mass * paint_rate
    labor_cost = area * Decimal("2") * Decimal("0.25") * labor_rate * Decimal("1.33")
    base_cost = (
        (metal_cost + labor_cost) * EXPECTED_MANUAL_INPUTS["R82"]
        + EXPECTED_MANUAL_INPUTS["O82"]
        + EXPECTED_MANUAL_INPUTS["P82"]
        + EXPECTED_MANUAL_INPUTS["S82"]
        + EXPECTED_Q82_VALUE
    ).quantize(Decimal("1"), rounding=ROUND_UP)
    return {
        "H82_width_m": width_m,
        "I82_height_m": height_m,
        "J82_depth_m": depth_m,
        "K82_sheet_area_m2": area,
        "L82_metal_mass_kg": metal_mass,
        "M82_metal_cost": metal_cost,
        "N82_labor_cost": labor_cost,
        "D82_base_cost": base_cost,
    }


def _resolve_checked_workbook(
    *,
    workbook_path: Path,
    workbook_sha256: str,
    thickness: Decimal,
) -> dict[str, Any]:
    workbook = load_workbook(
        workbook_path,
        read_only=True,
        data_only=False,
        keep_links=False,
    )
    try:
        if APPROVED_SHEET not in workbook.sheetnames:
            raise ResolutionError(f"required sheet {APPROVED_SHEET!r} is absent")
        sheet = workbook[APPROVED_SHEET]
        if sheet[f"A{APPROVED_ROW}"].value != APPROVED_SOURCE_LABEL:
            raise ResolutionError("A82 source label drift")

        dimensions = tuple(
            _decimal_from_numeric(sheet[f"{column}{APPROVED_ROW}"].value, cell=column)
            for column in ("E", "F", "G")
        )
        if dimensions != APPROVED_DIMENSIONS_MM:
            raise ResolutionError("E82:G82 dimensions drift")

        for cell, expected in EXPECTED_FORMULAS.items():
            _assert_formula(sheet, cell, expected)
        for cell, expected in EXPECTED_MANUAL_INPUTS.items():
            _assert_decimal(sheet, cell, expected)
        _assert_formula(sheet, "Q82", EXPECTED_Q82_FORMULA)

        paint_rate = _decimal_from_numeric(sheet["B2"].value, cell="B2")
        labor_rate = _decimal_from_numeric(sheet["B3"].value, cell="B3")
        if paint_rate <= 0 or labor_rate <= 0:
            raise ResolutionError("B2 and B3 must contain positive numeric values")

        roles = _calculate_roles(
            width_mm=dimensions[0],
            height_mm=dimensions[1],
            depth_mm=dimensions[2],
            thickness=thickness,
            paint_rate=paint_rate,
            labor_rate=labor_rate,
        )
        return {
            "schema_version": SCHEMA_VERSION,
            "status": "CUSTOM_SCHE_BASE_COST_RESOLUTION_VALIDATED",
            "errors": [],
            "cabinet_identity": {
                "internal_cabinet_code": APPROVED_INTERNAL_CABINET_CODE,
                "source_cabinet_label": APPROVED_SOURCE_LABEL,
                "installation": "built_in",
                "dimensions_mm": {
                    "width": _decimal_text(dimensions[0]),
                    "height": _decimal_text(dimensions[1]),
                    "depth": _decimal_text(dimensions[2]),
                },
                "metal_thickness_mm": _decimal_text(thickness),
                "shared_source_templates": list(APPROVED_SHARED_SOURCE_TEMPLATES),
            },
            "source_provenance": {
                "metal_workbook_path": str(workbook_path),
                "metal_workbook_sha256": workbook_sha256,
                "sheet": APPROVED_SHEET,
                "row": APPROVED_ROW,
                "source_cells": {
                    "label": "A82",
                    "dimensions": ["E82", "F82", "G82"],
                    "global_rates": ["B2", "B3"],
                    "formulas": list(EXPECTED_FORMULAS),
                    "manual_inputs": ["O82", "P82", "Q82", "R82", "S82"],
                },
                "excel_recalculation_executed": False,
            },
            "verified_formulas": dict(EXPECTED_FORMULAS),
            "verified_manual_inputs": {
                **{
                    cell: _decimal_text(value)
                    for cell, value in EXPECTED_MANUAL_INPUTS.items()
                },
                "Q82": {
                    "formula": EXPECTED_Q82_FORMULA,
                    "resolved_constant": _decimal_text(EXPECTED_Q82_VALUE),
                },
            },
            "checked_global_rates": {
                "B2": _decimal_text(paint_rate),
                "B3": _decimal_text(labor_rate),
            },
            "computed_formula_roles": {
                cell_role: _decimal_text(value) for cell_role, value in roles.items()
            },
            "base_cost": {
                "source_role": "D82",
                "value": _decimal_text(roles["D82_base_cost"]),
                "rounding": "ROUNDUP_TO_INTEGER",
                "excluded_output_roles": ["B82", "C82"],
            },
        }
    finally:
        workbook.close()


def _write_json_no_overwrite(output_path: Path, result: dict[str, Any]) -> None:
    if output_path.suffix.lower() != ".json":
        raise ResolutionError("output path must have .json suffix")
    if not output_path.parent.is_dir():
        raise ResolutionError("output parent directory does not exist")
    serialized = json.dumps(result, ensure_ascii=False, indent=2) + "\n"
    with output_path.open("x", encoding="utf-8", newline="\n") as output:
        output.write(serialized)


def resolve_custom_sche_cabinet_base_cost(
    *,
    metal_workbook_path: Path,
    expected_workbook_sha256: str,
    internal_cabinet_code: str,
    metal_thickness: str,
    expected_sheet: str,
    expected_row: int,
    output_json_path: Path | None = None,
) -> dict[str, Any]:
    """Validate source evidence and deterministically resolve the approved base cost."""

    try:
        thickness = _validate_request(
            expected_workbook_sha256=expected_workbook_sha256,
            internal_cabinet_code=internal_cabinet_code,
            metal_thickness=metal_thickness,
            expected_sheet=expected_sheet,
            expected_row=expected_row,
        )
        if output_json_path is not None and output_json_path.exists():
            raise ResolutionError("output path already exists; overwrite is forbidden")
        if not metal_workbook_path.is_file():
            raise ResolutionError("metal workbook does not exist")
        workbook_sha256_before = _sha256(metal_workbook_path)
        if workbook_sha256_before != expected_workbook_sha256:
            raise ResolutionError("metal workbook SHA-256 mismatch")

        result = _resolve_checked_workbook(
            workbook_path=metal_workbook_path,
            workbook_sha256=workbook_sha256_before,
            thickness=thickness,
        )
        workbook_sha256_after = _sha256(metal_workbook_path)
        if workbook_sha256_after != workbook_sha256_before:
            raise ResolutionError("metal workbook changed during resolution")
        if output_json_path is not None:
            result["output_json_path"] = str(output_json_path)
            _write_json_no_overwrite(output_json_path, result)
        return result
    except (OSError, ResolutionError) as exc:
        return _failure_result(
            workbook_path=metal_workbook_path,
            expected_workbook_sha256=expected_workbook_sha256,
            internal_cabinet_code=internal_cabinet_code,
            metal_thickness=metal_thickness,
            expected_sheet=expected_sheet,
            expected_row=expected_row,
            error=str(exc),
        )


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--metal-workbook", required=True, type=Path)
    parser.add_argument("--expected-workbook-sha256", required=True)
    parser.add_argument("--internal-cabinet-code", required=True)
    parser.add_argument("--metal-thickness", required=True)
    parser.add_argument("--expected-sheet", required=True)
    parser.add_argument("--expected-row", required=True, type=int)
    parser.add_argument("--output-json", type=Path)
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    result = resolve_custom_sche_cabinet_base_cost(
        metal_workbook_path=args.metal_workbook,
        expected_workbook_sha256=args.expected_workbook_sha256,
        internal_cabinet_code=args.internal_cabinet_code,
        metal_thickness=args.metal_thickness,
        expected_sheet=args.expected_sheet,
        expected_row=args.expected_row,
        output_json_path=args.output_json,
    )
    print(json.dumps(result, ensure_ascii=False, indent=2))
    return 0 if result["status"] == "CUSTOM_SCHE_BASE_COST_RESOLUTION_VALIDATED" else 1


if __name__ == "__main__":
    raise SystemExit(main())
