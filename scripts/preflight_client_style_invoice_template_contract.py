"""Read-only preflight for a client-style invoice template contract."""

from __future__ import annotations

import argparse
import hashlib
import json
import posixpath
import re
import zipfile
from collections.abc import Mapping, Sequence
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, cast
from xml.etree import ElementTree

from openpyxl import load_workbook
from openpyxl.utils.cell import column_index_from_string, coordinate_to_tuple

PROJECT_ROOT = Path(__file__).resolve().parents[1]
EXAMPLE_CONTRACT_RELATIVE_PATH = (
    Path("examples") / "client_style_invoice_template_contract.example.json"
)
REPORT_START = "CLIENT_STYLE_TEMPLATE_CONTRACT_PREFLIGHT_REPORT_START"
REPORT_END = "CLIENT_STYLE_TEMPLATE_CONTRACT_PREFLIGHT_REPORT_END"
MODE = "read-only client-style invoice template contract preflight"
COMMERCIAL_STATUS = "template preflight only; PASS is not client export approval"
HUMAN_APPROVAL = "required before generating or sending client-style invoice"
CONTRACT_PATH_POLICY = (
    "inside-Git example is documentation/test-only; production contract must "
    "be outside Git"
)

TOP_LEVEL_FIELDS = (
    "contract_id",
    "template_name",
    "template_version",
    "expected_sheet_name",
    "template_sha256",
    "allowed_extra_sheets",
    "print",
    "layout",
    "required_fixed_labels",
)
PRINT_FIELDS = (
    "orientation",
    "paper_size",
    "print_area_required",
)
LAYOUT_FIELDS = (
    "invoice_number_cell",
    "invoice_date_cell",
    "payer_cell",
    "object_cell",
    "table_header_row",
    "first_item_row",
    "item_columns",
    "amount_words_cell",
    "signer_name_cell",
    "signer_title_cell",
)
ITEM_COLUMN_FIELDS = (
    "index",
    "name",
    "unit",
    "quantity",
    "instruments_and_devices",
    "cabinet",
    "unit_price",
    "line_total",
)
CELL_LAYOUT_FIELDS = (
    "invoice_number_cell",
    "invoice_date_cell",
    "payer_cell",
    "amount_words_cell",
    "signer_name_cell",
    "signer_title_cell",
)
HASH_RE = re.compile(r"[0-9a-f]{64}\Z")
CELL_RE = re.compile(r"\$?[A-Za-z]{1,3}\$?[1-9][0-9]*\Z")
COLUMN_RE = re.compile(r"[A-Za-z]{1,3}\Z")
MAX_EXCEL_ROW = 1_048_576
MAX_EXCEL_COLUMN = 16_384
SPREADSHEET_NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
DOCUMENT_REL_NS = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
PACKAGE_REL_NS = "http://schemas.openxmlformats.org/package/2006/relationships"


@dataclass(frozen=True)
class PrintContract:
    orientation: str | None
    paper_size: str | None
    print_area_required: bool


@dataclass(frozen=True)
class LayoutContract:
    cells: Mapping[str, str | None]
    table_header_row: int
    first_item_row: int
    item_columns: Mapping[str, str]


@dataclass(frozen=True)
class FixedLabelContract:
    cell: str
    expected_text_contains: str


@dataclass(frozen=True)
class ActualPrintSetup:
    orientation: str | None
    paper_size: str | None
    has_print_area: bool


@dataclass(frozen=True)
class TemplateContract:
    contract_id: str
    template_name: str
    template_version: str
    expected_sheet_name: str
    template_sha256: str
    allowed_extra_sheets: tuple[str, ...]
    print_contract: PrintContract
    layout: LayoutContract
    required_fixed_labels: tuple[FixedLabelContract, ...]


@dataclass
class TemplateContractPreflightResult:
    template_xlsx: Path
    contract_json: Path
    status: str = "FAIL"
    checks: dict[str, str] = field(
        default_factory=lambda: {
            "input paths": "fail",
            "contract schema": "fail",
            "template hash": "fail",
            "workbook layout": "fail",
            "fixed labels": "fail",
            "print setup": "fail",
            "safety boundaries": "fail",
        }
    )
    red_flags: list[str] = field(default_factory=list)


def parse_args(argv: Sequence[str] | None = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Read-only preflight for a client-style XLSX template contract."
    )
    parser.add_argument("--template-xlsx", required=True, type=Path)
    parser.add_argument("--contract-json", required=True, type=Path)
    return parser.parse_args(argv)


def resolved(path: Path) -> Path:
    return path.expanduser().resolve(strict=False)


def is_inside_project(path: Path) -> bool:
    return resolved(path).is_relative_to(PROJECT_ROOT)


def documented_example_contract_path() -> Path:
    return resolved(PROJECT_ROOT / EXAMPLE_CONTRACT_RELATIVE_PATH)


def add_red_flag(result: TemplateContractPreflightResult, message: str) -> None:
    if message not in result.red_flags:
        result.red_flags.append(message)


def validate_input_paths(result: TemplateContractPreflightResult) -> bool:
    valid = True
    template_path = result.template_xlsx
    contract_path = result.contract_json

    if not template_path.is_file():
        valid = False
        add_red_flag(result, "template XLSX does not exist")
    if template_path.suffix.casefold() != ".xlsx":
        valid = False
        add_red_flag(result, "template suffix must be .xlsx")
    if is_inside_project(template_path):
        valid = False
        add_red_flag(result, "template XLSX must be outside the Git project")

    if not contract_path.is_file():
        valid = False
        add_red_flag(result, "contract JSON does not exist")
    if contract_path.suffix.casefold() != ".json":
        valid = False
        add_red_flag(result, "contract suffix must be .json")
    if (
        is_inside_project(contract_path)
        and contract_path != documented_example_contract_path()
    ):
        valid = False
        add_red_flag(
            result,
            "production contract JSON must be outside the Git project",
        )

    result.checks["input paths"] = "pass" if valid else "fail"
    return valid


def load_contract_json(
    path: Path,
    result: TemplateContractPreflightResult,
) -> Mapping[str, Any] | None:
    if not path.is_file():
        return None
    try:
        raw = json.loads(path.read_text(encoding="utf-8"))
    except UnicodeDecodeError:
        add_red_flag(result, "contract JSON must be valid UTF-8")
        return None
    except json.JSONDecodeError:
        add_red_flag(result, "contract JSON is invalid")
        return None
    except OSError:
        add_red_flag(result, "contract JSON could not be read")
        return None

    if not isinstance(raw, Mapping):
        add_red_flag(result, "contract JSON root must be an object")
        return None
    return cast(Mapping[str, Any], raw)


def has_required_fields(
    data: Mapping[str, Any],
    fields: Sequence[str],
    prefix: str,
    result: TemplateContractPreflightResult,
) -> bool:
    valid = True
    for field_name in fields:
        if field_name not in data:
            valid = False
            add_red_flag(result, f"contract field is missing: {prefix}{field_name}")
    return valid


def nonempty_string(
    data: Mapping[str, Any],
    field_name: str,
    prefix: str,
    result: TemplateContractPreflightResult,
) -> str | None:
    value = data.get(field_name)
    if not isinstance(value, str) or value.strip() == "":
        add_red_flag(
            result,
            f"contract field must be a non-empty string: {prefix}{field_name}",
        )
        return None
    return value


def optional_string(
    value: Any,
    field_name: str,
    result: TemplateContractPreflightResult,
) -> str | None:
    if value is None:
        return None
    if not isinstance(value, str) or value.strip() == "":
        add_red_flag(
            result,
            f"contract field must be a non-empty string or null: {field_name}",
        )
        return None
    return value


def positive_integer(
    data: Mapping[str, Any],
    field_name: str,
    prefix: str,
    result: TemplateContractPreflightResult,
) -> int | None:
    value = data.get(field_name)
    if not isinstance(value, int) or isinstance(value, bool) or value <= 0:
        add_red_flag(
            result,
            f"contract field must be a positive integer: {prefix}{field_name}",
        )
        return None
    return value


def valid_cell_coordinate(value: str) -> bool:
    if CELL_RE.fullmatch(value) is None:
        return False
    normalized = value.replace("$", "").upper()
    try:
        row, column = coordinate_to_tuple(normalized)
    except TypeError, ValueError:
        return False
    return 1 <= row <= MAX_EXCEL_ROW and 1 <= column <= MAX_EXCEL_COLUMN


def valid_column(value: str) -> bool:
    if COLUMN_RE.fullmatch(value) is None:
        return False
    try:
        column = column_index_from_string(value.upper())
    except ValueError:
        return False
    return 1 <= column <= MAX_EXCEL_COLUMN


def normalize_cell(value: str) -> str:
    return value.replace("$", "").upper()


def normalize_column(value: str) -> str:
    return value.upper()


def validate_print_contract(
    data: Any,
    result: TemplateContractPreflightResult,
) -> PrintContract | None:
    if not isinstance(data, Mapping):
        add_red_flag(result, "contract field must be an object: print")
        return None
    print_data = cast(Mapping[str, Any], data)
    if not has_required_fields(print_data, PRINT_FIELDS, "print.", result):
        return None

    orientation = optional_string(
        print_data.get("orientation"),
        "print.orientation",
        result,
    )
    paper_size = optional_string(
        print_data.get("paper_size"),
        "print.paper_size",
        result,
    )
    print_area_required = print_data.get("print_area_required")
    if not isinstance(print_area_required, bool):
        add_red_flag(
            result,
            "contract field must be boolean: print.print_area_required",
        )
        return None
    if print_data.get("orientation") is not None and orientation is None:
        return None
    if print_data.get("paper_size") is not None and paper_size is None:
        return None
    return PrintContract(
        orientation=orientation,
        paper_size=paper_size,
        print_area_required=print_area_required,
    )


def validate_layout_contract(
    data: Any,
    result: TemplateContractPreflightResult,
) -> LayoutContract | None:
    if not isinstance(data, Mapping):
        add_red_flag(result, "contract field must be an object: layout")
        return None
    layout_data = cast(Mapping[str, Any], data)
    if not has_required_fields(layout_data, LAYOUT_FIELDS, "layout.", result):
        return None

    valid = True
    cells: dict[str, str | None] = {}
    for field_name in CELL_LAYOUT_FIELDS:
        value = nonempty_string(layout_data, field_name, "layout.", result)
        if value is None:
            valid = False
            continue
        if not valid_cell_coordinate(value):
            valid = False
            add_red_flag(
                result,
                f"contract field must be a valid Excel cell: layout.{field_name}",
            )
            continue
        cells[field_name] = normalize_cell(value)

    object_value = layout_data.get("object_cell")
    if object_value is None:
        cells["object_cell"] = None
    elif not isinstance(object_value, str) or object_value.strip() == "":
        valid = False
        add_red_flag(
            result,
            "contract field must be a valid Excel cell or null: layout.object_cell",
        )
    elif not valid_cell_coordinate(object_value):
        valid = False
        add_red_flag(
            result,
            "contract field must be a valid Excel cell or null: layout.object_cell",
        )
    else:
        cells["object_cell"] = normalize_cell(object_value)

    table_header_row = positive_integer(
        layout_data,
        "table_header_row",
        "layout.",
        result,
    )
    first_item_row = positive_integer(
        layout_data,
        "first_item_row",
        "layout.",
        result,
    )
    if table_header_row is None or first_item_row is None:
        valid = False
    elif first_item_row <= table_header_row:
        valid = False
        add_red_flag(
            result,
            "layout.first_item_row must be greater than layout.table_header_row",
        )

    item_columns_data = layout_data.get("item_columns")
    item_columns: dict[str, str] = {}
    if not isinstance(item_columns_data, Mapping):
        valid = False
        add_red_flag(result, "contract field must be an object: layout.item_columns")
    else:
        typed_columns = cast(Mapping[str, Any], item_columns_data)
        if not has_required_fields(
            typed_columns,
            ITEM_COLUMN_FIELDS,
            "layout.item_columns.",
            result,
        ):
            valid = False
        for field_name in ITEM_COLUMN_FIELDS:
            value = typed_columns.get(field_name)
            if not isinstance(value, str) or not valid_column(value):
                valid = False
                add_red_flag(
                    result,
                    "contract field must be a valid Excel column: "
                    f"layout.item_columns.{field_name}",
                )
            else:
                item_columns[field_name] = normalize_column(value)

    if (
        not valid
        or table_header_row is None
        or first_item_row is None
        or len(cells) != len(CELL_LAYOUT_FIELDS) + 1
        or len(item_columns) != len(ITEM_COLUMN_FIELDS)
    ):
        return None
    return LayoutContract(
        cells=cells,
        table_header_row=table_header_row,
        first_item_row=first_item_row,
        item_columns=item_columns,
    )


def validate_fixed_labels(
    data: Any,
    result: TemplateContractPreflightResult,
) -> tuple[FixedLabelContract, ...] | None:
    if not isinstance(data, list) or not data:
        add_red_flag(
            result,
            "contract field must be a non-empty list: required_fixed_labels",
        )
        return None

    valid = True
    labels: list[FixedLabelContract] = []
    for index, raw_label in enumerate(data, start=1):
        prefix = f"required_fixed_labels[{index}]."
        if not isinstance(raw_label, Mapping):
            valid = False
            add_red_flag(
                result,
                f"contract fixed label must be an object: item {index}",
            )
            continue
        label = cast(Mapping[str, Any], raw_label)
        if not has_required_fields(
            label,
            ("cell", "expected_text_contains"),
            prefix,
            result,
        ):
            valid = False
            continue
        cell = nonempty_string(label, "cell", prefix, result)
        expected = nonempty_string(
            label,
            "expected_text_contains",
            prefix,
            result,
        )
        if cell is None or expected is None:
            valid = False
            continue
        if not valid_cell_coordinate(cell):
            valid = False
            add_red_flag(
                result,
                f"contract fixed label cell is invalid: item {index}",
            )
            continue
        labels.append(
            FixedLabelContract(
                cell=normalize_cell(cell),
                expected_text_contains=expected,
            )
        )

    if not valid or len(labels) != len(data):
        return None
    return tuple(labels)


def validate_contract_schema(
    data: Mapping[str, Any] | None,
    result: TemplateContractPreflightResult,
) -> TemplateContract | None:
    if data is None:
        result.checks["contract schema"] = "fail"
        return None
    if not has_required_fields(data, TOP_LEVEL_FIELDS, "", result):
        result.checks["contract schema"] = "fail"
        return None

    contract_id = nonempty_string(data, "contract_id", "", result)
    template_name = nonempty_string(data, "template_name", "", result)
    template_version = nonempty_string(data, "template_version", "", result)
    expected_sheet_name = nonempty_string(
        data,
        "expected_sheet_name",
        "",
        result,
    )
    template_sha256 = nonempty_string(data, "template_sha256", "", result)
    valid = all(
        value is not None
        for value in (
            contract_id,
            template_name,
            template_version,
            expected_sheet_name,
            template_sha256,
        )
    )
    raw_template_sha256 = data.get("template_sha256")
    if (
        not isinstance(raw_template_sha256, str)
        or HASH_RE.fullmatch(raw_template_sha256) is None
    ):
        valid = False
        add_red_flag(
            result,
            "template_sha256 must be exactly 64 lowercase hex characters",
        )

    allowed_raw = data.get("allowed_extra_sheets")
    allowed_extra_sheets: tuple[str, ...] = ()
    if not isinstance(allowed_raw, list):
        valid = False
        add_red_flag(result, "allowed_extra_sheets must be a list")
    elif not all(isinstance(item, str) and item.strip() for item in allowed_raw):
        valid = False
        add_red_flag(
            result,
            "allowed_extra_sheets must contain only non-empty strings",
        )
    elif len(set(allowed_raw)) != len(allowed_raw):
        valid = False
        add_red_flag(result, "allowed_extra_sheets must not contain duplicates")
    else:
        allowed_extra_sheets = tuple(cast(list[str], allowed_raw))

    print_contract = validate_print_contract(data.get("print"), result)
    layout_contract = validate_layout_contract(data.get("layout"), result)
    fixed_labels = validate_fixed_labels(data.get("required_fixed_labels"), result)
    if print_contract is None or layout_contract is None or fixed_labels is None:
        valid = False

    if (
        not valid
        or contract_id is None
        or template_name is None
        or template_version is None
        or expected_sheet_name is None
        or template_sha256 is None
        or print_contract is None
        or layout_contract is None
        or fixed_labels is None
    ):
        result.checks["contract schema"] = "fail"
        return None

    result.checks["contract schema"] = "pass"
    return TemplateContract(
        contract_id=contract_id,
        template_name=template_name,
        template_version=template_version,
        expected_sheet_name=expected_sheet_name,
        template_sha256=template_sha256,
        allowed_extra_sheets=allowed_extra_sheets,
        print_contract=print_contract,
        layout=layout_contract,
        required_fixed_labels=fixed_labels,
    )


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as source:
        for chunk in iter(lambda: source.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def verify_template_hash(
    contract: TemplateContract | None,
    result: TemplateContractPreflightResult,
) -> str | None:
    if contract is None or not result.template_xlsx.is_file():
        add_red_flag(result, "template hash verification could not run safely")
        return None
    try:
        actual_hash = sha256_file(result.template_xlsx)
    except OSError:
        add_red_flag(result, "template XLSX could not be hashed")
        return None
    if actual_hash != contract.template_sha256:
        add_red_flag(result, "template SHA256 does not match contract")
        return actual_hash
    result.checks["template hash"] = "pass"
    return actual_hash


def coordinate_within_sheet(
    coordinate: str,
    max_row: int,
    max_column: int,
) -> bool:
    row, column = coordinate_to_tuple(coordinate)
    return row <= max_row and column <= max_column


def validate_workbook_layout(
    workbook: Any,
    contract: TemplateContract,
    result: TemplateContractPreflightResult,
) -> Any | None:
    expected = contract.expected_sheet_name
    if expected not in workbook.sheetnames:
        add_red_flag(result, "expected worksheet is missing")
        return None

    extras = [name for name in workbook.sheetnames if name != expected]
    unexpected = [name for name in extras if name not in contract.allowed_extra_sheets]
    if unexpected:
        add_red_flag(result, "workbook contains an unexpected extra worksheet")
        return None

    worksheet = workbook[expected]
    max_row = int(worksheet.max_row)
    max_column = int(worksheet.max_column)
    valid = True
    for field_name, coordinate in contract.layout.cells.items():
        if coordinate is None:
            continue
        if not coordinate_within_sheet(coordinate, max_row, max_column):
            valid = False
            add_red_flag(
                result,
                f"layout cell is outside worksheet used range: {field_name}",
            )

    if contract.layout.table_header_row > max_row:
        valid = False
        add_red_flag(result, "table header row is outside worksheet used range")
    if contract.layout.first_item_row > max_row:
        valid = False
        add_red_flag(result, "first item row is outside worksheet used range")

    for field_name, column in contract.layout.item_columns.items():
        column_index = column_index_from_string(column)
        if column_index > max_column:
            valid = False
            add_red_flag(
                result,
                f"item column is outside worksheet used range: {field_name}",
            )

    for index, label in enumerate(contract.required_fixed_labels, start=1):
        if not coordinate_within_sheet(label.cell, max_row, max_column):
            valid = False
            add_red_flag(
                result,
                f"fixed label cell is outside worksheet used range: item {index}",
            )

    result.checks["workbook layout"] = "pass" if valid else "fail"
    return worksheet if valid else None


def verify_fixed_labels(
    worksheet: Any | None,
    contract: TemplateContract,
    result: TemplateContractPreflightResult,
) -> bool:
    if worksheet is None:
        add_red_flag(result, "fixed label checks could not run safely")
        return False

    valid = True
    for index, label in enumerate(contract.required_fixed_labels, start=1):
        actual = worksheet[label.cell].value
        actual_text = "" if actual is None else str(actual)
        if label.expected_text_contains not in actual_text:
            valid = False
            add_red_flag(result, f"required fixed label mismatch at item {index}")

    result.checks["fixed labels"] = "pass" if valid else "fail"
    return valid


def relationship_target(
    relationships_root: ElementTree.Element,
    relationship_id: str,
) -> str | None:
    relationship_tag = f"{{{PACKAGE_REL_NS}}}Relationship"
    for relationship in relationships_root.findall(relationship_tag):
        if relationship.get("Id") == relationship_id:
            return relationship.get("Target")
    return None


def worksheet_part_name(target: str) -> str:
    normalized = target.replace("\\", "/")
    if normalized.startswith("/"):
        return normalized.lstrip("/")
    return posixpath.normpath(posixpath.join("xl", normalized))


def read_actual_print_setup(
    template_path: Path,
    sheet_name: str,
) -> ActualPrintSetup:
    with zipfile.ZipFile(template_path, "r") as archive:
        workbook_root = ElementTree.fromstring(archive.read("xl/workbook.xml"))
        relationships_root = ElementTree.fromstring(
            archive.read("xl/_rels/workbook.xml.rels")
        )
        sheet_tag = f"{{{SPREADSHEET_NS}}}sheet"
        relationship_attribute = f"{{{DOCUMENT_REL_NS}}}id"
        sheets = workbook_root.findall(f"{{{SPREADSHEET_NS}}}sheets/{sheet_tag}")

        sheet_index: int | None = None
        relationship_id: str | None = None
        for index, sheet in enumerate(sheets):
            if sheet.get("name") == sheet_name:
                sheet_index = index
                relationship_id = sheet.get(relationship_attribute)
                break
        if sheet_index is None or relationship_id is None:
            raise ValueError("expected worksheet relationship is missing")

        target = relationship_target(relationships_root, relationship_id)
        if target is None:
            raise ValueError("expected worksheet part is missing")
        worksheet_root = ElementTree.fromstring(
            archive.read(worksheet_part_name(target))
        )
        page_setup = worksheet_root.find(f"{{{SPREADSHEET_NS}}}pageSetup")
        orientation = None if page_setup is None else page_setup.get("orientation")
        paper_size = None if page_setup is None else page_setup.get("paperSize")

        has_print_area = False
        defined_name_tag = f"{{{SPREADSHEET_NS}}}definedName"
        defined_names = workbook_root.find(f"{{{SPREADSHEET_NS}}}definedNames")
        if defined_names is not None:
            for defined_name in defined_names.findall(defined_name_tag):
                if (
                    defined_name.get("name") == "_xlnm.Print_Area"
                    and defined_name.get("localSheetId") == str(sheet_index)
                    and (defined_name.text or "").strip() != ""
                ):
                    has_print_area = True
                    break

    return ActualPrintSetup(
        orientation=orientation,
        paper_size=paper_size,
        has_print_area=has_print_area,
    )


def verify_print_setup(
    worksheet: Any | None,
    contract: TemplateContract,
    result: TemplateContractPreflightResult,
) -> bool:
    if worksheet is None:
        add_red_flag(result, "print setup checks could not run safely")
        return False

    valid = True
    expected = contract.print_contract
    try:
        actual = read_actual_print_setup(
            result.template_xlsx,
            contract.expected_sheet_name,
        )
    except (
        ElementTree.ParseError,
        KeyError,
        OSError,
        ValueError,
        zipfile.BadZipFile,
    ):
        add_red_flag(result, "print setup could not be read safely")
        return False

    if expected.orientation is not None:
        if (
            actual.orientation is None
            or actual.orientation.casefold() != expected.orientation.casefold()
        ):
            valid = False
            add_red_flag(result, "print orientation does not match contract")
    if expected.paper_size is not None:
        if actual.paper_size is None or actual.paper_size != expected.paper_size:
            valid = False
            add_red_flag(result, "print paper size does not match contract")
    if expected.print_area_required and not actual.has_print_area:
        valid = False
        add_red_flag(result, "print area is required but missing")

    result.checks["print setup"] = "pass" if valid else "fail"
    return valid


def inspect_workbook(
    contract: TemplateContract | None,
    hash_ok: bool,
    result: TemplateContractPreflightResult,
) -> None:
    if contract is None or not hash_ok:
        add_red_flag(result, "workbook checks could not run safely")
        return

    try:
        workbook = load_workbook(
            result.template_xlsx,
            read_only=True,
            data_only=False,
            keep_links=False,
        )
    except Exception:
        add_red_flag(result, "template XLSX could not be opened")
        return

    try:
        try:
            worksheet = validate_workbook_layout(workbook, contract, result)
            verify_fixed_labels(worksheet, contract, result)
            verify_print_setup(worksheet, contract, result)
        except Exception:
            add_red_flag(result, "workbook inspection failed safely")
    finally:
        workbook.close()


def verify_template_unchanged(
    hash_before: str | None,
    result: TemplateContractPreflightResult,
) -> None:
    if hash_before is None or not result.template_xlsx.is_file():
        add_red_flag(result, "template immutability check could not run safely")
        return
    try:
        hash_after = sha256_file(result.template_xlsx)
    except OSError:
        add_red_flag(result, "template immutability check could not run safely")
        return
    if hash_before != hash_after:
        add_red_flag(result, "template XLSX changed during preflight")
        return
    result.checks["safety boundaries"] = "pass"


def preflight(
    template_xlsx: Path,
    contract_json: Path,
) -> TemplateContractPreflightResult:
    result = TemplateContractPreflightResult(
        template_xlsx=resolved(template_xlsx),
        contract_json=resolved(contract_json),
    )
    validate_input_paths(result)
    raw_contract = load_contract_json(result.contract_json, result)
    contract = validate_contract_schema(raw_contract, result)
    hash_before = verify_template_hash(contract, result)
    hash_ok = result.checks["template hash"] == "pass"
    inspect_workbook(contract, hash_ok, result)
    verify_template_unchanged(hash_before, result)

    all_checks_pass = all(status == "pass" for status in result.checks.values())
    result.status = "PASS" if all_checks_pass and not result.red_flags else "FAIL"
    return result


def format_items(values: Sequence[str]) -> list[str]:
    return list(values) if values else ["none"]


def format_report(result: TemplateContractPreflightResult) -> str:
    lines = [
        REPORT_START,
        "",
        "Status:",
        result.status,
        "",
        "Mode:",
        MODE,
        "",
        "Checks:",
    ]
    lines.extend(f"{name}: {status}" for name, status in result.checks.items())
    lines.extend(["", "Red flags:"])
    lines.extend(format_items(result.red_flags))
    lines.extend(
        [
            "",
            "Contract path policy:",
            CONTRACT_PATH_POLICY,
            "",
            "Commercial status:",
            COMMERCIAL_STATUS,
            "",
            "Human Approval:",
            HUMAN_APPROVAL,
            "",
            REPORT_END,
        ]
    )
    return "\n".join(lines)


def main(argv: Sequence[str] | None = None) -> int:
    args = parse_args(argv)
    result = preflight(args.template_xlsx, args.contract_json)
    print(format_report(result))
    return 0 if result.status == "PASS" else 1


if __name__ == "__main__":
    raise SystemExit(main())
