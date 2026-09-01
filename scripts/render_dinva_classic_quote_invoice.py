"""Render a new DINVA classic quote/invoice XLSX from approved inputs."""

from __future__ import annotations

import argparse
import base64
import hashlib
import importlib.util
import json
import os
import uuid
from collections.abc import Mapping, Sequence
from datetime import UTC, date, datetime
from decimal import ROUND_HALF_UP, Decimal
from pathlib import Path
from types import ModuleType
from typing import Any, NoReturn, cast
from xml.etree import ElementTree
from xml.sax.saxutils import escape
from zipfile import ZIP_DEFLATED, BadZipFile, ZipFile

from openpyxl import Workbook  # type: ignore[import-untyped]
from openpyxl.cell.cell import Cell  # type: ignore[import-untyped]
from openpyxl.styles import (  # type: ignore[import-untyped]
    Alignment,
    Border,
    Color,
    Font,
    PatternFill,
    Side,
)

PROJECT_ROOT = Path(__file__).resolve().parents[1]
VALIDATOR_PATH = Path(__file__).with_name("validate_dinva_classic_quote_invoice.py")
PROFILE_SCHEMA_VERSION = "dinva_classic_presentation_profile.v0.1"
DOCUMENT_SCHEMA_VERSION = "dinva_quote_invoice_document.v0.1"
FAMILY = "DINVA_CLASSIC_QUOTE_INVOICE_V0_1"
TEST_MODE_ENV = "DINVA_RENDERER_TEST_MODE"
SPREADSHEET_NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
OFFICE_REL_NS = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
PACKAGE_REL_NS = "http://schemas.openxmlformats.org/package/2006/relationships"
CONTENT_TYPES_NS = "http://schemas.openxmlformats.org/package/2006/content-types"
CUSTOM_PROPS_NS = (
    "http://schemas.openxmlformats.org/officeDocument/2006/custom-properties"
)
VT_NS = "http://schemas.openxmlformats.org/officeDocument/2006/docPropsVTypes"
DRAWING_NS = "http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing"
DRAWING_MAIN_NS = "http://schemas.openxmlformats.org/drawingml/2006/main"
CUSTOM_PROPERTY_FORMAT_ID = "{D5CDD505-2E9C-101B-9397-08002B2CF9AE}"
PROFILE_KEYS = {
    "schema_version",
    "profile_id",
    "document_family",
    "artifact_status",
    "reference_provenance",
    "presentation_contract",
    "presentation_contract_fingerprint",
    "approval_provenance",
}
DOCUMENT_KEYS = {
    "schema_version",
    "document_family",
    "document_type",
    "document_id",
    "document_number",
    "document_date",
    "currency",
    "payer",
    "object_name",
    "basis",
    "apparatus_heading",
    "items",
    "approved_grand_total_kzt",
    "vat",
    "amount_words",
    "terms",
    "signatures",
    "approval_provenance",
}


class RendererError(ValueError):
    """A clean render would violate the approved contract."""


def fail(message: str) -> NoReturn:
    raise RendererError(message)


def require(condition: bool, message: str) -> None:
    if not condition:
        fail(message)


def resolved(path: Path) -> Path:
    return path.expanduser().resolve(strict=False)


def is_inside_project(path: Path) -> bool:
    return resolved(path).is_relative_to(resolved(PROJECT_ROOT))


def sha256_bytes(raw: bytes) -> str:
    return hashlib.sha256(raw).hexdigest()


def canonical_json(value: object) -> bytes:
    return json.dumps(
        value,
        ensure_ascii=False,
        sort_keys=True,
        separators=(",", ":"),
        allow_nan=False,
    ).encode("utf-8")


def reject_duplicate_keys(pairs: list[tuple[str, Any]]) -> dict[str, Any]:
    result: dict[str, Any] = {}
    for key, value in pairs:
        require(key not in result, f"duplicate JSON key: {key}")
        result[key] = value
    return result


def load_bound_json(
    path: Path, expected_sha256: str, label: str
) -> tuple[Path, bytes, dict[str, Any]]:
    actual = path.resolve(strict=True)
    require(not is_inside_project(actual), f"real {label} must be outside Git")
    raw = actual.read_bytes()
    require(sha256_bytes(raw) == expected_sha256, f"{label} SHA-256 mismatch")
    try:
        payload = json.loads(
            raw.decode("utf-8", errors="strict"),
            object_pairs_hook=reject_duplicate_keys,
        )
    except (UnicodeDecodeError, json.JSONDecodeError) as exc:
        raise RendererError(f"{label} is not strict UTF-8 JSON: {exc}") from exc
    require(isinstance(payload, Mapping), f"{label} root must be an object")
    return actual, raw, dict(payload)


def exact_keys(value: Mapping[str, Any], expected: set[str], label: str) -> None:
    require(set(value) == expected, f"{label} fields mismatch")


def mapping(value: Any, label: str) -> Mapping[str, Any]:
    require(isinstance(value, Mapping), f"{label} must be an object")
    return cast(Mapping[str, Any], value)


def text(value: Any, label: str, *, nullable: bool = False) -> str | None:
    if value is None and nullable:
        return None
    require(
        isinstance(value, str) and bool(value.strip()),
        f"{label} must be non-empty text",
    )
    return cast(str, value)


def integer(value: Any, label: str, *, minimum: int = 0) -> int:
    require(
        type(value) is int and value >= minimum,
        f"{label} must be an integer >= {minimum}",
    )
    return cast(int, value)


def test_mode(requested: bool) -> bool:
    return requested and os.environ.get(TEST_MODE_ENV) == "1"


def validate_profile(
    profile: Mapping[str, Any], *, allow_test_profile: bool
) -> Mapping[str, Any]:
    exact_keys(profile, PROFILE_KEYS, "profile")
    require(
        profile.get("schema_version") == PROFILE_SCHEMA_VERSION,
        "profile schema mismatch",
    )
    require(profile.get("profile_id") == FAMILY, "profile id mismatch")
    require(
        profile.get("document_family") == FAMILY, "unknown/unsupported profile family"
    )
    contract = mapping(profile.get("presentation_contract"), "presentation contract")
    fingerprint = sha256_bytes(canonical_json(contract))
    require(
        profile.get("presentation_contract_fingerprint") == fingerprint,
        "profile contract fingerprint mismatch",
    )
    approval = mapping(
        profile.get("approval_provenance"), "profile approval provenance"
    )
    if test_mode(allow_test_profile):
        require(
            approval.get("status") in {"DRAFT_UNAPPROVED", "APPROVED"},
            "test profile status invalid",
        )
    else:
        require(
            profile.get("artifact_status") == "IMMUTABLE_APPROVED_PROFILE",
            "profile is not immutable/approved",
        )
        require(approval.get("status") == "APPROVED", "profile is DRAFT/unapproved")
        require(
            approval.get("authority") == "IGOR_DIRECT_HUMAN_APPROVAL",
            "profile approval authority mismatch",
        )
        require(
            approval.get("approved_contract_fingerprint") == fingerprint,
            "profile approval fingerprint mismatch",
        )
    require(
        contract.get("contract_version") == "dinva_classic_presentation_contract.v0.1",
        "presentation contract version mismatch",
    )
    workbook = mapping(contract.get("workbook"), "workbook contract")
    sheets = workbook.get("sheets")
    require(
        isinstance(sheets, list)
        and len(sheets) == 1
        and mapping(cast(list[Any], sheets)[0], "primary sheet").get("role")
        == "PRIMARY_DOCUMENT"
        and isinstance(cast(list[Any], sheets)[0].get("name"), str),
        "sheet contract mismatch",
    )
    require(
        workbook.get("extra_sheets_allowed") is False, "extra sheets must be closed"
    )
    layout = mapping(contract.get("layout"), "layout contract")
    require(
        layout.get("table_header_row") == 15
        and type(layout.get("first_item_row")) is int
        and type(layout.get("section_row")) is int
        and layout.get("section_row") == cast(int, layout.get("first_item_row")) - 1
        and type(layout.get("item_capacity")) is int
        and cast(int, layout.get("item_capacity")) > 0,
        "classic table geometry mismatch",
    )
    assets = contract.get("assets")
    require(
        isinstance(assets, list) and len(assets) == 1,
        "asset contract must contain one logo",
    )
    asset = mapping(cast(list[Any], assets)[0], "logo asset")
    try:
        logo = base64.b64decode(cast(str, asset.get("data_base64")), validate=True)
    except (ValueError, TypeError) as exc:
        raise RendererError("logo asset base64 is invalid") from exc
    require(sha256_bytes(logo) == asset.get("sha256"), "logo asset SHA-256 mismatch")
    return contract


def validate_document(document: Mapping[str, Any], *, allow_test_profile: bool) -> None:
    exact_keys(document, DOCUMENT_KEYS, "document")
    require(
        document.get("schema_version") == DOCUMENT_SCHEMA_VERSION,
        "document schema mismatch",
    )
    require(
        document.get("document_family") == FAMILY, "unknown/unsupported document family"
    )
    require(
        document.get("document_type") in {"QUOTE", "INVOICE", "QUOTE_INVOICE"},
        "document type mismatch",
    )
    for field in ("document_id", "document_number", "payer", "apparatus_heading"):
        text(document.get(field), field)
    try:
        date.fromisoformat(cast(str, document.get("document_date")))
    except (TypeError, ValueError) as exc:
        raise RendererError("document date must be ISO date") from exc
    require(document.get("currency") == "KZT", "document currency mismatch")
    text(document.get("object_name"), "object_name", nullable=True)
    text(document.get("basis"), "basis", nullable=True)
    items = document.get("items")
    require(isinstance(items, list) and bool(items), "document items must be non-empty")
    expected_item_keys = {
        "position",
        "name",
        "unit",
        "quantity",
        "detailed_technical_composition",
        "apparatus",
        "enclosure",
        "approved_unit_price_kzt",
        "approved_line_total_kzt",
        "approval_reference",
    }
    total = 0
    for expected_position, raw_item in enumerate(cast(list[Any], items), start=1):
        item = mapping(raw_item, f"item {expected_position}")
        exact_keys(item, expected_item_keys, f"item {expected_position}")
        require(
            integer(item.get("position"), "item position", minimum=1)
            == expected_position,
            "item positions must be contiguous",
        )
        for field in (
            "name",
            "unit",
            "detailed_technical_composition",
            "apparatus",
            "enclosure",
            "approval_reference",
        ):
            text(item.get(field), f"item {expected_position}.{field}")
        composition = cast(str, item["detailed_technical_composition"])
        apparatus = cast(str, item["apparatus"])
        require(
            apparatus in composition,
            f"item {expected_position} apparatus is not exactly represented "
            "in detailed composition",
        )
        quantity = integer(item.get("quantity"), "quantity", minimum=1)
        unit_price = integer(item.get("approved_unit_price_kzt"), "unit price")
        line_total = integer(item.get("approved_line_total_kzt"), "line total")
        require(
            quantity * unit_price == line_total,
            f"item {expected_position} approved arithmetic mismatch",
        )
        total += line_total
    require(
        integer(document.get("approved_grand_total_kzt"), "grand total") == total,
        "approved grand total mismatch",
    )
    amount_words = mapping(document.get("amount_words"), "amount words")
    exact_keys(amount_words, {"amount_kzt", "approved_text"}, "amount words")
    require(
        integer(amount_words.get("amount_kzt"), "amount words amount") == total,
        "amount words amount mismatch",
    )
    text(amount_words.get("approved_text"), "amount words text")
    vat = mapping(document.get("vat"), "VAT")
    exact_keys(
        vat, {"rate_percent", "included", "approved_amount_kzt", "approved_text"}, "VAT"
    )
    rate = integer(vat.get("rate_percent"), "VAT rate")
    require(rate <= 100, "VAT rate exceeds 100")
    require(type(vat.get("included")) is bool, "VAT included must be boolean")
    divisor = Decimal(100 + rate) if vat["included"] else Decimal(100)
    expected_vat = int(
        (Decimal(total) * Decimal(rate) / divisor).quantize(
            Decimal("1"), rounding=ROUND_HALF_UP
        )
    )
    require(
        integer(vat.get("approved_amount_kzt"), "VAT amount") == expected_vat,
        "approved VAT arithmetic mismatch",
    )
    text(vat.get("approved_text"), "VAT text")
    terms = mapping(document.get("terms"), "terms")
    exact_keys(
        terms, {"payment", "delivery", "manufacturing_lead_time", "validity"}, "terms"
    )
    for field in ("payment", "delivery", "manufacturing_lead_time"):
        text(terms.get(field), f"terms.{field}")
    text(terms.get("validity"), "terms.validity", nullable=True)
    signatures = mapping(document.get("signatures"), "signatures")
    exact_keys(
        signatures,
        {"director_title", "director_name", "executor_title", "executor_name"},
        "signatures",
    )
    for field in signatures:
        text(signatures[field], f"signatures.{field}")
    approval = mapping(document.get("approval_provenance"), "document approval")
    exact_keys(
        approval,
        {
            "status",
            "authority",
            "approval_id",
            "approved_at",
            "source_sha256s",
            "rendering_authorized",
            "client_send_authorized",
        },
        "document approval",
    )
    require(approval.get("status") == "APPROVED", "document is not approved")
    if not test_mode(allow_test_profile):
        require(
            approval.get("authority") == "IGOR_DIRECT_HUMAN_APPROVAL",
            "document approval authority mismatch",
        )
    text(approval.get("authority"), "approval authority")
    text(approval.get("approval_id"), "approval id")
    text(approval.get("approved_at"), "approved_at")
    source_shas = approval.get("source_sha256s")
    require(
        isinstance(source_shas, list) and bool(source_shas),
        "document source hashes missing",
    )
    require(
        all(
            isinstance(item, str) and len(item) == 64
            for item in cast(list[Any], source_shas)
        ),
        "document source hash invalid",
    )
    require(
        approval.get("rendering_authorized") is True,
        "document rendering is not authorized",
    )
    require(
        approval.get("client_send_authorized") is False,
        "renderer cannot consume sending authorization",
    )


def color_from_spec(spec: Mapping[str, Any] | None) -> Color | None:
    if spec is None:
        return None
    kind = spec.get("type")
    value = spec.get("value")
    tint = float(spec.get("tint", 0))
    if kind == "rgb":
        return Color(rgb=cast(str, value), tint=tint)
    if kind == "indexed":
        return Color(indexed=int(cast(int, value)), tint=tint)
    if kind == "theme":
        return Color(theme=int(cast(int, value)), tint=tint)
    fail("unsupported profile color")


def apply_style(cell: Cell, raw_style: Mapping[str, Any]) -> None:
    font = mapping(raw_style.get("font"), "style font")
    fill = mapping(raw_style.get("fill"), "style fill")
    border = mapping(raw_style.get("border"), "style border")
    alignment = mapping(raw_style.get("alignment"), "style alignment")
    cell.font = Font(
        name=cast(str, font["name"]),
        size=float(cast(float, font["size"])),
        bold=bool(font["bold"]),
        italic=bool(font["italic"]),
        underline=cast(str | None, font["underline"]),
        color=color_from_spec(cast(Mapping[str, Any] | None, font["color"])),
    )
    cell.fill = PatternFill(
        fill_type=cast(str | None, fill["type"]),
        fgColor=color_from_spec(cast(Mapping[str, Any] | None, fill["foreground"])),
    )
    cell.border = Border(
        left=Side(style=cast(str | None, border["left"]), color="FF000000"),
        right=Side(style=cast(str | None, border["right"]), color="FF000000"),
        top=Side(style=cast(str | None, border["top"]), color="FF000000"),
        bottom=Side(style=cast(str | None, border["bottom"]), color="FF000000"),
    )
    cell.alignment = Alignment(
        horizontal=cast(str | None, alignment["horizontal"]),
        vertical=cast(str | None, alignment["vertical"]),
        wrap_text=bool(alignment["wrap_text"]),
        shrink_to_fit=bool(alignment["shrink_to_fit"]),
    )
    cell.number_format = cast(str, raw_style["number_format"])


def item_height(texts: Sequence[str], rule: Mapping[str, Any]) -> float:
    longest = max(len(value) for value in texts)
    base = float(rule["base"])
    threshold = integer(rule["characters_per_increment"], "height threshold", minimum=1)
    increment = float(rule["increment"])
    maximum = float(rule["maximum"])
    extra = max(0, (longest - 1) // threshold)
    return min(maximum, base + extra * increment)


def render_clean_workbook(
    profile: Mapping[str, Any],
    profile_sha256: str,
    document: Mapping[str, Any],
    document_sha256: str,
    output: Path,
) -> None:
    contract = mapping(profile["presentation_contract"], "presentation contract")
    layout = mapping(contract["layout"], "layout")
    styles = mapping(contract["styles"], "styles")
    fixed = mapping(contract["fixed_blocks"], "fixed blocks")
    workbook = Workbook()
    worksheet = workbook.active
    sheet_contract = cast(
        list[Mapping[str, Any]], mapping(contract["workbook"], "workbook")["sheets"]
    )[0]
    worksheet.title = cast(str, sheet_contract["name"])
    workbook.properties.creator = "DINVA deterministic renderer"
    fixed_time = datetime(2000, 1, 1, tzinfo=UTC)
    workbook.properties.created = fixed_time
    workbook.properties.modified = fixed_time
    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True
    workbook.calculation.calcMode = "auto"
    worksheet.sheet_view.showGridLines = False
    for column, width in mapping(layout["column_widths"], "column widths").items():
        worksheet.column_dimensions[column].width = float(cast(float, width))
    company = mapping(fixed["company"], "company block")
    for coordinate, value in company.items():
        worksheet[coordinate] = value
        apply_style(
            worksheet[coordinate],
            mapping(
                styles["company_title" if coordinate == "C2" else "company_info"],
                "company style",
            ),
        )
    worksheet["B9"] = (
        f"Счёт-КП № {document['document_number']} от {document['document_date']}"
    )
    worksheet["B10"] = f"Плательщик: {document['payer']}"
    worksheet["B11"] = f"Объект: {document['object_name'] or 'не указан'}"
    worksheet["B12"] = f"Основание / проект: {document['basis'] or 'не указано'}"
    worksheet["B13"] = "Статус документа: DRAFT XLSX"
    worksheet["G9"] = "ВНИМАНИЕ!"
    terms = mapping(document["terms"], "terms")
    worksheet["G10"] = f"Срок изготовления: {terms['manufacturing_lead_time']}"
    worksheet["G11"] = f"Оплата: {terms['payment']}"
    worksheet["G12"] = f"Поставка: {terms['delivery']}"
    worksheet["G13"] = "Technical PASS не является разрешением на отправку"
    for coordinate in ("B9", "B10", "B11", "B12", "B13"):
        apply_style(
            worksheet[coordinate], mapping(styles["company_info"], "metadata style")
        )
    for coordinate in ("G9", "G10", "G11", "G12", "G13"):
        apply_style(
            worksheet[coordinate],
            mapping(
                styles["warning" if coordinate == "G9" else "company_info"],
                "warning style",
            ),
        )
    header_row = integer(layout["table_header_row"], "table header row", minimum=1)
    headers = dict(mapping(fixed["table_headers"], "table headers"))
    headers["F"] = document["apparatus_heading"]
    for column in "BCDEFGHI":
        coordinate = f"{column}{header_row}"
        worksheet[coordinate] = headers[column]
        apply_style(
            worksheet[coordinate], mapping(styles["table_header"], "table header style")
        )
    first_row = integer(layout["first_item_row"], "first item row", minimum=1)
    section_row = integer(layout["section_row"], "section row", minimum=1)
    capacity = integer(layout["item_capacity"], "item capacity", minimum=1)
    items = cast(list[Mapping[str, Any]], document["items"])
    require(len(items) <= capacity, "document exceeds governed item capacity")
    height_rule = mapping(layout["item_height_rule"], "item height rule")
    section_value = (
        document["object_name"] or document["basis"] or document["document_id"]
    )
    worksheet[f"C{section_row}"] = f"{fixed['section_label_prefix']} {section_value}"
    apply_style(worksheet[f"C{section_row}"], mapping(styles["terms"], "section style"))
    style_by_column = {
        "B": "position",
        "C": "item_name",
        "D": "unit",
        "E": "quantity",
        "F": "technical_composition",
        "G": "enclosure",
        "H": "money",
        "I": "line_total",
    }
    for row in range(first_row, first_row + capacity):
        worksheet.row_dimensions[row].height = float(height_rule["base"])
        for column, style_name in style_by_column.items():
            apply_style(
                worksheet[f"{column}{row}"],
                mapping(styles[style_name], f"{column} item style"),
            )
    line_template = cast(
        str, mapping(contract["formulas"], "formulas")["line_total_template"]
    )
    for offset, item in enumerate(items):
        row = first_row + offset
        values = {
            "B": item["position"],
            "C": item["name"],
            "D": item["unit"],
            "E": item["quantity"],
            "F": item["detailed_technical_composition"],
            "G": item["enclosure"],
            "H": item["approved_unit_price_kzt"],
            "I": line_template.format(row=row),
        }
        for column, value in values.items():
            worksheet[f"{column}{row}"] = value
        worksheet.row_dimensions[row].height = item_height(
            [
                cast(str, item["name"]),
                cast(str, item["detailed_technical_composition"]),
                cast(str, item["enclosure"]),
            ],
            height_rule,
        )
    total_row = integer(layout["total_row"], "total row", minimum=1)
    worksheet[f"H{total_row}"] = fixed["total_label"]
    grand_template = cast(
        str, mapping(contract["formulas"], "formulas")["grand_total_template"]
    )
    worksheet[f"I{total_row}"] = grand_template.format(
        start=first_row, end=first_row + capacity - 1
    )
    apply_style(worksheet[f"H{total_row}"], mapping(styles["total"], "total style"))
    apply_style(worksheet[f"I{total_row}"], mapping(styles["total"], "total style"))
    vat_row = integer(layout["vat_row"], "VAT row", minimum=1)
    vat = mapping(document["vat"], "VAT")
    worksheet[f"H{vat_row}"] = vat["approved_text"]
    worksheet[f"I{vat_row}"] = vat["approved_amount_kzt"]
    apply_style(worksheet[f"H{vat_row}"], mapping(styles["money"], "VAT style"))
    apply_style(worksheet[f"I{vat_row}"], mapping(styles["money"], "VAT style"))
    amount_row = integer(layout["amount_words_row"], "amount row", minimum=1)
    worksheet[f"C{amount_row}"] = mapping(document["amount_words"], "amount words")[
        "approved_text"
    ]
    apply_style(
        worksheet[f"C{amount_row}"],
        mapping(styles["amount_words"], "amount words style"),
    )
    guard_lines = mapping(fixed["guard_lines"], "guard lines")
    term_values = [
        f"Срок действия: {terms['validity'] or 'не указан'}",
        f"Условия оплаты: {terms['payment']}. Условия поставки: {terms['delivery']}.",
        f"Срок изготовления: {terms['manufacturing_lead_time']}",
        guard_lines["specification"],
        guard_lines["no_send"],
    ]
    term_rows = cast(list[int], layout["terms_rows"])
    require(len(term_rows) == len(term_values), "terms region mismatch")
    for row, value in zip(term_rows, term_values, strict=True):
        worksheet[f"C{row}"] = value
        apply_style(
            worksheet[f"C{row}"],
            mapping(styles["terms"], "terms style"),
        )
    signatures = mapping(document["signatures"], "signatures")
    signature_rows = cast(list[int], layout["signature_rows"])
    require(len(signature_rows) == 3, "signature region mismatch")
    director_row, executor_row, review_row = signature_rows
    worksheet[f"B{director_row}"] = signatures["director_title"]
    worksheet[f"F{director_row}"] = signatures["director_name"]
    worksheet[f"B{executor_row}"] = signatures["executor_title"]
    worksheet[f"F{executor_row}"] = signatures["executor_name"]
    for coordinate in (
        f"B{director_row}",
        f"F{director_row}",
        f"B{executor_row}",
        f"F{executor_row}",
    ):
        apply_style(
            worksheet[coordinate], mapping(styles["signature"], "signature style")
        )
    worksheet[f"B{review_row}"] = guard_lines["review_date"]
    apply_style(
        worksheet[f"B{review_row}"], mapping(styles["signature"], "review style")
    )
    for merged_range in cast(
        list[str], mapping(layout["merged_cells"], "merged cells")["ranges"]
    ):
        worksheet.merge_cells(merged_range)
    print_contract = mapping(contract["print"], "print contract")
    worksheet.page_setup.paperSize = cast(str, print_contract["paper_size"])
    worksheet.page_setup.orientation = cast(str, print_contract["orientation"])
    worksheet.page_setup.scale = integer(
        print_contract["scale"], "page scale", minimum=1
    )
    worksheet.page_setup.fitToHeight = integer(
        print_contract["fit_to_height"], "fit height"
    )
    worksheet.sheet_properties.pageSetUpPr.fitToPage = bool(
        print_contract["fit_to_page"]
    )
    margins = mapping(print_contract["margins"], "margins")
    for name in ("left", "right", "top", "bottom", "header", "footer"):
        setattr(worksheet.page_margins, name, float(cast(float, margins[name])))
    final_row = integer(layout["final_row"], "final row", minimum=1)
    worksheet.print_area = f"B1:I{final_row}"
    workbook.save(output)
    workbook.close()
    asset = mapping(cast(list[Any], contract["assets"])[0], "logo asset")
    inject_governed_parts(
        output,
        base64.b64decode(cast(str, asset["data_base64"]), validate=True),
        mapping(asset["placement"], "logo placement"),
        {
            "DINVA_PROFILE_ID": cast(str, profile["profile_id"]),
            "DINVA_DOCUMENT_FAMILY": FAMILY,
            "DINVA_PROFILE_SHA256": profile_sha256,
            "DINVA_PRESENTATION_FINGERPRINT": cast(
                str, profile["presentation_contract_fingerprint"]
            ),
            "DINVA_DOCUMENT_SHA256": document_sha256,
        },
    )


def archive_parts(path: Path) -> dict[str, bytes]:
    try:
        with ZipFile(path) as archive:
            return {name: archive.read(name) for name in archive.namelist()}
    except BadZipFile as exc:
        raise RendererError(f"candidate is not a valid XLSX package: {exc}") from exc


def drawing_xml(placement: Mapping[str, Any]) -> bytes:
    require(placement.get("anchor_type") == "ONE_CELL", "logo anchor type mismatch")
    start = mapping(placement["from"], "logo from anchor")
    extent = mapping(placement["extent"], "logo extent")

    def marker(name: str, value: Mapping[str, Any]) -> str:
        column = integer(value["column"], "anchor column")
        column_offset = integer(value["column_offset"], "anchor column offset")
        row = integer(value["row"], "anchor row")
        row_offset = integer(value["row_offset"], "anchor row offset")
        return (
            f"<xdr:{name}><xdr:col>{column}</xdr:col>"
            f"<xdr:colOff>{column_offset}</xdr:colOff>"
            f"<xdr:row>{row}</xdr:row>"
            f"<xdr:rowOff>{row_offset}</xdr:rowOff>"
            f"</xdr:{name}>"
        )

    return (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        f'<xdr:wsDr xmlns:xdr="{DRAWING_NS}" xmlns:a="{DRAWING_MAIN_NS}" '
        f'xmlns:r="{OFFICE_REL_NS}"><xdr:oneCellAnchor>'
        + marker("from", start)
        + f'<xdr:ext cx="{integer(extent["cx"], "logo extent cx")}" '
        f'cy="{integer(extent["cy"], "logo extent cy")}"/>'
        + '<xdr:pic><xdr:nvPicPr><xdr:cNvPr id="1" name="DINVA classic logo"/>'
        '<xdr:cNvPicPr><a:picLocks noChangeAspect="1"/></xdr:cNvPicPr></xdr:nvPicPr>'
        '<xdr:blipFill><a:blip r:embed="rId1"/><a:stretch><a:fillRect/></a:stretch>'
        '</xdr:blipFill><xdr:spPr><a:prstGeom prst="rect"><a:avLst/></a:prstGeom>'
        "<a:noFill/><a:ln><a:noFill/></a:ln></xdr:spPr></xdr:pic><xdr:clientData/>"
        "</xdr:oneCellAnchor></xdr:wsDr>"
    ).encode("utf-8")


def custom_properties_xml(values: Mapping[str, str]) -> bytes:
    properties = []
    for pid, (name, value) in enumerate(sorted(values.items()), start=2):
        properties.append(
            f'<property fmtid="{CUSTOM_PROPERTY_FORMAT_ID}" pid="{pid}" '
            f'name="{escape(name)}"><vt:lpwstr>{escape(value)}</vt:lpwstr></property>'
        )
    return (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        f'<Properties xmlns="{CUSTOM_PROPS_NS}" xmlns:vt="{VT_NS}">'
        + "".join(properties)
        + "</Properties>"
    ).encode("utf-8")


def inject_governed_parts(
    path: Path,
    logo: bytes,
    placement: Mapping[str, Any],
    custom_properties: Mapping[str, str],
) -> None:
    parts = archive_parts(path)
    sheet_name = "xl/worksheets/sheet1.xml"
    root = ElementTree.fromstring(parts[sheet_name])
    drawing = ElementTree.SubElement(root, f"{{{SPREADSHEET_NS}}}drawing")
    drawing.set(f"{{{OFFICE_REL_NS}}}id", "rId1")
    ElementTree.register_namespace("", SPREADSHEET_NS)
    ElementTree.register_namespace("r", OFFICE_REL_NS)
    parts[sheet_name] = ElementTree.tostring(
        root, encoding="utf-8", xml_declaration=True
    )
    parts["xl/worksheets/_rels/sheet1.xml.rels"] = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        f'<Relationships xmlns="{PACKAGE_REL_NS}"><Relationship Id="rId1" '
        f'Type="{OFFICE_REL_NS}/drawing" Target="../drawings/drawing1.xml"/>'
        "</Relationships>"
    ).encode()
    parts["xl/drawings/drawing1.xml"] = drawing_xml(placement)
    parts["xl/drawings/_rels/drawing1.xml.rels"] = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        f'<Relationships xmlns="{PACKAGE_REL_NS}"><Relationship Id="rId1" '
        f'Type="{OFFICE_REL_NS}/image" Target="../media/image1.png"/>'
        "</Relationships>"
    ).encode()
    parts["xl/media/image1.png"] = logo
    content_types = ElementTree.fromstring(parts["[Content_Types].xml"])
    if not any(item.get("Extension") == "png" for item in content_types):
        ElementTree.SubElement(
            content_types,
            f"{{{CONTENT_TYPES_NS}}}Default",
            {"Extension": "png", "ContentType": "image/png"},
        )
    ElementTree.SubElement(
        content_types,
        f"{{{CONTENT_TYPES_NS}}}Override",
        {
            "PartName": "/xl/drawings/drawing1.xml",
            "ContentType": "application/vnd.openxmlformats-officedocument.drawing+xml",
        },
    )
    ElementTree.SubElement(
        content_types,
        f"{{{CONTENT_TYPES_NS}}}Override",
        {
            "PartName": "/docProps/custom.xml",
            "ContentType": (
                "application/vnd.openxmlformats-officedocument." "custom-properties+xml"
            ),
        },
    )
    parts["[Content_Types].xml"] = ElementTree.tostring(
        content_types, encoding="utf-8", xml_declaration=True
    )
    package_rels = ElementTree.fromstring(parts["_rels/.rels"])
    relationship_ids = {item.get("Id") for item in package_rels}
    rid_number = 1
    while f"rId{rid_number}" in relationship_ids:
        rid_number += 1
    ElementTree.SubElement(
        package_rels,
        f"{{{PACKAGE_REL_NS}}}Relationship",
        {
            "Id": f"rId{rid_number}",
            "Type": f"{OFFICE_REL_NS}/custom-properties",
            "Target": "docProps/custom.xml",
        },
    )
    parts["_rels/.rels"] = ElementTree.tostring(
        package_rels, encoding="utf-8", xml_declaration=True
    )
    parts["docProps/custom.xml"] = custom_properties_xml(custom_properties)
    for name in list(parts):
        if name == "xl/calcChain.xml":
            fail("clean renderer unexpectedly emitted calcChain")
    temporary = path.with_name(f".{path.stem}.{uuid.uuid4().hex}.rewrite.xlsx")
    try:
        with ZipFile(temporary, "w", ZIP_DEFLATED) as archive:
            for name in sorted(parts):
                archive.writestr(name, parts[name])
        temporary.replace(path)
    finally:
        temporary.unlink(missing_ok=True)


def load_validator() -> ModuleType:
    spec = importlib.util.spec_from_file_location(
        "dinva_classic_independent_validator", VALIDATOR_PATH
    )
    if spec is None or spec.loader is None:
        fail("independent validator is unavailable")
    module = importlib.util.module_from_spec(spec)
    import sys

    sys.modules[spec.name] = module
    spec.loader.exec_module(module)
    return module


def validate_output_path(output: Path) -> Path:
    path = resolved(output)
    require(path.suffix.casefold() == ".xlsx", "output suffix must be .xlsx")
    require(not is_inside_project(path), "output must be outside Git")
    require(path.parent.is_dir(), "output parent must already exist")
    require(not path.exists(), "output already exists")
    return path


def render(
    *,
    profile_path: Path,
    expected_profile_sha256: str,
    document_path: Path,
    expected_document_sha256: str,
    output: Path,
    allow_test_profile: bool = False,
) -> Path:
    profile_file, profile_raw, profile = load_bound_json(
        profile_path, expected_profile_sha256, "profile"
    )
    document_file, document_raw, document = load_bound_json(
        document_path, expected_document_sha256, "document"
    )
    validate_profile(profile, allow_test_profile=allow_test_profile)
    validate_document(document, allow_test_profile=allow_test_profile)
    output_path = validate_output_path(output)
    candidate = output_path.with_name(
        f".{output_path.stem}.{uuid.uuid4().hex}.candidate.xlsx"
    )
    try:
        render_clean_workbook(
            profile,
            expected_profile_sha256,
            document,
            expected_document_sha256,
            candidate,
        )
        validator = load_validator()
        validator.validate_or_raise(
            candidate,
            profile,
            expected_profile_sha256,
            document,
            expected_document_sha256,
            allow_test_profile=allow_test_profile,
        )
        require(
            profile_file.read_bytes() == profile_raw, "profile changed during render"
        )
        require(
            document_file.read_bytes() == document_raw, "document changed during render"
        )
        require(not output_path.exists(), "output appeared before atomic publish")
        os.link(candidate, output_path)
        validator.validate_or_raise(
            output_path,
            profile,
            expected_profile_sha256,
            document,
            expected_document_sha256,
            allow_test_profile=allow_test_profile,
        )
        require(
            profile_file.read_bytes() == profile_raw, "profile changed after publish"
        )
        require(
            document_file.read_bytes() == document_raw, "document changed after publish"
        )
    except (OSError, RendererError, ValueError) as exc:
        output_path.unlink(missing_ok=True)
        if isinstance(exc, RendererError):
            raise
        raise RendererError(f"clean render failed: {exc}") from exc
    finally:
        candidate.unlink(missing_ok=True)
    return output_path


def parse_args(argv: Sequence[str] | None = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--profile", type=Path, required=True)
    parser.add_argument("--profile-sha256", required=True)
    parser.add_argument("--document", type=Path, required=True)
    parser.add_argument("--document-sha256", required=True)
    parser.add_argument("--output", type=Path, required=True)
    parser.add_argument(
        "--test-only-allow-unapproved-profile",
        action="store_true",
        help=argparse.SUPPRESS,
    )
    return parser.parse_args(argv)


def main(argv: Sequence[str] | None = None) -> int:
    args = parse_args(argv)
    try:
        output = render(
            profile_path=cast(Path, args.profile),
            expected_profile_sha256=cast(str, args.profile_sha256),
            document_path=cast(Path, args.document),
            expected_document_sha256=cast(str, args.document_sha256),
            output=cast(Path, args.output),
            allow_test_profile=bool(args.test_only_allow_unapproved_profile),
        )
    except (OSError, RendererError) as exc:
        print(f"HOLD: {exc}")
        return 1
    print("DINVA_CLASSIC_RENDER=PASS_CANDIDATE_ONLY")
    print(f"OUTPUT={output}")
    print("CLIENT_SEND=CLOSED")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
