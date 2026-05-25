"""Spike check for openpyxl roundtripping the invoice/quotation template."""

from __future__ import annotations

import argparse
import copy
import hashlib
import posixpath
import subprocess
import sys
from collections.abc import Callable, Sequence
from pathlib import Path
from typing import cast
from xml.etree import ElementTree
from zipfile import ZipFile

from openpyxl import load_workbook  # type: ignore[import-untyped]
from openpyxl.worksheet.worksheet import Worksheet  # type: ignore[import-untyped]

SHEET_NAME = "Счёт-КП шаблон"
TARGET_CELL = "B13"
TARGET_TEXT = "Статус документа: Spike openpyxl test"
FORMULA_CELLS = tuple(f"I{row}" for row in range(17, 23))
SIGNATURE_RANGE = "B32:I34"
HEADER_RANGES = ("C2:I6", "B4:B6")
MEDIA_PART = "xl/media/image1.png"
DRAWINGS_PREFIX = "xl/drawings/"
CONTENT_TYPES_PART = "[Content_Types].xml"
WORKBOOK_PART = "xl/workbook.xml"
WORKBOOK_RELS_PART = "xl/_rels/workbook.xml.rels"
WORKSHEET_NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
RELATIONSHIP_NS = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
PACKAGE_RELATIONSHIP_NS = "http://schemas.openxmlformats.org/package/2006/relationships"
CONTENT_TYPES_NS = "http://schemas.openxmlformats.org/package/2006/content-types"
NS = {"main": WORKSHEET_NS, "rel": RELATIONSHIP_NS}
WARNING = (
    "Warning: openpyxl may rewrite xl/drawings XML during save; "
    "verify the output visually in Excel."
)


Check = tuple[str, bool, str]


def parse_args(argv: Sequence[str] | None = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "Create a one-cell openpyxl roundtrip copy of the Excel template "
            "and verify preservation of critical workbook parts."
        )
    )
    parser.add_argument("--template", required=True, type=Path)
    parser.add_argument("--output", required=True, type=Path)
    return parser.parse_args(argv)


def same_path(left: Path, right: Path) -> bool:
    return left.resolve(strict=False) == right.resolve(strict=False)


def file_sha256(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()


def formula_snapshot(sheet: Worksheet) -> dict[str, object]:
    return {cell: sheet[cell].value for cell in FORMULA_CELLS}


def merged_ranges(sheet: Worksheet) -> tuple[str, ...]:
    return tuple(str(cell_range) for cell_range in sheet.merged_cells.ranges)


def range_snapshot(sheet: Worksheet, range_ref: str) -> tuple[tuple[object, ...], ...]:
    return tuple(tuple(cell.value for cell in row) for row in sheet[range_ref])


def zip_parts(path: Path, predicate: Callable[[str], bool]) -> list[str]:
    with ZipFile(path) as archive:
        return sorted(name for name in archive.namelist() if predicate(name))


def zip_bytes(path: Path, part_name: str) -> bytes | None:
    try:
        with ZipFile(path) as archive:
            return archive.read(part_name)
    except KeyError:
        return None


def zip_part_names(path: Path) -> set[str]:
    with ZipFile(path) as archive:
        return set(archive.namelist())


def replace_zip_parts(path: Path, replacements: dict[str, bytes]) -> None:
    temp_path = path.with_name(f"{path.name}.tmp")
    with ZipFile(path, "r") as source, ZipFile(temp_path, "w") as target:
        copied = set()
        for item in source.infolist():
            if item.filename in replacements:
                target.writestr(item, replacements[item.filename])
            else:
                target.writestr(item, source.read(item.filename))
            copied.add(item.filename)

        for name, data in replacements.items():
            if name not in copied:
                target.writestr(name, data)

    temp_path.replace(path)


def normalize_workbook_target(target: str) -> str:
    if target.startswith("/"):
        return target.lstrip("/")
    return posixpath.normpath(posixpath.join("xl", target))


def workbook_sheet_part(path: Path, sheet_name: str) -> str:
    with ZipFile(path) as archive:
        workbook = ElementTree.fromstring(archive.read(WORKBOOK_PART))
        relationships = ElementTree.fromstring(archive.read(WORKBOOK_RELS_PART))

    targets = {
        relationship.get("Id"): normalize_workbook_target(
            relationship.get("Target", "")
        )
        for relationship in relationships.findall(
            f"{{{PACKAGE_RELATIONSHIP_NS}}}Relationship"
        )
    }

    for sheet in workbook.findall("main:sheets/main:sheet", NS):
        if sheet.get("name") != sheet_name:
            continue
        relationship_id = sheet.get(f"{{{RELATIONSHIP_NS}}}id")
        target = targets.get(relationship_id)
        if target is not None:
            return target
    raise KeyError(f"Sheet not found: {sheet_name}")


def worksheet_rels_part(sheet_part: str) -> str:
    directory, filename = posixpath.split(sheet_part)
    return posixpath.join(directory, "_rels", f"{filename}.rels")


def copy_template_drawings(template: Path, output: Path) -> None:
    template_sheet_part = workbook_sheet_part(template, SHEET_NAME)
    output_sheet_part = workbook_sheet_part(output, SHEET_NAME)
    template_rels_part = worksheet_rels_part(template_sheet_part)
    output_rels_part = worksheet_rels_part(output_sheet_part)

    with ZipFile(template) as template_archive, ZipFile(output) as output_archive:
        replacements = {
            name: template_archive.read(name)
            for name in template_archive.namelist()
            if name.startswith(DRAWINGS_PREFIX) or name.startswith("xl/media/")
        }
        if template_rels_part in template_archive.namelist():
            replacements[output_rels_part] = template_archive.read(template_rels_part)

        replacements[CONTENT_TYPES_PART] = merged_content_types(
            template_archive.read(CONTENT_TYPES_PART),
            output_archive.read(CONTENT_TYPES_PART),
        )
        replacements[output_sheet_part] = worksheet_with_template_drawing(
            template_archive.read(template_sheet_part),
            output_archive.read(output_sheet_part),
        )

    replace_zip_parts(output, replacements)


def merged_content_types(template_xml: bytes, output_xml: bytes) -> bytes:
    ElementTree.register_namespace("", CONTENT_TYPES_NS)
    template_root = ElementTree.fromstring(template_xml)
    output_root = ElementTree.fromstring(output_xml)

    existing_defaults = {
        item.get("Extension")
        for item in output_root.findall(f"{{{CONTENT_TYPES_NS}}}Default")
    }
    existing_overrides = {
        item.get("PartName")
        for item in output_root.findall(f"{{{CONTENT_TYPES_NS}}}Override")
    }

    for item in template_root.findall(f"{{{CONTENT_TYPES_NS}}}Default"):
        extension = item.get("Extension")
        if extension in existing_defaults:
            continue
        if extension in {"png", "jpg", "jpeg"}:
            output_root.append(copy.deepcopy(item))

    for item in template_root.findall(f"{{{CONTENT_TYPES_NS}}}Override"):
        part_name = item.get("PartName")
        if part_name in existing_overrides:
            continue
        if part_name and (
            part_name.startswith("/xl/drawings/") or part_name.startswith("/xl/media/")
        ):
            output_root.append(copy.deepcopy(item))

    return cast(
        bytes, ElementTree.tostring(output_root, encoding="utf-8", xml_declaration=True)
    )


def worksheet_with_template_drawing(template_xml: bytes, output_xml: bytes) -> bytes:
    ElementTree.register_namespace("", WORKSHEET_NS)
    ElementTree.register_namespace("r", RELATIONSHIP_NS)
    template_root = ElementTree.fromstring(template_xml)
    output_root = ElementTree.fromstring(output_xml)

    output_drawings = output_root.findall(f"{{{WORKSHEET_NS}}}drawing")
    if output_drawings:
        return output_xml

    template_drawings = template_root.findall(f"{{{WORKSHEET_NS}}}drawing")
    if not template_drawings:
        return output_xml

    insert_at = len(output_root)
    for index, child in enumerate(output_root):
        if child.tag.endswith("}pageMargins"):
            insert_at = index + 1

    for drawing in template_drawings:
        output_root.insert(insert_at, copy.deepcopy(drawing))
        insert_at += 1

    return cast(
        bytes, ElementTree.tostring(output_root, encoding="utf-8", xml_declaration=True)
    )


def pass_fail(ok: bool) -> str:
    return "PASS" if ok else "FAIL"


def add_check(checks: list[Check], name: str, ok: bool, detail: str = "") -> None:
    checks.append((name, ok, detail))


def inspect_output(output: Path) -> tuple[bool, str]:
    inspector = Path(__file__).with_name("inspect_excel_template.py")
    result = subprocess.run(
        [sys.executable, str(inspector), str(output)],
        capture_output=True,
        check=False,
        text=True,
    )
    if result.returncode == 0:
        return True, "inspector exited with code 0"
    detail = result.stderr.strip() or result.stdout.strip()
    return False, detail


def validate_args(template: Path, output: Path) -> Check | None:
    if not template.is_file():
        return ("template exists", False, f"file not found: {template}")
    if same_path(template, output):
        return ("output differs from template", False, "output matches template")
    if output.exists():
        return ("output does not already exist", False, f"file exists: {output}")
    return None


def run_roundtrip(template: Path, output: Path) -> list[Check]:
    checks: list[Check] = []
    template_bytes_before = template.read_bytes()
    template_hash_before = file_sha256(template)

    workbook = load_workbook(template, data_only=False)
    try:
        if SHEET_NAME not in workbook.sheetnames:
            add_check(checks, f"sheet exists: {SHEET_NAME}", False)
            return checks

        sheet = workbook[SHEET_NAME]
        sheet[TARGET_CELL] = TARGET_TEXT
        output.parent.mkdir(parents=True, exist_ok=True)
        workbook.save(output)
    finally:
        workbook.close()

    template_bytes_after = template.read_bytes()
    add_check(
        checks,
        "template bytes unchanged",
        template_bytes_after == template_bytes_before,
        f"sha256 before={template_hash_before} after={file_sha256(template)}",
    )
    add_check(checks, "output created", output.is_file(), str(output))
    copy_template_drawings(template, output)

    template_workbook = load_workbook(template, data_only=False)
    output_workbook = load_workbook(output, data_only=False)
    try:
        template_sheet = template_workbook[SHEET_NAME]
        output_sheet = output_workbook[SHEET_NAME]

        add_check(
            checks,
            "B13 contains spike text",
            output_sheet[TARGET_CELL].value == TARGET_TEXT,
            str(output_sheet[TARGET_CELL].value),
        )
        add_check(
            checks,
            "formulas I17:I22 unchanged",
            formula_snapshot(output_sheet) == formula_snapshot(template_sheet),
        )
        add_check(
            checks,
            "merged ranges unchanged",
            merged_ranges(output_sheet) == merged_ranges(template_sheet),
        )
        add_check(
            checks,
            "signatures B32:I34 unchanged",
            range_snapshot(output_sheet, SIGNATURE_RANGE)
            == range_snapshot(template_sheet, SIGNATURE_RANGE),
        )
        for range_ref in HEADER_RANGES:
            add_check(
                checks,
                f"header/requisites {range_ref} unchanged",
                range_snapshot(output_sheet, range_ref)
                == range_snapshot(template_sheet, range_ref),
            )
    finally:
        template_workbook.close()
        output_workbook.close()

    template_image = zip_bytes(template, MEDIA_PART)
    output_image = zip_bytes(output, MEDIA_PART)
    add_check(
        checks,
        f"{MEDIA_PART} exists",
        output_image is not None,
    )
    add_check(
        checks,
        f"{MEDIA_PART} bytes unchanged",
        output_image is not None and output_image == template_image,
    )

    drawing_parts = zip_parts(output, lambda name: name.startswith(DRAWINGS_PREFIX))
    add_check(
        checks,
        "xl/drawings/* parts exist",
        bool(drawing_parts),
        ", ".join(drawing_parts),
    )

    inspector_ok, inspector_detail = inspect_output(output)
    add_check(
        checks,
        "inspect_excel_template.py opens output",
        inspector_ok,
        inspector_detail,
    )
    return checks


def print_report(output: Path, checks: Sequence[Check]) -> None:
    print("openpyxl template roundtrip spike")
    print(f"Output: {output}")
    print(f"Written: {SHEET_NAME}!{TARGET_CELL} = {TARGET_TEXT}")
    print(WARNING)
    print()
    print("Checks:")
    for name, ok, detail in checks:
        suffix = f" - {detail}" if detail else ""
        print(f"- {pass_fail(ok)} {name}{suffix}")


def main(argv: Sequence[str] | None = None) -> int:
    args = parse_args(argv)
    template: Path = args.template
    output: Path = args.output

    argument_error = validate_args(template, output)
    if argument_error is not None:
        print_report(output, [argument_error])
        return 1

    try:
        checks = run_roundtrip(template, output)
    except Exception as error:
        checks = [("roundtrip execution", False, str(error))]

    print_report(output, checks)
    return 0 if checks and all(ok for _, ok, _ in checks) else 1


if __name__ == "__main__":
    raise SystemExit(main())
