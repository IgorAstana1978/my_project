"""Reconcile commercial CSV data with an existing XLSX without modifying either."""

from __future__ import annotations

import argparse
import csv
import importlib.util
import posixpath
import re
import sys
from collections.abc import Mapping, Sequence
from dataclasses import dataclass, field
from decimal import Decimal, InvalidOperation
from pathlib import Path
from types import ModuleType
from typing import Any, cast
from xml.etree import ElementTree
from zipfile import BadZipFile, ZipFile

from openpyxl import load_workbook  # type: ignore[import-untyped]

PROJECT_ROOT = Path(__file__).resolve().parents[1]
COMMERCIAL_PREFLIGHT_SCRIPT = (
    PROJECT_ROOT / "scripts" / "preflight_quote_commercial_input.py"
)
SHEET_NAME = "Счёт-КП шаблон"
CERTIFIED_CAPACITY = 100
ITEM_START_ROW = 17
ITEM_END_ROW = 116
TOTAL_ROW = 117
MAIN_NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
OFFICE_REL_NS = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
PACKAGE_REL_NS = "http://schemas.openxmlformats.org/package/2006/relationships"
EXPECTED_LOWER_MERGES = {
    "C119:I119",
    "C121:I121",
    "C122:I122",
    "C123:I123",
    "C124:I124",
    "C125:I125",
    "B127:E127",
    "F127:I127",
    "B128:E128",
    "F128:I128",
    "B129:I129",
}
FORMULA_COORDINATES = {
    *(f"I{row}" for row in range(ITEM_START_ROW, ITEM_END_ROW + 1)),
    f"I{TOTAL_ROW}",
}
PASS_NEXT = (
    "commercial draft structure reconciled for internal review only; manual Igor "
    "check and separate Human Approval are required before any client-ready use"
)
FAIL_NEXT = (
    "not safe for commercial draft review; fix CSV/XLSX mismatch and rerun "
    "reconciliation"
)


def load_sibling_module(module_name: str, path: Path) -> ModuleType:
    spec = importlib.util.spec_from_file_location(module_name, path)
    if spec is None or spec.loader is None:
        raise RuntimeError("commercial preflight helper could not be loaded")
    module = importlib.util.module_from_spec(spec)
    sys.modules[spec.name] = module
    spec.loader.exec_module(module)
    return module


commercial_preflight = cast(
    Any,
    load_sibling_module(
        "preflight_quote_commercial_input_for_reconciliation",
        COMMERCIAL_PREFLIGHT_SCRIPT,
    ),
)


@dataclass
class CommercialReconciliationResult:
    commercial_csv: Path
    draft_xlsx: Path
    template_capacity: int
    status: str = "PASS"
    row_count: int = 0
    checks: dict[str, str] = field(
        default_factory=lambda: {
            "commercial preflight": "fail",
            "workbook opens": "fail",
            "worksheet present": "fail",
            "capacity100 profile": "fail",
            "row count reconciliation": "fail",
            "quantity reconciliation": "fail",
            "unit price type": "fail",
            "unit price reconciliation": "fail",
            "line formulas": "fail",
            "total formula": "fail",
            "independent arithmetic": "fail",
            "unused rows": "fail",
            "protected areas": "fail",
            "cached formula values": "fail",
        }
    )
    warnings: list[str] = field(default_factory=list)
    failures: list[str] = field(default_factory=list)
    next_action: str = FAIL_NEXT


@dataclass(frozen=True)
class WorksheetXmlInspection:
    merge_refs: frozenset[str]
    cached_formula_values: Mapping[str, str]


def parse_args(argv: Sequence[str] | None = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "Read-only reconciliation of strict commercial CSV data with an "
            "existing certified capacity100 quote draft."
        )
    )
    parser.add_argument("--commercial-csv", required=True, type=Path)
    parser.add_argument("--draft-xlsx", required=True, type=Path)
    parser.add_argument("--template-capacity", required=True, type=int)
    return parser.parse_args(argv)


def resolved(path: Path) -> Path:
    return path.expanduser().resolve(strict=False)


def is_inside_project(path: Path) -> bool:
    return resolved(path).is_relative_to(PROJECT_ROOT)


def add_failure(result: CommercialReconciliationResult, message: str) -> None:
    result.failures.append(message)


def item_formula(row: int) -> str:
    return (
        f'=IF(OR(E{row}="",H{row}=""),"",' f'IFERROR(E{row}*H{row},"нужно уточнить"))'
    )


def total_formula() -> str:
    return '=IF(COUNT(I17:I116)=0,"нужно уточнить",SUM(I17:I116))'


def load_validated_commercial_rows(path: Path) -> list[dict[str, str]]:
    with path.open("r", encoding="utf-8-sig", newline="") as csv_file:
        reader = csv.DictReader(
            csv_file,
            delimiter=commercial_preflight.CSV_DELIMITER,
            strict=True,
        )
        return [dict(row) for row in reader]


def worksheet_part(archive: ZipFile) -> str:
    workbook_root = ElementTree.fromstring(archive.read("xl/workbook.xml"))
    sheets = workbook_root.find(f"{{{MAIN_NS}}}sheets")
    if sheets is None:
        raise ValueError("workbook sheets element is missing")

    relationship_id: str | None = None
    for sheet in sheets.findall(f"{{{MAIN_NS}}}sheet"):
        if sheet.get("name") == SHEET_NAME:
            relationship_id = sheet.get(f"{{{OFFICE_REL_NS}}}id")
            break
    if relationship_id is None:
        raise ValueError("worksheet relationship is missing")

    relationships_root = ElementTree.fromstring(
        archive.read("xl/_rels/workbook.xml.rels")
    )
    target: str | None = None
    for relationship in relationships_root.findall(f"{{{PACKAGE_REL_NS}}}Relationship"):
        if relationship.get("Id") == relationship_id:
            target = relationship.get("Target")
            break
    if target is None:
        raise ValueError("worksheet target is missing")
    if target.startswith("/"):
        return target.lstrip("/")
    return posixpath.normpath(posixpath.join("xl", target))


def inspect_worksheet_xml(path: Path) -> WorksheetXmlInspection:
    with ZipFile(path, "r") as archive:
        worksheet_root = ElementTree.fromstring(archive.read(worksheet_part(archive)))

    merge_refs = frozenset(
        reference
        for merge_cell in worksheet_root.findall(f".//{{{MAIN_NS}}}mergeCell")
        for reference in [merge_cell.get("ref")]
        if reference is not None
    )
    cached_formula_values: dict[str, str] = {}
    for cell in worksheet_root.findall(f".//{{{MAIN_NS}}}c"):
        coordinate = cell.get("r")
        formula = cell.find(f"{{{MAIN_NS}}}f")
        cached_value = cell.find(f"{{{MAIN_NS}}}v")
        if (
            coordinate is not None
            and formula is not None
            and cached_value is not None
            and cached_value.text not in (None, "")
        ):
            cached_formula_values[coordinate] = cached_value.text
    return WorksheetXmlInspection(
        merge_refs=merge_refs,
        cached_formula_values=cached_formula_values,
    )


def has_cell_value(value: Any) -> bool:
    return value not in (None, "")


def is_integer_cell(value: Any) -> bool:
    return isinstance(value, int) and not isinstance(value, bool)


def cached_integer_matches(value: str, expected: int) -> bool:
    try:
        decimal_value = Decimal(value)
    except InvalidOperation:
        return False
    return decimal_value.is_finite() and decimal_value == expected


def cached_values_match(
    cached_values: Mapping[str, str],
    line_totals: Sequence[int],
    grand_total: int,
) -> bool:
    expected_values = {
        f"I{ITEM_START_ROW + offset}": line_total
        for offset, line_total in enumerate(line_totals)
    }
    expected_values[f"I{TOTAL_ROW}"] = grand_total

    for coordinate, value in cached_values.items():
        if coordinate not in FORMULA_COORDINATES:
            continue
        if coordinate not in expected_values:
            return False
        if not cached_integer_matches(value, expected_values[coordinate]):
            return False
    return True


def finalize_status(result: CommercialReconciliationResult) -> None:
    if result.failures or any(status != "pass" for status in result.checks.values()):
        result.status = "FAIL"
        result.next_action = FAIL_NEXT
    else:
        result.status = "PASS"
        result.next_action = PASS_NEXT


def reconcile(
    commercial_csv: Path,
    draft_xlsx: Path,
    template_capacity: int,
) -> CommercialReconciliationResult:
    csv_path = resolved(commercial_csv)
    xlsx_path = resolved(draft_xlsx)
    result = CommercialReconciliationResult(
        commercial_csv=csv_path,
        draft_xlsx=xlsx_path,
        template_capacity=template_capacity,
    )

    preflight_result = commercial_preflight.preflight(csv_path)
    result.row_count = preflight_result.row_count
    if preflight_result.status != "PASS":
        add_failure(result, "commercial CSV did not pass the existing preflight")
        finalize_status(result)
        return result
    result.checks["commercial preflight"] = "pass"

    try:
        csv_rows = load_validated_commercial_rows(csv_path)
        quantities = [int(row["quantity"]) for row in csv_rows]
        prices = [int(row["unit_price_kzt"]) for row in csv_rows]
        line_totals = [
            quantity * price for quantity, price in zip(quantities, prices, strict=True)
        ]
        grand_total = sum(line_totals)
        result.checks["independent arithmetic"] = "pass"
    except OSError, csv.Error, KeyError, TypeError, ValueError:
        add_failure(result, "validated commercial CSV could not be read safely")
        finalize_status(result)
        return result

    if template_capacity != CERTIFIED_CAPACITY:
        add_failure(result, "only the certified capacity100 profile is supported")
    if xlsx_path.suffix.casefold() != ".xlsx":
        add_failure(result, "draft input must have .xlsx suffix")
    if not xlsx_path.is_file():
        add_failure(result, "draft XLSX does not exist")
    if is_inside_project(xlsx_path):
        add_failure(result, "draft XLSX must be outside the Git project")
    if result.failures:
        finalize_status(result)
        return result

    workbook = None
    try:
        workbook = load_workbook(xlsx_path, read_only=True, data_only=False)
        result.checks["workbook opens"] = "pass"
        if SHEET_NAME not in workbook.sheetnames:
            add_failure(result, "expected worksheet is missing")
            finalize_status(result)
            return result
        result.checks["worksheet present"] = "pass"
        result.checks["capacity100 profile"] = "pass"
        worksheet = workbook[SHEET_NAME]

        item_values = list(
            worksheet.iter_rows(
                min_row=ITEM_START_ROW,
                max_row=ITEM_END_ROW,
                min_col=3,
                max_col=9,
                values_only=True,
            )
        )
        filled_rows = [
            ITEM_START_ROW + offset
            for offset, values in enumerate(item_values)
            if any(has_cell_value(value) for value in values[:6])
        ]
        expected_filled_rows = list(
            range(ITEM_START_ROW, ITEM_START_ROW + result.row_count)
        )
        if filled_rows == expected_filled_rows:
            result.checks["row count reconciliation"] = "pass"
        else:
            add_failure(result, "filled item rows do not match commercial CSV rows")

        used_values = item_values[: result.row_count]
        quantity_ok = all(
            is_integer_cell(values[2]) and values[2] == expected
            for values, expected in zip(used_values, quantities, strict=True)
        )
        if quantity_ok:
            result.checks["quantity reconciliation"] = "pass"
        else:
            add_failure(result, "one or more item quantities do not reconcile")

        price_type_ok = all(is_integer_cell(values[5]) for values in used_values)
        if price_type_ok:
            result.checks["unit price type"] = "pass"
        else:
            add_failure(result, "one or more unit prices are not numeric integers")

        price_reconciliation_ok = all(
            is_integer_cell(values[5]) and values[5] == expected
            for values, expected in zip(used_values, prices, strict=True)
        )
        if price_reconciliation_ok:
            result.checks["unit price reconciliation"] = "pass"
        else:
            add_failure(result, "one or more unit prices do not reconcile")

        line_formulas_ok = all(
            values[6] == item_formula(row)
            for row, values in zip(
                range(ITEM_START_ROW, ITEM_END_ROW + 1),
                item_values,
                strict=True,
            )
        )
        if line_formulas_ok:
            result.checks["line formulas"] = "pass"
        else:
            add_failure(result, "one or more item formulas are missing or unexpected")

        if worksheet[f"I{TOTAL_ROW}"].value == total_formula():
            result.checks["total formula"] = "pass"
        else:
            add_failure(result, "total formula is missing or unexpected")

        unused_rows_ok = all(
            not any(has_cell_value(value) for value in values[:6])
            for values in item_values[result.row_count :]
        )
        if unused_rows_ok:
            result.checks["unused rows"] = "pass"
        else:
            add_failure(result, "unused item rows contain values in C:H")
    except Exception:
        add_failure(result, "workbook could not be opened or inspected")
        finalize_status(result)
        return result
    finally:
        if workbook is not None:
            workbook.close()

    try:
        xml_inspection = inspect_worksheet_xml(xlsx_path)
        lower_merges = frozenset(
            reference
            for reference in xml_inspection.merge_refs
            if any(
                int(character_group) >= TOTAL_ROW
                for character_group in re.findall(
                    r"[A-Z]+([0-9]+)",
                    reference,
                )
            )
        )
        if lower_merges == EXPECTED_LOWER_MERGES:
            result.checks["protected areas"] = "pass"
        else:
            add_failure(result, "protected lower-block merged ranges are unexpected")

        if cached_values_match(
            xml_inspection.cached_formula_values,
            line_totals,
            grand_total,
        ):
            result.checks["cached formula values"] = "pass"
            if xml_inspection.cached_formula_values:
                result.warnings.append(
                    "formula caches were present and matched independent arithmetic"
                )
        else:
            add_failure(result, "stale or unexpected cached formula values detected")
    except BadZipFile, ElementTree.ParseError, KeyError, OSError, ValueError:
        add_failure(result, "workbook package structure could not be inspected safely")

    finalize_status(result)
    return result


def format_list(values: Sequence[str]) -> list[str]:
    return list(values) if values else ["none"]


def format_report(result: CommercialReconciliationResult) -> str:
    lines = [
        "COMMERCIAL_QUOTE_RECONCILIATION_REPORT_START",
        "",
        "Input CSV:",
        str(result.commercial_csv),
        "",
        "Input XLSX:",
        str(result.draft_xlsx),
        "",
        "Status:",
        result.status,
        "",
        "Rows:",
        str(result.row_count),
        "",
        "Checks:",
    ]
    lines.extend(f"{name}: {status}" for name, status in result.checks.items())
    lines.extend(["", "Warnings:"])
    lines.extend(format_list(result.warnings))
    lines.extend(["", "Failures:"])
    lines.extend(format_list(result.failures))
    lines.extend(
        [
            "",
            "Next:",
            result.next_action,
            "",
            "Manual Igor check:",
            "required",
            "",
            "COMMERCIAL_QUOTE_RECONCILIATION_REPORT_END",
        ]
    )
    return "\n".join(lines)


def main(argv: Sequence[str] | None = None) -> int:
    args = parse_args(argv)
    result = reconcile(
        args.commercial_csv,
        args.draft_xlsx,
        args.template_capacity,
    )
    print(format_report(result))
    return 0 if result.status == "PASS" else 1


if __name__ == "__main__":
    raise SystemExit(main())
