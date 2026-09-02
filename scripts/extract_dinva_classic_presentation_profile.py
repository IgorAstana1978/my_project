"""Extract a DRAFT DINVA classic profile from governed evidence roles."""

from __future__ import annotations

import argparse
import base64
import hashlib
import json
import os
import re
import uuid
from collections.abc import Mapping, Sequence
from dataclasses import dataclass
from pathlib import Path
from typing import Any, NoReturn, cast
from xml.etree import ElementTree
from zipfile import BadZipFile, ZipFile

from openpyxl import load_workbook  # type: ignore[import-untyped]
from openpyxl.cell.cell import Cell  # type: ignore[import-untyped]

PROJECT_ROOT = Path(__file__).resolve().parents[1]
PROFILE_SCHEMA_VERSION = "dinva_classic_presentation_profile.v0.1"
PROFILE_ID = "DINVA_CLASSIC_QUOTE_INVOICE_V0_1"
CONTRACT_VERSION = "dinva_classic_presentation_contract.v0.1"
SHA256_RE = re.compile(r"[0-9a-f]{64}\Z")
UTC_TIMESTAMP_RE = re.compile(r"\d{4}-\d{2}-\d{2}T\d{2}:\d{2}:\d{2}Z\Z")
DECISION_SCHEMA_VERSION = "dinva_classic_canonical_logo_human_decision.v0.1"
DECISION_ARTIFACT_TYPE = "IMMUTABLE_HUMAN_DECISION_CAPTURE"
DECISION_ID = "IGOR-DINVA-CLASSIC-CANONICAL-LOGO-20260901-001"
DECISION_STATUS = "IGOR_DINVA_CLASSIC_CANONICAL_LOGO_APPROVED_NOT_APPLIED"
DECISION_AUTHORITY = "IGOR_DIRECT_HUMAN_APPROVAL"
DECISION_SCOPE = "BRAND_LOGO_ONLY"
DECISION_APPLICATION_STATUS = "NOT_APPLIED_TO_PROFILE"
APPROVED_CANONICAL_LOGO_DECISION_SHA256 = (
    "e7c043f19b7eb8606f59dd8e7de06b29ca4305cc1fe2362ecb93767dd589f63b"
)
CANONICAL_LOGO_RAW_SHA256 = (
    "28a6a59ae0a5ca274c206c70545f70b333cac0276a7c4dcbebbf9156f88e0fa8"
)
CANONICAL_LOGO_PIXEL_FINGERPRINT = (
    "81d979c4c158452cca8e3b40d23a4fd321538dfcef238b6f8133beb33a122846"
)
RUNTIME_DIFFERING_LOGO_RAW_SHA256 = (
    "18e0f9446c72f8aa80ea833df07c2e42eb830770a0186decc476c5f948987301"
)
DECISION_SAFETY = {
    "human_decision_recorded": True,
    "canonical_logo_approved": True,
    "profile_approval_authorized": False,
    "profile_generation_authorized": False,
    "profile_publication_authorized": False,
    "runtime_template_modification_authorized": False,
    "quote_generation_authorized": False,
    "invoice_generation_authorized": False,
    "xlsx_generation_authorized": False,
    "pdf_generation_authorized": False,
    "client_send_authorized": False,
    "procurement_authorized": False,
    "reserve_authorized": False,
    "prepayment_authorized": False,
    "payment_authorized": False,
    "production_authorized": False,
    "downstream_authorized": False,
}
DECISION_PUBLICATION_CONTROL = {
    "immutable": True,
    "no_overwrite": True,
    "new_directory_required": True,
    "atomic_publication": True,
    "input_toctou_recheck_required": True,
    "final_strict_json_reread_required": True,
    "rollback_on_failure": True,
    "authorization_token_required": True,
}
SPREADSHEET_NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
DRAWING_NS = "http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing"
NS = {"main": SPREADSHEET_NS, "xdr": DRAWING_NS}
FIXED_HEADERS = {
    "B": "№ п/п",
    "C": "Наименование",
    "D": "Ед.",
    "E": "Кол-во",
    "H": "Цена",
    "I": "Сумма",
}
RUNTIME_COMPANY_CELLS = ("C2", "C3", "B4", "B5", "B6", "G2", "G3", "G4", "G5", "G6")


class ProfileExtractionError(ValueError):
    """Evidence cannot produce one deterministic classic profile."""


@dataclass(frozen=True)
class ReferenceInput:
    path: Path
    expected_sha256: str


@dataclass(frozen=True)
class FamilyEvidence:
    path: Path
    sha256: str
    header_row: int
    first_item_row: int
    logo_sha256: str
    logo_raw: bytes


@dataclass(frozen=True)
class RuntimeEvidence:
    path: Path
    sha256: str
    contract: dict[str, Any]


@dataclass(frozen=True)
class CanonicalLogoDecisionEvidence:
    path: Path
    sha256: str
    payload: Mapping[str, Any]


def fail(message: str) -> NoReturn:
    raise ProfileExtractionError(message)


def require(condition: bool, message: str) -> None:
    if not condition:
        fail(message)


def resolved(path: Path) -> Path:
    return path.expanduser().resolve(strict=False)


def is_inside_project(path: Path) -> bool:
    return resolved(path).is_relative_to(resolved(PROJECT_ROOT))


def sha256_bytes(raw: bytes) -> str:
    return hashlib.sha256(raw).hexdigest()


def canonical_json(value: object) -> bytes:
    return json.dumps(
        value,
        ensure_ascii=False,
        sort_keys=True,
        separators=(",", ":"),
        allow_nan=False,
    ).encode("utf-8")


def reject_duplicate_keys(pairs: list[tuple[str, Any]]) -> dict[str, Any]:
    result: dict[str, Any] = {}
    for key, value in pairs:
        require(key not in result, f"duplicate JSON key: {key}")
        result[key] = value
    return result


def exact_keys(value: Mapping[str, Any], expected: set[str], label: str) -> None:
    require(set(value) == expected, f"{label} fields mismatch")


def mapping(value: object, label: str) -> Mapping[str, Any]:
    require(isinstance(value, Mapping), f"{label} must be an object")
    return cast(Mapping[str, Any], value)


def load_strict_json(raw: bytes, label: str) -> Mapping[str, Any]:
    try:
        text = raw.decode("utf-8")
        value = json.loads(text, object_pairs_hook=reject_duplicate_keys)
    except (UnicodeDecodeError, json.JSONDecodeError) as exc:
        raise ProfileExtractionError(
            f"{label} is not strict UTF-8 JSON: {exc}"
        ) from exc
    return mapping(value, label)


def contract_fingerprint(contract: Mapping[str, Any]) -> str:
    return sha256_bytes(canonical_json(contract))


def normalized_text(value: object) -> str:
    require(isinstance(value, str), "classic evidence text is missing")
    return " ".join(cast(str, value).split())


def normalized_header(value: object) -> str:
    return re.sub(r"-\s+", "-", normalized_text(value))


def color_spec(color: Any) -> dict[str, object] | None:
    if color is None or color.type is None:
        return None
    value: str | int
    if color.type == "rgb":
        value = str(color.rgb)
    elif color.type == "indexed":
        value = int(color.indexed)
    elif color.type == "theme":
        value = int(color.theme)
    else:
        fail(f"unsupported color type: {color.type}")
    return {"type": color.type, "value": value, "tint": float(color.tint or 0)}


def style_spec(cell: Cell) -> dict[str, Any]:
    require(
        cell.font.name == "Times New Roman", f"non-classic font at {cell.coordinate}"
    )
    return {
        "font": {
            "name": cell.font.name,
            "size": float(cell.font.sz or 0),
            "bold": bool(cell.font.b),
            "italic": bool(cell.font.i),
            "underline": cell.font.u,
            "color": color_spec(cell.font.color),
        },
        "fill": {
            "type": cell.fill.fill_type,
            "foreground": color_spec(cell.fill.fgColor),
        },
        "border": {
            "left": cell.border.left.style,
            "right": cell.border.right.style,
            "top": cell.border.top.style,
            "bottom": cell.border.bottom.style,
        },
        "alignment": {
            "horizontal": cell.alignment.horizontal,
            "vertical": cell.alignment.vertical,
            "wrap_text": bool(cell.alignment.wrap_text),
            "shrink_to_fit": bool(cell.alignment.shrink_to_fit),
        },
        "number_format": cell.number_format,
    }


def bound_path(reference: ReferenceInput, label: str) -> tuple[Path, str]:
    path = reference.path.resolve(strict=True)
    require(not is_inside_project(path), f"{label} must be outside Git")
    require(
        SHA256_RE.fullmatch(reference.expected_sha256) is not None,
        f"invalid {label} expected SHA-256",
    )
    actual = sha256_bytes(path.read_bytes())
    require(actual == reference.expected_sha256, f"{label} SHA-256 mismatch")
    return path, actual


def header_row(worksheet: Any) -> int:
    matches = [
        row
        for row in range(1, worksheet.max_row + 1)
        if isinstance(worksheet[f"B{row}"].value, str)
        and normalized_header(worksheet[f"B{row}"].value) == "№ п/п"
    ]
    require(len(matches) == 1, "classic table header row is missing or ambiguous")
    return matches[0]


def first_numbered_item_row(worksheet: Any, start: int) -> int:
    matches = [
        row
        for row in range(start, worksheet.max_row + 1)
        if type(worksheet[f"B{row}"].value) is int and worksheet[f"B{row}"].value > 0
    ]
    require(bool(matches), "classic evidence has no numbered item row")
    return matches[0]


def endpoint(node: ElementTree.Element, name: str) -> dict[str, int]:
    marker = node.find(f"xdr:{name}", NS)
    require(marker is not None, f"logo {name} marker is missing")

    def value(child: str) -> int:
        item = cast(ElementTree.Element, marker).find(f"xdr:{child}", NS)
        require(
            item is not None and item.text is not None, f"logo {name}.{child} missing"
        )
        return int(cast(str, cast(ElementTree.Element, item).text))

    return {
        "column": value("col"),
        "column_offset": value("colOff"),
        "row": value("row"),
        "row_offset": value("rowOff"),
    }


def exact_asset(path: Path) -> tuple[bytes, dict[str, Any]]:
    try:
        with ZipFile(path) as archive:
            media = sorted(
                name for name in archive.namelist() if name.startswith("xl/media/")
            )
            require(
                media == ["xl/media/image1.png"], "logo part is missing or ambiguous"
            )
            logo = archive.read(media[0])
            drawing = ElementTree.fromstring(archive.read("xl/drawings/drawing1.xml"))
    except (BadZipFile, KeyError, ElementTree.ParseError) as exc:
        raise ProfileExtractionError(
            f"asset contract could not be read: {exc}"
        ) from exc
    one = drawing.findall("xdr:oneCellAnchor", NS)
    two = drawing.findall("xdr:twoCellAnchor", NS)
    require(len(one) + len(two) == 1, "logo anchor is missing or ambiguous")
    if one:
        extent = one[0].find("xdr:ext", NS)
        require(extent is not None, "logo extent is missing")
        return logo, {
            "anchor_type": "ONE_CELL",
            "from": endpoint(one[0], "from"),
            "extent": {
                "cx": int(cast(str, cast(ElementTree.Element, extent).get("cx"))),
                "cy": int(cast(str, cast(ElementTree.Element, extent).get("cy"))),
            },
        }
    return logo, {
        "anchor_type": "TWO_CELL",
        "from": endpoint(two[0], "from"),
        "to": endpoint(two[0], "to"),
    }


def load_family_evidence(reference: ReferenceInput) -> FamilyEvidence:
    path, digest = bound_path(reference, "classic family evidence")
    try:
        workbook = load_workbook(
            path, data_only=False, read_only=False, keep_links=True
        )
    except (OSError, ValueError, BadZipFile) as exc:
        raise ProfileExtractionError(
            f"classic reference could not be opened: {exc}"
        ) from exc
    try:
        require(
            workbook.index(workbook.active) == 0, "classic active sheet is not first"
        )
        worksheet = workbook.active
        require(
            normalized_text(worksheet["C2"].value) == "ТОО «ДиН ВА-КЭС»",
            "unsupported/non-classic company block",
        )
        require(worksheet["G9"].value == "ВНИМАНИЕ!", "non-classic warning block")
        table_header = header_row(worksheet)
        for column, expected in FIXED_HEADERS.items():
            require(
                normalized_header(worksheet[f"{column}{table_header}"].value)
                == expected,
                f"non-classic table header: {column}",
            )
        require(
            worksheet.page_setup.orientation == "portrait", "non-classic orientation"
        )
        require(str(worksheet.page_setup.paperSize) == "9", "non-classic paper size")
        require(worksheet.page_setup.scale == 54, "non-classic page scale")
        first_item = first_numbered_item_row(worksheet, table_header + 1)
    finally:
        workbook.close()
    logo, _ = exact_asset(path)
    require(sha256_bytes(path.read_bytes()) == digest, "classic reference changed")
    return FamilyEvidence(
        path, digest, table_header, first_item, sha256_bytes(logo), logo
    )


def find_label_row(worksheet: Any, column: str, label: str) -> int:
    matches = [
        row
        for row in range(1, worksheet.max_row + 1)
        if isinstance(worksheet[f"{column}{row}"].value, str)
        and normalized_text(worksheet[f"{column}{row}"].value).startswith(label)
    ]
    require(len(matches) == 1, f"runtime label is missing or ambiguous: {label}")
    return matches[0]


def runtime_contract(path: Path, digest: str) -> dict[str, Any]:
    try:
        workbook = load_workbook(
            path, data_only=False, read_only=False, keep_links=True
        )
    except (OSError, ValueError, BadZipFile) as exc:
        raise ProfileExtractionError(
            f"runtime template could not be opened: {exc}"
        ) from exc
    try:
        require(len(workbook.worksheets) == 1, "runtime template must have one sheet")
        require(workbook.index(workbook.active) == 0, "runtime active sheet mismatch")
        worksheet = workbook.active
        sheet_name = worksheet.title
        require(
            normalized_text(worksheet["C2"].value) == "ТОО «ДиН ВА-КЭС»",
            "runtime template family mismatch",
        )
        require(worksheet["G9"].value == "ВНИМАНИЕ!", "runtime warning mismatch")
        table_header = header_row(worksheet)
        for column, expected in FIXED_HEADERS.items():
            require(
                normalized_header(worksheet[f"{column}{table_header}"].value)
                == expected,
                f"runtime table header mismatch: {column}",
            )
        first_item = first_numbered_item_row(worksheet, table_header + 1)
        total_row = find_label_row(worksheet, "H", "ИТОГО")
        amount_row = find_label_row(worksheet, "C", "Всего прописью:")
        director_row = find_label_row(worksheet, "B", "Директор")
        review_row = find_label_row(worksheet, "B", "Дата проверки:")
        require(
            first_item == table_header + 2, "runtime section/item geometry mismatch"
        )
        require(total_row > first_item, "runtime item capacity is invalid")
        require(amount_row == total_row + 2, "runtime amount-words geometry mismatch")
        require(director_row + 2 == review_row, "runtime signature geometry mismatch")
        require(
            worksheet.page_setup.orientation == "portrait",
            "runtime orientation mismatch",
        )
        require(str(worksheet.page_setup.paperSize) == "9", "runtime paper mismatch")
        require(worksheet.page_setup.scale == 54, "runtime scale mismatch")
        company = {
            coordinate: normalized_text(worksheet[coordinate].value)
            for coordinate in RUNTIME_COMPANY_CELLS
        }
        headers = {
            column: cast(str, worksheet[f"{column}{table_header}"].value)
            for column in "BCDEFGHI"
        }
        anchors = {
            "company_title": "C2",
            "company_info": "C3",
            "warning": "G9",
            "table_header": f"C{table_header}",
            "position": f"B{first_item}",
            "item_name": f"C{first_item}",
            "unit": f"D{first_item}",
            "quantity": f"E{first_item}",
            "technical_composition": f"F{first_item}",
            "enclosure": f"G{first_item}",
            "money": f"H{first_item}",
            "line_total": f"I{first_item}",
            "total": f"I{total_row}",
            "amount_words": f"C{amount_row}",
            "terms": f"C{amount_row + 2}",
            "signature": f"B{director_row}",
        }
        styles = {
            name: style_spec(worksheet[coordinate])
            for name, coordinate in anchors.items()
        }
        widths = {
            column: float(worksheet.column_dimensions[column].width)
            for column in "BCDEFGHI"
        }
        margins = {
            name: float(getattr(worksheet.page_margins, name))
            for name in ("left", "right", "top", "bottom", "header", "footer")
        }
        merges = sorted(str(value) for value in worksheet.merged_cells.ranges)
        line_formula = worksheet[f"I{first_item}"].value
        grand_formula = worksheet[f"I{total_row}"].value
        require(isinstance(line_formula, str), "runtime line formula missing")
        require(isinstance(grand_formula, str), "runtime total formula missing")
        line_template = cast(str, line_formula).replace(str(first_item), "{row}")
        capacity_end = total_row - 1
        grand_template = (
            cast(str, grand_formula)
            .replace(str(first_item), "{start}")
            .replace(str(capacity_end), "{end}")
        )
        guard_lines = {
            "specification": normalized_text(worksheet[f"C{amount_row + 5}"].value),
            "no_send": normalized_text(worksheet[f"C{amount_row + 6}"].value),
            "review_date": normalized_text(worksheet[f"B{review_row}"].value),
        }
        section_text = normalized_text(worksheet[f"C{first_item - 1}"].value)
        require(":" in section_text, "runtime section prefix missing")
        item_base_height = float(worksheet.row_dimensions[first_item].height)
        print_contract = {
            "paper_size": str(worksheet.page_setup.paperSize),
            "orientation": worksheet.page_setup.orientation,
            "scale": int(worksheet.page_setup.scale),
            "fit_to_page": bool(worksheet.sheet_properties.pageSetUpPr.fitToPage),
            "fit_to_height": int(worksheet.page_setup.fitToHeight or 0),
            "margins": margins,
            "print_area_columns": "B:I",
        }
    finally:
        workbook.close()
    logo, placement = exact_asset(path)
    require(
        placement.get("anchor_type") == "ONE_CELL", "runtime logo anchor type mismatch"
    )
    require(sha256_bytes(path.read_bytes()) == digest, "runtime template changed")
    return {
        "contract_version": CONTRACT_VERSION,
        "workbook": {
            "sheets": [{"name": sheet_name, "role": "PRIMARY_DOCUMENT"}],
            "active_sheet_index": 0,
            "extra_sheets_allowed": False,
        },
        "layout": {
            "table_columns": {
                "position": "B",
                "name": "C",
                "unit": "D",
                "quantity": "E",
                "technical_composition": "F",
                "enclosure": "G",
                "unit_price": "H",
                "line_total": "I",
            },
            "column_widths": widths,
            "company_rows": [2, 3, 4, 5, 6],
            "metadata_rows": [9, 10, 11, 12, 13],
            "table_header_row": table_header,
            "first_item_row": first_item,
            "family_evidence_first_item_rows": [],
            "section_row": first_item - 1,
            "item_capacity": total_row - first_item,
            "total_row": total_row,
            "vat_row": total_row + 1,
            "amount_words_row": amount_row,
            "terms_rows": [amount_row + offset for offset in (2, 3, 4, 5, 6)],
            "signature_rows": [director_row, director_row + 1, review_row],
            "final_row": review_row,
            "item_height_rule": {
                "base": item_base_height,
                "characters_per_increment": 90,
                "increment": item_base_height,
                "maximum": 360.0,
            },
            "merged_cells": {"mode": "EXACT", "ranges": merges},
        },
        "styles": styles,
        "fixed_blocks": {
            "company": company,
            "warning": {"G9": "ВНИМАНИЕ!"},
            "metadata_anchors": {
                "left": ["B9", "B10", "B11", "B12", "B13"],
                "right": ["G9", "G10", "G11", "G12", "G13"],
            },
            "table_headers": headers,
            "section_label_prefix": section_text.split(":", 1)[0] + ":",
            "total_label": "ИТОГО",
            "guard_lines": guard_lines,
        },
        "assets": [
            {
                "asset_id": "DINVA_CLASSIC_LOGO_V0_1",
                "media_type": "image/png",
                "sha256": sha256_bytes(logo),
                "data_base64": base64.b64encode(logo).decode("ascii"),
                "source_reference_sha256s": [digest],
                "placement": placement,
            }
        ],
        "print": print_contract,
        "formulas": {
            "line_total_template": line_template,
            "grand_total_template": grand_template,
            "calc_chain_policy": "OPTIONAL_BUT_MUST_BE_CONSISTENT",
        },
        "package": {
            "forbidden_part_prefixes": [
                "xl/externalLinks/",
                "xl/activeX/",
                "xl/embeddings/",
                "xl/connections",
                "customXml/",
            ],
            "forbidden_part_suffixes": ["vbaProject.bin", ".vml"],
            "external_relationships_allowed": False,
        },
        "variable_elements": [
            "document number/date",
            "payer/customer",
            "object/project/basis",
            "item count/order",
            "item names",
            "detailed technical composition",
            "apparatus",
            "enclosure",
            "approved prices/totals",
            "VAT",
            "amount words",
            "terms",
            "lead time",
            "signatures",
        ],
        "optional_elements": ["object", "basis", "validity"],
    }


def load_runtime_evidence(reference: ReferenceInput) -> RuntimeEvidence:
    path, digest = bound_path(reference, "certified runtime template evidence")
    return RuntimeEvidence(path, digest, runtime_contract(path, digest))


def validate_decision_source_binding_shape(binding: Mapping[str, Any]) -> None:
    exact_keys(
        binding,
        {
            "label",
            "role",
            "path",
            "expected_workbook_sha256",
            "actual_workbook_sha256",
            "media_part_path",
            "expected_logo_raw_sha256",
            "actual_logo_raw_sha256",
            "normalized_pixel_fingerprint",
            "native_dimensions",
            "decoded_mode",
        },
        "canonical-logo decision source binding",
    )
    require(
        isinstance(binding["label"], str) and bool(binding["label"]),
        "canonical-logo decision source label mismatch",
    )
    require(
        binding["role"]
        in {"CLASSIC_FAMILY_EVIDENCE", "CERTIFIED_RUNTIME_TEMPLATE_EVIDENCE"},
        "canonical-logo decision source role mismatch",
    )
    require(
        isinstance(binding["path"], str) and bool(binding["path"]),
        "canonical-logo decision source path mismatch",
    )
    for name in (
        "expected_workbook_sha256",
        "actual_workbook_sha256",
        "expected_logo_raw_sha256",
        "actual_logo_raw_sha256",
        "normalized_pixel_fingerprint",
    ):
        require(
            isinstance(binding[name], str)
            and SHA256_RE.fullmatch(cast(str, binding[name])) is not None,
            f"canonical-logo decision {name} mismatch",
        )
    require(
        binding["expected_workbook_sha256"] == binding["actual_workbook_sha256"],
        "canonical-logo decision workbook SHA binding mismatch",
    )
    require(
        binding["expected_logo_raw_sha256"] == binding["actual_logo_raw_sha256"],
        "canonical-logo decision logo SHA binding mismatch",
    )
    require(
        binding["media_part_path"] == "xl/media/image1.png",
        "canonical-logo decision media part mismatch",
    )
    dimensions = binding["native_dimensions"]
    require(
        isinstance(dimensions, list)
        and len(dimensions) == 2
        and all(type(value) is int and value > 0 for value in dimensions),
        "canonical-logo decision dimensions mismatch",
    )
    require(
        binding["decoded_mode"] in {"RGB", "RGBA"},
        "canonical-logo decision decoded mode mismatch",
    )


def validate_decision_payload(payload: Mapping[str, Any]) -> None:
    exact_keys(
        payload,
        {
            "schema_version",
            "artifact_type",
            "decision_id",
            "status",
            "authority",
            "approval_scope",
            "canonical_logo_application_status",
            "created_at_utc",
            "source_bindings",
            "canonical_logo_decision",
            "runtime_template_policy",
            "safety",
            "publication_control",
        },
        "canonical-logo Human Decision",
    )
    exact_top_level = {
        "schema_version": DECISION_SCHEMA_VERSION,
        "artifact_type": DECISION_ARTIFACT_TYPE,
        "decision_id": DECISION_ID,
        "status": DECISION_STATUS,
        "authority": DECISION_AUTHORITY,
        "approval_scope": DECISION_SCOPE,
        "canonical_logo_application_status": DECISION_APPLICATION_STATUS,
    }
    for name, expected in exact_top_level.items():
        require(
            payload[name] == expected,
            f"canonical-logo Human Decision {name} mismatch",
        )
    require(
        isinstance(payload["created_at_utc"], str)
        and UTC_TIMESTAMP_RE.fullmatch(cast(str, payload["created_at_utc"]))
        is not None,
        "canonical-logo Human Decision created_at_utc mismatch",
    )
    canonical = mapping(payload["canonical_logo_decision"], "canonical logo decision")
    require(
        canonical
        == {
            "approved_variant": "A_FAMILY_INVOICE519",
            "authoritative_brand_source": "CLASSIC_FAMILY_EMBEDDED_LOGO",
            "media_part_path": "xl/media/image1.png",
            "raw_sha256": CANONICAL_LOGO_RAW_SHA256,
            "normalized_pixel_fingerprint": CANONICAL_LOGO_PIXEL_FINGERPRINT,
            "native_dimensions": [200, 68],
            "decoded_mode": "RGB",
        },
        "canonical logo decision mismatch",
    )
    runtime_policy = mapping(
        payload["runtime_template_policy"], "runtime template policy"
    )
    require(
        runtime_policy
        == {
            "template_id": "capacity100_tuned_v4",
            "permitted_role": "RUNTIME_GEOMETRY_STYLE_LAYOUT_SOURCE_ONLY",
            "authoritative_brand_logo_source": False,
            "differing_logo_raw_sha256": RUNTIME_DIFFERING_LOGO_RAW_SHA256,
        },
        "canonical-logo Human Decision runtime policy mismatch",
    )
    require(
        mapping(payload["safety"], "canonical-logo decision safety") == DECISION_SAFETY,
        "canonical-logo Human Decision safety mismatch",
    )
    require(
        mapping(payload["publication_control"], "decision publication control")
        == DECISION_PUBLICATION_CONTROL,
        "canonical-logo Human Decision publication control mismatch",
    )
    bindings = payload["source_bindings"]
    require(
        isinstance(bindings, list) and len(bindings) == 4,
        "canonical-logo Human Decision source binding count mismatch",
    )
    normalized_bindings = [
        mapping(item, "canonical-logo decision source binding")
        for item in cast(list[object], bindings)
    ]
    for binding in normalized_bindings:
        validate_decision_source_binding_shape(binding)
    require(
        len({cast(str, item["label"]) for item in normalized_bindings}) == 4,
        "canonical-logo Human Decision source labels are not unique",
    )
    roles = [item["role"] for item in normalized_bindings]
    require(
        roles.count("CLASSIC_FAMILY_EVIDENCE") == 3
        and roles.count("CERTIFIED_RUNTIME_TEMPLATE_EVIDENCE") == 1,
        "canonical-logo Human Decision evidence role set mismatch",
    )


def load_canonical_logo_decision(
    reference: ReferenceInput,
) -> CanonicalLogoDecisionEvidence:
    path = reference.path.resolve(strict=True)
    require(
        not is_inside_project(path), "canonical-logo Human Decision must be outside Git"
    )
    require(
        SHA256_RE.fullmatch(reference.expected_sha256) is not None,
        "invalid canonical-logo Human Decision expected SHA-256",
    )
    require(
        reference.expected_sha256 == APPROVED_CANONICAL_LOGO_DECISION_SHA256,
        "canonical-logo Human Decision SHA-256 is not the approved artifact SHA-256",
    )
    raw = path.read_bytes()
    digest = sha256_bytes(raw)
    require(
        digest == reference.expected_sha256,
        "canonical-logo Human Decision SHA-256 mismatch",
    )
    payload = load_strict_json(raw, "canonical-logo Human Decision")
    validate_decision_payload(payload)
    require(path.read_bytes() == raw, "canonical-logo Human Decision changed")
    return CanonicalLogoDecisionEvidence(path, digest, payload)


def validate_decision_source_bindings(
    decision: CanonicalLogoDecisionEvidence,
    family: Sequence[FamilyEvidence],
    runtime: Sequence[RuntimeEvidence],
) -> None:
    bindings = cast(list[Mapping[str, Any]], decision.payload["source_bindings"])

    def identity(binding: Mapping[str, Any]) -> tuple[str, str, str, str, str, str]:
        return (
            cast(str, binding["role"]),
            cast(str, binding["path"]),
            cast(str, binding["expected_workbook_sha256"]),
            cast(str, binding["actual_workbook_sha256"]),
            cast(str, binding["expected_logo_raw_sha256"]),
            cast(str, binding["actual_logo_raw_sha256"]),
        )

    actual = sorted(identity(binding) for binding in bindings)
    expected = sorted(
        [
            (
                "CLASSIC_FAMILY_EVIDENCE",
                str(item.path),
                item.sha256,
                item.sha256,
                item.logo_sha256,
                item.logo_sha256,
            )
            for item in family
        ]
        + [
            (
                "CERTIFIED_RUNTIME_TEMPLATE_EVIDENCE",
                str(item.path),
                item.sha256,
                item.sha256,
                cast(str, item.contract["assets"][0]["sha256"]),
                cast(str, item.contract["assets"][0]["sha256"]),
            )
            for item in runtime
        ]
    )
    require(actual == expected, "canonical-logo Human Decision source binding mismatch")
    canonical = mapping(
        decision.payload["canonical_logo_decision"], "canonical logo decision"
    )
    runtime_policy = mapping(
        decision.payload["runtime_template_policy"], "runtime template policy"
    )
    for binding in bindings:
        if binding["role"] == "CLASSIC_FAMILY_EVIDENCE":
            require(
                binding["actual_logo_raw_sha256"] == canonical["raw_sha256"]
                and binding["normalized_pixel_fingerprint"]
                == canonical["normalized_pixel_fingerprint"]
                and binding["native_dimensions"] == canonical["native_dimensions"]
                and binding["decoded_mode"] == canonical["decoded_mode"],
                "canonical-logo Human Decision family logo binding mismatch",
            )
        else:
            require(
                binding["actual_logo_raw_sha256"]
                == runtime_policy["differing_logo_raw_sha256"],
                "canonical-logo Human Decision runtime logo binding mismatch",
            )


def contract_with_family_logo(
    contract: Mapping[str, Any], logo_raw: bytes, family_sha256s: Sequence[str]
) -> dict[str, Any]:
    result = cast(dict[str, Any], json.loads(canonical_json(contract).decode("utf-8")))
    asset = cast(dict[str, Any], result["assets"][0])
    asset["sha256"] = sha256_bytes(logo_raw)
    asset["data_base64"] = base64.b64encode(logo_raw).decode("ascii")
    asset["source_reference_sha256s"] = sorted(family_sha256s)
    return result


def extract_profile(
    references: Sequence[ReferenceInput],
    runtime_templates: Sequence[ReferenceInput],
    canonical_logo_human_decision: ReferenceInput | None = None,
) -> dict[str, Any]:
    require(len(references) >= 2, "at least two classic family references are required")
    require(bool(runtime_templates), "certified runtime template evidence is required")
    family = [load_family_evidence(reference) for reference in references]
    require(
        len({item.header_row for item in family}) == 1,
        "classic family header geometry lacks consensus",
    )
    require(
        len({item.logo_raw for item in family}) == 1,
        "classic family logo evidence lacks consensus",
    )
    family_logo_sha256 = family[0].logo_sha256
    family_logo_raw = family[0].logo_raw
    runtime = [load_runtime_evidence(reference) for reference in runtime_templates]
    logo_mismatch = any(
        item.contract["assets"][0]["sha256"] != family_logo_sha256 for item in runtime
    )
    decision = (
        load_canonical_logo_decision(canonical_logo_human_decision)
        if canonical_logo_human_decision is not None
        else None
    )
    if logo_mismatch:
        require(
            decision is not None,
            "runtime logo differs from classic family logo evidence; "
            "canonical-logo Human Decision is required",
        )
    if decision is not None:
        validate_decision_source_bindings(decision, family, runtime)
        require(
            family_logo_sha256 == CANONICAL_LOGO_RAW_SHA256,
            "classic family logo does not match approved canonical logo",
        )
        candidate_contracts = [
            contract_with_family_logo(
                item.contract,
                family_logo_raw,
                [evidence.sha256 for evidence in family],
            )
            for item in runtime
        ]
    else:
        candidate_contracts = [item.contract for item in runtime]
    contracts = {canonical_json(item) for item in candidate_contracts}
    require(
        len(contracts) == 1,
        "certified runtime template geometry/style lacks consensus",
    )
    contract = json.loads(next(iter(contracts)).decode("utf-8"))
    contract["layout"]["family_evidence_first_item_rows"] = sorted(
        item.first_item_row for item in family
    )
    fingerprint = contract_fingerprint(contract)
    provenance = [
        {
            "path": str(item.path),
            "expected_sha256": item.sha256,
            "actual_sha256": item.sha256,
            "role": "CLASSIC_FAMILY_EVIDENCE",
        }
        for item in family
    ] + [
        {
            "path": str(item.path),
            "expected_sha256": item.sha256,
            "actual_sha256": item.sha256,
            "role": "CERTIFIED_RUNTIME_TEMPLATE_EVIDENCE",
        }
        for item in runtime
    ]
    if decision is not None:
        require(
            sha256_bytes(decision.path.read_bytes()) == decision.sha256,
            "canonical-logo Human Decision changed",
        )
        provenance.append(
            {
                "path": str(decision.path),
                "expected_sha256": decision.sha256,
                "actual_sha256": decision.sha256,
                "role": "CANONICAL_LOGO_HUMAN_DECISION",
            }
        )
    provenance.sort(
        key=lambda item: (item["role"], item["actual_sha256"], item["path"])
    )
    return {
        "schema_version": PROFILE_SCHEMA_VERSION,
        "profile_id": PROFILE_ID,
        "document_family": PROFILE_ID,
        "artifact_status": "DRAFT_PROFILE_CANDIDATE",
        "reference_provenance": provenance,
        "presentation_contract": contract,
        "presentation_contract_fingerprint": fingerprint,
        "approval_provenance": {
            "status": "DRAFT_UNAPPROVED",
            "authority": None,
            "approval_id": None,
            "approved_at": None,
            "approved_contract_fingerprint": None,
        },
    }


def validate_output_path(path: Path) -> Path:
    output = resolved(path)
    require(output.suffix.casefold() == ".json", "profile output suffix must be .json")
    require(not is_inside_project(output), "profile output must be outside Git")
    require(output.parent.is_dir(), "profile output parent must already exist")
    require(not output.exists(), "profile output already exists")
    return output


def publish_profile(profile: Mapping[str, Any], output: Path) -> Path:
    output_path = validate_output_path(output)
    raw = canonical_json(profile) + b"\n"
    candidate = output_path.with_name(
        f".{output_path.stem}.{uuid.uuid4().hex}.candidate.json"
    )
    try:
        candidate.write_bytes(raw)
        require(candidate.read_bytes() == raw, "profile candidate reread mismatch")
        require(not output_path.exists(), "profile output appeared before publish")
        os.link(candidate, output_path)
        require(output_path.read_bytes() == raw, "profile final reread mismatch")
    except OSError as exc:
        raise ProfileExtractionError(
            f"profile no-overwrite publish failed: {exc}"
        ) from exc
    finally:
        candidate.unlink(missing_ok=True)
    return output_path


def parse_args(argv: Sequence[str] | None = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--reference", action="append", type=Path, required=True)
    parser.add_argument("--reference-sha256", action="append", required=True)
    parser.add_argument("--runtime-template", action="append", type=Path, required=True)
    parser.add_argument("--runtime-template-sha256", action="append", required=True)
    parser.add_argument("--canonical-logo-human-decision", type=Path)
    parser.add_argument("--canonical-logo-human-decision-sha256")
    parser.add_argument("--output-profile", type=Path, required=True)
    return parser.parse_args(argv)


def paired(paths: list[Path], hashes: list[str], label: str) -> list[ReferenceInput]:
    require(len(paths) == len(hashes), f"{label} path/SHA count mismatch")
    return [
        ReferenceInput(path, digest) for path, digest in zip(paths, hashes, strict=True)
    ]


def optional_reference(
    path: Path | None, digest: str | None, label: str
) -> ReferenceInput | None:
    require(
        (path is None) == (digest is None),
        f"{label} path and SHA-256 must be supplied together",
    )
    if path is None:
        return None
    return ReferenceInput(path, cast(str, digest))


def main(argv: Sequence[str] | None = None) -> int:
    args = parse_args(argv)
    try:
        profile = extract_profile(
            paired(
                cast(list[Path], args.reference),
                cast(list[str], args.reference_sha256),
                "reference",
            ),
            paired(
                cast(list[Path], args.runtime_template),
                cast(list[str], args.runtime_template_sha256),
                "runtime template",
            ),
            optional_reference(
                cast(Path | None, args.canonical_logo_human_decision),
                cast(str | None, args.canonical_logo_human_decision_sha256),
                "canonical-logo Human Decision",
            ),
        )
        output = publish_profile(profile, cast(Path, args.output_profile))
    except (OSError, ProfileExtractionError) as exc:
        print(f"HOLD: {exc}")
        return 1
    print("DINVA_CLASSIC_PROFILE_CANDIDATE=DRAFT_UNAPPROVED")
    print(
        "PRESENTATION_CONTRACT_FINGERPRINT="
        f"{profile['presentation_contract_fingerprint']}"
    )
    print(f"OUTPUT={output}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
