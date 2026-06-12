"""Minimal isolated extended invoice-quote writer.

This module has a minimal isolated CLI and is not connected to the v0.2 separate
layer. It writes only to an explicitly provided test/extended layout.
"""

from __future__ import annotations

import argparse
import importlib.util
import json
import sys
import uuid
from collections.abc import Callable, Mapping, Sequence
from dataclasses import dataclass
from pathlib import Path
from types import ModuleType
from typing import Any, NoReturn, cast

from openpyxl import load_workbook
from openpyxl.utils.cell import range_boundaries
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

PROJECT_ROOT = Path(__file__).resolve().parents[1]
DRAWING_MEDIA_SNAPSHOT_SCRIPT = PROJECT_ROOT / "scripts" / "drawing_media_snapshot.py"
OOXML_CELL_PATCHER_SCRIPT = PROJECT_ROOT / "scripts" / "ooxml_cell_patcher.py"
SHEET_NAME = "Счёт-КП шаблон"
NEEDS_CLARIFICATION = "нужно уточнить"
ITEM_ROW_HEIGHTS = (24, 42, 60, 78, 96, 114, 132, 150, 168)
ROW_HEIGHT_TEXT_WIDTHS = {
    "name": 24,
    "instruments_and_devices": 30,
    "cabinet_type_dimensions_material": 22,
}


class ExtendedFillError(Exception):
    """Expected validation, preflight, or generation error."""


@dataclass(frozen=True)
class ExtendedLayout:
    item_start_row: int
    item_end_row: int
    capacity: int
    total_row: int
    signature_range: str
    header_ranges: tuple[str, ...]
    formula_cells: tuple[str, ...]


@dataclass(frozen=True)
class WorkbookSnapshot:
    formulas: dict[str, Any]
    signature_values: dict[str, Any]
    header_values: dict[str, Any]
    merged_ranges: tuple[str, ...]


def fail(message: str) -> NoReturn:
    raise ExtendedFillError(message)


def load_sibling_module(module_name: str, path: Path) -> ModuleType:
    spec = importlib.util.spec_from_file_location(module_name, path)
    if spec is None or spec.loader is None:
        fail(f"could not load helper module: {path}")
    module = importlib.util.module_from_spec(spec)
    sys.modules[spec.name] = module
    spec.loader.exec_module(module)
    return module


drawing_media_snapshot = cast(
    Any,
    load_sibling_module(
        "drawing_media_snapshot_for_extended_writer",
        DRAWING_MEDIA_SNAPSHOT_SCRIPT,
    ),
)
DrawingMediaSnapshotError = drawing_media_snapshot.DrawingMediaSnapshotError
build_drawing_media_snapshot = cast(
    Callable[[Path], Any],
    drawing_media_snapshot.build_drawing_media_snapshot,
)
compare_drawing_media_snapshots = cast(
    Callable[[Any, Any], None],
    drawing_media_snapshot.compare_drawing_media_snapshots,
)
ooxml_cell_patcher = cast(
    Any,
    load_sibling_module(
        "ooxml_cell_patcher_for_extended_writer",
        OOXML_CELL_PATCHER_SCRIPT,
    ),
)
OoxmlCellPatcherError = ooxml_cell_patcher.OoxmlCellPatcherError
patch_existing_cells = cast(
    Callable[..., Path],
    ooxml_cell_patcher.patch_existing_cells,
)


def parse_args(argv: Sequence[str] | None = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Generate an isolated extended invoice-quote draft workbook."
    )
    parser.add_argument(
        "--payload-json",
        required=True,
        type=Path,
        help="Path to internal flat payload JSON",
    )
    parser.add_argument(
        "--layout-json",
        required=True,
        type=Path,
        help="Path to internal extended layout JSON",
    )
    parser.add_argument("--template", required=True, type=Path, help="Path to .xlsx")
    parser.add_argument("--output", required=True, type=Path, help="Output .xlsx path")
    return parser.parse_args(argv)


def resolved(path: Path) -> Path:
    return path.expanduser().resolve(strict=False)


def is_inside_project(path: Path) -> bool:
    return resolved(path).is_relative_to(PROJECT_ROOT)


def validate_layout(layout: ExtendedLayout) -> None:
    if layout.capacity < 1:
        fail(f"layout capacity must be positive: {layout.capacity}")
    if layout.item_end_row < layout.item_start_row:
        fail("layout item_end_row must be greater than or equal to item_start_row")
    actual_capacity = layout.item_end_row - layout.item_start_row + 1
    if layout.capacity != actual_capacity:
        fail(
            "layout capacity must equal item_end_row - item_start_row + 1: "
            f"{layout.capacity} != {actual_capacity}"
        )
    if layout.total_row in range(layout.item_start_row, layout.item_end_row + 1):
        fail("layout total_row must be outside item rows")
    if not layout.signature_range:
        fail("layout signature_range must be set")
    if not layout.header_ranges:
        fail("layout header_ranges must be set")
    if not layout.formula_cells:
        fail("layout formula_cells must be set")


def load_json_object(path: Path, label: str) -> Mapping[str, Any]:
    json_path = resolved(path)
    if not json_path.is_file():
        fail(f"{label} JSON does not exist: {json_path}")
    try:
        raw_data = json.loads(json_path.read_text(encoding="utf-8"))
    except json.JSONDecodeError as error:
        fail(f"{label} JSON is invalid: {error.msg}")
    if not isinstance(raw_data, Mapping):
        fail(f"{label} JSON must be an object")
    return cast(Mapping[str, Any], raw_data)


def required_int(data: Mapping[str, Any], key: str) -> int:
    value = data.get(key)
    if not isinstance(value, int) or isinstance(value, bool):
        fail(f"layout.{key} must be an integer")
    return value


def required_str(data: Mapping[str, Any], key: str) -> str:
    value = data.get(key)
    if not isinstance(value, str) or value == "":
        fail(f"layout.{key} must be a non-empty string")
    return value


def required_str_tuple(data: Mapping[str, Any], key: str) -> tuple[str, ...]:
    value = data.get(key)
    if not isinstance(value, Sequence) or isinstance(value, (str, bytes)):
        fail(f"layout.{key} must be a list of strings")
    items: list[str] = []
    for index, item in enumerate(value, start=1):
        if not isinstance(item, str) or item == "":
            fail(f"layout.{key}[{index}] must be a non-empty string")
        items.append(item)
    return tuple(items)


def layout_from_json(data: Mapping[str, Any]) -> ExtendedLayout:
    return ExtendedLayout(
        item_start_row=required_int(data, "item_start_row"),
        item_end_row=required_int(data, "item_end_row"),
        capacity=required_int(data, "capacity"),
        total_row=required_int(data, "total_row"),
        signature_range=required_str(data, "signature_range"),
        header_ranges=required_str_tuple(data, "header_ranges"),
        formula_cells=required_str_tuple(data, "formula_cells"),
    )


def require_items(payload: Mapping[str, Any]) -> Sequence[Mapping[str, Any]]:
    items = payload.get("items")
    if not isinstance(items, Sequence) or isinstance(items, (str, bytes)):
        fail("payload items must be a list")

    checked_items: list[Mapping[str, Any]] = []
    for index, item in enumerate(items, start=1):
        if not isinstance(item, Mapping) or not item:
            fail(f"payload items[{index}] must be a non-empty object")
        for key in ("name", "unit", "quantity"):
            if key not in item or item[key] in (None, ""):
                fail(f"payload items[{index}].{key} is required")
        checked_items.append(cast(Mapping[str, Any], item))
    if not checked_items:
        fail("payload items must not be empty")
    return tuple(checked_items)


def validate_capacity(
    items: Sequence[Mapping[str, Any]], layout: ExtendedLayout
) -> None:
    if len(items) > layout.capacity:
        fail(f"items count {len(items)} exceeds layout capacity {layout.capacity}")


def validate_template_and_output(template: Path, output: Path) -> tuple[Path, Path]:
    template_path = resolved(template)
    output_path = resolved(output)

    if not template_path.is_file():
        fail(f"template does not exist: {template_path}")
    if output_path.exists():
        fail(f"output already exists: {output_path}")
    if not output_path.parent.is_dir():
        fail(f"output parent directory does not exist: {output_path.parent}")
    if is_inside_project(output_path):
        fail(f"output is inside the Git project: {output_path}")
    if template_path == output_path:
        fail("output matches template")
    return template_path, output_path


def value_map(worksheet: Worksheet, ranges: Sequence[str]) -> dict[str, Any]:
    values: dict[str, Any] = {}
    for cell_range in ranges:
        min_column, min_row, max_column, max_row = range_boundaries(cell_range)
        for row in range(min_row, max_row + 1):
            for column in range(min_column, max_column + 1):
                cell = worksheet.cell(row=row, column=column)
                values[cell.coordinate] = cell.value
    return values


def snapshot_workbook(worksheet: Worksheet, layout: ExtendedLayout) -> WorkbookSnapshot:
    return WorkbookSnapshot(
        formulas={cell: worksheet[cell].value for cell in layout.formula_cells},
        signature_values=value_map(worksheet, (layout.signature_range,)),
        header_values=value_map(worksheet, layout.header_ranges),
        merged_ranges=tuple(str(item) for item in worksheet.merged_cells.ranges),
    )


def load_template_workbook(template: Path) -> Workbook:
    workbook = load_workbook(template, data_only=False)
    if SHEET_NAME not in workbook.sheetnames:
        fail(f'worksheet "{SHEET_NAME}" not found')
    return cast(Workbook, workbook)


def optional_text(value: Any) -> str:
    if value is None or value == "":
        return NEEDS_CLARIFICATION
    return str(value)


def visual_line_count(value: Any, width: int) -> int:
    text = optional_text(value)
    return sum(max(1, (len(line) + width - 1) // width) for line in text.split("\n"))


def height_for_visual_lines(lines: int) -> int:
    if lines <= 1:
        return 24
    if lines == 2:
        return 42
    if lines == 3:
        return 60
    if lines == 4:
        return 78
    if lines == 5:
        return 96
    if lines == 6:
        return 114
    if lines == 7:
        return 132
    if lines == 8:
        return 150
    return 168


def estimate_item_row_height(item: Mapping[str, Any]) -> int:
    visual_lines = max(
        visual_line_count(item.get(field), width)
        for field, width in ROW_HEIGHT_TEXT_WIDTHS.items()
    )
    return height_for_visual_lines(visual_lines)


def build_cell_updates(
    items: Sequence[Mapping[str, Any]],
    layout: ExtendedLayout,
) -> dict[str, str | int | None]:
    updates: dict[str, str | int | None] = {}
    for offset, item in enumerate(items):
        row = layout.item_start_row + offset
        updates[f"C{row}"] = str(item["name"])
        updates[f"D{row}"] = str(item["unit"])
        updates[f"E{row}"] = cast(int, item["quantity"])
        updates[f"F{row}"] = optional_text(item.get("instruments_and_devices"))
        updates[f"G{row}"] = optional_text(
            item.get("cabinet_type_dimensions_material"),
        )
        updates[f"H{row}"] = NEEDS_CLARIFICATION

    for row in range(layout.item_start_row + len(items), layout.item_end_row + 1):
        updates[f"C{row}"] = None
        updates[f"D{row}"] = None
        updates[f"E{row}"] = None
        updates[f"F{row}"] = None
        updates[f"G{row}"] = None
        updates[f"H{row}"] = None

    return updates


def build_row_hidden_updates(
    items: Sequence[Mapping[str, Any]],
    layout: ExtendedLayout,
) -> dict[int, bool]:
    used_rows = {layout.item_start_row + offset for offset in range(len(items))}
    return {
        row: row not in used_rows
        for row in range(layout.item_start_row, layout.item_end_row + 1)
    }


def build_row_height_updates(
    items: Sequence[Mapping[str, Any]],
    layout: ExtendedLayout,
) -> dict[int, int]:
    updates: dict[int, int] = {}
    for offset, item in enumerate(items):
        row = layout.item_start_row + offset
        height = estimate_item_row_height(item)
        if height not in ITEM_ROW_HEIGHTS:
            fail(f"unsupported item row height: {height}")
        updates[row] = height
    for row in range(layout.item_start_row + len(items), layout.item_end_row + 1):
        updates[row] = 24
    return updates


def verify_output(
    output: Path, layout: ExtendedLayout, before: WorkbookSnapshot
) -> None:
    workbook = load_template_workbook(output)
    worksheet = workbook[SHEET_NAME]
    after = snapshot_workbook(worksheet, layout)

    if after.formulas != before.formulas:
        fail("formula cells changed")
    if after.signature_values != before.signature_values:
        fail("signature range changed")
    if after.header_values != before.header_values:
        fail("header ranges changed")
    if after.merged_ranges != before.merged_ranges:
        fail("merged ranges changed")


def verified_drawing_media_snapshot(path: Path) -> Any:
    try:
        return build_drawing_media_snapshot(path)
    except DrawingMediaSnapshotError as error:
        fail(f"drawing/media verification failed: {error}")


def verify_drawing_media_output(before: Any, output: Path) -> None:
    after = verified_drawing_media_snapshot(output)
    try:
        compare_drawing_media_snapshots(before, after)
    except DrawingMediaSnapshotError as error:
        fail(f"drawing/media verification failed: {error}")


def generate_extended_workbook(
    template: Path,
    output: Path,
    payload: Mapping[str, Any],
    layout: ExtendedLayout,
) -> Path:
    validate_layout(layout)
    items = require_items(payload)
    validate_capacity(items, layout)
    template_path, output_path = validate_template_and_output(template, output)
    before_drawing_media = verified_drawing_media_snapshot(template_path)

    workbook = load_template_workbook(template_path)
    worksheet = workbook[SHEET_NAME]
    before = snapshot_workbook(worksheet, layout)
    cell_updates = build_cell_updates(items, layout)
    row_hidden_updates = build_row_hidden_updates(items, layout)
    row_height_updates = build_row_height_updates(items, layout)

    temporary_output = output_path.with_name(
        f".{output_path.stem}.{uuid.uuid4().hex}.tmp.xlsx"
    )
    try:
        try:
            patch_existing_cells(
                template=template_path,
                output=temporary_output,
                sheet_name=SHEET_NAME,
                updates=cell_updates,
                row_hidden_updates=row_hidden_updates,
                row_height_updates=row_height_updates,
            )
        except OoxmlCellPatcherError as error:
            fail(f"OOXML patching failed: {error}")
        verify_output(temporary_output, layout, before)
        verify_drawing_media_output(before_drawing_media, temporary_output)
        if output_path.exists():
            fail(f"output already exists: {output_path}")
        temporary_output.replace(output_path)
    except Exception:
        if temporary_output.exists():
            temporary_output.unlink()
        raise

    return output_path


def main(argv: Sequence[str] | None = None) -> int:
    args = parse_args(argv)
    try:
        payload = load_json_object(args.payload_json, "payload")
        layout_data = load_json_object(args.layout_json, "layout")
        layout = layout_from_json(layout_data)
        output = generate_extended_workbook(
            template=args.template,
            output=args.output,
            payload=payload,
            layout=layout,
        )
    except ExtendedFillError as error:
        print(f"ERROR: {error}", file=sys.stderr)
        return 1

    print(f"CREATED: {output}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
