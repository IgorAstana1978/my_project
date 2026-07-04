"""Export one approved commercial item to a client-style XLSX candidate."""

from __future__ import annotations

import argparse
import csv
import hashlib
import importlib.util
import json
import sys
import uuid
import zipfile
from collections.abc import Mapping, Sequence
from dataclasses import dataclass, field
from pathlib import Path
from types import ModuleType
from typing import Any, NoReturn, cast

from openpyxl import load_workbook

PROJECT_ROOT = Path(__file__).resolve().parents[1]
APPROVAL_PREFLIGHT_SCRIPT = (
    PROJECT_ROOT / "scripts" / "preflight_client_style_invoice_export.py"
)
TEMPLATE_PREFLIGHT_SCRIPT = (
    PROJECT_ROOT / "scripts" / "preflight_client_style_invoice_template_contract.py"
)
COMMERCIAL_PREFLIGHT_SCRIPT = (
    PROJECT_ROOT / "scripts" / "preflight_quote_commercial_input.py"
)
COMMERCIAL_WRITER_SCRIPT = (
    PROJECT_ROOT / "scripts" / "run_invoice_quote_commercial_from_csv.py"
)
OOXML_CELL_PATCHER_SCRIPT = PROJECT_ROOT / "scripts" / "ooxml_cell_patcher.py"

REPORT_START = "CLIENT_STYLE_INVOICE_EXPORT_REPORT_START"
REPORT_END = "CLIENT_STYLE_INVOICE_EXPORT_REPORT_END"
MODE = "one-item client-style invoice export candidate"
COMMERCIAL_STATUS = "client-style candidate only; PASS is not sending approval"
HUMAN_APPROVAL = "required before sending to client"

APPROVED_FIELDS = (
    "invoice_number",
    "invoice_date",
    "payer_name",
    "object_name",
    "vat_text_approved",
    "payment_terms_approved",
    "delivery_terms_approved",
    "validity_terms_approved",
    "return_terms_approved",
    "signer_name",
    "signer_title",
    "approval_note",
)
APPROVAL_HASH_FIELDS = {
    "commercial CSV": "commercial_csv_sha256",
    "internal draft XLSX": "internal_draft_xlsx_sha256",
    "template XLSX": "template_sha256",
}
TERM_PLACEHOLDERS = {
    "payment": ("[условия оплаты]", "payment_terms_approved"),
    "return": ("[условия возврата]", "return_terms_approved"),
    "validity": (
        "[условия изменения спецификации / срок действия]",
        "validity_terms_approved",
    ),
    "contract": ("[условия договора]", "approval_note"),
    "delivery": (
        "[условия поставки / срок изготовления]",
        "delivery_terms_approved",
    ),
}
INVOICE_PLACEHOLDER = "Счёт № [номер] от [дата]"
FORBIDDEN_PLACEHOLDERS = (
    "[номер]",
    "[дата]",
    "[наименование плательщика]",
    "[наименование позиции]",
    "[ед.]",
    "[кол-во]",
    "[приборы и аппараты согласно схеме]",
    "[тип шкафа]",
    "[цена]",
    "[сумма]",
    "[итого]",
    "[сумма прописью]",
    "[текст НДС]",
    "[условия оплаты]",
    "[условия возврата]",
    "[условия изменения спецификации / срок действия]",
    "[условия договора]",
    "[условия поставки / срок изготовления]",
    "[должность]",
    "[подписант]",
    "PLACEHOLDER",
)
FORBIDDEN_OLD_TOKENS = (
    "TDK Energy",
    "551",
    "44512",
    "EXW",
    "10-15",
    "10–15",
    "Сорок четыре тысячи",
    "НДС 16%",
    "РУ-АВР, IP54",
)


class ClientStyleExporterError(Exception):
    """Expected fail-closed exporter error."""


@dataclass(frozen=True)
class CommercialItem:
    name: str
    unit: str
    quantity: int
    instruments_and_devices: str
    cabinet_type_dimensions_material: str
    unit_price_kzt: int


@dataclass(frozen=True)
class ApprovedValues:
    invoice_number: str
    invoice_date: str
    payer_name: str
    object_name: str | None
    vat_text_approved: str
    payment_terms_approved: str
    delivery_terms_approved: str
    validity_terms_approved: str
    return_terms_approved: str
    signer_name: str
    signer_title: str
    approval_note: str
    input_hashes: Mapping[str, str]


@dataclass(frozen=True)
class TemplateLayout:
    expected_sheet_name: str
    invoice_number_cell: str
    invoice_date_cell: str
    payer_cell: str
    object_cell: str | None
    first_item_row: int
    item_columns: Mapping[str, str]
    amount_words_cell: str
    signer_name_cell: str
    signer_title_cell: str


@dataclass(frozen=True)
class TemplateTargets:
    total_row: int
    total_label_cell: str
    term_cells: Mapping[str, str]
    invoice_placeholder_cells: tuple[str, ...]


@dataclass
class ClientStyleExportResult:
    commercial_csv: Path
    internal_draft_xlsx: Path
    template_xlsx: Path
    template_contract_json: Path
    approval_json: Path
    output_xlsx: Path
    status: str = "FAIL"
    checks: dict[str, str] = field(
        default_factory=lambda: {
            "approval preflight": "fail",
            "template contract preflight": "fail",
            "commercial CSV": "fail",
            "export write": "fail",
            "reconciliation": "fail",
            "safety boundaries": "fail",
        }
    )
    red_flags: list[str] = field(default_factory=list)


def fail(message: str) -> NoReturn:
    raise ClientStyleExporterError(message)


def load_sibling_module(module_name: str, path: Path) -> ModuleType:
    spec = importlib.util.spec_from_file_location(module_name, path)
    if spec is None or spec.loader is None:
        fail(f"required helper could not be loaded: {path.name}")
    module = importlib.util.module_from_spec(spec)
    sys.modules[spec.name] = module
    spec.loader.exec_module(module)
    return module


approval_preflight = cast(
    Any,
    load_sibling_module(
        "preflight_client_style_invoice_export_for_exporter",
        APPROVAL_PREFLIGHT_SCRIPT,
    ),
)
template_contract_preflight = cast(
    Any,
    load_sibling_module(
        "preflight_client_style_invoice_template_contract_for_exporter",
        TEMPLATE_PREFLIGHT_SCRIPT,
    ),
)
commercial_csv_preflight = cast(
    Any,
    load_sibling_module(
        "preflight_quote_commercial_input_for_client_exporter",
        COMMERCIAL_PREFLIGHT_SCRIPT,
    ),
)
commercial_writer = cast(
    Any,
    load_sibling_module(
        "run_invoice_quote_commercial_from_csv_for_client_exporter",
        COMMERCIAL_WRITER_SCRIPT,
    ),
)
ooxml_cell_patcher = cast(
    Any,
    load_sibling_module(
        "ooxml_cell_patcher_for_client_exporter",
        OOXML_CELL_PATCHER_SCRIPT,
    ),
)
OoxmlCellPatcherError = ooxml_cell_patcher.OoxmlCellPatcherError
patch_existing_cells = ooxml_cell_patcher.patch_existing_cells


def parse_args(argv: Sequence[str] | None = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "Create one reconciled client-style XLSX candidate after approved "
            "input and template preflights."
        )
    )
    parser.add_argument("--commercial-csv", required=True, type=Path)
    parser.add_argument("--internal-draft-xlsx", required=True, type=Path)
    parser.add_argument("--template-xlsx", required=True, type=Path)
    parser.add_argument("--template-contract-json", required=True, type=Path)
    parser.add_argument("--approval-json", required=True, type=Path)
    parser.add_argument("--output-xlsx", required=True, type=Path)
    return parser.parse_args(argv)


def resolved(path: Path) -> Path:
    return path.expanduser().resolve(strict=False)


def is_inside_project(path: Path) -> bool:
    return resolved(path).is_relative_to(PROJECT_ROOT)


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as source:
        for chunk in iter(lambda: source.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def add_red_flag(result: ClientStyleExportResult, message: str) -> None:
    if message not in result.red_flags:
        result.red_flags.append(message)


def safe_prefixed_flags(prefix: str, values: Sequence[str]) -> list[str]:
    return [f"{prefix}: {value}" for value in values] or [f"{prefix}: failed"]


def load_json_object(path: Path, label: str) -> Mapping[str, Any]:
    try:
        data = json.loads(path.read_text(encoding="utf-8"))
    except OSError, UnicodeDecodeError, json.JSONDecodeError:
        fail(f"{label} could not be read safely")
    if not isinstance(data, dict):
        fail(f"{label} root must be an object")
    return cast(Mapping[str, Any], data)


def required_string(data: Mapping[str, Any], field_name: str, label: str) -> str:
    value = data.get(field_name)
    if not isinstance(value, str) or value.strip() == "":
        fail(f"{label} field is invalid: {field_name}")
    return value


def optional_string(
    data: Mapping[str, Any],
    field_name: str,
    label: str,
) -> str | None:
    value = data.get(field_name)
    if value is None:
        return None
    if not isinstance(value, str) or value.strip() == "":
        fail(f"{label} field is invalid: {field_name}")
    return value


def approval_values(data: Mapping[str, Any]) -> ApprovedValues:
    for field_name in APPROVED_FIELDS:
        if field_name not in data:
            fail(f"approval field is missing: {field_name}")
    input_hashes = {
        label: required_string(data, field_name, "approval")
        for label, field_name in APPROVAL_HASH_FIELDS.items()
    }
    return ApprovedValues(
        invoice_number=required_string(data, "invoice_number", "approval"),
        invoice_date=required_string(data, "invoice_date", "approval"),
        payer_name=required_string(data, "payer_name", "approval"),
        object_name=optional_string(data, "object_name", "approval"),
        vat_text_approved=required_string(data, "vat_text_approved", "approval"),
        payment_terms_approved=required_string(
            data,
            "payment_terms_approved",
            "approval",
        ),
        delivery_terms_approved=required_string(
            data,
            "delivery_terms_approved",
            "approval",
        ),
        validity_terms_approved=required_string(
            data,
            "validity_terms_approved",
            "approval",
        ),
        return_terms_approved=required_string(
            data,
            "return_terms_approved",
            "approval",
        ),
        signer_name=required_string(data, "signer_name", "approval"),
        signer_title=required_string(data, "signer_title", "approval"),
        approval_note=required_string(data, "approval_note", "approval"),
        input_hashes=input_hashes,
    )


def normalize_cell(value: str) -> str:
    return value.replace("$", "").upper()


def normalize_column(value: str) -> str:
    return value.replace("$", "").upper()


def template_layout(data: Mapping[str, Any]) -> TemplateLayout:
    expected_sheet_name = required_string(
        data,
        "expected_sheet_name",
        "template contract",
    )
    raw_layout = data.get("layout")
    if not isinstance(raw_layout, dict):
        fail("template contract layout is invalid")
    layout = cast(Mapping[str, Any], raw_layout)
    raw_item_columns = layout.get("item_columns")
    if not isinstance(raw_item_columns, dict):
        fail("template contract item columns are invalid")
    item_columns_data = cast(Mapping[str, Any], raw_item_columns)
    item_columns = {
        field_name: normalize_column(
            required_string(item_columns_data, field_name, "template contract")
        )
        for field_name in (
            "index",
            "name",
            "unit",
            "quantity",
            "instruments_and_devices",
            "cabinet",
            "unit_price",
            "line_total",
        )
    }
    first_item_row = layout.get("first_item_row")
    if isinstance(first_item_row, bool) or not isinstance(first_item_row, int):
        fail("template contract first item row is invalid")
    return TemplateLayout(
        expected_sheet_name=expected_sheet_name,
        invoice_number_cell=normalize_cell(
            required_string(layout, "invoice_number_cell", "template contract")
        ),
        invoice_date_cell=normalize_cell(
            required_string(layout, "invoice_date_cell", "template contract")
        ),
        payer_cell=normalize_cell(
            required_string(layout, "payer_cell", "template contract")
        ),
        object_cell=(
            None
            if layout.get("object_cell") is None
            else normalize_cell(
                required_string(layout, "object_cell", "template contract")
            )
        ),
        first_item_row=first_item_row,
        item_columns=item_columns,
        amount_words_cell=normalize_cell(
            required_string(layout, "amount_words_cell", "template contract")
        ),
        signer_name_cell=normalize_cell(
            required_string(layout, "signer_name_cell", "template contract")
        ),
        signer_title_cell=normalize_cell(
            required_string(layout, "signer_title_cell", "template contract")
        ),
    )


def load_commercial_item(path: Path) -> CommercialItem:
    try:
        with path.open("r", encoding="utf-8-sig", newline="") as csv_file:
            reader = csv.DictReader(
                csv_file,
                delimiter=commercial_csv_preflight.CSV_DELIMITER,
                strict=True,
            )
            if tuple(reader.fieldnames or ()) != tuple(
                commercial_csv_preflight.REQUIRED_COLUMNS
            ):
                fail("commercial CSV header does not match the strict schema")
            rows = [dict(row) for row in reader]
    except OSError, UnicodeDecodeError, csv.Error:
        fail("commercial CSV could not be read safely")
    if len(rows) != 1:
        fail("commercial CSV must contain exactly one item row")
    row = rows[0]
    if any(value is None or value.strip() == "" for value in row.values()):
        fail("commercial CSV contains an empty required field")
    if row["price_confirmed_by_igor"] != "yes":
        fail("commercial CSV price confirmation must be exact yes")
    if row["price_includes_vat"] != "yes":
        fail("commercial CSV VAT mode must be exact yes")
    try:
        quantity = int(row["quantity"])
        unit_price = int(row["unit_price_kzt"])
    except KeyError, TypeError, ValueError:
        fail("commercial CSV numeric fields are invalid")
    if quantity <= 0 or unit_price <= 0:
        fail("commercial CSV numeric fields must be positive integers")
    return CommercialItem(
        name=row["name"],
        unit=row["unit"],
        quantity=quantity,
        instruments_and_devices=row["instruments_and_devices"],
        cabinet_type_dimensions_material=row["cabinet_type_dimensions_material"],
        unit_price_kzt=unit_price,
    )


def find_exact_cells(worksheet: Any, expected_text: str) -> list[str]:
    found: list[str] = []
    for row in worksheet.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and cell.value.strip() == expected_text:
                found.append(cell.coordinate.upper())
    return found


def discover_template_targets(
    template_xlsx: Path,
    layout: TemplateLayout,
) -> TemplateTargets:
    try:
        workbook = load_workbook(template_xlsx, data_only=False, read_only=False)
    except Exception:
        fail("approved template could not be opened for target discovery")
    try:
        if layout.expected_sheet_name not in workbook.sheetnames:
            fail("approved template expected worksheet is missing")
        worksheet = workbook[layout.expected_sheet_name]
        term_cells: dict[str, str] = {}
        for label, (placeholder, _approval_field) in TERM_PLACEHOLDERS.items():
            matches = find_exact_cells(worksheet, placeholder)
            if len(matches) != 1:
                fail(f"template term placeholder is missing or duplicated: {label}")
            term_cells[label] = matches[0]

        invoice_cells = find_exact_cells(worksheet, INVOICE_PLACEHOLDER)
        if layout.invoice_number_cell not in invoice_cells:
            fail("template invoice placeholder does not match the contract")
        if not invoice_cells:
            fail("template invoice placeholder is missing")

        total_row = layout.first_item_row + 1
        total_matches = [
            cell.coordinate.upper()
            for cell in worksheet[total_row]
            if isinstance(cell.value, str)
            and cell.value.strip().rstrip(":").casefold() == "итого"
        ]
        if len(total_matches) != 1:
            fail("template total label is missing or duplicated")
        return TemplateTargets(
            total_row=total_row,
            total_label_cell=total_matches[0],
            term_cells=term_cells,
            invoice_placeholder_cells=tuple(invoice_cells),
        )
    finally:
        workbook.close()


def amount_words_for_total(total: int) -> str:
    try:
        words = cast(str, commercial_writer.integer_to_russian_words(total))
    except Exception:
        fail("total amount could not be converted to words safely")
    return f"{words[:1].upper() + words[1:]} тенге 00 тиын"


def approved_term_value(approved: ApprovedValues, field_name: str) -> str:
    return cast(str, getattr(approved, field_name))


def build_updates(
    layout: TemplateLayout,
    targets: TemplateTargets,
    item: CommercialItem,
    approved: ApprovedValues,
) -> tuple[dict[str, object], int, str]:
    total = item.quantity * item.unit_price_kzt
    amount_words = amount_words_for_total(total)
    invoice_text = f"Счёт № {approved.invoice_number} от {approved.invoice_date}"
    updates: dict[str, object] = {
        layout.payer_cell: f"Плательщик: {approved.payer_name}",
        f"{layout.item_columns['index']}{layout.first_item_row}": 1,
        f"{layout.item_columns['name']}{layout.first_item_row}": item.name,
        f"{layout.item_columns['unit']}{layout.first_item_row}": item.unit,
        f"{layout.item_columns['quantity']}{layout.first_item_row}": item.quantity,
        (
            f"{layout.item_columns['instruments_and_devices']}"
            f"{layout.first_item_row}"
        ): item.instruments_and_devices,
        f"{layout.item_columns['cabinet']}{layout.first_item_row}": (
            item.cabinet_type_dimensions_material
        ),
        f"{layout.item_columns['unit_price']}{layout.first_item_row}": (
            item.unit_price_kzt
        ),
        f"{layout.item_columns['line_total']}{layout.first_item_row}": total,
        f"{layout.item_columns['line_total']}{targets.total_row}": total,
        layout.amount_words_cell: (
            f"ВСЕГО: {amount_words}, {approved.vat_text_approved}"
        ),
        layout.signer_title_cell: approved.signer_title,
        layout.signer_name_cell: approved.signer_name,
    }
    for invoice_cell in targets.invoice_placeholder_cells:
        updates[invoice_cell] = invoice_text
    if layout.invoice_number_cell == layout.invoice_date_cell:
        updates[layout.invoice_number_cell] = invoice_text
    else:
        updates[layout.invoice_number_cell] = f"Счёт № {approved.invoice_number}"
        updates[layout.invoice_date_cell] = f"от {approved.invoice_date}"
    if layout.object_cell is not None:
        updates[layout.object_cell] = approved.object_name
    for label, (_placeholder, approval_field) in TERM_PLACEHOLDERS.items():
        updates[targets.term_cells[label]] = approved_term_value(
            approved,
            approval_field,
        )
    return updates, total, amount_words


def candidate_path_for(output: Path) -> Path:
    return output.with_name(f".{output.stem}.{uuid.uuid4().hex}.candidate.xlsx")


def capture_input_hashes(paths: Mapping[str, Path]) -> dict[str, str]:
    try:
        return {label: sha256_file(path) for label, path in paths.items()}
    except OSError:
        fail("input immutability hashes could not be captured")


def inputs_unchanged(
    paths: Mapping[str, Path],
    expected_hashes: Mapping[str, str],
) -> bool:
    try:
        return all(
            sha256_file(path) == expected_hashes[label] for label, path in paths.items()
        )
    except OSError:
        return False


def package_token_hits(path: Path, tokens: Sequence[str]) -> list[str]:
    hits: list[str] = []
    try:
        with zipfile.ZipFile(path) as archive:
            for name in archive.namelist():
                raw = archive.read(name)
                try:
                    text = raw.decode("utf-8")
                except UnicodeDecodeError:
                    continue
                for token in tokens:
                    if token.casefold() in text.casefold() and token not in hits:
                        hits.append(token)
    except OSError, zipfile.BadZipFile:
        fail("output XLSX package could not be scanned safely")
    return hits


def allowed_dynamic_tokens(
    item: CommercialItem,
    approved: ApprovedValues,
    total: int,
    amount_words: str,
    expected_invoice: str,
    expected_amount: str,
) -> set[str]:
    tokens = {
        approved.invoice_number,
        approved.invoice_date,
        approved.payer_name,
        approved.vat_text_approved,
        approved.payment_terms_approved,
        approved.delivery_terms_approved,
        approved.validity_terms_approved,
        approved.return_terms_approved,
        approved.signer_name,
        approved.signer_title,
        approved.approval_note,
        item.name,
        item.unit,
        item.instruments_and_devices,
        item.cabinet_type_dimensions_material,
        str(item.unit_price_kzt),
        str(total),
        amount_words,
        expected_invoice,
        expected_amount,
    }
    if approved.object_name is not None:
        tokens.add(approved.object_name)
    return {token for token in tokens if token}


def is_approved_dynamic_token(token: str, allowed_tokens: set[str]) -> bool:
    folded_token = token.casefold()
    return any(folded_token in value.casefold() for value in allowed_tokens)


def reconcile_candidate(
    candidate: Path,
    *,
    paths: Mapping[str, Path],
    input_hashes_before: Mapping[str, str],
    layout: TemplateLayout,
    targets: TemplateTargets,
    item: CommercialItem,
    approved: ApprovedValues,
    total: int,
    amount_words: str,
) -> list[str]:
    failures: list[str] = []
    if not candidate.is_file():
        return ["reconciliation output does not exist"]
    if is_inside_project(candidate):
        failures.append("reconciliation output is inside the Git project")

    expected_invoice = f"Счёт № {approved.invoice_number} от {approved.invoice_date}"
    expected_amount = f"ВСЕГО: {amount_words}, {approved.vat_text_approved}"
    try:
        workbook = load_workbook(candidate, data_only=False, read_only=False)
    except Exception:
        return [*failures, "reconciliation workbook could not be opened"]
    try:
        if layout.expected_sheet_name not in workbook.sheetnames:
            failures.append("reconciliation expected worksheet is missing")
            return failures
        worksheet = workbook[layout.expected_sheet_name]
        if worksheet[layout.invoice_number_cell].value != expected_invoice:
            failures.append("reconciliation invoice header mismatch")
        if (
            layout.invoice_date_cell != layout.invoice_number_cell
            and worksheet[layout.invoice_date_cell].value
            != f"от {approved.invoice_date}"
        ):
            failures.append("reconciliation invoice date mismatch")
        if worksheet[layout.payer_cell].value != (f"Плательщик: {approved.payer_name}"):
            failures.append("reconciliation payer mismatch")
        if (
            layout.object_cell is not None
            and worksheet[layout.object_cell].value != approved.object_name
        ):
            failures.append("reconciliation object mismatch")

        item_row = layout.first_item_row
        expected_item_values: dict[str, object] = {
            "index": 1,
            "name": item.name,
            "unit": item.unit,
            "quantity": item.quantity,
            "instruments_and_devices": item.instruments_and_devices,
            "cabinet": item.cabinet_type_dimensions_material,
            "unit_price": item.unit_price_kzt,
            "line_total": total,
        }
        for field_name, expected_value in expected_item_values.items():
            coordinate = f"{layout.item_columns[field_name]}{item_row}"
            if worksheet[coordinate].value != expected_value:
                failures.append(f"reconciliation item mismatch: {field_name}")
        total_cell = f"{layout.item_columns['line_total']}{targets.total_row}"
        if worksheet[total_cell].value != total:
            failures.append("reconciliation total mismatch")
        if worksheet[targets.total_label_cell].value.strip().rstrip(":").casefold() != (
            "итого"
        ):
            failures.append("reconciliation total label mismatch")
        if worksheet[layout.amount_words_cell].value != expected_amount:
            failures.append("reconciliation amount words or VAT mismatch")
        if worksheet[layout.signer_title_cell].value != approved.signer_title:
            failures.append("reconciliation signer title mismatch")
        if worksheet[layout.signer_name_cell].value != approved.signer_name:
            failures.append("reconciliation signer name mismatch")
        for label, (_placeholder, approval_field) in TERM_PLACEHOLDERS.items():
            if worksheet[targets.term_cells[label]].value != approved_term_value(
                approved,
                approval_field,
            ):
                failures.append(f"reconciliation approved term mismatch: {label}")
    finally:
        workbook.close()

    placeholder_hits = package_token_hits(candidate, FORBIDDEN_PLACEHOLDERS)
    if placeholder_hits:
        failures.append("reconciliation found forbidden placeholder")

    allowed_tokens = allowed_dynamic_tokens(
        item,
        approved,
        total,
        amount_words,
        expected_invoice,
        expected_amount,
    )
    legacy_hits = package_token_hits(candidate, FORBIDDEN_OLD_TOKENS)
    unapproved_legacy_hits = [
        token
        for token in legacy_hits
        if not is_approved_dynamic_token(token, allowed_tokens)
    ]
    if unapproved_legacy_hits:
        failures.append("reconciliation found unapproved legacy token")
    if not inputs_unchanged(paths, input_hashes_before):
        failures.append("reconciliation input file changed during export")
    for label, expected_hash in approved.input_hashes.items():
        path = paths[label]
        try:
            actual_hash = sha256_file(path)
        except OSError:
            failures.append("reconciliation approved input hash could not be read")
            continue
        if actual_hash != expected_hash:
            failures.append(f"reconciliation approved input hash mismatch: {label}")
    return failures


def remove_if_exists(path: Path) -> bool:
    if not path.exists():
        return True
    try:
        path.unlink()
    except OSError:
        return False
    return True


def finalize_failure(
    result: ClientStyleExportResult,
    *,
    candidate: Path | None,
    output_preexisted: bool,
    paths: Mapping[str, Path] | None = None,
    input_hashes_before: Mapping[str, str] | None = None,
) -> ClientStyleExportResult:
    clean = True
    if candidate is not None and not remove_if_exists(candidate):
        clean = False
        add_red_flag(result, "temporary candidate could not be removed")
    if not output_preexisted and result.output_xlsx.exists():
        if not remove_if_exists(result.output_xlsx):
            clean = False
            add_red_flag(result, "partial output could not be removed")
    if (
        paths is not None
        and input_hashes_before is not None
        and not inputs_unchanged(paths, input_hashes_before)
    ):
        clean = False
        add_red_flag(result, "an input file changed during failed export")
    result.checks["safety boundaries"] = "pass" if clean else "fail"
    return result


def run_export(
    commercial_csv: Path,
    internal_draft_xlsx: Path,
    template_xlsx: Path,
    template_contract_json: Path,
    approval_json: Path,
    output_xlsx: Path,
) -> ClientStyleExportResult:
    result = ClientStyleExportResult(
        commercial_csv=resolved(commercial_csv),
        internal_draft_xlsx=resolved(internal_draft_xlsx),
        template_xlsx=resolved(template_xlsx),
        template_contract_json=resolved(template_contract_json),
        approval_json=resolved(approval_json),
        output_xlsx=resolved(output_xlsx),
    )
    output_preexisted = result.output_xlsx.exists()
    candidate: Path | None = None
    input_paths: dict[str, Path] | None = None
    input_hashes_before: dict[str, str] | None = None

    try:
        approval_result = approval_preflight.preflight(
            result.commercial_csv,
            result.internal_draft_xlsx,
            result.template_xlsx,
            result.approval_json,
            result.output_xlsx,
        )
        if approval_result.status != "PASS":
            result.red_flags.extend(
                safe_prefixed_flags(
                    "approval preflight",
                    approval_result.red_flags,
                )
            )
            return finalize_failure(
                result,
                candidate=None,
                output_preexisted=output_preexisted,
            )
        result.checks["approval preflight"] = "pass"

        template_result = template_contract_preflight.preflight(
            result.template_xlsx,
            result.template_contract_json,
        )
        if template_result.status != "PASS":
            result.red_flags.extend(
                safe_prefixed_flags(
                    "template contract preflight",
                    template_result.red_flags,
                )
            )
            return finalize_failure(
                result,
                candidate=None,
                output_preexisted=output_preexisted,
            )
        result.checks["template contract preflight"] = "pass"

        input_paths = {
            "commercial CSV": result.commercial_csv,
            "internal draft XLSX": result.internal_draft_xlsx,
            "template XLSX": result.template_xlsx,
            "template contract JSON": result.template_contract_json,
            "approval JSON": result.approval_json,
        }
        input_hashes_before = capture_input_hashes(input_paths)
        approved = approval_values(
            load_json_object(result.approval_json, "approval JSON")
        )
        layout = template_layout(
            load_json_object(
                result.template_contract_json,
                "template contract JSON",
            )
        )

        commercial_result = commercial_csv_preflight.preflight(result.commercial_csv)
        if commercial_result.status != "PASS":
            result.red_flags.extend(
                safe_prefixed_flags(
                    "commercial CSV",
                    commercial_result.failures,
                )
            )
            return finalize_failure(
                result,
                candidate=None,
                output_preexisted=output_preexisted,
                paths=input_paths,
                input_hashes_before=input_hashes_before,
            )
        if commercial_result.row_count != 1:
            fail("commercial CSV must contain exactly one item row")
        item = load_commercial_item(result.commercial_csv)
        result.checks["commercial CSV"] = "pass"

        targets = discover_template_targets(result.template_xlsx, layout)
        updates, total, amount_words = build_updates(
            layout,
            targets,
            item,
            approved,
        )
        candidate = candidate_path_for(result.output_xlsx)
        patch_existing_cells(
            template=result.template_xlsx,
            output=candidate,
            sheet_name=layout.expected_sheet_name,
            updates=updates,
        )
        result.checks["export write"] = "pass"

        reconciliation_failures = reconcile_candidate(
            candidate,
            paths=input_paths,
            input_hashes_before=input_hashes_before,
            layout=layout,
            targets=targets,
            item=item,
            approved=approved,
            total=total,
            amount_words=amount_words,
        )
        if reconciliation_failures:
            result.red_flags.extend(reconciliation_failures)
            return finalize_failure(
                result,
                candidate=candidate,
                output_preexisted=output_preexisted,
                paths=input_paths,
                input_hashes_before=input_hashes_before,
            )
        result.checks["reconciliation"] = "pass"

        if result.output_xlsx.exists():
            fail("output XLSX appeared before atomic publish")
        try:
            candidate.rename(result.output_xlsx)
        except OSError:
            fail("reconciled candidate could not be published atomically")
        candidate = None
        result.checks["safety boundaries"] = "pass"
        result.status = "PASS"
        return result
    except (ClientStyleExporterError, OoxmlCellPatcherError) as error:
        add_red_flag(result, str(error))
    except Exception:
        add_red_flag(result, "unexpected internal exporter failure")
    return finalize_failure(
        result,
        candidate=candidate,
        output_preexisted=output_preexisted,
        paths=input_paths,
        input_hashes_before=input_hashes_before,
    )


def format_items(values: Sequence[str]) -> list[str]:
    return list(values) if values else ["none"]


def format_report(result: ClientStyleExportResult) -> str:
    lines = [
        REPORT_START,
        "",
        "Status:",
        result.status,
        "",
        "Mode:",
        MODE,
        "",
        "Checks:",
    ]
    lines.extend(f"{name}: {status}" for name, status in result.checks.items())
    lines.extend(["", "Red flags:"])
    lines.extend(format_items(result.red_flags))
    lines.extend(
        [
            "",
            "Output:",
            str(result.output_xlsx) if result.status == "PASS" else "not created",
            "",
            "Commercial status:",
            COMMERCIAL_STATUS,
            "",
            "Human Approval:",
            HUMAN_APPROVAL,
            "",
            REPORT_END,
        ]
    )
    return "\n".join(lines)


def main(argv: Sequence[str] | None = None) -> int:
    args = parse_args(argv)
    result = run_export(
        args.commercial_csv,
        args.internal_draft_xlsx,
        args.template_xlsx,
        args.template_contract_json,
        args.approval_json,
        args.output_xlsx,
    )
    print(format_report(result))
    return 0 if result.status == "PASS" else 1


if __name__ == "__main__":
    raise SystemExit(main())
