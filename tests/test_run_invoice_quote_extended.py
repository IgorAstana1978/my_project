import importlib.util
import json
import subprocess
import sys
import tempfile
from pathlib import Path
from types import ModuleType
from typing import Any, cast

from openpyxl import Workbook, load_workbook  # type: ignore[import-untyped]

PROJECT_ROOT = Path(__file__).resolve().parents[1]
RUNNER_SCRIPT = PROJECT_ROOT / "scripts" / "run_invoice_quote_extended.py"
SHEET_NAME = "Счёт-КП шаблон"


def load_script_module(module_name: str, path: Path) -> ModuleType:
    spec = importlib.util.spec_from_file_location(module_name, path)
    assert spec is not None
    assert spec.loader is not None
    module = importlib.util.module_from_spec(spec)
    sys.modules[spec.name] = module
    spec.loader.exec_module(module)
    return module


runner = cast(
    Any,
    load_script_module("run_invoice_quote_extended_for_test", RUNNER_SCRIPT),
)


def layout_json(capacity: int = 8) -> dict[str, Any]:
    item_start_row = 17
    item_end_row = item_start_row + capacity - 1
    total_row = item_start_row + capacity
    return {
        "item_start_row": item_start_row,
        "item_end_row": item_end_row,
        "capacity": capacity,
        "total_row": total_row,
        "signature_range": f"B{20 + capacity}:I{22 + capacity}",
        "header_ranges": ["C2:I6", "B4:B6"],
        "formula_cells": [f"I{row}" for row in range(17, 17 + capacity)]
        + [f"I{total_row}"],
    }


def write_extended_template(path: Path, capacity: int = 8) -> None:
    layout = layout_json(capacity)
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = SHEET_NAME

    for row in range(2, 7):
        for column in range(3, 10):
            worksheet.cell(row=row, column=column).value = f"header-{row}-{column}"
    for row in range(4, 7):
        worksheet.cell(row=row, column=2).value = f"req-{row}"

    worksheet.merge_cells(layout["signature_range"])
    worksheet.cell(row=20 + capacity, column=2).value = "signature"

    for row in range(layout["item_start_row"], layout["item_end_row"] + 1):
        worksheet[f"C{row}"] = "template item"
        worksheet[f"D{row}"] = "шт"
        worksheet[f"E{row}"] = 1
        worksheet[f"F{row}"] = "template instruments"
        worksheet[f"G{row}"] = "template cabinet"
        worksheet[f"H{row}"] = "template price"
        worksheet[f"I{row}"] = f"=E{row}*H{row}"
    worksheet[f"I{layout['total_row']}"] = (
        f"=SUM(I{layout['item_start_row']}:I{layout['item_end_row']})"
    )

    workbook.save(path)


def item(index: int) -> dict[str, Any]:
    return {
        "name": f"ВРУ-{index}",
        "unit": "шт.",
        "quantity": index,
        "instruments_and_devices": "нужно уточнить",
        "cabinet_type_dimensions_material": "нужно уточнить",
        "price_kzt": None,
        "price_confirmed_by_igor": False,
    }


def payload(items_count: int) -> dict[str, Any]:
    return {"items": [item(index) for index in range(1, items_count + 1)]}


def job_json(items_count: int, capacity: int = 8) -> dict[str, Any]:
    return {"payload": payload(items_count), "layout": layout_json(capacity)}


def write_json(path: Path, data: Any) -> None:
    path.write_text(json.dumps(data, ensure_ascii=False), encoding="utf-8")


def runner_args(job_json_path: Path, template: Path, output: Path) -> list[str]:
    return [
        "--job-json",
        str(job_json_path),
        "--template",
        str(template),
        "--output",
        str(output),
    ]


def run_runner(
    job_json_path: Path,
    template: Path,
    output: Path,
) -> subprocess.CompletedProcess[str]:
    return subprocess.run(
        [sys.executable, str(RUNNER_SCRIPT)]
        + runner_args(job_json_path, template, output),
        capture_output=True,
        text=True,
        check=False,
    )


def workbook_value(path: Path, cell: str) -> Any:
    workbook = load_workbook(path, data_only=False)
    worksheet = workbook[SHEET_NAME]
    return worksheet[cell].value


def test_runner_successfully_runs_six_item_job(tmp_path: Path) -> None:
    template = tmp_path / "extended_template.xlsx"
    job_json_path = tmp_path / "job.json"
    output_dir = tmp_path / "out"
    output_dir.mkdir()
    output = output_dir / "draft.xlsx"
    write_extended_template(template)
    write_json(job_json_path, job_json(6))

    result = run_runner(job_json_path, template, output)

    assert result.returncode == 0
    assert "CREATED:" in result.stdout
    assert "ERROR:" not in result.stderr
    assert output.is_file()
    assert workbook_value(output, "C22") == "ВРУ-6"


def test_runner_successfully_runs_fifty_item_job(tmp_path: Path) -> None:
    template = tmp_path / "extended_template.xlsx"
    job_json_path = tmp_path / "job.json"
    output_dir = tmp_path / "out"
    output_dir.mkdir()
    output = output_dir / "draft.xlsx"
    write_extended_template(template, capacity=50)
    write_json(job_json_path, job_json(50, capacity=50))

    result = run_runner(job_json_path, template, output)

    assert result.returncode == 0
    assert "CREATED:" in result.stdout
    assert "ERROR:" not in result.stderr
    assert output.is_file()
    assert workbook_value(output, "C66") == "ВРУ-50"
    assert workbook_value(output, "I67") == "=SUM(I17:I66)"


def test_runner_missing_job_json_returns_one(tmp_path: Path, capsys: Any) -> None:
    template = tmp_path / "extended_template.xlsx"
    output = tmp_path / "out" / "draft.xlsx"
    output.parent.mkdir()
    write_extended_template(template)

    exit_code = runner.main(
        runner_args(tmp_path / "missing_job.json", template, output)
    )

    captured = capsys.readouterr()
    assert exit_code == 1
    assert "ERROR:" in captured.err
    assert not output.exists()
    assert list(output.parent.iterdir()) == []


def test_runner_invalid_job_json_returns_one(tmp_path: Path, capsys: Any) -> None:
    template = tmp_path / "extended_template.xlsx"
    job_json_path = tmp_path / "job.json"
    output = tmp_path / "out" / "draft.xlsx"
    output.parent.mkdir()
    write_extended_template(template)
    job_json_path.write_text("{invalid", encoding="utf-8")

    exit_code = runner.main(runner_args(job_json_path, template, output))

    captured = capsys.readouterr()
    assert exit_code == 1
    assert "ERROR:" in captured.err
    assert not output.exists()
    assert list(output.parent.iterdir()) == []


def test_runner_root_json_must_be_object(tmp_path: Path, capsys: Any) -> None:
    template = tmp_path / "extended_template.xlsx"
    job_json_path = tmp_path / "job.json"
    output = tmp_path / "out" / "draft.xlsx"
    output.parent.mkdir()
    write_extended_template(template)
    write_json(job_json_path, [])

    exit_code = runner.main(runner_args(job_json_path, template, output))

    captured = capsys.readouterr()
    assert exit_code == 1
    assert "job JSON must be an object" in captured.err
    assert not output.exists()


def test_runner_requires_payload_object(tmp_path: Path, capsys: Any) -> None:
    template = tmp_path / "extended_template.xlsx"
    job_json_path = tmp_path / "job.json"
    output = tmp_path / "out" / "draft.xlsx"
    output.parent.mkdir()
    write_extended_template(template)
    write_json(job_json_path, {"layout": layout_json()})

    exit_code = runner.main(runner_args(job_json_path, template, output))

    captured = capsys.readouterr()
    assert exit_code == 1
    assert "job.payload is required" in captured.err
    assert not output.exists()


def test_runner_rejects_non_object_payload(tmp_path: Path, capsys: Any) -> None:
    template = tmp_path / "extended_template.xlsx"
    job_json_path = tmp_path / "job.json"
    output = tmp_path / "out" / "draft.xlsx"
    output.parent.mkdir()
    write_extended_template(template)
    write_json(job_json_path, {"payload": [], "layout": layout_json()})

    exit_code = runner.main(runner_args(job_json_path, template, output))

    captured = capsys.readouterr()
    assert exit_code == 1
    assert "job.payload must be an object" in captured.err
    assert not output.exists()


def test_runner_requires_layout_object(tmp_path: Path, capsys: Any) -> None:
    template = tmp_path / "extended_template.xlsx"
    job_json_path = tmp_path / "job.json"
    output = tmp_path / "out" / "draft.xlsx"
    output.parent.mkdir()
    write_extended_template(template)
    write_json(job_json_path, {"payload": payload(6)})

    exit_code = runner.main(runner_args(job_json_path, template, output))

    captured = capsys.readouterr()
    assert exit_code == 1
    assert "job.layout is required" in captured.err
    assert not output.exists()


def test_runner_rejects_non_object_layout(tmp_path: Path, capsys: Any) -> None:
    template = tmp_path / "extended_template.xlsx"
    job_json_path = tmp_path / "job.json"
    output = tmp_path / "out" / "draft.xlsx"
    output.parent.mkdir()
    write_extended_template(template)
    write_json(job_json_path, {"payload": payload(6), "layout": []})

    exit_code = runner.main(runner_args(job_json_path, template, output))

    captured = capsys.readouterr()
    assert exit_code == 1
    assert "job.layout must be an object" in captured.err
    assert not output.exists()


def test_runner_returns_downstream_error_for_capacity_overflow(tmp_path: Path) -> None:
    template = tmp_path / "extended_template.xlsx"
    job_json_path = tmp_path / "job.json"
    output_dir = tmp_path / "out"
    output_dir.mkdir()
    output = output_dir / "draft.xlsx"
    write_extended_template(template, capacity=6)
    write_json(job_json_path, job_json(7, capacity=6))

    result = run_runner(job_json_path, template, output)

    assert result.returncode == 1
    assert "ERROR:" in result.stderr
    assert "items count 7 exceeds layout capacity 6" in result.stderr
    assert not output.exists()
    assert list(output_dir.iterdir()) == []


def test_runner_existing_output_is_not_overwritten(tmp_path: Path) -> None:
    template = tmp_path / "extended_template.xlsx"
    job_json_path = tmp_path / "job.json"
    output_dir = tmp_path / "out"
    output_dir.mkdir()
    output = output_dir / "draft.xlsx"
    output.write_bytes(b"existing")
    write_extended_template(template)
    write_json(job_json_path, job_json(6))

    result = run_runner(job_json_path, template, output)

    assert result.returncode == 1
    assert "ERROR:" in result.stderr
    assert "output already exists" in result.stderr
    assert output.read_bytes() == b"existing"
    assert list(output_dir.iterdir()) == [output]


def test_runner_cleans_temp_payload_and_layout_after_success(
    tmp_path: Path,
    monkeypatch: Any,
) -> None:
    template = tmp_path / "extended_template.xlsx"
    job_json_path = tmp_path / "job.json"
    temp_parent = tmp_path / "runner-temp"
    temp_parent.mkdir()
    output = tmp_path / "out" / "draft.xlsx"
    output.parent.mkdir()
    created_dirs: list[Path] = []
    write_extended_template(template)
    write_json(job_json_path, job_json(6))

    def temporary_directory(*args: Any, **kwargs: Any) -> Any:
        kwargs["dir"] = temp_parent
        manager = tempfile.TemporaryDirectory(*args, **kwargs)
        created_dirs.append(Path(manager.name))
        return manager

    monkeypatch.setattr(runner, "TEMPORARY_DIRECTORY", temporary_directory)

    exit_code = runner.main(runner_args(job_json_path, template, output))

    assert exit_code == 0
    assert created_dirs
    assert all(not path.exists() for path in created_dirs)
    assert list(temp_parent.iterdir()) == []


def test_runner_cleans_temp_payload_and_layout_after_error(
    tmp_path: Path,
    monkeypatch: Any,
) -> None:
    template = tmp_path / "extended_template.xlsx"
    job_json_path = tmp_path / "job.json"
    temp_parent = tmp_path / "runner-temp"
    temp_parent.mkdir()
    output_dir = tmp_path / "out"
    output_dir.mkdir()
    output = output_dir / "draft.xlsx"
    created_dirs: list[Path] = []
    write_extended_template(template, capacity=6)
    write_json(job_json_path, job_json(7, capacity=6))

    def temporary_directory(*args: Any, **kwargs: Any) -> Any:
        kwargs["dir"] = temp_parent
        manager = tempfile.TemporaryDirectory(*args, **kwargs)
        created_dirs.append(Path(manager.name))
        return manager

    monkeypatch.setattr(runner, "TEMPORARY_DIRECTORY", temporary_directory)

    exit_code = runner.main(runner_args(job_json_path, template, output))

    assert exit_code == 1
    assert created_dirs
    assert all(not path.exists() for path in created_dirs)
    assert list(temp_parent.iterdir()) == []
    assert not output.exists()
    assert list(output_dir.iterdir()) == []


def test_no_real_xlsx_or_manifest_files_are_added_to_git() -> None:
    status = subprocess.run(
        ["git", "status", "--short", "--untracked-files=all"],
        cwd=PROJECT_ROOT,
        capture_output=True,
        text=True,
        check=True,
    )
    changed_paths = [
        line[3:].strip() for line in status.stdout.splitlines() if line.strip()
    ]
    assert not any(path.endswith(".xlsx") for path in changed_paths)
    assert not any(
        Path(path).name.casefold() == "manifest.json" for path in changed_paths
    )
