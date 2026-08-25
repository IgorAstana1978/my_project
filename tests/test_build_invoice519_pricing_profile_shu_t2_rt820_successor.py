from __future__ import annotations

import copy
import hashlib
import importlib.util
import json
import sys
from pathlib import Path
from types import ModuleType
from typing import Any, cast

import pytest
from openpyxl import Workbook  # type: ignore[import-untyped]

PROJECT_ROOT = Path(__file__).resolve().parents[1]
SCRIPT = (
    PROJECT_ROOT
    / "scripts"
    / "build_invoice519_pricing_profile_shu_t2_rt820_successor.py"
)
REAL_PARENT = Path(
    r"C:\Users\IgorN\Documents\production_ai_cases\CASE-QF-PROJECT-2024-086-SHU-T1-PRICING-PROFILE-SUCCESSOR-20260820-001\invoice519-pricing-profile-additive-successor.json"
)
REAL_TECHNICAL = Path(
    r"C:\Users\IgorN\Documents\production_ai_cases\CASE-QF-PROJECT-2024-086-SHU-T2-RT820-TECHNICAL-SUCCESSOR-20260824-001\price-calculator-input-v0.2-completed-shu-t2-rt820-successor.json"
)
REAL_DECISION = Path(
    r"C:\Users\IgorN\Documents\production_ai_cases\CASE-QF-PROJECT-2024-086-SHU-T2-RT820-SCOPE-DECISION-20260820-001\technical-shu-t2-rt820-scope-human-decision-v0.1.json"
)
ADDITIVE_TEST = (
    PROJECT_ROOT
    / "tests"
    / "test_build_invoice519_pricing_profile_additive_successor.py"
)


def load_module() -> ModuleType:
    spec = importlib.util.spec_from_file_location("shu_t2_pricing_builder", SCRIPT)
    assert spec is not None and spec.loader is not None
    module = importlib.util.module_from_spec(spec)
    sys.modules[spec.name] = module
    spec.loader.exec_module(module)
    return module


builder = load_module()


def load_additive_test_helpers() -> ModuleType:
    spec = importlib.util.spec_from_file_location(
        "additive_test_helpers", ADDITIVE_TEST
    )
    assert spec is not None and spec.loader is not None
    module = importlib.util.module_from_spec(spec)
    sys.modules[spec.name] = module
    spec.loader.exec_module(module)
    return module


additive_helpers = load_additive_test_helpers()


def sha256(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()


def write_json(path: Path, value: Any) -> None:
    path.write_text(
        json.dumps(value, ensure_ascii=False, indent=2) + "\n", encoding="utf-8"
    )


def make_workbook(path: Path, *, material: int = 15000, work: int = 900) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "КРН"
    sheet.cell(19, 1, "Терморегулятор RT-820")
    sheet.cell(19, 2, material)
    sheet.cell(19, 3, work)
    workbook.save(path)
    workbook.close()


def synthetic_parent_profile() -> dict[str, Any]:
    _, profile = additive_helpers.build()
    profile["authority"] = copy.deepcopy(builder.EXPECTED_AUTHORITY)
    profile["safety_flags"] = copy.deepcopy(builder.EXPECTED_SAFETY_FLAGS)
    profile["non_approvals"] = copy.deepcopy(builder.EXPECTED_NON_APPROVALS)
    scope = profile["current_completed_technical_scope"]
    scope["cabinet_groups"][2].update(
        {
            "product_name": "ШУ-Т2",
            "row_draft_ids": [
                "ROW-DRAFT-0020",
                "ROW-DRAFT-0021",
                "ROW-DRAFT-0022",
                "ROW-DRAFT-0023",
                "ROW-DRAFT-0024",
                "ROW-DRAFT-0025",
                "ROW-DRAFT-0026",
                "ROW-DRAFT-0027",
            ],
        }
    )
    for position_id, source_id, row_ids, row_indexes in (
        (
            "PRICE-POSITION-009",
            "TFE-016",
            ["ROW-DRAFT-0020", "ROW-DRAFT-0024"],
            [19, 23],
        ),
        (
            "PRICE-POSITION-023",
            "TFE-041",
            ["ROW-DRAFT-0021", "ROW-DRAFT-0025"],
            [20, 24],
        ),
        (
            "PRICE-POSITION-035",
            "TFE-061",
            ["ROW-DRAFT-0022", "ROW-DRAFT-0026"],
            [21, 25],
        ),
        (
            "PRICE-POSITION-047",
            "TFE-083",
            ["ROW-DRAFT-0023", "ROW-DRAFT-0027"],
            [22, 26],
        ),
    ):
        position = scope["pricing_positions"][int(position_id[-3:]) - 1]
        position.update(
            {
                "source_position_id": source_id,
                "product_name": "ШУ-Т2",
                "row_draft_ids": row_ids,
                "row_draft_json_paths": [
                    f"$.calculator_input_format.row_drafts[{index}]"
                    for index in row_indexes
                ],
                "composition_fingerprint_sha256": builder.OLD_FINGERPRINT,
                "approved_unit_price_kzt": None,
                "approved_unit_price_decision_status": "NOT_CALCULATED_NOT_APPROVED",
            }
        )
    scope["composition_fingerprints"][0] = {
        "fingerprint_sha256": builder.OLD_FINGERPRINT,
        "canonicalization": (
            "SHA256 UTF-8 canonical JSON of sorted "
            "component_code/component_qty/install_type tuples"
        ),
        "components": [builder.NEW_COMPONENTS[0], builder.NEW_COMPONENTS[2]],
        "source_position_ids": ["TFE-016", "TFE-041", "TFE-061", "TFE-083"],
        "pricing_position_ids": [
            "PRICE-POSITION-009",
            "PRICE-POSITION-023",
            "PRICE-POSITION-035",
            "PRICE-POSITION-047",
        ],
    }
    return cast(dict[str, Any], profile)


def synthetic_decision(parent_path: Path, parent_sha: str) -> dict[str, Any]:
    positions = builder._expected_positions()
    return {
        "schema_version": builder.DECISION_SCHEMA,
        "artifact_type": "IMMUTABLE_HUMAN_DECISION_CAPTURE",
        "project_id": builder.PROJECT_ID,
        "decision_id": builder.DECISION_ID,
        "status": builder.DECISION_STATUS,
        "authority": builder.AUTHORITY,
        "application_status": builder.APPLICATION_STATUS,
        "input_bindings": [
            {
                "role": "pricing_profile",
                "path": str(parent_path.resolve()),
                "expected_sha256": parent_sha,
                "actual_sha256": parent_sha,
            }
        ],
        "exact_scope": {
            "product": "ШУ-Т2",
            "cabinet_group_id": builder.TARGET_GROUP_ID,
            "positions": positions,
            "source_evidence_row_count": 8,
            "future_component_row_count": 4,
        },
        "rt820_contract": {
            "component_code": "EKF-RT-820",
            "component_qty_per_physical_cabinet": 1,
            "install_type": "temperature_relay_din_2mod",
            "module_width": 2,
            "source_range": "КРН!A19:C19",
            "source_label": "Терморегулятор RT-820",
            "material_kzt": 15000,
            "work_kzt": 900,
            "generic_work_432_prohibited": True,
            "family_fallback_prohibited": True,
            "fuzzy_fallback_prohibited": True,
            "similar_relay_fallback_prohibited": True,
        },
        "bundle_semantics": {
            "tst05_provenance_only": True,
            "separate_tst05_component_row": False,
            "separate_tst05_material_charge": False,
            "separate_tst05_work_charge": False,
            "separate_tst05_pricing_row": False,
        },
        "supersession": {
            "outside_cabinet_exclusion_count_must_be_derived": True,
            "outside_cabinet_exclusion_count_override_prohibited": True,
            "shu_t1_unchanged": True,
        },
    }


def synthetic_technical(decision_path: Path, decision_sha: str) -> dict[str, Any]:
    value = additive_helpers.technical_successor()
    group3 = value["cabinet_groups"][2]
    group3.update(
        {
            "product_name": "ШУ-Т2",
            "row_draft_ids": [
                *(f"ROW-DRAFT-{index:04d}" for index in range(20, 28)),
                *(f"ROW-DRAFT-{index:04d}" for index in range(113, 117)),
            ],
        }
    )
    expected_positions = builder._expected_positions()
    evidence_ids = [
        evidence
        for item in expected_positions
        for evidence in (item["relay_evidence_id"], item["sensor_evidence_id"])
    ]
    value["source"]["shu_t2_rt820_technical_successor"] = {
        "contract": builder.TECHNICAL_CONTRACT,
        "human_decision": {
            "path": str(decision_path.resolve()),
            "sha256": decision_sha,
            "decision_id": builder.DECISION_ID,
            "application_status": builder.APPLICATION_STATUS,
        },
        "technical_projection": {
            "row_ids": [f"ROW-DRAFT-{index:04d}" for index in range(113, 117)],
            "evidence_count": 8,
            "evidence_ids": evidence_ids,
            "outside_cabinet_membership_asserted": False,
            "outside_cabinet_count_transition_asserted": False,
        },
        "rt820_pricing_provenance_only": {
            "source_range": "КРН!A19:C19",
            "material_kzt": 15000,
            "work_kzt": 900,
            "pricing_calculation_executed": False,
            "generic_work_432_prohibited": True,
            "family_fallback_prohibited": True,
            "fuzzy_fallback_prohibited": True,
            "similar_relay_fallback_prohibited": True,
        },
        "append_only": True,
        "scope_expansion": False,
    }
    value["calculator_input_format"]["row_drafts"].extend(
        builder._expected_appended_rows()
    )
    value["completion"]["status"] = builder.TECHNICAL_STATUS
    value["completion"]["scope"] = {
        "component_groups": 35,
        "rows": "116/116",
        "cabinet_groups": "15/15",
        "duplicate_component_membership": 0,
        "duplicate_cabinet_membership": 0,
        "scope_expansion": False,
    }
    return cast(dict[str, Any], value)


def synthetic_contract(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> tuple[Any, Any, dict[str, Any], dict[str, Any], dict[str, Any]]:
    tmp_path.mkdir(parents=True, exist_ok=True)
    parent_data = synthetic_parent_profile()
    parent = tmp_path / "parent.json"
    technical = tmp_path / "technical.json"
    decision = tmp_path / "decision.json"
    workbook = tmp_path / "prices.xlsx"
    write_json(parent, parent_data)
    parent_sha = sha256(parent)
    decision_data = synthetic_decision(parent, parent_sha)
    write_json(decision, decision_data)
    decision_sha = sha256(decision)
    technical_data = synthetic_technical(decision, decision_sha)
    write_json(technical, technical_data)
    technical_sha = sha256(technical)
    make_workbook(workbook)
    workbook_sha = sha256(workbook)
    for name, value in (
        ("PARENT_PROFILE", parent),
        ("PARENT_PROFILE_SHA256", parent_sha),
        ("TECHNICAL_SUCCESSOR", technical),
        ("TECHNICAL_SUCCESSOR_SHA256", technical_sha),
        ("HUMAN_DECISION", decision),
        ("HUMAN_DECISION_SHA256", decision_sha),
        ("PRICING_WORKBOOK", workbook),
        ("PRICING_WORKBOOK_SHA256", workbook_sha),
    ):
        monkeypatch.setattr(builder, name, value)
    paths = builder.InputPaths(parent, technical, decision, workbook)
    shas = builder.ExpectedShas(parent_sha, technical_sha, decision_sha, workbook_sha)
    return paths, shas, parent_data, technical_data, decision_data


DELETE = object()


def mutate_path(value: Any, path: tuple[Any, ...], replacement: Any) -> None:
    target = value
    for key in path[:-1]:
        target = target[key]
    if replacement is DELETE:
        del target[path[-1]]
    else:
        target[path[-1]] = replacement


PAYLOAD_CLOSED_ENVELOPE_MUTATIONS = [
    ("application_applied", ("application_status",), "APPLIED"),
    ("application_missing", ("application_status",), DELETE),
    ("application_null", ("application_status",), None),
    ("application_other_string", ("application_status",), "false"),
    (
        "safety_flag",
        ("safety_flags", "calculator_run_authorized"),
        True,
    ),
    ("non_approvals", ("non_approvals", "project_total_approved"), True),
    ("status", ("status",), "APPLIED"),
    ("authority", ("authority", "no_scope_expansion"), False),
    ("immutable", ("immutable_state", "immutable"), False),
    ("scope_expansion", ("scope_expansion",), True),
    (
        "unrelated_scope",
        ("current_completed_technical_scope", "technical_scope_status"),
        "changed",
    ),
    (
        "authoritative_binding",
        ("authoritative_inputs", 0, "sha256"),
        "0" * 64,
    ),
    ("validation_summary", ("validation_summary", "duplicate_json_keys"), "FAIL"),
    ("extra_root", ("unexpected_root",), True),
    (
        "extra_scope",
        ("current_completed_technical_scope", "unexpected_scope"),
        True,
    ),
    (
        "target_approved_price",
        (
            "current_completed_technical_scope",
            "pricing_positions",
            8,
            "approved_unit_price_kzt",
        ),
        53763,
    ),
    (
        "target_approved_status",
        (
            "current_completed_technical_scope",
            "pricing_positions",
            8,
            "approved_unit_price_decision_status",
        ),
        "APPROVED",
    ),
    (
        "non_target_approved_price",
        (
            "current_completed_technical_scope",
            "pricing_positions",
            0,
            "approved_unit_price_kzt",
        ),
        53763,
    ),
    (
        "non_target_approved_status",
        (
            "current_completed_technical_scope",
            "pricing_positions",
            0,
            "approved_unit_price_decision_status",
        ),
        "APPROVED",
    ),
]


def test_synthetic_valid_successor_has_exact_controlled_replacement(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    paths, shas, parent, _, _ = synthetic_contract(tmp_path, monkeypatch)
    loaded = builder.load_and_validate_inputs(paths, shas)
    payload = builder.build_successor_payload(loaded)
    builder.validate_successor_payload(payload, loaded)
    scope = payload["current_completed_technical_scope"]
    assert scope["coverage"] == {
        "technical_cabinet_groups": 15,
        "section_aware_pricing_positions": 55,
        "physical_cabinets": 137,
        "composition_fingerprints": 11,
    }
    assert len(scope["cabinet_groups"]) == 15
    assert len(scope["pricing_positions"]) == 55
    assert (
        sum(item["physical_multiplicity"] for item in scope["pricing_positions"]) == 137
    )
    assert len(scope["composition_fingerprints"]) == 11
    assert scope["cabinet_groups"][2]["row_draft_ids"][-4:] == [
        "ROW-DRAFT-0113",
        "ROW-DRAFT-0114",
        "ROW-DRAFT-0115",
        "ROW-DRAFT-0116",
    ]
    assert (
        scope["cabinet_groups"][:2]
        == parent["current_completed_technical_scope"]["cabinet_groups"][:2]
    )
    assert (
        scope["cabinet_groups"][3:]
        == parent["current_completed_technical_scope"]["cabinet_groups"][3:]
    )
    assert builder.OLD_FINGERPRINT not in {
        item["fingerprint_sha256"] for item in scope["composition_fingerprints"]
    }
    merged = next(
        item
        for item in scope["composition_fingerprints"]
        if item["fingerprint_sha256"] == builder.NEW_FINGERPRINT
    )
    assert merged["pricing_position_ids"] == [
        "PRICE-POSITION-009",
        "PRICE-POSITION-023",
        "PRICE-POSITION-035",
        "PRICE-POSITION-047",
        "PRICE-POSITION-052",
        "PRICE-POSITION-053",
        "PRICE-POSITION-054",
        "PRICE-POSITION-055",
    ]
    assert payload["shu_t2_rt820_pricing_profile_successor"][
        "preliminary_not_approved_invariants"
    ] == {
        "status": "NOT_CALCULATED_NOT_APPROVED",
        "X_cabinet_base_kzt": 6936,
        "G_material_kzt": 20450,
        "H_work_kzt": 1764,
        "formula_base_kzt": 33240,
        "raw_unit_candidate_kzt": "53762.72702586206896551724138",
        "unit_candidate_kzt": 53763,
        "four_position_candidate_kzt": 215052,
        "delta_from_prior_checked_candidate_kzt": 122276,
        "preliminary_project_candidate_kzt": 11963792,
        "approved": False,
        "applied": False,
    }


@pytest.mark.parametrize(
    ("mutation", "message"),
    [
        (lambda value: value.update({"project_id": "other"}), "project"),
        (
            lambda value: value["current_completed_technical_scope"][
                "pricing_positions"
            ][0].update({"section": "99"}),
            "non-target pricing position",
        ),
        (
            lambda value: value["current_completed_technical_scope"]["cabinet_groups"][
                0
            ].update({"cabinet_base_kzt": 1}),
            "non-target cabinet group",
        ),
        (
            lambda value: value["current_completed_technical_scope"][
                "pricing_positions"
            ][8].update({"approved_unit_price_kzt": 53763}),
            "identity/state",
        ),
        (
            lambda value: value["current_completed_technical_scope"][
                "composition_fingerprints"
            ].append(
                copy.deepcopy(
                    value["current_completed_technical_scope"][
                        "composition_fingerprints"
                    ][-1]
                )
            ),
            "inventory",
        ),
    ],
)
def test_successor_validation_rejects_scope_price_and_duplicate_drift(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
    mutation: Any,
    message: str,
) -> None:
    paths, shas, _, _, _ = synthetic_contract(tmp_path, monkeypatch)
    loaded = builder.load_and_validate_inputs(paths, shas)
    payload = builder.build_successor_payload(loaded)
    mutation(payload)
    with pytest.raises(builder.ContractError, match=message):
        builder.validate_successor_payload(payload, loaded)


@pytest.mark.parametrize(
    ("_name", "path", "replacement"), PAYLOAD_CLOSED_ENVELOPE_MUTATIONS
)
def test_successor_validation_rejects_every_closed_envelope_drift(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
    _name: str,
    path: tuple[Any, ...],
    replacement: Any,
) -> None:
    paths, shas, _, _, _ = synthetic_contract(tmp_path, monkeypatch)
    loaded = builder.load_and_validate_inputs(paths, shas)
    payload = builder.build_successor_payload(loaded)
    mutate_path(payload, path, replacement)
    with pytest.raises(builder.ContractError):
        builder.validate_successor_payload(payload, loaded)


@pytest.mark.parametrize("flag", list(builder.EXPECTED_SAFETY_FLAGS))
@pytest.mark.parametrize("mode", ["wrong_boolean", "non_boolean", "missing"])
def test_successor_validation_rejects_each_safety_flag_drift(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
    flag: str,
    mode: str,
) -> None:
    paths, shas, _, _, _ = synthetic_contract(tmp_path, monkeypatch)
    loaded = builder.load_and_validate_inputs(paths, shas)
    payload = builder.build_successor_payload(loaded)
    flags = payload["safety_flags"]
    if mode == "wrong_boolean":
        flags[flag] = not builder.EXPECTED_SAFETY_FLAGS[flag]
    elif mode == "non_boolean":
        flags[flag] = "false"
    else:
        del flags[flag]
    with pytest.raises(builder.ContractError, match="safety flags"):
        builder.validate_successor_payload(payload, loaded)


@pytest.mark.parametrize(
    ("name", "mutation"),
    [
        (
            "missing_evidence",
            lambda row, rows: row["source_component_evidence_ids"].pop(),
        ),
        (
            "duplicate_evidence",
            lambda row, rows: row["source_component_evidence_ids"].__setitem__(
                1, row["source_component_evidence_ids"][0]
            ),
        ),
        (
            "swapped_evidence",
            lambda row, rows: (
                row["source_component_evidence_ids"].__setitem__(
                    1, rows[1]["source_component_evidence_ids"][1]
                ),
                rows[1]["source_component_evidence_ids"].__setitem__(1, "COMP-034"),
            ),
        ),
        (
            "wrong_technical_position",
            lambda row, rows: row["source_quantity"].update(
                {"technical_position_id": "TFE-041"}
            ),
        ),
        (
            "wrong_pricing_position",
            lambda row, rows: row["source_quantity"].update(
                {"pricing_position_id": "PRICE-POSITION-023"}
            ),
        ),
        (
            "wrong_section",
            lambda row, rows: row["source_quantity"].update({"section": "12"}),
        ),
        (
            "wrong_mapping_status",
            lambda row, rows: row.update({"mapping_status": "WRONG"}),
        ),
        (
            "wrong_component_label",
            lambda row, rows: row.update({"component_label": "wrong"}),
        ),
        ("extra_row_field", lambda row, rows: row.update({"extra": True})),
    ],
)
def test_technical_appended_rows_reject_membership_and_envelope_drift(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
    name: str,
    mutation: Any,
) -> None:
    _paths, shas, _, technical, _ = synthetic_contract(tmp_path, monkeypatch)
    rows = technical["calculator_input_format"]["row_drafts"][-4:]
    mutation(rows[0], rows)
    with pytest.raises(builder.ContractError, match="appended row"):
        builder.validate_technical(technical, shas.human_decision)


@pytest.mark.parametrize(
    "field",
    [
        "decision_id",
        "decision_kind",
        "quantity_per_individual_cabinet",
        "physical_multiplicity",
        "applies_once_per_cabinet",
        "multiply_by_member_count",
        "scope_expansion",
    ],
)
def test_technical_appended_rows_reject_each_quantity_and_multiplicity_drift(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
    field: str,
) -> None:
    _paths, shas, _, technical, _ = synthetic_contract(tmp_path, monkeypatch)
    row = technical["calculator_input_format"]["row_drafts"][-4]
    row["source_quantity"][field] = None
    with pytest.raises(builder.ContractError, match="appended row"):
        builder.validate_technical(technical, shas.human_decision)


@pytest.mark.parametrize(
    "field",
    [
        "manufacturer",
        "product",
        "manufacturer_article",
        "supply_form",
        "module_width_din",
        "TST05_evidence_included_as_provenance_only",
        "TST05_separate_component_row",
    ],
)
def test_technical_appended_rows_reject_each_signature_field_drift(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
    field: str,
) -> None:
    _paths, shas, _, technical, _ = synthetic_contract(tmp_path, monkeypatch)
    row = technical["calculator_input_format"]["row_drafts"][-4]
    row["approved_signature"][field] = None
    with pytest.raises(builder.ContractError, match="appended row"):
        builder.validate_technical(technical, shas.human_decision)


@pytest.mark.parametrize(
    "role",
    ["parent_profile", "technical_successor", "human_decision", "pricing_workbook"],
)
def test_each_path_and_sha_binding_is_exact(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch, role: str
) -> None:
    paths, shas, _, _, _ = synthetic_contract(tmp_path, monkeypatch)
    changed_paths = builder.InputPaths(
        *[
            tmp_path / "wrong" if field == role else getattr(paths, field)
            for field in paths.__dataclass_fields__
        ]
    )
    with pytest.raises(builder.ContractError, match="path binding"):
        builder.load_and_validate_inputs(changed_paths, shas)
    changed_shas = builder.ExpectedShas(
        *[
            "0" * 64 if field == role else getattr(shas, field)
            for field in shas.__dataclass_fields__
        ]
    )
    with pytest.raises(builder.ContractError, match="expected SHA binding"):
        builder.load_and_validate_inputs(paths, changed_shas)


def test_duplicate_json_key_and_workbook_source_drift_fail(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    paths, shas, _, _, _ = synthetic_contract(tmp_path, monkeypatch)
    paths.parent_profile.write_text('{"a":1,"a":2}', encoding="utf-8")
    monkeypatch.setattr(builder, "PARENT_PROFILE_SHA256", sha256(paths.parent_profile))
    duplicate_shas = builder.ExpectedShas(
        sha256(paths.parent_profile),
        shas.technical_successor,
        shas.human_decision,
        shas.pricing_workbook,
    )
    with pytest.raises(builder.ContractError, match="duplicate key"):
        builder.load_and_validate_inputs(paths, duplicate_shas)
    paths, shas, _, _, _ = synthetic_contract(tmp_path / "second", monkeypatch)
    make_workbook(paths.pricing_workbook, work=432)
    workbook_sha = sha256(paths.pricing_workbook)
    monkeypatch.setattr(builder, "PRICING_WORKBOOK_SHA256", workbook_sha)
    drifted = builder.ExpectedShas(
        shas.parent_profile,
        shas.technical_successor,
        shas.human_decision,
        workbook_sha,
    )
    with pytest.raises(builder.ContractError, match="A19:C19"):
        builder.load_and_validate_inputs(paths, drifted)


def test_synthetic_publication_is_atomic_and_no_overwrite(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    paths, shas, _, _, _ = synthetic_contract(tmp_path / "inputs", monkeypatch)
    output = tmp_path / "case" / builder.OUTPUT_FILENAME
    result = builder.publish_successor(paths, shas, output)
    assert output.read_bytes() == result.encoded
    assert sha256(output) == result.sha256
    assert set(output.parent.iterdir()) == {output}
    with pytest.raises(builder.ContractError, match="already exists"):
        builder.publish_successor(paths, shas, output)


def test_wrong_authorization_fails_before_publication(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    paths, shas, _, _, _ = synthetic_contract(tmp_path, monkeypatch)
    output = tmp_path / "case" / builder.OUTPUT_FILENAME
    argv = [
        "--parent-pricing-profile",
        str(paths.parent_profile),
        "--parent-pricing-profile-sha256",
        shas.parent_profile,
        "--technical-successor",
        str(paths.technical_successor),
        "--technical-successor-sha256",
        shas.technical_successor,
        "--human-decision",
        str(paths.human_decision),
        "--human-decision-sha256",
        shas.human_decision,
        "--pricing-workbook",
        str(paths.pricing_workbook),
        "--pricing-workbook-sha256",
        shas.pricing_workbook,
        "--output",
        str(output),
        "--authorization",
        "WRONG",
    ]
    with pytest.raises(builder.ContractError, match="authorization"):
        builder.main(argv)
    assert not output.parent.exists()


def test_toctou_and_post_link_validation_failure_roll_back_everything(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    paths, shas, _, _, _ = synthetic_contract(tmp_path / "first", monkeypatch)
    output = tmp_path / "case1" / builder.OUTPUT_FILENAME
    original_recheck = builder._recheck_inputs

    def fail_recheck(loaded: Any, phase: str) -> None:
        if phase == "pre-publication":
            raise builder.ContractError("synthetic TOCTOU")
        original_recheck(loaded, phase)

    monkeypatch.setattr(builder, "_recheck_inputs", fail_recheck)
    with pytest.raises(builder.ContractError, match="synthetic TOCTOU"):
        builder.publish_successor(paths, shas, output)
    assert not output.parent.exists()

    monkeypatch.setattr(builder, "_recheck_inputs", original_recheck)
    paths, shas, _, _, _ = synthetic_contract(tmp_path / "second", monkeypatch)
    output = tmp_path / "case2" / builder.OUTPUT_FILENAME
    link_completed = False
    original_link = builder.os.link
    original_validate = builder.validate_successor_payload

    def wrapped_link(source: Any, target: Any) -> None:
        nonlocal link_completed
        original_link(source, target)
        link_completed = True

    def fail_after_link(payload: Any, loaded: Any) -> None:
        original_validate(payload, loaded)
        if link_completed:
            raise builder.ContractError("synthetic final validation failure")

    monkeypatch.setattr(builder.os, "link", wrapped_link)
    monkeypatch.setattr(builder, "validate_successor_payload", fail_after_link)
    with pytest.raises(
        builder.ContractError, match="synthetic final validation failure"
    ):
        builder.publish_successor(paths, shas, output)
    assert link_completed is True
    assert not output.exists()
    assert not output.parent.exists()


@pytest.mark.parametrize(
    ("name", "path", "replacement"), PAYLOAD_CLOSED_ENVELOPE_MUTATIONS
)
def test_each_closed_envelope_post_link_mutation_rolls_back_everything(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
    name: str,
    path: tuple[Any, ...],
    replacement: Any,
) -> None:
    paths, shas, _, _, _ = synthetic_contract(tmp_path / "inputs", monkeypatch)
    output = tmp_path / "case" / builder.OUTPUT_FILENAME
    success_marker = output.with_suffix(".success")
    original_link = builder.os.link
    original_validate = builder.validate_successor_payload
    link_completed = False

    def wrapped_link(source: Any, target: Any) -> None:
        nonlocal link_completed
        original_link(source, target)
        link_completed = True

    def reject_mutated_final(payload: Any, loaded: Any) -> None:
        if not link_completed:
            original_validate(payload, loaded)
            return
        mutate_path(payload, path, replacement)
        try:
            original_validate(payload, loaded)
        except builder.ContractError as exc:
            raise builder.ContractError(
                f"synthetic post-link closed-envelope rejection: {name}"
            ) from exc
        raise AssertionError("post-link mutation unexpectedly passed validation")

    monkeypatch.setattr(builder.os, "link", wrapped_link)
    monkeypatch.setattr(builder, "validate_successor_payload", reject_mutated_final)
    with pytest.raises(
        builder.ContractError,
        match=f"synthetic post-link closed-envelope rejection: {name}",
    ):
        builder.publish_successor(paths, shas, output)
    assert link_completed is True
    assert not output.exists()
    assert not success_marker.exists()
    assert not output.parent.exists()


def test_foreign_replacement_is_preserved_and_cleanup_cannot_report_success(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    paths, shas, _, _, _ = synthetic_contract(tmp_path / "inputs", monkeypatch)
    output = tmp_path / "case" / builder.OUTPUT_FILENAME
    original_validate = builder.validate_successor_payload
    link_completed = False
    original_link = builder.os.link

    def wrapped_link(source: Any, target: Any) -> None:
        nonlocal link_completed
        original_link(source, target)
        link_completed = True

    def replace_after_link(payload: Any, loaded: Any) -> None:
        original_validate(payload, loaded)
        if link_completed and output.exists():
            output.unlink()
            output.write_text("foreign", encoding="utf-8")
            raise builder.ContractError("synthetic foreign replacement")

    monkeypatch.setattr(builder.os, "link", wrapped_link)
    monkeypatch.setattr(builder, "validate_successor_payload", replace_after_link)
    with pytest.raises(
        builder.ContractError, match="foreign final replacement preserved"
    ):
        builder.publish_successor(paths, shas, output)
    assert output.read_text(encoding="utf-8") == "foreign"


def test_real_inputs_read_only_validation_does_not_publish() -> None:
    if not all(
        path.is_file()
        for path in (
            builder.PARENT_PROFILE,
            builder.TECHNICAL_SUCCESSOR,
            builder.HUMAN_DECISION,
            builder.PRICING_WORKBOOK,
        )
    ):
        pytest.skip("case-scoped real inputs are unavailable in this environment")
    result = builder.validate_real_inputs_read_only(
        builder.InputPaths(
            builder.PARENT_PROFILE,
            builder.TECHNICAL_SUCCESSOR,
            builder.HUMAN_DECISION,
            builder.PRICING_WORKBOOK,
        ),
        builder.ExpectedShas(
            builder.PARENT_PROFILE_SHA256,
            builder.TECHNICAL_SUCCESSOR_SHA256,
            builder.HUMAN_DECISION_SHA256,
            builder.PRICING_WORKBOOK_SHA256,
        ),
    )
    assert result == {
        "status": "PASS",
        "coverage": builder.EXPECTED_COVERAGE,
        "fingerprint": builder.NEW_FINGERPRINT,
        "publication_called": False,
        "price_approved": False,
    }


def test_help_and_source_contain_exact_safety_contract(capsys: Any) -> None:
    with pytest.raises(SystemExit, match="0"):
        builder.parse_args(["--help"])
    help_text = capsys.readouterr().out
    for option in (
        "--parent-pricing-profile",
        "--technical-successor",
        "--human-decision",
        "--pricing-workbook",
        "--output",
        "--authorization",
    ):
        assert option in help_text
    source = SCRIPT.read_text(encoding="utf-8")
    assert "os.link(staging, output)" in source
    assert (
        "IGOR_SHU_T2_RT820_PRICING_PROFILE_SUCCESSOR_PUBLICATION_AUTHORIZED" in source
    )
    assert "git " not in source.casefold()
