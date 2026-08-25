import csv
import hashlib
import importlib.util
import py_compile
import subprocess
import sys
from decimal import Decimal
from pathlib import Path
from types import ModuleType
from typing import Any, cast

import pytest
from openpyxl import Workbook, load_workbook  # type: ignore[import-untyped]

PROJECT_ROOT = Path(__file__).resolve().parents[1]
SCRIPT = PROJECT_ROOT / "scripts" / "calc_quote_price_draft.py"
TECHNICAL_WORKFLOW = PROJECT_ROOT / "scripts" / "make_quote_capacity100_checked.ps1"
COMMERCIAL_WORKFLOW = (
    PROJECT_ROOT / "scripts" / "make_quote_capacity100_commercial_checked.ps1"
)


def load_script_module(module_name: str, path: Path) -> ModuleType:
    spec = importlib.util.spec_from_file_location(module_name, path)
    assert spec is not None
    assert spec.loader is not None
    module = importlib.util.module_from_spec(spec)
    sys.modules[spec.name] = module
    spec.loader.exec_module(module)
    return module


calculator = cast(
    Any,
    load_script_module("calc_quote_price_draft_for_test", SCRIPT),
)


def confirmed_rows() -> list[list[str]]:
    return [
        [
            "РУ-АВР / ЩРН-24",
            "CAB-KRN-24",
            "1.20",
            "EKF-VA47-29-1P",
            "4",
            "modular_1p",
        ],
        [
            "РУ-АВР / ЩРН-24",
            "CAB-KRN-24",
            "1.20",
            "EKF-VA47-29-3P",
            "3",
            "modular_3p",
        ],
        [
            "РУ-АВР / ЩРН-24",
            "CAB-KRN-24",
            "1.20",
            "EKF-RN-47",
            "1",
            "modular_1p",
        ],
    ]


def write_csv(path: Path, rows: list[list[str]] | None = None) -> None:
    with path.open("w", encoding="utf-8", newline="") as csv_file:
        writer = csv.writer(csv_file, delimiter=";", lineterminator="\n")
        writer.writerow(calculator.REQUIRED_COLUMNS)
        writer.writerows(confirmed_rows() if rows is None else rows)


def write_technical_csv(path: Path, rows: list[list[str]]) -> None:
    with path.open("w", encoding="utf-8", newline="") as csv_file:
        writer = csv.writer(csv_file, delimiter=";", lineterminator="\n")
        writer.writerow(calculator.TECHNICAL_COLUMNS)
        writer.writerows(rows)


def write_shu_t2_bound_csv(path: Path, rows: list[list[str]]) -> None:
    with path.open("w", encoding="utf-8", newline="") as csv_file:
        writer = csv.writer(csv_file, delimiter=";", lineterminator="\n")
        writer.writerow(calculator.SHU_T2_TECHNICAL_COLUMNS)
        writer.writerows(rows)


def write_workbook(
    path: Path,
    *,
    include_component: bool = True,
    include_cabinet: bool = True,
) -> None:
    workbook = Workbook()
    krn = workbook.active
    krn.title = "КРН"
    krn.append(["Наименование", "Материал", "Работа"])
    if include_component:
        krn.append(["ВА47 1 полюсный", 700, 216])
        krn.append(["ВА47 3 полюсный до 63А", 2200, 540])
        krn.append(["независимый расцепитель для ВА47 РН47", 7500, 216])
    if include_cabinet:
        krn["L8"] = "Корпус КРН-24 395х330х100"
        krn["M8"] = 7985

    forbidden = workbook.create_sheet("Прайс")
    forbidden.append(["ВА47 1 полюсный", 1, 1])
    forbidden.append(["ВА47 3 полюсный до 63А", 1, 1])
    forbidden.append(["независимый расцепитель для ВА47 РН47", 1, 1])
    forbidden["L8"] = "Корпус КРН-24 395х330х100"
    forbidden["M8"] = 1

    workbook.save(path)
    workbook.close()


def write_approved_workbook(path: Path) -> None:
    workbook = Workbook()
    krn = workbook.active
    krn.title = "КРН"
    krn["A5"] = "УЗО АД-32 1Р+N до 63А EKF"
    krn["B5"] = 4100
    krn["C5"] = 432
    krn["L9"] = "Корпус КРН-36 540х330х100"
    krn["M9"] = 9405
    krn["A14"] = "ВН-32 3Р 16-25-40-63-80-100А"
    krn["B14"] = 2750
    krn["C14"] = 540
    krn["A28"] = "УЗО АД-32 1Р+N до 63А 100мА-300мА EKF "
    krn["B28"] = 8000
    krn["C28"] = 432
    krn["A29"] = "УЗО АД-32 3Р+N до 63А 100мА-300мА EKF"
    krn["B29"] = 10000
    krn["C29"] = 540

    shr = workbook.create_sheet("ЩР")
    shr["A8"] = "ВА55/57/59, АМ1  3 полюсные от 16 до 63А"
    shr["B8"] = 13000
    shr["C8"] = 1800
    shr["L8"] = "800х600х250"
    shr["M8"] = 21336

    forbidden = workbook.create_sheet("Прайс")
    forbidden["A5"] = "УЗО АД-32 1Р+N до 63А EKF"
    forbidden["B5"] = 1
    forbidden["C5"] = 1
    workbook.save(path)
    workbook.close()


def technical_row(
    *,
    product_name: str = "TEST-PANEL",
    cabinet_code: str,
    component_code: str,
    install_type: str,
    component_label: str,
    cabinet_label: str,
) -> list[str]:
    return [
        product_name,
        cabinet_code,
        "1.20",
        component_code,
        "1",
        install_type,
        component_label,
        cabinet_label,
    ]


def calculate_mapping_case(
    tmp_path: Path,
    *,
    case_name: str,
    component_code: str,
    install_type: str,
    component_label: str,
    workbook_changes: dict[str, Any] | None = None,
) -> Any:
    workbook_path = tmp_path / f"{case_name}.xlsx"
    write_approved_workbook(workbook_path)
    if workbook_changes:
        workbook = load_workbook(workbook_path)
        worksheet = workbook["КРН"]
        for coordinate, value in workbook_changes.items():
            worksheet[coordinate] = value
        workbook.save(workbook_path)
        workbook.close()
    csv_path = tmp_path / f"{case_name}.csv"
    write_technical_csv(
        csv_path,
        [
            technical_row(
                cabinet_code="КРН-36",
                component_code=component_code,
                install_type=install_type,
                component_label=component_label,
                cabinet_label="КРН-36, 540×330×100 мм, металл",
            )
        ],
    )
    return calculate(workbook_path, csv_path)


def test_approved_requests_003_009_016_remain_pass(tmp_path: Path) -> None:
    workbook_path = tmp_path / "approved.xlsx"
    write_approved_workbook(workbook_path)
    cases = (
        (
            "RAW-VA88-32",
            "mccb_up_to_100a",
            "CHINT, автоматический выключатель 3P 63А",
            "ПР",
            "ПР 800×600×250 мм, металл",
            13000,
            1800,
        ),
        (
            "RAW-AVDT32",
            "diff_1p_n",
            "CHINT, АВДТ 2P C16/30мА",
            "КРН-36",
            "КРН-36, 540×330×100 мм, металл",
            4100,
            432,
        ),
        (
            "ANOTHER-RAW-CODE",
            "load_switch_3p",
            "CHINT, выключатель нагрузки 3P 32А",
            "КРН-36",
            "КРН-36, 540×330×100 мм, металл",
            2750,
            540,
        ),
    )

    for index, case in enumerate(cases):
        code, install_type, label, cabinet_code, cabinet_label, material, work = case
        csv_path = tmp_path / f"technical-{index}.csv"
        write_technical_csv(
            csv_path,
            [
                technical_row(
                    cabinet_code=cabinet_code,
                    component_code=code,
                    install_type=install_type,
                    component_label=label,
                    cabinet_label=cabinet_label,
                )
            ],
        )
        result = calculate(workbook_path, csv_path)
        assert result.status == "PASS"
        assert result.component_material_total == material
        assert result.work_total == work


def test_breaking_capacity_variants_are_parsed_exactly() -> None:
    for suffix in ("6кА", "6 кА", "6kA", "6 kA"):
        signature = calculator.parse_component_signature(
            f"CHINT, АВДТ 2P C16/30мА {suffix}",
            "diff_1p_n",
        )

        assert signature is not None
        assert signature.breaking_capacity_ka == Decimal("6")
        assert calculator.resolve_component_mapping(signature) is None


def test_explicit_6ka_without_approved_signature_fails_closed(
    tmp_path: Path,
    monkeypatch: Any,
) -> None:
    workbook_path = tmp_path / "approved.xlsx"
    write_approved_workbook(workbook_path)
    monkeypatch.setitem(
        calculator.COMPONENT_DEFINITIONS,
        "TEST-RCBO",
        calculator.ComponentDefinition(
            workbook_label="УЗО АД-32 1Р+N до 63А EKF",
            install_type="diff_1p_n",
        ),
    )
    csv_path = tmp_path / "composition.csv"
    write_technical_csv(
        csv_path,
        [
            technical_row(
                cabinet_code="CAB-KRN-24",
                component_code="TEST-RCBO",
                install_type="diff_1p_n",
                component_label="АВДТ 2P C16/30мА 6кА",
                cabinet_label="КРН-24, 395×330×100 мм, металл",
            )
        ],
    )

    result = calculate(workbook_path, csv_path)

    assert result.status == "FAIL"
    assert result.total_preliminary_price is None
    assert any(
        "unknown or ambiguous technical component mapping" in flag
        for flag in result.red_flags
    )


def test_explicit_6ka_blocks_fallback_when_full_signature_is_unparsed(
    tmp_path: Path,
    monkeypatch: Any,
) -> None:
    workbook_path = tmp_path / "approved.xlsx"
    write_approved_workbook(workbook_path)
    monkeypatch.setitem(
        calculator.COMPONENT_DEFINITIONS,
        "TEST-RCBO",
        calculator.ComponentDefinition(
            workbook_label="УЗО АД-32 1Р+N до 63А EKF",
            install_type="diff_1p_n",
        ),
    )
    csv_path = tmp_path / "composition.csv"
    write_technical_csv(
        csv_path,
        [
            technical_row(
                cabinet_code="CAB-KRN-24",
                component_code="TEST-RCBO",
                install_type="diff_1p_n",
                component_label="неразобранный аппарат 6 kA",
                cabinet_label="КРН-24, 395×330×100 мм, металл",
            )
        ],
    )

    result = calculate(workbook_path, csv_path)

    assert (
        calculator.parse_component_signature(
            "неразобранный аппарат 6 kA",
            "diff_1p_n",
        )
        is None
    )
    assert result.status == "FAIL"
    assert result.total_preliminary_price is None
    assert any(
        "unknown or ambiguous technical component mapping" in flag
        for flag in result.red_flags
    )


def test_component_mapping_005_like_ad32_fallback_fails_closed(
    tmp_path: Path,
) -> None:
    workbook_path = tmp_path / "approved.xlsx"
    write_approved_workbook(workbook_path)
    csv_path = tmp_path / "composition.csv"
    write_technical_csv(
        csv_path,
        [
            technical_row(
                cabinet_code="CAB-KRN-24",
                component_code="EKF-AD32-1P-N",
                install_type="diff_1p_n",
                component_label="АВДТ 4P C16/100мА",
                cabinet_label="КРН-24, 395×330×100 мм, металл",
            )
        ],
    )

    result = calculate(workbook_path, csv_path)

    assert result.status == "FAIL"
    assert result.total_preliminary_price is None
    assert any(
        "unknown or ambiguous technical component mapping" in flag
        for flag in result.red_flags
    )


def test_component_mapping_012_passes_for_approved_code_and_6ka_variants(
    tmp_path: Path,
) -> None:
    labels = (
        "АД12, 2P, C16, 30мА, 6кА",
        "АВДТ 2P C16/30мА 6 кА",
        "АВДТ 2P C16/30мА 6kA",
        "АВДТ 2P C16/30мА 6 kA",
    )
    for index, label in enumerate(labels):
        result = calculate_mapping_case(
            tmp_path,
            case_name=f"mapping-012-pass-{index}",
            component_code="EKF-AD32-1P-N",
            install_type="diff_1p_n",
            component_label=label,
        )

        assert result.status == "PASS"
        assert result.component_material_total == 4100
        assert result.work_total == 432


def test_component_mapping_012_rejects_wrong_code_or_signature(
    tmp_path: Path,
) -> None:
    cases = (
        ("WRONG-CODE", "АВДТ 2P C16/30мА 6кА"),
        ("EKF-AD32-1P-N", "АВДТ 2P C16/30мА 10кА"),
        ("EKF-AD32-1P-N", "АВДТ 2P C16/100мА 6кА"),
        ("EKF-AD32-1P-N", "АВДТ 2P C20/30мА 6кА"),
        ("EKF-AD32-1P-N", "АВДТ 4P C16/30мА 6кА"),
    )
    for index, (code, label) in enumerate(cases):
        result = calculate_mapping_case(
            tmp_path,
            case_name=f"mapping-012-fail-signature-{index}",
            component_code=code,
            install_type="diff_1p_n",
            component_label=label,
        )

        assert result.status == "FAIL"
        assert result.total_preliminary_price is None
        assert any("unknown or ambiguous" in flag for flag in result.red_flags)


def test_component_mapping_012_rejects_workbook_drift(tmp_path: Path) -> None:
    changes: tuple[dict[str, Any], ...] = (
        {"A5": "УЗО АД-32 1Р+N до 63А EKF "},
        {"B5": 4101},
        {"C5": 433},
    )
    for index, workbook_changes in enumerate(changes):
        result = calculate_mapping_case(
            tmp_path,
            case_name=f"mapping-012-fail-workbook-{index}",
            component_code="EKF-AD32-1P-N",
            install_type="diff_1p_n",
            component_label="АД12, 2P, C16, 30мА, 6кА",
            workbook_changes=workbook_changes,
        )

        assert result.status == "FAIL"
        assert result.total_preliminary_price is None
        assert any("approved component mapping" in flag for flag in result.red_flags)


def test_ad12_price_mapping_passes_only_exact_approved_identity(
    tmp_path: Path,
) -> None:
    label = "Дифференциальный автомат АД12 2Р 16А, 30мА — 1шт., C, 4,5кА"
    signature = calculator.parse_component_signature(label, "diff_1p_n")

    assert signature is not None
    mapping = calculator.resolve_component_mapping(
        signature,
        "EKF-AD12-1P-N-C16-30MA-4P5KA",
    )
    assert mapping is not None
    assert (mapping.sheet_name, mapping.row) == ("КРН", 5)
    assert mapping.component_code == "EKF-AD12-1P-N-C16-30MA-4P5KA"
    assert mapping.strict_raw_label is True

    result = calculate_mapping_case(
        tmp_path,
        case_name="ad12-price-mapping-pass",
        component_code="EKF-AD12-1P-N-C16-30MA-4P5KA",
        install_type="diff_1p_n",
        component_label=label,
    )

    assert result.status == "PASS"
    assert result.component_material_total == 4100
    assert result.work_total == 432


def test_ad12_price_mapping_rejects_wrong_code_install_type_or_identity(
    tmp_path: Path,
) -> None:
    approved_code = "EKF-AD12-1P-N-C16-30MA-4P5KA"
    cases = (
        (
            "WRONG-CODE",
            "diff_1p_n",
            "Дифференциальный автомат АД12 2Р 16А, 30мА, C, 4,5кА",
        ),
        (
            approved_code,
            "diff_3p_4p",
            "Дифференциальный автомат АД12 2Р 16А, 30мА, C, 4,5кА",
        ),
        (
            approved_code,
            "diff_1p_n",
            "Дифференциальный автомат АД12 2Р 20А, 30мА, C, 4,5кА",
        ),
        (
            approved_code,
            "diff_1p_n",
            "Дифференциальный автомат АД12 2Р 16А, 100мА, C, 4,5кА",
        ),
        (
            approved_code,
            "diff_1p_n",
            "Дифференциальный автомат АД12 4Р 16А, 30мА, C, 4,5кА",
        ),
    )
    for index, (code, install_type, label) in enumerate(cases):
        result = calculate_mapping_case(
            tmp_path,
            case_name=f"ad12-price-mapping-fail-identity-{index}",
            component_code=code,
            install_type=install_type,
            component_label=label,
        )

        assert result.status == "FAIL"
        assert result.total_preliminary_price is None
        assert any("unknown or ambiguous" in flag for flag in result.red_flags)


def test_ad12_price_mapping_rejects_breaking_capacity_or_family_fallback(
    tmp_path: Path,
) -> None:
    code = "EKF-AD12-1P-N-C16-30MA-4P5KA"
    labels = (
        "Дифференциальный автомат АД12 2Р 16А, 30мА, C, 6кА",
        "Дифференциальный автомат АД12 2Р 16А, 30мА, C, 10кА",
        "Дифференциальный автомат АД12 2Р 16А, 30мА, C",
    )
    for index, label in enumerate(labels):
        result = calculate_mapping_case(
            tmp_path,
            case_name=f"ad12-price-mapping-no-fallback-{index}",
            component_code=code,
            install_type="diff_1p_n",
            component_label=label,
        )

        assert result.status == "FAIL"
        assert result.total_preliminary_price is None
        assert any("unknown or ambiguous" in flag for flag in result.red_flags)


def test_ad12_price_mapping_rejects_workbook_drift_or_formulas(
    tmp_path: Path,
) -> None:
    changes: tuple[dict[str, Any], ...] = (
        {"A5": "УЗО АД-32 1Р+N до 63А EKF "},
        {"B5": 4101},
        {"C5": 433},
        {"B5": "=4100"},
        {"C5": "=432"},
    )
    for index, workbook_changes in enumerate(changes):
        result = calculate_mapping_case(
            tmp_path,
            case_name=f"ad12-price-mapping-fail-workbook-{index}",
            component_code="EKF-AD12-1P-N-C16-30MA-4P5KA",
            install_type="diff_1p_n",
            component_label=("Дифференциальный автомат АД12 2Р 16А, 30мА, C, 4,5кА"),
            workbook_changes=workbook_changes,
        )

        assert result.status == "FAIL"
        assert result.total_preliminary_price is None
        assert any("approved component mapping" in flag for flag in result.red_flags)


def test_existing_ad32_mapping_012_remains_exact_and_unchanged(
    tmp_path: Path,
) -> None:
    result = calculate_mapping_case(
        tmp_path,
        case_name="mapping-012-preserved",
        component_code="EKF-AD32-1P-N",
        install_type="diff_1p_n",
        component_label="АД12, 2P, C16, 30мА, 6кА",
    )

    assert result.status == "PASS"
    assert result.component_material_total == 4100
    assert result.work_total == 432
    assert calculator.RESOLVED_COMPONENT_MAPPING_PROVENANCE[
        "COMPONENT-MAPPING-012"
    ] == {
        "article": "DA32-6-16-30-ac-pro",
        "component_code": "EKF-AD32-1P-N",
        "pricing_decision_artifact_sha256": (
            calculator.PRICING_DECISION_ARTIFACT_SHA256
        ),
    }


def test_component_mapping_005_passes_only_exact_approved_path(
    tmp_path: Path,
) -> None:
    signature = calculator.parse_component_signature(
        "АВДТ-34, 4P, C16, 100мА",
        "diff_3p_4p",
    )
    assert signature is not None
    mapping = calculator.resolve_component_mapping(
        signature,
        "EKF-AVDT63N-3P-N-C16-100MA-6KA-S",
    )
    assert mapping is not None
    assert (mapping.sheet_name, mapping.row) == ("КРН", 28)

    result = calculate_mapping_case(
        tmp_path,
        case_name="mapping-005-pass",
        component_code="EKF-AVDT63N-3P-N-C16-100MA-6KA-S",
        install_type="diff_3p_4p",
        component_label="АВДТ-34, 4P, C16, 100мА",
    )

    assert result.status == "PASS"
    assert result.component_material_total == 8000
    assert result.work_total == 432


def test_component_mapping_005_rejects_wrong_code_or_signature(
    tmp_path: Path,
) -> None:
    approved_code = "EKF-AVDT63N-3P-N-C16-100MA-6KA-S"
    cases = (
        ("WRONG-CODE", "diff_3p_4p", "АВДТ-34, 4P, C16, 100мА"),
        (approved_code, "diff_3p_4p", "АВДТ-34, 4P, C20, 100мА"),
        (approved_code, "diff_3p_4p", "АВДТ-34, 4P, C16, 30мА"),
        (approved_code, "diff_3p_4p", "АВДТ-34, 2P, C16, 100мА"),
        (approved_code, "diff_1p_n", "АВДТ-34, 4P, C16, 100мА"),
    )
    for index, (code, install_type, label) in enumerate(cases):
        result = calculate_mapping_case(
            tmp_path,
            case_name=f"mapping-005-fail-signature-{index}",
            component_code=code,
            install_type=install_type,
            component_label=label,
        )

        assert result.status == "FAIL"
        assert result.total_preliminary_price is None
        assert any("unknown or ambiguous" in flag for flag in result.red_flags)


def test_component_mapping_005_rejects_row_28_drift_or_absence(
    tmp_path: Path,
) -> None:
    changes: tuple[dict[str, Any], ...] = (
        {"A28": "УЗО АД-32 1Р+N до 63А 100мА-300мА EKF"},
        {"B28": 8001},
        {"C28": 433},
        {"A28": None, "B28": None, "C28": None},
    )
    for index, workbook_changes in enumerate(changes):
        result = calculate_mapping_case(
            tmp_path,
            case_name=f"mapping-005-fail-workbook-{index}",
            component_code="EKF-AVDT63N-3P-N-C16-100MA-6KA-S",
            install_type="diff_3p_4p",
            component_label="АВДТ-34, 4P, C16, 100мА",
            workbook_changes=workbook_changes,
        )

        assert result.status == "FAIL"
        assert result.total_preliminary_price is None
        assert any("approved component mapping" in flag for flag in result.red_flags)


def test_component_mapping_005_cannot_use_row_29_or_legacy_fallback(
    tmp_path: Path,
    monkeypatch: Any,
) -> None:
    monkeypatch.setitem(
        calculator.COMPONENT_DEFINITIONS,
        "LEGACY-AVDT",
        calculator.ComponentDefinition(
            workbook_label="УЗО АД-32 3Р+N до 63А 100мА-300мА EKF",
            install_type="diff_3p_4p",
        ),
    )
    result = calculate_mapping_case(
        tmp_path,
        case_name="mapping-005-row-29-and-legacy",
        component_code="LEGACY-AVDT",
        install_type="diff_3p_4p",
        component_label="АВДТ-34, 4P, C16, 100мА",
        workbook_changes={"A28": None, "B28": None, "C28": None},
    )

    assert result.status == "FAIL"
    assert result.total_preliminary_price is None
    assert any("unknown or ambiguous" in flag for flag in result.red_flags)


def test_approved_cabinet_mappings_are_exact(tmp_path: Path) -> None:
    workbook_path = tmp_path / "approved.xlsx"
    write_approved_workbook(workbook_path)
    cases = (
        ("ПР", "ПР 800×600×250 мм, металл", 21336),
        ("КРН-36", "КРН-36, 540×330×100 мм, металл", 9405),
    )
    for index, (cabinet_code, cabinet_label, expected_price) in enumerate(cases):
        csv_path = tmp_path / f"cabinet-{index}.csv"
        is_pr = cabinet_code == "ПР"
        write_technical_csv(
            csv_path,
            [
                technical_row(
                    cabinet_code=cabinet_code,
                    component_code="RAW",
                    install_type=("mccb_up_to_100a" if is_pr else "diff_1p_n"),
                    component_label=(
                        "CHINT, автоматический выключатель 3P 63А"
                        if is_pr
                        else "CHINT, АВДТ 2P C20/30мА"
                    ),
                    cabinet_label=cabinet_label,
                )
            ],
        )
        assert calculate(workbook_path, csv_path).cabinet_price == expected_price


def test_approved_workbook_row_signature_mismatch_fails(tmp_path: Path) -> None:
    workbook_path = tmp_path / "mismatch.xlsx"
    write_approved_workbook(workbook_path)
    workbook = load_workbook(workbook_path)
    workbook["КРН"]["A5"] = "changed signature"
    workbook.save(workbook_path)
    workbook.close()
    csv_path = tmp_path / "composition.csv"
    write_technical_csv(
        csv_path,
        [
            technical_row(
                cabinet_code="КРН-36",
                component_code="RAW",
                install_type="diff_1p_n",
                component_label="CHINT, АВДТ 2P C16/30мА",
                cabinet_label="КРН-36, 540×330×100 мм, металл",
            )
        ],
    )
    result = calculate(workbook_path, csv_path)
    assert result.status == "FAIL"
    assert any("signature mismatch" in flag for flag in result.red_flags)


def test_unknown_technical_mapping_fails_closed(tmp_path: Path) -> None:
    workbook_path = tmp_path / "approved.xlsx"
    write_approved_workbook(workbook_path)
    csv_path = tmp_path / "composition.csv"
    write_technical_csv(
        csv_path,
        [
            technical_row(
                cabinet_code="КРН-36",
                component_code="RAW",
                install_type="diff_1p_n",
                component_label="CHINT, АВДТ 2P C25/30мА",
                cabinet_label="КРН-36, 540×330×100 мм, металл",
            )
        ],
    )
    result = calculate(workbook_path, csv_path)
    assert result.status == "FAIL"
    assert any("unknown or ambiguous" in flag for flag in result.red_flags)


def test_ambiguous_technical_mapping_fails_closed(
    tmp_path: Path,
    monkeypatch: Any,
) -> None:
    workbook_path = tmp_path / "approved.xlsx"
    write_approved_workbook(workbook_path)
    duplicate = calculator.APPROVED_COMPONENT_PRICE_MAPPINGS[1]
    monkeypatch.setattr(
        calculator,
        "APPROVED_COMPONENT_PRICE_MAPPINGS",
        calculator.APPROVED_COMPONENT_PRICE_MAPPINGS + (duplicate,),
    )
    csv_path = tmp_path / "composition.csv"
    write_technical_csv(
        csv_path,
        [
            technical_row(
                cabinet_code="КРН-36",
                component_code="RAW",
                install_type="diff_1p_n",
                component_label="CHINT, АВДТ 2P C16/30мА",
                cabinet_label="КРН-36, 540×330×100 мм, металл",
            )
        ],
    )
    result = calculate(workbook_path, csv_path)
    assert result.status == "FAIL"
    assert any("unknown or ambiguous" in flag for flag in result.red_flags)


def sha256(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()


def calculate(workbook_path: Path, csv_path: Path) -> Any:
    return calculator.calculate_price_draft(workbook_path, csv_path)


def write_shu_t1_workbook(path: Path, *, rt_work: int = 900) -> None:
    workbook = Workbook()
    krn = workbook.active
    krn.title = "КРН"
    krn["A5"] = "УЗО АД-32 1Р+N до 63А EKF"
    krn["B5"] = 4100
    krn["C5"] = 432
    krn["A13"] = "ВА47 2 полюсный"
    krn["B13"] = 1350
    krn["C13"] = 432
    krn["A19"] = "Терморегулятор RT-820"
    krn["B19"] = 15000
    krn["C19"] = rt_work
    krn["L6"] = "Корпус КРН-12 265х330х100"
    krn["M6"] = 6936
    workbook.save(path)
    workbook.close()


def shu_t1_rows() -> list[list[str]]:
    cabinet = "Корпус КРН-12 265×330×100 мм, металл"
    return [
        technical_row(
            product_name="ШУ-Т1",
            cabinet_code="CAB-KRN-12",
            component_code="EKF-RT-820",
            install_type="temperature_relay_din_2mod",
            component_label="Реле температуры RT-820 EKF PROxima с внешним датчиком",
            cabinet_label=cabinet,
        ),
        technical_row(
            product_name="ШУ-Т1",
            cabinet_code="CAB-KRN-12",
            component_code="EKF-AD12-1P-N-C16-30MA-4P5KA",
            install_type="diff_1p_n",
            component_label="АД12 Basic АВДТ 2P C16/30мА 4.5kA",
            cabinet_label=cabinet,
        ),
        technical_row(
            product_name="ШУ-Т1",
            cabinet_code="CAB-KRN-12",
            component_code="EKF-VA47-29-2P",
            install_type="modular_2p",
            component_label=("Автоматический выключатель ВА47-29 BASIC 2P C10 4.5kA"),
            cabinet_label=cabinet,
        ),
    ]


def shu_t2_rows() -> list[list[str]]:
    bindings = [
        calculator.SHU_T2_RT820_TECHNICAL_CONTRACT,
        calculator.SHU_T2_RT820_TECHNICAL_SHA256,
        calculator.SHU_T2_RT820_PROFILE_CONTRACT,
        calculator.SHU_T2_RT820_PROFILE_SHA256,
        calculator.SHU_T2_RT820_HUMAN_DECISION_SHA256,
    ]
    rows = shu_t1_rows()
    for row in rows:
        row[0] = "ШУ-Т2"
    return [[*row, *bindings] for row in rows]


def test_exact_shu_t1_rt820_mapping_preserves_exact_material_and_work(
    tmp_path: Path,
) -> None:
    workbook_path = tmp_path / "price.xlsx"
    csv_path = tmp_path / "composition.csv"
    write_shu_t1_workbook(workbook_path)
    write_technical_csv(csv_path, shu_t1_rows())

    result = calculate(workbook_path, csv_path)

    assert result.status == "PASS"
    assert result.cabinet_price == 6936
    assert result.component_material_total == 20450
    assert result.work_total == 1764
    # The standalone calculator reports its internal pre-tail draft. Invoice 519
    # applies the case tail and approved 53,763 KZT anchor in the checked runner.
    assert result.total_preliminary_price == 47783


def test_exact_bound_shu_t2_rt820_mapping_preserves_15000_and_900(
    tmp_path: Path,
) -> None:
    workbook_path = tmp_path / "price.xlsx"
    csv_path = tmp_path / "composition.csv"
    write_shu_t1_workbook(workbook_path)
    write_shu_t2_bound_csv(csv_path, shu_t2_rows())
    result = calculate(workbook_path, csv_path)
    assert result.status == "PASS"
    assert result.component_material_total == 20450
    assert result.work_total == 1764


@pytest.mark.parametrize("binding_index", [8, 9, 10, 11, 12])
def test_shu_t2_rt820_requires_every_exact_contract_binding(
    tmp_path: Path, binding_index: int
) -> None:
    workbook_path = tmp_path / "price.xlsx"
    csv_path = tmp_path / "composition.csv"
    write_shu_t1_workbook(workbook_path)
    rows = shu_t2_rows()
    rows[0][binding_index] = "not-an-exact-binding"
    write_shu_t2_bound_csv(csv_path, rows)
    result = calculate(workbook_path, csv_path)
    assert result.status == "FAIL"
    assert result.red_flags


@pytest.mark.parametrize(
    "profile_sha",
    ["a" * 64, "b" * 64, calculator.SHU_T2_RT820_PROFILE_SHA256.upper(), ""],
)
def test_standalone_calculator_rejects_every_non_frozen_profile_sha(
    tmp_path: Path, profile_sha: str
) -> None:
    workbook_path = tmp_path / "price.xlsx"
    csv_path = tmp_path / "composition.csv"
    write_shu_t1_workbook(workbook_path)
    rows = shu_t2_rows()
    rows[0][11] = profile_sha
    write_shu_t2_bound_csv(csv_path, rows)
    result = calculate(workbook_path, csv_path)
    assert result.status == "FAIL"
    assert result.red_flags


def test_rt820_remains_forbidden_for_other_product_even_with_new_bindings(
    tmp_path: Path,
) -> None:
    workbook_path = tmp_path / "price.xlsx"
    csv_path = tmp_path / "composition.csv"
    write_shu_t1_workbook(workbook_path)
    rows = shu_t2_rows()
    rows[0][0] = "OTHER-PRODUCT"
    write_shu_t2_bound_csv(csv_path, rows)
    result = calculate(workbook_path, csv_path)
    assert result.status == "FAIL"


@pytest.mark.parametrize(
    ("column", "value"),
    [
        (0, "OTHER-PROJECT-PRODUCT"),
        (2, "CAB-KRN-24"),
        (4, "SIMILAR-RELAY"),
        (5, "modular_2p"),
        (6, "Похожее реле температуры 2P"),
    ],
)
def test_rt820_exact_scope_identity_and_fallbacks_fail_closed(
    tmp_path: Path, column: int, value: str
) -> None:
    workbook_path = tmp_path / "price.xlsx"
    csv_path = tmp_path / "composition.csv"
    write_shu_t1_workbook(workbook_path)
    rows = shu_t1_rows()
    rows[0][column] = value
    write_technical_csv(csv_path, rows)
    result = calculate(workbook_path, csv_path)
    assert result.status == "FAIL"
    assert result.red_flags


def test_rt820_requires_exact_900_work_and_never_uses_generic_432(
    tmp_path: Path,
) -> None:
    workbook_path = tmp_path / "price.xlsx"
    csv_path = tmp_path / "composition.csv"
    write_shu_t1_workbook(workbook_path, rt_work=432)
    write_technical_csv(csv_path, shu_t1_rows())
    result = calculate(workbook_path, csv_path)
    assert result.status == "FAIL"
    assert any("price mismatch" in flag for flag in result.red_flags)


def test_rt820_label_free_csv_is_forbidden(tmp_path: Path) -> None:
    workbook_path = tmp_path / "price.xlsx"
    csv_path = tmp_path / "composition.csv"
    write_shu_t1_workbook(workbook_path)
    write_csv(
        csv_path,
        [
            [
                "ШУ-Т1",
                "CAB-KRN-12",
                "1.20",
                "EKF-RT-820",
                "1",
                "temperature_relay_din_2mod",
            ]
        ],
    )
    result = calculate(workbook_path, csv_path)
    assert result.status == "FAIL"
    assert any("label-free fallback" in flag for flag in result.red_flags)


def test_confirmed_reference_calculation_is_44512(tmp_path: Path) -> None:
    workbook_path = tmp_path / "price.xlsx"
    csv_path = tmp_path / "composition.csv"
    write_workbook(workbook_path)
    write_csv(csv_path)

    result = calculate(workbook_path, csv_path)

    assert result.status == "PASS"
    assert result.product_name == "РУ-АВР / ЩРН-24"
    assert result.cabinet_price == 7985
    assert result.component_material_total == 16900
    assert result.work_total == 2700
    assert result.consumables_factor == calculator.Decimal("1.20")
    assert result.base == calculator.Decimal("30965")
    assert result.total_preliminary_price == 44512


def test_forbidden_price_worksheet_is_never_selected(
    tmp_path: Path,
    monkeypatch: Any,
) -> None:
    workbook_path = tmp_path / "price.xlsx"
    csv_path = tmp_path / "composition.csv"
    write_workbook(workbook_path)
    write_csv(csv_path)
    real_load_workbook = calculator.load_workbook
    selected_sheets: list[str] = []

    class TrackingWorkbook:
        def __init__(self, wrapped: Any) -> None:
            self.wrapped = wrapped

        def __getitem__(self, name: str) -> Any:
            selected_sheets.append(name)
            if name == "Прайс":
                raise AssertionError("forbidden worksheet was selected")
            return self.wrapped[name]

        def close(self) -> None:
            self.wrapped.close()

    def tracking_load_workbook(*args: Any, **kwargs: Any) -> TrackingWorkbook:
        return TrackingWorkbook(real_load_workbook(*args, **kwargs))

    monkeypatch.setattr(calculator, "load_workbook", tracking_load_workbook)

    result = calculate(workbook_path, csv_path)

    assert result.status == "PASS"
    assert selected_sheets == ["КРН"]
    assert result.total_preliminary_price == 44512


def test_unknown_component_code_fails_with_red_flag(tmp_path: Path) -> None:
    workbook_path = tmp_path / "price.xlsx"
    csv_path = tmp_path / "composition.csv"
    write_workbook(workbook_path)
    rows = confirmed_rows()
    rows[0][3] = "UNKNOWN-COMPONENT"
    write_csv(csv_path, rows)

    result = calculate(workbook_path, csv_path)
    report = calculator.format_report(result)

    assert result.status == "FAIL"
    assert "component_code is not confirmed: UNKNOWN-COMPONENT" in report
    assert "ask Igor" in report
    assert result.total_preliminary_price is None


def test_unknown_cabinet_code_fails_with_red_flag(tmp_path: Path) -> None:
    workbook_path = tmp_path / "price.xlsx"
    csv_path = tmp_path / "composition.csv"
    write_workbook(workbook_path)
    rows = confirmed_rows()
    for row in rows:
        row[1] = "UNKNOWN-CABINET"
    write_csv(csv_path, rows)

    result = calculate(workbook_path, csv_path)
    report = calculator.format_report(result)

    assert result.status == "FAIL"
    assert "cabinet_code is not confirmed: UNKNOWN-CABINET" in report
    assert "ask Igor" in report
    assert result.cabinet_price is None


def test_missing_confirmed_component_price_fails(tmp_path: Path) -> None:
    workbook_path = tmp_path / "price.xlsx"
    csv_path = tmp_path / "composition.csv"
    write_workbook(workbook_path, include_component=False)
    write_csv(csv_path)

    result = calculate(workbook_path, csv_path)
    report = calculator.format_report(result)

    assert result.status == "FAIL"
    assert "component price row was not found in КРН" in report
    assert "ask Igor" in report


def test_missing_confirmed_cabinet_price_fails(tmp_path: Path) -> None:
    workbook_path = tmp_path / "price.xlsx"
    csv_path = tmp_path / "composition.csv"
    write_workbook(workbook_path, include_cabinet=False)
    write_csv(csv_path)

    result = calculate(workbook_path, csv_path)
    report = calculator.format_report(result)

    assert result.status == "FAIL"
    assert "cabinet price row was not found in КРН" in report
    assert "ask Igor" in report


def test_install_type_mismatch_fails_closed(tmp_path: Path) -> None:
    workbook_path = tmp_path / "price.xlsx"
    csv_path = tmp_path / "composition.csv"
    write_workbook(workbook_path)
    rows = confirmed_rows()
    rows[0][5] = "modular_3p"
    write_csv(csv_path, rows)

    result = calculate(workbook_path, csv_path)

    assert result.status == "FAIL"
    assert any("install_type does not match" in flag for flag in result.red_flags)


def test_calculation_does_not_create_or_change_xlsx(tmp_path: Path) -> None:
    workbook_path = tmp_path / "price.xlsx"
    csv_path = tmp_path / "composition.csv"
    write_workbook(workbook_path)
    write_csv(csv_path)
    xlsx_before = sorted(tmp_path.glob("*.xlsx"))
    workbook_hash_before = sha256(workbook_path)

    result = calculate(workbook_path, csv_path)

    assert result.status == "PASS"
    assert sorted(tmp_path.glob("*.xlsx")) == xlsx_before
    assert sha256(workbook_path) == workbook_hash_before


def test_report_contains_required_safety_boundaries(tmp_path: Path) -> None:
    workbook_path = tmp_path / "price.xlsx"
    csv_path = tmp_path / "composition.csv"
    write_workbook(workbook_path)
    write_csv(csv_path)

    report = calculator.format_report(calculate(workbook_path, csv_path))

    assert report.startswith("PRICE_CALCULATION_DRAFT_REPORT_START")
    assert report.endswith("PRICE_CALCULATION_DRAFT_REPORT_END")
    assert "Mode:\nread-only preliminary price draft" in report
    assert "Total preliminary price:\n44 512" in report
    assert "PASS is not commercial approval" in report
    assert "Igor approval required" in report
    assert "Manual Igor check:\nrequired" in report
    assert "Human Approval:" in report


def test_cli_prints_pass_report_and_returns_zero(tmp_path: Path) -> None:
    workbook_path = tmp_path / "price.xlsx"
    csv_path = tmp_path / "composition.csv"
    write_workbook(workbook_path)
    write_csv(csv_path)

    result = subprocess.run(
        [
            sys.executable,
            str(SCRIPT),
            "--price-workbook",
            str(workbook_path),
            "--input-csv",
            str(csv_path),
        ],
        capture_output=True,
        text=True,
        check=False,
    )

    assert result.returncode == 0
    assert "Status:\nPASS" in result.stdout
    assert "Cabinet price:\n7 985" in result.stdout
    assert "Component material total:\n16 900" in result.stdout
    assert "Work total:\n2 700" in result.stdout
    assert "Base:\n30 965" in result.stdout
    assert "Total preliminary price:\n44 512" in result.stdout


def test_existing_workflows_remain_isolated_from_calculator() -> None:
    calculator_text = SCRIPT.read_text(encoding="utf-8")
    technical_text = TECHNICAL_WORKFLOW.read_text(encoding="utf-8")
    commercial_text = COMMERCIAL_WORKFLOW.read_text(encoding="utf-8")

    assert "make_quote_capacity100_checked.ps1" not in calculator_text
    assert "make_quote_capacity100_commercial_checked.ps1" not in calculator_text
    assert "calc_quote_price_draft.py" not in technical_text
    assert "calc_quote_price_draft.py" not in commercial_text


def test_calculator_has_portable_exception_syntax_and_compiles(
    tmp_path: Path,
) -> None:
    py_compile.compile(
        str(SCRIPT),
        cfile=str(tmp_path / "calc_quote_price_draft.pyc"),
        doraise=True,
    )
    calculator_text = SCRIPT.read_text(encoding="utf-8")

    assert "except (OSError, ValueError):" in calculator_text
    assert "except OSError, ValueError:" not in calculator_text


def test_approved_component_definitions_are_exact() -> None:
    expected = {
        "EKF-VA47-29-2P": ("ВА47 2 полюсный", "modular_2p"),
        "EKF-VN-32-1P": ("ВН-32 1Р 16-25-40-63А", "load_switch_1p"),
        "EKF-VN-32-2P": ("ВН-32 2Р 16-25-40-63А", "load_switch_2p"),
        "EKF-VN-32-3P": (
            "ВН-32 3Р 16-25-40-63-80-100А",
            "load_switch_3p",
        ),
        "EKF-AD32-1P-N": ("УЗО АД-32 1Р+N до 63А EKF", "diff_1p_n"),
    }

    assert {
        code: (definition.workbook_label, definition.install_type)
        for code, definition in calculator.COMPONENT_DEFINITIONS.items()
        if code in expected
    } == expected


def test_approved_regular_cabinet_definitions_are_exact() -> None:
    assert calculator.CABINET_DEFINITIONS["CAB-KURN-038-24"] == (
        "Корпус КУРН-0,38-24 540х490х170"
    )
    assert calculator.CABINET_DEFINITIONS["CAB-KRN-18"] == ("Корпус КРН-18 265х440х100")
    assert calculator.CABINET_DEFINITIONS["CAB-KRN-12"] == ("Корпус КРН-12 265х330х100")
    assert calculator.CABINET_DEFINITIONS["CAB-KRN-24"] == ("Корпус КРН-24 395х330х100")
    assert calculator.CABINET_DEFINITIONS["CAB-SCHE-BI-900X900X120-M12"] == (
        "Встроенный ЩЭ, 900×900×120 мм, металл 1.2 мм"
    )


def test_seven_approved_templates_share_one_krn_12_code() -> None:
    templates = (
        "ШУ-Т2",
        "ЩАО-1Ж",
        "ЩАО-2Ж",
        "ЩАО-3Ж",
        "ЩО-1Ж",
        "ЩО-2Ж",
        "ЩО-3Ж",
    )

    assert {
        template: calculator.CABINET_SOURCE_TEMPLATE_CODES[template]
        for template in templates
    } == {template: "CAB-KRN-12" for template in templates}


def test_resolved_component_requests_have_exact_provenance() -> None:
    assert calculator.UNRESOLVED_COMPONENT_MAPPING_REQUESTS == {}
    assert calculator.PRICING_DECISION_ARTIFACT_SHA256 == (
        "777faed80c8ef92782378dd2a788160af8ad2252d8cb4f539560f15657a1d96e"
    )
    assert calculator.AD12_PRICE_MAPPING_DECISION_ARTIFACT_SHA256 == (
        "f67c0d79ec404a739ad5bdc3650a6259b9dc496a6f23ebffb7f29e7a9a24a17a"
    )
    assert calculator.AD12_PRICE_MAPPING_DECISION_ARTIFACT_SCHEMA == (
        "technical_ad12_price_mapping_human_decisions.v0.1"
    )
    assert calculator.AD12_PRICE_MAPPING_DECISION_ARTIFACT_STATUS == (
        "IGOR_AD12_SHARED_PRICE_MAPPING_APPROVED_NOT_APPLIED"
    )
    assert calculator.AD12_PRICE_MAPPING_DECISION_ID == (
        "IGOR-AD12-SHARED-PRICE-MAPPING-2024-086-001"
    )
    assert calculator.RESOLVED_COMPONENT_MAPPING_PROVENANCE == {
        "COMPONENT-MAPPING-005": {
            "article": "D63N46ES16C100",
            "component_code": "EKF-AVDT63N-3P-N-C16-100MA-6KA-S",
            "pricing_decision_artifact_sha256": (
                calculator.PRICING_DECISION_ARTIFACT_SHA256
            ),
        },
        "COMPONENT-MAPPING-012": {
            "article": "DA32-6-16-30-ac-pro",
            "component_code": "EKF-AD32-1P-N",
            "pricing_decision_artifact_sha256": (
                calculator.PRICING_DECISION_ARTIFACT_SHA256
            ),
        },
        "COMPONENT-MAPPING-009": {
            "article": "DA12-16-30-bas",
            "component_code": "EKF-AD12-1P-N-C16-30MA-4P5KA",
            "row_draft_ids": (
                "ROW-DRAFT-0024",
                "ROW-DRAFT-0025",
                "ROW-DRAFT-0026",
                "ROW-DRAFT-0027",
            ),
            "human_decision_artifact_sha256": (
                "5d6e0de7af052c959abff015f41081c8bddc10e834fe4f971e8a7d2e60f19c46"
            ),
            "pricing_decision_artifact_sha256": (
                calculator.AD12_PRICE_MAPPING_DECISION_ARTIFACT_SHA256
            ),
            "pricing_decision_artifact_schema": (
                calculator.AD12_PRICE_MAPPING_DECISION_ARTIFACT_SCHEMA
            ),
            "pricing_decision_artifact_status": (
                calculator.AD12_PRICE_MAPPING_DECISION_ARTIFACT_STATUS
            ),
            "pricing_decision_id": calculator.AD12_PRICE_MAPPING_DECISION_ID,
            "direct_human_shared_price_decision": True,
            "ad32_fallback_used_for_ad12": False,
            "scope_expansion": False,
        },
        "COMPONENT-MAPPING-016": {
            "article": "DA12-16-30-bas",
            "component_code": "EKF-AD12-1P-N-C16-30MA-4P5KA",
            "row_draft_ids": (
                "ROW-DRAFT-0074",
                "ROW-DRAFT-0075",
            ),
            "human_decision_artifact_sha256": (
                "5d6e0de7af052c959abff015f41081c8bddc10e834fe4f971e8a7d2e60f19c46"
            ),
            "pricing_decision_artifact_sha256": (
                calculator.AD12_PRICE_MAPPING_DECISION_ARTIFACT_SHA256
            ),
            "pricing_decision_artifact_schema": (
                calculator.AD12_PRICE_MAPPING_DECISION_ARTIFACT_SCHEMA
            ),
            "pricing_decision_artifact_status": (
                calculator.AD12_PRICE_MAPPING_DECISION_ARTIFACT_STATUS
            ),
            "pricing_decision_id": calculator.AD12_PRICE_MAPPING_DECISION_ID,
            "direct_human_shared_price_decision": True,
            "ad32_fallback_used_for_ad12": False,
            "scope_expansion": False,
        },
    }
    ad12_provenance = {
        request_id: calculator.RESOLVED_COMPONENT_MAPPING_PROVENANCE[request_id]
        for request_id in ("COMPONENT-MAPPING-009", "COMPONENT-MAPPING-016")
    }
    assert tuple(ad12_provenance) == (
        "COMPONENT-MAPPING-009",
        "COMPONENT-MAPPING-016",
    )
    assert tuple(
        row_id
        for provenance in ad12_provenance.values()
        for row_id in provenance["row_draft_ids"]
    ) == (
        "ROW-DRAFT-0024",
        "ROW-DRAFT-0025",
        "ROW-DRAFT-0026",
        "ROW-DRAFT-0027",
        "ROW-DRAFT-0074",
        "ROW-DRAFT-0075",
    )
    assert all(
        provenance["scope_expansion"] is False
        and provenance["ad32_fallback_used_for_ad12"] is False
        for provenance in ad12_provenance.values()
    )
    assert calculator.COMPONENT_DEFINITIONS["EKF-AD32-1P-N"].install_type == "diff_1p_n"
    assert (
        calculator.COMPONENT_DEFINITIONS["EKF-AD12-1P-N-C16-30MA-4P5KA"].install_type
        == "diff_1p_n"
    )
    assert (
        calculator.COMPONENT_DEFINITIONS["EKF-AD12-1P-N-C16-30MA-4P5KA"].workbook_label
        is None
    )
    assert "EKF-AVDT63N-3P-N-C16-100MA-6KA-S" not in calculator.COMPONENT_DEFINITIONS


def test_modular_technical_labels_with_breaking_capacity_parse_explicitly() -> None:
    cases = (
        (
            "Автоматический выключатель ВА47 1Р 20А — 2шт., C, 6кА",
            "modular_1p",
            "EKF-VA47-29-1P",
            1,
            20,
        ),
        (
            "Выключатель автоматический ВА47-63(C) 2Р 63/10A; 1шт, 6кА",
            "modular_2p",
            "EKF-VA47-29-2P",
            2,
            10,
        ),
        (
            "Автоматический выключатель ВА47 3Р 16А — 3шт., C, 6кА",
            "modular_3p",
            "EKF-VA47-29-3P",
            3,
            16,
        ),
    )

    for label, install_type, code, poles, rating in cases:
        signature = calculator.parse_component_signature(label, install_type)
        assert signature == calculator.TechnicalSignature(
            "mcb",
            poles,
            rating,
            None,
            "C",
            install_type,
            Decimal("6"),
        )
        assert calculator.technical_component_definition_matches(
            signature,
            code,
            install_type,
        )


def test_ad12_45ka_mapping_is_distinct_and_exact() -> None:
    label = "Дифференциальный автомат АД12 2Р 16А, 30мА — 1шт., C, 4,5кА"
    signature = calculator.parse_component_signature(label, "diff_1p_n")

    assert signature == calculator.TechnicalSignature(
        "rcbo",
        2,
        16,
        30,
        "C",
        "diff_1p_n",
        Decimal("4.5"),
    )
    assert calculator.technical_component_definition_matches(
        signature,
        "EKF-AD12-1P-N-C16-30MA-4P5KA",
        "diff_1p_n",
    )
    assert not calculator.technical_component_definition_matches(
        signature,
        "EKF-AD32-1P-N",
        "diff_1p_n",
    )


def test_five_approved_cabinet_technical_labels_are_exact() -> None:
    assert calculator.CABINET_TECHNICAL_LABELS == {
        "CAB-KURN-038-24": "Корпус КУРН-0,38-24 540×490×170 мм, металл",
        "CAB-KRN-18": "Корпус КРН-18 265×440×100 мм, металл",
        "CAB-KRN-12": "Корпус КРН-12 265×330×100 мм, металл",
        "CAB-KRN-24": "Корпус КРН-24 395×330×100 мм, металл",
        "CAB-SCHE-BI-900X900X120-M12": ("Встроенный ЩЭ, 900×900×120 мм, металл 1.2 мм"),
    }
    assert all(
        calculator.technical_cabinet_definition_matches(code, label)
        for code, label in calculator.CABINET_TECHNICAL_LABELS.items()
    )


def test_custom_sche_uses_checked_base_cost_without_price_row(
    tmp_path: Path,
) -> None:
    workbook_path = tmp_path / "prices.xlsx"
    write_workbook(workbook_path)
    csv_path = tmp_path / "sche.csv"
    write_technical_csv(
        csv_path,
        [
            technical_row(
                product_name="ЩЭ-3кв",
                cabinet_code="CAB-SCHE-BI-900X900X120-M12",
                component_code="EKF-VA47-29-1P",
                install_type="modular_1p",
                component_label=(
                    "Автоматический выключатель ВА47 1Р 20А — 1шт., C, 6кА"
                ),
                cabinet_label=("Встроенный ЩЭ, 900×900×120 мм, металл 1.2 мм"),
            )
        ],
    )

    result = calculator.calculate_price_draft(
        workbook_path,
        csv_path,
        custom_cabinet_base_cost=10000,
    )

    assert result.status == "PASS"
    assert result.cabinet_price == 10000


def test_invoice519_exact_pr_formula_anchors_are_calculated() -> None:
    for material, work, expected in ((14850, 3024, 54023), (17050, 3564, 59166)):
        result = calculator.calculate_invoice519_position_price(
            project_id="2024/086",
            profile_decision_id="IGOR-INVOICE519-PRICING-PROFILE-2024-086-001",
            formula_family="CURRENT_MODULAR_CASE_PROFILE",
            cabinet_code="CAB-KURN-038-24",
            cabinet_base_kzt=12557,
            additional_cabinet_cost_kzt=0,
            component_material_total_kzt=material,
            work_total_kzt=work,
            physical_multiplicity=1,
        )

        assert result.rounded_unit_price_kzt == expected
        assert result.position_total_kzt == expected
        assert result.apartment_component_kzt == 0
        assert result.unrounded_unit_price_kzt != Decimal(expected)


def test_invoice519_exact_sche_formula_anchors_are_calculated() -> None:
    cases = (
        (3, 9600, 2592, 80413),
        (4, 12800, 3456, 96270),
        (5, 16000, 4320, 112127),
        (6, 19200, 5184, 127984),
    )
    for apartments, material, work, expected in cases:
        result = calculator.calculate_invoice519_position_price(
            project_id="2024/086",
            profile_decision_id="IGOR-INVOICE519-PRICING-PROFILE-2024-086-001",
            formula_family="CURRENT_SCHE_CASE_PROFILE",
            cabinet_code="CAB-SCHE-BI-900X900X120-M12",
            cabinet_base_kzt=20305,
            additional_cabinet_cost_kzt=0,
            component_material_total_kzt=material,
            work_total_kzt=work,
            physical_multiplicity=2,
            apartment_count=apartments,
        )

        assert result.rounded_unit_price_kzt == expected
        assert result.position_total_kzt == expected * 2
        assert result.apartment_component_kzt == 5100 * apartments


def test_invoice519_formula_is_exact_scope_and_fail_closed() -> None:
    valid = {
        "project_id": "2024/086",
        "profile_decision_id": "IGOR-INVOICE519-PRICING-PROFILE-2024-086-001",
        "formula_family": "CURRENT_MODULAR_CASE_PROFILE",
        "cabinet_code": "CAB-KRN-12",
        "cabinet_base_kzt": 6936,
        "additional_cabinet_cost_kzt": 0,
        "component_material_total_kzt": 100,
        "work_total_kzt": 20,
        "physical_multiplicity": 1,
    }
    invalid_overrides = (
        {"project_id": "OTHER"},
        {"profile_decision_id": "OTHER"},
        {"cabinet_code": "UNKNOWN"},
        {"cabinet_base_kzt": 18762},
        {"additional_cabinet_cost_kzt": 1},
        {"component_material_total_kzt": -1},
        {"work_total_kzt": -1},
        {"physical_multiplicity": 0},
        {"formula_family": "RESERVED"},
        {"apartment_count": 3},
    )
    for override in invalid_overrides:
        arguments = valid | override
        try:
            calculator.calculate_invoice519_position_price(**arguments)
        except ValueError:
            pass
        else:
            raise AssertionError(f"expected fail-closed formula for {override}")

    for apartments in (None, 2, 7):
        arguments = valid | {
            "formula_family": "CURRENT_SCHE_CASE_PROFILE",
            "cabinet_code": "CAB-SCHE-BI-900X900X120-M12",
            "cabinet_base_kzt": 20305,
            "apartment_count": apartments,
        }
        try:
            calculator.calculate_invoice519_position_price(**arguments)
        except ValueError:
            pass
        else:
            raise AssertionError("invalid ЩЭ apartment scope must fail")
