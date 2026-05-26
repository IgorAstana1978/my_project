import importlib.util
import json
import re
import shutil
import subprocess
import sys
from copy import deepcopy
from pathlib import Path
from types import ModuleType
from typing import Any, cast
from zipfile import ZIP_DEFLATED, ZipFile

from openpyxl import Workbook, load_workbook  # type: ignore[import-untyped]

PROJECT_ROOT = Path(__file__).resolve().parents[1]
SCRIPT = PROJECT_ROOT / "scripts" / "fill_invoice_quote_draft.py"
SHEET_NAME = "Счёт-КП шаблон"
TINY_PNG = (
    b"\x89PNG\r\n\x1a\n\x00\x00\x00\rIHDR\x00\x00\x00\x01"
    b"\x00\x00\x00\x01\x08\x06\x00\x00\x00\x1f\x15\xc4\x89"
    b"\x00\x00\x00\rIDATx\x9cc\xf8\xff\xff?\x00\x05\xfe"
    b"\x02\xfeA\xe2d\x9a\x00\x00\x00\x00IEND\xaeB`\x82"
)


def load_draft_module() -> ModuleType:
    spec = importlib.util.spec_from_file_location(
        "fill_invoice_quote_draft_for_test", SCRIPT
    )
    assert spec is not None
    assert spec.loader is not None
    module = importlib.util.module_from_spec(spec)
    sys.modules[spec.name] = module
    spec.loader.exec_module(module)
    return module


draft = cast(Any, load_draft_module())


def valid_data() -> dict[str, Any]:
    return {
        "document": {"status": "черновик"},
        "customer": {"payer_name": "ТОО Ромашка"},
        "project": {
            "object_name": "Насосная станция",
            "basis_or_project": "нужно уточнить",
            "section_or_project_position": "Шкаф управления насосами",
        },
        "items": [
            {
                "name": "Шкаф управления насосом",
                "unit": "шт",
                "quantity": 1,
                "instruments_and_devices": "нужно уточнить",
                "cabinet_type_dimensions_material": "нужно уточнить",
                "price_kzt": None,
                "price_confirmed_by_igor": False,
            },
            {
                "name": "Шкаф распределительный",
                "unit": "шт",
                "quantity": 2,
                "price_kzt": None,
                "price_confirmed_by_igor": False,
            },
        ],
        "commercial_terms": {
            "total_amount_words": "нужно уточнить",
            "total_amount_confirmed_by_igor": False,
            "delivery_lead_time_working_days": None,
            "delivery_confirmed_by_igor": False,
        },
    }


def rewrite_xlsx(
    path: Path,
    updates: dict[str, bytes],
    removals: set[str] | None = None,
) -> None:
    removals = removals or set()
    temporary_path = path.with_suffix(".tmp.xlsx")
    with ZipFile(path) as source_archive:
        source_entries = {
            name: source_archive.read(name)
            for name in source_archive.namelist()
            if name not in updates and name not in removals
        }

    with ZipFile(temporary_path, "w", ZIP_DEFLATED) as target_archive:
        for name, content in source_entries.items():
            target_archive.writestr(name, content)
        for name, content in updates.items():
            target_archive.writestr(name, content)

    temporary_path.replace(path)


def content_types_with_drawing(content: str) -> str:
    if 'Extension="png"' not in content:
        content = content.replace(
            "</Types>",
            '<Default Extension="png" ContentType="image/png"/></Types>',
        )
    if 'PartName="/xl/drawings/drawing1.xml"' not in content:
        content = content.replace(
            "</Types>",
            (
                '<Override PartName="/xl/drawings/drawing1.xml" '
                'ContentType="application/vnd.openxmlformats-officedocument.'
                'drawing+xml"/></Types>'
            ),
        )
    return content


def add_valid_drawing_chain(path: Path) -> None:
    with ZipFile(path) as archive:
        sheet_xml = archive.read("xl/worksheets/sheet1.xml").decode("utf-8")
        content_types = archive.read("[Content_Types].xml").decode("utf-8")

    sheet_xml = sheet_xml.replace(
        "</worksheet>",
        (
            '<drawing xmlns:r="http://schemas.openxmlformats.org/'
            'officeDocument/2006/relationships" r:id="rId1"/>'
            "</worksheet>"
        ),
    )
    sheet_rels = (
        '<?xml version="1.0" encoding="UTF-8"?>'
        '<Relationships xmlns="http://schemas.openxmlformats.org/package/'
        '2006/relationships">'
        '<Relationship Id="rId1" Type="http://schemas.openxmlformats.org/'
        'officeDocument/2006/relationships/drawing" '
        'Target="../drawings/drawing1.xml"/>'
        "</Relationships>"
    )
    drawing_xml = (
        '<?xml version="1.0" encoding="UTF-8"?>'
        '<xdr:wsDr xmlns:xdr="http://schemas.openxmlformats.org/'
        'drawingml/2006/spreadsheetDrawing" '
        'xmlns:a="http://schemas.openxmlformats.org/drawingml/2006/main">'
        "<xdr:oneCellAnchor><xdr:from><xdr:col>0</xdr:col>"
        "<xdr:colOff>0</xdr:colOff><xdr:row>0</xdr:row>"
        "<xdr:rowOff>0</xdr:rowOff></xdr:from>"
        '<xdr:ext cx="1" cy="1"/>'
        '<xdr:pic><xdr:nvPicPr><xdr:cNvPr id="2" name="logo"/>'
        "<xdr:cNvPicPr/></xdr:nvPicPr><xdr:blipFill>"
        '<a:blip xmlns:r="http://schemas.openxmlformats.org/'
        'officeDocument/2006/relationships" r:embed="rId1"/>'
        "<a:stretch><a:fillRect/></a:stretch></xdr:blipFill>"
        "<xdr:spPr/></xdr:pic><xdr:clientData/></xdr:oneCellAnchor>"
        "</xdr:wsDr>"
    )
    drawing_rels = (
        '<?xml version="1.0" encoding="UTF-8"?>'
        '<Relationships xmlns="http://schemas.openxmlformats.org/package/'
        '2006/relationships">'
        '<Relationship Id="rId1" Type="http://schemas.openxmlformats.org/'
        'officeDocument/2006/relationships/image" '
        'Target="../media/image1.png"/>'
        "</Relationships>"
    )

    rewrite_xlsx(
        path,
        {
            "xl/worksheets/sheet1.xml": sheet_xml.encode("utf-8"),
            "xl/worksheets/_rels/sheet1.xml.rels": sheet_rels.encode("utf-8"),
            "xl/drawings/drawing1.xml": drawing_xml.encode("utf-8"),
            "xl/drawings/_rels/drawing1.xml.rels": drawing_rels.encode("utf-8"),
            "xl/media/image1.png": TINY_PNG,
            "[Content_Types].xml": content_types_with_drawing(content_types).encode(
                "utf-8"
            ),
        },
    )


def remove_worksheet_drawing_reference(path: Path) -> None:
    with ZipFile(path) as archive:
        sheet_xml = archive.read("xl/worksheets/sheet1.xml").decode("utf-8")
    sheet_xml = re.sub(r"<(?:\w+:)?drawing\b[^>]*/>", "", sheet_xml)
    rewrite_xlsx(path, {"xl/worksheets/sheet1.xml": sheet_xml.encode("utf-8")})


def remove_sheet_drawing_relationship(path: Path) -> None:
    rewrite_xlsx(path, {}, {"xl/worksheets/_rels/sheet1.xml.rels"})


def write_template(path: Path) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = SHEET_NAME

    for merged_range in (
        "B9:F9",
        "G9:I13",
        "B10:F10",
        "B11:F11",
        "B12:F12",
        "B13:F13",
        "C24:I24",
        "C26:I30",
        "B32:I34",
    ):
        worksheet.merge_cells(merged_range)

    for row in range(2, 7):
        for column in range(3, 10):
            worksheet.cell(row=row, column=column).value = f"header-{row}-{column}"
    for row in range(4, 7):
        worksheet.cell(row=row, column=2).value = f"req-{row}"
    worksheet["B32"] = "signature-32"

    worksheet["B10"] = "Плательщик: шаблон"
    worksheet["B11"] = "Объект: шаблон"
    worksheet["B12"] = "Основание / проект: шаблон"
    worksheet["B13"] = "Статус документа: шаблон"
    worksheet["C16"] = "Раздел шаблона"
    worksheet["C17"] = "Позиция шаблона"
    worksheet["D17"] = "шт"
    worksheet["E17"] = 99
    worksheet["F17"] = "Прибор шаблона"
    worksheet["G17"] = "Шкаф шаблона"
    worksheet["H17"] = "Цена шаблона"
    worksheet["H22"] = "Не писать напрямую"
    worksheet["C24"] = "Всего прописью: шаблон"
    worksheet["C26"] = "Примечание шаблона"

    for row in range(17, 22):
        worksheet[f"I{row}"] = f"=E{row}*H{row}"
    worksheet["I22"] = "=SUM(I17:I21)"

    workbook.save(path)
    add_valid_drawing_chain(path)


def write_json(path: Path, data: dict[str, Any]) -> None:
    path.write_text(json.dumps(data, ensure_ascii=False), encoding="utf-8")


def run_script(
    template: Path, input_json: Path, output: Path
) -> subprocess.CompletedProcess[str]:
    return subprocess.run(
        [
            sys.executable,
            str(SCRIPT),
            "--template",
            str(template),
            "--input",
            str(input_json),
            "--output",
            str(output),
        ],
        cwd=PROJECT_ROOT,
        capture_output=True,
        text=True,
        check=False,
    )


def workbook_values(path: Path, cells: list[str]) -> dict[str, Any]:
    workbook = load_workbook(path, data_only=False)
    worksheet = workbook[SHEET_NAME]
    return {cell: worksheet[cell].value for cell in cells}


def merged_ranges(path: Path) -> tuple[str, ...]:
    workbook = load_workbook(path, data_only=False)
    worksheet = workbook[SHEET_NAME]
    return tuple(str(item) for item in worksheet.merged_cells.ranges)


def formula_values(path: Path) -> dict[str, Any]:
    return workbook_values(path, [f"I{row}" for row in range(17, 23)])


def prepare_case(tmp_path: Path) -> tuple[Path, Path, Path, dict[str, Any]]:
    template = tmp_path / "template.xlsx"
    input_json = tmp_path / "input.json"
    output = tmp_path / "output.xlsx"
    data = valid_data()
    write_template(template)
    write_json(input_json, data)
    return template, input_json, output, data


def assert_rejected(result: subprocess.CompletedProcess[str], text: str) -> None:
    assert result.returncode == 1
    assert text in result.stderr


def failed_check(results: list[Any], name: str) -> bool:
    return any(result.name == name and not result.passed for result in results)


def test_successfully_fills_temporary_xlsx(tmp_path: Path) -> None:
    template, input_json, output, _data = prepare_case(tmp_path)

    result = run_script(template, input_json, output)

    assert result.returncode == 0, result.stderr
    assert "PASS: output создан" in result.stdout
    assert "PASS: sheet1.xml содержит drawing reference" in result.stdout
    assert (
        "PASS: xl/worksheets/_rels/sheet1.xml.rels содержит relationship " "на drawing"
    ) in result.stdout
    values = workbook_values(
        output,
        [
            "B10",
            "B11",
            "B12",
            "B13",
            "C16",
            "C17",
            "D17",
            "E17",
            "F17",
            "G17",
            "H17",
            "C18",
            "D18",
            "E18",
            "F18",
            "G18",
            "H18",
            "C19",
            "D19",
            "E19",
            "F19",
            "G19",
            "H19",
        ],
    )
    assert values["B10"] == "Плательщик: ТОО Ромашка"
    assert values["B11"] == "Объект: Насосная станция"
    assert values["B12"] == "Основание / проект: нужно уточнить"
    assert values["B13"] == "Статус документа: Черновик"
    assert values["C16"] == "Раздел шаблона"
    assert values["C17"] == "Шкаф управления насосом"
    assert values["D17"] == "шт"
    assert values["E17"] == 1
    assert values["F17"] == "нужно уточнить"
    assert values["G17"] == "нужно уточнить"
    assert values["H17"] == "нужно уточнить"
    assert values["C18"] == "Шкаф распределительный"
    assert values["D18"] == "шт"
    assert values["E18"] == 2
    assert values["F18"] == "нужно уточнить"
    assert values["G18"] == "нужно уточнить"
    assert values["H18"] == "нужно уточнить"
    assert values["C19"] == "нужно уточнить"
    assert values["D19"] == "шт"
    assert values["E19"] == 1
    assert values["F19"] == "нужно уточнить"
    assert values["G19"] == "нужно уточнить"
    assert values["H19"] == "нужно уточнить"


def test_rejects_more_than_five_items(tmp_path: Path) -> None:
    template, input_json, output, data = prepare_case(tmp_path)
    item = deepcopy(data["items"][0])
    data["items"] = [deepcopy(item) for _index in range(6)]
    write_json(input_json, data)

    result = run_script(template, input_json, output)

    assert_rejected(result, "items больше 5")
    assert not output.exists()


def test_rejects_non_draft_status(tmp_path: Path) -> None:
    template, input_json, output, data = prepare_case(tmp_path)
    data["document"]["status"] = "финальный"
    write_json(input_json, data)

    result = run_script(template, input_json, output)

    assert_rejected(result, 'document.status должен быть "черновик"')
    assert not output.exists()


def test_rejects_non_null_price(tmp_path: Path) -> None:
    template, input_json, output, data = prepare_case(tmp_path)
    data["items"][0]["price_kzt"] = 123
    write_json(input_json, data)

    result = run_script(template, input_json, output)

    assert_rejected(result, "price_kzt должен быть null")
    assert not output.exists()


def test_rejects_confirmed_price(tmp_path: Path) -> None:
    template, input_json, output, data = prepare_case(tmp_path)
    data["items"][0]["price_confirmed_by_igor"] = True
    write_json(input_json, data)

    result = run_script(template, input_json, output)

    assert_rejected(result, "price_confirmed_by_igor должен быть false")
    assert not output.exists()


def test_rejects_delivery_time_without_confirmation(tmp_path: Path) -> None:
    template, input_json, output, data = prepare_case(tmp_path)
    data["commercial_terms"]["delivery_lead_time_working_days"] = 10
    data["commercial_terms"]["delivery_confirmed_by_igor"] = False
    write_json(input_json, data)

    result = run_script(template, input_json, output)

    assert_rejected(result, "delivery_lead_time_working_days заполнен")
    assert not output.exists()


def test_rejects_output_equal_to_template(tmp_path: Path) -> None:
    template, input_json, _output, _data = prepare_case(tmp_path)

    result = run_script(template, input_json, template)

    assert_rejected(result, "output совпадает с template")


def test_rejects_output_inside_project(tmp_path: Path) -> None:
    template, input_json, _output, _data = prepare_case(tmp_path)
    output = PROJECT_ROOT / "fill_invoice_quote_draft_blocked_output.xlsx"

    result = run_script(template, input_json, output)

    assert_rejected(result, "output находится внутри Git-проекта")
    assert not output.exists()


def test_preserves_formulas_and_merged_ranges(tmp_path: Path) -> None:
    template, input_json, output, _data = prepare_case(tmp_path)
    before_formulas = formula_values(template)
    before_merged_ranges = merged_ranges(template)

    result = run_script(template, input_json, output)

    assert result.returncode == 0, result.stderr
    assert formula_values(output) == before_formulas
    assert merged_ranges(output) == before_merged_ranges


def test_preserves_forbidden_cells(tmp_path: Path) -> None:
    template, input_json, output, _data = prepare_case(tmp_path)
    forbidden_cells = [
        "I17",
        "I18",
        "I19",
        "I20",
        "I21",
        "I22",
        "H22",
        "C24",
        "C26",
        "B32",
        "C2",
        "I6",
        "B4",
        "B6",
    ]
    before_values = workbook_values(template, forbidden_cells)

    result = run_script(template, input_json, output)

    assert result.returncode == 0, result.stderr
    assert workbook_values(output, forbidden_cells) == before_values


def test_c16_is_not_overwritten_by_project_position(tmp_path: Path) -> None:
    template, input_json, output, _data = prepare_case(tmp_path)

    result = run_script(template, input_json, output)

    assert result.returncode == 0, result.stderr
    assert workbook_values(output, ["C16"])["C16"] == "Раздел шаблона"


def test_empty_item_rows_are_filled_with_safe_values(tmp_path: Path) -> None:
    template, input_json, output, _data = prepare_case(tmp_path)

    result = run_script(template, input_json, output)

    assert result.returncode == 0, result.stderr
    values = workbook_values(
        output,
        [
            "C19",
            "D19",
            "E19",
            "F19",
            "G19",
            "H19",
            "C20",
            "D20",
            "E20",
            "F20",
            "G20",
            "H20",
            "C21",
            "D21",
            "E21",
            "F21",
            "G21",
            "H21",
        ],
    )
    for row in range(19, 22):
        assert values[f"C{row}"] == "нужно уточнить"
        assert values[f"D{row}"] == "шт"
        assert values[f"E{row}"] == 1
        assert values[f"F{row}"] == "нужно уточнить"
        assert values[f"G{row}"] == "нужно уточнить"
        assert values[f"H{row}"] == "нужно уточнить"


def test_orphan_drawings_do_not_pass_verification(tmp_path: Path) -> None:
    template, _input_json, output, _data = prepare_case(tmp_path)
    shutil.copyfile(template, output)
    remove_sheet_drawing_relationship(output)
    before = draft.snapshot_template(template)

    results = draft.verify_output(template, output, before)

    assert failed_check(
        results,
        "xl/worksheets/_rels/sheet1.xml.rels содержит relationship на drawing",
    )


def test_missing_worksheet_drawing_reference_returns_fail(tmp_path: Path) -> None:
    template, _input_json, output, _data = prepare_case(tmp_path)
    shutil.copyfile(template, output)
    remove_worksheet_drawing_reference(output)
    before = draft.snapshot_template(template)

    results = draft.verify_output(template, output, before)

    assert failed_check(results, "sheet1.xml содержит drawing reference")
