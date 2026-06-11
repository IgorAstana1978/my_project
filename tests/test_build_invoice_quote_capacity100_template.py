import hashlib
import importlib.util
import re
import subprocess
import sys
from pathlib import Path
from types import ModuleType
from typing import Any, cast
from xml.etree import ElementTree
from zipfile import ZIP_DEFLATED, ZipFile

from openpyxl import load_workbook  # type: ignore[import-untyped]

PROJECT_ROOT = Path(__file__).resolve().parents[1]
SCRIPT = PROJECT_ROOT / "scripts" / "build_invoice_quote_capacity100_template.py"
SNAPSHOT_SCRIPT = PROJECT_ROOT / "scripts" / "drawing_media_snapshot.py"
SHEET_NAME = "Счёт-КП шаблон"
NS = {"main": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
TINY_PNG = (
    b"\x89PNG\r\n\x1a\n\x00\x00\x00\rIHDR\x00\x00\x00\x01"
    b"\x00\x00\x00\x01\x08\x06\x00\x00\x00\x1f\x15\xc4\x89"
    b"\x00\x00\x00\rIDATx\x9cc\xf8\xff\xff?\x00\x05\xfe"
    b"\x02\xfeA\xe2d\x9a\x00\x00\x00\x00IEND\xaeB`\x82"
)


def load_script_module(module_name: str, path: Path) -> ModuleType:
    spec = importlib.util.spec_from_file_location(module_name, path)
    assert spec is not None
    assert spec.loader is not None
    module = importlib.util.module_from_spec(spec)
    sys.modules[spec.name] = module
    spec.loader.exec_module(module)
    return module


builder = cast(Any, load_script_module("capacity_builder_for_test", SCRIPT))
snapshot = cast(
    Any,
    load_script_module("drawing_media_snapshot_for_capacity_test", SNAPSHOT_SCRIPT),
)


def formula(row: int) -> str:
    return f'IF(OR(E{row}="",H{row}=""),"",IFERROR(E{row}*H{row},"нужно уточнить"))'


def total_formula(start: int = 17, end: int = 21) -> str:
    return f'IF(COUNT(I{start}:I{end})=0,"нужно уточнить",SUM(I{start}:I{end}))'


def cell(column: str, row: int, body: str = "", style: int = 12) -> str:
    return f'<c r="{column}{row}" s="{style}">{body}</c>'


def row_xml(row: int, item_number: int | None = None) -> str:
    if item_number is None:
        values = "".join(cell(column, row) for column in "ABCDEFGHI")
    else:
        values = (
            cell("A", row, "", 1)
            + cell("B", row, f"<v>{item_number}</v>", 12)
            + cell("C", row, "", 13)
            + f'<c r="D{row}" s="12" t="inlineStr"><is><t>шт.</t></is></c>'
            + cell("E", row, "", 12)
            + cell("F", row, "", 14)
            + cell("G", row, "", 14)
            + cell("H", row, "", 12)
            + cell("I", row, f"<f>{formula(row)}</f>", 12)
        )
    return f'<row r="{row}" spans="1:9" ht="54" customHeight="1">{values}</row>'


def source_sheet_xml(*, sheet_name: str = SHEET_NAME) -> bytes:
    del sheet_name
    rows = [
        '<row r="15"><c r="B15" s="1" t="inlineStr"><is><t>№</t></is></c></row>',
        '<row r="16"><c r="C16" s="1" t="inlineStr"><is><t>Раздел</t></is></c></row>',
    ]
    rows.extend(row_xml(row, row - 16) for row in range(17, 22))
    rows.append(
        '<row r="22" ht="27.95" customHeight="1">'
        '<c r="H22" s="15" t="inlineStr"><is><t>ИТОГО</t></is></c>'
        f'<c r="I22" s="12" t="str"><f>{total_formula()}</f></c></row>'
    )
    rows.append('<row r="23" ht="8.1" customHeight="1"/>')
    rows.append(
        '<row r="24" ht="24" customHeight="1">'
        '<c r="C24" s="16" t="inlineStr"><is><t>Всего прописью</t></is></c></row>'
    )
    rows.append('<row r="25" ht="8.1" customHeight="1"/>')
    for row in range(26, 31):
        rows.append(
            f'<row r="{row}" ht="23.1" customHeight="1">'
            f'<c r="C{row}" s="16" t="inlineStr"><is>'
            f"<t>Примечание {row}</t></is></c></row>"
        )
    rows.append('<row r="31" ht="8.1" customHeight="1"/>')
    rows.append(
        '<row r="32"><c r="B32" t="inlineStr"><is>' "<t>Директор</t></is></c></row>"
    )
    rows.append(
        '<row r="33"><c r="B33" t="inlineStr"><is>' "<t>Исполнитель</t></is></c></row>"
    )
    rows.append(
        '<row r="34"><c r="B34" t="inlineStr"><is>'
        "<t>Дата проверки</t></is></c></row>"
    )
    rows.append('<row r="35" ht="8.1" customHeight="1"/>')
    rows.append('<row r="36" ht="8.1" customHeight="1"/>')
    return (
        '<?xml version="1.0" encoding="UTF-8"?>'
        '<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" '
        'xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships" '
        'xmlns:mc="http://schemas.openxmlformats.org/markup-compatibility/2006" '
        'xmlns:x14ac="http://schemas.microsoft.com/office/spreadsheetml/2009/9/ac" '
        'mc:Ignorable="x14ac">'
        '<dimension ref="A1:I36"/>'
        '<sheetViews><sheetView workbookViewId="0"/></sheetViews>'
        '<sheetFormatPr defaultRowHeight="14.25"/>'
        "<sheetData>" + "".join(rows) + "</sheetData>"
        '<mergeCells count="11">'
        '<mergeCell ref="C24:I24"/><mergeCell ref="C26:I26"/>'
        '<mergeCell ref="C27:I27"/><mergeCell ref="C28:I28"/>'
        '<mergeCell ref="C29:I29"/><mergeCell ref="C30:I30"/>'
        '<mergeCell ref="B32:E32"/><mergeCell ref="F32:I32"/>'
        '<mergeCell ref="B33:E33"/><mergeCell ref="F33:I33"/>'
        '<mergeCell ref="B34:I34"/></mergeCells>'
        '<pageMargins left="0.7" right="0.7" top="0.75" bottom="0.75" '
        'header="0.3" footer="0.3"/>'
        '<drawing r:id="rId1"/></worksheet>'
    ).encode()


def package_parts(
    sheet_xml: bytes | None = None,
    workbook_name: str = SHEET_NAME,
) -> dict[str, bytes]:
    sheet_xml = source_sheet_xml() if sheet_xml is None else sheet_xml
    workbook_xml = (
        '<?xml version="1.0" encoding="UTF-8"?>'
        '<workbook xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" '
        'xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships">'
        '<sheets><sheet name="'
        + workbook_name
        + '" sheetId="1" r:id="rId1"/></sheets><calcPr calcId="1"/></workbook>'
    ).encode()
    return {
        "[Content_Types].xml": (
            b'<?xml version="1.0" encoding="UTF-8"?>'
            b'<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">'
            b'<Default Extension="xml" ContentType="application/xml"/>'
            b'<Default Extension="rels" ContentType="application/'
            b'vnd.openxmlformats-package.relationships+xml"/>'
            b'<Default Extension="png" ContentType="image/png"/>'
            b'<Override PartName="/xl/workbook.xml" ContentType="application/'
            b'vnd.openxmlformats-officedocument.spreadsheetml.sheet.main+xml"/>'
            b'<Override PartName="/xl/styles.xml" ContentType="application/'
            b'vnd.openxmlformats-officedocument.spreadsheetml.styles+xml"/>'
            b'<Override PartName="/xl/calcChain.xml" ContentType="application/'
            b'vnd.openxmlformats-officedocument.spreadsheetml.calcChain+xml"/>'
            b'<Override PartName="/xl/worksheets/sheet1.xml" '
            b'ContentType="application/vnd.openxmlformats-officedocument.'
            b'spreadsheetml.worksheet+xml"/>'
            b'<Override PartName="/xl/drawings/drawing1.xml" '
            b'ContentType="application/vnd.openxmlformats-officedocument.'
            b'drawing+xml"/>'
            b"</Types>"
        ),
        "_rels/.rels": (
            b'<?xml version="1.0" encoding="UTF-8"?>'
            b'<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
            b'<Relationship Id="rId1" Type="http://schemas.openxmlformats.'
            b'org/officeDocument/2006/relationships/officeDocument" '
            b'Target="xl/workbook.xml"/>'
            b"</Relationships>"
        ),
        "xl/workbook.xml": workbook_xml,
        "xl/styles.xml": (
            b'<?xml version="1.0" encoding="UTF-8"?>'
            b'<styleSheet xmlns="http://schemas.openxmlformats.org/'
            b'spreadsheetml/2006/main">'
            b'<fonts count="8">'
            b'<font><name val="Calibri"/><sz val="11"/><color theme="1"/></font>'
            b'<font><name val="Arial"/><sz val="10"/><b/></font>'
            b'<font><name val="Calibri"/><sz val="16"/><i/></font>'
            b'<font><name val="Calibri"/><sz val="11"/><u/></font>'
            b'<font><name val="Calibri"/><sz val="11"/><color rgb="FFFF0000"/></font>'
            b'<font><name val="Calibri"/><sz val="10"/></font>'
            b'<font><name val="Calibri"/><sz val="10"/><b/><i/></font>'
            b'<font><name val="Calibri"/><sz val="11"/><family val="2"/></font>'
            b"</fonts>"
            b'<fills count="1"><fill><patternFill patternType="none"/></fill></fills>'
            b'<borders count="1"><border><left/><right/><top/><bottom/>'
            b"</border></borders>"
            b'<cellStyleXfs count="1"><xf numFmtId="0" fontId="0" fillId="0" '
            b'borderId="0"/></cellStyleXfs>'
            b'<cellXfs count="18">'
            + b"".join(
                (
                    f'<xf numFmtId="0" fontId="{font_id}" fillId="0" '
                    'borderId="0" xfId="0" applyFont="1"/>'
                ).encode()
                for font_id in [0, 1, 2, 3, 2, 3, 4, 1, 5, 6, 3, 1, 1, 1, 1, 3, 3, 0]
            )
            + b"</cellXfs>"
            b'<cellStyles count="1"><cellStyle name="Normal" xfId="0" builtinId="0"/>'
            b"</cellStyles></styleSheet>"
        ),
        "xl/_rels/workbook.xml.rels": (
            b'<?xml version="1.0" encoding="UTF-8"?>'
            b'<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
            b'<Relationship Id="rId1" Type="http://schemas.openxmlformats.'
            b'org/officeDocument/2006/relationships/worksheet" '
            b'Target="worksheets/sheet1.xml"/>'
            b'<Relationship Id="rId2" Type="http://schemas.openxmlformats.'
            b'org/officeDocument/2006/relationships/calcChain" '
            b'Target="calcChain.xml"/>'
            b"</Relationships>"
        ),
        "xl/calcChain.xml": (
            b'<?xml version="1.0" encoding="UTF-8"?>'
            b'<calcChain xmlns="http://schemas.openxmlformats.org/'
            b'spreadsheetml/2006/main">'
            b'<c r="I21" i="1" l="1"/><c r="I20" i="1"/>'
            b'<c r="I19" i="1"/><c r="I18" i="1"/>'
            b'<c r="I17" i="1"/><c r="I22" i="1" s="1"/>'
            b"</calcChain>"
        ),
        "xl/worksheets/sheet1.xml": sheet_xml,
        "xl/worksheets/_rels/sheet1.xml.rels": (
            b'<?xml version="1.0" encoding="UTF-8"?>'
            b'<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
            b'<Relationship Id="rId1" Type="http://schemas.openxmlformats.'
            b'org/officeDocument/2006/relationships/drawing" '
            b'Target="../drawings/drawing1.xml"/>'
            b"</Relationships>"
        ),
        "xl/drawings/drawing1.xml": (
            b'<?xml version="1.0" encoding="UTF-8"?>'
            b'<xdr:wsDr xmlns:xdr="http://schemas.openxmlformats.org/'
            b'drawingml/2006/spreadsheetDrawing" '
            b'xmlns:a="http://schemas.openxmlformats.org/drawingml/2006/main" '
            b'xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships">'
            b"<xdr:oneCellAnchor><xdr:from><xdr:col>0</xdr:col><xdr:colOff>0</xdr:colOff>"
            b"<xdr:row>1</xdr:row><xdr:rowOff>0</xdr:rowOff></xdr:from>"
            b'<xdr:ext cx="1" cy="1"/><xdr:pic><xdr:nvPicPr>'
            b'<xdr:cNvPr id="2" name="logo"/>'
            b'<xdr:cNvPicPr/></xdr:nvPicPr><xdr:blipFill><a:blip r:embed="rId1"/>'
            b"</xdr:blipFill><xdr:spPr/></xdr:pic><xdr:clientData/></xdr:oneCellAnchor>"
            b"</xdr:wsDr>"
        ),
        "xl/drawings/_rels/drawing1.xml.rels": (
            b'<?xml version="1.0" encoding="UTF-8"?>'
            b'<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
            b'<Relationship Id="rId1" Type="http://schemas.openxmlformats.'
            b'org/officeDocument/2006/relationships/image" '
            b'Target="../media/image1.png"/>'
            b"</Relationships>"
        ),
        "xl/media/image1.png": TINY_PNG,
    }


def write_package(path: Path, parts: dict[str, bytes] | None = None) -> None:
    with ZipFile(path, "w", ZIP_DEFLATED) as archive:
        for name, content in (package_parts() if parts is None else parts).items():
            archive.writestr(name, content)


def sha(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()


def output_path(tmp_path: Path) -> Path:
    directory = tmp_path / "out"
    directory.mkdir()
    return directory / "capacity100.xlsx"


def run_builder(
    source: Path,
    output: Path,
    expected_sha: str,
    *,
    project_root: Path | None = None,
) -> subprocess.CompletedProcess[str]:
    env = None
    if project_root is not None:
        env = {"CAPACITY100_BUILDER_PROJECT_ROOT": str(project_root)}
    return subprocess.run(
        [
            sys.executable,
            str(SCRIPT),
            "--source",
            str(source),
            "--output",
            str(output),
            "--expected-source-sha256",
            expected_sha,
        ],
        capture_output=True,
        text=True,
        check=False,
        env=env,
    )


def worksheet_root(path: Path) -> ElementTree.Element:
    with ZipFile(path) as archive:
        return ElementTree.fromstring(archive.read("xl/worksheets/sheet1.xml"))


def worksheet_bytes(path: Path) -> bytes:
    with ZipFile(path) as archive:
        return archive.read("xl/worksheets/sheet1.xml")


def cells(root: ElementTree.Element) -> dict[str, ElementTree.Element]:
    return {
        cast(str, cell_element.get("r")): cell_element
        for cell_element in root.findall(".//main:c", NS)
        if cell_element.get("r")
    }


def column_number(column: str) -> int:
    number = 0
    for letter in column:
        number = number * 26 + ord(letter) - ord("A") + 1
    return number


def split_ref(reference: str) -> tuple[str, int]:
    match = re.fullmatch(r"([A-Z]{1,3})([0-9]+)", reference)
    assert match is not None
    return match.group(1), int(match.group(2))


def cell_value(cell_element: ElementTree.Element, tag: str) -> str | None:
    child = cell_element.find(f"main:{tag}", NS)
    return None if child is None else child.text


def assert_strict_cell_records(path: Path) -> None:
    parts = zip_parts(path)
    root = ElementTree.fromstring(parts["xl/worksheets/sheet1.xml"])
    styles = ElementTree.fromstring(parts["xl/styles.xml"])
    cell_xfs = styles.find("main:cellXfs", NS)
    assert cell_xfs is not None
    style_count = len(cell_xfs.findall("main:xf", NS))
    shared_count = 0
    if "xl/sharedStrings.xml" in parts:
        shared_strings = ElementTree.fromstring(parts["xl/sharedStrings.xml"])
        shared_count = len(shared_strings.findall("main:si", NS))
    seen: set[str] = set()
    previous_row = 0
    for row_element in root.findall("main:sheetData/main:row", NS):
        row_number = int(cast(str, row_element.get("r")))
        assert row_number > previous_row
        previous_row = row_number
        previous_column = 0
        for cell_element in row_element.findall("main:c", NS):
            reference = cast(str, cell_element.get("r"))
            column, cell_row = split_ref(reference)
            assert cell_row == row_number
            current_column = column_number(column)
            assert current_column > previous_column
            previous_column = current_column
            assert reference not in seen
            seen.add(reference)

            style = cell_element.get("s")
            if style is not None:
                assert int(style) < style_count
            cell_type = cell_element.get("t")
            formula_element = cell_element.find("main:f", NS)
            value_element = cell_element.find("main:v", NS)
            inline_element = cell_element.find("main:is", NS)
            if formula_element is not None:
                assert cell_type is None
            if cell_type == "inlineStr":
                assert inline_element is not None
                assert inline_element.find("main:t", NS) is not None
            elif cell_type == "s":
                assert value_element is not None
                assert value_element.text is not None
                assert 0 <= int(value_element.text) < shared_count
            elif cell_type == "str":
                assert value_element is not None
            elif cell_type is not None:
                assert list(cell_element)


def zip_parts(path: Path) -> dict[str, bytes]:
    with ZipFile(path) as archive:
        return {name: archive.read(name) for name in archive.namelist()}


def tuned_style_bytes(styles_xml: bytes) -> bytes:
    return styles_xml.replace(b'val="10"', b'val="12"').replace(
        b'val="11"',
        b'val="12"',
    )


def assert_no_partial(output: Path) -> None:
    assert not output.exists()
    assert list(output.parent.iterdir()) == []


def test_cli_builds_capacity100_template_and_preserves_source(tmp_path: Path) -> None:
    source = tmp_path / "source.xlsx"
    output = output_path(tmp_path)
    write_package(source)
    source_sha = sha(source)
    before_snapshot = snapshot.build_drawing_media_snapshot(source)
    before_parts = zip_parts(source)

    result = run_builder(source, output, source_sha, project_root=tmp_path / "git")

    assert result.returncode == 0
    assert "CREATED:" in result.stdout
    assert output.is_file()
    assert sha(source) == source_sha
    assert_strict_cell_records(output)
    root = worksheet_root(output)
    dimension = root.find("main:dimension", NS)
    assert dimension is not None
    assert dimension.get("ref") == "A1:I131"
    workbook = load_workbook(output, data_only=False, read_only=True)
    sheet = workbook[SHEET_NAME]
    assert sheet.max_row == 131
    all_cells = cells(root)
    for row in range(17, 117):
        assert f"A{row}" in all_cells
        assert cell_value(all_cells[f"B{row}"], "v") == str(row - 16)
        assert sheet[f"D{row}"].value == "шт."
        formula_cell = all_cells[f"I{row}"]
        assert formula_cell.get("t") is None
        assert formula_cell.find("main:v", NS) is None
        assert cell_value(formula_cell, "f") == formula(row)
    for row in range(22, 117):
        for column in "CEFGH":
            element = all_cells[f"{column}{row}"]
            assert element.find("main:v", NS) is None
            assert element.find("main:f", NS) is None
            assert element.find("main:is", NS) is None
    assert all_cells["I117"].get("t") is None
    assert all_cells["I117"].find("main:v", NS) is None
    assert cell_value(all_cells["I117"], "f") == total_formula(17, 116)
    assert cell_value(all_cells["C119"], "t") is None
    assert sheet["C119"].value == "Всего прописью"
    merge_cells = root.find("main:mergeCells", NS)
    assert merge_cells is not None
    merge_refs = {
        cast(str, item.get("ref")) for item in merge_cells.findall("main:mergeCell", NS)
    }
    assert {
        "C119:I119",
        "C121:I121",
        "C122:I122",
        "C123:I123",
        "C124:I124",
        "C125:I125",
        "B127:E127",
        "F127:I127",
        "B128:E128",
        "F128:I128",
        "B129:I129",
    }.issubset(merge_refs)
    rows = root.findall(".//main:row", NS)
    for row_number in range(17, 117):
        row_element = next(row for row in rows if row.get("r") == str(row_number))
        assert row_element.get("ht") == "24"
        assert row_element.get("customHeight") == "1"
    assert all_cells["C22"].get("s") == all_cells["C19"].get("s")
    assert all_cells["F22"].get("s") == all_cells["F19"].get("s")
    assert b'mc:Ignorable="x14ac"' in worksheet_bytes(output)
    after_snapshot = snapshot.build_drawing_media_snapshot(output)
    snapshot.compare_drawing_media_snapshots(before_snapshot, after_snapshot)
    after_parts = zip_parts(output)
    assert "xl/calcChain.xml" not in after_parts
    assert b"calcChain" not in after_parts["[Content_Types].xml"]
    assert b"calcChain" not in after_parts["xl/_rels/workbook.xml.rels"]
    assert b'val="10"' not in after_parts["xl/styles.xml"]
    assert b'val="11"' not in after_parts["xl/styles.xml"]
    assert b'val="12"' in after_parts["xl/styles.xml"]
    assert after_parts["xl/styles.xml"] == tuned_style_bytes(
        before_parts["xl/styles.xml"]
    )
    assert set(after_parts) == set(before_parts) - {"xl/calcChain.xml"}
    allowed_changed = {
        "xl/worksheets/sheet1.xml",
        "[Content_Types].xml",
        "xl/styles.xml",
        "xl/_rels/workbook.xml.rels",
        "xl/calcChain.xml",
    }
    for name, content in before_parts.items():
        if name not in allowed_changed:
            assert after_parts[name] == content, name


def assert_builder_error(
    tmp_path: Path,
    parts: dict[str, bytes],
    expected: str,
) -> None:
    source = tmp_path / "source.xlsx"
    output = output_path(tmp_path)
    write_package(source, parts)
    result = run_builder(source, output, sha(source), project_root=tmp_path / "git")
    assert result.returncode == 1
    assert expected in result.stderr
    assert_no_partial(output)


def test_missing_source_fails_closed(tmp_path: Path) -> None:
    output = output_path(tmp_path)
    result = run_builder(
        tmp_path / "missing.xlsx",
        output,
        "abc",
        project_root=tmp_path / "git",
    )
    assert result.returncode == 1
    assert "source does not exist" in result.stderr
    assert_no_partial(output)


def test_wrong_expected_sha_fails_closed(tmp_path: Path) -> None:
    source = tmp_path / "source.xlsx"
    output = output_path(tmp_path)
    write_package(source)
    result = run_builder(source, output, "0" * 64, project_root=tmp_path / "git")
    assert result.returncode == 1
    assert "source SHA-256 mismatch" in result.stderr
    assert_no_partial(output)


def test_existing_output_fails_closed(tmp_path: Path) -> None:
    source = tmp_path / "source.xlsx"
    output = output_path(tmp_path)
    write_package(source)
    output.write_bytes(b"existing")
    result = run_builder(source, output, sha(source), project_root=tmp_path / "git")
    assert result.returncode == 1
    assert "output already exists" in result.stderr
    assert output.read_bytes() == b"existing"


def test_output_inside_git_fails_closed(tmp_path: Path) -> None:
    source = tmp_path / "source.xlsx"
    output = PROJECT_ROOT / "blocked_capacity100.xlsx"
    write_package(source)
    result = run_builder(source, output, sha(source), project_root=PROJECT_ROOT)
    assert result.returncode == 1
    assert "output is inside the Git project" in result.stderr
    assert not output.exists()


def test_output_matches_source_fails_closed(tmp_path: Path) -> None:
    source = tmp_path / "source.xlsx"
    write_package(source)
    result = run_builder(source, source, sha(source), project_root=tmp_path / "git")
    assert result.returncode == 1
    assert "output matches source" in result.stderr


def test_missing_output_directory_fails_closed(tmp_path: Path) -> None:
    source = tmp_path / "source.xlsx"
    output = tmp_path / "missing" / "out.xlsx"
    write_package(source)
    result = run_builder(source, output, sha(source), project_root=tmp_path / "git")
    assert result.returncode == 1
    assert "output parent directory does not exist" in result.stderr
    assert not output.exists()


def test_wrong_sheet_name_fails_closed(tmp_path: Path) -> None:
    parts = package_parts(workbook_name="Wrong")
    assert_builder_error(tmp_path, parts, "worksheet not found")


def test_wrong_dimension_fails_closed(tmp_path: Path) -> None:
    parts = package_parts(source_sheet_xml().replace(b"A1:I36", b"A1:I35", 1))
    assert_builder_error(tmp_path, parts, "unexpected source dimension")


def test_missing_row_19_fails_closed(tmp_path: Path) -> None:
    xml = re.sub(rb'<row r="19".*?</row>', b"", source_sheet_xml(), count=1)
    assert_builder_error(tmp_path, package_parts(xml), "source item row is missing: 19")


def test_structurally_different_item_rows_fail_closed(tmp_path: Path) -> None:
    xml = source_sheet_xml().replace(
        b'<row r="20" spans="1:9" ht="54"',
        b'<row r="20" spans="1:9" ht="40"',
        1,
    )
    assert_builder_error(
        tmp_path,
        package_parts(xml),
        "source item row has unexpected height: 20",
    )


def test_wrong_item_formula_fails_closed(tmp_path: Path) -> None:
    xml = source_sheet_xml().replace(formula(18).encode(), b"E18*H18", 1)
    assert_builder_error(tmp_path, package_parts(xml), "unexpected item formula in I18")


def test_wrong_total_formula_fails_closed(tmp_path: Path) -> None:
    xml = source_sheet_xml().replace(total_formula().encode(), b"SUM(I17:I21)", 1)
    assert_builder_error(
        tmp_path,
        package_parts(xml),
        "unexpected source total formula",
    )


def test_merge_crossing_insertion_boundary_fails_closed(tmp_path: Path) -> None:
    xml = source_sheet_xml().replace(
        b'<mergeCell ref="C24:I24"/>',
        b'<mergeCell ref="C20:I24"/>',
        1,
    )
    assert_builder_error(
        tmp_path,
        package_parts(xml),
        "merge range crosses insertion boundary",
    )


def test_unsupported_features_fail_closed(tmp_path: Path) -> None:
    xml = source_sheet_xml().replace(
        b"</worksheet>",
        b'<dataValidations count="0"/></worksheet>',
        1,
    )
    assert_builder_error(tmp_path, package_parts(xml), "unsupported worksheet feature")


def test_calc_chain_part_without_content_type_fails_closed(tmp_path: Path) -> None:
    parts = package_parts()
    parts["[Content_Types].xml"] = parts["[Content_Types].xml"].replace(
        b'<Override PartName="/xl/calcChain.xml" ContentType="application/'
        b'vnd.openxmlformats-officedocument.spreadsheetml.calcChain+xml"/>',
        b"",
        1,
    )
    assert_builder_error(
        tmp_path,
        parts,
        "calcChain content type override is missing",
    )


def test_calc_chain_part_without_relationship_fails_closed(tmp_path: Path) -> None:
    parts = package_parts()
    parts["xl/_rels/workbook.xml.rels"] = parts["xl/_rels/workbook.xml.rels"].replace(
        b'<Relationship Id="rId2" Type="http://schemas.openxmlformats.'
        b'org/officeDocument/2006/relationships/calcChain" '
        b'Target="calcChain.xml"/>',
        b"",
        1,
    )
    assert_builder_error(
        tmp_path,
        parts,
        "calcChain relationship is missing",
    )


def test_calc_chain_relationship_without_part_fails_closed(tmp_path: Path) -> None:
    parts = package_parts()
    del parts["xl/calcChain.xml"]
    assert_builder_error(
        tmp_path,
        parts,
        "calcChain relationship exists but calcChain part is missing",
    )


def test_calc_chain_unexpected_relationship_target_fails_closed(
    tmp_path: Path,
) -> None:
    parts = package_parts()
    parts["xl/_rels/workbook.xml.rels"] = parts["xl/_rels/workbook.xml.rels"].replace(
        b'Target="calcChain.xml"',
        b'Target="worksheets/calcChain.xml"',
        1,
    )
    assert_builder_error(
        tmp_path,
        parts,
        "calcChain relationship target is unexpected",
    )


def test_calc_chain_unsupported_structure_fails_closed(tmp_path: Path) -> None:
    parts = package_parts()
    parts["xl/calcChain.xml"] = parts["xl/calcChain.xml"].replace(
        b"<calcChain ",
        b'<calcChain unsupported="1" ',
        1,
    )
    assert_builder_error(tmp_path, parts, "calcChain structure is unsupported")


def test_invalid_content_types_fails_closed(tmp_path: Path) -> None:
    parts = package_parts()
    parts["[Content_Types].xml"] = b"<Types><broken></Types>"
    assert_builder_error(tmp_path, parts, "invalid XML part [Content_Types].xml")


def test_drawing_anchor_in_shifted_area_fails_closed(tmp_path: Path) -> None:
    parts = package_parts()
    parts["xl/drawings/drawing1.xml"] = parts["xl/drawings/drawing1.xml"].replace(
        b"<xdr:row>1</xdr:row>",
        b"<xdr:row>21</xdr:row>",
        1,
    )
    assert_builder_error(tmp_path, parts, "drawing anchor is in the shifted area")


def test_temporary_write_error_cleans_partial_output(
    tmp_path: Path,
    monkeypatch: Any,
) -> None:
    source = tmp_path / "source.xlsx"
    output = output_path(tmp_path)
    write_package(source)
    monkeypatch.setattr(builder, "PROJECT_ROOT", tmp_path / "git")

    def fail_write(
        output_parts: Any,
        temporary_output: Path,
    ) -> None:
        del output_parts
        temporary_output.write_bytes(b"partial")
        raise builder.CapacityTemplateBuilderError("forced temporary write failure")

    monkeypatch.setattr(builder, "write_package", fail_write)

    try:
        builder.build_capacity100_template(
            source=source,
            output=output,
            expected_source_sha256=sha(source),
        )
    except builder.CapacityTemplateBuilderError as error:
        assert "forced temporary write failure" in str(error)
    else:
        raise AssertionError("builder should fail")
    assert_no_partial(output)


def test_no_real_xlsx_or_manifest_files_are_added_to_git() -> None:
    status = subprocess.run(
        ["git", "status", "--short", "--untracked-files=all"],
        cwd=PROJECT_ROOT,
        capture_output=True,
        text=True,
        check=True,
    )
    changed_paths = [
        line[3:].strip() for line in status.stdout.splitlines() if line.strip()
    ]
    assert not any(
        path.endswith(".xlsx") and not path.startswith(".tmp_pytest")
        for path in changed_paths
    )
    assert not any(
        Path(path).name.casefold() == "manifest.json" for path in changed_paths
    )
