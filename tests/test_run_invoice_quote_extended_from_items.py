import importlib.util
import json
import subprocess
import sys
import tempfile
from pathlib import Path
from types import ModuleType
from typing import Any, cast

from openpyxl import Workbook, load_workbook  # type: ignore[import-untyped]

PROJECT_ROOT = Path(__file__).resolve().parents[1]
BRIDGE_SCRIPT = PROJECT_ROOT / "scripts" / "run_invoice_quote_extended_from_items.py"
SHEET_NAME = "Счёт-КП шаблон"


def load_script_module(module_name: str, path: Path) -> ModuleType:
    spec = importlib.util.spec_from_file_location(module_name, path)
    assert spec is not None
    assert spec.loader is not None
    module = importlib.util.module_from_spec(spec)
    sys.modules[spec.name] = module
    spec.loader.exec_module(module)
    return module


bridge = cast(
    Any,
    load_script_module("run_invoice_quote_extended_from_items_for_test", BRIDGE_SCRIPT),
)


def layout_json(capacity: int = 8) -> dict[str, Any]:
    item_start_row = 17
    item_end_row = item_start_row + capacity - 1
    total_row = item_start_row + capacity
    return {
        "item_start_row": item_start_row,
        "item_end_row": item_end_row,
        "capacity": capacity,
        "total_row": total_row,
        "signature_range": f"B{20 + capacity}:I{22 + capacity}",
        "header_ranges": ["C2:I6", "B4:B6"],
        "formula_cells": [f"I{row}" for row in range(17, 17 + capacity)]
        + [f"I{total_row}"],
    }


def write_extended_template(path: Path, capacity: int = 8) -> None:
    layout = layout_json(capacity)
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = SHEET_NAME

    for row in range(2, 7):
        for column in range(3, 10):
            worksheet.cell(row=row, column=column).value = f"header-{row}-{column}"
    for row in range(4, 7):
        worksheet.cell(row=row, column=2).value = f"req-{row}"

    worksheet.merge_cells(layout["signature_range"])
    worksheet.cell(row=20 + capacity, column=2).value = "signature"

    for row in range(layout["item_start_row"], layout["item_end_row"] + 1):
        worksheet[f"C{row}"] = "template item"
        worksheet[f"D{row}"] = "шт"
        worksheet[f"E{row}"] = 1
        worksheet[f"F{row}"] = "template instruments"
        worksheet[f"G{row}"] = "template cabinet"
        worksheet[f"H{row}"] = "template price"
        worksheet[f"I{row}"] = f"=E{row}*H{row}"
    worksheet[f"I{layout['total_row']}"] = (
        f"=SUM(I{layout['item_start_row']}:I{layout['item_end_row']})"
    )

    workbook.save(path)


def item(index: int) -> dict[str, Any]:
    return {
        "name": f"ВРУ-{index}",
        "unit": "шт.",
        "quantity": index,
        "instruments_and_devices": "нужно уточнить",
        "cabinet_type_dimensions_material": "нужно уточнить",
        "price_kzt": None,
        "price_confirmed_by_igor": False,
    }


def items_json(items_count: int) -> dict[str, Any]:
    return {"items": [item(index) for index in range(1, items_count + 1)]}


def write_json(path: Path, data: Any) -> None:
    path.write_text(json.dumps(data, ensure_ascii=False), encoding="utf-8")


def bridge_args(
    items_json_path: Path,
    template: Path,
    capacity: int,
    output: Path,
) -> list[str]:
    return [
        "--items-json",
        str(items_json_path),
        "--template",
        str(template),
        "--template-capacity",
        str(capacity),
        "--output",
        str(output),
    ]


def run_bridge(
    items_json_path: Path,
    template: Path,
    capacity: int,
    output: Path,
) -> subprocess.CompletedProcess[str]:
    return subprocess.run(
        [sys.executable, str(BRIDGE_SCRIPT)]
        + bridge_args(items_json_path, template, capacity, output),
        capture_output=True,
        text=True,
        check=False,
    )


def workbook_value(path: Path, cell: str) -> Any:
    workbook = load_workbook(path, data_only=False)
    worksheet = workbook[SHEET_NAME]
    return worksheet[cell].value


def test_bridge_successfully_runs_six_item_job(tmp_path: Path) -> None:
    template = tmp_path / "extended_template.xlsx"
    items_json_path = tmp_path / "items.json"
    output_dir = tmp_path / "out"
    output_dir.mkdir()
    output = output_dir / "draft.xlsx"
    write_extended_template(template)
    write_json(items_json_path, items_json(6))

    result = run_bridge(items_json_path, template, 8, output)

    assert result.returncode == 0
    assert "CREATED:" in result.stdout
    assert "ERROR:" not in result.stderr
    assert output.is_file()
    assert workbook_value(output, "C22") == "ВРУ-6"
    assert workbook_value(output, "I25") == "=SUM(I17:I24)"


def test_bridge_successfully_runs_fifty_item_job(tmp_path: Path) -> None:
    template = tmp_path / "extended_template.xlsx"
    items_json_path = tmp_path / "items.json"
    output_dir = tmp_path / "out"
    output_dir.mkdir()
    output = output_dir / "draft.xlsx"
    write_extended_template(template, capacity=50)
    write_json(items_json_path, items_json(50))

    result = run_bridge(items_json_path, template, 50, output)

    assert result.returncode == 0
    assert "CREATED:" in result.stdout
    assert "ERROR:" not in result.stderr
    assert output.is_file()
    assert workbook_value(output, "C66") == "ВРУ-50"
    assert workbook_value(output, "I67") == "=SUM(I17:I66)"


def test_bridge_missing_items_json_returns_one(tmp_path: Path, capsys: Any) -> None:
    template = tmp_path / "extended_template.xlsx"
    output = tmp_path / "out" / "draft.xlsx"
    output.parent.mkdir()
    write_extended_template(template)

    exit_code = bridge.main(
        bridge_args(tmp_path / "missing_items.json", template, 8, output)
    )

    captured = capsys.readouterr()
    assert exit_code == 1
    assert "ERROR:" in captured.err
    assert not output.exists()
    assert list(output.parent.iterdir()) == []


def test_bridge_invalid_items_json_returns_one(tmp_path: Path, capsys: Any) -> None:
    template = tmp_path / "extended_template.xlsx"
    items_json_path = tmp_path / "items.json"
    output = tmp_path / "out" / "draft.xlsx"
    output.parent.mkdir()
    write_extended_template(template)
    items_json_path.write_text("{invalid", encoding="utf-8")

    exit_code = bridge.main(bridge_args(items_json_path, template, 8, output))

    captured = capsys.readouterr()
    assert exit_code == 1
    assert "items JSON is invalid" in captured.err
    assert not output.exists()
    assert list(output.parent.iterdir()) == []


def test_bridge_root_json_must_be_object(tmp_path: Path, capsys: Any) -> None:
    template = tmp_path / "extended_template.xlsx"
    items_json_path = tmp_path / "items.json"
    output = tmp_path / "out" / "draft.xlsx"
    output.parent.mkdir()
    write_extended_template(template)
    write_json(items_json_path, [])

    exit_code = bridge.main(bridge_args(items_json_path, template, 8, output))

    captured = capsys.readouterr()
    assert exit_code == 1
    assert "items JSON must be an object" in captured.err
    assert not output.exists()


def test_bridge_requires_items(tmp_path: Path, capsys: Any) -> None:
    template = tmp_path / "extended_template.xlsx"
    items_json_path = tmp_path / "items.json"
    output = tmp_path / "out" / "draft.xlsx"
    output.parent.mkdir()
    write_extended_template(template)
    write_json(items_json_path, {})

    exit_code = bridge.main(bridge_args(items_json_path, template, 8, output))

    captured = capsys.readouterr()
    assert exit_code == 1
    assert "items is required" in captured.err
    assert not output.exists()


def test_bridge_rejects_non_list_items(tmp_path: Path, capsys: Any) -> None:
    template = tmp_path / "extended_template.xlsx"
    items_json_path = tmp_path / "items.json"
    output = tmp_path / "out" / "draft.xlsx"
    output.parent.mkdir()
    write_extended_template(template)
    write_json(items_json_path, {"items": {}})

    exit_code = bridge.main(bridge_args(items_json_path, template, 8, output))

    captured = capsys.readouterr()
    assert exit_code == 1
    assert "items must be a list" in captured.err
    assert not output.exists()


def test_bridge_rejects_empty_items(tmp_path: Path, capsys: Any) -> None:
    template = tmp_path / "extended_template.xlsx"
    items_json_path = tmp_path / "items.json"
    output = tmp_path / "out" / "draft.xlsx"
    output.parent.mkdir()
    write_extended_template(template)
    write_json(items_json_path, {"items": []})

    exit_code = bridge.main(bridge_args(items_json_path, template, 8, output))

    captured = capsys.readouterr()
    assert exit_code == 1
    assert "items must not be empty" in captured.err
    assert not output.exists()


def test_bridge_rejects_zero_capacity(tmp_path: Path, capsys: Any) -> None:
    template = tmp_path / "extended_template.xlsx"
    items_json_path = tmp_path / "items.json"
    output = tmp_path / "out" / "draft.xlsx"
    output.parent.mkdir()
    write_extended_template(template)
    write_json(items_json_path, items_json(1))

    exit_code = bridge.main(bridge_args(items_json_path, template, 0, output))

    captured = capsys.readouterr()
    assert exit_code == 1
    assert "template_capacity must be positive" in captured.err
    assert not output.exists()


def test_bridge_rejects_negative_capacity(tmp_path: Path, capsys: Any) -> None:
    template = tmp_path / "extended_template.xlsx"
    items_json_path = tmp_path / "items.json"
    output = tmp_path / "out" / "draft.xlsx"
    output.parent.mkdir()
    write_extended_template(template)
    write_json(items_json_path, items_json(1))

    exit_code = bridge.main(bridge_args(items_json_path, template, -1, output))

    captured = capsys.readouterr()
    assert exit_code == 1
    assert "template_capacity must be positive" in captured.err
    assert not output.exists()


def test_bridge_rejects_bool_capacity() -> None:
    try:
        bridge.validate_template_capacity(True)
    except bridge.ItemsBridgeError as error:
        assert "template_capacity must be a positive integer" in str(error)
    else:
        raise AssertionError("bool capacity should fail")


def test_bridge_rejects_overflow_before_downstream(tmp_path: Path, capsys: Any) -> None:
    template = tmp_path / "extended_template.xlsx"
    items_json_path = tmp_path / "items.json"
    output_dir = tmp_path / "out"
    output_dir.mkdir()
    output = output_dir / "draft.xlsx"
    write_extended_template(template, capacity=50)
    write_json(items_json_path, items_json(51))

    exit_code = bridge.main(bridge_args(items_json_path, template, 50, output))

    captured = capsys.readouterr()
    assert exit_code == 1
    assert "items count 51 exceeds template capacity 50" in captured.err
    assert not output.exists()
    assert list(output_dir.iterdir()) == []


def test_bridge_existing_output_is_not_overwritten(tmp_path: Path) -> None:
    template = tmp_path / "extended_template.xlsx"
    items_json_path = tmp_path / "items.json"
    output_dir = tmp_path / "out"
    output_dir.mkdir()
    output = output_dir / "draft.xlsx"
    output.write_bytes(b"existing")
    write_extended_template(template)
    write_json(items_json_path, items_json(6))

    result = run_bridge(items_json_path, template, 8, output)

    assert result.returncode == 1
    assert "ERROR:" in result.stderr
    assert "output already exists" in result.stderr
    assert output.read_bytes() == b"existing"
    assert list(output_dir.iterdir()) == [output]


def test_bridge_cleans_temp_job_after_success(
    tmp_path: Path,
    monkeypatch: Any,
) -> None:
    template = tmp_path / "extended_template.xlsx"
    items_json_path = tmp_path / "items.json"
    temp_parent = tmp_path / "bridge-temp"
    temp_parent.mkdir()
    output = tmp_path / "out" / "draft.xlsx"
    output.parent.mkdir()
    created_dirs: list[Path] = []
    write_extended_template(template)
    write_json(items_json_path, items_json(6))

    def temporary_directory(*args: Any, **kwargs: Any) -> Any:
        kwargs["dir"] = temp_parent
        manager = tempfile.TemporaryDirectory(*args, **kwargs)
        created_dirs.append(Path(manager.name))
        return manager

    monkeypatch.setattr(bridge, "TEMPORARY_DIRECTORY", temporary_directory)

    exit_code = bridge.main(bridge_args(items_json_path, template, 8, output))

    assert exit_code == 0
    assert created_dirs
    assert all(not path.exists() for path in created_dirs)
    assert list(temp_parent.iterdir()) == []


def test_bridge_cleans_temp_job_after_downstream_error(
    tmp_path: Path,
    monkeypatch: Any,
) -> None:
    template = tmp_path / "extended_template.xlsx"
    items_json_path = tmp_path / "items.json"
    temp_parent = tmp_path / "bridge-temp"
    temp_parent.mkdir()
    output_dir = tmp_path / "out"
    output_dir.mkdir()
    output = output_dir / "draft.xlsx"
    output.write_bytes(b"existing")
    created_dirs: list[Path] = []
    write_extended_template(template)
    write_json(items_json_path, items_json(6))

    def temporary_directory(*args: Any, **kwargs: Any) -> Any:
        kwargs["dir"] = temp_parent
        manager = tempfile.TemporaryDirectory(*args, **kwargs)
        created_dirs.append(Path(manager.name))
        return manager

    monkeypatch.setattr(bridge, "TEMPORARY_DIRECTORY", temporary_directory)

    exit_code = bridge.main(bridge_args(items_json_path, template, 8, output))

    assert exit_code == 1
    assert created_dirs
    assert all(not path.exists() for path in created_dirs)
    assert list(temp_parent.iterdir()) == []
    assert output.read_bytes() == b"existing"
    assert list(output_dir.iterdir()) == [output]


def test_no_real_xlsx_or_manifest_files_are_added_to_git() -> None:
    status = subprocess.run(
        ["git", "status", "--short", "--untracked-files=all"],
        cwd=PROJECT_ROOT,
        capture_output=True,
        text=True,
        check=True,
    )
    changed_paths = [
        line[3:].strip() for line in status.stdout.splitlines() if line.strip()
    ]
    assert not any(path.endswith(".xlsx") for path in changed_paths)
    assert not any(
        Path(path).name.casefold() == "manifest.json" for path in changed_paths
    )
