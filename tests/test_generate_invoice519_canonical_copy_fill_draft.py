from __future__ import annotations

import ast
import hashlib
import importlib.util
import json
import py_compile
import sys
from collections.abc import Callable, Mapping
from dataclasses import dataclass
from pathlib import Path
from types import ModuleType
from typing import Any, cast
from xml.etree import ElementTree
from zipfile import ZIP_DEFLATED, ZipFile

import pytest
from openpyxl import Workbook, load_workbook  # type: ignore[import-untyped]
from openpyxl.styles import Font, PatternFill  # type: ignore[import-untyped]

PROJECT_ROOT = Path(__file__).resolve().parents[1]
SCRIPT = PROJECT_ROOT / "scripts" / "generate_invoice519_canonical_copy_fill_draft.py"
RUNBOOK = PROJECT_ROOT / "docs" / "invoice519_canonical_copy_fill_draft_runbook.md"


def load_generator() -> ModuleType:
    spec = importlib.util.spec_from_file_location(
        "generate_invoice519_canonical_copy_fill_draft_for_test", SCRIPT
    )
    assert spec is not None and spec.loader is not None
    module = importlib.util.module_from_spec(spec)
    sys.modules[spec.name] = module
    spec.loader.exec_module(module)
    return module


writer = load_generator()


@dataclass(frozen=True)
class SyntheticCase:
    ledger: Path
    ledger_sha: str
    yauo: Path
    yauo_sha: str
    canonical: Path
    canonical_sha: str
    output_owner: Path


def sha256(raw: bytes) -> str:
    return hashlib.sha256(raw).hexdigest()


def encoded(payload: Mapping[str, Any]) -> bytes:
    return (json.dumps(payload, ensure_ascii=False, indent=2) + "\n").encode("utf-8")


def line_prices() -> list[int]:
    return [1] * 87 + [writer.APPROVED_TOTAL_KZT - 87]


def ledger_payload() -> dict[str, Any]:
    positions = []
    for position, (row, price) in enumerate(
        zip(writer.POSITION_ROWS, line_prices(), strict=True), start=1
    ):
        positions.append(
            {
                "invoice_position_number": position,
                "quantity": 1,
                "approved_unit_price_kzt": price,
                "approved_position_total_kzt": price,
                "currency": "KZT",
                "pricing_provenance": {
                    "partition": "FROZEN_55",
                    "evidence_role": "SYNTHETIC_TEST_EVIDENCE",
                    "source_binding_role": "pricing_profile",
                    "source_reference": f"POSITION-{position:03d}",
                    "allocation_method": "DIRECT_CHECKED_POSITION_PRICE",
                    "price_recalculated": False,
                },
                "technical_description_reference": {
                    "source_binding_role": "canonical_invoice_519",
                    "worksheet": writer.WORKSHEET,
                    "row": row,
                    "composition_status": (
                        "UNCHANGED_FROM_PRICE_APPLICATION_PREDECESSOR"
                    ),
                },
            }
        )
    return {
        "schema_version": "invoice519_commercial_pricing_ledger.v0.1",
        "status": (
            "IGOR_INVOICE519_88_POSITION_PRICING_LEDGER_READY_QUOTE_NOT_GENERATED"
        ),
        "application_status": "APPLIED",
        "positions": positions,
        "ledger_summary": {
            "position_count": 88,
            "approved_total_kzt": writer.APPROVED_TOTAL_KZT,
            "derived_line_total_kzt": writer.APPROVED_TOTAL_KZT,
            "frozen_55_subtotal_kzt": 11_963_792,
            "checked_missing_33_subtotal_kzt": 7_535_394,
            "duplicates": 0,
            "missing": 0,
            "extra": 0,
            "unit_price_allocation_used": False,
            "price_recalculation_used": False,
            "technical_composition_changed": False,
        },
        "reconciliation": {
            "coverage": {"covered": 88, "total": 88, "overlap": 0, "uncovered": 0},
            "combined_total_kzt": writer.APPROVED_TOTAL_KZT,
            "frozen_55_recalculated": False,
        },
        "price_grain": {
            "unit_prices_recalculated": False,
            "arbitrary_allocation_used": False,
        },
        "technical_composition": {"status": "UNCHANGED_FROM_PREDECESSOR"},
        "safety": {
            "quote_generation_authorized": False,
            "invoice_generation_authorized": False,
            "quote_or_invoice_publication_authorized": False,
            "client_send_authorized": False,
            "procurement_authorized": False,
            "reserve_authorized": False,
            "prepayment_authorized": False,
            "production_authorized": False,
            "downstream_authorized": False,
        },
    }


def yauo_payload() -> dict[str, Any]:
    return {
        "schema_version": "invoice519_yauo_enclosure_human_decision.v0.1",
        "status": ("IGOR_INVOICE519_YAUO_ENCLOSURE_APPROVED_NOT_APPLIED_TO_QUOTE"),
        "approval_scope": "ENCLOSURE_DIMENSIONS_ONLY",
        "quote_application_status": "NOT_APPLIED",
        "technical_decision": {
            "invoice_position_number": 87,
            "product_identity": "YAUO9601_3474",
            "field": "enclosure_dimensions",
            "previous_value": "450×300×250 mm",
            "approved_value": "400×300×250 mm",
            "change_scope": "POSITION_87_ENCLOSURE_ONLY",
            "quote_application_status": "NOT_APPLIED",
        },
        "safety": {
            "human_decision_recorded": True,
            "technical_decision_recorded": True,
            "quote_application_authorized": False,
            "quote_generation_authorized": False,
            "invoice_generation_authorized": False,
            "quote_publication_authorized": False,
            "invoice_publication_authorized": False,
            "client_send_authorized": False,
            "procurement_authorized": False,
            "reserve_authorized": False,
            "prepayment_authorized": False,
            "production_authorized": False,
            "downstream_authorized": False,
        },
    }


def write_canonical(path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = writer.WORKSHEET
    workbook.create_sheet("Лист2")["A1"] = "preserve auxiliary sheet 2"
    workbook.create_sheet("Лист3")["A1"] = "preserve auxiliary sheet 3"
    sheet.merge_cells("C9:F9")
    sheet.freeze_panes = "C17"
    sheet.print_title_rows = "1:15"
    sheet.print_options.horizontalCentered = True
    sheet.page_setup.orientation = "landscape"
    sheet.page_setup.fitToWidth = 1
    sheet.column_dimensions["C"].width = 28
    sheet.row_dimensions[17].height = 31
    sheet[writer.LEAD_TIME_CELL] = writer.CANONICAL_LEAD_TIME_VALUE
    sheet[writer.AMOUNT_WORDS_CELL] = writer.CANONICAL_AMOUNT_WORDS_VALUE
    sheet[writer.TOTAL_CELL] = writer.CANONICAL_TOTAL_VALUE
    for section_number, row in enumerate(writer.SECTION_ROWS, start=1):
        sheet[f"C{row}"] = f"Секция {section_number}"
    for position, row in enumerate(writer.POSITION_ROWS, start=1):
        sheet[f"B{row}"] = position
        sheet[f"C{row}"] = f"Позиция {position} — подробное описание шкафа"
        sheet[f"D{row}"] = "шт."
        sheet[f"E{row}"] = 1
        sheet[f"F{row}"] = f"Подробный технический состав {position}"
        sheet[f"G{row}"] = f"Корпус {position}"
        sheet[f"H{row}"] = 1000 + position
        sheet[f"I{row}"] = f"=H{row}*E{row}"
        for column in "BCDEFGHI":
            sheet[f"{column}{row}"].font = Font(name="Arial", size=9, bold=True)
            sheet[f"{column}{row}"].fill = PatternFill(
                fill_type="solid", fgColor="FFF2CC"
            )
    sheet[writer.YAUO_CELL] = writer.CANONICAL_YAUO_VALUE
    workbook.save(path)


def prepare_case(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> SyntheticCase:
    inputs = tmp_path / "inputs"
    inputs.mkdir()
    ledger = inputs / "ledger.json"
    ledger_raw = encoded(ledger_payload())
    ledger.write_bytes(ledger_raw)
    yauo = inputs / "yauo.json"
    yauo_raw = encoded(yauo_payload())
    yauo.write_bytes(yauo_raw)
    canonical = inputs / "canonical.xlsx"
    write_canonical(canonical)
    canonical_raw = canonical.read_bytes()
    output_owner = tmp_path / "cases"
    output_owner.mkdir()
    monkeypatch.setattr(writer, "LEDGER_PATH", ledger.resolve())
    monkeypatch.setattr(writer, "LEDGER_SHA256", sha256(ledger_raw))
    monkeypatch.setattr(writer, "YAUO_DECISION_PATH", yauo.resolve())
    monkeypatch.setattr(writer, "YAUO_DECISION_SHA256", sha256(yauo_raw))
    monkeypatch.setattr(writer, "CANONICAL_WORKBOOK_PATH", canonical.resolve())
    monkeypatch.setattr(writer, "CANONICAL_WORKBOOK_SHA256", sha256(canonical_raw))
    monkeypatch.setattr(writer, "REPO_ROOT", tmp_path / "unrelated-repository")
    monkeypatch.setattr(
        writer.ledger_publisher, "validate_payload", lambda _value: None
    )
    monkeypatch.setattr(writer.yauo_publisher, "validate_payload", lambda _value: None)
    return SyntheticCase(
        ledger,
        sha256(ledger_raw),
        yauo,
        sha256(yauo_raw),
        canonical,
        sha256(canonical_raw),
        output_owner,
    )


def output_path(case: SyntheticCase, name: str = "new-case") -> Path:
    return case.output_owner / name / cast(str, writer.OUTPUT_FILENAME)


def arguments(case: SyntheticCase, output: Path, token: str) -> list[str]:
    return [
        "--commercial-pricing-ledger",
        str(case.ledger),
        "--commercial-pricing-ledger-sha256",
        case.ledger_sha,
        "--yauo-enclosure-human-decision",
        str(case.yauo),
        "--yauo-enclosure-human-decision-sha256",
        case.yauo_sha,
        "--canonical-invoice-519",
        str(case.canonical),
        "--canonical-invoice-519-sha256",
        case.canonical_sha,
        "--output",
        str(output),
        "--authorization",
        token,
    ]


def zip_parts(path: Path) -> dict[str, bytes]:
    with ZipFile(path) as archive:
        return {name: archive.read(name) for name in archive.namelist()}


def worksheet_xml(path: Path) -> bytes:
    with ZipFile(path) as archive:
        part = writer.patcher.worksheet_part_for_sheet(archive, writer.WORKSHEET)
        return cast(bytes, archive.read(part))


def masked_cells(raw: bytes, coordinates: set[str]) -> bytes:
    masked = bytearray(raw)
    ranges = writer.patcher.cell_ranges(raw)
    replacements = []
    for coordinate in coordinates:
        matches = ranges[coordinate]
        assert len(matches) == 1
        replacements.append((matches[0].start, matches[0].end, coordinate))
    for start, end, coordinate in sorted(replacements, reverse=True):
        masked[start:end] = f"__CELL_{coordinate}__".encode("ascii")
    return bytes(masked)


def styles(path: Path, coordinates: set[str]) -> dict[str, str | None]:
    root = ElementTree.fromstring(worksheet_xml(path))
    cells = {
        cell.get("r"): cell
        for cell in root.findall(".//main:c", writer.patcher.NS)
        if cell.get("r") is not None
    }
    return {coordinate: cells[coordinate].get("s") for coordinate in coordinates}


def mutate_json(path: Path, mutation: Callable[[dict[str, Any]], None]) -> str:
    payload = json.loads(path.read_text(encoding="utf-8"))
    mutation(payload)
    raw = encoded(payload)
    path.write_bytes(raw)
    return sha256(raw)


def test_script_and_test_compile(tmp_path: Path) -> None:
    py_compile.compile(str(SCRIPT), cfile=str(tmp_path / "writer.pyc"), doraise=True)
    py_compile.compile(
        str(Path(__file__)), cfile=str(tmp_path / "test.pyc"), doraise=True
    )


def test_positive_copy_fill_preserves_canonical_package_and_styles(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
    capsys: pytest.CaptureFixture[str],
) -> None:
    case = prepare_case(tmp_path, monkeypatch)
    output = output_path(case)
    before_parts = zip_parts(case.canonical)
    allowlist = set(writer.build_updates(writer.ledger_lines(ledger_payload())))
    before_styles = styles(case.canonical, allowlist)
    assert writer.main(arguments(case, output, writer.GENERATION_AUTHORIZATION)) == 0
    report = capsys.readouterr().out
    assert "GENERATED_LOCAL_DRAFT_IMMUTABLE_NO_OVERWRITE" in report
    assert "MODIFIED_CELLS=180" in report
    assert "CLIENT_SEND=CLOSED" in report
    assert list(output.parent.iterdir()) == [output]

    after_parts = zip_parts(output)
    changed_part = writer.load_canonical(
        case.canonical, case.canonical_sha
    ).worksheet_part
    assert set(after_parts) == set(before_parts)
    for name, raw in before_parts.items():
        if name != changed_part:
            assert after_parts[name] == raw, name
    assert masked_cells(before_parts[changed_part], allowlist) == masked_cells(
        after_parts[changed_part], allowlist
    )
    assert styles(output, allowlist) == before_styles

    workbook = load_workbook(output, data_only=False, read_only=False)
    try:
        sheet = workbook[writer.WORKSHEET]
        for row, price in zip(writer.POSITION_ROWS, line_prices(), strict=True):
            assert sheet[f"H{row}"].value == price
            assert sheet[f"I{row}"].value == price
        assert sheet[writer.TOTAL_CELL].value == writer.APPROVED_TOTAL_KZT
        assert (
            sheet[writer.AMOUNT_WORDS_CELL].value == writer.APPROVED_AMOUNT_WORDS_VALUE
        )
        assert sheet[writer.LEAD_TIME_CELL].value == writer.APPROVED_LEAD_TIME_VALUE
        assert sheet[writer.YAUO_CELL].value == writer.APPROVED_YAUO_VALUE
        assert tuple(str(value) for value in sheet.merged_cells.ranges) == ("C9:F9",)
        assert sheet.freeze_panes == "C17"
        assert sheet.print_title_rows == "$1:$15"
        assert sheet.page_setup.orientation == "landscape"
    finally:
        workbook.close()


def test_exact_cell_map_allowlist_and_amount_words() -> None:
    lines = writer.ledger_lines(ledger_payload())
    updates = writer.build_updates(lines)
    assert len(lines) == 88
    assert tuple(line.row for line in lines) == writer.POSITION_ROWS
    assert writer.SECTION_ROWS == (32, 39, 57, 64, 79, 86, 103, 110)
    assert len(updates) == 180
    assert updates["I113"] == 19_499_186
    assert updates["G10"] == "Срок изготовления 30–40 рабочих дней"
    assert updates["G111"] == "Накладной 400х300х250 металл 1,2мм"
    assert "Девятнадцать миллионов" in cast(str, updates["C115"])


def test_authorization_fails_before_generation(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    case = prepare_case(tmp_path, monkeypatch)
    output = output_path(case)
    with pytest.raises(writer.GeneratorError, match="exact Invoice 519"):
        writer.main(arguments(case, output, "WRONG"))
    assert not output.parent.exists()


@pytest.mark.parametrize("binding", ["ledger", "yauo", "canonical"])
def test_wrong_input_path_fails_closed(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch, binding: str
) -> None:
    case = prepare_case(tmp_path, monkeypatch)
    alternate = tmp_path / f"alternate-{binding}"
    source = cast(Path, getattr(case, binding))
    alternate.write_bytes(source.read_bytes())
    values = {
        "ledger_path": case.ledger,
        "ledger_sha256": case.ledger_sha,
        "yauo_decision_path": case.yauo,
        "yauo_decision_sha256": case.yauo_sha,
        "canonical_workbook": case.canonical,
        "canonical_workbook_sha256": case.canonical_sha,
        "output": output_path(case),
    }
    key = {
        "ledger": "ledger_path",
        "yauo": "yauo_decision_path",
        "canonical": "canonical_workbook",
    }[binding]
    values[key] = alternate
    with pytest.raises(writer.GeneratorError, match="path binding mismatch"):
        writer.generate_draft(**values)
    assert not output_path(case).parent.exists()


@pytest.mark.parametrize("binding", ["ledger", "yauo", "canonical"])
def test_wrong_sha_argument_fails_closed(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch, binding: str
) -> None:
    case = prepare_case(tmp_path, monkeypatch)
    values = {
        "ledger_path": case.ledger,
        "ledger_sha256": case.ledger_sha,
        "yauo_decision_path": case.yauo,
        "yauo_decision_sha256": case.yauo_sha,
        "canonical_workbook": case.canonical,
        "canonical_workbook_sha256": case.canonical_sha,
        "output": output_path(case),
    }
    key = {
        "ledger": "ledger_sha256",
        "yauo": "yauo_decision_sha256",
        "canonical": "canonical_workbook_sha256",
    }[binding]
    values[key] = "0" * 64
    with pytest.raises(writer.GeneratorError, match="SHA"):
        writer.generate_draft(**values)
    assert not output_path(case).parent.exists()


def test_upstream_strict_contract_failure_is_wrapped(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    case = prepare_case(tmp_path, monkeypatch)

    def reject(_payload: Any) -> None:
        raise ValueError("schema drift")

    monkeypatch.setattr(writer.ledger_publisher, "validate_payload", reject)
    with pytest.raises(
        writer.GeneratorError, match="strict contract failed: schema drift"
    ):
        writer.generate_draft(
            ledger_path=case.ledger,
            ledger_sha256=case.ledger_sha,
            yauo_decision_path=case.yauo,
            yauo_decision_sha256=case.yauo_sha,
            canonical_workbook=case.canonical,
            canonical_workbook_sha256=case.canonical_sha,
            output=output_path(case),
        )


@pytest.mark.parametrize(
    ("field", "bad_value", "message"),
    [
        ("status", "WRONG", "ledger status"),
        ("application_status", "NOT_APPLIED", "ledger not APPLIED"),
        ("ledger_summary.approved_total_kzt", 1, "approved_total_kzt"),
        (
            "ledger_summary.unit_price_allocation_used",
            True,
            "unit_price_allocation_used",
        ),
        ("reconciliation.coverage.overlap", 1, "coverage"),
        ("reconciliation.frozen_55_recalculated", True, "frozen 55"),
        ("price_grain.unit_prices_recalculated", True, "price calculation"),
        ("technical_composition.status", "CHANGED", "technical composition"),
        ("safety.client_send_authorized", True, "client_send_authorized"),
    ],
)
def test_ledger_boundaries_fail_closed(
    field: str, bad_value: Any, message: str
) -> None:
    payload = ledger_payload()
    target: dict[str, Any] = payload
    keys = field.split(".")
    for key in keys[:-1]:
        target = cast(dict[str, Any], target[key])
    target[keys[-1]] = bad_value
    with pytest.raises(writer.GeneratorError, match=message):
        writer.validate_ledger_payload(payload)


@pytest.mark.parametrize(
    ("mutation", "message"),
    [
        (lambda values: values.pop(), "count"),
        (
            lambda values: values[0].__setitem__("invoice_position_number", 2),
            "membership/order",
        ),
        (
            lambda values: values[0].__setitem__("approved_position_total_kzt", 2),
            "multiplicity",
        ),
        (
            lambda values: values[0]["pricing_provenance"].__setitem__(
                "price_recalculated", True
            ),
            "recalculated",
        ),
        (
            lambda values: values[0]["pricing_provenance"].__setitem__(
                "allocation_method", "PROPORTIONAL"
            ),
            "arbitrary allocation",
        ),
        (
            lambda values: values[0]["technical_description_reference"].__setitem__(
                "row", 18
            ),
            "row mapping",
        ),
    ],
)
def test_ledger_line_membership_price_and_provenance_fail_closed(
    mutation: Callable[[list[dict[str, Any]]], None], message: str
) -> None:
    payload = ledger_payload()
    positions = cast(list[dict[str, Any]], payload["positions"])
    mutation(positions)
    with pytest.raises(writer.GeneratorError, match=message):
        writer.ledger_lines(payload)


@pytest.mark.parametrize(
    ("field", "bad_value", "message"),
    [
        ("status", "WRONG", "YAUO status"),
        ("approval_scope", "ALL", "approval scope"),
        ("quote_application_status", "APPLIED", "quote application"),
        ("technical_decision.invoice_position_number", 88, "exact decision"),
        ("technical_decision.product_identity", "OTHER", "exact decision"),
        ("technical_decision.previous_value", "450×300×200 mm", "exact decision"),
        ("technical_decision.approved_value", "400×300×200 mm", "exact decision"),
        ("safety.client_send_authorized", True, "safety boundary"),
    ],
)
def test_yauo_boundaries_fail_closed(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
    field: str,
    bad_value: Any,
    message: str,
) -> None:
    case = prepare_case(tmp_path, monkeypatch)

    def mutate(payload: dict[str, Any]) -> None:
        target = payload
        keys = field.split(".")
        for key in keys[:-1]:
            target = cast(dict[str, Any], target[key])
        target[keys[-1]] = bad_value

    digest = mutate_json(case.yauo, mutate)
    monkeypatch.setattr(writer, "YAUO_DECISION_SHA256", digest)
    with pytest.raises(writer.GeneratorError, match=message):
        writer.load_yauo_decision(case.yauo, digest)


@pytest.mark.parametrize(
    ("coordinate", "bad_value", "message"),
    [
        ("E17", 2, "quantity mismatch"),
        ("B17", 2, "position mismatch"),
        ("G10", "Срок изготовления неизвестен", "lead-time"),
        ("G111", "Накладной 400х300х250 металл 1,2мм", "YAUO"),
        ("C115", "Другая сумма", "amount-words"),
        ("I113", 1, "total cell"),
    ],
)
def test_canonical_semantic_drift_fails_closed(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
    coordinate: str,
    bad_value: Any,
    message: str,
) -> None:
    case = prepare_case(tmp_path, monkeypatch)
    workbook = load_workbook(case.canonical)
    workbook[writer.WORKSHEET][coordinate] = bad_value
    workbook.save(case.canonical)
    workbook.close()
    digest = sha256(case.canonical.read_bytes())
    monkeypatch.setattr(writer, "CANONICAL_WORKBOOK_SHA256", digest)
    canonical = writer.load_canonical(case.canonical, digest)
    with pytest.raises(writer.GeneratorError, match=message):
        writer.validate_canonical_map(canonical, writer.ledger_lines(ledger_payload()))


def test_invalid_canonical_package_fails_closed(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    case = prepare_case(tmp_path, monkeypatch)
    case.canonical.write_bytes(b"not-an-xlsx")
    digest = sha256(case.canonical.read_bytes())
    monkeypatch.setattr(writer, "CANONICAL_WORKBOOK_SHA256", digest)
    with pytest.raises(writer.GeneratorError, match="valid XLSX"):
        writer.load_canonical(case.canonical, digest)


def test_output_filename_directory_collision_and_repo_policy(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    case = prepare_case(tmp_path, monkeypatch)
    with pytest.raises(writer.GeneratorError, match="filename"):
        writer.validate_output_path(case.output_owner / "new" / "wrong.xlsx")
    existing = case.output_owner / "existing"
    existing.mkdir()
    with pytest.raises(writer.GeneratorError, match="already exists"):
        writer.validate_output_path(existing / writer.OUTPUT_FILENAME)
    monkeypatch.setattr(writer, "REPO_ROOT", case.output_owner)
    with pytest.raises(writer.GeneratorError, match="outside repository"):
        writer.validate_output_path(
            case.output_owner / "inside" / writer.OUTPUT_FILENAME
        )


def test_toctou_failure_rolls_back_new_directory(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    case = prepare_case(tmp_path, monkeypatch)
    output = output_path(case)
    original = writer.verify_inputs_unchanged
    calls = 0

    def fail_second(inputs: Any) -> None:
        nonlocal calls
        calls += 1
        if calls == 2:
            raise writer.GeneratorError("TOCTOU bytes changed: ledger")
        original(inputs)

    monkeypatch.setattr(writer, "verify_inputs_unchanged", fail_second)
    with pytest.raises(writer.GeneratorError, match="TOCTOU bytes changed"):
        writer.generate_draft(
            ledger_path=case.ledger,
            ledger_sha256=case.ledger_sha,
            yauo_decision_path=case.yauo,
            yauo_decision_sha256=case.yauo_sha,
            canonical_workbook=case.canonical,
            canonical_workbook_sha256=case.canonical_sha,
            output=output,
        )
    assert not output.parent.exists()


def test_candidate_validation_failure_rolls_back(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    case = prepare_case(tmp_path, monkeypatch)
    output = output_path(case)

    def reject(*_args: Any, **_kwargs: Any) -> bytes:
        raise writer.GeneratorError("candidate rejected")

    monkeypatch.setattr(writer, "validate_candidate", reject)
    with pytest.raises(writer.GeneratorError, match="candidate rejected"):
        writer.generate_draft(
            ledger_path=case.ledger,
            ledger_sha256=case.ledger_sha,
            yauo_decision_path=case.yauo,
            yauo_decision_sha256=case.yauo_sha,
            canonical_workbook=case.canonical,
            canonical_workbook_sha256=case.canonical_sha,
            output=output,
        )
    assert not output.parent.exists()


def test_post_link_validation_failure_rolls_back_final_and_staging(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    case = prepare_case(tmp_path, monkeypatch)
    output = output_path(case)
    original = writer.validate_candidate
    calls = 0

    def fail_second(*args: Any, **kwargs: Any) -> bytes:
        nonlocal calls
        calls += 1
        if calls == 2:
            raise writer.GeneratorError("post-link rejected")
        return cast(bytes, original(*args, **kwargs))

    monkeypatch.setattr(writer, "validate_candidate", fail_second)
    with pytest.raises(writer.GeneratorError, match="post-link rejected"):
        writer.generate_draft(
            ledger_path=case.ledger,
            ledger_sha256=case.ledger_sha,
            yauo_decision_path=case.yauo,
            yauo_decision_sha256=case.yauo_sha,
            canonical_workbook=case.canonical,
            canonical_workbook_sha256=case.canonical_sha,
            output=output,
        )
    assert not output.parent.exists()


def test_atomic_link_failure_rolls_back(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    case = prepare_case(tmp_path, monkeypatch)
    output = output_path(case)

    def reject_link(_source: Any, _target: Any) -> None:
        raise OSError("link denied")

    monkeypatch.setattr(writer.os, "link", reject_link)
    with pytest.raises(writer.GeneratorError, match="atomic no-overwrite"):
        writer.generate_draft(
            ledger_path=case.ledger,
            ledger_sha256=case.ledger_sha,
            yauo_decision_path=case.yauo,
            yauo_decision_sha256=case.yauo_sha,
            canonical_workbook=case.canonical,
            canonical_workbook_sha256=case.canonical_sha,
            output=output,
        )
    assert not output.parent.exists()


def test_existing_output_is_never_overwritten(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    case = prepare_case(tmp_path, monkeypatch)
    output = output_path(case)
    output.parent.mkdir()
    output.write_bytes(b"preserve")
    with pytest.raises(writer.GeneratorError, match="directory already exists"):
        writer.generate_draft(
            ledger_path=case.ledger,
            ledger_sha256=case.ledger_sha,
            yauo_decision_path=case.yauo,
            yauo_decision_sha256=case.yauo_sha,
            canonical_workbook=case.canonical,
            canonical_workbook_sha256=case.canonical_sha,
            output=output,
        )
    assert output.read_bytes() == b"preserve"


def test_source_and_runbook_keep_real_generation_and_downstream_closed() -> None:
    source = SCRIPT.read_text(encoding="utf-8")
    tree = ast.parse(source)
    imports = {
        alias.name
        for node in ast.walk(tree)
        if isinstance(node, ast.Import)
        for alias in node.names
    }
    assert not imports.intersection({"requests", "smtplib", "subprocess"})
    assert "pdf" not in source.casefold()
    assert "CLIENT_SEND=CLOSED" in source
    runbook = RUNBOOK.read_text(encoding="utf-8")
    assert writer.GENERATION_AUTHORIZATION in runbook
    assert "Do not invoke it during implementation" in runbook
    assert "180-cell allowlist" in runbook
    assert not any(PROJECT_ROOT.glob("*.xlsx"))


def test_candidate_rejects_unexpected_part_and_style_change(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    case = prepare_case(tmp_path, monkeypatch)
    canonical = writer.load_canonical(case.canonical, case.canonical_sha)
    updates = writer.build_updates(writer.ledger_lines(ledger_payload()))
    expected = writer.expected_patched_worksheet(canonical, updates)
    candidate = tmp_path / "candidate.xlsx"
    writer.patcher.write_patched_package(
        canonical.parts, canonical.worksheet_part, expected, candidate
    )
    rewritten = tmp_path / "rewritten.xlsx"
    with ZipFile(candidate) as source, ZipFile(rewritten, "w", ZIP_DEFLATED) as target:
        for name in source.namelist():
            target.writestr(name, source.read(name))
        target.writestr("unexpected.xml", "<unexpected/>")
    with pytest.raises(writer.GeneratorError, match="part membership"):
        writer.validate_candidate(rewritten, canonical, expected, updates)

    cells = writer.patcher.cell_ranges(expected)
    target_cell = cells["H17"][0]
    changed = (
        expected[: target_cell.start]
        + target_cell.xml.replace(b's="', b's="999', 1)
        + expected[target_cell.end :]
    )
    writer.patcher.write_patched_package(
        canonical.parts, canonical.worksheet_part, changed, candidate
    )
    with pytest.raises(writer.GeneratorError, match="worksheet patch"):
        writer.validate_candidate(candidate, canonical, expected, updates)


def test_verify_inputs_unchanged_reports_missing_file(tmp_path: Path) -> None:
    missing = tmp_path / "missing"
    with pytest.raises(writer.GeneratorError, match="TOCTOU reread failed"):
        writer.verify_inputs_unchanged(((missing, b"expected", "missing input"),))


def test_rollback_preserves_foreign_final(tmp_path: Path) -> None:
    directory = tmp_path / "new"
    directory.mkdir()
    output = directory / writer.OUTPUT_FILENAME
    output.write_bytes(b"foreign")
    blockers = writer._rollback(output, None, True, (999, 999))
    assert blockers[0] == "foreign final replacement preserved"
    assert "output directory cleanup failed" in blockers[1]
    assert output.read_bytes() == b"foreign"


def test_output_owner_must_exist(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    monkeypatch.setattr(writer, "REPO_ROOT", tmp_path / "repo")
    output = tmp_path / "missing-owner" / "case" / writer.OUTPUT_FILENAME
    with pytest.raises(writer.GeneratorError, match="owner must already exist"):
        writer.validate_output_path(output)


def test_duplicate_json_key_fails_strict_loading(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    case = prepare_case(tmp_path, monkeypatch)
    case.ledger.write_text('{"status":"A","status":"B"}', encoding="utf-8")
    digest = sha256(case.ledger.read_bytes())
    monkeypatch.setattr(writer, "LEDGER_SHA256", digest)
    with pytest.raises(writer.GeneratorError, match="strict contract failed"):
        writer.load_ledger(case.ledger, digest)


def test_actual_input_byte_drift_fails_before_directory_creation(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    case = prepare_case(tmp_path, monkeypatch)
    case.yauo.write_bytes(case.yauo.read_bytes() + b"\n")
    with pytest.raises(writer.GeneratorError, match="SHA-256 mismatch"):
        writer.generate_draft(
            ledger_path=case.ledger,
            ledger_sha256=case.ledger_sha,
            yauo_decision_path=case.yauo,
            yauo_decision_sha256=case.yauo_sha,
            canonical_workbook=case.canonical,
            canonical_workbook_sha256=case.canonical_sha,
            output=output_path(case),
        )
    assert not output_path(case).parent.exists()


def test_generated_result_sha_size_and_allowlist(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    case = prepare_case(tmp_path, monkeypatch)
    output = output_path(case)
    result = writer.generate_draft(
        ledger_path=case.ledger,
        ledger_sha256=case.ledger_sha,
        yauo_decision_path=case.yauo,
        yauo_decision_sha256=case.yauo_sha,
        canonical_workbook=case.canonical,
        canonical_workbook_sha256=case.canonical_sha,
        output=output,
    )
    raw = output.read_bytes()
    assert result.output == output.resolve()
    assert result.sha256 == sha256(raw)
    assert result.size == len(raw)
    assert len(result.modified_cells) == 180
    assert result.modified_cells == tuple(sorted(result.modified_cells))


def test_os_link_is_no_overwrite_primitive() -> None:
    source = SCRIPT.read_text(encoding="utf-8")
    assert "os.link(staging, output_path)" in source
    assert "replace(output_path)" not in source
    assert "rename(output_path)" not in source
