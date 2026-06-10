import importlib.util
import re
import subprocess
import sys
from pathlib import Path
from types import ModuleType
from typing import Any, cast
from xml.etree import ElementTree
from zipfile import ZIP_DEFLATED, ZipFile

from openpyxl import Workbook, load_workbook  # type: ignore[import-untyped]
from openpyxl.styles import Font  # type: ignore[import-untyped]

PROJECT_ROOT = Path(__file__).resolve().parents[1]
PATCHER_SCRIPT = PROJECT_ROOT / "scripts" / "ooxml_cell_patcher.py"
SNAPSHOT_SCRIPT = PROJECT_ROOT / "scripts" / "drawing_media_snapshot.py"
SHEET_NAME = "Счёт-КП шаблон"
SPREADSHEET_NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
NS = {"main": SPREADSHEET_NS}
TARGET_CELLS = {"C17", "D17", "E17", "F17", "G17", "H17"}
TINY_PNG = (
    b"\x89PNG\r\n\x1a\n\x00\x00\x00\rIHDR\x00\x00\x00\x01"
    b"\x00\x00\x00\x01\x08\x06\x00\x00\x00\x1f\x15\xc4\x89"
    b"\x00\x00\x00\rIDATx\x9cc\xf8\xff\xff?\x00\x05\xfe"
    b"\x02\xfeA\xe2d\x9a\x00\x00\x00\x00IEND\xaeB`\x82"
)


def load_script_module(module_name: str, path: Path) -> ModuleType:
    spec = importlib.util.spec_from_file_location(module_name, path)
    assert spec is not None
    assert spec.loader is not None
    module = importlib.util.module_from_spec(spec)
    sys.modules[spec.name] = module
    spec.loader.exec_module(module)
    return module


patcher = cast(Any, load_script_module("ooxml_cell_patcher_for_test", PATCHER_SCRIPT))
snapshot = cast(
    Any,
    load_script_module("drawing_media_snapshot_for_patcher_test", SNAPSHOT_SCRIPT),
)


def rewrite_xlsx(
    path: Path,
    updates: dict[str, bytes],
    removals: set[str] | None = None,
) -> None:
    removals = removals or set()
    temporary_path = path.with_suffix(".tmp.xlsx")
    with ZipFile(path) as source_archive:
        source_entries = {
            name: source_archive.read(name)
            for name in source_archive.namelist()
            if name not in updates and name not in removals
        }

    with ZipFile(temporary_path, "w", ZIP_DEFLATED) as target_archive:
        for name, content in source_entries.items():
            target_archive.writestr(name, content)
        for name, content in updates.items():
            target_archive.writestr(name, content)

    temporary_path.replace(path)


def content_types_with_drawing(content: str) -> str:
    if 'Extension="png"' not in content:
        content = content.replace(
            "</Types>",
            '<Default Extension="png" ContentType="image/png"/></Types>',
        )
    if 'PartName="/xl/drawings/drawing1.xml"' not in content:
        content = content.replace(
            "</Types>",
            (
                '<Override PartName="/xl/drawings/drawing1.xml" '
                'ContentType="application/vnd.openxmlformats-officedocument.'
                'drawing+xml"/></Types>'
            ),
        )
    return content


def add_valid_drawing_chain(path: Path) -> None:
    with ZipFile(path) as archive:
        sheet_xml = archive.read("xl/worksheets/sheet1.xml").decode("utf-8")
        content_types = archive.read("[Content_Types].xml").decode("utf-8")

    sheet_xml = sheet_xml.replace(
        "</worksheet>",
        (
            '<drawing xmlns:r="http://schemas.openxmlformats.org/'
            'officeDocument/2006/relationships" r:id="rId1"/>'
            "</worksheet>"
        ),
    )
    sheet_rels = (
        '<?xml version="1.0" encoding="UTF-8"?>'
        '<Relationships xmlns="http://schemas.openxmlformats.org/package/'
        '2006/relationships">'
        '<Relationship Id="rId1" Type="http://schemas.openxmlformats.org/'
        'officeDocument/2006/relationships/drawing" '
        'Target="../drawings/drawing1.xml"/>'
        "</Relationships>"
    )
    drawing_xml = (
        '<?xml version="1.0" encoding="UTF-8"?>'
        '<xdr:wsDr xmlns:xdr="http://schemas.openxmlformats.org/drawingml/'
        '2006/spreadsheetDrawing" '
        'xmlns:a="http://schemas.openxmlformats.org/drawingml/2006/main">'
        "<xdr:oneCellAnchor><xdr:from><xdr:col>0</xdr:col>"
        "<xdr:colOff>0</xdr:colOff><xdr:row>0</xdr:row>"
        "<xdr:rowOff>0</xdr:rowOff></xdr:from>"
        '<xdr:ext cx="1" cy="1"/>'
        '<xdr:pic><xdr:nvPicPr><xdr:cNvPr id="2" name="logo"/>'
        "<xdr:cNvPicPr/></xdr:nvPicPr><xdr:blipFill>"
        '<a:blip xmlns:r="http://schemas.openxmlformats.org/'
        'officeDocument/2006/relationships" r:embed="rId1"/>'
        "<a:stretch><a:fillRect/></a:stretch></xdr:blipFill>"
        "<xdr:spPr/></xdr:pic><xdr:clientData/></xdr:oneCellAnchor>"
        "</xdr:wsDr>"
    )
    drawing_rels = (
        '<?xml version="1.0" encoding="UTF-8"?>'
        '<Relationships xmlns="http://schemas.openxmlformats.org/package/'
        '2006/relationships">'
        '<Relationship Id="rId1" Type="http://schemas.openxmlformats.org/'
        'officeDocument/2006/relationships/image" '
        'Target="../media/image1.png"/>'
        "</Relationships>"
    )

    rewrite_xlsx(
        path,
        {
            "xl/worksheets/sheet1.xml": sheet_xml.encode("utf-8"),
            "xl/worksheets/_rels/sheet1.xml.rels": sheet_rels.encode("utf-8"),
            "xl/drawings/drawing1.xml": drawing_xml.encode("utf-8"),
            "xl/drawings/_rels/drawing1.xml.rels": drawing_rels.encode("utf-8"),
            "xl/media/image1.png": TINY_PNG,
            "[Content_Types].xml": content_types_with_drawing(content_types).encode(
                "utf-8"
            ),
        },
    )


def write_template(path: Path) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = SHEET_NAME
    worksheet.merge_cells("B20:I22")
    worksheet.freeze_panes = "C17"
    worksheet.print_title_rows = "1:16"
    worksheet.print_options.horizontalCentered = True
    worksheet.page_setup.orientation = "landscape"
    worksheet.page_setup.fitToWidth = 1
    worksheet.page_margins.left = 0.25
    worksheet.page_margins.right = 0.25
    worksheet.page_margins.top = 0.5
    worksheet.page_margins.bottom = 0.5
    worksheet.row_dimensions[17].height = 27
    worksheet.column_dimensions["C"].width = 32
    for row in range(17, 19):
        for column in range(3, 10):
            cell = worksheet.cell(row=row, column=column)
            cell.value = "template"
            cell.font = Font(bold=True)
    for row in range(17, 19):
        worksheet[f"I{row}"] = f"=E{row}*H{row}"
    workbook.save(path)
    add_valid_drawing_chain(path)


def patch_output_path(tmp_path: Path) -> Path:
    output_dir = tmp_path / "out"
    output_dir.mkdir()
    return output_dir / "patched.xlsx"


def updates() -> dict[str, object]:
    return {
        "C17": '  ВРУ-0012 "Тест"\nстрока & < >  ',
        "D17": "шт.",
        "E17": 2,
        "F17": "Автоматы",
        "G17": "Шкаф 600x400x250",
        "H17": "нужно уточнить",
    }


def zip_parts(path: Path) -> dict[str, bytes]:
    with ZipFile(path) as archive:
        return {name: archive.read(name) for name in archive.namelist()}


def worksheet_part(path: Path) -> str:
    with ZipFile(path) as archive:
        return cast(str, patcher.worksheet_part_for_sheet(archive, SHEET_NAME))


def worksheet_root(path: Path) -> ElementTree.Element:
    part = worksheet_part(path)
    with ZipFile(path) as archive:
        return ElementTree.fromstring(archive.read(part))


def worksheet_xml_bytes(path: Path) -> bytes:
    with ZipFile(path) as archive:
        return archive.read(worksheet_part(path))


def target_cell_ranges(
    worksheet_xml: bytes,
    coordinates: set[str],
) -> dict[str, Any]:
    ranges = patcher.cell_ranges(worksheet_xml)
    selected = {}
    for coordinate in coordinates:
        matches = ranges[coordinate]
        assert len(matches) == 1
        selected[coordinate] = matches[0]
    return selected


def cell_xml_bytes(worksheet_xml: bytes, coordinate: str) -> bytes:
    ranges = target_cell_ranges(worksheet_xml, {coordinate})
    cell_range = ranges[coordinate]
    return worksheet_xml[cell_range.start : cell_range.end]


def mask_target_cells(worksheet_xml: bytes, coordinates: set[str]) -> bytes:
    masked = bytearray(worksheet_xml)
    ranges = target_cell_ranges(worksheet_xml, coordinates)
    for coordinate, cell_range in sorted(
        ranges.items(),
        key=lambda item: item[1].start,
        reverse=True,
    ):
        placeholder = f"__PATCHED_CELL_{coordinate}__".encode("ascii")
        masked[cell_range.start : cell_range.end] = placeholder
    return bytes(masked)


def root_start_tag(worksheet_xml: bytes) -> bytes:
    start = worksheet_xml.find(b"<worksheet")
    assert start != -1
    end = worksheet_xml.find(b">", start)
    assert end != -1
    return worksheet_xml[start : end + 1]


def cell_elements(path: Path) -> dict[str, ElementTree.Element]:
    cells: dict[str, ElementTree.Element] = {}
    for cell in worksheet_root(path).findall(".//main:c", NS):
        coordinate = cell.get("r")
        if coordinate:
            cells[coordinate] = cell
    return cells


def cell_styles(path: Path, cells: set[str]) -> dict[str, str | None]:
    elements = cell_elements(path)
    return {cell: elements[cell].get("s") for cell in cells}


def cell_xml_map(path: Path) -> dict[str, bytes]:
    return {
        coordinate: ElementTree.tostring(cell)
        for coordinate, cell in cell_elements(path).items()
    }


def assert_non_target_parts_unchanged(template: Path, output: Path) -> None:
    template_parts = zip_parts(template)
    output_parts = zip_parts(output)
    changed_part = worksheet_part(template)
    assert set(output_parts) == set(template_parts)
    for name, content in template_parts.items():
        if name != changed_part:
            assert output_parts[name] == content, name


def assert_non_target_cells_unchanged(template: Path, output: Path) -> None:
    before_xml = worksheet_xml_bytes(template)
    after_xml = worksheet_xml_bytes(output)
    assert mask_target_cells(before_xml, TARGET_CELLS) == mask_target_cells(
        after_xml,
        TARGET_CELLS,
    )


def complex_worksheet_xml(sheet_xml: bytes) -> bytes:
    sheet_xml = sheet_xml.replace(
        b'<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/'
        b'2006/main">',
        (
            b'<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/'
            b'2006/main" xmlns:mc="http://schemas.openxmlformats.org/'
            b'markup-compatibility/2006" xmlns:x14ac="http://schemas.microsoft.'
            b'com/office/spreadsheetml/2009/9/ac" xmlns:xr="http://schemas.'
            b'microsoft.com/office/spreadsheetml/2014/revision" xmlns:xr2='
            b'"http://schemas.microsoft.com/office/spreadsheetml/2015/revision2" '
            b'mc:Ignorable="x14ac xr xr2" xr:uid="{11111111-1111-1111-1111-'
            b'111111111111}">'
        ),
        1,
    )
    sheet_xml = sheet_xml.replace(
        b'<row r="17"',
        b'<row r="17" hidden="1" x14ac:dyDescent="0.25"',
        1,
    )
    sheet_xml = sheet_xml.replace(
        b"</worksheet>",
        (
            b'<extLst><ext uri="{TEST-EXT}" xmlns:mx="http://example.com/'
            b'custom-main"><mx:payload xr2:uid="{22222222-2222-2222-2222-'
            b'222222222222}">keep-me</mx:payload></ext></extLst>'
            b"</worksheet>"
        ),
        1,
    )
    return sheet_xml


def add_complex_worksheet_markup(path: Path) -> None:
    sheet_part = worksheet_part(path)
    with ZipFile(path) as archive:
        sheet_xml = archive.read(sheet_part)
    rewrite_xlsx(path, {sheet_part: complex_worksheet_xml(sheet_xml)})


def make_cell_self_closing(path: Path, coordinate: str) -> None:
    sheet_part = worksheet_part(path)
    sheet_xml = worksheet_xml_bytes(path)
    cell_range = target_cell_ranges(sheet_xml, {coordinate})[coordinate]
    replacement = b'<c r="' + coordinate.encode("ascii") + b'" s="1" customWidth="1"/>'
    rewrite_xlsx(
        path,
        {
            sheet_part: (
                sheet_xml[: cell_range.start]
                + replacement
                + sheet_xml[cell_range.end :]
            )
        },
    )


def replace_cell_xml(path: Path, coordinate: str, replacement: bytes) -> None:
    sheet_part = worksheet_part(path)
    sheet_xml = worksheet_xml_bytes(path)
    cell_range = target_cell_ranges(sheet_xml, {coordinate})[coordinate]
    rewrite_xlsx(
        path,
        {
            sheet_part: (
                sheet_xml[: cell_range.start]
                + replacement
                + sheet_xml[cell_range.end :]
            )
        },
    )


def assert_patcher_error(expected: str, call: Any) -> None:
    try:
        call()
    except patcher.OoxmlCellPatcherError as error:
        assert expected in str(error)
    else:
        raise AssertionError("patcher call should fail")


def test_patch_existing_cells_preserves_drawing_media_and_other_parts(
    tmp_path: Path,
) -> None:
    template = tmp_path / "template.xlsx"
    output = patch_output_path(tmp_path)
    write_template(template)
    before_snapshot = snapshot.build_drawing_media_snapshot(template)
    before_styles = cell_styles(template, TARGET_CELLS)

    result = patcher.patch_existing_cells(
        template=template,
        output=output,
        sheet_name=SHEET_NAME,
        updates=updates(),
    )

    assert result == output
    assert output.is_file()
    workbook = load_workbook(output, data_only=False)
    worksheet = workbook[SHEET_NAME]
    assert worksheet["C17"].value == '  ВРУ-0012 "Тест"\nстрока & < >  '
    assert worksheet["D17"].value == "шт."
    assert worksheet["E17"].value == 2
    assert isinstance(worksheet["E17"].value, int)
    assert worksheet["I17"].value == "=E17*H17"
    assert worksheet["I18"].value == "=E18*H18"
    assert tuple(str(item) for item in worksheet.merged_cells.ranges) == ("B20:I22",)
    assert worksheet.page_setup.orientation == "landscape"
    assert worksheet.page_setup.fitToWidth == 1
    assert worksheet.freeze_panes == "C17"
    assert cell_styles(output, TARGET_CELLS) == before_styles
    after_snapshot = snapshot.build_drawing_media_snapshot(output)
    snapshot.compare_drawing_media_snapshots(before_snapshot, after_snapshot)
    assert_non_target_parts_unchanged(template, output)
    assert_non_target_cells_unchanged(template, output)
    output_xml = zip_parts(output)[worksheet_part(output)].decode("utf-8")
    assert "<drawing" in output_xml
    assert "pageSetup" in output_xml
    assert "mergeCell" in output_xml
    assert "0012" in output_xml
    assert "&amp;" in output_xml
    assert "&lt;" in output_xml
    assert "&gt;" in output_xml
    assert 'xml:space="preserve"' in output_xml


def test_patch_existing_cells_preserves_worksheet_bytes_outside_target_cells(
    tmp_path: Path,
) -> None:
    template = tmp_path / "template.xlsx"
    output = patch_output_path(tmp_path)
    write_template(template)
    add_complex_worksheet_markup(template)
    before_snapshot = snapshot.build_drawing_media_snapshot(template)
    before_xml = worksheet_xml_bytes(template)
    before_root = root_start_tag(before_xml)
    before_masked = mask_target_cells(before_xml, TARGET_CELLS)
    before_cells = {
        coordinate: cell_xml_bytes(before_xml, coordinate)
        for coordinate in TARGET_CELLS
    }

    patcher.patch_existing_cells(
        template=template,
        output=output,
        sheet_name=SHEET_NAME,
        updates=updates(),
    )

    after_xml = worksheet_xml_bytes(output)
    after_masked = mask_target_cells(after_xml, TARGET_CELLS)
    after_cells = {
        coordinate: cell_xml_bytes(after_xml, coordinate) for coordinate in TARGET_CELLS
    }
    assert after_masked == before_masked
    assert root_start_tag(after_xml) == before_root
    assert b'mc:Ignorable="x14ac xr xr2"' in after_xml
    assert b"keep-me" in after_xml
    assert b"<drawing " in after_xml
    assert b"<mergeCell " in after_xml
    assert b'hidden="1"' in after_xml
    assert b'x14ac:dyDescent="0.25"' in after_xml
    assert b"<pageMargins " in after_xml
    assert b"<pageSetup " in after_xml
    assert b"<printOptions " in after_xml
    for coordinate, before_cell_xml in before_cells.items():
        assert after_cells[coordinate] != before_cell_xml, coordinate
    assert_non_target_parts_unchanged(template, output)
    after_snapshot = snapshot.build_drawing_media_snapshot(output)
    snapshot.compare_drawing_media_snapshots(before_snapshot, after_snapshot)


def test_self_closing_target_cell_is_patched_without_touching_other_bytes(
    tmp_path: Path,
) -> None:
    template = tmp_path / "template.xlsx"
    output = patch_output_path(tmp_path)
    write_template(template)
    make_cell_self_closing(template, "C17")
    before_xml = worksheet_xml_bytes(template)

    patcher.patch_existing_cells(
        template=template,
        output=output,
        sheet_name=SHEET_NAME,
        updates={"C17": "  self closing  "},
    )

    after_xml = worksheet_xml_bytes(output)
    assert mask_target_cells(after_xml, {"C17"}) == mask_target_cells(
        before_xml,
        {"C17"},
    )
    patched_cell = cell_xml_bytes(after_xml, "C17")
    assert b't="inlineStr"' in patched_cell
    assert b'customWidth="1"' in patched_cell
    assert b"  self closing  " in patched_cell
    assert b'xml:space="preserve"' in patched_cell


def test_formula_cached_value_and_existing_cell_types_are_replaced_in_target_only(
    tmp_path: Path,
) -> None:
    template = tmp_path / "template.xlsx"
    output = patch_output_path(tmp_path)
    write_template(template)
    replace_cell_xml(
        template,
        "C17",
        b'<c r="C17" s="1"><f>SUM(A1:A2)</f><v>2</v></c>',
    )
    replace_cell_xml(
        template,
        "D17",
        b'<c r="D17" s="1" t="s"><v>0</v></c>',
    )
    replace_cell_xml(
        template,
        "F17",
        (b'<c r="F17" s="1" t="inlineStr"><is><t>old</t></is>' b"</c>"),
    )
    before_xml = worksheet_xml_bytes(template)

    patcher.patch_existing_cells(
        template=template,
        output=output,
        sheet_name=SHEET_NAME,
        updates={"C17": 9, "D17": "новый текст", "F17": "inline"},
    )

    after_xml = worksheet_xml_bytes(output)
    assert mask_target_cells(after_xml, {"C17", "D17", "F17"}) == mask_target_cells(
        before_xml,
        {"C17", "D17", "F17"},
    )
    c17 = cell_xml_bytes(after_xml, "C17")
    d17 = cell_xml_bytes(after_xml, "D17")
    f17 = cell_xml_bytes(after_xml, "F17")
    assert b"<f>" not in c17
    assert b"<v>9</v>" in c17
    assert b"t=" not in c17
    assert b't="inlineStr"' in d17
    assert "новый текст".encode() in d17
    assert b't="inlineStr"' in f17
    assert b">inline<" in f17
    assert b's="1"' in c17
    assert b's="1"' in d17
    assert b's="1"' in f17


def test_duplicate_normalized_input_coordinate_fails_closed(tmp_path: Path) -> None:
    template = tmp_path / "template.xlsx"
    output = patch_output_path(tmp_path)
    write_template(template)

    assert_patcher_error(
        "duplicate normalized cell coordinate: C17",
        lambda: patcher.patch_existing_cells(
            template=template,
            output=output,
            sheet_name=SHEET_NAME,
            updates={"c17": "первое", "C17": "второе"},
        ),
    )
    assert not output.exists()


def test_duplicate_coordinate_in_worksheet_xml_fails_closed(tmp_path: Path) -> None:
    template = tmp_path / "template.xlsx"
    output = patch_output_path(tmp_path)
    write_template(template)
    sheet_part = worksheet_part(template)
    sheet_xml = worksheet_xml_bytes(template)
    c18 = cell_xml_bytes(sheet_xml, "C18")
    sheet_xml = sheet_xml.replace(c18, c18.replace(b'r="C18"', b'r="C17"'), 1)
    rewrite_xlsx(template, {sheet_part: sheet_xml})

    assert_patcher_error(
        "duplicate cell coordinate in worksheet XML: C17",
        lambda: patcher.patch_existing_cells(
            template=template,
            output=output,
            sheet_name=SHEET_NAME,
            updates={"C17": "ambiguous"},
        ),
    )
    assert not output.exists()


def test_missing_template_fails_closed(tmp_path: Path) -> None:
    output = patch_output_path(tmp_path)

    assert_patcher_error(
        "template does not exist",
        lambda: patcher.patch_existing_cells(
            template=tmp_path / "missing.xlsx",
            output=output,
            sheet_name=SHEET_NAME,
            updates=updates(),
        ),
    )
    assert not output.exists()


def test_invalid_zip_template_fails_closed(tmp_path: Path) -> None:
    template = tmp_path / "template.xlsx"
    output = patch_output_path(tmp_path)
    template.write_bytes(b"not a zip")

    assert_patcher_error(
        "invalid xlsx ZIP package",
        lambda: patcher.patch_existing_cells(
            template=template,
            output=output,
            sheet_name=SHEET_NAME,
            updates=updates(),
        ),
    )
    assert not output.exists()


def test_missing_sheet_fails_closed(tmp_path: Path) -> None:
    template = tmp_path / "template.xlsx"
    output = patch_output_path(tmp_path)
    write_template(template)

    assert_patcher_error(
        "worksheet not found",
        lambda: patcher.patch_existing_cells(
            template=template,
            output=output,
            sheet_name="Missing",
            updates=updates(),
        ),
    )
    assert not output.exists()


def test_missing_cell_fails_closed_and_leaves_no_partial_output(tmp_path: Path) -> None:
    template = tmp_path / "template.xlsx"
    output = patch_output_path(tmp_path)
    write_template(template)

    assert_patcher_error(
        "cell does not exist: C999",
        lambda: patcher.patch_existing_cells(
            template=template,
            output=output,
            sheet_name=SHEET_NAME,
            updates={"C999": "missing"},
        ),
    )
    assert not output.exists()
    assert list(output.parent.iterdir()) == []


def test_invalid_cell_coordinate_fails_closed(tmp_path: Path) -> None:
    template = tmp_path / "template.xlsx"
    output = patch_output_path(tmp_path)
    write_template(template)

    assert_patcher_error(
        "invalid cell coordinate",
        lambda: patcher.patch_existing_cells(
            template=template,
            output=output,
            sheet_name=SHEET_NAME,
            updates={"17C": "bad"},
        ),
    )
    assert not output.exists()


def test_unsupported_value_type_fails_closed(tmp_path: Path) -> None:
    template = tmp_path / "template.xlsx"
    output = patch_output_path(tmp_path)
    write_template(template)

    assert_patcher_error(
        "unsupported value type for C17",
        lambda: patcher.patch_existing_cells(
            template=template,
            output=output,
            sheet_name=SHEET_NAME,
            updates={"C17": 1.5},
        ),
    )
    assert not output.exists()


def test_bool_value_type_fails_closed(tmp_path: Path) -> None:
    template = tmp_path / "template.xlsx"
    output = patch_output_path(tmp_path)
    write_template(template)

    assert_patcher_error(
        "unsupported value type for C17",
        lambda: patcher.patch_existing_cells(
            template=template,
            output=output,
            sheet_name=SHEET_NAME,
            updates={"C17": True},
        ),
    )
    assert not output.exists()


def test_existing_output_fails_closed_without_overwrite(tmp_path: Path) -> None:
    template = tmp_path / "template.xlsx"
    output = patch_output_path(tmp_path)
    write_template(template)
    output.write_bytes(b"existing")

    assert_patcher_error(
        "output already exists",
        lambda: patcher.patch_existing_cells(
            template=template,
            output=output,
            sheet_name=SHEET_NAME,
            updates=updates(),
        ),
    )
    assert output.read_bytes() == b"existing"


def test_output_inside_project_fails_closed(tmp_path: Path) -> None:
    template = tmp_path / "template.xlsx"
    output = PROJECT_ROOT / "blocked_ooxml_output.xlsx"
    write_template(template)

    assert_patcher_error(
        "output is inside the Git project",
        lambda: patcher.patch_existing_cells(
            template=template,
            output=output,
            sheet_name=SHEET_NAME,
            updates=updates(),
        ),
    )
    assert not output.exists()


def test_output_matches_template_fails_closed(tmp_path: Path) -> None:
    template = tmp_path / "template.xlsx"
    write_template(template)

    assert_patcher_error(
        "output matches template",
        lambda: patcher.patch_existing_cells(
            template=template,
            output=template,
            sheet_name=SHEET_NAME,
            updates=updates(),
        ),
    )


def test_missing_output_directory_fails_closed(tmp_path: Path) -> None:
    template = tmp_path / "template.xlsx"
    output = tmp_path / "missing-dir" / "patched.xlsx"
    write_template(template)

    assert_patcher_error(
        "output parent directory does not exist",
        lambda: patcher.patch_existing_cells(
            template=template,
            output=output,
            sheet_name=SHEET_NAME,
            updates=updates(),
        ),
    )
    assert not output.exists()


def test_temporary_write_error_cleans_partial_output(
    tmp_path: Path,
    monkeypatch: Any,
) -> None:
    template = tmp_path / "template.xlsx"
    output = patch_output_path(tmp_path)
    write_template(template)

    def fail_write(
        template_parts: Any,
        worksheet_part: str,
        worksheet_xml: bytes,
        temporary_output: Path,
    ) -> None:
        temporary_output.write_bytes(b"partial")
        raise patcher.OoxmlCellPatcherError("forced temporary write failure")

    monkeypatch.setattr(patcher, "write_patched_package", fail_write)

    assert_patcher_error(
        "forced temporary write failure",
        lambda: patcher.patch_existing_cells(
            template=template,
            output=output,
            sheet_name=SHEET_NAME,
            updates=updates(),
        ),
    )
    assert not output.exists()
    assert list(output.parent.iterdir()) == []


def test_empty_updates_fail_closed(tmp_path: Path) -> None:
    template = tmp_path / "template.xlsx"
    output = patch_output_path(tmp_path)
    write_template(template)

    assert_patcher_error(
        "updates must not be empty",
        lambda: patcher.patch_existing_cells(
            template=template,
            output=output,
            sheet_name=SHEET_NAME,
            updates={},
        ),
    )
    assert not output.exists()


def test_no_real_xlsx_or_manifest_files_are_added_to_git() -> None:
    status = subprocess.run(
        ["git", "status", "--short", "--untracked-files=all"],
        cwd=PROJECT_ROOT,
        capture_output=True,
        text=True,
        check=True,
    )
    changed_paths = [
        re.sub(r"^.. ", "", line).strip()
        for line in status.stdout.splitlines()
        if line.strip()
    ]
    assert not any(path.endswith(".xlsx") for path in changed_paths)
    assert not any(
        Path(path).name.casefold() == "manifest.json" for path in changed_paths
    )
