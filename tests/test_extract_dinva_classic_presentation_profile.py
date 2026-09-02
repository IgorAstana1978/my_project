from __future__ import annotations

import base64
import copy
import hashlib
import importlib.util
import json
import sys
from pathlib import Path
from types import ModuleType
from typing import Any

import pytest
from openpyxl import Workbook  # type: ignore[import-untyped]
from openpyxl.styles import (  # type: ignore[import-untyped]
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)

ROOT = Path(__file__).resolve().parents[1]
TINY_PNG = base64.b64decode(
    "iVBORw0KGgoAAAANSUhEUgAAAAEAAAABCAQAAAC1HAwCAAAAC0lEQVR42mNk+A8AAQUB"
    "AScY42YAAAAASUVORK5CYII="
)
DIFFERING_PNG = TINY_PNG + b"\x00"
TEST_CANONICAL_PIXEL_FINGERPRINT = "a" * 64
TEST_RUNTIME_PIXEL_FINGERPRINT = "b" * 64
PLACEMENT = {
    "anchor_type": "ONE_CELL",
    "from": {"column": 0, "column_offset": 76200, "row": 1, "row_offset": 66675},
    "extent": {"cx": 781050, "cy": 428625},
}
RUNTIME_MERGES = [
    "C2:F2",
    "G2:I2",
    "C3:F3",
    "G3:I3",
    "B4:F4",
    "G4:I4",
    "B5:F5",
    "G5:I5",
    "B6:F6",
    "G6:I6",
    "B9:F9",
    "G9:I9",
    "B10:F10",
    "G10:I10",
    "B11:F11",
    "G11:I11",
    "B12:F12",
    "G12:I12",
    "B13:F13",
    "G13:I13",
    "C22:I22",
    "C24:I24",
    "C25:I25",
    "C26:I26",
    "C27:I27",
    "C28:I28",
    "B30:E30",
    "F30:I30",
    "B31:E31",
    "F31:I31",
    "B32:I32",
]


def load_script(name: str) -> ModuleType:
    path = ROOT / "scripts" / f"{name}.py"
    spec = importlib.util.spec_from_file_location(f"test_{name}", path)
    assert spec is not None and spec.loader is not None
    module = importlib.util.module_from_spec(spec)
    sys.modules[spec.name] = module
    spec.loader.exec_module(module)
    return module


def sha(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()


def apply(cell: Any, *, bold: bool = False, red: bool = False) -> None:
    cell.font = Font(
        name="Times New Roman",
        size=11,
        bold=bold,
        color="FFFF0000" if red else "FF000000",
    )
    cell.fill = PatternFill(fill_type=None)
    cell.border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    cell.number_format = "General"


def base_page(sheet: Any) -> None:
    sheet.page_setup.paperSize = "9"
    sheet.page_setup.orientation = "portrait"
    sheet.page_setup.scale = 54
    sheet.page_setup.fitToHeight = 0
    sheet.sheet_properties.pageSetUpPr.fitToPage = True


def write_reference(
    path: Path,
    *,
    case_marker: str,
    first_item_row: int = 16,
    classic: bool = True,
    renderer: ModuleType | None = None,
) -> str:
    renderer = renderer or load_script("render_dinva_classic_quote_invoice")
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Лист1"
    sheet["C2"] = "ТОО «ДиН ВА-КЭС»" if classic else "Другая компания"
    sheet["G9"] = "ВНИМАНИЕ!"
    apply(sheet["C2"], bold=True)
    apply(sheet["G9"], bold=True, red=True)
    headers = {
        "B": "№ п/п",
        "C": "Наименование",
        "D": "Ед.",
        "E": "Кол-во",
        "F": f"Приборы {case_marker}",
        "G": f"Шкаф {case_marker}",
        "H": "Цена",
        "I": "Сумма",
    }
    for column, value in headers.items():
        sheet[f"{column}15"] = value
        apply(sheet[f"{column}15"], bold=True)
    row = first_item_row
    values: dict[str, Any] = {
        f"B{row}": 1,
        f"C{row}": f"Case-only {case_marker}",
        f"D{row}": "шт.",
        f"E{row}": 1,
        f"F{row}": f"Composition {case_marker}",
        f"G{row}": f"Enclosure {case_marker}",
        f"H{row}": 100,
        f"I{row}": f"=E{row}*H{row}",
    }
    for coordinate, value in values.items():
        sheet[coordinate] = value
        apply(sheet[coordinate])
    base_page(sheet)
    workbook.save(path)
    workbook.close()
    renderer.inject_governed_parts(path, TINY_PNG, PLACEMENT, {"fixture": case_marker})
    return sha(path)


def write_runtime_template(
    path: Path,
    *,
    renderer: ModuleType | None = None,
    width_b: float = 10.0,
    logo: bytes = TINY_PNG,
) -> str:
    renderer = renderer or load_script("render_dinva_classic_quote_invoice")
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Счёт-КП шаблон"
    company = {
        "C2": "ТОО «ДиН ВА-КЭС»",
        "C3": "Республика Казахстан",
        "B4": "г. Астана",
        "B5": "Тел: +7",
        "B6": "info@dinva.kz",
        "G2": "БИН 000",
        "G3": "Банк",
        "G4": "IBAN KZ00",
        "G5": "BIK TEST",
        "G6": "КБЕ: 17",
    }
    for coordinate, value in company.items():
        sheet[coordinate] = value
        apply(sheet[coordinate], bold=coordinate == "C2")
    metadata = {
        "B9": "Черновик счёта-КП",
        "G9": "ВНИМАНИЕ!",
        "B10": "Плательщик",
        "G10": "Не отправлять",
        "B11": "Объект",
        "G11": "Проверить",
        "B12": "Основание",
        "G12": "Не запускать",
        "B13": "Статус",
        "G13": "Внутренний черновик",
    }
    for coordinate, value in metadata.items():
        sheet[coordinate] = value
        apply(sheet[coordinate], bold=coordinate == "G9", red=coordinate == "G9")
    headers = {
        "B": "№\nп/п",
        "C": "Наименование",
        "D": "Ед.",
        "E": "Кол-\nво",
        "F": "Применяемые приборы и аппараты согласно схемы",
        "G": "Тип шкафа, габариты ВхШхГ, материал",
        "H": "Цена",
        "I": "Сумма",
    }
    for column, value in headers.items():
        sheet[f"{column}15"] = value
        apply(sheet[f"{column}15"], bold=True)
    sheet["C16"] = "Раздел / объект / позиция проекта: нужно уточнить"
    apply(sheet["C16"])
    for row in range(17, 20):
        for column in "BCDEFGHI":
            apply(sheet[f"{column}{row}"])
        sheet.row_dimensions[row].height = 24
    sheet["B17"] = 1
    sheet["D17"] = "шт."
    sheet["I17"] = '=IF(OR(E17="",H17=""),"",IFERROR(E17*H17,"нужно уточнить"))'
    sheet["H20"] = "ИТОГО"
    sheet["I20"] = '=IF(COUNT(I17:I19)=0,"нужно уточнить",SUM(I17:I19))'
    sheet["H21"] = "НДС"
    sheet["I21"] = 0
    sheet["C22"] = "Всего прописью: нужно уточнить"
    sheet["C24"] = "Счёт действителен: нужно уточнить"
    sheet["C25"] = "Условия оплаты и поставки: нужно уточнить"
    sheet["C26"] = "Предположительный срок изготовления: нужно уточнить"
    sheet["C27"] = "Спецификация и условия подлежат проверке"
    sheet["C28"] = "Документ внутренний. Клиенту не отправлять"
    sheet["B30"] = "Директор"
    sheet["F30"] = "Директор Тест"
    sheet["B31"] = "Исполнитель:"
    sheet["F31"] = "Исполнитель Тест"
    sheet["B32"] = "Дата проверки: ____ / ____ / 2026"
    for coordinate in ("H20", "I20"):
        apply(sheet[coordinate], bold=True)
    for coordinate in ("H21", "I21"):
        apply(sheet[coordinate])
    for row in (22, 24, 25, 26, 27, 28):
        apply(sheet[f"C{row}"])
    for coordinate in ("B30", "F30", "B31", "F31", "B32"):
        apply(sheet[coordinate], bold=True)
    for column, width in zip(
        "BCDEFGHI", (width_b, 28, 7, 13, 35, 24, 14, 15), strict=True
    ):
        sheet.column_dimensions[column].width = width
    for merged_range in RUNTIME_MERGES:
        sheet.merge_cells(merged_range)
    base_page(sheet)
    workbook.save(path)
    workbook.close()
    renderer.inject_governed_parts(path, logo, PLACEMENT, {"fixture": "runtime"})
    return sha(path)


@pytest.fixture
def modules() -> tuple[ModuleType, ModuleType]:
    return (
        load_script("extract_dinva_classic_presentation_profile"),
        load_script("render_dinva_classic_quote_invoice"),
    )


def synthetic_inputs(
    tmp_path: Path,
    extractor: ModuleType,
    renderer: ModuleType,
    *,
    runtime_logo: bytes = TINY_PNG,
) -> tuple[list[Any], list[Any]]:
    first = tmp_path / "463.xlsx"
    second = tmp_path / "519.xlsx"
    third = tmp_path / "551.xlsx"
    runtime = tmp_path / "tuned-v4.xlsx"
    first_sha = write_reference(first, case_marker="463", renderer=renderer)
    second_sha = write_reference(
        second, case_marker="519", first_item_row=17, renderer=renderer
    )
    third_sha = write_reference(third, case_marker="551", renderer=renderer)
    runtime_sha = write_runtime_template(runtime, renderer=renderer, logo=runtime_logo)
    return (
        [
            extractor.ReferenceInput(first, first_sha),
            extractor.ReferenceInput(second, second_sha),
            extractor.ReferenceInput(third, third_sha),
        ],
        [extractor.ReferenceInput(runtime, runtime_sha)],
    )


def configure_synthetic_decision_constants(extractor: ModuleType) -> None:
    assert (
        extractor.APPROVED_CANONICAL_LOGO_DECISION_SHA256
        == "e7c043f19b7eb8606f59dd8e7de06b29ca4305cc1fe2362ecb93767dd589f63b"
    )
    assert (
        extractor.CANONICAL_LOGO_RAW_SHA256
        == "28a6a59ae0a5ca274c206c70545f70b333cac0276a7c4dcbebbf9156f88e0fa8"
    )
    assert (
        extractor.CANONICAL_LOGO_PIXEL_FINGERPRINT
        == "81d979c4c158452cca8e3b40d23a4fd321538dfcef238b6f8133beb33a122846"
    )
    assert (
        extractor.RUNTIME_DIFFERING_LOGO_RAW_SHA256
        == "18e0f9446c72f8aa80ea833df07c2e42eb830770a0186decc476c5f948987301"
    )
    extractor.__dict__["CANONICAL_LOGO_RAW_SHA256"] = hashlib.sha256(
        TINY_PNG
    ).hexdigest()
    extractor.__dict__["CANONICAL_LOGO_PIXEL_FINGERPRINT"] = (
        TEST_CANONICAL_PIXEL_FINGERPRINT
    )
    extractor.__dict__["RUNTIME_DIFFERING_LOGO_RAW_SHA256"] = hashlib.sha256(
        DIFFERING_PNG
    ).hexdigest()


def decision_binding(
    reference: Any,
    *,
    label: str,
    role: str,
    logo_sha256: str,
    pixel_fingerprint: str,
    dimensions: list[int],
    mode: str,
) -> dict[str, Any]:
    return {
        "label": label,
        "role": role,
        "path": str(reference.path.resolve()),
        "expected_workbook_sha256": reference.expected_sha256,
        "actual_workbook_sha256": reference.expected_sha256,
        "media_part_path": "xl/media/image1.png",
        "expected_logo_raw_sha256": logo_sha256,
        "actual_logo_raw_sha256": logo_sha256,
        "normalized_pixel_fingerprint": pixel_fingerprint,
        "native_dimensions": dimensions,
        "decoded_mode": mode,
    }


def decision_payload(
    extractor: ModuleType, family: list[Any], runtime: list[Any]
) -> dict[str, Any]:
    family_bindings = [
        decision_binding(
            reference,
            label=label,
            role="CLASSIC_FAMILY_EVIDENCE",
            logo_sha256=extractor.CANONICAL_LOGO_RAW_SHA256,
            pixel_fingerprint=extractor.CANONICAL_LOGO_PIXEL_FINGERPRINT,
            dimensions=[200, 68],
            mode="RGB",
        )
        for reference, label in zip(
            family, ("Invoice463", "Invoice519", "Invoice551"), strict=True
        )
    ]
    runtime_bindings = [
        decision_binding(
            reference,
            label="capacity100_tuned_v4",
            role="CERTIFIED_RUNTIME_TEMPLATE_EVIDENCE",
            logo_sha256=extractor.RUNTIME_DIFFERING_LOGO_RAW_SHA256,
            pixel_fingerprint=TEST_RUNTIME_PIXEL_FINGERPRINT,
            dimensions=[115, 43],
            mode="RGBA",
        )
        for reference in runtime
    ]
    return {
        "schema_version": extractor.DECISION_SCHEMA_VERSION,
        "artifact_type": extractor.DECISION_ARTIFACT_TYPE,
        "decision_id": extractor.DECISION_ID,
        "status": extractor.DECISION_STATUS,
        "authority": extractor.DECISION_AUTHORITY,
        "approval_scope": extractor.DECISION_SCOPE,
        "canonical_logo_application_status": extractor.DECISION_APPLICATION_STATUS,
        "created_at_utc": "2026-09-02T06:45:08Z",
        "source_bindings": family_bindings + runtime_bindings,
        "canonical_logo_decision": {
            "approved_variant": "A_FAMILY_INVOICE519",
            "authoritative_brand_source": "CLASSIC_FAMILY_EMBEDDED_LOGO",
            "media_part_path": "xl/media/image1.png",
            "raw_sha256": extractor.CANONICAL_LOGO_RAW_SHA256,
            "normalized_pixel_fingerprint": (
                extractor.CANONICAL_LOGO_PIXEL_FINGERPRINT
            ),
            "native_dimensions": [200, 68],
            "decoded_mode": "RGB",
        },
        "runtime_template_policy": {
            "template_id": "capacity100_tuned_v4",
            "permitted_role": "RUNTIME_GEOMETRY_STYLE_LAYOUT_SOURCE_ONLY",
            "authoritative_brand_logo_source": False,
            "differing_logo_raw_sha256": (extractor.RUNTIME_DIFFERING_LOGO_RAW_SHA256),
        },
        "safety": copy.deepcopy(extractor.DECISION_SAFETY),
        "publication_control": copy.deepcopy(extractor.DECISION_PUBLICATION_CONTROL),
    }


def write_decision(path: Path, extractor: ModuleType, payload: dict[str, Any]) -> Any:
    raw = (json.dumps(payload, ensure_ascii=False, indent=2) + "\n").encode("utf-8")
    path.write_bytes(raw)
    digest = hashlib.sha256(raw).hexdigest()
    extractor.__dict__["APPROVED_CANONICAL_LOGO_DECISION_SHA256"] = digest
    return extractor.ReferenceInput(path, digest)


def test_extraction_uses_certified_runtime_geometry_and_is_draft(
    tmp_path: Path, modules: tuple[ModuleType, ModuleType]
) -> None:
    extractor, renderer = modules
    project_root = tmp_path / "synthetic-repo"
    project_root.mkdir()
    extractor.__dict__["PROJECT_ROOT"] = project_root
    family, runtime = synthetic_inputs(tmp_path, extractor, renderer)
    left = extractor.extract_profile(family, runtime)
    right = extractor.extract_profile(list(reversed(family)), runtime)
    assert left == right
    layout = left["presentation_contract"]["layout"]
    assert layout["table_header_row"] == 15
    assert layout["first_item_row"] == 17
    assert layout["family_evidence_first_item_rows"] == [16, 16, 17]
    assert layout["merged_cells"]["ranges"] == sorted(RUNTIME_MERGES)
    assert left["artifact_status"] == "DRAFT_PROFILE_CANDIDATE"
    assert left["approval_provenance"]["status"] == "DRAFT_UNAPPROVED"
    encoded = json.dumps(left, ensure_ascii=False)
    assert "Case-only 463" not in encoded
    assert "Case-only 519" not in encoded
    assert "Case-only 551" not in encoded


def test_extractor_rejects_sha_nonclassic_and_runtime_geometry_conflict(
    tmp_path: Path, modules: tuple[ModuleType, ModuleType]
) -> None:
    extractor, renderer = modules
    project_root = tmp_path / "synthetic-repo"
    project_root.mkdir()
    extractor.__dict__["PROJECT_ROOT"] = project_root
    family, runtime = synthetic_inputs(tmp_path, extractor, renderer)
    with pytest.raises(extractor.ProfileExtractionError, match="SHA-256 mismatch"):
        extractor.extract_profile(
            [extractor.ReferenceInput(family[0].path, "0" * 64), family[1]], runtime
        )
    nonclassic = tmp_path / "637.xlsx"
    nonclassic_sha = write_reference(
        nonclassic, case_marker="637", classic=False, renderer=renderer
    )
    with pytest.raises(
        extractor.ProfileExtractionError, match="unsupported/non-classic"
    ):
        extractor.extract_profile(
            [family[0], extractor.ReferenceInput(nonclassic, nonclassic_sha)], runtime
        )
    conflicting = tmp_path / "conflicting-runtime.xlsx"
    conflicting_sha = write_runtime_template(
        conflicting, renderer=renderer, width_b=11.0
    )
    runtime_copy = tmp_path / "matching-runtime-copy.xlsx"
    runtime_copy.write_bytes(runtime[0].path.read_bytes())
    matching_runtime = runtime + [
        extractor.ReferenceInput(runtime_copy, sha(runtime_copy))
    ]
    assert extractor.extract_profile(family, matching_runtime)
    with pytest.raises(extractor.ProfileExtractionError, match="lacks consensus"):
        extractor.extract_profile(
            family,
            runtime + [extractor.ReferenceInput(conflicting, conflicting_sha)],
        )


def test_extractor_rejects_runtime_logo_not_bound_to_family_consensus(
    tmp_path: Path, modules: tuple[ModuleType, ModuleType]
) -> None:
    extractor, renderer = modules
    project_root = tmp_path / "synthetic-repo"
    project_root.mkdir()
    extractor.__dict__["PROJECT_ROOT"] = project_root
    family, runtime = synthetic_inputs(
        tmp_path, extractor, renderer, runtime_logo=DIFFERING_PNG
    )
    with pytest.raises(
        extractor.ProfileExtractionError,
        match="canonical-logo Human Decision is required",
    ):
        extractor.extract_profile(family, runtime)


def test_exact_decision_applies_family_logo_only_and_remains_draft(
    tmp_path: Path, modules: tuple[ModuleType, ModuleType]
) -> None:
    extractor, renderer = modules
    project_root = tmp_path / "synthetic-repo"
    project_root.mkdir()
    extractor.__dict__["PROJECT_ROOT"] = project_root
    configure_synthetic_decision_constants(extractor)
    family, runtime = synthetic_inputs(
        tmp_path, extractor, renderer, runtime_logo=DIFFERING_PNG
    )
    decision_path = tmp_path / "canonical-logo-decision.json"
    decision = write_decision(
        decision_path, extractor, decision_payload(extractor, family, runtime)
    )

    left = extractor.extract_profile(family, runtime, decision)
    right = extractor.extract_profile(list(reversed(family)), runtime, decision)

    assert left == right
    assert set(left) == {
        "schema_version",
        "profile_id",
        "document_family",
        "artifact_status",
        "reference_provenance",
        "presentation_contract",
        "presentation_contract_fingerprint",
        "approval_provenance",
    }
    asset = left["presentation_contract"]["assets"][0]
    assert base64.b64decode(asset["data_base64"]) == TINY_PNG
    assert base64.b64decode(asset["data_base64"]) != DIFFERING_PNG
    assert asset["sha256"] == hashlib.sha256(TINY_PNG).hexdigest()
    assert asset["placement"] == PLACEMENT
    assert asset["source_reference_sha256s"] == sorted(
        reference.expected_sha256 for reference in family
    )
    decision_provenance = [
        item
        for item in left["reference_provenance"]
        if item["role"] == "CANONICAL_LOGO_HUMAN_DECISION"
    ]
    assert decision_provenance == [
        {
            "path": str(decision_path.resolve()),
            "expected_sha256": decision.expected_sha256,
            "actual_sha256": decision.expected_sha256,
            "role": "CANONICAL_LOGO_HUMAN_DECISION",
        }
    ]
    assert left["artifact_status"] == "DRAFT_PROFILE_CANDIDATE"
    assert left["approval_provenance"] == {
        "status": "DRAFT_UNAPPROVED",
        "authority": None,
        "approval_id": None,
        "approved_at": None,
        "approved_contract_fingerprint": None,
    }
    assert left["presentation_contract_fingerprint"] == (
        extractor.contract_fingerprint(left["presentation_contract"])
    )
    assert list(tmp_path.glob("*.json")) == [decision_path]


def test_decision_sha_duplicate_keys_and_cli_pair_are_fail_closed(
    tmp_path: Path, modules: tuple[ModuleType, ModuleType]
) -> None:
    extractor, _ = modules
    project_root = tmp_path / "synthetic-repo"
    project_root.mkdir()
    extractor.__dict__["PROJECT_ROOT"] = project_root
    configure_synthetic_decision_constants(extractor)
    placeholder = extractor.ReferenceInput(tmp_path / "source.xlsx", "1" * 64)
    payload = decision_payload(
        extractor, [placeholder, placeholder, placeholder], [placeholder]
    )
    decision_path = tmp_path / "decision.json"
    decision = write_decision(decision_path, extractor, payload)
    assert extractor.load_canonical_logo_decision(decision).sha256 == (
        decision.expected_sha256
    )
    copied_path = tmp_path / "decision-copy.json"
    copied_path.write_bytes(decision_path.read_bytes())
    assert (
        extractor.load_canonical_logo_decision(
            extractor.ReferenceInput(copied_path, decision.expected_sha256)
        ).path
        == copied_path.resolve()
    )
    with pytest.raises(extractor.ProfileExtractionError, match="invalid.*SHA-256"):
        extractor.load_canonical_logo_decision(
            extractor.ReferenceInput(decision_path, "invalid")
        )
    with pytest.raises(extractor.ProfileExtractionError, match="approved artifact"):
        extractor.load_canonical_logo_decision(
            extractor.ReferenceInput(decision_path, "0" * 64)
        )

    duplicate_text = decision_path.read_text(encoding="utf-8").replace(
        '  "status":', '  "status": "DUPLICATE",\n  "status":', 1
    )
    duplicate_path = tmp_path / "duplicate.json"
    duplicate_path.write_text(duplicate_text, encoding="utf-8", newline="\n")
    extractor.__dict__["APPROVED_CANONICAL_LOGO_DECISION_SHA256"] = sha(duplicate_path)
    with pytest.raises(extractor.ProfileExtractionError, match="duplicate JSON key"):
        extractor.load_canonical_logo_decision(
            extractor.ReferenceInput(duplicate_path, sha(duplicate_path))
        )

    assert extractor.optional_reference(None, None, "decision") is None
    assert extractor.optional_reference(
        decision_path, decision.expected_sha256, "decision"
    ) == extractor.ReferenceInput(decision_path, decision.expected_sha256)
    with pytest.raises(extractor.ProfileExtractionError, match="supplied together"):
        extractor.optional_reference(decision_path, None, "decision")
    with pytest.raises(extractor.ProfileExtractionError, match="supplied together"):
        extractor.optional_reference(None, decision.expected_sha256, "decision")


def test_only_exact_approved_decision_bytes_are_accepted(
    tmp_path: Path, modules: tuple[ModuleType, ModuleType]
) -> None:
    extractor, _ = modules
    project_root = tmp_path / "synthetic-repo"
    project_root.mkdir()
    extractor.__dict__["PROJECT_ROOT"] = project_root
    configure_synthetic_decision_constants(extractor)
    placeholder = extractor.ReferenceInput(tmp_path / "source.xlsx", "1" * 64)
    payload = decision_payload(
        extractor, [placeholder, placeholder, placeholder], [placeholder]
    )
    approved_path = tmp_path / "approved.json"
    approved = write_decision(approved_path, extractor, payload)
    approved_raw = approved_path.read_bytes()

    label_mutation = copy.deepcopy(payload)
    label_mutation["source_bindings"][0]["label"] = "OtherwiseValidLabelChange"
    created_at_mutation = copy.deepcopy(payload)
    created_at_mutation["created_at_utc"] = "2026-09-02T06:45:09Z"
    mutations = {
        "otherwise-valid.json": (
            json.dumps(label_mutation, ensure_ascii=False, indent=2) + "\n"
        ).encode("utf-8"),
        "whitespace-only.json": approved_raw + b"\n",
        "created-at-only.json": (
            json.dumps(created_at_mutation, ensure_ascii=False, indent=2) + "\n"
        ).encode("utf-8"),
    }
    for name, raw in mutations.items():
        path = tmp_path / name
        path.write_bytes(raw)
        recomputed = hashlib.sha256(raw).hexdigest()
        with pytest.raises(extractor.ProfileExtractionError, match="approved artifact"):
            extractor.load_canonical_logo_decision(
                extractor.ReferenceInput(path, recomputed)
            )

    wrong_actual_path = tmp_path / "wrong-actual-bytes.json"
    wrong_actual_path.write_bytes(approved_raw + b" ")
    with pytest.raises(extractor.ProfileExtractionError, match="SHA-256 mismatch"):
        extractor.load_canonical_logo_decision(
            extractor.ReferenceInput(wrong_actual_path, approved.expected_sha256)
        )


@pytest.mark.parametrize(
    ("field_path", "bad_value", "message"),
    [
        (("schema_version",), "wrong", "schema_version"),
        (("status",), "wrong", "status"),
        (("authority",), "wrong", "authority"),
        (("approval_scope",), "wrong", "approval_scope"),
        (
            ("canonical_logo_application_status",),
            "wrong",
            "canonical_logo_application_status",
        ),
        (
            ("canonical_logo_decision", "raw_sha256"),
            "0" * 64,
            "canonical logo decision",
        ),
        (
            ("canonical_logo_decision", "normalized_pixel_fingerprint"),
            "0" * 64,
            "canonical logo decision",
        ),
        (
            ("runtime_template_policy", "differing_logo_raw_sha256"),
            "0" * 64,
            "runtime policy",
        ),
        (
            ("safety", "profile_generation_authorized"),
            True,
            "safety",
        ),
    ],
)
def test_decision_exact_contract_rejects_mutation(
    tmp_path: Path,
    modules: tuple[ModuleType, ModuleType],
    field_path: tuple[str, ...],
    bad_value: object,
    message: str,
) -> None:
    extractor, _ = modules
    project_root = tmp_path / "synthetic-repo"
    project_root.mkdir()
    extractor.__dict__["PROJECT_ROOT"] = project_root
    configure_synthetic_decision_constants(extractor)
    placeholder = extractor.ReferenceInput(tmp_path / "source.xlsx", "1" * 64)
    payload = decision_payload(
        extractor, [placeholder, placeholder, placeholder], [placeholder]
    )
    target: dict[str, Any] = payload
    for key in field_path[:-1]:
        target = target[key]
    target[field_path[-1]] = bad_value
    decision = write_decision(tmp_path / "decision.json", extractor, payload)
    with pytest.raises(extractor.ProfileExtractionError, match=message):
        extractor.load_canonical_logo_decision(decision)


@pytest.mark.parametrize("mutation", ["path", "sha", "role"])
def test_decision_source_binding_is_exact(
    tmp_path: Path,
    modules: tuple[ModuleType, ModuleType],
    mutation: str,
) -> None:
    extractor, renderer = modules
    project_root = tmp_path / "synthetic-repo"
    project_root.mkdir()
    extractor.__dict__["PROJECT_ROOT"] = project_root
    configure_synthetic_decision_constants(extractor)
    family, runtime = synthetic_inputs(
        tmp_path, extractor, renderer, runtime_logo=DIFFERING_PNG
    )
    payload = decision_payload(extractor, family, runtime)
    binding = payload["source_bindings"][0]
    if mutation == "path":
        binding["path"] = str(tmp_path / "substituted.xlsx")
    elif mutation == "sha":
        binding["expected_workbook_sha256"] = "0" * 64
        binding["actual_workbook_sha256"] = "0" * 64
    else:
        binding["role"] = "CERTIFIED_RUNTIME_TEMPLATE_EVIDENCE"
    decision = write_decision(tmp_path / "decision.json", extractor, payload)
    with pytest.raises(extractor.ProfileExtractionError, match="source|role"):
        extractor.extract_profile(family, runtime, decision)


def test_decision_rejects_unbound_missing_or_extra_evidence(
    tmp_path: Path, modules: tuple[ModuleType, ModuleType]
) -> None:
    extractor, renderer = modules
    project_root = tmp_path / "synthetic-repo"
    project_root.mkdir()
    extractor.__dict__["PROJECT_ROOT"] = project_root
    configure_synthetic_decision_constants(extractor)
    family, runtime = synthetic_inputs(
        tmp_path, extractor, renderer, runtime_logo=DIFFERING_PNG
    )
    decision = write_decision(
        tmp_path / "decision.json",
        extractor,
        decision_payload(extractor, family, runtime),
    )
    with pytest.raises(extractor.ProfileExtractionError, match="source binding"):
        extractor.extract_profile(family[:-1], runtime, decision)

    runtime_copy_path = tmp_path / "unbound-runtime-copy.xlsx"
    runtime_copy_path.write_bytes(runtime[0].path.read_bytes())
    runtime_with_extra = runtime + [
        extractor.ReferenceInput(runtime_copy_path, sha(runtime_copy_path))
    ]
    with pytest.raises(extractor.ProfileExtractionError, match="source binding"):
        extractor.extract_profile(family, runtime_with_extra, decision)


def test_profile_publication_is_no_overwrite_and_outside_git(
    tmp_path: Path, modules: tuple[ModuleType, ModuleType]
) -> None:
    extractor, _ = modules
    project_root = tmp_path / "synthetic-repo"
    project_root.mkdir()
    extractor.__dict__["PROJECT_ROOT"] = project_root
    output = tmp_path / "profile.json"
    assert extractor.publish_profile({"draft": True}, output) == output
    with pytest.raises(extractor.ProfileExtractionError, match="already exists"):
        extractor.publish_profile({"draft": True}, output)
    with pytest.raises(extractor.ProfileExtractionError, match="outside Git"):
        extractor.publish_profile({"draft": True}, project_root / "profile.json")
