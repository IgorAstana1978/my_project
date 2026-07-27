import importlib.util
import json
import sys
from pathlib import Path
from types import ModuleType
from typing import Any, cast

import pytest
from openpyxl import Workbook, load_workbook  # type: ignore[import-untyped]
from pypdf import PdfReader, PdfWriter

PROJECT_ROOT = Path(__file__).resolve().parents[1]
SCRIPTS_DIR = PROJECT_ROOT / "scripts"
CORE_SCRIPT = SCRIPTS_DIR / "project_spec_extraction.py"
OPERATOR_SCRIPT = SCRIPTS_DIR / "extract_mixed_source_composition.py"
sys.path.insert(0, str(SCRIPTS_DIR))


def load_module(name: str, path: Path) -> ModuleType:
    spec = importlib.util.spec_from_file_location(name, path)
    assert spec is not None
    assert spec.loader is not None
    module = importlib.util.module_from_spec(spec)
    sys.modules[spec.name] = module
    spec.loader.exec_module(module)
    return module


extraction = cast(Any, load_module("project_spec_extraction_for_test", CORE_SCRIPT))
operator = cast(Any, load_module("mixed_source_operator_for_test", OPERATOR_SCRIPT))


def pdf_escape(value: str) -> str:
    return value.replace("\\", "\\\\").replace("(", "\\(").replace(")", "\\)")


def write_text_pdf(
    path: Path,
    pages: list[list[tuple[str, int, int]]],
    *,
    image_for_empty: bool = True,
) -> Path:
    objects: dict[int, bytes] = {}
    objects[1] = b"<< /Type /Catalog /Pages 2 0 R >>"
    page_ids = [4 + index * 2 for index in range(len(pages))]
    kids = " ".join(f"{page_id} 0 R" for page_id in page_ids)
    objects[2] = f"<< /Type /Pages /Kids [{kids}] /Count {len(pages)} >>".encode()
    objects[3] = b"<< /Type /Font /Subtype /Type1 /BaseFont /Helvetica >>"
    image_id = 4 + len(pages) * 2
    if image_for_empty and any(not lines for lines in pages):
        image_data = b"\x80"
        objects[image_id] = (
            b"<< /Type /XObject /Subtype /Image /Width 1 /Height 1 "
            b"/ColorSpace /DeviceGray /BitsPerComponent 8 /Length 1 >>\n"
            b"stream\n" + image_data + b"\nendstream"
        )
    for index, lines in enumerate(pages):
        page_id = page_ids[index]
        content_id = page_id + 1
        commands = [
            f"BT /F1 12 Tf {x} {y} Td ({pdf_escape(text)}) Tj ET"
            for text, x, y in lines
        ]
        content = "\n".join(commands).encode("ascii")
        if not lines and image_for_empty:
            content = b"q 100 0 0 100 72 600 cm /Im1 Do Q"
        xobject = (
            f" /XObject << /Im1 {image_id} 0 R >>"
            if not lines and image_for_empty
            else ""
        )
        objects[page_id] = (
            "<< /Type /Page /Parent 2 0 R /MediaBox [0 0 612 792] "
            f"/Resources << /Font << /F1 3 0 R >>{xobject} >> "
            f"/Contents {content_id} 0 R >>"
        ).encode()
        objects[content_id] = (
            f"<< /Length {len(content)} >>\nstream\n".encode()
            + content
            + b"\nendstream"
        )

    output = bytearray(b"%PDF-1.4\n%synthetic\n")
    offsets = [0]
    for object_number in range(1, max(objects) + 1):
        offsets.append(len(output))
        output.extend(f"{object_number} 0 obj\n".encode())
        output.extend(objects[object_number])
        output.extend(b"\nendobj\n")
    xref_offset = len(output)
    output.extend(f"xref\n0 {len(offsets)}\n".encode())
    output.extend(b"0000000000 65535 f \n")
    for offset in offsets[1:]:
        output.extend(f"{offset:010d} 00000 n \n".encode())
    output.extend(
        (
            f"trailer\n<< /Size {len(offsets)} /Root 1 0 R >>\n"
            f"startxref\n{xref_offset}\n%%EOF\n"
        ).encode()
    )
    path.write_bytes(output)
    return path


def write_spec_workbook(
    path: Path,
    rows_by_sheet: dict[str, list[list[object]]],
    *,
    repeat_header: bool = False,
) -> Path:
    workbook = Workbook()
    default = workbook.active
    workbook.remove(default)
    header = [
        "Board",
        "Board name",
        "Board quantity",
        "Item",
        "Model",
        "Brand",
        "Rating",
        "Unit",
        "Quantity",
        "Note",
    ]
    for sheet_name, rows in rows_by_sheet.items():
        sheet = workbook.create_sheet(sheet_name)
        sheet.append(header)
        for index, row in enumerate(rows, start=1):
            if repeat_header and index == 2:
                sheet.append(header)
            sheet.append(row)
    workbook.save(path)
    workbook.close()
    return path


def standard_row(
    board: str = "VRU-1",
    item: str = "Breaker VA47 100A",
    quantity: object = 2,
    rating: str = "100A",
    board_quantity: object = 1,
    model: str = "VA47",
    brand: str = "EKF",
    note: str = "synthetic",
) -> list[object]:
    return [
        board,
        f"Panel {board}",
        board_quantity,
        item,
        model,
        brand,
        rating,
        "pcs",
        quantity,
        note,
    ]


def text_page(
    board: str = "VRU-1", rating: str = "100A", quantity: int = 2
) -> list[tuple[str, int, int]]:
    return [
        (f"Switchboard schedule {board} qty 1", 72, 720),
        (f"Breaker VA47 EKF {rating} {quantity} pcs", 72, 690),
        ("Project note: verify enclosure and supply boundary", 72, 660),
    ]


def first_item(draft: dict[str, Any]) -> dict[str, Any]:
    return cast(dict[str, Any], cast(list[dict[str, Any]], draft["items"])[0])


def first_component(draft: dict[str, Any]) -> dict[str, Any]:
    return cast(dict[str, Any], first_item(draft)["components"][0])


def synthetic_component(
    label: str,
    *,
    quantity: int | float | None = 1,
    model: str | None = None,
    rating: str | None = None,
    raw_text: str | None = None,
    conflicts: list[dict[str, Any]] | None = None,
) -> Any:
    return extraction.ComponentCandidate(
        label=label,
        quantity=quantity,
        unit="шт.",
        model=model,
        brand=None,
        rating=rating,
        note=None,
        provenance=[
            extraction.Provenance(
                source_file="synthetic.txt",
                source_type="manual",
                locator="row=1",
                raw_text=raw_text or label,
                confidence=0.9,
                reason="bounded synthetic regression fixture",
            )
        ],
        confidence=0.9,
        conflicts=conflicts or [],
    )


def write_section_intake(
    path: Path,
    documents: list[dict[str, str]],
    *,
    project_id: str = "2024/086",
) -> Path:
    path.write_text(
        json.dumps(
            {
                "intake_version": "section_aware_extraction_intake.v0.1",
                "project_id": project_id,
                "source_documents": documents,
            },
            ensure_ascii=False,
        ),
        encoding="utf-8",
    )
    return path


def section_document(
    path: Path,
    document_id: str,
    *,
    section_id: str = "13",
    discipline: str = "ЭОМ",
) -> dict[str, str]:
    return {
        "path": path.name,
        "source_document_id": document_id,
        "section_id": section_id,
        "discipline": discipline,
        "source_role": "project_pdf",
    }


def test_pdf_with_one_switchboard(tmp_path: Path) -> None:
    pdf = write_text_pdf(tmp_path / "project.pdf", [text_page()])

    result = extraction.extract_pdf(pdf)

    assert result.pages[0]["status"] == "text_available"
    assert [board.normalized for board in result.boards] == ["VRU-1"]


def test_pdf_with_multiple_switchboards(tmp_path: Path) -> None:
    page = text_page("VRU-1") + text_page("AVR-2")
    pdf = write_text_pdf(tmp_path / "project.pdf", [page])

    result = extraction.extract_pdf(pdf)

    assert {board.normalized for board in result.boards} == {"VRU-1", "AVR-2"}


def test_pdf_switchboards_on_different_pages(tmp_path: Path) -> None:
    pdf = write_text_pdf(
        tmp_path / "project.pdf", [text_page("VRU-1"), text_page("AVR-2")]
    )

    result = extraction.extract_pdf(pdf)

    assert len(result.pages) == 2
    assert {value.page for board in result.boards for value in board.provenance} == {
        1,
        2,
    }


def test_pdf_specification_component_row(tmp_path: Path) -> None:
    pdf = write_text_pdf(tmp_path / "project.pdf", [text_page()])

    result = extraction.extract_pdf(pdf)

    assert result.boards[0].components[0].quantity == 2
    assert result.boards[0].components[0].brand == "EKF"


def test_pdf_broken_block_order_is_low_confidence(tmp_path: Path) -> None:
    page = [
        ("Switchboard schedule VRU-1 qty 1", 72, 500),
        ("Breaker VA47 EKF 100A 2 pcs", 72, 700),
        ("Project note long enough for extraction", 72, 400),
        ("Supply boundary must be checked", 72, 650),
    ]
    pdf = write_text_pdf(tmp_path / "project.pdf", [page])

    result = extraction.extract_pdf(pdf)

    assert result.pages[0]["status"] == "low_text_confidence"
    assert result.pages[0]["block_order_suspect"] is True


def test_pdf_without_text_layer_is_not_success(tmp_path: Path) -> None:
    pdf = write_text_pdf(tmp_path / "scan.pdf", [[]])

    result = extraction.extract_pdf(pdf)

    assert result.pages[0]["status"] == "image_only"
    assert result.boards == []


def test_blank_pdf_page_is_unreadable_not_image_only(tmp_path: Path) -> None:
    pdf = write_text_pdf(tmp_path / "blank.pdf", [[]], image_for_empty=False)

    result = extraction.extract_pdf(pdf)

    assert result.pages[0]["status"] == "unreadable"


def test_mixed_pdf_keeps_text_and_image_only_pages(tmp_path: Path) -> None:
    pdf = write_text_pdf(tmp_path / "mixed.pdf", [text_page(), []])

    result = extraction.extract_pdf(pdf)

    assert [page["status"] for page in result.pages] == ["text_available", "image_only"]
    assert len(result.boards) == 1


def test_corrupt_pdf_is_classified(tmp_path: Path) -> None:
    pdf = tmp_path / "corrupt.pdf"
    pdf.write_bytes(b"not a pdf")

    result = extraction.extract_pdf(pdf)

    assert result.status == "corrupt"
    assert result.pages == [{"page": 0, "status": "corrupt"}]


def test_protected_pdf_is_classified(tmp_path: Path) -> None:
    source = write_text_pdf(tmp_path / "source.pdf", [text_page()])
    reader = PdfReader(source)
    writer = PdfWriter()
    writer.append_pages_from_reader(reader)
    writer.encrypt("secret")
    protected = tmp_path / "protected.pdf"
    with protected.open("wb") as output:
        writer.write(output)

    result = extraction.extract_pdf(protected)

    assert result.status == "encrypted_or_protected"
    assert result.pages == [{"page": 0, "status": "encrypted_or_protected"}]


def test_pdf_provenance_has_page_and_locator(tmp_path: Path) -> None:
    pdf = write_text_pdf(tmp_path / "project.pdf", [text_page()])

    result = extraction.extract_pdf(pdf)
    provenance = result.boards[0].provenance[0].as_dict()

    assert provenance["page"] == 1
    assert "page=1" in provenance["locator"]
    assert provenance["block_coordinates"] != ""


def test_workbook_with_one_switchboard(tmp_path: Path) -> None:
    workbook = write_spec_workbook(tmp_path / "spec.xlsx", {"Spec": [standard_row()]})

    result = extraction.extract_workbook(workbook)

    assert [board.normalized for board in result.boards] == ["VRU-1"]
    assert len(result.boards[0].components) == 1


def test_workbook_with_multiple_switchboards_on_sheet(tmp_path: Path) -> None:
    workbook = write_spec_workbook(
        tmp_path / "spec.xlsx",
        {"Spec": [standard_row("VRU-1"), standard_row("AVR-2")]},
    )

    result = extraction.extract_workbook(workbook)

    assert {board.normalized for board in result.boards} == {"VRU-1", "AVR-2"}


def test_workbook_switchboards_on_different_sheets(tmp_path: Path) -> None:
    workbook = write_spec_workbook(
        tmp_path / "spec.xlsx",
        {"Input": [standard_row("VRU-1")], "Output": [standard_row("AVR-2")]},
    )

    result = extraction.extract_workbook(workbook)

    assert len(result.sheets) == 2
    assert {board.normalized for board in result.boards} == {"VRU-1", "AVR-2"}


def test_workbook_merged_board_cells(tmp_path: Path) -> None:
    workbook_path = write_spec_workbook(
        tmp_path / "spec.xlsx",
        {"Spec": [standard_row(), standard_row(item="Contactor KM1", quantity=1)]},
    )
    workbook = load_workbook(workbook_path)
    sheet = workbook["Spec"]
    sheet.merge_cells("A2:A3")
    workbook.save(workbook_path)
    workbook.close()

    result = extraction.extract_workbook(workbook_path)

    assert len(result.boards) == 1
    assert len(result.boards[0].components) == 2


def test_workbook_repeated_headers_are_skipped(tmp_path: Path) -> None:
    workbook = write_spec_workbook(
        tmp_path / "spec.xlsx",
        {"Spec": [standard_row(), standard_row(item="Contactor KM1", quantity=1)]},
        repeat_header=True,
    )

    result = extraction.extract_workbook(workbook)

    assert len(result.boards[0].components) == 2


def test_workbook_missing_quantity_requires_review(tmp_path: Path) -> None:
    workbook = write_spec_workbook(
        tmp_path / "spec.xlsx", {"Spec": [standard_row(quantity=None)]}
    )

    result = extraction.extract_workbook(workbook)

    assert result.boards[0].components[0].quantity is None
    assert (
        "missing or doubtful component quantity"
        in result.boards[0].components[0].red_flags
    )


def test_workbook_possible_duplicate_requires_review(tmp_path: Path) -> None:
    workbook = write_spec_workbook(
        tmp_path / "spec.xlsx", {"Spec": [standard_row(), standard_row()]}
    )

    result = extraction.extract_workbook(workbook)

    assert "possible duplicate workbook row" in result.boards[0].components[1].red_flags


def test_workbook_provenance_has_sheet_row_and_cells(tmp_path: Path) -> None:
    workbook = write_spec_workbook(tmp_path / "spec.xlsx", {"Spec": [standard_row()]})

    result = extraction.extract_workbook(workbook)
    provenance = result.boards[0].components[0].provenance[0].as_dict()

    assert provenance["sheet"] == "Spec"
    assert provenance["row"] == 2
    assert provenance["cell_range"] == "A2:J2"


def mixed_artifacts(
    tmp_path: Path,
    *,
    pdf_board: str = "VRU-1",
    workbook_board: str = "VRU-1",
    pdf_rating: str = "100A",
    workbook_rating: str = "100A",
    pdf_quantity: int = 2,
    workbook_quantity: object = 2,
) -> Any:
    pdf = write_text_pdf(
        tmp_path / "project.pdf",
        [text_page(pdf_board, pdf_rating, pdf_quantity)],
    )
    workbook = write_spec_workbook(
        tmp_path / "spec.xlsx",
        {
            "Spec": [
                standard_row(
                    workbook_board,
                    quantity=workbook_quantity,
                    rating=workbook_rating,
                )
            ]
        },
    )
    return extraction.build_artifacts(pdf, workbook)


def sparse_pdf_mixed_artifacts(
    tmp_path: Path,
    workbook_row: list[object],
) -> Any:
    pdf = write_text_pdf(
        tmp_path / "project.pdf",
        [
            [
                ("Switchboard schedule VRU-1 qty 1", 72, 720),
                ("Breaker VA47 2 pcs", 72, 690),
                ("Project note: verify supply boundary", 72, 660),
            ]
        ],
    )
    workbook = write_spec_workbook(tmp_path / "spec.xlsx", {"Spec": [workbook_row]})
    return extraction.build_artifacts(pdf, workbook)


def test_same_switchboard_is_matched_between_sources(tmp_path: Path) -> None:
    artifacts = mixed_artifacts(tmp_path)

    assert artifacts.summary["switchboards_matched"] == 1
    assert len(artifacts.draft["items"]) == 1


def test_pdf_designation_and_workbook_composition_are_combined(tmp_path: Path) -> None:
    artifacts = mixed_artifacts(tmp_path)
    item = first_item(artifacts.draft)

    assert item["normalized_designation"] == "VRU-1"
    assert {value["source_type"] for value in item["provenance"]} == {"pdf", "workbook"}
    assert item["components"]


def test_matching_component_quantity_merges_without_quantity_conflict(
    tmp_path: Path,
) -> None:
    artifacts = mixed_artifacts(tmp_path)
    component = first_component(artifacts.draft)

    assert component["quantity_guess"] == 2
    assert not any("quantity" in value["type"] for value in component["conflicts"])


def test_workbook_brand_enriches_sparse_pdf_component(tmp_path: Path) -> None:
    artifacts = sparse_pdf_mixed_artifacts(
        tmp_path,
        standard_row(
            item="Breaker VA47",
            rating="",
            brand="EKF",
            note="",
        ),
    )
    component = first_component(artifacts.draft)

    assert component["brand_guess"] == "EKF"
    assert not any(
        value["type"] == "component_brand_mismatch" for value in component["conflicts"]
    )


def test_workbook_rating_enriches_sparse_pdf_component(tmp_path: Path) -> None:
    artifacts = sparse_pdf_mixed_artifacts(
        tmp_path,
        standard_row(
            item="Breaker VA47",
            rating="100A",
            brand="",
            note="",
        ),
    )
    component = first_component(artifacts.draft)

    assert component["rating_guess"] == "100A"
    assert not any(
        value["type"] == "component_rating_mismatch" for value in component["conflicts"]
    )


def test_workbook_note_enriches_sparse_pdf_component(tmp_path: Path) -> None:
    artifacts = sparse_pdf_mixed_artifacts(
        tmp_path,
        standard_row(
            item="Breaker VA47",
            rating="",
            brand="",
            note="Mount on DIN rail",
        ),
    )
    component = first_component(artifacts.draft)

    assert component["note_guess"] == "Mount on DIN rail"
    assert {value["source_type"] for value in component["provenance"]} == {
        "pdf",
        "workbook",
    }


def test_different_notes_create_conflict_without_automatic_choice(
    tmp_path: Path,
) -> None:
    workbook = write_spec_workbook(
        tmp_path / "spec.xlsx",
        {
            "Spec": [
                standard_row(note="Install left"),
                standard_row(note="Install right"),
            ]
        },
    )

    artifacts = extraction.build_artifacts(None, workbook)
    component = first_component(artifacts.draft)

    assert component["note_guess"] is None
    assert any(
        value["type"] == "component_note_mismatch" for value in component["conflicts"]
    )


def test_component_quantity_conflict_is_not_silently_resolved(tmp_path: Path) -> None:
    artifacts = mixed_artifacts(tmp_path, pdf_quantity=2, workbook_quantity=3)
    component = first_component(artifacts.draft)

    assert component["quantity_guess"] is None
    assert any(
        value["type"] == "component_quantity_mismatch"
        for value in component["conflicts"]
    )


def test_component_rating_conflict_is_not_silently_resolved(tmp_path: Path) -> None:
    artifacts = mixed_artifacts(tmp_path, pdf_rating="100A", workbook_rating="125A")
    component = first_component(artifacts.draft)

    assert component["rating_guess"] is None
    assert any(
        value["type"] == "component_rating_mismatch" for value in component["conflicts"]
    )


def test_switchboard_only_in_pdf_is_unmatched(tmp_path: Path) -> None:
    artifacts = mixed_artifacts(tmp_path, workbook_board="AVR-2")

    assert artifacts.summary["switchboards_unmatched"] == 2
    assert any(
        value["type"] == "switchboard_present_in_one_source"
        for item in artifacts.draft["items"]
        for value in item["conflicts"]
    )


def test_switchboard_only_in_workbook_is_unmatched(tmp_path: Path) -> None:
    artifacts = mixed_artifacts(tmp_path, pdf_board="AVR-2")

    assert artifacts.summary["switchboards_unmatched"] == 2
    assert {item["normalized_designation"] for item in artifacts.draft["items"]} == {
        "AVR-2",
        "VRU-1",
    }


def test_ambiguous_designations_are_not_merged(tmp_path: Path) -> None:
    artifacts = mixed_artifacts(tmp_path, pdf_board="VRU-1A", workbook_board="VRU-1")

    assert artifacts.summary["switchboards_matched"] == 0
    assert any(
        value["type"] == "ambiguous_switchboard_match"
        for item in artifacts.draft["items"]
        for value in item["conflicts"]
    )


def test_similar_significant_numbers_are_not_merged(tmp_path: Path) -> None:
    artifacts = mixed_artifacts(tmp_path, pdf_board="VRU-1", workbook_board="VRU-11")

    assert artifacts.summary["switchboards_matched"] == 0
    assert len(artifacts.draft["items"]) == 2


def test_provenance_survives_merge(tmp_path: Path) -> None:
    artifacts = mixed_artifacts(tmp_path)
    component = first_component(artifacts.draft)

    assert {value["source_type"] for value in component["provenance"]} == {
        "pdf",
        "workbook",
    }
    assert all(value["locator"] for value in component["provenance"])


def test_conflict_appears_in_existing_review_card(tmp_path: Path) -> None:
    pdf = write_text_pdf(tmp_path / "project.pdf", [text_page(quantity=2)])
    workbook = write_spec_workbook(
        tmp_path / "spec.xlsx", {"Spec": [standard_row(quantity=3)]}
    )
    output_dir = tmp_path / "output"

    result = operator.run_operator(pdf, workbook, output_dir)
    markdown = (output_dir / operator.REVIEW_NAME).read_text(encoding="utf-8")

    assert result.status == "PASS", result.red_flags
    assert "## Требует проверки Игоря" in markdown
    assert "different component quantities between sources" in markdown


def test_workbook_note_is_visible_in_existing_review_card(tmp_path: Path) -> None:
    pdf = write_text_pdf(
        tmp_path / "project.pdf",
        [
            [
                ("Switchboard schedule VRU-1 qty 1", 72, 720),
                ("Breaker VA47 2 pcs", 72, 690),
                ("Project note: verify supply boundary", 72, 660),
            ]
        ],
    )
    workbook = write_spec_workbook(
        tmp_path / "spec.xlsx",
        {
            "Spec": [
                standard_row(
                    item="Breaker VA47",
                    rating="",
                    brand="",
                    note="Mount on DIN rail",
                )
            ]
        },
    )
    output_dir = tmp_path / "output"

    result = operator.run_operator(pdf, workbook, output_dir)
    markdown = (output_dir / operator.REVIEW_NAME).read_text(encoding="utf-8")

    assert result.status == "PASS", result.red_flags
    assert "note_guess" in markdown
    assert "Mount on DIN rail" in markdown


def test_human_approval_is_never_set_automatically(tmp_path: Path) -> None:
    artifacts = mixed_artifacts(tmp_path)
    safety = artifacts.draft["safety"]

    assert all(value is False for key, value in safety.items() if key != "status")
    serialized = json.dumps(artifacts.draft)
    for forbidden in (
        '"composition_approved": true',
        '"price_confirmed_by_igor": true',
        '"commercial_csv_approved": true',
        '"quote_approved": true',
        '"client_send_approved": true',
        '"procurement_approved": true',
        '"production_approved": true',
    ):
        assert forbidden not in serialized


def test_generated_draft_passes_existing_preliminary_validator(tmp_path: Path) -> None:
    pdf = write_text_pdf(tmp_path / "project.pdf", [text_page()])
    output_dir = tmp_path / "output"

    result = operator.run_operator(pdf, None, output_dir)

    assert result.status == "PASS", result.red_flags
    assert result.checks["preliminary draft validation"] == "pass"
    assert result.checks["source bundle verification and review card"] == "pass"


def test_old_pdf_mode_keeps_v01_contract_without_section_fields(tmp_path: Path) -> None:
    pdf = write_text_pdf(tmp_path / "project.pdf", [text_page()])

    draft = extraction.build_artifacts(pdf, None).draft

    assert draft["schema_version"] == "preliminary_composition_draft.v0.1"
    serialized = json.dumps(draft, ensure_ascii=False)
    for field_name in (
        "project_id",
        "section_id",
        "discipline",
        "source_document_id",
        "source_role",
        "source_designation",
    ):
        assert f'"{field_name}"' not in serialized


def test_same_designation_in_different_sections_stays_separate(tmp_path: Path) -> None:
    first_pdf = write_text_pdf(tmp_path / "section-12.pdf", [text_page()])
    second_pdf = write_text_pdf(tmp_path / "section-13.pdf", [text_page()])
    intake = write_section_intake(
        tmp_path / "intake.json",
        [
            section_document(first_pdf, "section-12-eom", section_id="12"),
            section_document(second_pdf, "section-13-eom", section_id="13"),
        ],
    )

    draft = operator.build_section_aware_artifacts(intake).draft

    assert len(draft["items"]) == 2
    assert {item["section_id"] for item in draft["items"]} == {"12", "13"}
    assert {item["normalized_designation"] for item in draft["items"]} == {"VRU-1"}


def test_same_designation_in_different_disciplines_stays_separate(
    tmp_path: Path,
) -> None:
    eom_pdf = write_text_pdf(tmp_path / "eom.pdf", [text_page()])
    eof_pdf = write_text_pdf(tmp_path / "eof.pdf", [text_page()])
    intake = write_section_intake(
        tmp_path / "intake.json",
        [
            section_document(eom_pdf, "section-11-eom", section_id="11"),
            section_document(
                eof_pdf,
                "section-11-eof",
                section_id="11",
                discipline="ЭОФ",
            ),
        ],
    )

    items = operator.build_section_aware_artifacts(intake).draft["items"]

    assert len(items) == 2
    assert {item["discipline"] for item in items} == {"ЭОМ", "ЭОФ"}


def test_same_designation_in_different_source_documents_stays_separate(
    tmp_path: Path,
) -> None:
    first_pdf = write_text_pdf(tmp_path / "part-a.pdf", [text_page()])
    second_pdf = write_text_pdf(tmp_path / "part-b.pdf", [text_page()])
    intake = write_section_intake(
        tmp_path / "intake.json",
        [
            section_document(first_pdf, "part-a"),
            section_document(second_pdf, "part-b"),
        ],
    )

    items = operator.build_section_aware_artifacts(intake).draft["items"]

    assert len(items) == 2
    assert {item["source_document_id"] for item in items} == {"part-a", "part-b"}


def test_same_identity_merges_evidence_from_multiple_pages(tmp_path: Path) -> None:
    pdf = write_text_pdf(tmp_path / "two-pages.pdf", [text_page(), text_page()])
    intake = write_section_intake(
        tmp_path / "intake.json", [section_document(pdf, "two-pages")]
    )

    item = first_item(operator.build_section_aware_artifacts(intake).draft)

    assert {value["page"] for value in item["provenance"]} == {1, 2}
    assert {value["source_document_id"] for value in item["provenance"]} == {
        "two-pages"
    }


@pytest.mark.parametrize(
    "missing_field",
    ["source_document_id", "section_id", "discipline", "source_role"],
)
def test_section_intake_rejects_missing_document_metadata(
    tmp_path: Path, missing_field: str
) -> None:
    pdf = write_text_pdf(tmp_path / "project.pdf", [text_page()])
    document = section_document(pdf, "project")
    del document[missing_field]
    intake = write_section_intake(tmp_path / "intake.json", [document])

    with pytest.raises(operator.ExtractionError, match="missing required fields"):
        operator.build_section_aware_artifacts(intake)


def test_section_intake_rejects_missing_project_id(tmp_path: Path) -> None:
    pdf = write_text_pdf(tmp_path / "project.pdf", [text_page()])
    intake = write_section_intake(
        tmp_path / "intake.json", [section_document(pdf, "project")], project_id=""
    )

    with pytest.raises(operator.ExtractionError, match="project_id"):
        operator.build_section_aware_artifacts(intake)


def test_source_document_id_format_and_uniqueness_are_fail_closed(
    tmp_path: Path,
) -> None:
    first_pdf = write_text_pdf(tmp_path / "first.pdf", [text_page()])
    second_pdf = write_text_pdf(tmp_path / "second.pdf", [text_page()])
    invalid = write_section_intake(
        tmp_path / "invalid.json", [section_document(first_pdf, "bad id")]
    )
    duplicate = write_section_intake(
        tmp_path / "duplicate.json",
        [
            section_document(first_pdf, "same-id", section_id="12"),
            section_document(second_pdf, "same-id", section_id="13"),
        ],
    )

    with pytest.raises(operator.ExtractionError, match="must match"):
        operator.build_section_aware_artifacts(invalid)
    with pytest.raises(operator.ExtractionError, match="conflicting metadata"):
        operator.build_section_aware_artifacts(duplicate)


def test_section_provenance_resolves_canonical_source_and_item(tmp_path: Path) -> None:
    pdf = write_text_pdf(tmp_path / "project.pdf", [text_page()])
    intake = write_section_intake(
        tmp_path / "intake.json", [section_document(pdf, "section-13-eom")]
    )

    draft = operator.build_section_aware_artifacts(intake).draft
    source = draft["source"]["source_documents"][0]
    item = draft["items"][0]
    component = item["components"][0]

    assert source["source_document_id"] == item["source_document_id"]
    assert item["provenance"][0]["item_id"] == item["item_id"]
    assert component["provenance"][0]["item_id"] == item["item_id"]
    assert component["provenance"][0]["component_id"] == component["component_id"]
    assert Path(source["resolved_path"]) == pdf.resolve()
    assert "resolved_path" not in item
    assert "resolved_path" not in component


def test_section_aware_operator_passes_new_validator(tmp_path: Path) -> None:
    pdf = write_text_pdf(tmp_path / "project.pdf", [text_page()])
    intake = write_section_intake(
        tmp_path / "intake.json", [section_document(pdf, "section-13-eom")]
    )
    output = tmp_path / "section-output"

    result = operator.run_operator(None, None, output, section_aware_intake=intake)

    assert result.status == "PASS", result.red_flags
    draft = json.loads((output / operator.DRAFT_NAME).read_text(encoding="utf-8"))
    assert draft["schema_version"] == operator.SECTION_AWARE_SCHEMA_VERSION


def test_section_aware_and_v01_inputs_are_mutually_exclusive(tmp_path: Path) -> None:
    pdf = write_text_pdf(tmp_path / "project.pdf", [text_page()])
    intake = write_section_intake(
        tmp_path / "intake.json", [section_document(pdf, "project")]
    )

    result = operator.run_operator(
        pdf,
        None,
        tmp_path / "output",
        section_aware_intake=intake,
    )

    assert result.status == "FAIL"
    assert "cannot be combined" in result.red_flags[0]
    with pytest.raises(SystemExit):
        operator.parse_args(
            [
                "--project-pdf",
                str(pdf),
                "--section-aware-intake",
                str(intake),
                "--output-dir",
                str(tmp_path / "cli-output"),
            ]
        )


def test_cli_requires_at_least_one_source(tmp_path: Path) -> None:
    result = operator.run_operator(None, None, tmp_path / "output")

    assert result.status == "FAIL"
    assert "at least one source" in result.red_flags[0]
    assert not result.output_dir.exists()


def test_cli_rejects_output_inside_git(tmp_path: Path) -> None:
    pdf = write_text_pdf(tmp_path / "project.pdf", [text_page()])
    output = PROJECT_ROOT / "unsafe-generated-output"

    result = operator.run_operator(pdf, None, output)

    assert result.status == "FAIL"
    assert not output.exists()


def test_pdf_only_operator_smoke(tmp_path: Path) -> None:
    pdf = write_text_pdf(tmp_path / "project.pdf", [text_page()])
    output = tmp_path / "pdf-output"

    result = operator.run_operator(pdf, None, output)

    assert result.status == "PASS", result.red_flags
    assert (output / operator.REVIEW_NAME).is_file()


def test_workbook_only_operator_smoke(tmp_path: Path) -> None:
    workbook = write_spec_workbook(tmp_path / "spec.xlsx", {"Spec": [standard_row()]})
    output = tmp_path / "workbook-output"

    result = operator.run_operator(None, workbook, output)

    assert result.status == "PASS", result.red_flags
    assert (output / operator.REVIEW_NAME).is_file()


def test_mixed_source_operator_smoke(tmp_path: Path) -> None:
    pdf = write_text_pdf(tmp_path / "project.pdf", [text_page()])
    workbook = write_spec_workbook(tmp_path / "spec.xlsx", {"Spec": [standard_row()]})
    output = tmp_path / "mixed-output"

    result = operator.run_operator(pdf, workbook, output)

    assert result.status == "PASS"
    assert result.summary["switchboards_matched"] == 1
    assert (output / operator.MANIFEST_NAME).is_file()
    assert (output / operator.DRAFT_NAME).is_file()
    assert (output / operator.REVIEW_NAME).is_file()


def pdf_block(text: str, x: float, y: float, number: int) -> Any:
    return extraction.PdfBlock(text, x, y, number)


def test_schematic_qf_line_with_qf0_and_qf1_is_split() -> None:
    segments = extraction.split_qf_segments_from_text(
        "QF0ВН-323P 25ААВР QF1АВДТ32 2PC16/30мА розеточная сеть"
    )

    assert len(segments) == 2
    assert segments[0].startswith("QF0")
    assert segments[1].startswith("QF1")


def test_schematic_title_without_space_finds_board_designations() -> None:
    assert extraction.schematic_board_designations(
        "Формат А4 Принципиальная схема группового щитаЩО6"
    ) == ["ЩО6"]
    assert extraction.schematic_board_designations(
        "Формат А3 Принципиальная схема группового щитаНЩР17"
    ) == ["НЩР17"]
    assert extraction.schematic_board_designations(
        "Формат А3 Принципиальная схема группового щитаАВР17"
    ) == ["АВР17"]


def test_source_reference_to_vru_is_classified_as_reference() -> None:
    assert extraction.is_source_reference("от ВРУ 3А")
    assert extraction.is_source_reference("питание от ВРУ 3А")
    assert extraction.is_source_reference("существующая группа от ВРУ")
    assert not extraction.is_source_reference(
        "Принципиальная схема группового щитаВРУ1"
    )


def test_real_schematic_title_vru_is_not_globally_blocked() -> None:
    blocks = [
        pdf_block("Принципиальная схема группового щита", 10, 10, 1),
        pdf_block("ВРУ", 20, 20, 2),
        pdf_block("1", 30, 20, 3),
    ]

    titles = extraction.find_schematic_board_titles(
        blocks, "Принципиальная схема группового щитаВРУ1"
    )

    assert [title.normalized for title in titles] == ["ВРУ-1"]


def test_qf_apparatus_tokens_are_parsed_deterministically() -> None:
    cases = {
        "QF0 ВА88-32 3P 63А": ("ВА88-32", "3P", "63А"),
        "QF0 ВН-32 3P 25А": ("ВН-32", "3P", "25А"),
        "QF1 АВДТ32 2P C16/30мА": ("АВДТ32", "2P", "C16/30мА"),
        "QF2 АВДТ32 2P C20/30мА": ("АВДТ32", "2P", "C20/30мА"),
    }

    for raw, expected in cases.items():
        parsed = extraction.parse_qf_apparatus(raw)

        assert (parsed["model"], parsed["poles"], parsed["rating"]) == expected


def test_plan_without_schematic_evidence_does_not_create_schematic_title() -> None:
    blocks = [
        pdf_block("План розеточных сетей 2-го этажа", 10, 10, 1),
        pdf_block("ЩО.6", 20, 20, 2),
        pdf_block("АВР", 30, 20, 3),
        pdf_block("НЩР", 40, 20, 4),
    ]

    assert extraction.find_schematic_board_titles(blocks, "") == []


def test_multiple_schematic_zones_can_leave_qf_assignment_ambiguous() -> None:
    titles = [
        extraction.SchematicBoardTitle(
            "ЩО6",
            "ЩО-6",
            "Принципиальная схема группового щита ЩО6",
            pdf_block("", 0, 0, 1),
        ),
        extraction.SchematicBoardTitle(
            "НЩР17",
            "НЩР-17",
            "Принципиальная схема группового щита НЩР17",
            pdf_block("", 20, 0, 2),
        ),
    ]
    segment = extraction.QfSegment(
        "QF1",
        "QF1 АВДТ32 2P C16/30мА",
        pdf_block("", 10, 0, 3),
    )

    assert extraction.nearest_schematic_board(segment, titles) is None


def test_qf_segments_from_blocks_reports_detected_segments() -> None:
    segments = extraction.qf_segments_from_blocks(
        [
            pdf_block("QF0", 72, 690, 1),
            pdf_block("ВА88-32", 72, 680, 2),
            pdf_block("3P 63А", 72, 670, 3),
            pdf_block("QF1", 172, 690, 4),
            pdf_block("АВДТ32 2P", 172, 680, 5),
            pdf_block("C16/30мА", 172, 670, 6),
        ],
        "",
    )

    assert [segment.designation for segment in segments] == ["QF0", "QF1"]


def test_schematic_qf_component_key_keeps_same_model_qfs_distinct() -> None:
    first = extraction.qf_component_from_segment(
        extraction.QfSegment(
            "QF1",
            "QF1 АВДТ32 2P C16/30мА",
            pdf_block("QF1", 1, 1, 1),
        ),
        "project.pdf",
        1,
        0.82,
    )
    second = extraction.qf_component_from_segment(
        extraction.QfSegment(
            "QF2",
            "QF2 АВДТ32 2P C16/30мА",
            pdf_block("QF2", 2, 1, 2),
        ),
        "project.pdf",
        1,
        0.82,
    )

    assert extraction.component_key(first) == "qf:QF1"
    assert extraction.component_key(second) == "qf:QF2"


def test_schematic_qf_provenance_keeps_page_raw_segment_and_locator() -> None:
    component = extraction.qf_component_from_segment(
        extraction.QfSegment(
            "QF1",
            "QF1 АВДТ32 2P C20/30мА",
            pdf_block("QF1", 72, 690, 3),
        ),
        "project.pdf",
        1,
        0.82,
    )

    provenance = component.provenance[0].as_dict()

    assert provenance["page"] == 1
    assert "QF1" in provenance["raw_text"]
    assert "qf=QF1" in provenance["locator"]


def specification_page(
    rows: list[list[tuple[str, int, int]]],
    *,
    include_quantity_header: bool = True,
) -> list[tuple[str, int, int]]:
    header = [
        ("Position", 50, 740),
        ("Description", 150, 740),
        ("Type/model", 300, 740),
        ("Unit", 430, 740),
    ]
    if include_quantity_header:
        header.append(("Quantity", 500, 740))
    return header + [fragment for row in rows for fragment in row]


def specification_blocks(
    rows: list[list[tuple[str, int, int]]],
    *,
    include_quantity_header: bool = True,
) -> list[Any]:
    return [
        pdf_block(text, x, y, number)
        for number, (text, x, y) in enumerate(
            specification_page(
                rows,
                include_quantity_header=include_quantity_header,
            ),
            start=1,
        )
    ]


def test_gated_specification_extracts_top_level_fields_and_provenance() -> None:
    blocks = specification_blocks(
        [
            [("1.1", 50, 650), ("VRU-1", 55, 640), ("pcs", 430, 645), ("2", 500, 645)],
            [("1.2", 50, 550), ("AVR-2", 55, 540), ("pcs", 430, 545), ("3", 500, 545)],
            [("1.3", 50, 450), ("SHRS-3", 55, 440), ("pcs", 430, 445), ("4", 500, 445)],
        ]
    )

    result = extraction.extract_gated_specification_page(
        "stable-specification.pdf", 1, blocks, "usable synthetic text layer"
    )

    assert result.gated
    assert [board.normalized for board in result.boards] == [
        "VRU-1",
        "AVR-2",
        "SHRS-3",
    ]
    assert [board.quantity for board in result.boards] == [2, 3, 4], [
        board.red_flags for board in result.boards
    ]
    assert all(not board.components for board in result.boards)
    provenance = result.boards[0].provenance[0].as_dict()
    assert provenance["page"] == 1
    assert "unit=pcs" in provenance["locator"]
    assert provenance["raw_text"] == "1.1 | pcs | 2 | VRU-1"


def test_gated_specification_ambiguous_quantities_stay_null() -> None:
    blocks = specification_blocks(
        [
            [
                ("1.1", 50, 650),
                ("VRU-1", 55, 640),
                ("pcs", 430, 645),
                ("2", 500, 646),
                ("7", 500, 642),
            ],
            [
                ("1.2", 50, 550),
                ("AVR-2", 55, 540),
                ("8", 0, 0),
                ("pcs", 430, 545),
                ("3", 500, 545),
            ],
            [("1.3", 50, 450), ("SHRS-3", 55, 440), ("pcs", 430, 445), ("4", 500, 445)],
        ]
    )

    boards = extraction.extract_gated_specification_page(
        "ambiguous-quantity.pdf", 1, blocks, "usable synthetic text layer"
    ).boards

    assert [board.quantity for board in boards] == [None, None, 4], [
        board.red_flags for board in boards
    ]
    assert "multiple numeric fragments" in " ".join(boards[0].red_flags)
    assert "orphan zero-coordinate" in " ".join(boards[1].red_flags)


def test_gated_specification_excludes_component_tail() -> None:
    blocks = specification_blocks(
        [
            [("1.1", 50, 650), ("VRU-1", 55, 640), ("pcs", 430, 645), ("1", 500, 645)],
            [
                ("Breaker VA47 component tail 12 pcs", 150, 570),
                ("VA47-63", 300, 570),
            ],
            [("1.2", 50, 550), ("AVR-2", 55, 540), ("pcs", 430, 545), ("2", 500, 545)],
            [("1.3", 50, 450), ("SHRS-3", 55, 440), ("pcs", 430, 445), ("3", 500, 445)],
        ]
    )

    boards = extraction.extract_gated_specification_page(
        "tail-specification.pdf", 1, blocks, "usable synthetic text layer"
    ).boards

    assert [board.normalized for board in boards] == ["VRU-1", "AVR-2", "SHRS-3"]
    assert all(not board.components for board in boards)
    assert all(
        "component tail" not in provenance.raw_text
        for board in boards
        for provenance in board.provenance
    )


def test_broken_specification_gate_and_model_substring_fail_closed() -> None:
    blocks = specification_blocks(
        [
            [("8. Functional group heading", 150, 680)],
            [
                ("1.1", 50, 650),
                ("Щит распределительный настенный с дверцей на 12 мод. IP65", 150, 640),
                ("ЩРН-12", 300, 640),
                ("шт.", 430, 640),
                ("1", 500, 640),
            ],
            [
                ("1.2", 50, 550),
                ("Вводное устройство для лифтов", 150, 540),
                ("ЯРВ-100", 300, 540),
                ("шт.", 430, 540),
                ("2", 500, 540),
            ],
            [
                ("1.3", 50, 450),
                ("ЩК", 55, 440),
                (
                    "Корпус настенный с дверью, размеры 600х400х200",
                    150,
                    440,
                ),
                ("ЩРВ-П-18", 300, 440),
                ("шт.", 430, 440),
                ("80", 500, 440),
            ],
            [
                ("1.4", 50, 350),
                ("АВР-9", 300, 340),
                ("шт.", 430, 340),
                ("1", 500, 340),
            ],
        ]
    )

    result = extraction.extract_gated_specification_page(
        "valid-specification.pdf", 1, blocks, "usable synthetic text layer"
    )

    assert result.gated
    assert [board.normalized for board in result.boards] == ["ЯРВ-100", "ЩК"]
    assert [board.quantity for board in result.boards] == [2, 80]
    assert all(board.normalized != "ЩР" for board in result.boards)
    assert sum("generic enclosure" in value for value in result.diagnostics) == 1
    assert (
        sum("description cluster is empty" in value for value in result.diagnostics)
        == 1
    )
    yarv_provenance = result.boards[0].provenance[0].raw_text
    assert "Вводное устройство для лифтов" in yarv_provenance
    assert "Functional group heading" not in yarv_provenance
    assert extraction.specification_literal_designation("ЩРВ-П-18") is None
    assert extraction.specification_board_model("ЩРВ-П-18") == "ЩРВ-П-18"
    assert extraction.normalize_specification_model("ЩРВ-П-18") == "ЩРВ-П-18"


def test_four_page_v01_semantics_are_unchanged_except_created_at(
    tmp_path: Path,
) -> None:
    pdf = write_text_pdf(
        tmp_path / "four-page-v01.pdf",
        [text_page(board=f"VRU-{index}") for index in range(1, 5)],
    )

    default_draft = extraction.build_artifacts(pdf, None).draft
    explicit_legacy_draft = extraction.build_artifacts(
        pdf,
        None,
        specification_rows=False,
    ).draft
    default_created_at = default_draft.pop("created_at")
    explicit_created_at = explicit_legacy_draft.pop("created_at")

    assert default_created_at
    assert explicit_created_at
    assert default_draft == explicit_legacy_draft
    assert default_draft["schema_version"] == "preliminary_composition_draft.v0.1"


def test_exact_n_pe_bus_has_no_false_rating_blocker() -> None:
    component = synthetic_component("Шина N/PE", quantity=2)

    draft = extraction.component_to_draft(component, "COMP-NPE-001")

    assert draft["component_id"] == "COMP-NPE-001"
    assert draft["quantity_guess"] == 2
    assert "rating_guess" not in draft["missing_fields"]
    assert draft["field_applicability"] == [
        {
            "field": "rating_guess",
            "status": "NOT_APPLICABLE_WITH_REASON",
            "reason": (
                "Exact normalized N/PE bus identity has no separate rating field."
            ),
            "source": "contract",
        }
    ]


def test_n_pe_bus_missing_quantity_stays_missing_without_default() -> None:
    component = synthetic_component("Шина PE", quantity=None)

    draft = extraction.component_to_draft(component, "COMP-NPE-002")

    assert draft["quantity_guess"] is None
    assert "quantity_guess" in draft["missing_fields"]
    assert "rating_guess" not in draft["missing_fields"]


def test_ambiguous_bus_identity_does_not_use_non_applicability() -> None:
    component = synthetic_component("Шина N дополнительная", quantity=1)

    draft = extraction.component_to_draft(component, "COMP-NPE-UNKNOWN")

    assert "field_applicability" not in draft
    assert "rating_guess" in draft["missing_fields"]


def test_meter_model_semantics_preserve_section_specific_variants() -> None:
    first = synthetic_component(
        "Счетчик электроэнергии",
        model="Mercury 230 ART-01",
    )
    second = synthetic_component(
        "Счетчик электроэнергии",
        model="Mercury 230 ART-02",
    )

    drafts = [
        extraction.component_to_draft(first, "SECTION-12-COMP-001"),
        extraction.component_to_draft(second, "SECTION-13-COMP-001"),
    ]

    assert [value["model_guess"] for value in drafts] == [
        "Mercury 230 ART-01",
        "Mercury 230 ART-02",
    ]
    assert all("rating_guess" not in value["missing_fields"] for value in drafts)
    assert all(
        value["field_applicability"][0]["status"] == "MODEL_OR_TYPE_SEMANTICS"
        for value in drafts
    )


def test_meter_with_missing_model_fails_closed_without_variant_repair() -> None:
    component = synthetic_component("Счетчик электроэнергии", model=None)

    draft = extraction.component_to_draft(component, "COMP-METER-BROKEN")

    assert draft["model_guess"] is None
    assert "rating_guess" in draft["missing_fields"]
    assert draft["field_applicability"][0]["status"] == ("UNRESOLVED_TECHNICAL_DETAIL")
    assert draft["field_applicability"][0]["source"] == "unresolved"


@pytest.mark.parametrize(
    ("line", "expected_model"),
    [
        ("Регулятор температуры РТ 007S 1 шт", "РТ 007S"),
        ("Датчик температуры TST05 1 шт", "TST05"),
    ],
)
def test_exact_contextual_model_type_is_model_not_rating(
    line: str,
    expected_model: str,
) -> None:
    provenance = extraction.Provenance(
        source_file="shu-t2.txt",
        source_type="manual",
        locator="row=1",
        raw_text=line,
        confidence=0.9,
        reason="bounded synthetic ШУ-Т2 row",
    )

    component = extraction.component_from_text(line, provenance, 0.9)

    assert component is not None
    assert component.model == expected_model
    assert component.rating is None
    assert component.provenance[0].raw_text == line
    draft = extraction.component_to_draft(component, "COMP-SHU-T2")
    assert draft["model_guess"] == expected_model
    assert "rating_guess" not in draft["missing_fields"]
    assert draft["field_applicability"][0]["status"] == ("MODEL_OR_TYPE_SEMANTICS")
    assert draft["field_applicability"][0]["source"] == "raw_model_semantics"


@pytest.mark.parametrize(
    "line",
    [
        "Регулятор температуры РТ 007S РТ 007S 1 шт",
        "Датчик температуры TST0 1 шт",
    ],
)
def test_multiple_or_partial_model_candidates_fail_closed(line: str) -> None:
    provenance = extraction.Provenance(
        source_file="shu-t2.txt",
        source_type="manual",
        locator="row=1",
        raw_text=line,
        confidence=0.9,
        reason="bounded ambiguous ШУ-Т2 row",
    )

    component = extraction.component_from_text(line, provenance, 0.9)

    assert component is not None
    assert component.model not in {"РТ 007S", "TST05"}
    draft = extraction.component_to_draft(component, "COMP-SHU-T2-AMBIGUOUS")
    assert "rating_guess" in draft["missing_fields"]
    assert draft["field_applicability"][0]["status"] == ("UNRESOLVED_TECHNICAL_DETAIL")


def test_explicit_model_semantics_are_evaluated_per_provenance_row() -> None:
    component = synthetic_component(
        "Датчик температуры TST05",
        model="TST05",
        raw_text="Датчик температуры TST05 1 шт",
    )
    component.provenance.append(
        extraction.Provenance(
            source_file="second-source.txt",
            source_type="manual",
            locator="row=2",
            raw_text="Датчик температуры TST05 1 шт",
            confidence=0.9,
            reason="independent bounded synthetic row",
        )
    )

    draft = extraction.component_to_draft(component, "COMP-SHU-T2-MULTI-SOURCE")

    assert draft["model_guess"] == "TST05"
    assert len(draft["provenance"]) == 2
    assert "rating_guess" not in draft["missing_fields"]
    assert draft["field_applicability"][0]["status"] == "MODEL_OR_TYPE_SEMANTICS"


def test_tst05_quantity_conflict_is_not_corrected() -> None:
    first = synthetic_component(
        "Датчик температуры TST05",
        quantity=5,
        model="TST05",
        raw_text="Датчик температуры TST05 5 шт",
    )
    second = synthetic_component(
        "Датчик температуры TST05",
        quantity=1,
        model="TST05",
        raw_text="Датчик температуры TST05 1 шт",
    )
    board = extraction.BoardCandidate(
        designation="ШУ-Т2",
        normalized="ШУ-Т2",
        title="ШУ-Т2",
        quantity=1,
        provenance=first.provenance,
        confidence=0.9,
    )

    merged, _count = extraction.merge_components(board, [first, second], False)

    assert len(merged) == 1
    assert merged[0].quantity is None
    assert any(
        conflict["field"] == "quantity_guess" for conflict in merged[0].conflicts
    )


@pytest.mark.parametrize(
    "component_id",
    [
        "COMP-040",
        "COMP-137",
        "COMP-187",
        "COMP-034",
        "COMP-088",
        "COMP-131",
        "COMP-181",
    ],
)
def test_frozen_quantity_conflicts_remain_unresolved(component_id: str) -> None:
    component = synthetic_component(
        "Шина N" if component_id in {"COMP-040", "COMP-137", "COMP-187"} else "TST05",
        quantity=None,
        model=None if component_id.startswith("COMP-0") else "TST05",
        conflicts=[
            {
                "conflict_id": f"{component_id}-QTY",
                "type": "component_quantity_mismatch",
                "field": "quantity_guess",
                "message": "bounded frozen quantity conflict",
                "sources": [],
            }
        ],
    )

    draft = extraction.component_to_draft(component, component_id)

    assert draft["component_id"] == component_id
    assert draft["quantity_guess"] is None
    assert "quantity_guess" in draft["missing_fields"]
    assert draft["conflicts"][0]["field"] == "quantity_guess"


def test_bounded_policy_addresses_expected_rating_record_counts() -> None:
    components = [
        *[synthetic_component("Шина N") for _index in range(29)],
        *[
            synthetic_component("Счетчик электроэнергии", model=f"METER-{index}")
            for index in range(16)
        ],
        *[
            synthetic_component(
                "Датчик температуры TST05",
                model="TST05",
                raw_text="Датчик температуры TST05 1 шт",
            )
            for _index in range(8)
        ],
    ]

    drafts = [
        extraction.component_to_draft(component, f"COMP-{index:03d}")
        for index, component in enumerate(components, start=1)
    ]

    assert len(drafts) == 53
    assert (
        sum(
            value["field_applicability"][0]["status"] == "NOT_APPLICABLE_WITH_REASON"
            for value in drafts
        )
        == 29
    )
    assert (
        sum(
            value["field_applicability"][0]["status"] == "MODEL_OR_TYPE_SEMANTICS"
            for value in drafts
        )
        == 24
    )
    assert all("rating_guess" not in value["missing_fields"] for value in drafts)
