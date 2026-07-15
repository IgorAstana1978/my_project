import csv
import hashlib
import importlib.util
import json
import subprocess
import sys
import zipfile
from pathlib import Path
from types import ModuleType
from typing import Any, cast

import pytest
from openpyxl import Workbook, load_workbook  # type: ignore[import-untyped]
from openpyxl.styles import Alignment  # type: ignore[import-untyped]

PROJECT_ROOT = Path(__file__).resolve().parents[1]
SCRIPT = PROJECT_ROOT / "scripts" / "run_invoice_quote_commercial_from_csv.py"
OLD_FIVE_COLUMN_SCRIPT = (
    PROJECT_ROOT / "scripts" / "run_invoice_quote_extended_from_csv.py"
)
SYNTHETIC_LOGO_BYTES = b"synthetic-logo-bytes"
SYNTHETIC_LOGO_SHA256 = hashlib.sha256(SYNTHETIC_LOGO_BYTES).hexdigest()


def load_script_module(module_name: str, path: Path) -> ModuleType:
    spec = importlib.util.spec_from_file_location(module_name, path)
    assert spec is not None
    assert spec.loader is not None
    module = importlib.util.module_from_spec(spec)
    sys.modules[spec.name] = module
    spec.loader.exec_module(module)
    return module


writer = cast(
    Any,
    load_script_module("run_invoice_quote_commercial_from_csv_for_test", SCRIPT),
)


def commercial_row(
    index: int = 1,
    *,
    quantity: str | None = None,
    unit_price: str | None = None,
    vat_mode: str = "no",
    confirmation: str = "yes",
) -> list[str]:
    return [
        f"SYNTHETIC-ITEM-{index}",
        "шт.",
        quantity if quantity is not None else str(index),
        f"SYNTHETIC-DEVICES-{index}",
        f"SYNTHETIC-CABINET-{index}",
        unit_price if unit_price is not None else str(index * 1000),
        vat_mode,
        confirmation,
    ]


def write_commercial_csv(path: Path, rows: list[list[str]]) -> None:
    with path.open("w", encoding="utf-8", newline="") as csv_file:
        csv_writer = csv.writer(csv_file, delimiter=";", lineterminator="\n")
        csv_writer.writerow(writer.commercial_preflight.REQUIRED_COLUMNS)
        csv_writer.writerows(rows)


def write_capacity100_template(
    path: Path,
    *,
    wrong_item_formula: bool = False,
    with_drawing_parts: bool = True,
) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = writer.SHEET_NAME
    worksheet[writer.DOCUMENT_LINE_CELL] = (
        "Черновик счёта-КП № ____ от «__» ______ 2026 года"
    )
    worksheet[writer.PAYER_CELL] = "Плательщик: нужно уточнить"
    worksheet[writer.OBJECT_CELL] = "Объект: нужно уточнить"
    worksheet[writer.BASIS_PROJECT_CELL] = "Основание / проект: нужно уточнить"
    worksheet[writer.SECTION_CELL] = "Раздел / объект / позиция проекта: нужно уточнить"
    worksheet["B13"] = "Статус документа: Черновик"
    worksheet[writer.VALIDITY_CELL] = "Срок действия: нужно уточнить"
    worksheet[writer.PAYMENT_DELIVERY_CELL] = (
        "Условия оплаты: нужно уточнить. Условия поставки: нужно уточнить."
    )
    worksheet[writer.MANUFACTURING_CELL] = (
        "Ориентировочный срок изготовления: нужно уточнить."
    )
    worksheet[writer.VAT_LABEL_CELL] = writer.VAT_LABEL_FORMULA
    worksheet[writer.VAT_AMOUNT_CELL] = writer.VAT_AMOUNT_FORMULA
    worksheet[writer.VAT_AMOUNT_CELL].number_format = "#,##0.##"
    worksheet[writer.VAT_RATE_CELL] = writer.VAT_RATE_PLACEHOLDER
    worksheet[writer.VAT_RATE_CELL].number_format = "0"
    worksheet.row_dimensions[131].hidden = True
    worksheet.sheet_properties.pageSetUpPr.fitToPage = True
    worksheet.page_setup.paperSize = writer.PRINT_PAGE_SETUP["paperSize"]
    worksheet.page_setup.scale = int(writer.PRINT_PAGE_SETUP["scale"])
    worksheet.page_setup.fitToHeight = int(writer.PRINT_PAGE_SETUP["fitToHeight"])
    worksheet.page_setup.orientation = writer.PRINT_PAGE_SETUP["orientation"]
    for name, value in writer.PRINT_PAGE_MARGINS.items():
        setattr(worksheet.page_margins, name, value)

    for row in range(writer.ITEM_START_ROW, writer.ITEM_END_ROW + 1):
        worksheet[f"C{row}"] = "template item"
        worksheet[f"D{row}"] = "template unit"
        worksheet[f"E{row}"] = 1
        worksheet[f"F{row}"] = "template devices"
        worksheet[f"G{row}"] = "template cabinet"
        worksheet[f"H{row}"] = "template internal price placeholder"
        worksheet[f"I{row}"] = writer.commercial_reconciliation.item_formula(row)
        worksheet[f"J{row}"] = '=""'
        worksheet[f"J{row}"].alignment = Alignment(wrap_text=True)
    if wrong_item_formula:
        worksheet[f"I{writer.ITEM_START_ROW}"] = "=E17*H17"

    worksheet[f"I{writer.commercial_reconciliation.TOTAL_ROW}"] = (
        writer.commercial_reconciliation.total_formula()
    )
    worksheet[f"C{writer.AMOUNT_WORDS_ROW}"] = (
        "Всего прописью: template internal placeholder"
    )
    worksheet.column_dimensions["J"].width = 30.5703125
    worksheet.column_dimensions["K"].width = 11.5703125
    worksheet["K131"] = '=""'
    for merge_range in writer.commercial_reconciliation.EXPECTED_LOWER_MERGES:
        worksheet.merge_cells(merge_range)

    workbook.save(path)
    workbook.close()
    if with_drawing_parts:
        with zipfile.ZipFile(path) as archive:
            parts = {name: archive.read(name) for name in archive.namelist()}
        content_types = parts["[Content_Types].xml"]
        content_types = content_types.replace(
            b"</Types>",
            (
                b'<Default Extension="png" ContentType="image/png"/>'
                b'<Override PartName="/xl/drawings/drawing1.xml" '
                b'ContentType="application/vnd.openxmlformats-officedocument.'
                b'drawing+xml"/></Types>'
            ),
        )
        parts["[Content_Types].xml"] = content_types
        parts["xl/worksheets/sheet1.xml"] = parts["xl/worksheets/sheet1.xml"].replace(
            b"</worksheet>",
            (
                b'<drawing xmlns:r="http://schemas.openxmlformats.org/'
                b'officeDocument/2006/relationships" r:id="rId1"/>'
                b"</worksheet>"
            ),
        )
        parts["xl/worksheets/_rels/sheet1.xml.rels"] = (
            b'<?xml version="1.0" encoding="UTF-8"?>'
            b'<Relationships xmlns="http://schemas.openxmlformats.org/'
            b'package/2006/relationships">'
            b'<Relationship Id="rId1" Type="http://schemas.openxmlformats.org/'
            b'officeDocument/2006/relationships/drawing" '
            b'Target="../drawings/drawing1.xml"/></Relationships>'
        )
        parts["xl/drawings/drawing1.xml"] = (
            b'<?xml version="1.0" encoding="UTF-8"?>'
            b'<xdr:wsDr xmlns:xdr="http://schemas.openxmlformats.org/drawingml/'
            b'2006/spreadsheetDrawing" xmlns:a="http://schemas.openxmlformats.org/'
            b'drawingml/2006/main" xmlns:r="http://schemas.openxmlformats.org/'
            b'officeDocument/2006/relationships"><xdr:oneCellAnchor>'
            b"<xdr:from><xdr:col>0</xdr:col><xdr:colOff>76200</xdr:colOff>"
            b"<xdr:row>1</xdr:row><xdr:rowOff>66675</xdr:rowOff></xdr:from>"
            b'<xdr:ext cx="781050" cy="428625"/><xdr:pic><xdr:nvPicPr>'
            b'<xdr:cNvPr id="2" name="synthetic-logo"/><xdr:cNvPicPr/>'
            b'</xdr:nvPicPr><xdr:blipFill><a:blip r:embed="rId1"/>'
            b"</xdr:blipFill><xdr:spPr/></xdr:pic><xdr:clientData/>"
            b"</xdr:oneCellAnchor></xdr:wsDr>"
        )
        parts["xl/drawings/_rels/drawing1.xml.rels"] = (
            b'<?xml version="1.0" encoding="UTF-8"?>'
            b'<Relationships xmlns="http://schemas.openxmlformats.org/'
            b'package/2006/relationships"><Relationship Id="rId1" '
            b'Type="http://schemas.openxmlformats.org/officeDocument/2006/'
            b'relationships/image" Target="../media/image1.png"/>'
            b"</Relationships>"
        )
        parts["xl/media/image1.png"] = SYNTHETIC_LOGO_BYTES
        with zipfile.ZipFile(path, "w", zipfile.ZIP_DEFLATED) as archive:
            for name, content in parts.items():
                archive.writestr(name, content)


def output_path(tmp_path: Path) -> Path:
    output_directory = tmp_path / "out"
    output_directory.mkdir()
    return output_directory / "commercial-internal-draft.xlsx"


def writer_args(
    commercial_csv: Path,
    template: Path,
    output: Path,
    capacity: int = writer.CERTIFIED_CAPACITY,
    metadata: Path | None = None,
) -> list[str]:
    args = [
        "--commercial-csv",
        str(commercial_csv),
        "--template",
        str(template),
        "--template-capacity",
        str(capacity),
        "--output",
        str(output),
    ]
    if metadata is not None:
        args.extend(["--quote-metadata-json", str(metadata)])
    return args


def run_cli(
    commercial_csv: Path,
    template: Path,
    output: Path,
    capacity: int = writer.CERTIFIED_CAPACITY,
    metadata: Path | None = None,
) -> subprocess.CompletedProcess[str]:
    command = [sys.executable, str(SCRIPT)]
    if metadata is not None:
        module_name = "commercial_writer_synthetic_logo_contract"
        bootstrap = (
            "import importlib.util,sys;"
            f"spec=importlib.util.spec_from_file_location({module_name!r},{str(SCRIPT)!r});"
            "module=importlib.util.module_from_spec(spec);"
            "sys.modules[spec.name]=module;"
            "spec.loader.exec_module(module);"
            f"module.CERTIFIED_LOGO_SHA256={SYNTHETIC_LOGO_SHA256!r};"
            "raise SystemExit(module.main(sys.argv[1:]))"
        )
        command = [sys.executable, "-c", bootstrap]
    return subprocess.run(
        [
            *command,
            *writer_args(commercial_csv, template, output, capacity, metadata),
        ],
        capture_output=True,
        text=True,
        check=False,
    )


def candidate_paths(output: Path) -> list[Path]:
    return list(output.parent.glob(".*.candidate.xlsx"))


def rewrite_package_part(
    path: Path,
    part_name: str,
    replacement: bytes | None,
) -> None:
    with zipfile.ZipFile(path) as archive:
        parts = {name: archive.read(name) for name in archive.namelist()}
    if replacement is None:
        parts.pop(part_name)
    else:
        parts[part_name] = replacement
    with zipfile.ZipFile(path, "w", zipfile.ZIP_DEFLATED) as archive:
        for name, content in parts.items():
            archive.writestr(name, content)


def quote_metadata_payload() -> dict[str, object]:
    return {
        "schema_version": writer.QUOTE_METADATA_SCHEMA_VERSION,
        "document_number": "463",
        "document_date": "2026-07-10",
        "payer_name": "ТОО «Rich energy»",
        "payment_terms": "100% предоплата",
        "manufacturing_lead_time": "7–10 рабочих дней",
        "delivery_terms": "EXW, г. Астана",
        "vat_rate_percent": 16,
        "validity_period": None,
        "object_name": None,
        "basis_project": None,
        "item_notes": [],
    }


def write_quote_metadata(path: Path, payload: object) -> None:
    path.write_text(
        json.dumps(payload, ensure_ascii=False),
        encoding="utf-8",
    )


def test_valid_commercial_csv_creates_reconciled_internal_draft(
    tmp_path: Path,
) -> None:
    commercial_csv = tmp_path / "commercial.csv"
    template = tmp_path / "capacity100.xlsx"
    output = output_path(tmp_path)
    rows = [
        commercial_row(1, quantity="2", unit_price="100000"),
        commercial_row(2, quantity="4", unit_price="50000"),
    ]
    write_commercial_csv(commercial_csv, rows)
    write_capacity100_template(template)

    template_workbook = load_workbook(template, data_only=False)
    template_sheet = template_workbook[writer.SHEET_NAME]
    line_formulas_before = {
        f"I{row}": template_sheet[f"I{row}"].value
        for row in range(writer.ITEM_START_ROW, writer.ITEM_END_ROW + 1)
    }
    total_formula_before = template_sheet[
        f"I{writer.commercial_reconciliation.TOTAL_ROW}"
    ].value
    template_workbook.close()

    result = run_cli(commercial_csv, template, output)

    assert result.returncode == 0
    assert result.stderr == ""
    assert "Status:\nPASS" in result.stdout
    assert "Mode:\ninternal draft only" in result.stdout
    assert "commercial reconciliation: pass" in result.stdout
    assert "Manual Igor check:\nrequired" in result.stdout
    assert "Human Approval:\nseparate approval required" in result.stdout
    assert output.is_file()
    assert candidate_paths(output) == []

    workbook = load_workbook(output, data_only=False)
    worksheet = workbook[writer.SHEET_NAME]
    assert worksheet["C17"].value == rows[0][0]
    assert worksheet["D17"].value == rows[0][1]
    assert worksheet["E17"].value == int(rows[0][2])
    assert worksheet["F17"].value == rows[0][3]
    assert worksheet["G17"].value == rows[0][4]
    assert worksheet["H17"].value == int(rows[0][5])
    assert isinstance(worksheet["H17"].value, int)
    assert not isinstance(worksheet["H17"].value, bool)
    assert worksheet["H17"].number_format == writer.NUMBER_FORMAT_CODE
    assert worksheet["I17"].data_type == "f"
    assert worksheet["I17"].number_format == writer.NUMBER_FORMAT_CODE
    assert worksheet[f"I{writer.commercial_reconciliation.TOTAL_ROW}"].data_type == "f"
    assert (
        worksheet[f"I{writer.commercial_reconciliation.TOTAL_ROW}"].number_format
        == writer.NUMBER_FORMAT_CODE
    )
    assert all(
        worksheet[f"H{row}"].number_format == writer.NUMBER_FORMAT_CODE
        and worksheet[f"I{row}"].number_format == writer.NUMBER_FORMAT_CODE
        for row in range(writer.ITEM_START_ROW, writer.ITEM_END_ROW + 1)
    )
    assert {
        f"I{row}": worksheet[f"I{row}"].value
        for row in range(writer.ITEM_START_ROW, writer.ITEM_END_ROW + 1)
    } == line_formulas_before
    assert (
        worksheet[f"I{writer.commercial_reconciliation.TOTAL_ROW}"].value
        == total_formula_before
    )
    assert worksheet[f"C{writer.AMOUNT_WORDS_ROW}"].value == (
        "Всего прописью: Четыреста тысяч тенге 00 тиын"
    )
    assert worksheet[writer.DOCUMENT_LINE_CELL].value == (
        "Черновик счёта-КП № ____ от «__» ______ 2026 года"
    )
    assert worksheet[writer.PAYER_CELL].value == "Плательщик: нужно уточнить"
    assert worksheet[writer.SECTION_CELL].value == (
        "Раздел / объект / позиция проекта: нужно уточнить"
    )
    assert worksheet.row_dimensions[writer.SECTION_ROW].hidden is False
    assert worksheet[writer.VAT_RATE_CELL].value == writer.VAT_RATE_PLACEHOLDER
    assert worksheet[writer.VAT_LABEL_CELL].value == writer.VAT_LABEL_FORMULA
    assert worksheet[writer.VAT_AMOUNT_CELL].value == writer.VAT_AMOUNT_FORMULA
    assert worksheet["B13"].value == "Статус документа: Черновик"
    assert all(
        worksheet.cell(row=row, column=column).value not in {"yes", "no"}
        for row in range(1, worksheet.max_row + 1)
        for column in range(1, worksheet.max_column + 1)
    )
    workbook.close()


def test_quote_metadata_populates_certified_fields_and_preserves_template_vat_formula(
    tmp_path: Path,
) -> None:
    commercial_csv = tmp_path / "commercial.csv"
    template = tmp_path / "capacity100-v4.xlsx"
    metadata = tmp_path / "quote-metadata.json"
    output = output_path(tmp_path)
    write_commercial_csv(commercial_csv, [commercial_row(unit_price="209553")])
    write_capacity100_template(template)
    write_quote_metadata(metadata, quote_metadata_payload())
    template_hash = hashlib.sha256(template.read_bytes()).hexdigest()

    result = run_cli(commercial_csv, template, output, metadata=metadata)

    assert result.returncode == 0, result.stdout
    assert hashlib.sha256(template.read_bytes()).hexdigest() == template_hash
    workbook = load_workbook(output, data_only=False)
    worksheet = workbook[writer.SHEET_NAME]
    assert worksheet[writer.DOCUMENT_LINE_CELL].value == (
        "Черновик счёта-КП № 463 от «10» июля 2026 года"
    )
    assert worksheet[writer.PAYER_CELL].value == "Плательщик: ТОО «Rich energy»"
    assert worksheet[writer.OBJECT_CELL].value is None
    assert worksheet[writer.BASIS_PROJECT_CELL].value is None
    assert worksheet[writer.SECTION_CELL].value is None
    assert worksheet.row_dimensions[writer.SECTION_ROW].hidden is True
    assert worksheet[writer.VALIDITY_CELL].value is None
    assert worksheet[writer.PAYMENT_DELIVERY_CELL].value == (
        "Условия оплаты: 100% предоплата. " "Условия поставки: EXW, г. Астана."
    )
    assert worksheet[writer.MANUFACTURING_CELL].value == (
        "Ориентировочный срок изготовления: 7–10 рабочих дней."
    )
    assert worksheet[writer.VAT_RATE_CELL].value == 16
    assert isinstance(worksheet[writer.VAT_RATE_CELL].value, int)
    assert worksheet[writer.VAT_LABEL_CELL].value == writer.VAT_LABEL_FORMULA
    assert worksheet[writer.VAT_AMOUNT_CELL].value == writer.VAT_AMOUNT_FORMULA
    assert "I117" in worksheet[writer.VAT_AMOUNT_CELL].value
    assert worksheet[writer.VAT_AMOUNT_CELL].data_type == "f"
    assert worksheet[writer.VAT_AMOUNT_CELL].number_format == "#,##0.##"
    assert worksheet["H17"].value == 209553
    assert isinstance(worksheet["H17"].value, int)
    assert all(
        symbol not in worksheet[cell].number_format
        for cell in ("H17", "I17", "I117", writer.VAT_AMOUNT_CELL)
        for symbol in ("₽", "₸")
    )
    assert all(
        ".00" not in worksheet[cell].number_format
        for cell in ("H17", "I17", "I117", writer.VAT_AMOUNT_CELL)
    )
    assert worksheet["B13"].value == "Статус документа: Черновик"
    assert worksheet.row_dimensions[131].hidden is True
    workbook.close()


def test_quote_metadata_populates_object_basis_and_two_position_notes(
    tmp_path: Path,
) -> None:
    commercial_csv = tmp_path / "commercial.csv"
    template = tmp_path / "capacity100-v4.xlsx"
    metadata = tmp_path / "quote-metadata.json"
    output = output_path(tmp_path)
    rows = [commercial_row(index) for index in range(1, 4)]
    payload = quote_metadata_payload()
    payload.update(
        {
            "object_name": "Бизнес-центр Rich energy",
            "basis_project": "Проект RE-463",
            "item_notes": [
                {"item_number": 1, "text": "ВН 3Р 16А заменён на ВН 3Р 20А."},
                {"item_number": 2, "text": "ВН 3Р 25А заменён на ВН 3Р 32А."},
            ],
        }
    )
    write_commercial_csv(commercial_csv, rows)
    write_capacity100_template(template)
    write_quote_metadata(metadata, payload)

    result = run_cli(commercial_csv, template, output, metadata=metadata)

    assert result.returncode == 0, result.stdout
    workbook = load_workbook(output, data_only=False)
    worksheet = workbook[writer.SHEET_NAME]
    assert worksheet[writer.OBJECT_CELL].value == "Объект: Бизнес-центр Rich energy"
    assert worksheet[writer.BASIS_PROJECT_CELL].value == (
        "Основание / проект: Проект RE-463"
    )
    assert worksheet[writer.SECTION_CELL].value == (
        "Раздел / объект / позиция проекта: нужно уточнить"
    )
    assert worksheet.row_dimensions[writer.SECTION_ROW].hidden is False
    assert worksheet["J17"].value == "ВН 3Р 16А заменён на ВН 3Р 20А."
    assert worksheet["J18"].value == "ВН 3Р 25А заменён на ВН 3Р 32А."
    assert worksheet["J19"].value == '=""'
    assert worksheet["F17"].value == rows[0][3]
    assert worksheet["H17"].value == int(rows[0][5])
    assert worksheet["I17"].data_type == "f"
    assert worksheet.column_dimensions["J"].width == pytest.approx(30.5703125)
    assert worksheet.column_dimensions["K"].width == pytest.approx(11.5703125)
    assert worksheet.calculate_dimension().endswith("K131")
    assert worksheet.row_dimensions[17].height >= writer.BASE_ITEM_ROW_HEIGHT
    assert worksheet.row_dimensions[18].height >= writer.BASE_ITEM_ROW_HEIGHT
    workbook.close()


def test_native_page_setup_contract_is_preserved_in_metadata_output(
    tmp_path: Path,
) -> None:
    commercial_csv = tmp_path / "commercial.csv"
    template = tmp_path / "capacity100-v4.xlsx"
    metadata = tmp_path / "quote-metadata.json"
    output = output_path(tmp_path)
    write_commercial_csv(commercial_csv, [commercial_row()])
    write_capacity100_template(template)
    write_quote_metadata(metadata, quote_metadata_payload())

    result = run_cli(commercial_csv, template, output, metadata=metadata)

    assert result.returncode == 0, result.stdout
    workbook = load_workbook(output, data_only=False)
    worksheet = workbook[writer.SHEET_NAME]
    assert worksheet.print_area == ""
    assert worksheet.sheet_properties.pageSetUpPr.fitToPage is True
    assert worksheet.page_setup.paperSize == int(writer.PRINT_PAGE_SETUP["paperSize"])
    assert worksheet.page_setup.scale == int(writer.PRINT_PAGE_SETUP["scale"])
    assert worksheet.page_setup.fitToHeight == int(
        writer.PRINT_PAGE_SETUP["fitToHeight"]
    )
    assert worksheet.page_setup.orientation == writer.PRINT_PAGE_SETUP["orientation"]
    for name, value in writer.PRINT_PAGE_MARGINS.items():
        assert getattr(worksheet.page_margins, name) == pytest.approx(value)
    workbook.close()


@pytest.mark.parametrize(
    ("item_notes", "expected_failure"),
    [
        (
            [
                {"item_number": 1, "text": "first"},
                {"item_number": 1, "text": "duplicate"},
            ],
            "duplicated",
        ),
        ([{"item_number": 2, "text": "out of range"}], "out of range"),
        ([{"item_number": 1, "text": ""}], "non-empty string"),
        ([{"item_number": 1, "text": "note", "extra": True}], "unknown fields"),
    ],
)
def test_invalid_item_notes_fail_closed_without_output(
    tmp_path: Path,
    item_notes: list[dict[str, object]],
    expected_failure: str,
) -> None:
    commercial_csv = tmp_path / "commercial.csv"
    template = tmp_path / "capacity100-v4.xlsx"
    metadata = tmp_path / "quote-metadata.json"
    output = output_path(tmp_path)
    payload = quote_metadata_payload()
    payload["item_notes"] = item_notes
    write_commercial_csv(commercial_csv, [commercial_row()])
    write_capacity100_template(template)
    write_quote_metadata(metadata, payload)

    result = run_cli(commercial_csv, template, output, metadata=metadata)

    assert result.returncode == 1
    assert expected_failure in result.stdout
    assert not output.exists()
    assert candidate_paths(output) == []


def test_metadata_writer_preserves_logo_media_drawings_and_relationships(
    tmp_path: Path,
) -> None:
    commercial_csv = tmp_path / "commercial.csv"
    template = tmp_path / "capacity100-v4-with-logo.xlsx"
    metadata = tmp_path / "quote-metadata.json"
    output = output_path(tmp_path)
    write_commercial_csv(commercial_csv, [commercial_row()])
    write_capacity100_template(template, with_drawing_parts=True)
    write_quote_metadata(metadata, quote_metadata_payload())

    with zipfile.ZipFile(template) as archive:
        protected_parts = {
            name: archive.read(name)
            for name in archive.namelist()
            if name.startswith("xl/media/")
            or name.startswith("xl/drawings/")
            or name.endswith("sheet1.xml.rels")
        }
    result = run_cli(commercial_csv, template, output, metadata=metadata)
    with zipfile.ZipFile(output) as archive:
        output_parts = {name: archive.read(name) for name in protected_parts}

    assert result.returncode == 0, result.stdout
    assert protected_parts
    assert output_parts == protected_parts


def test_certified_logo_layout_contract_passes(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    template = tmp_path / "capacity100-v4.xlsx"
    write_capacity100_template(template)
    assert writer.CERTIFIED_LOGO_SHA256 == (
        "18e0f9446c72f8aa80ea833df07c2e42eb830770a0186decc476c5f948987301"
    )
    monkeypatch.setattr(writer, "CERTIFIED_LOGO_SHA256", SYNTHETIC_LOGO_SHA256)

    writer.validate_metadata_template_contract(template)

    assert writer.CERTIFIED_LOGO_PART == "xl/media/image1.png"
    assert writer.CERTIFIED_DRAWING_PART == "xl/drawings/drawing1.xml"
    assert writer.CERTIFIED_LOGO_EXTENT == {"cx": "781050", "cy": "428625"}


@pytest.mark.parametrize(
    ("part_name", "replacement", "expected_failure"),
    [
        (
            "xl/media/image1.png",
            None,
            "certified logo/drawing part is missing",
        ),
        (
            "xl/media/image1.png",
            b"modified-logo-bytes",
            "certified logo bytes are unexpected",
        ),
        (
            "xl/drawings/_rels/drawing1.xml.rels",
            (
                b'<?xml version="1.0" encoding="UTF-8"?>'
                b'<Relationships xmlns="http://schemas.openxmlformats.org/'
                b'package/2006/relationships"><Relationship Id="rId1" '
                b'Type="http://schemas.openxmlformats.org/officeDocument/2006/'
                b'relationships/image" Target="../media/broken.png"/>'
                b"</Relationships>"
            ),
            "drawing image relationship is broken",
        ),
    ],
)
def test_broken_certified_logo_contract_fails_before_output(
    tmp_path: Path,
    part_name: str,
    replacement: bytes | None,
    expected_failure: str,
) -> None:
    commercial_csv = tmp_path / "commercial.csv"
    template = tmp_path / "capacity100-v4.xlsx"
    metadata = tmp_path / "quote-metadata.json"
    output = output_path(tmp_path)
    write_commercial_csv(commercial_csv, [commercial_row()])
    write_capacity100_template(template)
    rewrite_package_part(template, part_name, replacement)
    write_quote_metadata(metadata, quote_metadata_payload())

    result = run_cli(commercial_csv, template, output, metadata=metadata)

    assert result.returncode == 1
    assert expected_failure in result.stdout
    assert "candidate generation: fail" in result.stdout
    assert not output.exists()
    assert candidate_paths(output) == []


@pytest.mark.parametrize(
    ("metadata_bytes", "expected_failure"),
    [
        (b"{", "quote metadata JSON is malformed"),
        (b'{"payer_name":"\xff"}', "strict UTF-8"),
        (
            json.dumps(
                {**quote_metadata_payload(), "unexpected": True},
                ensure_ascii=False,
            ).encode("utf-8"),
            "unknown fields",
        ),
        (
            json.dumps(
                {
                    key: value
                    for key, value in quote_metadata_payload().items()
                    if key != "payer_name"
                },
                ensure_ascii=False,
            ).encode("utf-8"),
            "missing required fields",
        ),
        (
            json.dumps(
                {**quote_metadata_payload(), "object_name": "   "},
                ensure_ascii=False,
            ).encode("utf-8"),
            "null or a non-empty string: object_name",
        ),
        (
            json.dumps(
                {**quote_metadata_payload(), "basis_project": "\t"},
                ensure_ascii=False,
            ).encode("utf-8"),
            "null or a non-empty string: basis_project",
        ),
    ],
)
def test_invalid_quote_metadata_fails_closed_without_output(
    tmp_path: Path,
    metadata_bytes: bytes,
    expected_failure: str,
) -> None:
    commercial_csv = tmp_path / "commercial.csv"
    template = tmp_path / "capacity100-v4.xlsx"
    metadata = tmp_path / "quote-metadata.json"
    output = output_path(tmp_path)
    write_commercial_csv(commercial_csv, [commercial_row()])
    write_capacity100_template(template)
    metadata.write_bytes(metadata_bytes)

    result = run_cli(commercial_csv, template, output, metadata=metadata)

    assert result.returncode == 1
    assert expected_failure in result.stdout
    assert not output.exists()
    assert candidate_paths(output) == []


def test_grand_total_words_use_independent_python_arithmetic() -> None:
    rows = [
        {"quantity": "2", "unit_price_kzt": "100000"},
        {"quantity": "4", "unit_price_kzt": "50000"},
    ]

    grand_total = writer.calculate_grand_total(rows)

    assert grand_total == 400000
    assert writer.amount_words_text(grand_total) == (
        "Всего прописью: Четыреста тысяч тенге 00 тиын"
    )


def test_invalid_commercial_csv_fails_without_output(tmp_path: Path) -> None:
    commercial_csv = tmp_path / "commercial.csv"
    template = tmp_path / "capacity100.xlsx"
    output = output_path(tmp_path)
    row = commercial_row(unit_price="1.5")
    write_commercial_csv(commercial_csv, [row])
    write_capacity100_template(template)

    result = run_cli(commercial_csv, template, output)

    assert result.returncode == 1
    assert "Status:\nFAIL" in result.stdout
    assert "commercial preflight: fail" in result.stdout
    assert "1.5" not in result.stdout
    assert not output.exists()
    assert candidate_paths(output) == []


def test_existing_output_fails_without_overwrite(tmp_path: Path) -> None:
    commercial_csv = tmp_path / "commercial.csv"
    template = tmp_path / "capacity100.xlsx"
    output = output_path(tmp_path)
    existing_content = b"existing-output"
    output.write_bytes(existing_content)
    write_commercial_csv(commercial_csv, [commercial_row()])
    write_capacity100_template(template)

    result = run_cli(commercial_csv, template, output)

    assert result.returncode == 1
    assert "output already exists" in result.stdout
    assert output.read_bytes() == existing_content
    assert candidate_paths(output) == []


def test_output_inside_git_fails_before_candidate_generation(
    tmp_path: Path,
    monkeypatch: Any,
) -> None:
    fake_project_root = tmp_path / "repo"
    fake_project_root.mkdir()
    commercial_csv = tmp_path / "commercial.csv"
    template = tmp_path / "capacity100.xlsx"
    output = fake_project_root / "blocked.xlsx"
    write_commercial_csv(commercial_csv, [commercial_row()])
    write_capacity100_template(template)
    monkeypatch.setattr(writer, "PROJECT_ROOT", fake_project_root)

    result = writer.run_commercial_writer(
        commercial_csv,
        template,
        writer.CERTIFIED_CAPACITY,
        output,
    )
    report = writer.format_report(result)

    assert result.status == "FAIL"
    assert "output is inside the Git project" in report
    assert not output.exists()
    assert candidate_paths(output) == []


def test_reconciliation_failure_prevents_output_and_cleans_candidate(
    tmp_path: Path,
) -> None:
    commercial_csv = tmp_path / "commercial.csv"
    template = tmp_path / "capacity100-wrong-formula.xlsx"
    output = output_path(tmp_path)
    row = commercial_row()
    write_commercial_csv(commercial_csv, [row])
    write_capacity100_template(template, wrong_item_formula=True)

    result = run_cli(commercial_csv, template, output)

    assert result.returncode == 1
    assert "Status:\nFAIL" in result.stdout
    assert "candidate generation: pass" in result.stdout
    assert "commercial reconciliation: fail" in result.stdout
    assert "reconciliation line formulas: fail" in result.stdout
    assert not output.exists()
    assert candidate_paths(output) == []


def test_reports_do_not_leak_commercial_values_or_full_rows(
    tmp_path: Path,
) -> None:
    commercial_csv = tmp_path / "commercial.csv"
    template = tmp_path / "capacity100.xlsx"
    output = output_path(tmp_path)
    secret_price = "987654321"
    row = commercial_row(unit_price=secret_price)
    write_commercial_csv(commercial_csv, [row])
    write_capacity100_template(template)

    result = run_cli(commercial_csv, template, output)

    assert result.returncode == 0
    assert secret_price not in result.stdout
    secret_total = int(row[2]) * int(secret_price)
    assert str(secret_total) not in result.stdout
    assert writer.integer_to_russian_words(secret_total) not in result.stdout
    assert ";".join(row) not in result.stdout
    assert row[0] not in result.stdout
    assert row[3] not in result.stdout


def test_preflight_failure_report_does_not_leak_commercial_values(
    tmp_path: Path,
) -> None:
    commercial_csv = tmp_path / "commercial.csv"
    template = tmp_path / "capacity100.xlsx"
    output = output_path(tmp_path)
    secret_price = "876543219"
    row = commercial_row(unit_price=secret_price, confirmation="no")
    write_commercial_csv(commercial_csv, [row])
    write_capacity100_template(template)

    result = run_cli(commercial_csv, template, output)

    assert result.returncode == 1
    assert secret_price not in result.stdout
    assert ";".join(row) not in result.stdout
    assert row[0] not in result.stdout
    assert not output.exists()


def test_only_certified_capacity100_is_accepted(tmp_path: Path) -> None:
    commercial_csv = tmp_path / "commercial.csv"
    template = tmp_path / "capacity100.xlsx"
    output = output_path(tmp_path)
    write_commercial_csv(commercial_csv, [commercial_row()])
    write_capacity100_template(template)

    result = run_cli(commercial_csv, template, output, capacity=99)

    assert result.returncode == 1
    assert "only the certified capacity100 profile is supported" in result.stdout
    assert not output.exists()
    assert candidate_paths(output) == []


def test_old_five_column_workflow_contract_is_unchanged() -> None:
    old_workflow = cast(
        Any,
        load_script_module(
            "run_invoice_quote_extended_from_csv_commercial_writer_contract_test",
            OLD_FIVE_COLUMN_SCRIPT,
        ),
    )

    assert old_workflow.REQUIRED_COLUMNS == (
        "name",
        "unit",
        "quantity",
        "instruments_and_devices",
        "cabinet_type_dimensions_material",
    )
    assert "unit_price_kzt" not in old_workflow.REQUIRED_COLUMNS
