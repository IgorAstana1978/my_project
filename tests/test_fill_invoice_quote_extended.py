import hashlib
import importlib.util
import inspect
import json
import subprocess
import sys
from pathlib import Path
from types import ModuleType
from typing import Any, cast
from zipfile import ZIP_DEFLATED, ZipFile

from openpyxl import Workbook, load_workbook  # type: ignore[import-untyped]

PROJECT_ROOT = Path(__file__).resolve().parents[1]
SCRIPT = PROJECT_ROOT / "scripts" / "fill_invoice_quote_extended.py"
DRAFT_SCRIPT = PROJECT_ROOT / "scripts" / "fill_invoice_quote_draft.py"
SNAPSHOT_SCRIPT = PROJECT_ROOT / "scripts" / "drawing_media_snapshot.py"
TINY_PNG = (
    b"\x89PNG\r\n\x1a\n\x00\x00\x00\rIHDR\x00\x00\x00\x01"
    b"\x00\x00\x00\x01\x08\x06\x00\x00\x00\x1f\x15\xc4\x89"
    b"\x00\x00\x00\rIDATx\x9cc\xf8\xff\xff?\x00\x05\xfe"
    b"\x02\xfeA\xe2d\x9a\x00\x00\x00\x00IEND\xaeB`\x82"
)


def load_script_module(module_name: str, path: Path) -> ModuleType:
    spec = importlib.util.spec_from_file_location(module_name, path)
    assert spec is not None
    assert spec.loader is not None
    module = importlib.util.module_from_spec(spec)
    sys.modules[spec.name] = module
    spec.loader.exec_module(module)
    return module


extended = cast(Any, load_script_module("fill_invoice_quote_extended_for_test", SCRIPT))
snapshot = cast(
    Any,
    load_script_module("drawing_media_snapshot_for_extended_test", SNAPSHOT_SCRIPT),
)


def file_sha256(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()


def extended_layout(capacity: int = 8) -> Any:
    return extended.ExtendedLayout(
        item_start_row=17,
        item_end_row=17 + capacity - 1,
        capacity=capacity,
        total_row=17 + capacity,
        signature_range=f"B{20 + capacity}:I{22 + capacity}",
        header_ranges=("C2:I6", "B4:B6"),
        formula_cells=tuple(
            [f"I{row}" for row in range(17, 17 + capacity)] + [f"I{17 + capacity}"]
        ),
    )


def write_extended_template(path: Path, capacity: int = 8) -> Any:
    layout = extended_layout(capacity)
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = extended.SHEET_NAME

    for row in range(2, 7):
        for column in range(3, 10):
            worksheet.cell(row=row, column=column).value = f"header-{row}-{column}"
    for row in range(4, 7):
        worksheet.cell(row=row, column=2).value = f"req-{row}"

    worksheet.merge_cells(layout.signature_range)
    worksheet.cell(row=20 + capacity, column=2).value = "signature"

    for row in range(layout.item_start_row, layout.item_end_row + 1):
        worksheet[f"C{row}"] = "template item"
        worksheet[f"D{row}"] = "шт"
        worksheet[f"E{row}"] = 1
        worksheet[f"F{row}"] = "template instruments"
        worksheet[f"G{row}"] = "template cabinet"
        worksheet[f"H{row}"] = "template price"
        worksheet[f"I{row}"] = f"=E{row}*H{row}"
    worksheet[f"I{layout.total_row}"] = (
        f"=SUM(I{layout.item_start_row}:I{layout.item_end_row})"
    )

    workbook.save(path)
    return layout


def rewrite_xlsx(path: Path, updates: dict[str, bytes]) -> None:
    temporary_path = path.with_suffix(".tmp.xlsx")
    with ZipFile(path) as source_archive:
        source_entries = {
            name: source_archive.read(name)
            for name in source_archive.namelist()
            if name not in updates
        }
    with ZipFile(temporary_path, "w", ZIP_DEFLATED) as target_archive:
        for name, content in source_entries.items():
            target_archive.writestr(name, content)
        for name, content in updates.items():
            target_archive.writestr(name, content)
    temporary_path.replace(path)


def content_types_with_drawing(content: str) -> str:
    if 'Extension="png"' not in content:
        content = content.replace(
            "</Types>",
            '<Default Extension="png" ContentType="image/png"/></Types>',
        )
    if 'PartName="/xl/drawings/drawing1.xml"' not in content:
        content = content.replace(
            "</Types>",
            (
                '<Override PartName="/xl/drawings/drawing1.xml" '
                'ContentType="application/vnd.openxmlformats-officedocument.'
                'drawing+xml"/></Types>'
            ),
        )
    return content


def add_valid_drawing_chain(path: Path) -> None:
    with ZipFile(path) as archive:
        sheet_xml = archive.read("xl/worksheets/sheet1.xml").decode("utf-8")
        content_types = archive.read("[Content_Types].xml").decode("utf-8")

    sheet_xml = sheet_xml.replace(
        "</worksheet>",
        (
            '<drawing xmlns:r="http://schemas.openxmlformats.org/'
            'officeDocument/2006/relationships" r:id="rId1"/>'
            "</worksheet>"
        ),
    )
    sheet_rels = (
        '<?xml version="1.0" encoding="UTF-8"?>'
        '<Relationships xmlns="http://schemas.openxmlformats.org/package/'
        '2006/relationships">'
        '<Relationship Id="rId1" Type="http://schemas.openxmlformats.org/'
        'officeDocument/2006/relationships/drawing" '
        'Target="../drawings/drawing1.xml"/>'
        "</Relationships>"
    )
    drawing_xml = (
        '<?xml version="1.0" encoding="UTF-8"?>'
        '<xdr:wsDr xmlns:xdr="http://schemas.openxmlformats.org/drawingml/'
        '2006/spreadsheetDrawing" '
        'xmlns:a="http://schemas.openxmlformats.org/drawingml/2006/main">'
        "<xdr:oneCellAnchor><xdr:from><xdr:col>0</xdr:col>"
        "<xdr:colOff>0</xdr:colOff><xdr:row>0</xdr:row>"
        "<xdr:rowOff>0</xdr:rowOff></xdr:from>"
        '<xdr:ext cx="1" cy="1"/>'
        '<xdr:pic><xdr:nvPicPr><xdr:cNvPr id="2" name="logo"/>'
        "<xdr:cNvPicPr/></xdr:nvPicPr><xdr:blipFill>"
        '<a:blip xmlns:r="http://schemas.openxmlformats.org/'
        'officeDocument/2006/relationships" r:embed="rId1"/>'
        "<a:stretch><a:fillRect/></a:stretch></xdr:blipFill>"
        "<xdr:spPr/></xdr:pic><xdr:clientData/></xdr:oneCellAnchor>"
        "</xdr:wsDr>"
    )
    drawing_rels = (
        '<?xml version="1.0" encoding="UTF-8"?>'
        '<Relationships xmlns="http://schemas.openxmlformats.org/package/'
        '2006/relationships">'
        '<Relationship Id="rId1" Type="http://schemas.openxmlformats.org/'
        'officeDocument/2006/relationships/image" '
        'Target="../media/image1.png"/>'
        "</Relationships>"
    )

    rewrite_xlsx(
        path,
        {
            "xl/worksheets/sheet1.xml": sheet_xml.encode("utf-8"),
            "xl/worksheets/_rels/sheet1.xml.rels": sheet_rels.encode("utf-8"),
            "xl/drawings/drawing1.xml": drawing_xml.encode("utf-8"),
            "xl/drawings/_rels/drawing1.xml.rels": drawing_rels.encode("utf-8"),
            "xl/media/image1.png": TINY_PNG,
            "[Content_Types].xml": content_types_with_drawing(content_types).encode(
                "utf-8"
            ),
        },
    )


def item(index: int) -> dict[str, Any]:
    return {
        "name": f"ВРУ-{index}",
        "unit": "шт.",
        "quantity": index,
        "instruments_and_devices": "нужно уточнить",
        "cabinet_type_dimensions_material": "нужно уточнить",
        "price_kzt": None,
        "price_confirmed_by_igor": False,
    }


def payload(items_count: int) -> dict[str, Any]:
    return {"items": [item(index) for index in range(1, items_count + 1)]}


def layout_json(layout: Any) -> dict[str, Any]:
    return {
        "item_start_row": layout.item_start_row,
        "item_end_row": layout.item_end_row,
        "capacity": layout.capacity,
        "total_row": layout.total_row,
        "signature_range": layout.signature_range,
        "header_ranges": list(layout.header_ranges),
        "formula_cells": list(layout.formula_cells),
    }


def write_json(path: Path, data: dict[str, Any]) -> None:
    path.write_text(json.dumps(data, ensure_ascii=False), encoding="utf-8")


def cli_args(
    payload_json: Path,
    layout_json_path: Path,
    template: Path,
    output: Path,
) -> list[str]:
    return [
        "--payload-json",
        str(payload_json),
        "--layout-json",
        str(layout_json_path),
        "--template",
        str(template),
        "--output",
        str(output),
    ]


def workbook_values(path: Path) -> dict[str, Any]:
    workbook = load_workbook(path, data_only=False)
    worksheet = workbook[extended.SHEET_NAME]
    return {
        "C17": worksheet["C17"].value,
        "C22": worksheet["C22"].value,
        "I17": worksheet["I17"].value,
        "I25": worksheet["I25"].value,
        "B28": worksheet["B28"].value,
        "C2": worksheet["C2"].value,
    }


def merged_ranges(path: Path) -> tuple[str, ...]:
    workbook = load_workbook(path, data_only=False)
    worksheet = workbook[extended.SHEET_NAME]
    return tuple(str(item) for item in worksheet.merged_cells.ranges)


def test_extended_template_is_created_in_tmp_path(tmp_path: Path) -> None:
    template = tmp_path / "extended_template.xlsx"

    layout = write_extended_template(template)

    assert template.is_file()
    assert not template.is_relative_to(PROJECT_ROOT)
    assert layout.capacity == 8


def test_writes_six_items_to_extended_template(tmp_path: Path) -> None:
    template = tmp_path / "extended_template.xlsx"
    output_dir = tmp_path / "out"
    output_dir.mkdir()
    output = output_dir / "draft.xlsx"
    layout = write_extended_template(template)

    result = extended.generate_extended_workbook(
        template=template,
        output=output,
        payload=payload(6),
        layout=layout,
    )

    assert result == output
    assert output.is_file()
    values = workbook_values(output)
    assert values["C17"] == "ВРУ-1"
    assert values["C22"] == "ВРУ-6"
    assert values["I17"] == "=E17*H17"
    assert values["I25"] == "=SUM(I17:I24)"
    assert values["B28"] == "signature"
    assert values["C2"] == "header-2-3"


def test_merged_ranges_are_preserved_after_generation(tmp_path: Path) -> None:
    template = tmp_path / "extended_template.xlsx"
    output_dir = tmp_path / "out"
    output_dir.mkdir()
    output = output_dir / "draft.xlsx"
    layout = write_extended_template(template)
    before = merged_ranges(template)

    extended.generate_extended_workbook(
        template=template,
        output=output,
        payload=payload(6),
        layout=layout,
    )

    assert before == (layout.signature_range,)
    assert merged_ranges(output) == before


def test_generated_template_without_drawing_media_creates_output(
    tmp_path: Path,
) -> None:
    template = tmp_path / "extended_template.xlsx"
    output_dir = tmp_path / "out"
    output_dir.mkdir()
    output = output_dir / "draft.xlsx"
    layout = write_extended_template(template)

    extended.generate_extended_workbook(
        template=template,
        output=output,
        payload=payload(6),
        layout=layout,
    )

    assert output.is_file()
    assert workbook_values(output)["I25"] == "=SUM(I17:I24)"
    assert merged_ranges(output) == (layout.signature_range,)


def test_real_drawing_media_round_trip_preserves_template_parts(
    tmp_path: Path,
) -> None:
    template = tmp_path / "extended_template.xlsx"
    output_dir = tmp_path / "out"
    output_dir.mkdir()
    output = output_dir / "draft.xlsx"
    layout = write_extended_template(template)
    add_valid_drawing_chain(template)
    before_snapshot = snapshot.build_drawing_media_snapshot(template)

    extended.generate_extended_workbook(
        template=template,
        output=output,
        payload=payload(6),
        layout=layout,
    )

    assert output.is_file()
    after_snapshot = snapshot.build_drawing_media_snapshot(output)
    snapshot.compare_drawing_media_snapshots(before_snapshot, after_snapshot)
    values = workbook_values(output)
    assert values["C17"] == "ВРУ-1"
    assert values["C22"] == "ВРУ-6"
    assert values["I17"] == "=E17*H17"
    assert values["I25"] == "=SUM(I17:I24)"
    assert values["B28"] == "signature"
    assert values["C2"] == "header-2-3"
    assert merged_ranges(output) == (layout.signature_range,)


def test_build_cell_updates_maps_items_and_unused_rows() -> None:
    layout = extended_layout()
    custom_items = payload(6)["items"]
    custom_items[0]["name"] = "0012"
    custom_items[0]["instruments_and_devices"] = "Автомат\nQF1"
    custom_items[5]["cabinet_type_dimensions_material"] = "Шкаф ВРУ"

    updates = extended.build_cell_updates(custom_items, layout)

    assert updates["C17"] == "0012"
    assert updates["D17"] == "шт."
    assert updates["E17"] == 1
    assert isinstance(updates["E17"], int)
    assert updates["F17"] == "Автомат\nQF1"
    assert updates["G17"] == "нужно уточнить"
    assert updates["H17"] == "нужно уточнить"
    assert updates["C22"] == "ВРУ-6"
    assert updates["G22"] == "Шкаф ВРУ"
    assert updates["H22"] == "нужно уточнить"
    assert updates["C23"] == "нужно уточнить"
    assert updates["D23"] == "шт"
    assert updates["E23"] == 1
    assert updates["F23"] == "нужно уточнить"
    assert updates["G23"] == "нужно уточнить"
    assert updates["H23"] == "нужно уточнить"
    assert updates["C24"] == "нужно уточнить"
    assert updates["D24"] == "шт"
    assert updates["E24"] == 1


def test_generate_extended_workbook_uses_ooxml_patcher(
    tmp_path: Path,
    monkeypatch: Any,
) -> None:
    template = tmp_path / "extended_template.xlsx"
    output_dir = tmp_path / "out"
    output_dir.mkdir()
    output = output_dir / "draft.xlsx"
    layout = write_extended_template(template)
    original_patch = extended.patch_existing_cells
    calls: list[dict[str, Any]] = []
    verification_seen: list[Path] = []

    def tracking_patch_existing_cells(**kwargs: Any) -> Path:
        calls.append(kwargs)
        assert kwargs["template"] == template
        assert kwargs["sheet_name"] == extended.SHEET_NAME
        temporary_output = kwargs["output"]
        assert temporary_output != output
        assert temporary_output.parent == output.parent
        assert temporary_output.name.startswith(f".{output.stem}.")
        assert temporary_output.name.endswith(".tmp.xlsx")
        assert not output.exists()
        updates = kwargs["updates"]
        assert updates["C17"] == "ВРУ-1"
        assert updates["E17"] == 1
        assert isinstance(updates["E17"], int)
        assert updates["C22"] == "ВРУ-6"
        assert updates["H22"] == "нужно уточнить"
        return cast(Path, original_patch(**kwargs))

    original_verify_output = extended.verify_output

    def tracking_verify_output(
        temporary_output: Path,
        verify_layout: Any,
        before: Any,
    ) -> None:
        assert temporary_output.exists()
        assert not output.exists()
        verification_seen.append(temporary_output)
        original_verify_output(temporary_output, verify_layout, before)

    monkeypatch.setattr(extended, "patch_existing_cells", tracking_patch_existing_cells)
    monkeypatch.setattr(extended, "verify_output", tracking_verify_output)

    result = extended.generate_extended_workbook(
        template=template,
        output=output,
        payload=payload(6),
        layout=layout,
    )

    assert result == output
    assert output.is_file()
    assert len(calls) == 1
    assert len(verification_seen) == 1
    assert not verification_seen[0].exists()


def test_generate_extended_workbook_has_no_openpyxl_save_call() -> None:
    source = inspect.getsource(extended.generate_extended_workbook)

    assert ".save(" not in source


def test_drawing_media_snapshot_uses_template_and_temporary_output(
    tmp_path: Path,
    monkeypatch: Any,
) -> None:
    template = tmp_path / "extended_template.xlsx"
    output_dir = tmp_path / "out"
    output_dir.mkdir()
    output = output_dir / "draft.xlsx"
    layout = write_extended_template(template)
    snapshot_paths: list[Path] = []
    compared: list[tuple[Any, Any]] = []

    def fake_build_drawing_media_snapshot(path: Path) -> str:
        snapshot_paths.append(path)
        return f"snapshot:{path.name}"

    def fake_compare_drawing_media_snapshots(before: Any, after: Any) -> None:
        compared.append((before, after))
        temporary_output = snapshot_paths[-1]
        assert before == f"snapshot:{template.name}"
        assert after == f"snapshot:{temporary_output.name}"
        assert temporary_output != output
        assert temporary_output.parent == output.parent
        assert temporary_output.name.startswith(f".{output.stem}.")
        assert temporary_output.name.endswith(".tmp.xlsx")
        assert temporary_output.exists()
        assert not output.exists()

    monkeypatch.setattr(
        extended,
        "build_drawing_media_snapshot",
        fake_build_drawing_media_snapshot,
    )
    monkeypatch.setattr(
        extended,
        "compare_drawing_media_snapshots",
        fake_compare_drawing_media_snapshots,
    )

    result = extended.generate_extended_workbook(
        template=template,
        output=output,
        payload=payload(6),
        layout=layout,
    )

    assert result == output
    assert output.is_file()
    assert snapshot_paths[0] == template
    assert len(snapshot_paths) == 2
    assert snapshot_paths[1] != output
    assert compared == [
        (f"snapshot:{template.name}", f"snapshot:{snapshot_paths[1].name}")
    ]


def test_template_drawing_media_snapshot_error_fails_before_temp_output(
    tmp_path: Path,
    monkeypatch: Any,
) -> None:
    template = tmp_path / "extended_template.xlsx"
    output_dir = tmp_path / "out"
    output_dir.mkdir()
    output = output_dir / "draft.xlsx"
    layout = write_extended_template(template)

    def fail_template_snapshot(_path: Path) -> Any:
        raise extended.DrawingMediaSnapshotError("template media missing")

    monkeypatch.setattr(
        extended,
        "build_drawing_media_snapshot",
        fail_template_snapshot,
    )

    try:
        extended.generate_extended_workbook(
            template=template,
            output=output,
            payload=payload(6),
            layout=layout,
        )
    except extended.ExtendedFillError as error:
        assert "drawing/media verification failed" in str(error)
        assert "template media missing" in str(error)
    else:
        raise AssertionError("template snapshot error should fail")

    assert not output.exists()
    assert list(output_dir.iterdir()) == []


def test_ooxml_patcher_error_becomes_extended_error_and_cleans_outputs(
    tmp_path: Path,
    monkeypatch: Any,
) -> None:
    template = tmp_path / "extended_template.xlsx"
    output_dir = tmp_path / "out"
    output_dir.mkdir()
    output = output_dir / "draft.xlsx"
    layout = write_extended_template(template)

    def fail_patch_existing_cells(**_kwargs: Any) -> Path:
        raise extended.OoxmlCellPatcherError("cell does not exist: C17")

    monkeypatch.setattr(extended, "patch_existing_cells", fail_patch_existing_cells)

    try:
        extended.generate_extended_workbook(
            template=template,
            output=output,
            payload=payload(6),
            layout=layout,
        )
    except extended.ExtendedFillError as error:
        assert "OOXML patching failed: cell does not exist: C17" in str(error)
    else:
        raise AssertionError("OOXML patching error should fail")

    assert not output.exists()
    assert list(output_dir.iterdir()) == []


def test_cli_returns_one_without_traceback_for_ooxml_patching_error(
    tmp_path: Path,
    monkeypatch: Any,
    capsys: Any,
) -> None:
    template = tmp_path / "extended_template.xlsx"
    payload_json = tmp_path / "payload.json"
    layout_json_path = tmp_path / "layout.json"
    output_dir = tmp_path / "out"
    output_dir.mkdir()
    output = output_dir / "draft.xlsx"
    layout = write_extended_template(template)
    write_json(payload_json, payload(6))
    write_json(layout_json_path, layout_json(layout))

    def fail_patch_existing_cells(**_kwargs: Any) -> Path:
        raise extended.OoxmlCellPatcherError("cell does not exist: C17")

    monkeypatch.setattr(extended, "patch_existing_cells", fail_patch_existing_cells)

    exit_code = extended.main(
        cli_args(payload_json, layout_json_path, template, output)
    )

    assert exit_code == 1
    captured = capsys.readouterr()
    assert "ERROR:" in captured.err
    assert "OOXML patching failed: cell does not exist: C17" in captured.err
    assert "Traceback" not in captured.err
    assert not output.exists()
    assert list(output_dir.iterdir()) == []


def test_temporary_output_drawing_media_snapshot_error_cleans_temp_output(
    tmp_path: Path,
    monkeypatch: Any,
) -> None:
    template = tmp_path / "extended_template.xlsx"
    output_dir = tmp_path / "out"
    output_dir.mkdir()
    output = output_dir / "draft.xlsx"
    layout = write_extended_template(template)

    def fail_temporary_output_snapshot(path: Path) -> object:
        if path == template:
            return object()
        raise extended.DrawingMediaSnapshotError("temporary output unreadable")

    monkeypatch.setattr(
        extended,
        "build_drawing_media_snapshot",
        fail_temporary_output_snapshot,
    )

    try:
        extended.generate_extended_workbook(
            template=template,
            output=output,
            payload=payload(6),
            layout=layout,
        )
    except extended.ExtendedFillError as error:
        assert "drawing/media verification failed" in str(error)
        assert "temporary output unreadable" in str(error)
    else:
        raise AssertionError("temporary output snapshot error should fail")

    assert not output.exists()
    assert list(output_dir.iterdir()) == []


def test_drawing_media_compare_error_cleans_temp_output(
    tmp_path: Path,
    monkeypatch: Any,
) -> None:
    template = tmp_path / "extended_template.xlsx"
    output_dir = tmp_path / "out"
    output_dir.mkdir()
    output = output_dir / "draft.xlsx"
    layout = write_extended_template(template)

    def fake_build_drawing_media_snapshot(_path: Path) -> object:
        return object()

    def fail_drawing_media_compare(_before: Any, _after: Any) -> None:
        raise extended.DrawingMediaSnapshotError("media file hash changed")

    monkeypatch.setattr(
        extended,
        "build_drawing_media_snapshot",
        fake_build_drawing_media_snapshot,
    )
    monkeypatch.setattr(
        extended,
        "compare_drawing_media_snapshots",
        fail_drawing_media_compare,
    )

    try:
        extended.generate_extended_workbook(
            template=template,
            output=output,
            payload=payload(6),
            layout=layout,
        )
    except extended.ExtendedFillError as error:
        assert "drawing/media verification failed" in str(error)
        assert "media file hash changed" in str(error)
    else:
        raise AssertionError("drawing/media compare error should fail")

    assert not output.exists()
    assert list(output_dir.iterdir()) == []


def test_cli_returns_one_without_traceback_for_drawing_media_error(
    tmp_path: Path,
) -> None:
    template = tmp_path / "invalid_template.xlsx"
    payload_json = tmp_path / "payload.json"
    layout_json_path = tmp_path / "layout.json"
    output_dir = tmp_path / "out"
    output_dir.mkdir()
    output = output_dir / "draft.xlsx"
    layout = extended_layout()
    template.write_bytes(b"not a zip")
    write_json(payload_json, payload(6))
    write_json(layout_json_path, layout_json(layout))

    result = subprocess.run(
        [sys.executable, str(SCRIPT)]
        + cli_args(payload_json, layout_json_path, template, output),
        capture_output=True,
        text=True,
        check=False,
    )

    assert result.returncode == 1
    assert "ERROR:" in result.stderr
    assert "drawing/media verification failed" in result.stderr
    assert "invalid xlsx ZIP package" in result.stderr
    assert "Traceback" not in result.stderr
    assert not output.exists()
    assert list(output_dir.iterdir()) == []


def test_merged_range_change_fails_closed_and_removes_temp_output(
    tmp_path: Path,
    monkeypatch: Any,
) -> None:
    template = tmp_path / "extended_template.xlsx"
    output_dir = tmp_path / "out"
    output_dir.mkdir()
    output = output_dir / "draft.xlsx"
    layout = write_extended_template(template)

    def fail_merged_range_verification(
        _output: Path,
        _layout: Any,
        _before: Any,
    ) -> None:
        raise extended.ExtendedFillError("merged ranges changed")

    monkeypatch.setattr(extended, "verify_output", fail_merged_range_verification)

    try:
        extended.generate_extended_workbook(
            template=template,
            output=output,
            payload=payload(6),
            layout=layout,
        )
    except extended.ExtendedFillError as error:
        assert "merged ranges changed" in str(error)
    else:
        raise AssertionError("merged range change should fail")

    assert not output.exists()
    assert list(output_dir.iterdir()) == []


def test_cli_successfully_creates_output_for_six_items(tmp_path: Path) -> None:
    template = tmp_path / "extended_template.xlsx"
    payload_json = tmp_path / "payload.json"
    layout_json_path = tmp_path / "layout.json"
    output_dir = tmp_path / "out"
    output_dir.mkdir()
    output = output_dir / "draft.xlsx"
    layout = write_extended_template(template)
    write_json(payload_json, payload(6))
    write_json(layout_json_path, layout_json(layout))

    exit_code = extended.main(
        cli_args(payload_json, layout_json_path, template, output)
    )

    assert exit_code == 0
    assert output.is_file()
    assert workbook_values(output)["C22"] == "ВРУ-6"


def test_cli_capacity_overflow_returns_one_without_output(tmp_path: Path) -> None:
    template = tmp_path / "extended_template.xlsx"
    payload_json = tmp_path / "payload.json"
    layout_json_path = tmp_path / "layout.json"
    output_dir = tmp_path / "out"
    output_dir.mkdir()
    output = output_dir / "draft.xlsx"
    layout = write_extended_template(template, capacity=6)
    write_json(payload_json, payload(7))
    write_json(layout_json_path, layout_json(layout))

    exit_code = extended.main(
        cli_args(payload_json, layout_json_path, template, output)
    )

    assert exit_code == 1
    assert not output.exists()
    assert list(output_dir.iterdir()) == []


def test_cli_existing_output_returns_one_without_overwrite(tmp_path: Path) -> None:
    template = tmp_path / "extended_template.xlsx"
    payload_json = tmp_path / "payload.json"
    layout_json_path = tmp_path / "layout.json"
    output_dir = tmp_path / "out"
    output_dir.mkdir()
    output = output_dir / "draft.xlsx"
    output.write_bytes(b"existing")
    layout = write_extended_template(template)
    write_json(payload_json, payload(6))
    write_json(layout_json_path, layout_json(layout))

    exit_code = extended.main(
        cli_args(payload_json, layout_json_path, template, output)
    )

    assert exit_code == 1
    assert output.read_bytes() == b"existing"
    assert list(output_dir.iterdir()) == [output]


def test_cli_missing_payload_json_returns_one(tmp_path: Path) -> None:
    template = tmp_path / "extended_template.xlsx"
    payload_json = tmp_path / "missing_payload.json"
    layout_json_path = tmp_path / "layout.json"
    output_dir = tmp_path / "out"
    output_dir.mkdir()
    output = output_dir / "draft.xlsx"
    layout = write_extended_template(template)
    write_json(layout_json_path, layout_json(layout))

    exit_code = extended.main(
        cli_args(payload_json, layout_json_path, template, output)
    )

    assert exit_code == 1
    assert not output.exists()
    assert list(output_dir.iterdir()) == []


def test_cli_invalid_payload_json_returns_one(tmp_path: Path) -> None:
    template = tmp_path / "extended_template.xlsx"
    payload_json = tmp_path / "payload.json"
    layout_json_path = tmp_path / "layout.json"
    output_dir = tmp_path / "out"
    output_dir.mkdir()
    output = output_dir / "draft.xlsx"
    layout = write_extended_template(template)
    payload_json.write_text("{invalid", encoding="utf-8")
    write_json(layout_json_path, layout_json(layout))

    exit_code = extended.main(
        cli_args(payload_json, layout_json_path, template, output)
    )

    assert exit_code == 1
    assert not output.exists()
    assert list(output_dir.iterdir()) == []


def test_cli_missing_layout_json_returns_one(tmp_path: Path) -> None:
    template = tmp_path / "extended_template.xlsx"
    payload_json = tmp_path / "payload.json"
    layout_json_path = tmp_path / "missing_layout.json"
    output_dir = tmp_path / "out"
    output_dir.mkdir()
    output = output_dir / "draft.xlsx"
    write_extended_template(template)
    write_json(payload_json, payload(6))

    exit_code = extended.main(
        cli_args(payload_json, layout_json_path, template, output)
    )

    assert exit_code == 1
    assert not output.exists()
    assert list(output_dir.iterdir()) == []


def test_cli_invalid_layout_json_returns_one(tmp_path: Path) -> None:
    template = tmp_path / "extended_template.xlsx"
    payload_json = tmp_path / "payload.json"
    layout_json_path = tmp_path / "layout.json"
    output_dir = tmp_path / "out"
    output_dir.mkdir()
    output = output_dir / "draft.xlsx"
    write_extended_template(template)
    write_json(payload_json, payload(6))
    layout_json_path.write_text("{invalid", encoding="utf-8")

    exit_code = extended.main(
        cli_args(payload_json, layout_json_path, template, output)
    )

    assert exit_code == 1
    assert not output.exists()
    assert list(output_dir.iterdir()) == []


def test_cli_removes_temp_output_when_generation_fails(
    tmp_path: Path,
    monkeypatch: Any,
) -> None:
    template = tmp_path / "extended_template.xlsx"
    payload_json = tmp_path / "payload.json"
    layout_json_path = tmp_path / "layout.json"
    output_dir = tmp_path / "out"
    output_dir.mkdir()
    output = output_dir / "draft.xlsx"
    layout = write_extended_template(template)
    write_json(payload_json, payload(6))
    write_json(layout_json_path, layout_json(layout))

    def fail_verification(
        _output: Path,
        _layout: Any,
        _before: Any,
    ) -> None:
        raise extended.ExtendedFillError("forced verification failure")

    monkeypatch.setattr(extended, "verify_output", fail_verification)

    exit_code = extended.main(
        cli_args(payload_json, layout_json_path, template, output)
    )

    assert exit_code == 1
    assert not output.exists()
    assert list(output_dir.iterdir()) == []


def test_cli_subprocess_successfully_creates_output_for_six_items(
    tmp_path: Path,
) -> None:
    template = tmp_path / "extended_template.xlsx"
    payload_json = tmp_path / "payload.json"
    layout_json_path = tmp_path / "layout.json"
    output_dir = tmp_path / "out"
    output_dir.mkdir()
    output = output_dir / "draft.xlsx"
    layout = write_extended_template(template)
    write_json(payload_json, payload(6))
    write_json(layout_json_path, layout_json(layout))

    result = subprocess.run(
        [sys.executable, str(SCRIPT)]
        + cli_args(payload_json, layout_json_path, template, output),
        capture_output=True,
        text=True,
        check=False,
    )

    assert result.returncode == 0
    assert "CREATED:" in result.stdout
    assert "ERROR:" not in result.stderr
    assert output.is_file()
    assert workbook_values(output)["C22"] == "ВРУ-6"


def test_cli_subprocess_scales_to_fifty_items_when_layout_has_capacity(
    tmp_path: Path,
) -> None:
    template = tmp_path / "extended_template.xlsx"
    payload_json = tmp_path / "payload.json"
    layout_json_path = tmp_path / "layout.json"
    output_dir = tmp_path / "out"
    output_dir.mkdir()
    output = output_dir / "draft.xlsx"
    layout = write_extended_template(template, capacity=50)
    write_json(payload_json, payload(50))
    write_json(layout_json_path, layout_json(layout))

    result = subprocess.run(
        [sys.executable, str(SCRIPT)]
        + cli_args(payload_json, layout_json_path, template, output),
        capture_output=True,
        text=True,
        check=False,
    )

    assert result.returncode == 0
    assert "CREATED:" in result.stdout
    assert "ERROR:" not in result.stderr
    assert output.is_file()
    workbook = load_workbook(output, data_only=False)
    worksheet = workbook[extended.SHEET_NAME]
    assert worksheet["C17"].value == "ВРУ-1"
    assert worksheet["C66"].value == "ВРУ-50"
    assert worksheet["I67"].value == "=SUM(I17:I66)"
    assert worksheet["C2"].value == "header-2-3"
    assert worksheet["B70"].value == "signature"


def test_cli_subprocess_missing_payload_returns_one_without_output(
    tmp_path: Path,
) -> None:
    template = tmp_path / "extended_template.xlsx"
    payload_json = tmp_path / "missing_payload.json"
    layout_json_path = tmp_path / "layout.json"
    output_dir = tmp_path / "out"
    output_dir.mkdir()
    output = output_dir / "draft.xlsx"
    layout = write_extended_template(template)
    write_json(layout_json_path, layout_json(layout))

    result = subprocess.run(
        [sys.executable, str(SCRIPT)]
        + cli_args(payload_json, layout_json_path, template, output),
        capture_output=True,
        text=True,
        check=False,
    )

    assert result.returncode == 1
    assert "ERROR:" in result.stderr
    assert not output.exists()
    assert list(output_dir.iterdir()) == []


def test_items_above_capacity_stop_before_output(tmp_path: Path) -> None:
    template = tmp_path / "extended_template.xlsx"
    output_dir = tmp_path / "out"
    output_dir.mkdir()
    output = output_dir / "draft.xlsx"
    layout = write_extended_template(template, capacity=6)

    try:
        extended.generate_extended_workbook(
            template=template,
            output=output,
            payload=payload(7),
            layout=layout,
        )
    except extended.ExtendedFillError as error:
        assert "items count 7 exceeds layout capacity 6" in str(error)
    else:
        raise AssertionError("items above capacity should fail")

    assert not output.exists()
    assert list(output_dir.iterdir()) == []


def test_existing_output_is_not_overwritten(tmp_path: Path) -> None:
    template = tmp_path / "extended_template.xlsx"
    output = tmp_path / "out" / "draft.xlsx"
    output.parent.mkdir()
    output.write_bytes(b"existing")
    layout = write_extended_template(template)

    try:
        extended.generate_extended_workbook(
            template=template,
            output=output,
            payload=payload(6),
            layout=layout,
        )
    except extended.ExtendedFillError as error:
        assert "output already exists" in str(error)
    else:
        raise AssertionError("existing output should fail")

    assert output.read_bytes() == b"existing"
    assert list(output.parent.iterdir()) == [output]


def test_temp_output_is_removed_when_verification_fails(
    tmp_path: Path,
    monkeypatch: Any,
) -> None:
    template = tmp_path / "extended_template.xlsx"
    output_dir = tmp_path / "out"
    output_dir.mkdir()
    output = output_dir / "draft.xlsx"
    layout = write_extended_template(template)

    def fail_verification(
        _output: Path,
        _layout: Any,
        _before: Any,
    ) -> None:
        raise extended.ExtendedFillError("forced verification failure")

    monkeypatch.setattr(extended, "verify_output", fail_verification)

    try:
        extended.generate_extended_workbook(
            template=template,
            output=output,
            payload=payload(6),
            layout=layout,
        )
    except extended.ExtendedFillError as error:
        assert "forced verification failure" in str(error)
    else:
        raise AssertionError("verification failure should fail")

    assert not output.exists()
    assert list(output_dir.iterdir()) == []


def test_layout_fails_closed_before_output(tmp_path: Path) -> None:
    template = tmp_path / "extended_template.xlsx"
    output_dir = tmp_path / "out"
    output_dir.mkdir()
    output = output_dir / "draft.xlsx"
    write_extended_template(template)
    bad_layout = extended.ExtendedLayout(
        item_start_row=17,
        item_end_row=20,
        capacity=6,
        total_row=21,
        signature_range="B28:I30",
        header_ranges=("C2:I6",),
        formula_cells=("I17",),
    )

    try:
        extended.generate_extended_workbook(
            template=template,
            output=output,
            payload=payload(6),
            layout=bad_layout,
        )
    except extended.ExtendedFillError as error:
        assert "layout capacity must equal" in str(error)
    else:
        raise AssertionError("conflicting layout should fail")

    assert not output.exists()
    assert list(output_dir.iterdir()) == []


def test_old_mvp_sha256_stays_unchanged(tmp_path: Path) -> None:
    before = file_sha256(DRAFT_SCRIPT)
    template = tmp_path / "extended_template.xlsx"
    output_dir = tmp_path / "out"
    output_dir.mkdir()
    layout = write_extended_template(template)

    extended.generate_extended_workbook(
        template=template,
        output=output_dir / "draft.xlsx",
        payload=payload(6),
        layout=layout,
    )

    assert file_sha256(DRAFT_SCRIPT) == before


def test_no_real_xlsx_or_manifest_files_are_added_to_git() -> None:
    status = subprocess.run(
        ["git", "status", "--short", "--untracked-files=all"],
        cwd=PROJECT_ROOT,
        capture_output=True,
        text=True,
        check=True,
    )
    changed_paths = [
        line[3:].strip() for line in status.stdout.splitlines() if line.strip()
    ]
    assert not any(path.endswith(".xlsx") for path in changed_paths)
    assert not any(
        Path(path).name.casefold() == "manifest.json" for path in changed_paths
    )
