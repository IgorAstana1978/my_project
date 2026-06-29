import csv
import importlib.util
import subprocess
import sys
from pathlib import Path
from types import ModuleType
from typing import Any, cast

from openpyxl import Workbook  # type: ignore[import-untyped]

PROJECT_ROOT = Path(__file__).resolve().parents[1]
SCRIPT = PROJECT_ROOT / "scripts" / "inspect_quote_commercial_reconciliation.py"
OLD_FIVE_COLUMN_SCRIPT = (
    PROJECT_ROOT / "scripts" / "run_invoice_quote_extended_from_csv.py"
)


def load_script_module(module_name: str, path: Path) -> ModuleType:
    spec = importlib.util.spec_from_file_location(module_name, path)
    assert spec is not None
    assert spec.loader is not None
    module = importlib.util.module_from_spec(spec)
    sys.modules[spec.name] = module
    spec.loader.exec_module(module)
    return module


inspector = cast(
    Any,
    load_script_module("inspect_quote_commercial_reconciliation_for_test", SCRIPT),
)


def valid_row(index: int = 1) -> list[str]:
    return [
        f"SYNTHETIC-ITEM-{index}",
        "шт.",
        str(index),
        f"SYNTHETIC-DEVICES-{index}",
        f"SYNTHETIC-CABINET-{index}",
        str(index * 1000),
        "yes",
        "yes",
    ]


def write_csv(path: Path, rows: list[list[str]]) -> None:
    with path.open("w", encoding="utf-8", newline="") as csv_file:
        writer = csv.writer(csv_file, delimiter=";", lineterminator="\n")
        writer.writerow(inspector.commercial_preflight.REQUIRED_COLUMNS)
        writer.writerows(rows)


def write_workbook(
    path: Path,
    rows: list[list[str]],
    *,
    quantity_override: int | None = None,
    price_override: int | str | None = None,
    missing_used_rows: int = 0,
    wrong_item_formula: bool = False,
    wrong_total_formula: bool = False,
    extra_used_row: bool = False,
) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = inspector.SHEET_NAME

    rows_to_write = rows[: len(rows) - missing_used_rows]
    for offset, row in enumerate(rows_to_write):
        excel_row = inspector.ITEM_START_ROW + offset
        worksheet[f"C{excel_row}"] = row[0]
        worksheet[f"D{excel_row}"] = row[1]
        worksheet[f"E{excel_row}"] = (
            quantity_override
            if offset == 0 and quantity_override is not None
            else int(row[2])
        )
        worksheet[f"F{excel_row}"] = row[3]
        worksheet[f"G{excel_row}"] = row[4]
        worksheet[f"H{excel_row}"] = (
            price_override
            if offset == 0 and price_override is not None
            else int(row[5])
        )

    if extra_used_row:
        extra_row = inspector.ITEM_START_ROW + len(rows)
        worksheet[f"C{extra_row}"] = "EXTRA-SYNTHETIC-ROW"

    for excel_row in range(inspector.ITEM_START_ROW, inspector.ITEM_END_ROW + 1):
        worksheet[f"I{excel_row}"] = inspector.item_formula(excel_row)
    if wrong_item_formula:
        worksheet[f"I{inspector.ITEM_START_ROW}"] = "=E17*H17"
    worksheet[f"I{inspector.TOTAL_ROW}"] = (
        "=SUM(I17:I116)" if wrong_total_formula else inspector.total_formula()
    )

    for merge_range in inspector.EXPECTED_LOWER_MERGES:
        worksheet.merge_cells(merge_range)

    workbook.save(path)
    workbook.close()


def report_for(csv_path: Path, xlsx_path: Path) -> str:
    result = inspector.reconcile(csv_path, xlsx_path, 100)
    return cast(str, inspector.format_report(result))


def assert_status(csv_path: Path, xlsx_path: Path, expected: str) -> str:
    report = report_for(csv_path, xlsx_path)
    assert f"Status:\n{expected}" in report
    return report


def test_valid_synthetic_csv_and_xlsx_pass(tmp_path: Path) -> None:
    csv_path = tmp_path / "commercial.csv"
    xlsx_path = tmp_path / "draft.xlsx"
    rows = [valid_row(1), valid_row(2)]
    write_csv(csv_path, rows)
    write_workbook(xlsx_path, rows)

    report = assert_status(csv_path, xlsx_path, "PASS")

    for check in inspector.CommercialReconciliationResult(
        csv_path,
        xlsx_path,
        100,
    ).checks:
        assert f"{check}: pass" in report
    assert inspector.PASS_NEXT in report
    assert "Manual Igor check:\nrequired" in report


def test_quantity_mismatch_fails(tmp_path: Path) -> None:
    csv_path = tmp_path / "commercial.csv"
    xlsx_path = tmp_path / "draft.xlsx"
    rows = [valid_row()]
    write_csv(csv_path, rows)
    write_workbook(xlsx_path, rows, quantity_override=2)

    report = assert_status(csv_path, xlsx_path, "FAIL")

    assert "quantity reconciliation: fail" in report


def test_price_mismatch_fails_without_printing_price(tmp_path: Path) -> None:
    csv_path = tmp_path / "commercial.csv"
    xlsx_path = tmp_path / "draft.xlsx"
    rows = [valid_row()]
    write_csv(csv_path, rows)
    write_workbook(xlsx_path, rows, price_override=2000)

    report = assert_status(csv_path, xlsx_path, "FAIL")

    assert "unit price type: pass" in report
    assert "unit price reconciliation: fail" in report
    assert "2000" not in report


def test_price_stored_as_string_fails(tmp_path: Path) -> None:
    csv_path = tmp_path / "commercial.csv"
    xlsx_path = tmp_path / "draft.xlsx"
    rows = [valid_row()]
    write_csv(csv_path, rows)
    write_workbook(xlsx_path, rows, price_override="1000")

    report = assert_status(csv_path, xlsx_path, "FAIL")

    assert "unit price type: fail" in report
    assert "unit price reconciliation: fail" in report


def test_wrong_item_formula_fails(tmp_path: Path) -> None:
    csv_path = tmp_path / "commercial.csv"
    xlsx_path = tmp_path / "draft.xlsx"
    rows = [valid_row()]
    write_csv(csv_path, rows)
    write_workbook(xlsx_path, rows, wrong_item_formula=True)

    report = assert_status(csv_path, xlsx_path, "FAIL")

    assert "line formulas: fail" in report


def test_missing_item_formula_fails(tmp_path: Path) -> None:
    csv_path = tmp_path / "commercial.csv"
    xlsx_path = tmp_path / "draft.xlsx"
    rows = [valid_row()]
    write_csv(csv_path, rows)
    write_workbook(xlsx_path, rows)
    workbook = inspector.load_workbook(xlsx_path)
    worksheet = workbook[inspector.SHEET_NAME]
    worksheet["I18"] = None
    workbook.save(xlsx_path)
    workbook.close()

    report = assert_status(csv_path, xlsx_path, "FAIL")

    assert "line formulas: fail" in report


def test_wrong_total_formula_fails(tmp_path: Path) -> None:
    csv_path = tmp_path / "commercial.csv"
    xlsx_path = tmp_path / "draft.xlsx"
    rows = [valid_row()]
    write_csv(csv_path, rows)
    write_workbook(xlsx_path, rows, wrong_total_formula=True)

    report = assert_status(csv_path, xlsx_path, "FAIL")

    assert "total formula: fail" in report


def test_row_count_mismatch_fails(tmp_path: Path) -> None:
    csv_path = tmp_path / "commercial.csv"
    xlsx_path = tmp_path / "draft.xlsx"
    rows = [valid_row(1), valid_row(2)]
    write_csv(csv_path, rows)
    write_workbook(xlsx_path, rows, missing_used_rows=1)

    report = assert_status(csv_path, xlsx_path, "FAIL")

    assert "row count reconciliation: fail" in report


def test_extra_used_row_after_csv_rows_fails(tmp_path: Path) -> None:
    csv_path = tmp_path / "commercial.csv"
    xlsx_path = tmp_path / "draft.xlsx"
    rows = [valid_row()]
    write_csv(csv_path, rows)
    write_workbook(xlsx_path, rows, extra_used_row=True)

    report = assert_status(csv_path, xlsx_path, "FAIL")

    assert "row count reconciliation: fail" in report
    assert "unused rows: fail" in report


def test_invalid_commercial_csv_fails(tmp_path: Path) -> None:
    csv_path = tmp_path / "commercial.csv"
    xlsx_path = tmp_path / "draft.xlsx"
    invalid_row = valid_row()
    invalid_row[5] = "1.5"
    write_csv(csv_path, [invalid_row])
    write_workbook(xlsx_path, [valid_row()])

    report = assert_status(csv_path, xlsx_path, "FAIL")

    assert "commercial preflight: fail" in report
    assert "1.5" not in report


def test_missing_xlsx_fails(tmp_path: Path) -> None:
    csv_path = tmp_path / "commercial.csv"
    xlsx_path = tmp_path / "missing.xlsx"
    write_csv(csv_path, [valid_row()])

    report = assert_status(csv_path, xlsx_path, "FAIL")

    assert "workbook opens: fail" in report
    assert "draft XLSX does not exist" in report


def test_stale_cached_formula_value_is_rejected() -> None:
    assert not inspector.cached_values_match({"I17": "999"}, [1000], 1000)


def test_report_markers_present(tmp_path: Path) -> None:
    csv_path = tmp_path / "commercial.csv"
    xlsx_path = tmp_path / "draft.xlsx"
    rows = [valid_row()]
    write_csv(csv_path, rows)
    write_workbook(xlsx_path, rows)

    report = report_for(csv_path, xlsx_path)

    assert report.startswith("COMMERCIAL_QUOTE_RECONCILIATION_REPORT_START")
    assert report.endswith("COMMERCIAL_QUOTE_RECONCILIATION_REPORT_END")


def test_report_does_not_include_prices_totals_or_full_rows(tmp_path: Path) -> None:
    csv_path = tmp_path / "commercial.csv"
    xlsx_path = tmp_path / "draft.xlsx"
    row = valid_row()
    row[0] = "SECRET-SYNTHETIC-ITEM"
    row[3] = "SECRET-SYNTHETIC-DEVICES"
    row[5] = "987654321"
    write_csv(csv_path, [row])
    write_workbook(xlsx_path, [row])

    report = assert_status(csv_path, xlsx_path, "PASS")

    assert "SECRET-SYNTHETIC-ITEM" not in report
    assert "SECRET-SYNTHETIC-DEVICES" not in report
    assert "987654321" not in report
    assert ";".join(row) not in report


def test_inputs_are_not_modified(tmp_path: Path) -> None:
    csv_path = tmp_path / "commercial.csv"
    xlsx_path = tmp_path / "draft.xlsx"
    rows = [valid_row()]
    write_csv(csv_path, rows)
    write_workbook(xlsx_path, rows)
    csv_before = csv_path.read_bytes()
    xlsx_before = xlsx_path.read_bytes()

    report = assert_status(csv_path, xlsx_path, "PASS")

    assert "Status:\nPASS" in report
    assert csv_path.read_bytes() == csv_before
    assert xlsx_path.read_bytes() == xlsx_before


def test_cli_exit_codes_match_status(tmp_path: Path) -> None:
    csv_path = tmp_path / "commercial.csv"
    xlsx_path = tmp_path / "draft.xlsx"
    rows = [valid_row()]
    write_csv(csv_path, rows)
    write_workbook(xlsx_path, rows)

    pass_result = subprocess.run(
        [
            sys.executable,
            str(SCRIPT),
            "--commercial-csv",
            str(csv_path),
            "--draft-xlsx",
            str(xlsx_path),
            "--template-capacity",
            "100",
        ],
        capture_output=True,
        text=True,
        check=False,
    )
    fail_result = subprocess.run(
        [
            sys.executable,
            str(SCRIPT),
            "--commercial-csv",
            str(csv_path),
            "--draft-xlsx",
            str(tmp_path / "missing.xlsx"),
            "--template-capacity",
            "100",
        ],
        capture_output=True,
        text=True,
        check=False,
    )

    assert pass_result.returncode == 0
    assert "Status:\nPASS" in pass_result.stdout
    assert fail_result.returncode == 1
    assert "Status:\nFAIL" in fail_result.stdout


def test_old_five_column_workflow_contract_is_unchanged() -> None:
    old_workflow = cast(
        Any,
        load_script_module(
            "run_invoice_quote_extended_from_csv_contract_test",
            OLD_FIVE_COLUMN_SCRIPT,
        ),
    )

    assert old_workflow.REQUIRED_COLUMNS == (
        "name",
        "unit",
        "quantity",
        "instruments_and_devices",
        "cabinet_type_dimensions_material",
    )
    assert "unit_price_kzt" not in old_workflow.REQUIRED_COLUMNS
