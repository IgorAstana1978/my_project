from __future__ import annotations

import hashlib
import importlib.util
import json
import sys
from pathlib import Path
from types import ModuleType
from typing import Any, cast
from zipfile import ZipFile

import pytest
from openpyxl import load_workbook  # type: ignore[import-untyped]

ROOT = Path(__file__).resolve().parents[1]


def load_file(name: str, path: Path) -> ModuleType:
    spec = importlib.util.spec_from_file_location(name, path)
    assert spec is not None and spec.loader is not None
    module = importlib.util.module_from_spec(spec)
    sys.modules[name] = module
    spec.loader.exec_module(module)
    return module


def canonical_file(path: Path, value: object) -> str:
    raw = (
        json.dumps(
            value,
            ensure_ascii=False,
            sort_keys=True,
            separators=(",", ":"),
            allow_nan=False,
        )
        + "\n"
    ).encode()
    path.write_bytes(raw)
    return hashlib.sha256(raw).hexdigest()


def approved_document() -> dict[str, Any]:
    composition = "QF1 — автомат 16 А; KM1 — контактор 25 А\nБез сокращений"
    return {
        "schema_version": "dinva_quote_invoice_document.v0.1",
        "document_family": "DINVA_CLASSIC_QUOTE_INVOICE_V0_1",
        "document_type": "QUOTE_INVOICE",
        "document_id": "SYNTHETIC-001",
        "document_number": "TEST-001",
        "document_date": "2026-09-01",
        "currency": "KZT",
        "payer": "Synthetic Customer",
        "object_name": "Synthetic Object",
        "basis": None,
        "apparatus_heading": "Применяемые приборы и аппараты согласно схемы",
        "items": [
            {
                "position": 1,
                "name": "Шкаф управления",
                "unit": "шт.",
                "quantity": 2,
                "detailed_technical_composition": composition,
                "apparatus": "QF1 — автомат 16 А",
                "enclosure": "IP54, 600×400×250, металл",
                "approved_unit_price_kzt": 100000,
                "approved_line_total_kzt": 200000,
                "approval_reference": "SYNTHETIC-APPROVAL-1",
            }
        ],
        "approved_grand_total_kzt": 200000,
        "vat": {
            "rate_percent": 12,
            "included": True,
            "approved_amount_kzt": 21429,
            "approved_text": "В том числе НДС 12%",
        },
        "amount_words": {
            "amount_kzt": 200000,
            "approved_text": "Двести тысяч тенге 00 тиын",
        },
        "terms": {
            "payment": "50% предоплата",
            "delivery": "Самовывоз",
            "manufacturing_lead_time": "20 рабочих дней",
            "validity": "10 календарных дней",
        },
        "signatures": {
            "director_title": "Директор",
            "director_name": "Тестовый Директор",
            "executor_title": "Исполнитель",
            "executor_name": "Тестовый Исполнитель",
        },
        "approval_provenance": {
            "status": "APPROVED",
            "authority": "SYNTHETIC_TEST_AUTHORITY",
            "approval_id": "SYNTHETIC-ONLY",
            "approved_at": "2026-09-01T00:00:00Z",
            "source_sha256s": ["1" * 64],
            "rendering_authorized": True,
            "client_send_authorized": False,
        },
    }


def make_case(tmp_path: Path) -> dict[str, Any]:
    extractor_test = load_file(
        "dinva_extractor_test_helpers",
        ROOT / "tests" / "test_extract_dinva_classic_presentation_profile.py",
    )
    extractor = load_file(
        "dinva_extractor_for_render_tests",
        ROOT / "scripts" / "extract_dinva_classic_presentation_profile.py",
    )
    renderer = load_file(
        "dinva_renderer_for_tests",
        ROOT / "scripts" / "render_dinva_classic_quote_invoice.py",
    )
    synthetic_root = tmp_path / "synthetic-repo"
    synthetic_root.mkdir(exist_ok=True)
    extractor.__dict__["PROJECT_ROOT"] = synthetic_root
    renderer.__dict__["PROJECT_ROOT"] = synthetic_root
    first = tmp_path / "463.xlsx"
    second = tmp_path / "519.xlsx"
    runtime = tmp_path / "tuned-v4.xlsx"
    first_sha = extractor_test.write_reference(
        first, case_marker="463", renderer=renderer
    )
    second_sha = extractor_test.write_reference(
        second, case_marker="519", first_item_row=17, renderer=renderer
    )
    runtime_sha = extractor_test.write_runtime_template(runtime, renderer=renderer)
    profile = extractor.extract_profile(
        [
            extractor.ReferenceInput(first, first_sha),
            extractor.ReferenceInput(second, second_sha),
        ],
        [extractor.ReferenceInput(runtime, runtime_sha)],
    )
    document = approved_document()
    profile_path = tmp_path / "profile.json"
    document_path = tmp_path / "document.json"
    return {
        "renderer": renderer,
        "profile": profile,
        "document": document,
        "profile_path": profile_path,
        "document_path": document_path,
        "profile_sha": canonical_file(profile_path, profile),
        "document_sha": canonical_file(document_path, document),
        "references": (first, second, runtime),
    }


def render_case(case: dict[str, Any], output: Path) -> Path:
    renderer = case["renderer"]
    return cast(
        Path,
        renderer.render(
            profile_path=case["profile_path"],
            expected_profile_sha256=case["profile_sha"],
            document_path=case["document_path"],
            expected_document_sha256=case["document_sha"],
            output=output,
            allow_test_profile=True,
        ),
    )


def test_clean_render_preserves_exact_business_content_and_asset(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    monkeypatch.setenv("DINVA_RENDERER_TEST_MODE", "1")
    case = make_case(tmp_path)
    output = render_case(case, tmp_path / "output.xlsx")
    assert output.read_bytes() not in {
        reference.read_bytes() for reference in case["references"]
    }
    workbook = load_workbook(output, data_only=False)
    try:
        sheet = workbook["Счёт-КП шаблон"]
        item = case["document"]["items"][0]
        assert sheet["F17"].value == item["detailed_technical_composition"]
        assert sheet["I17"].value == (
            '=IF(OR(E17="",H17=""),"",IFERROR(E17*H17,"нужно уточнить"))'
        )
        assert sheet["I20"].value == (
            '=IF(COUNT(I17:I19)=0,"нужно уточнить",SUM(I17:I19))'
        )
        assert {str(value) for value in sheet.merged_cells.ranges} == set(
            case["profile"]["presentation_contract"]["layout"]["merged_cells"]["ranges"]
        )
        assert workbook.sheetnames == ["Счёт-КП шаблон"]
    finally:
        workbook.close()
    with ZipFile(output) as archive:
        names = set(archive.namelist())
        asset = case["profile"]["presentation_contract"]["assets"][0]
        assert (
            hashlib.sha256(archive.read("xl/media/image1.png")).hexdigest()
            == asset["sha256"]
        )
        assert "xl/calcChain.xml" not in names
        assert not any("externalLinks" in name for name in names)


@pytest.mark.parametrize(
    ("mutation", "message"),
    [
        (lambda case: case.update(profile_sha="0" * 64), "profile SHA-256 mismatch"),
        (
            lambda case: case["profile"].update(document_family="UNKNOWN"),
            "unsupported profile family",
        ),
        (
            lambda case: case["profile"].update(
                presentation_contract_fingerprint="0" * 64
            ),
            "fingerprint mismatch",
        ),
        (
            lambda case: case["document"].pop("payer"),
            "document fields mismatch",
        ),
    ],
)
def test_renderer_rejects_bad_bindings_and_malformed_input(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
    mutation: Any,
    message: str,
) -> None:
    monkeypatch.setenv("DINVA_RENDERER_TEST_MODE", "1")
    case = make_case(tmp_path)
    mutation(case)
    if case["profile_sha"] != "0" * 64:
        case["profile_sha"] = canonical_file(case["profile_path"], case["profile"])
    case["document_sha"] = canonical_file(case["document_path"], case["document"])
    with pytest.raises(case["renderer"].RendererError, match=message):
        render_case(case, tmp_path / "bad.xlsx")


def test_renderer_requires_approved_profile_by_default(tmp_path: Path) -> None:
    case = make_case(tmp_path)
    with pytest.raises(case["renderer"].RendererError, match="not immutable"):
        case["renderer"].render(
            profile_path=case["profile_path"],
            expected_profile_sha256=case["profile_sha"],
            document_path=case["document_path"],
            expected_document_sha256=case["document_sha"],
            output=tmp_path / "blocked.xlsx",
        )


def test_renderer_no_overwrite_outside_git_cleanup_and_toctou(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    monkeypatch.setenv("DINVA_RENDERER_TEST_MODE", "1")
    case = make_case(tmp_path)
    output = tmp_path / "exists.xlsx"
    output.write_bytes(b"occupied")
    with pytest.raises(case["renderer"].RendererError, match="already exists"):
        render_case(case, output)
    case["renderer"].PROJECT_ROOT = tmp_path / "synthetic-repo"
    with pytest.raises(case["renderer"].RendererError, match="outside Git"):
        render_case(case, tmp_path / "synthetic-repo" / "inside.xlsx")
    case["renderer"].PROJECT_ROOT = tmp_path / "another-synthetic-repo"
    original = case["renderer"].render_clean_workbook

    def mutate_after_render(*args: Any, **kwargs: Any) -> None:
        original(*args, **kwargs)
        case["document_path"].write_text("{}", encoding="utf-8")

    monkeypatch.setattr(case["renderer"], "render_clean_workbook", mutate_after_render)
    final = tmp_path / "toctou.xlsx"
    with pytest.raises(case["renderer"].RendererError, match="changed during render"):
        render_case(case, final)
    assert not final.exists()
    assert not list(tmp_path.glob(".*.candidate.xlsx"))
