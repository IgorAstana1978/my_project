import hashlib
import importlib.util
import json
import sys
import zipfile
from collections.abc import Mapping
from pathlib import Path
from types import ModuleType
from typing import Any, cast
from xml.etree import ElementTree

import pytest
from openpyxl import Workbook, load_workbook  # type: ignore[import-untyped]
from openpyxl.styles import Alignment  # type: ignore[import-untyped]
from openpyxl.worksheet.page import PageMargins  # type: ignore[import-untyped]

PROJECT_ROOT = Path(__file__).resolve().parents[1]
SCRIPT = PROJECT_ROOT / "scripts" / "checked_clientize_quote.py"
LAUNCHER = PROJECT_ROOT / "scripts" / "run_checked_clientization.ps1"
SOURCE_NOTE = (
    "ВН 3Р 25А заменён на ВН 3Р 32А — номинал 25А отсутствует в линейке CHINT."
)
CLIENT_NOTE = (
    "Примечание: ВН 3Р 25А заменён на ВН 3Р 32А, поскольку номинал 25А "
    "отсутствует в линейке CHINT."
)


def load_script_module(name: str, path: Path) -> ModuleType:
    spec = importlib.util.spec_from_file_location(name, path)
    assert spec is not None
    assert spec.loader is not None
    module = importlib.util.module_from_spec(spec)
    sys.modules[spec.name] = module
    spec.loader.exec_module(module)
    return module


clientizer = cast(
    Any,
    load_script_module("checked_clientize_quote_for_test", SCRIPT),
)


def sha256(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()


def item_rows() -> list[dict[str, Any]]:
    return [
        {
            "row": 17,
            "source_name": "BOARD-A",
            "client_name": "BOARD-A-CLIENT",
            "unit": "шт.",
            "quantity": 1,
            "instruments_and_devices": "CHINT DEVICES A",
            "cabinet_type_dimensions_material": "CABINET-A, металл",
            "unit_price_kzt": 100_000,
            "source_note": None,
            "client_note": None,
        },
        {
            "row": 18,
            "source_name": "BOARD-B",
            "client_name": "BOARD-B-CLIENT",
            "unit": "шт.",
            "quantity": 2,
            "instruments_and_devices": "CHINT DEVICES B 3P 32A",
            "cabinet_type_dimensions_material": "CABINET-B, металл",
            "unit_price_kzt": 50_000,
            "source_note": None,
            "client_note": None,
        },
        {
            "row": 19,
            "source_name": "BOARD-C",
            "client_name": "BOARD-C-CLIENT",
            "unit": "шт.",
            "quantity": 1,
            "instruments_and_devices": "CHINT DEVICES C 3P 32A",
            "cabinet_type_dimensions_material": "CABINET-C, металл",
            "unit_price_kzt": 90_000,
            "source_note": None,
            "client_note": None,
        },
    ]


def item_formula(row: int) -> str:
    return (
        f'=IF(OR(E{row}="",H{row}=""),"",' f'IFERROR(E{row}*H{row},"нужно уточнить"))'
    )


def add_synthetic_drawing_relationship(path: Path) -> None:
    with zipfile.ZipFile(path) as source:
        parts = {name: source.read(name) for name in source.namelist()}
    sheet = ElementTree.fromstring(parts["xl/worksheets/sheet1.xml"])
    drawing = ElementTree.SubElement(sheet, f"{{{clientizer.SPREADSHEET_NS}}}drawing")
    drawing.set(
        "{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id",
        "rId1",
    )
    ElementTree.register_namespace("", clientizer.SPREADSHEET_NS)
    ElementTree.register_namespace(
        "r", "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
    )
    parts["xl/worksheets/sheet1.xml"] = ElementTree.tostring(
        sheet, encoding="utf-8", xml_declaration=True
    )
    parts["xl/worksheets/_rels/sheet1.xml.rels"] = (
        b'<?xml version="1.0" encoding="UTF-8"?>'
        b'<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
        b'<Relationship Id="rId1" '
        b'Type="http://schemas.openxmlformats.org/'
        b'officeDocument/2006/relationships/drawing" '
        b'Target="../drawings/drawing1.xml"/></Relationships>'
    )
    parts["xl/drawings/drawing1.xml"] = (
        b'<?xml version="1.0" encoding="UTF-8"?>'
        b"<xdr:wsDr "
        b'xmlns:xdr="http://schemas.openxmlformats.org/'
        b'drawingml/2006/spreadsheetDrawing" '
        b'xmlns:a="http://schemas.openxmlformats.org/drawingml/2006/main"/>'
    )
    content_types = ElementTree.fromstring(parts["[Content_Types].xml"])
    ElementTree.SubElement(
        content_types,
        "{http://schemas.openxmlformats.org/package/2006/content-types}Override",
        {
            "PartName": "/xl/drawings/drawing1.xml",
            "ContentType": "application/vnd.openxmlformats-officedocument.drawing+xml",
        },
    )
    parts["[Content_Types].xml"] = ElementTree.tostring(
        content_types, encoding="utf-8", xml_declaration=True
    )
    with zipfile.ZipFile(path, "w", zipfile.ZIP_DEFLATED) as target:
        for name, content in parts.items():
            target.writestr(name, content)


def write_internal_draft(path: Path) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = clientizer.SHEET_NAME
    worksheet["B9"] = "Черновик счёта-КП № 463 от «10» июля 2026 года"
    worksheet["B10"] = "Плательщик: ТОО «Rich energy»"
    worksheet["B13"] = "Статус документа: Черновик"
    worksheet["F15"] = "Приборы и аппараты\nпроизводства CHINT"
    worksheet["G9"] = "ВНИМАНИЕ!"
    worksheet["G10"] = "Клиенту не отправлять без подтверждения Игоря"
    worksheet["G11"] = "Перед отправкой проверить цену, состав и срок у Игоря"
    worksheet["G12"] = "Закупку и цех не запускать по черновику"
    worksheet["G13"] = "Документ является внутренним черновиком"

    rows = item_rows()
    by_row = {item["row"]: item for item in rows}
    for row in range(clientizer.ITEM_START_ROW, clientizer.ITEM_END_ROW + 1):
        worksheet[f"B{row}"] = row - clientizer.ITEM_START_ROW + 1
        for column in "CDEFGH":
            worksheet[f"{column}{row}"].alignment = Alignment(vertical="top")
        worksheet[f"I{row}"] = item_formula(row)
        worksheet[f"J{row}"] = '=""'
        item = by_row.get(row)
        if item is None:
            continue
        worksheet[f"C{row}"] = item["source_name"]
        worksheet[f"D{row}"] = item["unit"]
        worksheet[f"E{row}"] = item["quantity"]
        worksheet[f"F{row}"] = item["instruments_and_devices"]
        worksheet[f"G{row}"] = item["cabinet_type_dimensions_material"]
        worksheet[f"H{row}"] = item["unit_price_kzt"]
    worksheet["I117"] = '=IF(COUNT(I17:I116)=0,"нужно уточнить",SUM(I17:I116))'
    worksheet["H118"] = (
        '=IF(NOT(ISNUMBER($A$131)),"","В том числе НДС "' '&TEXT($A$131,"0")&"%")'
    )
    worksheet["I118"] = (
        '=IF(OR(NOT(ISNUMBER(I117)),NOT(ISNUMBER($A$131))),"",'
        "I117*$A$131/(100+$A$131))"
    )
    worksheet["C119"] = "Всего прописью: Двести девяносто тысяч тенге 00 тиын"
    worksheet["C121"] = '=""'
    worksheet["C122"] = (
        "Условия оплаты: 100% предоплата. " "Условия поставки: EXW, г. Астана."
    )
    worksheet["C123"] = "Ориентировочный срок изготовления: 7–10 рабочих дней."
    worksheet["C124"] = (
        "Спецификация и условия подлежат проверке перед отправкой клиенту."
    )
    worksheet["C125"] = (
        "Документ является внутренним черновиком. "
        "Клиенту не отправлять без подтверждения Игоря."
    )
    worksheet["B129"] = "Дата проверки: ____ / ____ / 2026"
    worksheet["A131"] = 16
    for row in (17, 18, 19):
        worksheet.row_dimensions[row].height = 51
    for row in range(20, clientizer.ITEM_END_ROW + 1):
        worksheet.row_dimensions[row].height = 24
        worksheet.row_dimensions[row].hidden = True
    worksheet.column_dimensions["C"].width = 18
    worksheet.column_dimensions["F"].width = 48
    for cell_range in (
        "B9:F9",
        "B10:F10",
        "G9:I9",
        "G10:I10",
        "C119:I119",
        "C124:I124",
        "C125:I125",
    ):
        worksheet.merge_cells(cell_range)
    worksheet.print_area = "B2:I129"
    worksheet.page_margins = PageMargins(
        left=0.43307086614173229,
        right=0.23622047244094491,
        top=0.35433070866141736,
        bottom=0.74803149606299213,
        header=0.31496062992125984,
        footer=0.31496062992125984,
    )
    worksheet.page_setup.paperSize = "9"
    worksheet.page_setup.scale = 54
    worksheet.page_setup.fitToHeight = 0
    worksheet.page_setup.orientation = "portrait"
    workbook.save(path)
    workbook.close()
    add_synthetic_drawing_relationship(path)


def add_unreferenced_shared_guard(path: Path) -> None:
    with zipfile.ZipFile(path) as source:
        parts = {name: source.read(name) for name in source.namelist()}
    parts[clientizer.SHARED_STRINGS_PART] = (
        b'<?xml version="1.0" encoding="UTF-8"?>'
        b'<sst xmlns="http://schemas.openxmlformats.org/'
        b'spreadsheetml/2006/main" count="1" uniqueCount="1">'
        b"<si><t>internal draft stale token</t></si></sst>"
    ).replace(b"internal draft stale token", "Черновик stale token".encode())
    with zipfile.ZipFile(path, "w", zipfile.ZIP_DEFLATED) as target:
        for name, content in parts.items():
            target.writestr(name, content)


def approval_payload(internal_draft: Path) -> dict[str, Any]:
    return {
        "schema_version": clientizer.SCHEMA_VERSION,
        "approval_id": "SYNTHETIC-APPROVAL",
        "approved_by": "SYNTHETIC-IGOR",
        "approved_at": "2099-01-01T12:00:00+05:00",
        "internal_draft_xlsx_sha256": sha256(internal_draft),
        "invoice_number": "463",
        "invoice_date": "2026-07-10",
        "payer_name": "ТОО «Rich energy»",
        "apparatus_manufacturer": "CHINT",
        "vat_rate_percent": 16,
        "vat_amount_kzt": "40000.00",
        "commercial_total_kzt": 290_000,
        "payment_terms": "100% предоплата",
        "delivery_terms": "EXW, г. Астана",
        "manufacturing_lead_time": "7–10 рабочих дней",
        "manufacturing_lead_time_approved_by": "SYNTHETIC-PTO-ENGINEER",
        "manufacturing_lead_time_approved_at": "2099-01-01T11:30:00+05:00",
        "manufacturing_lead_time_approval_role": "pto_engineer",
        "validity_period": None,
        "amount_words_text": ("Всего прописью: Двести девяносто тысяч тенге 00 тиын"),
        "commercial_price_approved": "yes",
        "clientization_approved": "yes",
        "sending_approved": "no",
        "items": item_rows(),
    }


def write_json(path: Path, payload: Mapping[str, Any]) -> None:
    path.write_text(
        json.dumps(payload, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )


def build_case(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> dict[str, Any]:
    repo = tmp_path / "repo"
    outside = tmp_path / "outside"
    repo.mkdir()
    outside.mkdir()
    monkeypatch.setattr(clientizer, "PROJECT_ROOT", repo)
    monkeypatch.setattr(clientizer.ooxml_patcher, "PROJECT_ROOT", repo)
    internal_draft = outside / "internal-draft.xlsx"
    approval = outside / "approval.json"
    output = outside / "client-candidate.xlsx"
    write_internal_draft(internal_draft)
    add_unreferenced_shared_guard(internal_draft)
    payload = approval_payload(internal_draft)
    write_json(approval, payload)
    return {
        "repo": repo,
        "internal_draft": internal_draft,
        "approval": approval,
        "output": output,
        "payload": payload,
    }


def refresh_approval(case: Mapping[str, Any]) -> None:
    case["payload"]["internal_draft_xlsx_sha256"] = sha256(case["internal_draft"])
    write_json(case["approval"], case["payload"])


def set_source_cell(case: Mapping[str, Any], coordinate: str, value: object) -> None:
    workbook = load_workbook(case["internal_draft"])
    workbook[clientizer.SHEET_NAME][coordinate] = value
    workbook.save(case["internal_draft"])
    workbook.close()
    refresh_approval(case)


def approve_item_note(
    case: Mapping[str, Any],
    row: int,
    source_note: str = SOURCE_NOTE,
    client_note: str = CLIENT_NOTE,
) -> None:
    item = next(item for item in case["payload"]["items"] if item["row"] == row)
    item["source_note"] = source_note
    item["client_note"] = client_note
    set_source_cell(case, f"J{row}", source_note)


def add_unreferenced_shared_text(case: Mapping[str, Any], value: str) -> None:
    path = case["internal_draft"]
    with zipfile.ZipFile(path) as source:
        parts = {name: source.read(name) for name in source.namelist()}
    if clientizer.SHARED_STRINGS_PART in parts:
        root = ElementTree.fromstring(parts[clientizer.SHARED_STRINGS_PART])
    else:
        root = ElementTree.Element(
            f"{{{clientizer.SPREADSHEET_NS}}}sst",
            {"count": "0", "uniqueCount": "0"},
        )
    item = ElementTree.SubElement(root, f"{{{clientizer.SPREADSHEET_NS}}}si")
    text = ElementTree.SubElement(item, f"{{{clientizer.SPREADSHEET_NS}}}t")
    text.text = value
    for attribute in ("count", "uniqueCount"):
        root.set(attribute, str(int(root.get(attribute, "0")) + 1))
    parts[clientizer.SHARED_STRINGS_PART] = ElementTree.tostring(
        root, encoding="utf-8", xml_declaration=True
    )
    with zipfile.ZipFile(path, "w", zipfile.ZIP_DEFLATED) as target:
        for name, content in parts.items():
            target.writestr(name, content)
    refresh_approval(case)


def run_case(case: Mapping[str, Any]) -> Any:
    return clientizer.run_clientization(
        case["internal_draft"], case["approval"], case["output"]
    )


def assert_no_output_or_temporary_candidate(case: Mapping[str, Any]) -> None:
    assert not case["output"].exists()
    assert list(case["output"].parent.glob(".*.candidate.xlsx")) == []
    assert list(case["output"].parent.glob(".*.sanitize.tmp.xlsx")) == []


def package_text(path: Path) -> str:
    texts: list[str] = []
    with zipfile.ZipFile(path) as archive:
        for name in archive.namelist():
            try:
                texts.append(archive.read(name).decode("utf-8"))
            except UnicodeDecodeError:
                continue
    return "\n".join(texts)


def package_part(path: Path, part_name: str) -> bytes:
    with zipfile.ZipFile(path) as archive:
        return archive.read(part_name)


def rewrite_worksheet(
    path: Path,
    worksheet_part: str,
    mutation: str,
) -> None:
    with zipfile.ZipFile(path) as source:
        parts = {name: source.read(name) for name in source.namelist()}
    root = ElementTree.fromstring(parts[worksheet_part])
    namespace = {"main": clientizer.SPREADSHEET_NS}
    if mutation == "cell":
        node = root.find(".//main:c[@r='C122']", namespace)
        assert node is not None
        for child in list(node):
            node.remove(child)
        node.set("t", "inlineStr")
        inline = ElementTree.SubElement(node, f"{{{clientizer.SPREADSHEET_NS}}}is")
        text = ElementTree.SubElement(inline, f"{{{clientizer.SPREADSHEET_NS}}}t")
        text.text = "tampered terms"
    elif mutation == "style":
        node = root.find(".//main:c[@r='H17']", namespace)
        assert node is not None
        node.set("s", "999")
    elif mutation == "layout":
        node = root.find("main:pageSetup", namespace)
        assert node is not None
        node.set("scale", "55")
    elif mutation == "duplicate_client_note_j20":
        node = root.find(".//main:c[@r='J20']", namespace)
        assert node is not None
        for child in list(node):
            node.remove(child)
        node.set("t", "inlineStr")
        inline = ElementTree.SubElement(node, f"{{{clientizer.SPREADSHEET_NS}}}is")
        text = ElementTree.SubElement(inline, f"{{{clientizer.SPREADSHEET_NS}}}t")
        text.text = CLIENT_NOTE
    elif mutation in {"formula_guard", "formula_item_name", "formula_title"}:
        coordinate, cached_value = {
            "formula_guard": ("G12", None),
            "formula_item_name": ("C17", "BOARD-A-CLIENT"),
            "formula_title": ("B9", "Счёт-КП № 463 от «10» июля 2026 года"),
        }[mutation]
        node = root.find(f".//main:c[@r='{coordinate}']", namespace)
        assert node is not None
        for child in list(node):
            node.remove(child)
        node.set("t", "str")
        formula = ElementTree.SubElement(node, f"{{{clientizer.SPREADSHEET_NS}}}f")
        formula.text = '"synthetic hidden formula"'
        cached = ElementTree.SubElement(node, f"{{{clientizer.SPREADSHEET_NS}}}v")
        cached.text = cached_value
    else:
        raise AssertionError(f"unknown mutation: {mutation}")
    ElementTree.register_namespace("", clientizer.SPREADSHEET_NS)
    parts[worksheet_part] = ElementTree.tostring(
        root, encoding="utf-8", xml_declaration=True
    )
    with zipfile.ZipFile(path, "w", zipfile.ZIP_DEFLATED) as target:
        for name, content in parts.items():
            target.writestr(name, content)


def install_post_patch_mutation(
    monkeypatch: pytest.MonkeyPatch,
    mutation: str,
) -> None:
    original = clientizer.sanitize_unreferenced_shared_strings

    def sanitize_then_mutate(
        path: Path,
        worksheet_part: str,
        approval: Any,
    ) -> None:
        original(path, worksheet_part, approval)
        rewrite_worksheet(path, worksheet_part, mutation)

    monkeypatch.setattr(
        clientizer,
        "sanitize_unreferenced_shared_strings",
        sanitize_then_mutate,
    )


def test_valid_multi_item_clientization_is_checked_and_atomic(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    case = build_case(tmp_path, monkeypatch)
    source_hash = sha256(case["internal_draft"])

    result = run_case(case)

    assert result.status == "PASS"
    assert result.item_count == 3
    assert result.approval_id == "SYNTHETIC-APPROVAL"
    assert result.approved_by == "SYNTHETIC-IGOR"
    assert result.manufacturing_lead_time_approved_by == "SYNTHETIC-PTO-ENGINEER"
    assert result.manufacturing_lead_time_approved_at == "2099-01-01T11:30:00+05:00"
    assert result.manufacturing_lead_time_approval_role == "pto_engineer"
    assert all(status == "pass" for status in result.checks.values())
    assert sha256(case["internal_draft"]) == source_hash
    assert case["output"].is_file()
    output_text = package_text(case["output"]).casefold()
    assert not any(token in output_text for token in clientizer.FORBIDDEN_TOKENS)
    candidate_snapshot = clientizer.load_snapshot(case["output"])
    formula_free_coordinates = {
        "B9",
        *clientizer.GUARD_CELLS,
        *(f"J{row}" for row in range(17, 117)),
        *(f"C{row}" for row in (17, 18, 19)),
    }
    assert all(
        clientizer.cell(candidate_snapshot, coordinate).formula is None
        for coordinate in formula_free_coordinates
    )

    workbook = load_workbook(case["output"], data_only=False, read_only=False)
    worksheet = workbook[clientizer.SHEET_NAME]
    assert worksheet["B9"].value == "Счёт-КП № 463 от «10» июля 2026 года"
    assert [worksheet[f"C{row}"].value for row in (17, 18, 19)] == [
        "BOARD-A-CLIENT",
        "BOARD-B-CLIENT",
        "BOARD-C-CLIENT",
    ]
    assert worksheet["F18"].value == "CHINT DEVICES B 3P 32A"
    assert worksheet["J18"].value is None
    assert worksheet["J19"].value is None
    assert worksheet["I117"].value == (
        '=IF(COUNT(I17:I116)=0,"нужно уточнить",SUM(I17:I116))'
    )
    assert "B9:F9" in {str(value) for value in worksheet.merged_cells.ranges}
    assert worksheet.row_dimensions[20].hidden is True
    assert worksheet.row_dimensions[20].height == 24
    assert worksheet.page_setup.scale == 54
    assert worksheet.page_setup.orientation == "portrait"
    assert worksheet["H17"].style_id > 0
    workbook.close()
    source_snapshot = clientizer.load_snapshot(case["internal_draft"])
    relationship_part = (
        Path(source_snapshot.worksheet_part).parent
        / "_rels"
        / f"{Path(source_snapshot.worksheet_part).name}.rels"
    ).as_posix()
    assert package_part(case["internal_draft"], relationship_part) == package_part(
        case["output"], relationship_part
    )


@pytest.mark.parametrize(
    ("mutation", "message"),
    [
        ("missing_by", "approval fields do not match the strict schema"),
        ("blank_by", "manufacturing_lead_time_approved_by"),
        ("naive_at", "manufacturing_lead_time_approved_at must include a timezone"),
        ("wrong_role", "manufacturing_lead_time_approval_role must be pto_engineer"),
    ],
)
def test_invalid_pto_approval_metadata_fails_closed(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
    mutation: str,
    message: str,
) -> None:
    case = build_case(tmp_path, monkeypatch)
    if mutation == "missing_by":
        del case["payload"]["manufacturing_lead_time_approved_by"]
    elif mutation == "blank_by":
        case["payload"]["manufacturing_lead_time_approved_by"] = "  "
    elif mutation == "naive_at":
        case["payload"]["manufacturing_lead_time_approved_at"] = "2099-01-01T11:30:00"
    else:
        case["payload"]["manufacturing_lead_time_approval_role"] = "sales"
    write_json(case["approval"], case["payload"])

    result = run_case(case)

    assert result.status == "FAIL"
    assert message in "\n".join(result.failures)
    assert_no_output_or_temporary_candidate(case)


def test_c123_mismatch_fails_closed(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    case = build_case(tmp_path, monkeypatch)
    set_source_cell(case, "C123", "Ориентировочный срок изготовления: 99 дней.")

    result = run_case(case)

    assert result.status == "FAIL"
    assert "manufacturing lead time does not match approval" in "\n".join(
        result.failures
    )
    assert result.checks["candidate generation"] == "fail"
    assert_no_output_or_temporary_candidate(case)


def test_arbitrary_item_name_with_explicit_notes_passes(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    case = build_case(tmp_path, monkeypatch)
    case["payload"]["items"][1]["client_name"] = "ARBITRARY-CLIENT-NAME"
    approve_item_note(case, 18)
    add_unreferenced_shared_text(case, SOURCE_NOTE)

    result = run_case(case)

    assert result.status == "PASS"
    assert all(status == "pass" for status in result.checks.values())
    workbook = load_workbook(case["output"], data_only=False, read_only=False)
    worksheet = workbook[clientizer.SHEET_NAME]
    assert all(
        worksheet.cell(row=row, column=column).value is None
        for row in range(9, 14)
        for column in range(7, 10)
    )
    assert worksheet["J17"].value is None
    assert worksheet["J18"].value == CLIENT_NOTE
    assert worksheet["J19"].value is None
    assert worksheet["C123"].value == (
        "Ориентировочный срок изготовления: 7–10 рабочих дней."
    )
    assert all(
        worksheet[f"J{row}"].value is None
        for row in range(20, clientizer.ITEM_END_ROW + 1)
    )
    workbook.close()
    output_text = package_text(case["output"]).casefold()
    assert CLIENT_NOTE.casefold() in output_text
    assert SOURCE_NOTE.casefold() not in output_text
    assert not any(token in output_text for token in clientizer.FORBIDDEN_TOKENS)


def test_named_items_with_null_notes_do_not_gain_notes(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    case = build_case(tmp_path, monkeypatch)
    case["payload"]["items"][1]["client_name"] = "НЩР-17"
    case["payload"]["items"][2]["client_name"] = "АВР-17"
    write_json(case["approval"], case["payload"])

    result = run_case(case)

    assert result.status == "PASS"
    workbook = load_workbook(case["output"], data_only=False, read_only=False)
    worksheet = workbook[clientizer.SHEET_NAME]
    assert worksheet["J18"].value is None
    assert worksheet["J19"].value is None
    workbook.close()


def test_arbitrary_item_receives_exact_approved_client_note(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    case = build_case(tmp_path, monkeypatch)
    approve_item_note(case, 17, "source note for arbitrary item", "CLIENT NOTE: exact.")

    result = run_case(case)

    assert result.status == "PASS"
    candidate = clientizer.load_snapshot(case["output"])
    assert clientizer.cell(candidate, "J17").value == "CLIENT NOTE: exact."
    assert clientizer.cell(candidate, "J17").formula is None


def test_source_note_equal_to_client_note_passes_only_in_approved_j(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    case = build_case(tmp_path, monkeypatch)
    same_note = "Exact approved source and client note"
    approve_item_note(case, 18, same_note, same_note)

    result = run_case(case)

    assert result.status == "PASS"
    candidate = clientizer.load_snapshot(case["output"])
    matching = {
        coordinate
        for coordinate, value in candidate.cells.items()
        if value.value == same_note
    }
    assert matching == {"J18"}
    assert clientizer.cell(candidate, "J18").formula is None


def test_source_note_substring_of_client_note_passes_only_in_approved_j(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    case = build_case(tmp_path, monkeypatch)
    source_note = "approved source fragment"
    client_note = f"Client prefix: {source_note}; client suffix."
    approve_item_note(case, 18, source_note, client_note)

    result = run_case(case)

    assert result.status == "PASS"
    candidate = clientizer.load_snapshot(case["output"])
    assert clientizer.cell(candidate, "J18").value == client_note
    assert clientizer.cell(candidate, "J18").formula is None
    assert all(
        source_note not in value.value
        for coordinate, value in candidate.cells.items()
        if coordinate != "J18" and isinstance(value.value, str)
    )


def test_same_client_note_approved_for_two_j_cells_passes(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    case = build_case(tmp_path, monkeypatch)
    shared_client_note = "Same exact approved client note"
    approve_item_note(case, 18, "source note row 18", shared_client_note)
    approve_item_note(case, 19, "source note row 19", shared_client_note)

    result = run_case(case)

    assert result.status == "PASS"
    candidate = clientizer.load_snapshot(case["output"])
    matching = {
        coordinate
        for coordinate, value in candidate.cells.items()
        if value.value == shared_client_note
    }
    assert matching == {"J18", "J19"}
    assert clientizer.cell(candidate, "J18").formula is None
    assert clientizer.cell(candidate, "J19").formula is None


def test_source_note_mismatch_fails_before_candidate_generation(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    case = build_case(tmp_path, monkeypatch)
    approve_item_note(case, 18)
    case["payload"]["items"][1]["source_note"] = "different approved source note"
    write_json(case["approval"], case["payload"])

    result = run_case(case)

    assert result.status == "FAIL"
    assert "internal draft source note mismatch: J18" in result.failures
    assert result.checks["candidate generation"] == "fail"
    assert not case["output"].exists()


@pytest.mark.parametrize(
    ("source_note", "client_note"),
    [(SOURCE_NOTE, None), (None, CLIENT_NOTE)],
)
def test_note_pair_must_be_both_null_or_both_strings(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
    source_note: str | None,
    client_note: str | None,
) -> None:
    case = build_case(tmp_path, monkeypatch)
    case["payload"]["items"][1]["source_note"] = source_note
    case["payload"]["items"][1]["client_note"] = client_note
    write_json(case["approval"], case["payload"])

    result = run_case(case)

    assert result.status == "FAIL"
    assert "must both be null or strings" in "\n".join(result.failures)
    assert not case["output"].exists()


@pytest.mark.parametrize("field", ["source_note", "client_note"])
def test_note_fields_reject_empty_strings(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
    field: str,
) -> None:
    case = build_case(tmp_path, monkeypatch)
    case["payload"]["items"][1]["source_note"] = SOURCE_NOTE
    case["payload"]["items"][1]["client_note"] = CLIENT_NOTE
    case["payload"]["items"][1][field] = "  "
    write_json(case["approval"], case["payload"])

    result = run_case(case)

    assert result.status == "FAIL"
    assert f"{field} must be null or a non-empty string" in "\n".join(result.failures)
    assert not case["output"].exists()


def test_client_note_with_internal_guard_fails_schema_validation(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    case = build_case(tmp_path, monkeypatch)
    case["payload"]["items"][1]["source_note"] = SOURCE_NOTE
    case["payload"]["items"][1]["client_note"] = "Клиенту не отправлять"
    write_json(case["approval"], case["payload"])

    result = run_case(case)

    assert result.status == "FAIL"
    assert "client_note contains an internal forbidden token" in "\n".join(
        result.failures
    )
    assert not case["output"].exists()


def test_v0_1_approval_is_unsupported(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    case = build_case(tmp_path, monkeypatch)
    case["payload"]["schema_version"] = "checked_clientization_approval.v0.1"
    write_json(case["approval"], case["payload"])

    result = run_case(case)

    assert result.status == "FAIL"
    assert "schema_version is unsupported" in "\n".join(result.failures)
    assert not case["output"].exists()


def test_missing_approved_manufacturing_lead_time_fails_closed(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    case = build_case(tmp_path, monkeypatch)
    case["payload"]["manufacturing_lead_time"] = ""
    write_json(case["approval"], case["payload"])

    result = run_case(case)

    assert result.status == "FAIL"
    assert "non-empty string: manufacturing_lead_time" in "\n".join(result.failures)
    assert not case["output"].exists()


def test_hash_mismatch_fails_without_candidate_or_output(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    case = build_case(tmp_path, monkeypatch)
    case["payload"]["internal_draft_xlsx_sha256"] = "0" * 64
    write_json(case["approval"], case["payload"])

    result = run_case(case)

    assert result.status == "FAIL"
    assert "SHA-256 does not match" in "\n".join(result.failures)
    assert not case["output"].exists()
    assert list(case["output"].parent.glob(".*.candidate.xlsx")) == []


def test_source_reconciliation_rejects_changed_commercial_value(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    case = build_case(tmp_path, monkeypatch)
    workbook = load_workbook(case["internal_draft"])
    workbook[clientizer.SHEET_NAME]["H18"] = 50_001
    workbook.save(case["internal_draft"])
    workbook.close()
    case["payload"]["internal_draft_xlsx_sha256"] = sha256(case["internal_draft"])
    write_json(case["approval"], case["payload"])

    result = run_case(case)

    assert result.status == "FAIL"
    assert "internal draft item mismatch: H18" in result.failures
    assert not case["output"].exists()


def test_unapproved_extra_item_row_fails_before_candidate_generation(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    case = build_case(tmp_path, monkeypatch)
    workbook = load_workbook(case["internal_draft"])
    worksheet = workbook[clientizer.SHEET_NAME]
    worksheet["C20"] = "UNAPPROVED-BOARD"
    worksheet["D20"] = "шт."
    worksheet["E20"] = 1
    worksheet["F20"] = "UNAPPROVED COMPOSITION"
    worksheet["G20"] = "UNAPPROVED CABINET"
    worksheet["H20"] = 1
    workbook.save(case["internal_draft"])
    workbook.close()
    case["payload"]["internal_draft_xlsx_sha256"] = sha256(case["internal_draft"])
    write_json(case["approval"], case["payload"])

    result = run_case(case)

    assert result.status == "FAIL"
    assert "unapproved item field: C20" in "\n".join(result.failures)
    assert result.checks["candidate generation"] == "fail"
    assert not case["output"].exists()
    assert list(case["output"].parent.glob(".*.candidate.xlsx")) == []


def test_missing_required_guard_fails_before_candidate_generation(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    case = build_case(tmp_path, monkeypatch)
    workbook = load_workbook(case["internal_draft"])
    workbook[clientizer.SHEET_NAME]["G12"] = None
    workbook.save(case["internal_draft"])
    workbook.close()
    case["payload"]["internal_draft_xlsx_sha256"] = sha256(case["internal_draft"])
    write_json(case["approval"], case["payload"])

    result = run_case(case)

    assert result.status == "FAIL"
    assert "certified cell is missing: G12" in result.failures
    assert result.checks["candidate generation"] == "fail"
    assert not case["output"].exists()
    assert list(case["output"].parent.glob(".*.candidate.xlsx")) == []


@pytest.mark.parametrize("path_kind", ["input", "approval", "output"])
def test_input_or_output_inside_git_fails_closed(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
    path_kind: str,
) -> None:
    case = build_case(tmp_path, monkeypatch)
    if path_kind == "input":
        inside = case["repo"] / "internal-draft.xlsx"
        inside.write_bytes(case["internal_draft"].read_bytes())
        case["internal_draft"] = inside
        case["payload"]["internal_draft_xlsx_sha256"] = sha256(inside)
        write_json(case["approval"], case["payload"])
    elif path_kind == "approval":
        inside = case["repo"] / "approval.json"
        write_json(inside, case["payload"])
        case["approval"] = inside
    else:
        case["output"] = case["repo"] / "client-candidate.xlsx"

    result = run_case(case)

    assert result.status == "FAIL"
    assert "outside Git" in "\n".join(result.failures)
    assert not case["output"].exists()
    assert list(case["output"].parent.glob(".*.candidate.xlsx")) == []


def test_residual_guard_outside_certified_cells_prevents_publish(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    case = build_case(tmp_path, monkeypatch)
    workbook = load_workbook(case["internal_draft"])
    workbook[clientizer.SHEET_NAME]["K130"] = "Черновик residual guard"
    workbook.save(case["internal_draft"])
    workbook.close()
    case["payload"]["internal_draft_xlsx_sha256"] = sha256(case["internal_draft"])
    write_json(case["approval"], case["payload"])

    result = run_case(case)

    assert result.status == "FAIL"
    assert "candidate still contains an internal guard token" in result.failures
    assert not case["output"].exists()


def test_source_note_remaining_elsewhere_in_package_prevents_publish(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    case = build_case(tmp_path, monkeypatch)
    approve_item_note(case, 18)
    set_source_cell(case, "K130", SOURCE_NOTE)

    result = run_case(case)

    assert result.status == "FAIL"
    assert "note text exists outside its exact approved J location" in "\n".join(
        result.failures
    )
    assert_no_output_or_temporary_candidate(case)


def test_duplicate_client_note_in_k130_prevents_publish(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    case = build_case(tmp_path, monkeypatch)
    approve_item_note(case, 18)
    set_source_cell(case, "K130", CLIENT_NOTE)

    result = run_case(case)

    assert result.status == "FAIL"
    assert "note text exists outside its exact approved J location" in "\n".join(
        result.failures
    )
    assert_no_output_or_temporary_candidate(case)


def test_prefixed_client_note_in_k130_prevents_publish(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    case = build_case(tmp_path, monkeypatch)
    approve_item_note(case, 18)
    set_source_cell(case, "K130", f"prefix {CLIENT_NOTE}")

    result = run_case(case)

    assert result.status == "FAIL"
    assert "note text exists outside its exact approved J location" in "\n".join(
        result.failures
    )
    assert_no_output_or_temporary_candidate(case)


def test_source_note_with_suffix_in_other_cell_prevents_publish(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    case = build_case(tmp_path, monkeypatch)
    approve_item_note(case, 18)
    set_source_cell(case, "K130", f"prefix {SOURCE_NOTE} suffix")

    result = run_case(case)

    assert result.status == "FAIL"
    assert "note text exists outside its exact approved J location" in "\n".join(
        result.failures
    )
    assert_no_output_or_temporary_candidate(case)


def test_duplicate_client_note_in_other_j_prevents_publish(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    case = build_case(tmp_path, monkeypatch)
    approve_item_note(case, 18)
    install_post_patch_mutation(monkeypatch, "duplicate_client_note_j20")

    result = run_case(case)

    assert result.status == "FAIL"
    assert "note text exists outside its exact approved J location" in "\n".join(
        result.failures
    )
    assert result.checks["candidate generation"] == "pass"
    assert_no_output_or_temporary_candidate(case)


def test_unreferenced_duplicate_client_note_prevents_publish(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    case = build_case(tmp_path, monkeypatch)
    approve_item_note(case, 18)
    add_unreferenced_shared_text(case, CLIENT_NOTE)

    result = run_case(case)

    assert result.status == "FAIL"
    assert "unapproved shared-string location" in "\n".join(result.failures)
    assert_no_output_or_temporary_candidate(case)


def test_unknown_unapproved_source_note_is_not_allowed_automatically(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    case = build_case(tmp_path, monkeypatch)
    set_source_cell(case, "J18", "unknown technical note")

    result = run_case(case)

    assert result.status == "FAIL"
    assert "source note must use empty formula: J18" in "\n".join(result.failures)
    assert result.checks["candidate generation"] == "fail"
    assert not case["output"].exists()


@pytest.mark.parametrize(
    ("mutation", "message"),
    [
        ("cell", "non-allowed worksheet cell: C122"),
        ("style", "non-allowed worksheet cell: H17"),
        ("layout", "worksheet layout or print settings changed"),
    ],
)
def test_post_patch_non_allowed_worksheet_mutation_fails_and_cleans_candidate(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
    mutation: str,
    message: str,
) -> None:
    case = build_case(tmp_path, monkeypatch)
    install_post_patch_mutation(monkeypatch, mutation)

    result = run_case(case)

    assert result.status == "FAIL"
    assert message in "\n".join(result.failures)
    assert result.checks["candidate generation"] == "pass"
    assert result.checks["candidate reconciliation"] == "fail"
    assert not case["output"].exists()
    assert list(case["output"].parent.glob(".*.candidate.xlsx")) == []


@pytest.mark.parametrize(
    ("mutation", "message"),
    [
        ("formula_guard", "guard cell must not contain a formula: G12"),
        ("formula_item_name", "item name contains a formula: C17"),
        ("formula_title", "invoice title must not contain a formula: B9"),
    ],
)
def test_post_patch_formula_in_client_facing_cell_fails_and_cleans_candidate(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
    mutation: str,
    message: str,
) -> None:
    case = build_case(tmp_path, monkeypatch)
    install_post_patch_mutation(monkeypatch, mutation)

    result = run_case(case)

    assert result.status == "FAIL"
    assert message in "\n".join(result.failures)
    assert result.checks["candidate generation"] == "pass"
    assert result.checks["candidate reconciliation"] == "fail"
    assert not case["output"].exists()
    assert list(case["output"].parent.glob(".*.candidate.xlsx")) == []


@pytest.mark.parametrize(
    ("field", "value", "message"),
    [
        ("commercial_price_approved", "no", "must be exact yes"),
        ("clientization_approved", "no", "must be exact yes"),
        ("sending_approved", "yes", "must remain exact no"),
    ],
)
def test_human_approval_gates_fail_closed(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
    field: str,
    value: str,
    message: str,
) -> None:
    case = build_case(tmp_path, monkeypatch)
    case["payload"][field] = value
    write_json(case["approval"], case["payload"])

    result = run_case(case)

    assert result.status == "FAIL"
    assert message in "\n".join(result.failures)
    assert not case["output"].exists()


def test_existing_output_is_not_overwritten(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    case = build_case(tmp_path, monkeypatch)
    case["output"].write_bytes(b"existing")

    result = run_case(case)

    assert result.status == "FAIL"
    assert case["output"].read_bytes() == b"existing"
    assert "output XLSX already exists" in result.failures


def test_launcher_calls_only_checked_transformer() -> None:
    content = LAUNCHER.read_text(encoding="utf-8")

    assert "checked_clientize_quote.py" in content
    assert "--internal-draft-xlsx" in content
    assert "--approval-json" in content
    assert "--output-xlsx" in content
    for forbidden in ("pdf", "calculator", "procurement", "production", "git"):
        assert f"& {forbidden}" not in content.casefold()


def test_report_does_not_print_commercial_values(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    case = build_case(tmp_path, monkeypatch)
    result = run_case(case)

    report = clientizer.format_report(result)

    assert "290000" not in report
    assert "40000.00" not in report
    assert "CHINT DEVICES" not in report
    assert "SYNTHETIC-APPROVAL" in report
    assert "SYNTHETIC-IGOR" in report
    assert "SYNTHETIC-PTO-ENGINEER" in report
    assert "2099-01-01T11:30:00+05:00" in report
    assert "pto_engineer" in report
    assert clientizer.REPORT_START in report
    assert clientizer.REPORT_END in report
