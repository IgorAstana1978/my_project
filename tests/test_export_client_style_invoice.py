import csv
import hashlib
import importlib.util
import json
import sys
import zipfile
from collections.abc import Mapping, Sequence
from pathlib import Path
from types import ModuleType
from typing import Any, cast

import pytest
from openpyxl import Workbook, load_workbook  # type: ignore[import-untyped]

PROJECT_ROOT = Path(__file__).resolve().parents[1]
SCRIPT = PROJECT_ROOT / "scripts" / "export_client_style_invoice.py"
PRODUCTION_ENTRY_POINTS = (
    PROJECT_ROOT / "scripts" / "run_invoice_quote_commercial_from_csv.py",
    PROJECT_ROOT / "scripts" / "make_quote_capacity100_commercial_checked.ps1",
    PROJECT_ROOT / "scripts" / "calc_quote_price_draft.py",
    PROJECT_ROOT / "scripts" / "make_quote_capacity100_checked.ps1",
)
CSV_COLUMNS = (
    "name",
    "unit",
    "quantity",
    "instruments_and_devices",
    "cabinet_type_dimensions_material",
    "unit_price_kzt",
    "price_includes_vat",
    "price_confirmed_by_igor",
)


def load_script_module(module_name: str, path: Path) -> ModuleType:
    spec = importlib.util.spec_from_file_location(module_name, path)
    assert spec is not None
    assert spec.loader is not None
    module = importlib.util.module_from_spec(spec)
    sys.modules[spec.name] = module
    spec.loader.exec_module(module)
    return module


exporter = cast(
    Any,
    load_script_module("export_client_style_invoice_for_test", SCRIPT),
)


def sha256(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()


def default_row() -> dict[str, str]:
    return {
        "name": "SAFE-ITEM",
        "unit": "шт.",
        "quantity": "2",
        "instruments_and_devices": "SAFE-DEVICES",
        "cabinet_type_dimensions_material": "SAFE-CABINET",
        "unit_price_kzt": "12345",
        "price_includes_vat": "yes",
        "price_confirmed_by_igor": "yes",
    }


def write_commercial_csv(
    path: Path,
    rows: Sequence[Mapping[str, str]],
    *,
    columns: Sequence[str] = CSV_COLUMNS,
) -> None:
    with path.open("w", encoding="utf-8", newline="") as csv_file:
        writer = csv.writer(csv_file, delimiter=";", lineterminator="\n")
        writer.writerow(columns)
        for row in rows:
            writer.writerow([row[column] for column in columns])


def write_simple_xlsx(path: Path) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Internal"
    worksheet["A1"] = "synthetic internal draft"
    workbook.save(path)
    workbook.close()


def write_template(path: Path) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Лист1"
    worksheet["C9"] = "Счёт № [номер] от [дата]"
    worksheet["C10"] = "Плательщик: [наименование плательщика]"
    worksheet["B15"] = "№ п/п"
    worksheet["C15"] = "Наименование"
    worksheet["D15"] = "Ед."
    worksheet["E15"] = "Кол-во"
    worksheet["F15"] = "Приборы"
    worksheet["G15"] = "Шкаф"
    worksheet["H15"] = "Цена"
    worksheet["I15"] = "Сумма"
    worksheet["B16"] = 1
    worksheet["C16"] = "[наименование позиции]"
    worksheet["D16"] = "[ед.]"
    worksheet["E16"] = "[кол-во]"
    worksheet["F16"] = "[приборы и аппараты согласно схеме]"
    worksheet["G16"] = "[тип шкафа]"
    worksheet["H16"] = "[цена]"
    worksheet["I16"] = "[сумма]"
    worksheet["C17"] = "ИТОГО"
    worksheet["I17"] = "[итого]"
    worksheet["C19"] = "ВСЕГО: [сумма прописью]; [текст НДС]"
    worksheet["C20"] = "[условия оплаты]"
    worksheet["C21"] = "[условия возврата]"
    worksheet["C22"] = "[условия изменения спецификации / срок действия]"
    worksheet["C23"] = "[условия договора]"
    worksheet["C25"] = "[условия поставки / срок изготовления]"
    worksheet["C29"] = "[должность]"
    worksheet["F29"] = "[подписант]"
    worksheet["E30"] = "Счёт № [номер] от [дата]"
    worksheet.print_area = "B2:I30"
    worksheet.page_setup.orientation = "portrait"
    worksheet.page_setup.paperSize = "9"
    workbook.save(path)
    workbook.close()


def contract_data(template: Path) -> dict[str, Any]:
    return {
        "contract_id": "SYNTHETIC-CONTRACT",
        "template_name": "SYNTHETIC-TEMPLATE",
        "template_version": "SYNTHETIC-V1",
        "expected_sheet_name": "Лист1",
        "template_sha256": sha256(template),
        "allowed_extra_sheets": [],
        "print": {
            "orientation": "portrait",
            "paper_size": "9",
            "print_area_required": True,
        },
        "layout": {
            "invoice_number_cell": "C9",
            "invoice_date_cell": "C9",
            "payer_cell": "C10",
            "object_cell": None,
            "table_header_row": 15,
            "first_item_row": 16,
            "item_columns": {
                "index": "B",
                "name": "C",
                "unit": "D",
                "quantity": "E",
                "instruments_and_devices": "F",
                "cabinet": "G",
                "unit_price": "H",
                "line_total": "I",
            },
            "amount_words_cell": "C19",
            "signer_name_cell": "F29",
            "signer_title_cell": "C29",
        },
        "required_fixed_labels": [
            {"cell": "B15", "expected_text_contains": "№"},
            {"cell": "C15", "expected_text_contains": "Наименование"},
            {"cell": "I15", "expected_text_contains": "Сумма"},
        ],
    }


def approval_data(
    commercial_csv: Path,
    internal_draft: Path,
    template: Path,
) -> dict[str, Any]:
    return {
        "approval_id": "SYNTHETIC-APPROVAL",
        "approved_by": "SYNTHETIC-APPROVER",
        "approved_at": "2099-01-01T12:00:00+00:00",
        "commercial_csv_sha256": sha256(commercial_csv),
        "internal_draft_xlsx_sha256": sha256(internal_draft),
        "template_sha256": sha256(template),
        "invoice_number": "INV-9001",
        "invoice_date": "2099-01-01",
        "payer_name": "SAFE-PAYER",
        "object_name": None,
        "vat_text_approved": "НДС включён по согласованному режиму",
        "payment_terms_approved": "Согласованные условия оплаты",
        "delivery_terms_approved": "Согласованные условия поставки",
        "validity_terms_approved": "Согласованный срок действия",
        "return_terms_approved": "Согласованные условия возврата",
        "signer_name": "SAFE-SIGNER",
        "signer_title": "SAFE-TITLE",
        "approval_note": "Согласованные условия договора",
    }


def write_json(path: Path, data: Mapping[str, Any]) -> None:
    path.write_text(
        json.dumps(data, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )


def build_case(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> dict[str, Any]:
    tmp_path.mkdir(parents=True, exist_ok=True)
    repo_root = tmp_path / "repo"
    repo_root.mkdir()
    outside = tmp_path / "outside"
    outside.mkdir()

    for module in (
        exporter,
        exporter.approval_preflight,
        exporter.template_contract_preflight,
        exporter.commercial_csv_preflight,
        exporter.ooxml_cell_patcher,
    ):
        monkeypatch.setattr(module, "PROJECT_ROOT", repo_root)

    commercial_csv = outside / "commercial.csv"
    internal_draft = outside / "internal-draft.xlsx"
    template = outside / "client-template.xlsx"
    contract = outside / "template-contract.json"
    approval = outside / "approval.json"
    output = outside / "client-candidate.xlsx"

    write_commercial_csv(commercial_csv, [default_row()])
    write_simple_xlsx(internal_draft)
    write_template(template)
    contract_payload = contract_data(template)
    write_json(contract, contract_payload)
    approval_payload = approval_data(commercial_csv, internal_draft, template)
    write_json(approval, approval_payload)

    return {
        "repo_root": repo_root,
        "commercial_csv": commercial_csv,
        "internal_draft": internal_draft,
        "template": template,
        "contract": contract,
        "approval": approval,
        "output": output,
        "contract_payload": contract_payload,
        "approval_payload": approval_payload,
    }


def refresh_approval(case: dict[str, Any]) -> None:
    payload = case["approval_payload"]
    payload["commercial_csv_sha256"] = sha256(case["commercial_csv"])
    payload["internal_draft_xlsx_sha256"] = sha256(case["internal_draft"])
    payload["template_sha256"] = sha256(case["template"])
    write_json(case["approval"], payload)


def refresh_template_hashes(case: dict[str, Any]) -> None:
    template_hash = sha256(case["template"])
    case["contract_payload"]["template_sha256"] = template_hash
    write_json(case["contract"], case["contract_payload"])
    case["approval_payload"]["template_sha256"] = template_hash
    write_json(case["approval"], case["approval_payload"])


def add_template_leftover(
    case: dict[str, Any],
    value: str,
    *,
    cell: str = "J40",
) -> None:
    workbook = load_workbook(case["template"])
    try:
        workbook["Лист1"][cell] = value
        workbook.save(case["template"])
    finally:
        workbook.close()
    refresh_template_hashes(case)


def run_case(case: Mapping[str, Any]) -> Any:
    return exporter.run_export(
        case["commercial_csv"],
        case["internal_draft"],
        case["template"],
        case["contract"],
        case["approval"],
        case["output"],
    )


def package_text(path: Path) -> str:
    values: list[str] = []
    with zipfile.ZipFile(path) as archive:
        for name in archive.namelist():
            try:
                values.append(archive.read(name).decode("utf-8"))
            except UnicodeDecodeError:
                continue
    return "\n".join(values)


def test_valid_one_item_export_passes_and_creates_output(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    case = build_case(tmp_path, monkeypatch)

    result = run_case(case)

    assert result.status == "PASS"
    assert case["output"].is_file()
    assert all(value == "pass" for value in result.checks.values())


def test_approval_preflight_failure_creates_no_output(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    case = build_case(tmp_path, monkeypatch)
    case["approval_payload"]["commercial_csv_sha256"] = "0" * 64
    write_json(case["approval"], case["approval_payload"])

    result = run_case(case)

    assert result.status == "FAIL"
    assert result.checks["approval preflight"] == "fail"
    assert not case["output"].exists()


def test_template_contract_preflight_failure_creates_no_output(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    case = build_case(tmp_path, monkeypatch)
    case["contract_payload"]["template_sha256"] = "0" * 64
    write_json(case["contract"], case["contract_payload"])

    result = run_case(case)

    assert result.status == "FAIL"
    assert result.checks["approval preflight"] == "pass"
    assert result.checks["template contract preflight"] == "fail"
    assert not case["output"].exists()


def test_two_commercial_rows_fail(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    case = build_case(tmp_path, monkeypatch)
    second = default_row()
    second["name"] = "SAFE-ITEM-2"
    write_commercial_csv(case["commercial_csv"], [default_row(), second])
    refresh_approval(case)

    result = run_case(case)

    assert result.status == "FAIL"
    assert "exactly one item row" in "\n".join(result.red_flags)
    assert not case["output"].exists()


def test_commercial_csv_extra_column_fails(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    case = build_case(tmp_path, monkeypatch)
    row = default_row()
    row["unexpected"] = "value"
    write_commercial_csv(
        case["commercial_csv"],
        [row],
        columns=(*CSV_COLUMNS, "unexpected"),
    )
    refresh_approval(case)

    result = run_case(case)

    assert result.status == "FAIL"
    assert result.checks["commercial CSV"] == "fail"
    assert not case["output"].exists()


def test_unconfirmed_price_fails(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    case = build_case(tmp_path, monkeypatch)
    row = default_row()
    row["price_confirmed_by_igor"] = "no"
    write_commercial_csv(case["commercial_csv"], [row])
    refresh_approval(case)

    result = run_case(case)

    assert result.status == "FAIL"
    assert "price_confirmed_by_igor" in "\n".join(result.red_flags)
    assert not case["output"].exists()


def test_price_without_included_vat_fails(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    case = build_case(tmp_path, monkeypatch)
    row = default_row()
    row["price_includes_vat"] = "no"
    write_commercial_csv(case["commercial_csv"], [row])
    refresh_approval(case)

    result = run_case(case)

    assert result.status == "FAIL"
    assert "VAT mode must be exact yes" in "\n".join(result.red_flags)
    assert not case["output"].exists()


def test_existing_output_fails_without_overwrite(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    case = build_case(tmp_path, monkeypatch)
    case["output"].write_bytes(b"existing output")

    result = run_case(case)

    assert result.status == "FAIL"
    assert case["output"].read_bytes() == b"existing output"


def test_output_inside_git_fails(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    case = build_case(tmp_path, monkeypatch)
    case["output"] = case["repo_root"] / "forbidden-output.xlsx"

    result = run_case(case)

    assert result.status == "FAIL"
    assert not case["output"].exists()
    assert "outside the Git project" in "\n".join(result.red_flags)


def test_reconciliation_failure_removes_partial_candidate(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    case = build_case(tmp_path, monkeypatch)
    monkeypatch.setattr(
        exporter,
        "reconcile_candidate",
        lambda *args, **kwargs: ["forced reconciliation failure"],
    )

    result = run_case(case)

    assert result.status == "FAIL"
    assert not case["output"].exists()
    assert not list(case["output"].parent.glob(".*.candidate.xlsx"))


def test_template_and_all_inputs_remain_unchanged(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    case = build_case(tmp_path, monkeypatch)
    inputs = (
        case["commercial_csv"],
        case["internal_draft"],
        case["template"],
        case["contract"],
        case["approval"],
    )
    before = {path: sha256(path) for path in inputs}

    result = run_case(case)

    assert result.status == "PASS"
    assert {path: sha256(path) for path in inputs} == before


def test_output_contains_no_template_placeholders(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    case = build_case(tmp_path, monkeypatch)

    result = run_case(case)
    text = package_text(case["output"])

    assert result.status == "PASS"
    for token in exporter.FORBIDDEN_PLACEHOLDERS:
        assert token.casefold() not in text.casefold()


def test_output_contains_no_forbidden_old_tokens(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    case = build_case(tmp_path, monkeypatch)

    result = run_case(case)
    text = package_text(case["output"])

    assert result.status == "PASS"
    for token in exporter.FORBIDDEN_OLD_TOKENS:
        assert token.casefold() not in text.casefold()


def test_approved_unit_price_44512_passes(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    case = build_case(tmp_path, monkeypatch)
    row = default_row()
    row["quantity"] = "1"
    row["unit_price_kzt"] = "44512"
    write_commercial_csv(case["commercial_csv"], [row])
    refresh_approval(case)

    result = run_case(case)
    workbook = load_workbook(case["output"], data_only=False)
    try:
        worksheet = workbook["Лист1"]
        assert worksheet["H16"].value == 44512
        assert worksheet["I16"].value == 44512
        assert worksheet["I17"].value == 44512
        assert str(worksheet["C19"].value).startswith(
            "ВСЕГО: Сорок четыре тысячи пятьсот двенадцать тенге 00 тиын, "
        )
    finally:
        workbook.close()
    assert result.status == "PASS"


def test_approved_vat_legacy_text_passes_only_from_approval(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    case = build_case(tmp_path, monkeypatch)
    case["approval_payload"]["vat_text_approved"] = "НДС 16%"
    write_json(case["approval"], case["approval_payload"])

    result = run_case(case)
    workbook = load_workbook(case["output"], data_only=False)
    try:
        assert str(workbook["Лист1"]["C19"].value).endswith(", НДС 16%")
    finally:
        workbook.close()
    assert result.status == "PASS"


def test_unapproved_vat_legacy_text_in_template_fails(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    case = build_case(tmp_path, monkeypatch)
    add_template_leftover(case, "НДС 16%")

    result = run_case(case)

    assert result.status == "FAIL"
    assert "unapproved legacy token" in "\n".join(result.red_flags)
    assert not case["output"].exists()


def test_unapproved_payer_legacy_text_in_template_fails(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    case = build_case(tmp_path, monkeypatch)
    add_template_leftover(case, "TDK Energy")

    result = run_case(case)

    assert result.status == "FAIL"
    assert "unapproved legacy token" in "\n".join(result.red_flags)
    assert not case["output"].exists()


def test_placeholder_in_template_remains_strictly_forbidden(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    case = build_case(tmp_path, monkeypatch)
    add_template_leftover(case, "[номер]")

    result = run_case(case)

    assert result.status == "FAIL"
    assert "forbidden placeholder" in "\n".join(result.red_flags)
    assert not case["output"].exists()


def test_amount_words_match_numeric_total(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    case = build_case(tmp_path, monkeypatch)

    result = run_case(case)
    workbook = load_workbook(case["output"], data_only=False)
    try:
        worksheet = workbook["Лист1"]
        assert worksheet["I16"].value == 24690
        assert worksheet["I17"].value == 24690
        assert worksheet["C19"].value == (
            "ВСЕГО: Двадцать четыре тысячи шестьсот девяносто тенге "
            "00 тиын, НДС включён по согласованному режиму"
        )
    finally:
        workbook.close()
    assert result.status == "PASS"


def test_invoice_number_date_and_payer_come_from_approval(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    case = build_case(tmp_path, monkeypatch)

    result = run_case(case)
    workbook = load_workbook(case["output"], data_only=False)
    try:
        worksheet = workbook["Лист1"]
        assert worksheet["C9"].value == "Счёт № INV-9001 от 2099-01-01"
        assert worksheet["C10"].value == "Плательщик: SAFE-PAYER"
        assert worksheet["E30"].value == "Счёт № INV-9001 от 2099-01-01"
    finally:
        workbook.close()
    assert result.status == "PASS"


def test_signer_comes_from_approval(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    case = build_case(tmp_path, monkeypatch)

    result = run_case(case)
    workbook = load_workbook(case["output"], data_only=False)
    try:
        worksheet = workbook["Лист1"]
        assert worksheet["C29"].value == "SAFE-TITLE"
        assert worksheet["F29"].value == "SAFE-SIGNER"
    finally:
        workbook.close()
    assert result.status == "PASS"


def test_existing_workflows_do_not_reference_client_exporter() -> None:
    for path in PRODUCTION_ENTRY_POINTS:
        assert path.is_file()
        assert "export_client_style_invoice" not in path.read_text(encoding="utf-8")


def test_report_has_safety_boundaries_without_full_terms(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    case = build_case(tmp_path, monkeypatch)

    result = run_case(case)
    report = cast(str, exporter.format_report(result))

    assert exporter.REPORT_START in report
    assert exporter.REPORT_END in report
    assert "Status:\nPASS" in report
    assert "PASS is not sending approval" in report
    assert "required before sending to client" in report
    for field_name in (
        "payment_terms_approved",
        "delivery_terms_approved",
        "validity_terms_approved",
        "return_terms_approved",
        "approval_note",
    ):
        assert case["approval_payload"][field_name] not in report


def test_main_returns_zero_for_pass_and_one_for_fail(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
    capsys: pytest.CaptureFixture[str],
) -> None:
    passing = build_case(tmp_path / "pass", monkeypatch)
    passing_args = [
        "--commercial-csv",
        str(passing["commercial_csv"]),
        "--internal-draft-xlsx",
        str(passing["internal_draft"]),
        "--template-xlsx",
        str(passing["template"]),
        "--template-contract-json",
        str(passing["contract"]),
        "--approval-json",
        str(passing["approval"]),
        "--output-xlsx",
        str(passing["output"]),
    ]
    assert exporter.main(passing_args) == 0
    assert "Status:\nPASS" in capsys.readouterr().out

    failing = build_case(tmp_path / "fail", monkeypatch)
    failing["approval_payload"]["template_sha256"] = "0" * 64
    write_json(failing["approval"], failing["approval_payload"])
    failing_args = [
        "--commercial-csv",
        str(failing["commercial_csv"]),
        "--internal-draft-xlsx",
        str(failing["internal_draft"]),
        "--template-xlsx",
        str(failing["template"]),
        "--template-contract-json",
        str(failing["contract"]),
        "--approval-json",
        str(failing["approval"]),
        "--output-xlsx",
        str(failing["output"]),
    ]
    assert exporter.main(failing_args) == 1
    assert "Status:\nFAIL" in capsys.readouterr().out
